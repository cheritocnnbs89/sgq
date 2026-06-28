# modules/routes_gastos_tarjetas.py
from __future__ import annotations
from decimal import Decimal, ROUND_HALF_UP
import html
import re
from flask import app, session, request, render_template
import re
import sqlite3
import os
import uuid
from itertools import zip_longest
from numpy import where
from werkzeug.utils import secure_filename
from flask import jsonify, request, current_app
from io import BytesIO
from datetime import datetime
from flask import send_file
from flask import jsonify, request
from flask import jsonify
import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os, json, requests
from flask import jsonify
from datetime import datetime
import os, uuid
from werkzeug.utils import secure_filename

import xml.etree.ElementTree as ET
import html
import re
 


import xml.etree.ElementTree as ET
import html
import re
import html
import re
import xml.etree.ElementTree as ET

 
from io import BytesIO
from datetime import datetime
import sqlite3

from flask import send_file, request, session, current_app
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
 
from flask import request, render_template, redirect, url_for, flash,current_app, session
from jinja2 import TemplateNotFound
from werkzeug.utils import secure_filename

from .security import require_login, require_permission, has_permission
from .db import get_db
from .config import TABLE_GASTOS
from .config import TABLE_GASTOS as CFG_TABLE_GASTOS


# ✅ Importa módulos (no símbolos sueltos) para evitar ImportError por inicialización parcial
from . import gastos_helpers as gh
from . import gastos_exports as gx
from . import routes_gatos_mail_notify as mail
 
    # routes_gastos_tarjeta.py (o donde tengas tus rutas)
import os, json, requests
from flask import jsonify
from datetime import datetime
# NUEVO (reemplazo)
import os, time, sqlite3
from datetime import datetime
from flask import render_template, request, redirect, url_for, flash
from flask import session as flask_session
from flask import current_app

from werkzeug.utils import secure_filename
from .db import get_db
from .security import require_login, require_permission
from . import gastos_helpers as gh
TABLE_GASTOS = CFG_TABLE_GASTOS or "gastos_tarjeta"
from decimal import Decimal



import xml.etree.ElementTree as ET
import html
import html  # ya lo tienes arriba
from math import ceil
from flask import request, render_template, redirect, url_for, flash



import xml.etree.ElementTree as ET
import html
import re




# =========================
# IVA -> Indicador (editable)
# =========================
IVA_INDICADOR_MAP = {
    Decimal("0.00"):  "C0",  # IVA 0% => CE
    Decimal("15.00"): "CE",  # IVA 15% => CE
    Decimal("8.00"):  "CH",  # IVA 8%  => CH
    # Cuando confirmes:
    # Decimal("10.00"): "XX",
    # Decimal("5.00"):  "YY",
}

IVA_INDICADOR_DEFAULT = "CE"  # estándar para otros (cámbialo a "OT" si prefieres)

def _get_session_empresa_id(conn, uid=None):
    """
    Devuelve empresa_id del usuario logueado.
    Primero intenta sesión; si no existe, consulta BD.
    """
    emp = session.get("empresa_id")
    if emp:
        try:
            return int(emp)
        except Exception:
            pass

    uid = uid or session.get("usuario_id") or session.get("user_id")
    if not uid:
        return None

    cur = conn.cursor()
    cur.execute("""
        SELECT empresa_id
        FROM usuarios
        WHERE id = ?
    """, (int(uid),))
    r = cur.fetchone()

    if not r:
        return None

    try:
        emp = r["empresa_id"]
    except Exception:
        emp = r[0]

    try:
        return int(emp) if emp is not None else None
    except Exception:
        return None


def _apply_empresa_scope(conn, where, args, user_alias="u", uid=None):
    """
    Aplica filtro de empresa a todo el módulo de gastos.
    Admin ve todo.
    Los demás solo ven usuarios/gastos de su misma empresa.
    Requiere que el SELECT tenga JOIN usuarios u ON u.id = g.usuario_id
    o el alias indicado en user_alias.
    """
    role_name = (session.get("rol") or "").lower().strip()
    is_admin = (role_name == "admin") or bool(session.get("is_admin"))

    if is_admin:
        return

    empresa_id = _get_session_empresa_id(conn, uid)
    if not empresa_id:
        # Seguridad: si no tiene empresa asignada, no ve nada.
        where.append("1=0")
        return

    where.append(f"COALESCE({user_alias}.empresa_id, -1) = ?")
    args.append(empresa_id)


def _q2(x: float | int | str | None) -> Decimal:
    try:
        return Decimal(str(x or "0")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return Decimal("0.00")

def indicador_por_tarifa(tarifa: Decimal) -> str:
    tarifa = tarifa.quantize(Decimal("0.01"))
    return IVA_INDICADOR_MAP.get(tarifa, IVA_INDICADOR_DEFAULT)

def _log_export_params(tag: str):
    try:
        current_app.logger.warning(
            "[%s] args=%s form=%s json=%s qs=%s user_id=%s rol=%s",
            tag,
            dict(request.args),
            dict(request.form),
            request.get_json(silent=True),
            request.query_string.decode("utf-8", "ignore"),
            session.get("user_id"),
            session.get("rol"),
        )
    except Exception as e:
        current_app.logger.exception("[%s] log failed: %s", tag, e)

IVA_INDICADOR_DEFAULT = "OT"   # estándar para otros (puede ser "CE" si así lo decides)
def get_param_map(conn, group_id: str) -> dict:
    """
    Devuelve un dict normalizado {NOMBRE_NORMALIZADO: VALOR} para un group_id.
    Solo activos. Si hay duplicados por nombre, toma el de menor orden/id.
    """
    cur = conn.cursor()
    cur.execute("""
        SELECT pv.nombre, pv.valor
        FROM param_values pv
        WHERE pv.group_id = ?
         ORDER BY pv.orden ASC, pv.id ASC
    """, (group_id,))
    rows = cur.fetchall() or []

    def norm_key(s: str) -> str:
        return (s or "").strip().upper().replace("_", " ")

    out = {}
    for row in rows:
        try:
            nombre = row["nombre"]
            valor = row["valor"]
        except Exception:
            nombre = row[0]
            valor = row[1]

        key = norm_key(nombre)
        if not key:
            continue

        if key not in out:
            out[key] = (valor or "").strip()

    return out
 
def get_sap_config_from_db(conn) -> dict:
    """
    Lee credenciales/config SAP desde BD usando group_id.
    Normaliza nombres como:
      URL SAP / URL_SAP
      URL SAP QAS / URL_SAP_QAS
      USUARIO
      CLAVE
      SAP CLIENT / SAP_CLIENT
    """
    GROUP = "5791"

    m = get_param_map(conn, GROUP)

    missing = []

    def req(k):
        v = (m.get(k) or "").strip()
        if not v:
            missing.append(k)
        return v

    cfg = {
        "SAP_URL": req("URL SAP"),
        "SAP_URL_QAS": (m.get("URL SAP QAS") or "").strip(),
        "SAP_USER": req("USUARIO"),
        "SAP_PASS": req("CLAVE"),
        "SAP_CLIENT": (m.get("SAP CLIENT") or "400").strip(),
    }

    if missing:
        raise ValueError(
            f"Faltan parámetros SAP en BD (grupo '{GROUP}'): {', '.join(missing)} | disponibles={list(m.keys())}"
        )

    if not cfg["SAP_CLIENT"]:
        cfg["SAP_CLIENT"] = "400"

    return cfg


def apply_facturas_xml_search(where_parts: list[str], params: list, q: str, alias: str = "f") -> None:
    """
    Agrega al WHERE una búsqueda unificada por:
    - Razón social emisor
    - RUC emisor
    - Clave de acceso
    - Nº autorización
    - Secuencial
    - Nº factura formateado: estab-pto-secuencial(9)
    - Nº factura sin guiones: estabpto + secuencial(9)
    """
    q = (q or "").strip()
    if not q:
        return

    q_upper = q.upper()
    like_upper = f"%{q_upper}%"
    like_raw = f"%{q}%"
    q_digits = re.sub(r"\D+", "", q)

    numero_fmt = (
        f"(COALESCE({alias}.estab,'') + '-' + COALESCE({alias}.pto_emi,'') + '-' + "
        f"RIGHT(REPLICATE('0', 9) + CAST(CAST(COALESCE({alias}.secuencial,'0') AS INT) AS VARCHAR(9)), 9))"
    )
    numero_digits = (
        f"(COALESCE({alias}.estab,'') + COALESCE({alias}.pto_emi,'') + "
        f"RIGHT(REPLICATE('0', 9) + CAST(CAST(COALESCE({alias}.secuencial,'0') AS INT) AS VARCHAR(9)), 9))"
    )
    ruc_digits = (
        f"REPLACE(REPLACE(REPLACE(COALESCE({alias}.ruc_emisor,''),' ',''),'-',''),'.','')"
    )

    sub = [
        f"UPPER(COALESCE({alias}.razon_social_emisor,'')) LIKE ?",
        f"UPPER(COALESCE({alias}.ruc_emisor,'')) LIKE ?",
        f"UPPER(COALESCE({alias}.clave_acceso,'')) LIKE ?",
        f"UPPER(COALESCE({alias}.numero_autorizacion,'')) LIKE ?",
        f"CAST(COALESCE({alias}.secuencial,'') AS TEXT) LIKE ?",
        f"{numero_fmt} LIKE ?",
    ]
    params.extend([like_upper, like_upper, like_upper, like_upper, like_raw, like_raw])

    if q_digits:
        sub.append(f"{numero_digits} LIKE ?")
        params.append(f"%{q_digits}%")

        sub.append(f"{ruc_digits} LIKE ?")
        params.append(f"%{q_digits}%")

    where_parts.append("(" + " OR ".join(sub) + ")")


# ==========================================================
# Helpers: gerente real (último jefe válido) + username
# ==========================================================
def _get_gerente_username_real(
    conn: sqlite3.Connection,
    user_id: int | None
) -> str:
    """
    TU regla:
    'el jefe se saca validando el jefe de mi jefe y cuando no se encuentre
    alguien más arriba se conoce como gerente'

    => gerente = último jefe válido arriba en la cadena.
    """
    gid = _get_ultimo_jefe_id(conn, user_id)
    return _get_username(conn, gid) if gid else ""





def obtener_siguiente_secuencia(conn, nombre_secuencia):
    cur = conn.cursor()
    # 1. Bloqueamos y actualizamos el valor en un solo paso
    cur.execute("UPDATE secuencias_sap SET ultimo_valor = ultimo_valor + 1 WHERE nombre = ?", (nombre_secuencia,))
    # 2. Obtenemos el nuevo valor
    cur.execute("SELECT ultimo_valor FROM secuencias_sap WHERE nombre = ?", (nombre_secuencia,))
    res = cur.fetchone()
    if res: 
        # Formateamos a 10 dígitos con ceros a la izquierda
        return f"{res['ultimo_valor']:09d}"
    return "0000000000"
 
 
def _get_ultimo_jefe_id(
    conn: sqlite3.Connection,
    user_id: int | None,
    *,
    fallback_to_self: bool = False
) -> int | None:
    """
    Sube por usuarios.jefe_id hasta que:
    - no haya más jefe_id
    - o se detecte loop
    Devuelve el último jefe válido encontrado (activo).

    Si fallback_to_self=True:
    - si el usuario no tiene jefe_id, devuelve el mismo user_id (útil cuando el creador ya es gerente).
    """
    if not user_id:
        return None

    cur = conn.cursor()
    cur.execute("SELECT jefe_id FROM usuarios WHERE id=?", (int(user_id),))
    row = cur.fetchone()
    if not row:
        return None

    jefe_id = row["jefe_id"]
    if not jefe_id:
        return int(user_id) if fallback_to_self else None

    seen = set()
    last_valid = None

    while jefe_id and jefe_id not in seen:
        seen.add(jefe_id)

        cur.execute("""
            SELECT id, jefe_id
            FROM usuarios
            WHERE id=?
              AND COALESCE(disabled,0)=0
        """, (int(jefe_id),))
        j = cur.fetchone()
        if not j:
            break

        last_valid = int(j["id"])
        jefe_id = j["jefe_id"]

    # Si por alguna razón no encontró un jefe activo en la cadena,
    # y fallback_to_self=True, cae al mismo usuario.
    if last_valid is None and fallback_to_self:
        return int(user_id)

    return last_valid


def _get_username(conn: sqlite3.Connection, user_id: int | None) -> str:
    if not user_id:
        return ""
    cur = conn.cursor()
    cur.execute(
        "SELECT COALESCE(username,'') AS u FROM usuarios WHERE id=?",
        (user_id,)
    )
    r = cur.fetchone()
    return (r["u"] or "") if r else ""


def _get_gerente_username_real(conn: sqlite3.Connection, user_id: int | None) -> str:
    """
    Regla:
    - el gerente es el último jefe válido en la cadena de jefes
    """
    gid = _get_ultimo_jefe_id(conn, user_id)
    return _get_username(conn, gid) if gid else ""


def _role_lower():
    return (session.get("rol") or "").lower().strip()

def _is_super():
    return _role_lower() == "admin" or bool(session.get("is_admin"))

def _can_approve_area(area: str) -> bool:
    r = _role_lower()
    a = (area or "").lower().strip()
    if _is_super():
        return True
    if a == "gg":
        return r == "gerente general"
    if a == "gf":
        return r == "gerente financiero"
    if a == "ga":
        return r in ("gerente", "gerente de área", "gerente de area")
    return False

 

 

def _rol_norm(x):
    return (x or "").strip().lower()

def _can_gastos(action: str) -> bool:
    """
    Permiso efectivo para 'gastos_tarjeta' sin depender de security.py.
    - admin / is_admin: TRUE
    - si existe session['permissions']: usa eso
    - fallback: intenta has_permission / has_perm con distintas firmas
    """
    rol = _rol_norm(session.get("rol"))
    if rol == "admin" or bool(session.get("is_admin")):
        return True

    # 1) Si tienes permisos precargados en sesión (tu app.py los arma con build_permissions)
    perms = session.get("permissions") or {}
    # soporta: {'gastos_tarjeta': {'aprobar': True, ...}}
    try:
        if bool(perms.get("gastos_tarjeta", {}).get(action)):
            return True
        # a veces viene como "Gastos Tarjeta" o variaciones
        if bool(perms.get("Gastos Tarjeta", {}).get(action)):
            return True
    except Exception:
        pass

    # 2) Fallback: intenta funciones globales si existen en este módulo
    # has_permission puede ser (modulo, accion) o (rol, modulo, accion)
    try:
        return bool(has_permission("gastos_tarjeta", action))  # type: ignore
    except TypeError:
        try:
            return bool(has_permission(session.get("rol"), "gastos_tarjeta", action))  # type: ignore
        except Exception:
            pass
    except Exception:
        pass

    # has_perm a veces es (modulo, action) o (opcion, action)
    try:
        return bool(has_perm("gastos_tarjeta", action))  # type: ignore
    except Exception:
        pass

    return False


def _aprobar_gasto_core(cur, conn, gasto_id, area, value, uid, rol):
    # Traer metadatos necesarios
    cur.execute(f"""
        SELECT g.sap_contabilizacion,
               g.usuario_id,
               COALESCE(g.gg_aprobado, 0) AS gg_aprobado,
               COALESCE(g.gf_aprobado, 0) AS gf_aprobado
        FROM {TABLE_GASTOS} g
        WHERE g.id = ?
    """, (gasto_id,))
    meta = cur.fetchone()
    if not meta:
        return False, "Gasto no encontrado"

    if (meta["sap_contabilizacion"] or "").strip():
        return False, "No se puede modificar: ya tiene Doc. SAP."

    # Reglas de bloqueo
    if area == "gf" and int(meta["gg_aprobado"] or 0) == 1:
        return False, "Bloqueado: Gerente General ya aprobó."

    if area == "ga" and rol in ("gerente", "gerente de área", "gerente de area"):
        subordinados = obtener_subordinados(conn, uid)
        if meta["usuario_id"] not in subordinados:
            return False, "No puede aprobar gastos fuera de su jerarquía."

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if area == "gg":
        cur.execute(f"""UPDATE {TABLE_GASTOS}
                        SET gg_aprobado=?, gg_aprobado_por=?, gg_aprobado_at=?
                        WHERE id=?""",
                    (value, uid, now if value else None, gasto_id))
    elif area == "gf":
        cur.execute(f"""UPDATE {TABLE_GASTOS}
                        SET gf_aprobado=?, gf_aprobado_por=?, gf_aprobado_at=?
                        WHERE id=?""",
                    (value, uid, now if value else None, gasto_id))
 

    else:  # ga
        cur.execute(f"""UPDATE {TABLE_GASTOS}
                        SET ga_aprobado=?, ga_aprobado_por=?, ga_aprobado_at=?
                        WHERE id=?""",
                    (value, uid, now if value else None, gasto_id))

    return True, None
def parse_sri_xml(raw: bytes | str):
    """
    Devuelve (header_dict, [detalles_dict]) a partir de un XML de SRI.

    Soporta:
      - XML 'bonito': <autorizacion><comprobante>...&lt;factura&gt;...</comprobante>
      - XML 'roto' del SRI donde el cierre </comprobante> viene escapado
        y nunca se cierra realmente la etiqueta <comprobante>.
    """

    # 1) raw -> texto, limpiando BOM
    if isinstance(raw, bytes):
        text = raw.decode("utf-8-sig", errors="ignore")
    else:
        text = raw

    text = text.replace("\ufeff", "").strip()

    # Helper: sacar la factura interna desde el bloque <comprobante>
    def _extract_inner_from_comprobante(comp_block: str) -> ET.Element:
        inner_xml = html.unescape(comp_block.strip())
        inner_xml = re.sub(r"<\?xml[^>]+\?>", "", inner_xml)
        inner_xml = inner_xml.replace("\ufeff", "").strip()

        start = inner_xml.find("<factura")
        if start == -1:
            start = inner_xml.find("<notaCredito")
        if start == -1:
            start = inner_xml.find("<notaDebito")
        if start > 0:
            inner_xml = inner_xml[start:]

        for tag in ("factura", "notaCredito", "notaDebito"):
            end_marker = f"</{tag}>"
            end = inner_xml.rfind(end_marker)
            if end != -1:
                inner_xml = inner_xml[: end + len(end_marker)]
                break

        return ET.fromstring(inner_xml)

    # 2) Intentar parsear el XML completo
    numero_aut = None
    fecha_aut = None

    try:
        root = ET.fromstring(text)
    except ET.ParseError as e:
        if "mismatched tag" in str(e):
            start = text.find("<comprobante>")
            if start == -1:
                raise
            start += len("<comprobante>")

            end = text.rfind("</mensajes>")
            if end == -1:
                end = text.rfind("</autorizacion>")
                if end == -1:
                    end = len(text)

            comp_block = text[start:end]
            inner = _extract_inner_from_comprobante(comp_block)
        else:
            raise
    else:
        tag = root.tag.split("}", 1)[-1]

        if tag == "autorizacion":
            numero_aut = (root.findtext("numeroAutorizacion") or "").strip()
            fecha_aut = (root.findtext("fechaAutorizacion") or "").strip()
            comp_str = root.findtext("comprobante") or ""
            inner = _extract_inner_from_comprobante(comp_str)
        else:
            inner = root

    # 3) A partir de la factura interna, armar header y detalles
    def tag_name(el: ET.Element) -> str:
        return el.tag.split("}", 1)[-1]

    doc_tag = tag_name(inner)

    infoTrib = inner.find("./infoTributaria")
    if infoTrib is None:
        raise ValueError("XML sin infoTributaria")

    cod_doc = (infoTrib.findtext("codDoc") or "").strip()
    tipo = {
        "01": "FACTURA",
        "04": "NOTA_CREDITO",
        "05": "NOTA_DEBITO",
    }.get(cod_doc, doc_tag.upper())

    if doc_tag == "factura":
        info = inner.find("./infoFactura")
        detalles_root = inner.find("./detalles")
    elif doc_tag == "notaDebito":
        info = inner.find("./infoNotaDebito")
        detalles_root = inner.find("./motivos")
    else:
        info = inner.find("./infoFactura") or inner.find("./infoNotaDebito")
        detalles_root = inner.find("./detalles") or inner.find("./motivos")

    if info is None:
        raise ValueError("XML sin infoFactura/infoNotaDebito")

    h = {
        "clave_acceso": (infoTrib.findtext("claveAcceso") or "").strip(),
        "numero_autorizacion": numero_aut,
        "cod_doc": cod_doc,
        "tipo_comprobante": tipo,
        "fecha_emision": (info.findtext("fechaEmision") or "").strip(),
        "fecha_autorizacion": fecha_aut,
        "ruc_emisor": (infoTrib.findtext("ruc") or "").strip(),
        "razon_social_emisor": (infoTrib.findtext("razonSocial") or "").strip(),
        "ruc_cliente": (info.findtext("identificacionComprador") or "").strip(),
        "razon_social_cliente": (info.findtext("razonSocialComprador") or "").strip(),
        "estab": (infoTrib.findtext("estab") or "").strip(),
        "pto_emi": (infoTrib.findtext("ptoEmi") or "").strip(),
        "secuencial": (infoTrib.findtext("secuencial") or "").strip(),
        "subtotal": _to_float(info.findtext("totalSinImpuestos")),
        "descuento": _to_float(info.findtext("totalDescuento")),
        "moneda": (info.findtext("moneda") or "").strip(),
        
    }
    # ✅ Propina (viene en infoFactura)
    h["propina"] = _to_float(info.findtext("propina"))  # ej: 33.62
    # opcional: si tu sistema maneja esto como “servicios_10”
    h["servicios_10"] = h["propina"]


    iva_total = 0.0
    base_iva = 0.0
    base_0 = 0.0
    base_15 = 0.0
    tarifas_detectadas = set()

    MAP_IVA_PORC = {
        "0": 0.0,
        "2": 12.0,
        "3": 14.0,
        "4": 15.0,
        "8": 8.0,   # ✅ IVA 8%
    }


    for ti in info.findall("./totalConImpuestos/totalImpuesto"):
        codigo = (ti.findtext("codigo") or "").strip()
        if codigo != "2":  # IVA
            continue

        # 1) intenta leer tarifa si existiera
        tarifa = _to_float(ti.findtext("tarifa"))

        # 2) si no existe, usa codigoPorcentaje
        if not tarifa:
            cp = (ti.findtext("codigoPorcentaje") or "").strip()
            tarifa = MAP_IVA_PORC.get(cp, 0.0)

        base = _to_float(ti.findtext("baseImponible"))
        valor = _to_float(ti.findtext("valor"))

        tarifas_detectadas.add(tarifa)
        iva_total += valor

        if tarifa and tarifa > 0:
            base_iva += base
            if abs(tarifa - 15.0) < 0.01:
                base_15 += base
        else:
            base_0 += base

    h["iva"] = iva_total
    h["base_iva"] = base_iva
    h["subtotal_0"] = base_0
    h["subtotal_15"] = base_15  # ✅ CLAVE para que no quede NULL
    h["iva_tarifa"] = max(tarifas_detectadas) if tarifas_detectadas else 0.0
    h["total"] = _to_float(info.findtext("importeTotal")) or (
        (h["subtotal"] or 0) + (h["iva"] or 0) + (h.get("propina") or 0)
    )

    # Total (si existe en el XML úsalo; si no, fallback simple)
    h["total"] = _to_float(info.findtext("importeTotal")) or (h["subtotal"] + h["iva"])

    detalles = []

    if detalles_root is not None:
        if doc_tag == "factura":
            for d in detalles_root.findall("./detalle"):
                impuestos = d.findall("./impuestos/impuesto")

                iva_line = 0.0
                base_line = 0.0

                for i in impuestos:
                    codigo = (i.findtext("codigo") or "").strip()
                    if codigo != "2":
                        continue
                    iva_line += _to_float(i.findtext("valor"))
                    base_line += _to_float(i.findtext("baseImponible"))

                precio_sin_imp = _to_float(d.findtext("precioTotalSinImpuesto"))

                detalles.append({
                    "codigo_principal": (d.findtext("codigoPrincipal") or "").strip(),
                    "descripcion": (d.findtext("descripcion") or "").strip(),
                    "cantidad": _to_float(d.findtext("cantidad")),
                    "precio_unitario": _to_float(d.findtext("precioUnitario")),
                    "descuento": _to_float(d.findtext("descuento")),
                    "base_imponible": base_line or precio_sin_imp,
                    "iva": iva_line,
                    "total_linea": precio_sin_imp + iva_line,
                })

        elif doc_tag == "notaDebito":
            for m in detalles_root.findall("./motivo"):
                valor = _to_float(m.findtext("valor"))
                detalles.append({
                    "codigo_principal": "",
                    "descripcion": (m.findtext("razon") or "").strip(),
                    "cantidad": 1,
                    "precio_unitario": valor,
                    "descuento": 0.0,
                    "base_imponible": valor,
                    "iva": 0.0,
                    "total_linea": valor,
                })

    if not h["clave_acceso"]:
        raise ValueError("No se encontró claveAcceso en el XML")

    return h, detalles


def guess_gerente_area(conn: sqlite3.Connection, user_id: int | None) -> int | None:
    gerente = get_ultimo_jefe_activo(conn, user_id)
    if gerente:
        return gerente

    # Fallback antiguo por rol en departamento
    cur = conn.cursor()
    cur.execute("SELECT departamento_id FROM usuarios WHERE id = ?", (user_id,))
    u = cur.fetchone()
    if not u or not u["departamento_id"]:
        return None
    BOSS_ROLES = (
        'jefe',
        'gerente',
        'gerente general',
        'gerente financiero',
        'coordinador',
        'admin','usuario'
    )
    
    depto_id = u["departamento_id"]
    roles = [r.lower() for r in BOSS_ROLES]

    cur.execute(f"""
        SELECT TOP 1 id
        FROM usuarios
        WHERE departamento_id = ?
          AND LOWER(rol) IN ({",".join(["?"] * len(roles))})
          AND COALESCE(disabled, 0) = 0
        ORDER BY id
     """, (depto_id, *roles))

    row = cur.fetchone()
    return row["id"] if row else None



def get_ultimo_jefe_activo(conn: sqlite3.Connection, user_id: int | None) -> int | None:
    """
    Retorna el último jefe activo en la cadena de usuarios.jefe_id.
    Si hay bucles, se detiene por seguridad.
    """
    if not user_id:
        return None

    cur = conn.cursor()

    # traer jefe inicial
    cur.execute("SELECT jefe_id FROM usuarios WHERE id = ?", (user_id,))
    row = cur.fetchone()
    if not row:
        return None

    jefe_id = row["jefe_id"]
    if not jefe_id:
        return None

    seen = set()
    last_valid = None

    while jefe_id and jefe_id not in seen:
        seen.add(jefe_id)

        cur.execute("""
            SELECT id, jefe_id
            FROM usuarios
            WHERE id = ?
              AND COALESCE(disabled, 0) = 0
        """, (jefe_id,))
        j = cur.fetchone()

        if not j:
            break

        last_valid = j["id"]
        jefe_id = j["jefe_id"]

    return last_valid


def sanitize_sri_xml(text: str) -> str:
    """
    Arregla XML de SRI donde:
      - hay <autorizacion><comprobante> ... &lt;/comprobante&gt; ... </autorizacion>
      - pero NO existe el cierre real </comprobante>
    """
    # Si no es el wrapper de autorizacion, lo devolvemos tal cual
    if "<autorizacion" not in text:
        return text

    # Si ya tiene cierre real, no tocamos nada
    if "</comprobante>" in text:
        return text

    # Si ni siquiera hay <comprobante>, no tocamos nada
    if "<comprobante>" not in text:
        return text

    start_tag = "<comprobante>"
    start = text.find(start_tag)

    # Buscamos el cierre escapado
    end_marker = "&lt;/comprobante&gt;"
    end = text.find(end_marker, start)
    if end == -1:
        # No encontramos el patrón, mejor no tocar
        return text

    # Contenido escapado dentro de <comprobante> ... &lt;/comprobante&gt;
    inner_esc = text[start + len(start_tag):end]

    # Lo convertimos a XML real (<factura>, <notaDebito>, etc.)
    inner_xml = html.unescape(inner_esc.strip())

    # Lo envolvemos en CDATA para que sea texto seguro en el XML de autorizacion
    cdata = "<![CDATA[" + inner_xml + "]]>"

    # Armamos XML corregido
    new_text = (
        text[:start] +
        start_tag + cdata + "</comprobante>" +
        text[end + len(end_marker):]
    )
    return new_text

 

def _to_float(txt: str | None) -> float:
    if not txt:
        return 0.0
    t = txt.replace(',', '.').strip()
    try:
        return float(t)
    except ValueError:
        return 0.0

def _fecha_sql(col='fecha_emision'):
    return (
        f"TRY_CONVERT(date, "
        f"SUBSTRING({col}, 7, 4) + '-' + SUBSTRING({col}, 4, 2) + '-' + SUBSTRING({col}, 1, 2))"
    )

def _extraer_error_sap(resp):
    """
    Busca mensajes de error en la respuesta de SAP.
    Considera tipo 'E' (Error) o 'A' (Abort).
    Devuelve el texto del mensaje o None si no hay error.
    """
    if not isinstance(resp, (dict, list)):
        return None

    items = resp if isinstance(resp, list) else [resp]

    for item in items:
        if not isinstance(item, dict):
            continue
        tipo = (item.get("tipo") or item.get("Tipo") or "").strip().upper()
        if tipo in ("E", "A"):
            msg = (
                item.get("mensaje")
                or item.get("Mensaje")
                or item.get("message")
                or "Error devuelto por SAP"
            )
            return msg

    return None



# helpers_org.py (o dentro de routes_gastos_tarjetas.py si aún no separas)
def obtener_gerente_real(conn, user_id):
    """
    Sube por la cadena jerárquica usando jefe_id.
    En cada nivel busca si existe un usuario con rol 'gerente' 
    en el departamento del jefe actual.
    Si lo encuentra -> retorna el gerente.
    Si no existe gerente en ningún nivel -> retorna el último jefe.
    Si no hay jefes -> retorna None.
    """

    gerente_roles = ('gerente', 'gerente de área', 'gerente area')
    MAX_NIVELES = 10  # evita loops por datos corruptos
    
    print("\n=== INICIO BÚSQUEDA DE GERENTE PARA USER:", user_id, "===")

    # obtener usuario inicial
    usuario = conn.execute("""
        SELECT id, rol, departamento_id, jefe_id
        FROM usuarios
        WHERE id = ?
    """, (user_id,)).fetchone()

    if not usuario:
        print("Usuario no existe en BD.")
        return None

    jefe_id = usuario["jefe_id"]
    ultimo_jefe = None
    nivel = 1

    while jefe_id and nivel <= MAX_NIVELES:

        print(f"\n-- NIVEL {nivel}: jefe_id = {jefe_id}")

        # obtener datos del jefe en este nivel
        jefe = conn.execute("""
            SELECT id, rol, departamento_id, jefe_id
            FROM usuarios
            WHERE id = ?
        """, (jefe_id,)).fetchone()

        if not jefe:
            print("No existe registro del jefe:", jefe_id)
            break

        print("  Jefe encontrado:", dict(jefe))

        ultimo_jefe = jefe  # por si no encontramos gerente

        depto_jefe = jefe["departamento_id"]
        print("  Departamento del jefe:", depto_jefe)

        # buscar gerente del departamento
        placeholders = ",".join("?" for _ in gerente_roles)

        gerente = conn.execute(f"""
            SELECT top 1 id, rol, departamento_id, jefe_id
            FROM usuarios
            WHERE departamento_id = ?
              AND lower(rol) IN ({placeholders})
            
        """, (depto_jefe, *[r.lower() for r in gerente_roles])).fetchone()

        if gerente:
            print("✔ GERENTE ENCONTRADO EN ESTE NIVEL:", dict(gerente))
            return gerente

        # subir otro nivel
        jefe_id = jefe["jefe_id"]
        nivel += 1

    # si salimos del loop sin gerente:
    if ultimo_jefe:
        print("⚠ No se encontró gerente, se retorna el último jefe:", dict(ultimo_jefe))
        return ultimo_jefe

    print("⚠ No existe jefe ni gerente para este usuario.")
    return None


def obtener_subordinados(conn, jefe_id):
    """
    Retorna TODOS los subordinados (directos e indirectos)
    de un jefe usando recorrido recursivo.
    """
    result = set()
    pendientes = [jefe_id]

    cur = conn.cursor()

    while pendientes:
        actual = pendientes.pop()

        cur.execute("""
            SELECT id
            FROM usuarios
            WHERE jefe_id = ?
        """, (actual,))
        hijos = [row["id"] for row in cur.fetchall()]

        for h in hijos:
            if h not in result:
                result.add(h)
                pendientes.append(h)

    return list(result)

 
def register_gastos_routes(app):
    @app.post('/reembolsos/xml/seedbilling-consumir', endpoint='facturas_xml_seedbilling_consumir')
    @require_login
    @require_permission('facturas_xml_list', 'ver')
    def facturas_xml_seedbilling_consumir():
        """
        Consumo manual del API SeedBilling desde la pantalla Facturas XML.
        Ejecuta el mismo proceso automático del worker.
        """
        from flask import jsonify, current_app
        from modules.scheduler.seedbilling_xml_job import process_seedbilling_facturas_recibidas
        from modules.scheduler.scheduler_repository import get_db_standalone

        conn = None

        try:
            conn = get_db_standalone()
            result = process_seedbilling_facturas_recibidas(conn)

            current_app.logger.info(
                "[SEEDBILLING_MANUAL] resultado=%s",
                result
            )

            return jsonify(
                ok=True,
                msg=(
                    "Consumo SeedBilling finalizado. "
                    f"Recibidos: {result.get('recibidos', 0)} | "
                    f"Quimpac: {result.get('quimpac', 0)} | "
                    f"Insertados: {result.get('insertados', 0)} | "
                    f"Duplicados: {result.get('duplicados', 0)} | "
                    f"Otras empresas omitidas: {result.get('otras_empresas', 0)} | "
                    f"Marcados Quimpac: {result.get('marcados_entregados_quimpac', 0)} | "
                    f"Marcados otras empresas: {result.get('marcados_entregados_otras', 0)} | "
                    f"Errores: {result.get('errores', 0)}"
                ),
                result=result
            )

        except Exception as e:
            current_app.logger.exception("[SEEDBILLING_MANUAL] error")
            return jsonify(
                ok=False,
                msg=f"Error consumiendo SeedBilling: {e}"
            ), 500

        finally:
            try:
                if conn:
                    conn.close()
            except Exception:
                pass


   # ==========================================================
    # ENDPOINT
    # ==========================================================
    @app.route('/reembolsos/gastos/export/reporte.xlsx', methods=['GET'], endpoint='export_gastos_reporte_excel')
    @require_login
    @require_permission('gastos_tarjeta', 'exportar')
    def export_gastos_reporte_excel():
        from datetime import datetime
        from io import BytesIO
        from flask import send_file, request, session, current_app
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        conn = get_db()
        try:
            cur = conn.cursor()

            role_name = (session.get('rol') or '').lower().strip()
            uid = session.get('usuario_id') or session.get('user_id')
            is_admin = (role_name == "admin") or bool(session.get("is_admin"))

            # ==========================================================
            # 0) Si viene ids desde el FRONT -> priorizar
            # ==========================================================
            ids_raw = (request.args.get("ids") or "").strip()
            ids_req = [int(x) for x in ids_raw.split(",") if x.strip().isdigit()]

            # ==========================================================
            # 1) Filtros SOLO por request.args
            # ==========================================================
            filtros = {}
            where = []
            args = []

            where.append("COALESCE(g.inactivo,0)=0")

            # ---- fechas
            desde = (request.args.get("desde") or "").strip()
            hasta = (request.args.get("hasta") or "").strip()
            filtros["desde"] = desde
            filtros["hasta"] = hasta

            if desde:
                where.append("CAST(g.fecha AS date) >= CAST(? AS date)")
                args.append(desde)

            if hasta:
                where.append("CAST(g.fecha AS date) <= CAST(? AS date)")
                args.append(hasta)

            # ---- proveedor_id
            proveedor_id = (request.args.get("proveedor_id") or "").strip()
            filtros["proveedor_id"] = proveedor_id
            filtros["proveedor"] = (request.args.get("proveedor") or "").strip()

            if proveedor_id.isdigit():
                where.append("g.proveedor_id = ?")
                args.append(int(proveedor_id))

            # ---- descripción
            descripcion = (request.args.get("descripcion") or "").strip()
            filtros["descripcion"] = descripcion

            if descripcion:
                where.append("(LOWER(g.motivo) LIKE ?)")
                args.append(f"%{descripcion.lower()}%")

            # ---- CCB
            ccb_req = (request.args.get("ccb") or "").strip()
            if ccb_req in ("0", "1"):
                where.append("COALESCE(g.ccb,0)=?")
                args.append(int(ccb_req))
                filtros["ccb"] = ccb_req
            else:
                filtros["ccb"] = ""

            # ---- Tipo
            tipo = (request.args.get("tipo") or "").strip().lower()
            TIPOS_RESTRINGIDOS = ("caja_chica", "reembolso")

            if tipo == "caja_chica":
                where.append("COALESCE(g.es_caja_chica,0)=1")
            elif tipo == "reembolso":
                where.append("COALESCE(g.reembolso_vendedor,0)=1 AND COALESCE(g.es_caja_chica,0)=0")
            elif tipo == "tarjeta":
                where.append("COALESCE(g.es_caja_chica,0)=0 AND COALESCE(g.reembolso_vendedor,0)=0")
            elif tipo == "tarjeta_online":
                where.append("""
                    COALESCE(g.es_caja_chica,0)=0
                    AND COALESCE(g.reembolso_vendedor,0)=0
                    AND COALESCE(g.tarjeta_sin_soporte,0)=1
                """)
            elif tipo == "tarjeta_boletos":
                where.append("""
                    COALESCE(g.es_caja_chica,0)=0
                    AND COALESCE(g.reembolso_vendedor,0)=0
                    AND COALESCE(g.boletos_aereos,0)=1
                """)

            filtros["tipo"] = tipo

            # ---- pendientes
            pendientes = (request.args.get("pendientes") or request.args.get("pend") or "").strip()
            filtros["pendientes"] = pendientes

            if pendientes == "1":
                if tipo in TIPOS_RESTRINGIDOS:
                    where.append("""
                        COALESCE(g.ga_aprobado,0)=0
                        AND COALESCE(g.gg_aprobado,0)=0
                        AND COALESCE(g.gf_aprobado,0)=0
                    """)
                else:
                    where.append("""
                        (
                            COALESCE(g.ga_aprobado,0)=0
                            OR COALESCE(g.gg_aprobado,0)=0
                            OR COALESCE(g.gf_aprobado,0)=0
                        )
                    """)

            # ---- usuario_id
            usuario_id = (request.args.get("usuario_id") or "").strip()
            if usuario_id.isdigit():
                where.append("g.usuario_id = ?")
                args.append(int(usuario_id))
                filtros["usuario_id"] = int(usuario_id)
            else:
                filtros["usuario_id"] = ""

            # ---- gerente_id
            gerente_id_raw = (request.args.get("gerente_id") or "").strip()
            if gerente_id_raw.isdigit():
                gerente_id_sel = int(gerente_id_raw)
                filtros["gerente_id"] = gerente_id_sel
            else:
                gerente_id_sel = None
                filtros["gerente_id"] = ""

            # ==========================================================
            # Scope por empresa para export
            # ==========================================================
            _apply_empresa_scope(conn, where, args, user_alias="u", uid=uid)

            current_app.logger.warning("[EXPORT_REPORTE_EXCEL_FILTERS] args_url=%s", dict(request.args))
            current_app.logger.warning("[EXPORT_REPORTE_EXCEL_FILTERS] filtros=%s", filtros)

            # ==========================================================
            # 2) Obtener IDs a exportar
            # ==========================================================
            gasto_ids = []

            base_from_ids = f"""
                FROM {TABLE_GASTOS} g
                LEFT JOIN terceros t ON t.id = g.proveedor_id
                LEFT JOIN usuarios u ON u.id = g.usuario_id
            """

            if ids_req:
                placeholders_ids = ",".join("?" * len(ids_req))

                sql_ids = f"""
                    SELECT
                        g.id AS gasto_id,
                        CAST(g.fecha AS date) AS fecha_ord
                    {base_from_ids}
                    WHERE COALESCE(g.inactivo,0)=0
                    AND g.id IN ({placeholders_ids})
                """

                ids_args = list(ids_req)

                if not is_admin:
                    empresa_id = _get_session_empresa_id(conn, uid)
                    if empresa_id:
                        sql_ids += " AND COALESCE(u.empresa_id,-1)=?"
                        ids_args.append(empresa_id)
                    else:
                        sql_ids += " AND 1=0"

                sql_ids += " ORDER BY fecha_ord DESC, gasto_id DESC"

                cur.execute(sql_ids, ids_args)
                gasto_ids = [r["gasto_id"] for r in cur.fetchall()]

            else:
                if isinstance(gerente_id_sel, int):
                    sub = obtener_subordinados(conn, gerente_id_sel) or []
                    scope = set(int(x) for x in sub if str(x).isdigit())
                    scope.add(gerente_id_sel)

                    if scope:
                        placeholders = ",".join("?" * len(scope))
                        where.append(f"g.usuario_id IN ({placeholders})")
                        args.extend(list(scope))
                    else:
                        where.append("1=0")

                sql_ids = f"""
                    SELECT
                        g.id AS gasto_id,
                        CAST(g.fecha AS date) AS fecha_ord
                    {base_from_ids}
                """

                if where:
                    sql_ids += " WHERE " + " AND ".join(where)

                sql_ids += " ORDER BY fecha_ord DESC, gasto_id DESC"

                current_app.logger.warning("[EXPORT_REPORTE_EXCEL_IDS] SQL=%s", sql_ids)
                current_app.logger.warning("[EXPORT_REPORTE_EXCEL_IDS] ARGS=%s", args)

                cur.execute(sql_ids, args)
                gasto_ids = [r["gasto_id"] for r in cur.fetchall()]

            current_app.logger.warning(
                "[EXPORT_REPORTE_EXCEL] ids_count=%s ids_sample=%s",
                len(gasto_ids), gasto_ids[:20]
            )

            # ==========================================================
            # 3) Traer cabecera + detalle
            # ==========================================================
            rows = []

            if gasto_ids:
                placeholders = ",".join("?" * len(gasto_ids))

                base_from = f"""
                    FROM {TABLE_GASTOS} g
                    LEFT JOIN terceros t ON t.id = g.proveedor_id
                    LEFT JOIN usuarios u ON u.id = g.usuario_id
                    LEFT JOIN gastos_tarjeta_detalle d ON d.gasto_id = g.id
                """

                select_cols = """
                    g.id AS gasto_id,
                    g.fecha,
                    g.anio, g.mes, g.dia,
                    COALESCE(NULLIF(LTRIM(RTRIM(g.motivo)), ''), '') AS detalle_general,
                    COALESCE(g.ccb, 0) AS ccb,
                    COALESCE(t.identificacion, '') AS proveedor_identificacion,
                    COALESCE(t.nombre, g.proveedor, '') AS proveedor_nombre,
                    COALESCE(g.numero_factura, '') AS numero_factura,
                    COALESCE(g.orden_compra, '') AS orden_compra,

                    COALESCE(g.subtotal_factura, 0) AS subtotal_factura,
                    COALESCE(g.servicios_10, 0) AS servicios_10,
                    COALESCE(g.iva, 0) AS iva_cab,
                    COALESCE(g.total_con_iva, 0) AS total_cab,

                    COALESCE(d.observacion, '') AS det_observacion,
                    COALESCE(d.motivo, '') AS det_motivo_gasto,
                    COALESCE(d.centro_costo, '') AS det_centro_costo,
                    COALESCE(d.indicador, '') AS det_indicador,
                    COALESCE(d.subtotal_factura, 0) AS det_subtotal,
                    COALESCE(d.servicios_10, 0) AS det_servicios,
                    COALESCE(d.iva, 0) AS det_iva,
                    COALESCE(d.total_con_iva, 0) AS det_total,
                    COALESCE(d.descripcion, '') AS det_descripcion,

                    COALESCE(u.username, '') AS usuario_creador,
                    '' AS gerente_real,

                    g.usuario_id,
                    COALESCE(g.ga_aprobado, 0) AS ga_aprobado,
                    COALESCE(g.gg_aprobado, 0) AS gg_aprobado,
                    COALESCE(g.gf_aprobado, 0) AS gf_aprobado,
                    COALESCE(g.sap_contabilizacion, '') AS sap_contabilizacion,
                    COALESCE(g.boletos_aereos, 0) AS boletos_aereos,

                    CASE
                        WHEN COALESCE(g.es_caja_chica,0)=1 THEN 'caja_chica'
                        WHEN COALESCE(g.reembolso_vendedor,0)=1 THEN 'reembolso'
                        WHEN COALESCE(g.boletos_aereos,0)=1 THEN 'tarjeta_boletos'
                        WHEN COALESCE(g.tarjeta_sin_soporte,0)=1 THEN 'tarjeta_online'
                        ELSE 'tarjeta'
                    END AS tipo_gasto
                """

                sql = f"""
                    SELECT {select_cols}
                    {base_from}
                    WHERE g.id IN ({placeholders})
                    ORDER BY CAST(g.fecha AS date) DESC, g.id DESC, COALESCE(d.id,0) ASC
                """

                cur.execute(sql, gasto_ids)
                rows = [dict(r) for r in cur.fetchall()]

            # ==========================================================
            # 4) gerente_real
            # ==========================================================
            gerente_cache = {}

            for r in rows:
                uid_row = r.get("usuario_id")
                if not uid_row:
                    r["gerente_real"] = ""
                    continue

                try:
                    uid_row_int = int(uid_row)
                except Exception:
                    r["gerente_real"] = ""
                    continue

                if uid_row_int not in gerente_cache:
                    try:
                        gerente_cache[uid_row_int] = _get_gerente_username_real(conn, uid_row_int) or ""
                    except Exception:
                        gerente_cache[uid_row_int] = ""

                r["gerente_real"] = gerente_cache[uid_row_int]

            # ==========================================================
            # 5) Construir Excel
            # ==========================================================
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte"

            ws["A1"] = "REPORTE — Gastos con tarjeta (Detalle por línea)"
            ws["A1"].font = Font(bold=True, size=14)

            ws["A2"] = "Fecha de generación:"
            ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            ws["A3"] = "Usuario de generación:"
            ws["B3"] = session.get("usuario") or session.get("username") or "—"

            ws["A4"] = "Filtros:"
            filtros_txt = []

            for k, v in (filtros or {}).items():
                if v in (None, "", [], {}):
                    continue
                filtros_txt.append(f"{k}={v}")

            if ids_req:
                filtros_txt.append(f"ids={len(ids_req)} (visibles)")

            ws["B4"] = " | ".join(filtros_txt) if filtros_txt else "—"

            start_row = 6

            cols = [
                ("fecha", "Fecha"),
                ("anio", "Año"),
                ("mes", "Mes"),
                ("dia", "Día"),
                ("tipo_gasto", "Tipo"),
                ("detalle_general", "Detalle general"),
                ("ccb", "CCB"),
                ("proveedor_identificacion", "Identificación proveedor"),
                ("proveedor_nombre", "Proveedor"),
                ("numero_factura", "N° factura"),
                ("orden_compra", "N° Orden de compra"),

                ("det_observacion", "Observación (línea)"),
                ("det_motivo_gasto", "Motivo gasto (línea)"),
                ("det_centro_costo", "Centro de costo (línea)"),
                ("det_indicador", "Indicador (línea)"),
                ("det_subtotal", "Subtotal fac. (línea)"),
                ("det_servicios", "Serv./prop. (línea)"),
                ("det_iva", "IVA (línea)"),
                ("det_total", "Total con IVA (línea)"),
                ("det_descripcion", "Descripción (línea)"),

                ("subtotal_factura", "Subtotal factura (cab)"),
                ("servicios_10", "Servicios y propinas (cab)"),
                ("iva_cab", "IVA (cab)"),
                ("total_cab", "Total (cab)"),

                ("usuario_creador", "Usuario creador"),
                ("gerente_real", "Gerente"),
                ("ga_aprobado", "Aprob. GA"),
                ("gg_aprobado", "Aprob. GG"),
                ("gf_aprobado", "Aprob. GF"),
                ("sap_contabilizacion", "Doc. SAP"),
                ("gasto_id", "ID Gasto"),
            ]

            for c, (_, label) in enumerate(cols, start=1):
                cell = ws.cell(row=start_row, column=c, value=label)
                cell.font = Font(bold=True)

            rr = start_row + 1

            for row in rows:
                for c, (key, _) in enumerate(cols, start=1):
                    ws.cell(row=rr, column=c, value=row.get(key, ""))
                rr += 1

            ws.freeze_panes = ws[f"A{start_row + 1}"]
            ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(cols))}{max(start_row, rr - 1)}"

            for i in range(1, len(cols) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 18

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)

            return send_file(
                bio,
                as_attachment=True,
                download_name="reporte_gastos_tarjeta_detallado.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        finally:
            try:
                conn.close()
            except Exception:
                pass
  
  
    # LISTA
    from flask import render_template, request, session, redirect, url_for, current_app
    from datetime import date

    from flask import render_template, request, session, current_app, redirect, url_for
    from datetime import date

    from flask import render_template, request, session, current_app, redirect, url_for
    from datetime import date

    @app.route('/reembolsos/gastos', methods=['GET'], endpoint='lista_gastos')
    @require_permission('gastos_tarjeta', 'ver')
    @require_login
    def lista_gastos():
        conn = get_db()
        cur = conn.cursor()

        def _row_to_dict(row, columns=None):
            if row is None:
                return None
            try:
                return dict(row)
            except Exception:
                if columns:
                    return {columns[i]: row[i] for i in range(len(columns))}
                return None

        try:
            role_name = (session.get('rol') or '').lower().strip()
            uid = session.get('usuario_id') or session.get('user_id')

            try:
                cur.execute("""
                    SELECT id, nombre, identificacion
                    FROM terceros
                    WHERE UPPER(LTRIM(RTRIM(tipo))) = 'P'
                    AND COALESCE(activo, 1) = 1
                    ORDER BY nombre
                """)
                proveedores = cur.fetchall()
            except Exception:
                proveedores = []

            # ==========================================================
            # EVITAR QUE "pendientes=1" SE QUEDE PEGADO AL DAR "BUSCAR"
            # ==========================================================
            accion = (request.args.get("accion") or "").lower().strip()

            def _redirect_without_pendientes():
                q = request.args.to_dict(flat=True)
                q.pop("pendientes", None)
                q.pop("pend", None)
                q.pop("pend_view", None)
                q.pop("accion", None)
                return redirect(url_for("lista_gastos", **q))

            if accion == "buscar" and ((request.args.get("pendientes") == "1") or (request.args.get("pend") == "1")):
                return _redirect_without_pendientes()

            # ==========================================================
            # BOTÓN "PENDIENTES": forzar MES + activar pendientes=1
            # ==========================================================
            def _redirect_pendientes_mes():
                hoy = date.today()
                inicio_mes = hoy.replace(day=1)

                q = request.args.to_dict(flat=True)
                q.pop("pend_view", None)
                q.pop("accion", None)
                q.pop("desde", None)
                q.pop("hasta", None)

                q["pendientes"] = "1"
                q["desde"] = inicio_mes.isoformat()
                q["hasta"] = hoy.isoformat()

                return redirect(url_for("lista_gastos", **q))

            if request.args.get("pend_view") == "1":
                return _redirect_pendientes_mes()

            # ==========================================================
            # FECHAS POR DEFECTO EN BACKEND
            # - evita que cargue TODO y luego el JS reenvíe
            # ==========================================================
            qs = request.args.to_dict(flat=True)
            desde_req = (qs.get("desde") or "").strip()
            hasta_req = (qs.get("hasta") or "").strip()
            pendientes_req = (qs.get("pendientes") or qs.get("pend") or "").strip()
            special = (qs.get("ccb") == "1") or (pendientes_req == "1")

            if not desde_req and not hasta_req and not special:
                hoy = date.today().isoformat()
                qs["desde"] = hoy
                qs["hasta"] = hoy
                qs.pop("accion", None)
                return redirect(url_for("lista_gastos", **qs))

            # ==========================================================
            # Filtros base existentes
            # ==========================================================
            filtros, where, args, _ = gh.collect_gastos_filters(request, session)

            # ==========================================================
            # Filtro CCB: '' (todos) | '1' (con CCB) | '0' (sin CCB)
            # ==========================================================
            ccb_req = (request.args.get('ccb') or '').strip()
            if ccb_req in ('0', '1'):
                where.append("COALESCE(g.ccb,0)=?")
                args.append(int(ccb_req))
                filtros['ccb'] = ccb_req
            else:
                filtros['ccb'] = ''

            tipo = (request.args.get('tipo') or '').strip().lower()
            tipo_raw = request.args.get('tipo')

            current_app.logger.warning("DEBUG tipo_raw=%r tipo_norm=%r args=%r", tipo_raw, tipo, dict(request.args))

            TIPOS_RESTRINGIDOS = ('caja_chica', 'reembolso')

            if tipo == 'caja_chica':
                where.append("COALESCE(g.es_caja_chica,0)=1")
            elif tipo == 'reembolso':
                where.append("COALESCE(g.reembolso_vendedor,0)=1 AND COALESCE(g.es_caja_chica,0)=0")
            elif tipo == 'tarjeta':
                where.append("COALESCE(g.es_caja_chica,0)=0 AND COALESCE(g.reembolso_vendedor,0)=0")
            elif tipo == 'tarjeta_online':
                where.append("""
                    COALESCE(g.es_caja_chica,0)=0
                    AND COALESCE(g.reembolso_vendedor,0)=0
                    AND COALESCE(g.tarjeta_sin_soporte,0)=1
                """)
            elif tipo == 'tarjeta_boletos':
                where.append("""
                    COALESCE(g.es_caja_chica,0)=0
                    AND COALESCE(g.reembolso_vendedor,0)=0
                    AND COALESCE(g.boletos_aereos,0)=1
                """)

            filtros['tipo'] = tipo

            gerente_id_req = (request.args.get("gerente_id") or "").strip()
            if gerente_id_req.isdigit():
                filtros["gerente_id"] = int(gerente_id_req)
            else:
                filtros["gerente_id"] = ""

            pendientes = (request.args.get("pendientes") or request.args.get("pend") or "").strip()
            current_app.logger.info("ROL_SESSION=%r role_name=%r pendientes=%r", session.get("rol"), role_name, pendientes)

            if pendientes == "1":
                if tipo in TIPOS_RESTRINGIDOS and role_name in ("gerente general", "gerente financiero"):
                    where.append("1=0")
                else:
                    if role_name in ("gerente", "gerente de área", "gerente de area"):
                        where.append("COALESCE(g.ga_aprobado,0)=0")
                    elif role_name == "gerente general":
                        where.append("COALESCE(g.ga_aprobado,0)=1 AND COALESCE(g.gg_aprobado,0)=0")
                    #elif role_name == "gerente financiero":
                        #where.append("COALESCE(g.gg_aprobado,0)=1 AND COALESCE(g.gf_aprobado,0)=0")
                    elif role_name == "gerente financiero":
                        where.append("COALESCE(g.ga_aprobado,0)=1 AND COALESCE(g.gf_aprobado,0)=0")
                    else:
                        where.append("1=0")

            descripcion = (request.args.get('descripcion') or '').strip()
            if descripcion:
                where.append("(g.motivo LIKE ?)")
                args.append(f"%{descripcion}%")
                filtros['descripcion'] = descripcion
            else:
                filtros['descripcion'] = ''

            PRIV_ALL = ('admin', 'coordinador')
            if tipo not in TIPOS_RESTRINGIDOS:
                PRIV_ALL = ('admin', 'coordinador', 'gerente general', 'gerente financiero')

            GERENTE_ROLES = ('gerente', 'gerente de área', 'gerente de area')
            allowed_ids = None

            scope_ids_set = set()
            if role_name in ('gerente general', 'gerente financiero'):
                scope_ids_set = set(obtener_subordinados(conn, uid) or [])
                if uid:
                    scope_ids_set.add(uid)

                placeholders = ",".join("?" * len(scope_ids_set)) if scope_ids_set else "?"
                where.append(
                    f"""(
                        (COALESCE(g.es_caja_chica,0)=0 AND COALESCE(g.reembolso_vendedor,0)=0)
                        OR g.usuario_id IN ({placeholders})
                    )"""
                )
                args.extend(list(scope_ids_set) if scope_ids_set else [uid])
                allowed_ids = scope_ids_set.copy()

            if role_name in PRIV_ALL:
                pass

            elif role_name in ('gerente general', 'gerente financiero'):
                if allowed_ids is None:
                    allowed_ids = {uid}

            elif role_name in GERENTE_ROLES:
                gerente_id = uid
                subordinados = obtener_subordinados(conn, gerente_id) or []
                allowed_ids = set(subordinados)
                if gerente_id:
                    allowed_ids.add(gerente_id)

                if not allowed_ids:
                    allowed_ids = {gerente_id}

                placeholders = ",".join("?" * len(allowed_ids))
                where.append(f"g.usuario_id IN ({placeholders})")
                args.extend(list(allowed_ids))

            else:
                where.append("g.usuario_id = ?")
                args.append(uid)
                allowed_ids = {uid}

            usuario_id = (request.args.get('usuario_id') or '').strip()
            if usuario_id.isdigit():
                uid_req = int(usuario_id)

                if role_name in PRIV_ALL:
                    where.append("g.usuario_id = ?")
                    args.append(uid_req)
                    filtros['usuario_id'] = uid_req

                elif role_name in GERENTE_ROLES:
                    if allowed_ids and uid_req in allowed_ids:
                        where.append("g.usuario_id = ?")
                        args.append(uid_req)
                        filtros['usuario_id'] = uid_req
                    else:
                        filtros['usuario_id'] = ''

                elif role_name in ('gerente general', 'gerente financiero'):
                    if allowed_ids and uid_req in allowed_ids:
                        where.append("g.usuario_id = ?")
                        args.append(uid_req)
                        filtros['usuario_id'] = uid_req
                    else:
                        filtros['usuario_id'] = ''

                else:
                    if allowed_ids and uid_req in allowed_ids:
                        where.append("g.usuario_id = ?")
                        args.append(uid_req)
                        filtros['usuario_id'] = uid_req
                    else:
                        filtros['usuario_id'] = ''
            else:
                filtros['usuario_id'] = ''

            base_from = """
                FROM gastos_tarjeta g
                LEFT JOIN terceros t ON t.id = g.proveedor_id
                LEFT JOIN usuarios u ON u.id = g.usuario_id
                LEFT JOIN usuarios uga ON uga.id = g.ga_aprobado_por
                LEFT JOIN usuarios ugg ON ugg.id = g.gg_aprobado_por
                LEFT JOIN usuarios ugf ON ugf.id = g.gf_aprobado_por
            """

            select_cols = """
                g.id,
                g.fecha,
                g.motivo,
                g.centro_costo,
                g.orden_compra,
                g.con_soporte,
                g.sin_soporte,
                g.subtotal_factura,
                g.servicios_10,
                g.subtotal_sin_iva,
                g.iva,
                g.total_con_iva,
                g.archivo,
                g.proveedor_id,
                g.usuario_id,
                g.ccb,
                t.identificacion,
                COALESCE(t.nombre, g.proveedor) AS proveedor_nombre,
                COALESCE(g.gg_aprobado, 0) AS gg_aprobado,
                COALESCE(g.gf_aprobado, 0) AS gf_aprobado,
                COALESCE(g.ga_aprobado, 0) AS ga_aprobado,
                COALESCE(g.tarjeta_sin_soporte,0) AS tarjeta_sin_soporte,
                COALESCE(g.boletos_aereos,0) AS boletos_aereos,

                (
                    SELECT STRING_AGG(m.x, ', ')
                    FROM (
                        SELECT DISTINCT COALESCE(
                            pv.nombre,
                            CAST(d.motivo AS NVARCHAR(MAX))
                        ) AS x
                        FROM gastos_tarjeta_detalle d
                        LEFT JOIN param_values pv
                        ON pv.group_id = 1
                        AND CAST(pv.valor AS NVARCHAR(MAX)) = CAST(d.motivo AS NVARCHAR(MAX))
                        WHERE d.gasto_id = g.id
                    ) AS m
                ) AS motivos_detalle,

                u.username AS usuario_username,
                u.departamento_id AS usuario_departamento_id,
                g.sap_contabilizacion,
                uga.username AS ga_aprobado_por_username,
                ugg.username AS gg_aprobado_por_username,
                ugf.username AS gf_aprobado_por_username,
                t.codigo_sap AS proveedor_codigo_sap,

                CASE
                    WHEN COALESCE(g.es_caja_chica,0)=1 THEN 'caja_chica'
                    WHEN COALESCE(g.reembolso_vendedor,0)=1 THEN 'reembolso'
                    WHEN COALESCE(g.boletos_aereos,0)=1 THEN 'tarjeta_boletos'
                    WHEN COALESCE(g.tarjeta_sin_soporte,0)=1 THEN 'tarjeta_online'
                    ELSE 'tarjeta'
                END AS tipo_gasto,

                CASE
                    WHEN LTRIM(RTRIM(COALESCE(g.archivo,''))) <> '' THEN 1
                    WHEN EXISTS (
                        SELECT TOP 1 1
                        FROM gastos_tarjeta_archivos a
                        WHERE a.gasto_id = g.id
                    ) THEN 1
                    ELSE 0
                END AS has_adjuntos
            """

            sql = f"SELECT {select_cols} {base_from}"
            _apply_empresa_scope(conn, where, args, user_alias="u", uid=uid)
            all_where = list(where)
            all_where.append("COALESCE(g.inactivo,0)=0")

            if all_where:
                sql += " WHERE " + " AND ".join(all_where)

            sql += " ORDER BY CAST(g.fecha AS date) DESC, g.id DESC"

            current_app.logger.warning("DEBUG SQL=%s", sql)
            current_app.logger.warning("DEBUG WHERE=%s | ARGS=%s", all_where, args)

            try:
                cur.execute(sql, args)
                desc = [c[0] for c in cur.description]
                rows = [_row_to_dict(r, desc) for r in cur.fetchall()]
            except Exception:
                current_app.logger.exception("Error ejecutando lista_gastos")
                rows = []

            def _user_display_name(user_id: int) -> str:
                if not user_id:
                    return ""
                try:
                    c = conn.cursor()
                    c.execute("""
                        SELECT
                            LTRIM(RTRIM(COALESCE(nombre_completo,''))) AS nombre,
                            LTRIM(RTRIM(COALESCE(username,''))) AS username
                        FROM usuarios
                        WHERE id = ?
                    """, (user_id,))
                    r = c.fetchone()
                    if not r:
                        return ""

                    try:
                        nombre = (r["nombre"] or "").strip()
                        username = (r["username"] or "").strip()
                    except Exception:
                        nombre = (r[0] or "").strip()
                        username = (r[1] or "").strip()

                    return nombre if nombre else username
                except Exception:
                    return ""

            rows2 = []
            is_gg_gf = role_name in ("gerente general", "gerente financiero", "coordinador", "admin")
            ROLE_VER_GERENTE = role_name in ("gerente general", "gerente financiero", "coordinador", "admin")

            for d in rows:
                d = dict(d or {})

                tipo_gasto = (d.get('tipo_gasto') or '').strip().lower()
                usuario_gasto = d.get('usuario_id')

                es_tarjeta = tipo_gasto in ('tarjeta', 'tarjeta_online', 'tarjeta_boletos')
                es_restringido = tipo_gasto in ('caja_chica', 'reembolso')

                d['can_act_as_gf_tarjeta'] = int(d.get('can_act_as_gf_tarjeta') or 0)
                d['can_act_as_ga_tarjeta'] = int(d.get('can_act_as_ga_tarjeta') or 0)
                d['can_act_as_ga_restringido'] = int(d.get('can_act_as_ga_restringido') or 0)
                d['ga_actor'] = (d.get('ga_actor') or 'ga')

                if es_tarjeta and role_name == 'gerente financiero':
                    if usuario_gasto in scope_ids_set:
                        d['can_act_as_ga_tarjeta'] = 1
                        d['can_act_as_gf_tarjeta'] = 1
                        d['ga_actor'] = 'gf'

                if es_tarjeta and role_name == 'gerente general':
                    if usuario_gasto in scope_ids_set:
                        d['can_act_as_ga_tarjeta'] = 1
                        d['ga_actor'] = 'gg'

                if es_restringido and role_name in ('gerente financiero', 'gerente general'):
                    if usuario_gasto in scope_ids_set:
                        d['can_act_as_ga_restringido'] = 1

                if ROLE_VER_GERENTE:
                    uid_gasto = int(usuario_gasto or 0)
                    try:
                        gerente_id = _get_ultimo_jefe_id(conn, uid_gasto)
                    except Exception:
                        gerente_id = None

                    gerente_id = int(gerente_id or 0)

                    if not gerente_id and uid_gasto:
                        gerente_id = uid_gasto

                    d["gerente_id"] = gerente_id
                    d["gerente_nombre"] = _user_display_name(gerente_id) if gerente_id else ""
                else:
                    d["gerente_id"] = 0
                    d["gerente_nombre"] = ""

                rows2.append(d)

            rows = rows2

            if ROLE_VER_GERENTE and filtros.get("gerente_id"):
                gid_req = int(filtros["gerente_id"])
                rows = [d for d in rows if int(d.get("gerente_id") or 0) == gid_req]

            gerentes_reg = []
            if is_gg_gf:
                seen = set()
                for d in rows:
                    gid = int(d.get("gerente_id") or 0)
                    gnm = (d.get("gerente_nombre") or "").strip()
                    if gid and gid not in seen:
                        seen.add(gid)
                        gerentes_reg.append({"id": gid, "nombre": gnm})
                gerentes_reg.sort(key=lambda x: (x["nombre"] or "").lower())

            usuarios_reg = []
            try:
                seen_u = set()
                for d in rows:
                    u_id = int(d.get("usuario_id") or 0)
                    u_nm = (d.get("usuario_username") or "").strip()
                    if u_id and u_id not in seen_u:
                        seen_u.add(u_id)
                        usuarios_reg.append({"id": u_id, "username": u_nm})
                usuarios_reg.sort(key=lambda x: (x["username"] or "").lower())
            except Exception:
                usuarios_reg = []

            def _f(x):
                try:
                    return float(x or 0)
                except Exception:
                    return 0.0

            totales = {k: 0.0 for k in (
                'con_soporte', 'sin_soporte', 'subtotal_factura',
                'servicios_10', 'subtotal_sin_iva', 'iva', 'total_con_iva'
            )}

            for r in rows:
                for k in totales.keys():
                    totales[k] += _f(r.get(k))

            is_admin = (role_name == 'admin') or bool(session.get('is_admin'))

            can_approve_gg = is_admin or (role_name == 'gerente general')
            can_approve_gf = is_admin or (role_name == 'gerente financiero')
            can_approve_ga = is_admin or (role_name in ('gerente', 'gerente de área', 'gerente de area'))

            readonly_view = (role_name == 'coordinador')

            return render_template(
                'gastos_lista.html',
                gastos=rows,
                filtros=filtros,
                proveedores=proveedores,
                usuarios_reg=usuarios_reg,
                gerentes_reg=gerentes_reg,
                totales=type('T', (), totales)(),
                can_approve_gg=can_approve_gg,
                can_approve_gf=can_approve_gf,
                can_approve_ga=can_approve_ga,
                readonly_view=readonly_view,
                usuario=session.get('usuario'),
                rol=session.get('rol'),
                active_page='gastos_tarjeta'
            )

        finally:
            try:
                conn.close()
            except Exception:
                pass
    
    # EJEMPLO: tu ruta actual (ajusta el nombre real)
    
    from flask import render_template, request, session, redirect, url_for, current_app
    from datetime import date

    @app.route('/reembolsos/gastos/reporte', methods=['GET'], endpoint='reporte_gastos')
    @require_permission('gastos_tarjeta', 'ver')
    @require_login
    def reporte_gastos():
        conn = get_db()
        cur = conn.cursor()

        role_name = (session.get('rol') or '').lower().strip()
        uid_session = session.get('usuario_id') or session.get('user_id')

        # --- Proveedores para datalist ---
        try:
            cur.execute("""
                SELECT id, nombre, identificacion
                FROM terceros
                WHERE UPPER(TRIM(tipo))='P'
                AND COALESCE(activo,1)=1
                ORDER BY nombre
            """)
            proveedores = cur.fetchall()
        except Exception:
            proveedores = []

        # ==========================================================
        # ✅ EVITAR QUE "pendientes=1" SE QUEDE PEGADO AL DAR "BUSCAR"
        # ==========================================================
        accion = (request.args.get("accion") or "").lower().strip()

        def _redirect_without_pendientes():
            q = request.args.to_dict(flat=True)
            q.pop("pendientes", None)
            q.pop("pend", None)
            q.pop("pend_view", None)
            q.pop("accion", None)
            return redirect(url_for("reporte_gastos", **q))

        if accion == "buscar" and ((request.args.get("pendientes") == "1") or (request.args.get("pend") == "1")):
            return _redirect_without_pendientes()

        # ==========================================================
        # ✅ BOTÓN "PENDIENTES": forzar MES + activar pendientes=1
        # ==========================================================
        from datetime import date

        def _redirect_pendientes_mes():
            hoy = date.today()
            inicio_mes = hoy.replace(day=1)

            q = request.args.to_dict(flat=True)
            q.pop("pend_view", None)
            q.pop("accion", None)
            q.pop("desde", None)
            q.pop("hasta", None)

            q["pendientes"] = "1"
            q["desde"] = inicio_mes.isoformat()
            q["hasta"] = hoy.isoformat()
            return redirect(url_for("reporte_gastos", **q))

        if request.args.get("pend_view") == "1":
            return _redirect_pendientes_mes()

        # ==========================================================
        # ✅ FILTROS: SOLO lo que viene por request.args (SIN session scope)
        # ==========================================================
        filtros = {}
        where = []
        args = []

        # ✅ siempre excluir inactivos
        where.append("COALESCE(g.inactivo,0)=0")

        # ---- fechas ----
        desde = (request.args.get("desde") or "").strip()
        hasta = (request.args.get("hasta") or "").strip()
        filtros["desde"] = desde
        filtros["hasta"] = hasta

        if desde:
            where.append("date(g.fecha) >= date(?)")
            args.append(desde)
        if hasta:
            where.append("date(g.fecha) <= date(?)")
            args.append(hasta)

        # ---- proveedor_id (hidden) ----
        proveedor_id = (request.args.get("proveedor_id") or "").strip()
        filtros["proveedor_id"] = proveedor_id
        filtros["proveedor"] = (request.args.get("proveedor") or "").strip()

        if proveedor_id.isdigit():
            where.append("g.proveedor_id = ?")
            args.append(int(proveedor_id))

        # ---- descripción / motivo general ----
        descripcion = (request.args.get("descripcion") or "").strip()
        filtros["descripcion"] = descripcion
        if descripcion:
            where.append("(g.motivo LIKE ?)")
            args.append(f"%{descripcion}%")

        # ---- CCB ----
        ccb_req = (request.args.get("ccb") or "").strip()
        if ccb_req in ("0", "1"):
            where.append("COALESCE(g.ccb,0)=?")
            args.append(int(ccb_req))
            filtros["ccb"] = ccb_req
        else:
            filtros["ccb"] = ""

        # ---- Tipo ----
        tipo = (request.args.get("tipo") or "").strip().lower()
        TIPOS_RESTRINGIDOS = ("caja_chica", "reembolso")

        if tipo == "caja_chica":
            where.append("COALESCE(g.es_caja_chica,0)=1")
        elif tipo == "reembolso":
            where.append("COALESCE(g.reembolso_vendedor,0)=1 AND COALESCE(g.es_caja_chica,0)=0")
        elif tipo == "tarjeta":
            where.append("COALESCE(g.es_caja_chica,0)=0 AND COALESCE(g.reembolso_vendedor,0)=0")
        elif tipo == "tarjeta_online":
            where.append("""
                COALESCE(g.es_caja_chica,0)=0
                AND COALESCE(g.reembolso_vendedor,0)=0
                AND COALESCE(g.tarjeta_sin_soporte,0)=1
            """)
        elif tipo == "tarjeta_boletos":
            where.append("""
                COALESCE(g.es_caja_chica,0)=0
                AND COALESCE(g.reembolso_vendedor,0)=0
                AND COALESCE(g.boletos_aereos,0)=1
            """)
        filtros["tipo"] = tipo

        # ---- pendientes (SIN rol) ----
        pendientes = (request.args.get("pendientes") or request.args.get("pend") or "").strip()
        filtros["pendientes"] = pendientes

        if pendientes == "1":
            if tipo in TIPOS_RESTRINGIDOS:
                where.append("""
                    COALESCE(g.ga_aprobado,0)=0
                    AND COALESCE(g.gg_aprobado,0)=0
                    AND COALESCE(g.gf_aprobado,0)=0
                """)
            else:
                where.append("""
                    (
                        COALESCE(g.ga_aprobado,0)=0
                        OR COALESCE(g.gg_aprobado,0)=0
                        OR COALESCE(g.gf_aprobado,0)=0
                    )
                """)

        # ---- usuario_id (SOLO si viene en filtros) ----
        usuario_id = (request.args.get("usuario_id") or "").strip()
        if usuario_id.isdigit():
            where.append("g.usuario_id = ?")
            args.append(int(usuario_id))
            filtros["usuario_id"] = int(usuario_id)
        else:
            filtros["usuario_id"] = ""

        # ---- gerente_id (se filtra después en Python) ----
        gerente_id_req = (request.args.get("gerente_id") or "").strip()
        if gerente_id_req.isdigit():
            filtros["gerente_id"] = int(gerente_id_req)
        else:
            filtros["gerente_id"] = ""

        # ==========================================================
        # SQL base
        # ==========================================================
        base_from = """
            FROM gastos_tarjeta g
            LEFT JOIN terceros t ON t.id = g.proveedor_id
            LEFT JOIN usuarios u ON u.id = g.usuario_id
            LEFT JOIN usuarios uga ON uga.id = g.ga_aprobado_por
            LEFT JOIN usuarios ugg ON ugg.id = g.gg_aprobado_por
            LEFT JOIN usuarios ugf ON ugf.id = g.gf_aprobado_por
        """

        select_cols = """
            g.id, g.fecha, g.motivo, g.centro_costo, g.orden_compra,
            g.con_soporte, g.sin_soporte, g.subtotal_factura,
            g.servicios_10, g.subtotal_sin_iva, g.iva, g.total_con_iva, g.archivo,
            g.proveedor_id, g.usuario_id, g.ccb, t.identificacion,
            COALESCE(t.nombre, g.proveedor) AS proveedor_nombre,

            COALESCE(g.gg_aprobado, 0) AS gg_aprobado,
            COALESCE(g.gf_aprobado, 0) AS gf_aprobado,
            COALESCE(g.ga_aprobado, 0) AS ga_aprobado,

            COALESCE(g.tarjeta_sin_soporte,0) AS tarjeta_sin_soporte,
            COALESCE(g.boletos_aereos,0) AS boletos_aereos,

            (
                SELECT STRING_AGG(x, ', ')
                FROM (
                    SELECT DISTINCT COALESCE(pv.nombre, CAST(d.motivo AS TEXT)) AS x
                    FROM gastos_tarjeta_detalle d
                    LEFT JOIN param_values pv
                    ON pv.group_id = 1
                    AND CAST(pv.valor AS TEXT) = CAST(d.motivo AS TEXT)
                    WHERE d.gasto_id = g.id
                )
            ) AS motivos_detalle,

            u.username AS usuario_username,
            u.departamento_id AS usuario_departamento_id,

            g.sap_contabilizacion,
            uga.username AS ga_aprobado_por_username,
            ugg.username AS gg_aprobado_por_username,
            ugf.username AS gf_aprobado_por_username,

            t.codigo_sap AS proveedor_codigo_sap,

            CASE
                WHEN COALESCE(g.es_caja_chica,0)=1 THEN 'caja_chica'
                WHEN COALESCE(g.reembolso_vendedor,0)=1 THEN 'reembolso'
                WHEN COALESCE(g.boletos_aereos,0)=1 THEN 'tarjeta_boletos'
                WHEN COALESCE(g.tarjeta_sin_soporte,0)=1 THEN 'tarjeta_online'
                ELSE 'tarjeta'
            END AS tipo_gasto,

            CASE
                WHEN TRIM(COALESCE(g.archivo,'')) <> '' THEN 1
                WHEN EXISTS (
                    SELECT TOP 1 1 FROM gastos_tarjeta_archivos a
                    WHERE a.gasto_id = g.id
                     
                ) THEN 1
                ELSE 0
            END AS has_adjuntos
        """
        _apply_empresa_scope(conn, where, args, user_alias="u", uid=uid_session)
        sql = f"SELECT {select_cols} {base_from}"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY CAST(g.fecha AS date) DESC, g.id DESC"

        current_app.logger.warning("DEBUG SQL=%s", sql)
        current_app.logger.warning("DEBUG WHERE=%s | ARGS=%s", where, args)

        try:
            cur.execute(sql, args)
            rows = cur.fetchall()
        except Exception:
            rows = []

        # ==========================================================
        # Helper nombre de usuario
        # ==========================================================
        def _user_display_name(user_id: int) -> str:
            if not user_id:
                return ""
            try:
                c = conn.cursor()
                c.execute("""
                    SELECT TRIM(COALESCE(nombre_completo,'')) AS nombre,
                        TRIM(COALESCE(username,'')) AS username
                    FROM usuarios
                    WHERE id = ?
                """, (user_id,))
                r = c.fetchone()
                if not r:
                    return ""
                nombre = (r[0] or "").strip()
                return nombre if nombre else (r[1] or "").strip()
            except Exception:
                return ""

        # ==========================================================
        # Gerente por fila + filtro gerente en Python
        # ==========================================================
        rows2 = []
        for r in rows:
            d = dict(r)
            usuario_gasto = int(d.get("usuario_id") or 0)

            try:
                gerente_id = _get_ultimo_jefe_id(conn, usuario_gasto)
            except Exception:
                gerente_id = None

            gerente_id = int(gerente_id or 0)
            if not gerente_id and usuario_gasto:
                gerente_id = usuario_gasto

            d["gerente_id"] = gerente_id
            d["gerente_nombre"] = _user_display_name(gerente_id) if gerente_id else ""

            rows2.append(d)

        rows = rows2

        if filtros.get("gerente_id"):
            gid_req = int(filtros["gerente_id"])
            rows = [d for d in rows if int(d.get("gerente_id") or 0) == gid_req]

        # Dropdown Gerentes
        gerentes_reg, seen = [], set()
        for d in rows:
            gid = int(d.get("gerente_id") or 0)
            gnm = (d.get("gerente_nombre") or "").strip()
            if gid and gid not in seen:
                seen.add(gid)
                gerentes_reg.append({"id": gid, "nombre": gnm})
        gerentes_reg.sort(key=lambda x: (x["nombre"] or "").lower())

        # Dropdown Usuarios
        usuarios_reg, seen_u = [], set()
        for d in rows:
            u_id = int(d.get("usuario_id") or 0)
            u_nm = (d.get("usuario_username") or "").strip()
            if u_id and u_id not in seen_u:
                seen_u.add(u_id)
                usuarios_reg.append({"id": u_id, "username": u_nm})
        usuarios_reg.sort(key=lambda x: (x["username"] or "").lower())

        # Totales
        def _f(x):
            try:
                return float(x or 0)
            except Exception:
                return 0.0

        totales = {k: 0.0 for k in (
            'con_soporte', 'sin_soporte', 'subtotal_factura',
            'servicios_10', 'subtotal_sin_iva', 'iva', 'total_con_iva'
        )}
        for rr in rows:
            for k in totales.keys():
                totales[k] += _f(rr.get(k))

        # flags para template
        is_admin = (role_name == 'admin') or bool(session.get('is_admin'))
        can_approve_gg = is_admin or (role_name == 'gerente general')
        can_approve_gf = is_admin or (role_name == 'gerente financiero')
        can_approve_ga = is_admin or (role_name in ('gerente', 'gerente de área', 'gerente de area'))
        readonly_view = (role_name == 'coordinador')

        try:
            conn.close()
        except Exception:
            pass

        return render_template(
            'gastos_reporte.html',
            gastos=rows,
            filtros=filtros,
            proveedores=proveedores,
            usuarios_reg=usuarios_reg,
            gerentes_reg=gerentes_reg,
            totales=type('T', (), totales)(),
            can_approve_gg=can_approve_gg,
            can_approve_gf=can_approve_gf,
            can_approve_ga=can_approve_ga,
            readonly_view=readonly_view,
            usuario=session.get('usuario'),
            rol=session.get('rol'),
            active_page='gastos_tarjeta'
        )
    


    @app.route('/reembolsos/gastos/pendientes-aprobacion', methods=['GET'], endpoint='gastos_pendientes_aprobacion')
    @require_permission('gastos_pendientes_aprobacion', 'ver')
    @require_login
    def gastos_pendientes_aprobacion():
        conn = get_db()
        cur = conn.cursor()

        def _row_to_dict(row, columns=None):
            if row is None:
                return None
            try:
                return dict(row)
            except Exception:
                if columns:
                    return {columns[i]: row[i] for i in range(len(columns))}
                return None

        try:
            # --- Rol y usuario actual ---
            role_name = (session.get('rol') or '').lower().strip()
            uid = session.get('usuario_id') or session.get('user_id')
            from datetime import date
            # ==========================================================
            # FECHAS POR DEFECTO EN BACKEND
            # Evita cargar TODO y luego que JS reenvíe el formulario
            # ==========================================================
            qs = request.args.to_dict(flat=True)

            desde_req = (qs.get("desde") or "").strip()
            hasta_req = (qs.get("hasta") or "").strip()

            special = (
                qs.get("ccb") == "1"
                or qs.get("pendientes") == "1"
                or qs.get("pend") == "1"
                or qs.get("pend_view") == "1"
            )

            if not desde_req and not hasta_req and not special:
                hoy = date.today().isoformat()
                qs["desde"] = hoy
                qs["hasta"] = hoy
                qs.pop("accion", None)

                return redirect(url_for("gastos_pendientes_aprobacion", **qs))

            # ✅ SOLO GERENTES (capa extra de seguridad)
            ALLOWED = ("gerente", "gerente de área", "gerente de area", "gerente general", "gerente financiero")
            if role_name not in ALLOWED:
                try:
                    conn.close()
                except Exception:
                    pass
                flash("No tiene permisos para acceder a pendientes de aprobación.", "warning")
                return redirect(url_for("dashboard"))

            # --- Proveedores para datalist ---
            try:
                cur.execute("""
                    SELECT id, nombre, identificacion
                    FROM terceros
                    WHERE UPPER(LTRIM(RTRIM(tipo))) = 'P'
                    AND COALESCE(activo, 1) = 1
                    ORDER BY nombre
                """)
                proveedores = cur.fetchall()
            except Exception:
                proveedores = []

            # ==========================================================
            # Filtros base existentes
            # ==========================================================
            filtros, where, args, _ = gh.collect_gastos_pendientes_aprobacion_filters(request, session)

            # ✅ Forzar SIEMPRE pendientes=1 en esta pantalla
            filtros["pendientes"] = "1"

            # ==========================================================
            # ✅ Filtro CCB: '' (todos) | '1' (con CCB) | '0' (sin CCB)
            # ==========================================================
            ccb_req = (request.args.get('ccb') or '').strip()
            if ccb_req in ('0', '1'):
                where.append("COALESCE(g.ccb,0)=?")
                args.append(int(ccb_req))
                filtros['ccb'] = ccb_req
            else:
                filtros['ccb'] = ''

            tipo = (request.args.get('tipo') or '').strip().lower()
            tipo_raw = request.args.get('tipo')

            current_app.logger.warning("DEBUG tipo_raw=%r tipo_norm=%r args=%r", tipo_raw, tipo, dict(request.args))

            TIPOS_RESTRINGIDOS = ('caja_chica', 'reembolso')

            if tipo == 'caja_chica':
                where.append("COALESCE(g.es_caja_chica,0)=1")
            elif tipo == 'reembolso':
                where.append("COALESCE(g.reembolso_vendedor,0)=1 AND COALESCE(g.es_caja_chica,0)=0")
            elif tipo == 'tarjeta':
                where.append("COALESCE(g.es_caja_chica,0)=0 AND COALESCE(g.reembolso_vendedor,0)=0")
            elif tipo == 'tarjeta_online':
                where.append("""
                    COALESCE(g.es_caja_chica,0)=0
                    AND COALESCE(g.reembolso_vendedor,0)=0
                    AND COALESCE(g.tarjeta_sin_soporte,0)=1
                """)
            elif tipo == 'tarjeta_boletos':
                where.append("""
                    COALESCE(g.es_caja_chica,0)=0
                    AND COALESCE(g.reembolso_vendedor,0)=0
                    AND COALESCE(g.boletos_aereos,0)=1
                """)

            filtros['tipo'] = tipo

            # ==========================================================
            # ✅ filtro por GERENTE (solo GG/GF, se aplica después)
            # ==========================================================
            gerente_id_req = (request.args.get("gerente_id") or "").strip()
            if gerente_id_req.isdigit():
                filtros["gerente_id"] = int(gerente_id_req)
            else:
                filtros["gerente_id"] = ""

            pendientes = "1"
            current_app.logger.info("ROL_SESSION=%r role_name=%r pendientes=%r", session.get("rol"), role_name, pendientes)

            # --- Filtro por descripción / detalle ---
            descripcion = (request.args.get('descripcion') or '').strip()
            if descripcion:
                where.append("(g.motivo LIKE ?)")
                args.append(f"%{descripcion}%")
                filtros['descripcion'] = descripcion
            else:
                filtros['descripcion'] = ''

            # ==========================================================
            # ✅ REGLAS DE VISIBILIDAD
            # ==========================================================
            PRIV_ALL = ('admin', 'coordinador')
            if tipo not in TIPOS_RESTRINGIDOS:
                PRIV_ALL = ('admin', 'coordinador', 'gerente general', 'gerente financiero')

            GERENTE_ROLES = ('gerente', 'gerente de área', 'gerente de area')

            allowed_ids = None
            scope_ids_set = set()

            if role_name in ('gerente general', 'gerente financiero'):
                scope_ids_set = set(obtener_subordinados(conn, uid) or [])
                if uid:
                    scope_ids_set.add(uid)

                placeholders = ",".join("?" * len(scope_ids_set)) if scope_ids_set else "?"
                where.append(
                    f"""(
                        (COALESCE(g.es_caja_chica,0)=0 AND COALESCE(g.reembolso_vendedor,0)=0)
                        OR g.usuario_id IN ({placeholders})
                    )"""
                )
                args.extend(list(scope_ids_set) if scope_ids_set else [uid])

                allowed_ids = scope_ids_set.copy()

            if role_name in PRIV_ALL:
                pass

            elif role_name in ('gerente general', 'gerente financiero'):
                if allowed_ids is None:
                    allowed_ids = {uid}

            elif role_name in GERENTE_ROLES:
                gerente_id = uid
                subordinados = obtener_subordinados(conn, gerente_id) or []
                allowed_ids = set(subordinados)
                if gerente_id:
                    allowed_ids.add(gerente_id)

                if not allowed_ids:
                    allowed_ids = {gerente_id}

                placeholders = ",".join("?" * len(allowed_ids))
                where.append(f"g.usuario_id IN ({placeholders})")
                args.extend(list(allowed_ids))

            else:
                where.append("g.usuario_id = ?")
                args.append(uid)
                allowed_ids = {uid}

            # --- Filtro por usuario (sin escapar del alcance) ---
            usuario_id = (request.args.get('usuario_id') or '').strip()
            if usuario_id.isdigit():
                uid_req = int(usuario_id)

                if role_name in PRIV_ALL:
                    where.append("g.usuario_id = ?")
                    args.append(uid_req)
                    filtros['usuario_id'] = uid_req

                elif role_name in GERENTE_ROLES:
                    if allowed_ids and uid_req in allowed_ids:
                        where.append("g.usuario_id = ?")
                        args.append(uid_req)
                        filtros['usuario_id'] = uid_req
                    else:
                        filtros['usuario_id'] = ''

                elif role_name in ('gerente general', 'gerente financiero'):
                    if allowed_ids and uid_req in allowed_ids:
                        where.append("g.usuario_id = ?")
                        args.append(uid_req)
                        filtros['usuario_id'] = uid_req
                    else:
                        filtros['usuario_id'] = ''

                else:
                    if allowed_ids and uid_req in allowed_ids:
                        where.append("g.usuario_id = ?")
                        args.append(uid_req)
                        filtros['usuario_id'] = uid_req
                    else:
                        filtros['usuario_id'] = ''
            else:
                filtros['usuario_id'] = ''

            # --- SQL base ---
            base_from = """
                FROM gastos_tarjeta g
                LEFT JOIN terceros t ON t.id = g.proveedor_id
                LEFT JOIN usuarios u ON u.id = g.usuario_id
                LEFT JOIN usuarios uga ON uga.id = g.ga_aprobado_por
                LEFT JOIN usuarios ugg ON ugg.id = g.gg_aprobado_por
                LEFT JOIN usuarios ugf ON ugf.id = g.gf_aprobado_por
            """

            select_cols = """
                g.id,
                g.fecha,
                g.motivo,
                g.centro_costo,
                g.orden_compra,
                g.con_soporte,
                g.sin_soporte,
                g.subtotal_factura,
                g.servicios_10,
                g.subtotal_sin_iva,
                g.iva,
                g.total_con_iva,
                g.archivo,
                g.proveedor_id,
                g.usuario_id,
                g.ccb,
                t.identificacion,
                COALESCE(t.nombre, g.proveedor) AS proveedor_nombre,
                COALESCE(g.gg_aprobado, 0) AS gg_aprobado,
                COALESCE(g.gf_aprobado, 0) AS gf_aprobado,
                COALESCE(g.ga_aprobado, 0) AS ga_aprobado,
                COALESCE(g.tarjeta_sin_soporte,0) AS tarjeta_sin_soporte,
                COALESCE(g.boletos_aereos,0) AS boletos_aereos,

                (
                    SELECT STRING_AGG(m.x, ', ')
                    FROM (
                        SELECT DISTINCT COALESCE(
                            pv.nombre,
                            CAST(d.motivo AS NVARCHAR(MAX))
                        ) AS x
                        FROM gastos_tarjeta_detalle d
                        LEFT JOIN param_values pv
                        ON pv.group_id = 1
                        AND CAST(pv.valor AS NVARCHAR(MAX)) = CAST(d.motivo AS NVARCHAR(MAX))
                        WHERE d.gasto_id = g.id
                    ) AS m
                ) AS motivos_detalle,

                u.username AS usuario_username,
                u.departamento_id AS usuario_departamento_id,
                g.sap_contabilizacion,
                uga.username AS ga_aprobado_por_username,
                ugg.username AS gg_aprobado_por_username,
                ugf.username AS gf_aprobado_por_username,
                t.codigo_sap AS proveedor_codigo_sap,

                CASE
                    WHEN COALESCE(g.es_caja_chica,0)=1 THEN 'caja_chica'
                    WHEN COALESCE(g.reembolso_vendedor,0)=1 THEN 'reembolso'
                    WHEN COALESCE(g.boletos_aereos,0)=1 THEN 'tarjeta_boletos'
                    WHEN COALESCE(g.tarjeta_sin_soporte,0)=1 THEN 'tarjeta_online'
                    ELSE 'tarjeta'
                END AS tipo_gasto,

                CASE
                    WHEN LTRIM(RTRIM(COALESCE(g.archivo,''))) <> '' THEN 1
                    WHEN EXISTS (
                        SELECT TOP 1 1
                        FROM gastos_tarjeta_archivos a
                        WHERE a.gasto_id = g.id
                    ) THEN 1
                    ELSE 0
                END AS has_adjuntos
            """

            sql = f"SELECT {select_cols} {base_from}"
            _apply_empresa_scope(conn, where, args, user_alias="u", uid=uid)
            all_where = list(where)
            all_where.append("COALESCE(g.inactivo,0)=0")

            if all_where:
                sql += " WHERE " + " AND ".join(all_where)

            sql += " ORDER BY CAST(g.fecha AS date) DESC, g.id DESC"

            current_app.logger.warning("DEBUG SQL=%s", sql)
            current_app.logger.warning("DEBUG WHERE=%s | ARGS=%s", all_where, args)

            try:
                cur.execute(sql, args)
                desc = [c[0] for c in cur.description]
                rows = [_row_to_dict(r, desc) for r in cur.fetchall()]
            except Exception:
                current_app.logger.exception("Error ejecutando gastos_pendientes_aprobacion")
                rows = []

            # ==========================================================
            # Helper: nombre usuario
            # ==========================================================
            def _user_display_name(user_id: int) -> str:
                if not user_id:
                    return ""
                try:
                    c = conn.cursor()
                    c.execute("""
                        SELECT
                            LTRIM(RTRIM(COALESCE(nombre_completo,''))) AS nombre,
                            LTRIM(RTRIM(COALESCE(username,''))) AS username
                        FROM usuarios
                        WHERE id = ?
                    """, (user_id,))
                    r = c.fetchone()
                    if not r:
                        return ""

                    try:
                        nombre = (r["nombre"] or "").strip()
                        username = (r["username"] or "").strip()
                    except Exception:
                        nombre = (r[0] or "").strip()
                        username = (r[1] or "").strip()

                    return nombre if nombre else username
                except Exception:
                    return ""

            # ==========================================================
            # FLAGS POR FILA + gerente por fila
            # ==========================================================
            rows2 = []
            is_gg_gf = role_name in ("gerente general", "gerente financiero", "coordinador", "admin")
            ROLE_VER_GERENTE = role_name in ("gerente general", "gerente financiero", "coordinador", "admin")

            for d in rows:
                d = dict(d or {})

                tipo_gasto = (d.get('tipo_gasto') or '').strip().lower()
                usuario_gasto = d.get('usuario_id')

                # ✅ Regla negocio: TARJETA ONLINE => GA efectivo es GG
                if tipo_gasto == "tarjeta_online":
                    d["ga_actor"] = "gg"

                es_tarjeta = tipo_gasto in ('tarjeta', 'tarjeta_online', 'tarjeta_boletos')
                es_restringido = tipo_gasto in ('caja_chica', 'reembolso')

                d['can_act_as_gf_tarjeta'] = int(d.get('can_act_as_gf_tarjeta') or 0)
                d['can_act_as_ga_tarjeta'] = int(d.get('can_act_as_ga_tarjeta') or 0)
                d['can_act_as_ga_restringido'] = int(d.get('can_act_as_ga_restringido') or 0)
                d['ga_actor'] = (d.get('ga_actor') or 'ga')

                if es_tarjeta and role_name == 'gerente financiero':
                    if usuario_gasto in scope_ids_set:
                        d['can_act_as_ga_tarjeta'] = 1
                        d['can_act_as_gf_tarjeta'] = 1

                        # SOLO si NO es tarjeta_online, GF puede ser GA actor
                        if tipo_gasto != "tarjeta_online":
                            d['ga_actor'] = 'gf'

                if es_tarjeta and role_name == 'gerente general':
                    if usuario_gasto in scope_ids_set:
                        d['can_act_as_ga_tarjeta'] = 1
                        d['ga_actor'] = 'gg'

                if es_restringido and role_name in ('gerente financiero', 'gerente general'):
                    if usuario_gasto in scope_ids_set:
                        d['can_act_as_ga_restringido'] = 1

                if ROLE_VER_GERENTE:
                    uid_gasto = int(usuario_gasto or 0)
                    try:
                        gerente_id = _get_ultimo_jefe_id(conn, uid_gasto)
                    except Exception:
                        gerente_id = None

                    gerente_id = int(gerente_id or 0)
                    if not gerente_id and uid_gasto:
                        gerente_id = uid_gasto

                    d["gerente_id"] = gerente_id
                    d["gerente_nombre"] = _user_display_name(gerente_id) if gerente_id else ""
                else:
                    d["gerente_id"] = 0
                    d["gerente_nombre"] = ""

                rows2.append(d)

            rows = rows2

            def _is_pendiente(d):
                ga = int(d.get("ga_aprobado") or 0)
                gg = int(d.get("gg_aprobado") or 0)
                gf = int(d.get("gf_aprobado") or 0)

                tipo = (d.get("tipo_gasto") or "").strip().lower()
                es_restringido = tipo in ("caja_chica", "reembolso")
                es_tarjeta = tipo in ("tarjeta", "tarjeta_online", "tarjeta_boletos")

                ga_eff_restringido = 1 if (ga == 1 or gg == 1 or gf == 1) else 0

                if role_name in ("gerente", "gerente de área", "gerente de area"):
                    return ga == 0

                if role_name == "gerente general":
                    if es_restringido:
                        if int(d.get("can_act_as_ga_restringido") or 0) == 1:
                            return ga_eff_restringido == 0
                        return False

                    if es_tarjeta:
                        if int(d.get("can_act_as_ga_tarjeta") or 0) == 1:
                            return ga == 0

                        return (ga == 1 and gg == 0)

                    return False

                if role_name == "gerente financiero":
                    if es_restringido:
                        if int(d.get("can_act_as_ga_restringido") or 0) == 1:
                            return ga_eff_restringido == 0
                        return False

                    if es_tarjeta:
                        # GF actúa como GA -> la pendiente se mide con GF
                        if int(d.get("can_act_as_ga_tarjeta") or 0) == 1 and (str(d.get("ga_actor") or "").lower() == "gf"):
                            return (gf == 0)

                        # Tarjeta online (GA=GG): GF la ve pendiente hasta aprobar GF
                        if (tipo == "tarjeta_online") and (str(d.get("ga_actor") or "").lower() == "gg"):
                            return (gf == 0)

                        # Flujo normal: GG ya aprobó y GF aún no
                        return (ga == 1 and gf == 0)

                    return False

                return False

            rows = [d for d in rows if _is_pendiente(d)]

            if ROLE_VER_GERENTE and filtros.get("gerente_id"):
                gid_req = int(filtros["gerente_id"])
                rows = [d for d in rows if int(d.get("gerente_id") or 0) == gid_req]

            gerentes_reg = []
            if is_gg_gf:
                seen = set()
                for d in rows:
                    gid = int(d.get("gerente_id") or 0)
                    gnm = (d.get("gerente_nombre") or "").strip()
                    if gid and gid not in seen:
                        seen.add(gid)
                        gerentes_reg.append({"id": gid, "nombre": gnm})
                gerentes_reg.sort(key=lambda x: (x["nombre"] or "").lower())

            usuarios_reg = []
            try:
                seen_u = set()
                for d in rows:
                    u_id = int(d.get("usuario_id") or 0)
                    u_nm = (d.get("usuario_username") or "").strip()
                    if u_id and u_id not in seen_u:
                        seen_u.add(u_id)
                        usuarios_reg.append({"id": u_id, "username": u_nm})
                usuarios_reg.sort(key=lambda x: (x["username"] or "").lower())
            except Exception:
                usuarios_reg = []

            def _f(x):
                try:
                    return float(x or 0)
                except Exception:
                    return 0.0

            totales = {k: 0.0 for k in (
                'con_soporte', 'sin_soporte', 'subtotal_factura',
                'servicios_10', 'subtotal_sin_iva', 'iva', 'total_con_iva'
            )}

            for r in rows:
                for k in totales.keys():
                    totales[k] += _f(r.get(k))

            is_admin = (role_name == 'admin') or bool(session.get('is_admin'))

            can_approve_gg = is_admin or (role_name == 'gerente general')
            can_approve_gf = is_admin or (role_name == 'gerente financiero')
            can_approve_ga = is_admin or (role_name in ('gerente', 'gerente de área', 'gerente de area'))

            readonly_view = (role_name == 'coordinador')

            return render_template(
                'gastos_pendientes_aprobacion.html',
                gastos=rows,
                filtros=filtros,
                proveedores=proveedores,
                usuarios_reg=usuarios_reg,
                gerentes_reg=gerentes_reg,
                totales=type('T', (), totales)(),
                can_approve_gg=can_approve_gg,
                can_approve_gf=can_approve_gf,
                can_approve_ga=can_approve_ga,
                readonly_view=readonly_view,
                usuario=session.get('usuario'),
                rol=session.get('rol'),
                active_page='gastos_pendientes'
            )

        finally:
            try:
                conn.close()
            except Exception:
                pass    


    from flask import render_template, request
    from jinja2 import TemplateNotFound

    @app.route('/reembolsos/gastos/<int:gid>/adjuntos', methods=['GET'], endpoint='ver_gasto_adjuntos')
    @require_login
    @require_permission('gastos_tarjeta', 'ver')
    def ver_gasto_adjuntos(gid):
        conn = get_db()
        try:
            conn.row_factory = sqlite3.Row
        except Exception:
            pass
        cur = conn.cursor()

        # Traer gasto (para validar existencia y legacy archivo)
        cur.execute(f"SELECT id, archivo FROM {TABLE_GASTOS} WHERE id=?", (gid,))
        g = cur.fetchone()
        if not g:
            try: conn.close()
            except Exception: pass
            return "<div class='text-muted'>Gasto no encontrado.</div>", 404

        # Multi-adjuntos
        adjuntos = []
        try:
            cur.execute("""
                SELECT filename
                FROM gastos_tarjeta_archivos
                WHERE gasto_id=?
                ORDER BY id
            """, (gid,))
            adjuntos = [r["filename"] for r in cur.fetchall() if r and r["filename"]]
        except Exception:
            adjuntos = []

        # Legacy archivo (columna gastos_tarjeta.archivo)
        legacy = (g["archivo"] or "").strip()
        if legacy and legacy not in adjuntos:
            adjuntos.insert(0, legacy)

        try:
            conn.close()
        except Exception:
            pass

        # Render como “fragmento” para meter dentro del modal
        try:
            return render_template("gastos_adjuntos_popup.html", gid=gid, adjuntos=adjuntos)
        except TemplateNotFound:
            # fallback simple si no creas template
            items = "".join(
                f"<a class='list-group-item list-group-item-action d-flex justify-content-between align-items-center' "
                f"href='/static/uploads/{a}' target='_blank' rel='noopener'>"
                f"<span class='text-truncate'><i class='bi bi-paperclip me-2'></i>{a.split('/')[-1]}</span>"
                f"<span class='badge bg-primary rounded-pill'>Abrir</span></a>"
                for a in adjuntos
            )
            if not items:
                return "<div class='text-muted'>Sin adjuntos.</div>"
            return f"<div class='list-group'>{items}</div><small class='text-muted d-block mt-2'>Se abrirán en una nueva pestaña.</small>"

  
  
  
    @app.get("/reembolsos/xml/<int:fid>", endpoint="factura_xml_detalle")
    @require_login
    @require_permission("gastos_tarjeta", "ver")
    def factura_xml_detalle(fid):
        conn = get_db(); cur = conn.cursor()

        cur.execute("SELECT * FROM facturas_xml WHERE id=?", (fid,))
        f = cur.fetchone()
        if not f:
            conn.close()
            return "Factura no encontrada", 404

        cur.execute("SELECT * FROM facturas_xml_det WHERE factura_id=? ORDER BY id", (fid,))
        detalles = cur.fetchall()
        conn.close()

        popup = request.args.get("popup") == "1"
        if popup:
            return render_template("factura_xml_detalle_popup.html", f=f, detalles=detalles)

        return render_template("factura_xml_detalle.html", f=f, detalles=detalles)


    # ✅ Reemplaza AMBAS definiciones de ver_gasto / ver_gasto2 por esta única:
    @app.route('/reembolsos/gastos/<int:gid>/ver', methods=['GET'], endpoint='ver_gasto2')
    @require_login
    @require_permission('gastos_tarjeta', 'ver')
    def ver_gasto2(gid):
        conn = get_db()
        cur = conn.cursor()

        cur.execute(f"""
            SELECT
                g.*,
                t.identificacion AS proveedor_identificacion
            FROM {TABLE_GASTOS} g
            LEFT JOIN terceros t ON t.id = g.proveedor_id
            WHERE g.id = ?
        """, (gid,))

        row = cur.fetchone()
        if not row:
            conn.close()
            flash('Gasto no encontrado.', 'warning')
            return redirect(url_for('lista_gastos'))

        # ✅ OUTER APPLY con TOP 1 evita duplicados cuando param_values
        #    tiene más de un registro con el mismo valor en el mismo group_id
        cur.execute("""
            SELECT
                d.id,
                d.descripcion,
                d.observacion,
                d.centro_costo,
                pc.nombre AS centro_nombre,
                d.motivo,
                pm.nombre AS motivo_nombre,
                d.indicador,
                d.con_soporte,
                d.sin_soporte,
                d.subtotal_factura,
                d.servicios_10,
                d.subtotal_sin_iva,
                d.iva,
                d.total_con_iva
            FROM gastos_tarjeta_detalle d
            OUTER APPLY (
                SELECT TOP 1 nombre
                FROM param_values
                WHERE group_id = 7
                AND activo = 1
                AND TRIM(COALESCE(valor,'')) = TRIM(COALESCE(d.centro_costo,''))
                ORDER BY id
            ) pc
            OUTER APPLY (
                SELECT TOP 1 nombre
                FROM param_values
                WHERE group_id = 1
                AND activo = 1
                AND TRIM(COALESCE(valor,'')) = TRIM(COALESCE(d.motivo,''))
                ORDER BY id
            ) pm
            WHERE d.gasto_id = ?
            ORDER BY d.id
        """, (gid,))
        dets = cur.fetchall()

        adjuntos = []
        try:
            cur.execute("""
                SELECT filename
                FROM gastos_tarjeta_archivos
                WHERE gasto_id = ?
                ORDER BY id
            """, (gid,))
            adjuntos = [r['filename'] for r in cur.fetchall() if r and r['filename']]
        except Exception:
            adjuntos = []

        conn.close()

        popup = request.args.get('popup') == '1'
        tpls = ('gastosdetalle_popup.html',) if popup else ('gastosdetalle.html', 'gasto_detalle.html')
        for tpl in tpls:
            try:
                return render_template(tpl, g=row, detalles=dets, popup=popup, adjuntos=adjuntos)
            except TemplateNotFound:
                continue

        return f"<h1>Detalle de gasto #{row['id']}</h1>"
    # DETALLE (visual)
    @app.route('/reembolsos/gastos/<int:gid>/ver', methods=['GET'], endpoint='ver_gasto')
    @require_login
    @require_permission('gastos_tarjeta', 'ver')
    def ver_gasto(gid):
        conn = get_db(); cur = conn.cursor()
        cur.execute(f"""
            SELECT id, fecha, motivo, proveedor, centro_costo,
                   con_soporte, sin_soporte, subtotal_factura,
                   servicios_10, subtotal_sin_iva, iva, total_con_iva, archivo, ccb, usuario_id
            FROM {TABLE_GASTOS} WHERE id=?
        """, (gid,))
        row = cur.fetchone()
        if not row:
            conn.close()
            flash('Gasto no encontrado.', 'warning')
            return redirect(url_for('lista_gastos'))

        cur.execute("""
            SELECT id, descripcion, observacion, centro_costo, motivo, indicador,
                   con_soporte, sin_soporte, subtotal_factura, servicios_10,
                   subtotal_sin_iva, iva, total_con_iva
            FROM gastos_tarjeta_detalle
            WHERE gasto_id=?
            ORDER BY id
        """, (gid,))
        dets = cur.fetchall()
        conn.close()

        for tpl in ('gastosdetalle.html', 'gasto_detalle.html'):
            try:
                return render_template(tpl, g=row, detalles=dets)
            except TemplateNotFound:
                continue
        return f"<h1>Detalle de gasto #{row['id']}</h1>"

   
   
    @app.route('/reembolsos/gastos/<int:gid>/eliminar', methods=['POST'])
    @app.route('/reembolsos/gastos/<int:gasto_id>/eliminar', methods=['POST'])
    @require_login
    @require_permission('gastos_tarjeta', 'eliminar')
    def eliminar_gasto(gid=None, gasto_id=None):
        _id = gid if gid is not None else gasto_id
        conn = get_db(); cur = conn.cursor()
        snapshot = {}

        try:
            # 1) Snapshot antes de borrar
            cur.execute(f"SELECT * FROM {TABLE_GASTOS} WHERE id=?", (_id,))
            row = cur.fetchone()
            if row:
                snapshot = dict(row)

            factura_xml_id = snapshot.get('factura_xml_id')

            # 2) Borrar gasto
            cur.execute(f"DELETE FROM {TABLE_GASTOS} WHERE id=?", (_id,))

            # 3) ✅ Si estaba amarrado a una factura XML, liberar estado si ya nadie la usa
            if factura_xml_id:
                cur.execute("""
                    UPDATE facturas_xml
                    SET estado = 'PENDIENTE'
                    WHERE id = ?
                    AND NOT EXISTS (
                        SELECT 1
                            FROM gastos_tarjeta gt
                            WHERE gt.factura_xml_id = facturas_xml.id
                    )
                """, (factura_xml_id,))

            conn.commit()

            # (debug útil)
            current_app.logger.info(
                "Eliminar gasto id=%s | factura_xml_id=%s | UPDATE facturas_xml aplicado",
                _id, factura_xml_id
            )

        except Exception:
            conn.rollback()
            flash('No se pudo eliminar el registro.', 'danger')
            conn.close()
            return redirect(url_for('lista_gastos'))
        finally:
            conn.close()

        try:
            mail.notify_gasto_deleted(app, snapshot, session.get('usuario_id'))
        except Exception:
            app.logger.exception("No se pudo enviar la notificación de eliminación")

        return redirect(url_for('lista_gastos', status='deleted', t=int(time.time())))

   
   # ----------------------------------------------
    # APROBAR (GA/GG/GF)
    # ----------------------------------------------
 
    RETRY_ATTEMPTS = 3
    RETRY_SLEEP_SEC = 0.35

    @app.post("/reembolsos/gastos/aprobar-masivo", endpoint="aprobar_gasto_masivo")
    @require_login
    def aprobar_gasto_masivo():
        current_app.logger.warning(
            "[APROBAR_MASIVO] HIT path=%s content_type=%s",
            request.path, request.content_type
        )

        from datetime import datetime

        payload = request.get_json(silent=True) or {}
        area = (payload.get("area") or "").lower().strip()
        ids = payload.get("ids") or []
        value = 1 if str(payload.get("value")).lower() in ("1", "true", "on", "yes") else 0

        current_app.logger.warning(
            "[APROBAR_MASIVO] HIT area=%r ids_len=%s value=%r rol=%r uid=%r",
            area, len(ids) if isinstance(ids, list) else None, value,
            session.get("rol"), session.get("usuario_id")
        )

        if area not in ("ga", "gg", "gf"):
            return jsonify(ok=False, msg="Área inválida"), 400
        if not isinstance(ids, list) or not ids:
            return jsonify(ok=False, msg="Sin elementos a aprobar"), 400

        rol = (session.get("rol") or "").lower().strip()
        is_admin = (rol == "admin") or bool(session.get("is_admin"))

        if area == "gg" and not (is_admin or rol == "gerente general"):
            return jsonify(ok=False, msg="No puede aprobar como Gerente General"), 403
        if area == "gf" and not (is_admin or rol == "gerente financiero"):
            return jsonify(ok=False, msg="No puede aprobar como Gerente Financiero"), 403

        GA_ROLES = ("gerente", "gerente de área", "gerente de area")
        GA_ACT_AS_ROLES = ("gerente general", "gerente financiero")

        if area == "ga" and not (is_admin or rol in GA_ROLES or rol in GA_ACT_AS_ROLES):
            return jsonify(ok=False, msg="No puede aprobar como Gerente de área"), 403

        uid = session.get("user_id") or session.get("uid") or session.get("usuario_id") or None
        if not uid:
            return jsonify(ok=False, msg="Sesión inválida: no se encontró usuario aprobador"), 401

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ids_int = []
        failed = []
        for gid in ids:
            try:
                ids_int.append(int(gid))
            except Exception:
                failed.append({"id": gid, "msg": "ID inválido"})

        ids_int = list(dict.fromkeys(ids_int))
        if not ids_int:
            return jsonify(ok=False, msg="No hay IDs válidos para procesar", failed=failed[:50]), 400

        conn = get_db()
        ok_ids = []
        skipped = []

        try:
            cur = conn.cursor()

            placeholders = ",".join(["?"] * len(ids_int))
            cur.execute(f"""
                SELECT
                    g.id,
                    LTRIM(RTRIM(COALESCE(CAST(g.sap_contabilizacion AS NVARCHAR(50)), ''))) AS sap_contabilizacion,
                    g.usuario_id,

                    COALESCE(g.ga_aprobado,0) AS ga_aprobado,
                    COALESCE(g.gg_aprobado,0) AS gg_aprobado,
                    COALESCE(g.gf_aprobado,0) AS gf_aprobado,

                    COALESCE(g.reembolso_vendedor,0) AS reembolso_vendedor,
                    COALESCE(g.es_caja_chica,0) AS es_caja_chica
                FROM {TABLE_GASTOS} g
                WHERE g.id IN ({placeholders})
            """, ids_int)
            rows = cur.fetchall()

            by_id = {}
            for r in rows:
                try:
                    by_id[int(r["id"])] = dict(r)
                except Exception:
                    by_id[int(r[0])] = {
                        "id": r[0],
                        "sap_contabilizacion": r[1],
                        "usuario_id": r[2],
                        "ga_aprobado": r[3],
                        "gg_aprobado": r[4],
                        "gf_aprobado": r[5],
                        "reembolso_vendedor": r[6],
                        "es_caja_chica": r[7],
                    }

            for gid in ids_int:
                if gid not in by_id:
                    skipped.append({"id": gid, "msg": "No existe"})

            subordinados = None
            ga_act_as_tarjeta = False

            if area == "ga" and (not is_admin):
                if rol in ("gerente", "gerente de área", "gerente de area"):
                    subordinados = set(obtener_subordinados(conn, uid) or [])
                    subordinados.add(int(uid))
                elif rol in ("gerente general", "gerente financiero"):
                    ga_act_as_tarjeta = True

            def _has_sap(x) -> bool:
                s = str(x or "").strip()
                return bool(s) and s.lower() not in ("0", "none", "null")

            def _es_tarjeta(row) -> bool:
                return (int(row.get("reembolso_vendedor") or 0) == 0) and (int(row.get("es_caja_chica") or 0) == 0)

            def _actua_como_ga_por_ultimo_jefe(row) -> bool:
                try:
                    ultimo_jefe = _get_ultimo_jefe_id(conn, int(row.get("usuario_id") or 0))
                    return bool(ultimo_jefe) and int(ultimo_jefe) == int(uid)
                except Exception:
                    return False

            if value == 0 and not is_admin:
                return jsonify(ok=False, msg="Solo admin puede desmarcar aprobaciones"), 403

            for gid in ids_int:
                row = by_id.get(gid)
                if not row:
                    continue

                if _has_sap(row.get("sap_contabilizacion")):
                    skipped.append({"id": gid, "msg": "Bloqueado: ya tiene Doc. SAP"})
                    continue

                ga_ok = int(row.get("ga_aprobado") or 0) == 1
                gg_ok = int(row.get("gg_aprobado") or 0) == 1
                gf_ok = int(row.get("gf_aprobado") or 0) == 1

                es_auto_tipo = (int(row.get("reembolso_vendedor") or 0) == 1) or (int(row.get("es_caja_chica") or 0) == 1)
                es_tarjeta = _es_tarjeta(row)

                gf_es_ga = (rol == "gerente financiero") and es_tarjeta and _actua_como_ga_por_ultimo_jefe(row)
                gg_es_ga = (rol == "gerente general") and es_tarjeta and _actua_como_ga_por_ultimo_jefe(row)

                if subordinados is not None:
                    if int(row.get("usuario_id") or 0) not in subordinados:
                        skipped.append({"id": gid, "msg": "Fuera de su jerarquía (GA)"})
                        continue

                if ga_act_as_tarjeta:
                    if not es_tarjeta:
                        skipped.append({"id": gid, "msg": "No aplica: solo tarjeta"})
                        continue

                    ultimo_jefe = _get_ultimo_jefe_id(conn, int(row.get("usuario_id") or 0))
                    if not ultimo_jefe or int(ultimo_jefe) != int(uid):
                        skipped.append({"id": gid, "msg": "No es último jefe (GA tarjeta)"})
                        continue

                effective_area = area
                if area == "ga" and es_auto_tipo and (not is_admin):
                    if rol == "gerente financiero":
                        effective_area = "gf"
                    elif rol == "gerente general":
                        effective_area = "gg"

                if area == "ga":
                    if value == 1 and ga_ok:
                        skipped.append({"id": gid, "msg": "Ya aprobado por GA"})
                        continue

                elif area == "gg":
                    if value == 1 and es_tarjeta and (not gg_es_ga) and (not ga_ok):
                        skipped.append({"id": gid, "msg": "Requiere aprobación GA antes de GG"})
                        continue
                    if value == 1 and gg_ok:
                        skipped.append({"id": gid, "msg": "Ya aprobado por GG"})
                        continue

                elif area == "gf":
                    if value == 1 and (not es_auto_tipo) and es_tarjeta:
                        if not gf_es_ga:
                            if not ga_ok:
                                skipped.append({"id": gid, "msg": "Requiere aprobación GA"})
                                continue
                    if value == 1 and gf_ok:
                        skipped.append({"id": gid, "msg": "Ya aprobado por GF"})
                        continue

                dual_ga_gf = (effective_area == "gf") and es_tarjeta and gf_es_ga and (value == 1)
                dual_ga_gg = (effective_area == "gg") and es_tarjeta and gg_es_ga and (value == 1)

                GG_SET_GA_EN_AUTO_TIPO = True
                dual_ga_gg_auto = (
                    GG_SET_GA_EN_AUTO_TIPO
                    and (effective_area == "gg")
                    and es_auto_tipo
                    and (value == 1)
                )

                if dual_ga_gf:
                    col = "ga+gf"
                elif dual_ga_gg or dual_ga_gg_auto:
                    col = "ga+gg"
                else:
                    col = {"ga": "ga", "gg": "gg", "gf": "gf"}[effective_area]

                if dual_ga_gf:
                    cur.execute(f"""
                        UPDATE {TABLE_GASTOS}
                        SET
                            ga_aprobado = 1,
                            ga_aprobado_por = ?,
                            ga_aprobado_at = ?,
                            gf_aprobado = 1,
                            gf_aprobado_por = ?,
                            gf_aprobado_at = ?
                        WHERE id = ?
                        AND (COALESCE(ga_aprobado,0) <> 1 OR COALESCE(gf_aprobado,0) <> 1)
                    """, (uid, now, uid, now, gid))

                elif dual_ga_gg or dual_ga_gg_auto:
                    cur.execute(f"""
                        UPDATE {TABLE_GASTOS}
                        SET
                            ga_aprobado = 1,
                            ga_aprobado_por = ?,
                            ga_aprobado_at = ?,
                            gg_aprobado = 1,
                            gg_aprobado_por = ?,
                            gg_aprobado_at = ?
                        WHERE id = ?
                        AND (COALESCE(ga_aprobado,0) <> 1 OR COALESCE(gg_aprobado,0) <> 1)
                    """, (uid, now, uid, now, gid))

                else:
                    col_simple = {"ga": "ga", "gg": "gg", "gf": "gf"}[effective_area]
                    cur.execute(f"""
                        UPDATE {TABLE_GASTOS}
                        SET
                            {col_simple}_aprobado = ?,
                            {col_simple}_aprobado_por = CASE WHEN ?=1 THEN ? ELSE NULL END,
                            {col_simple}_aprobado_at = CASE WHEN ?=1 THEN ? ELSE NULL END
                        WHERE id = ?
                        AND COALESCE({col_simple}_aprobado,0) <> ?
                    """, (value, value, uid, value, now, gid, value))

                changed = int(cur.rowcount or 0)

                current_app.logger.warning(
                    "[APROBAR_MASIVO] gid=%s area=%s effective_area=%s col=%s value=%s "
                    "flags: ga_ok=%s gg_ok=%s gf_ok=%s es_tarjeta=%s es_auto_tipo=%s gf_es_ga=%s gg_es_ga=%s "
                    "rowcount=%s changed=%s",
                    gid, area, effective_area, col, value,
                    ga_ok, gg_ok, gf_ok, es_tarjeta, es_auto_tipo, gf_es_ga, gg_es_ga,
                    (cur.rowcount or 0), changed
                )

                if changed > 0:
                    ok_ids.append(gid)
                else:
                    skipped.append({"id": gid, "msg": "Sin cambios (ya estaba en ese estado)"})

            conn.commit()

        except Exception as e:
            conn.rollback()
            current_app.logger.exception("[APROBAR_MASIVO] error")
            return jsonify(ok=False, msg=str(e)), 500
        finally:
            try:
                conn.close()
            except Exception:
                pass

        if value == 1 and ok_ids:
            try:
                from modules.scheduler_jobs import enqueue_gasto_approved
                conn2 = get_db()
                try:
                    for gid_int in ok_ids:
                        enqueue_gasto_approved(conn2, int(gid_int), str(area), uid)
                    conn2.commit()
                finally:
                    conn2.close()
            except Exception:
                current_app.logger.exception("[APROBAR_MASIVO] enqueue notify error")

        return jsonify(
            ok=True,
            area=area,
            value=value,
            approved_count=len(ok_ids),
            updated=len(ok_ids),
            skipped_count=len(skipped),
            failed_count=len(failed),
            skipped=skipped[:100],
            failed=failed[:50],
        ), 200


   
    @app.post("/reembolsos/gastos/<int:gasto_id>/aprobar", endpoint="aprobar_gasto")
    @require_login
    def aprobar_gasto(gasto_id):
        try:
            from datetime import datetime

            payload = request.get_json(silent=True) or {}
            area = (payload.get("area") or "").lower().strip()
            value = 1 if str(payload.get("value")).lower() in ("1", "true", "on", "yes") else 0

            current_app.logger.warning(
                "[APROBAR] HIT gasto_id=%s area=%s value=%s rol=%r is_admin=%r",
                gasto_id, area, value, session.get("rol"), session.get("is_admin")
            )

            if area not in ("ga", "gg", "gf"):
                return jsonify(ok=False, msg="Área inválida"), 400

            rol = (session.get("rol") or "").lower().strip()
            is_admin = (rol == "admin") or bool(session.get("is_admin"))

            if area == "gg" and not (is_admin or rol == "gerente general"):
                return jsonify(ok=False, msg="No puede aprobar como Gerente General"), 403
            if area == "gf" and not (is_admin or rol == "gerente financiero"):
                return jsonify(ok=False, msg="No puede aprobar como Gerente Financiero"), 403

            GA_ROLES = ("gerente", "gerente de área", "gerente de area")
            GA_ACT_AS_ROLES = ("gerente general", "gerente financiero")

            if area == "ga" and not (is_admin or rol in GA_ROLES or rol in GA_ACT_AS_ROLES):
                return jsonify(ok=False, msg="No puede aprobar como Gerente de área"), 403

            uid = session.get("user_id") or session.get("uid") or session.get("usuario_id") or None
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            conn = get_db()
            try:
                cur = conn.cursor()

                cur.execute(f"""
                    SELECT
                        LTRIM(RTRIM(COALESCE(CAST(g.sap_contabilizacion AS NVARCHAR(50)), ''))) AS sap_contabilizacion,
                        g.usuario_id,
                        COALESCE(g.ga_aprobado,0) AS ga_aprobado,
                        COALESCE(g.gg_aprobado,0) AS gg_aprobado,
                        COALESCE(g.gf_aprobado,0) AS gf_aprobado,
                        COALESCE(g.reembolso_vendedor,0) AS reembolso_vendedor,
                        COALESCE(g.es_caja_chica,0) AS es_caja_chica
                    FROM {TABLE_GASTOS} g
                    WHERE g.id = ?
                """, (gasto_id,))
                meta = cur.fetchone()

                if not meta:
                    return jsonify(ok=False, msg="Gasto no encontrado"), 404

                try:
                    sap_doc = (meta["sap_contabilizacion"] or "").strip()
                    usuario_meta = meta["usuario_id"]
                    reembolso_vendedor = int(meta["reembolso_vendedor"] or 0)
                    es_caja_chica = int(meta["es_caja_chica"] or 0)
                except Exception:
                    sap_doc = str(meta[0] or "").strip()
                    usuario_meta = meta[1]
                    reembolso_vendedor = int(meta[5] or 0)
                    es_caja_chica = int(meta[6] or 0)

                if sap_doc:
                    return jsonify(ok=False, msg="No se puede modificar: ya tiene Doc. SAP."), 409

                if area == "ga" and (not is_admin):
                    if rol in ("gerente", "gerente de área", "gerente de area"):
                        subordinados = set(obtener_subordinados(conn, uid) or [])
                        subordinados.add(int(uid))
                        if int(usuario_meta or 0) not in subordinados:
                            return jsonify(ok=False, msg="No puede aprobar gastos fuera de su jerarquía."), 403

                    elif rol in ("gerente general", "gerente financiero"):
                        es_tarjeta = (reembolso_vendedor == 0) and (es_caja_chica == 0)
                        if not es_tarjeta:
                            return jsonify(ok=False, msg="No autorizado para GA en este tipo de gasto."), 403

                        ultimo_jefe = _get_ultimo_jefe_id(conn, int(usuario_meta) if usuario_meta else None)
                        if not ultimo_jefe or int(ultimo_jefe) != int(uid):
                            return jsonify(ok=False, msg="No puede aprobar como GA: no es el último jefe del usuario."), 403

                    else:
                        return jsonify(ok=False, msg="No autorizado (GA)."), 403

                es_auto_tipo = (reembolso_vendedor == 1) or (es_caja_chica == 1)

                effective_area = area
                if area == "ga" and es_auto_tipo and (not is_admin):
                    if rol == "gerente financiero":
                        effective_area = "gf"
                    elif rol == "gerente general":
                        effective_area = "gg"

                col = {"gg": "gg", "gf": "gf", "ga": "ga"}[effective_area]
                cur.execute(f"""
                    UPDATE {TABLE_GASTOS}
                    SET
                        {col}_aprobado = ?,
                        {col}_aprobado_por = CASE WHEN ?=1 THEN ? ELSE NULL END,
                        {col}_aprobado_at = CASE WHEN ?=1 THEN ? ELSE NULL END
                    WHERE id = ?
                    AND COALESCE({col}_aprobado,0) <> ?
                """, (value, value, uid, value, now, gasto_id, value))

                changed = cur.rowcount or 0
                conn.commit()

                if changed == 0:
                    return jsonify(ok=True, value=value, already=True)

                if value == 1:
                    try:
                        from modules.scheduler_jobs import enqueue_gasto_approved

                        conn_notify = get_db()
                        try:
                            enqueue_gasto_approved(
                                conn_notify,
                                int(gasto_id),
                                str(effective_area),
                                uid
                            )
                            conn_notify.commit()
                        finally:
                            try:
                                conn_notify.close()
                            except Exception:
                                pass

                    except Exception:
                        current_app.logger.exception("[APROBAR] Error encolando notificación")               

                return jsonify(ok=True, value=value)

            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        except Exception as e:
            current_app.logger.exception("[APROBAR] ERROR %s", e)
            return jsonify(ok=False, msg=f"Error interno: {e}"), 500
        
    @app.route('/reembolsos/gastos/nuevo', methods=['GET', 'POST'], endpoint='nuevo_gasto')
    @require_login
    @require_permission('gastos_tarjeta', 'crear')
    def nuevo_gasto():
        def render_nuevo(form_override=None, categoria='danger', msg=None):
            if msg:
                flash(msg, categoria)

            form_data = form_override if form_override is not None else {}

            if 'tipo_registro' not in form_data or not (form_data.get('tipo_registro') or '').strip():
                if tiene_caja_chica:
                    form_data['tipo_registro'] = 'caja_chica'
                elif es_comercial_qp:
                    form_data['tipo_registro'] = 'reembolso'
                else:
                    form_data['tipo_registro'] = ''

            return render_template(
                'gastos_form.html',
                modo='nuevo',
                proveedores=proveedores,
                form=form_data,
                archivo=None,
                usuario=session.get('usuario'),
                rol=session.get('rol'),
                active_page='gastos_tarjeta',
                tiene_caja_chica=tiene_caja_chica,
                tipo_caja_chica=tipo_caja_chica,
                es_comercial_qp=es_comercial_qp
            )

        conn = get_db()
        cur = conn.cursor()
        uid = session.get('usuario_id') or session.get('user_id')

        try:
            # 1. Datos del usuario (Caja Chica y Departamento)
            
            cur.execute("""
                SELECT
                    COALESCE(u.tiene_caja_chica, 0) AS tiene_caja_chica,
                    COALESCE(u.tipo_caja_chica, 'NINGUNA') AS tipo_caja_chica,
                    d.nombre AS depto_nombre
                FROM usuarios u
                LEFT JOIN departamentos d ON u.departamento_id = d.id
                WHERE u.id = ?
            """, (uid,))
            res_user = cur.fetchone()

            tiene_caja_chica = bool(res_user['tiene_caja_chica']) if res_user else False
            tipo_caja_chica = (res_user['tipo_caja_chica'] or 'NINGUNA').strip().upper() if res_user else 'NINGUNA'
            nombre_depto = (res_user['depto_nombre'] or "").upper() if res_user else ""
            cur.execute("""
                SELECT d.nombre
                FROM usuarios u
                JOIN departamentos d ON u.departamento_id = d.id
                WHERE u.id = ?
            """, (uid,))
            res_dept = cur.fetchone()
            nombre_depto = (res_dept['nombre'] if res_dept else "SIN DEPARTAMENTO")
            es_comercial_qp = (nombre_depto.upper() == 'COMERCIAL QP')

            # Regla negocio
            es_reembolso_vendedor = bool(tiene_caja_chica) or bool(es_comercial_qp)

            # 2. Catálogo de proveedores
            try:
                cur.execute("""
                    SELECT id, nombre
                    FROM terceros
                    WHERE UPPER(LTRIM(RTRIM(tipo))) = 'P'
                    AND COALESCE(activo, 1) = 1
                    ORDER BY nombre
                """)
                proveedores = cur.fetchall()
            except Exception:
                proveedores = []

            if request.method == 'GET':
                form_preload = {}
                from_xml_raw = (request.args.get('from_xml') or '').strip()
                detalles_xml = []
                if from_xml_raw.isdigit():
                    try:
                        _xml_cur = conn.cursor()
                        _xml_cur.execute("""
                            SELECT
                                f.id,
                                COALESCE(f.estab,'') + '-' + COALESCE(f.pto_emi,'') + '-' +
                                RIGHT(REPLICATE('0',9) + CAST(CAST(COALESCE(f.secuencial,'0') AS INT) AS VARCHAR(9)), 9)
                                    AS nro_factura,
                                f.clave_acceso,
                                f.fecha_emision,
                                f.fecha_autorizacion,
                                f.razon_social_emisor,
                                f.ruc_emisor,
                                f.total
                            FROM facturas_xml f
                            WHERE f.id = ?
                        """, (int(from_xml_raw),))
                        xml_row = _xml_cur.fetchone()
                        if xml_row:
                            cols = [c[0] for c in _xml_cur.description]
                            xr = {cols[i]: xml_row[i] for i in range(len(cols))}
                            # fecha_emision: puede ser string dd/mm/yyyy, yyyy-mm-dd, o date object
                            fe_raw = xr.get('fecha_emision')
                            anio_pre = mes_pre = dia_pre = ''
                            if fe_raw:
                                import datetime as _dt
                                if isinstance(fe_raw, (_dt.date, _dt.datetime)):
                                    anio_pre = str(fe_raw.year)
                                    mes_pre  = f"{fe_raw.month:02d}"
                                    dia_pre  = f"{fe_raw.day:02d}"
                                else:
                                    fe = str(fe_raw).strip()
                                    if len(fe) == 10 and fe[2] == '/':  # dd/mm/yyyy
                                        dia_pre, mes_pre, anio_pre = fe[:2], fe[3:5], fe[6:]
                                    elif len(fe) == 10 and fe[4] == '-':  # yyyy-mm-dd
                                        anio_pre, mes_pre, dia_pre = fe[:4], fe[5:7], fe[8:]
                            form_preload = {
                                'factura_xml_id': str(xr.get('id') or ''),
                                'numero_factura': str(xr.get('nro_factura') or ''),
                                'clave_autorizacion': str(xr.get('clave_acceso') or ''),
                                'fecha_autorizacion': str(xr.get('fecha_autorizacion') or ''),
                                'proveedor': str(xr.get('razon_social_emisor') or ''),
                                'proveedor_identificacion': str(xr.get('ruc_emisor') or ''),
                                'h_total_con_iva': str(xr.get('total') or ''),
                                'anio': anio_pre,
                                'mes': mes_pre,
                                'dia': dia_pre,
                            }
                        # Cargar líneas de detalle desde facturas_xml_det
                        _xml_cur.execute("""
                            SELECT
                                descripcion,
                                cantidad,
                                precio_unitario,
                                COALESCE(descuento, 0)      AS descuento,
                                COALESCE(base_imponible, 0) AS base_imponible,
                                COALESCE(iva, 0)            AS iva,
                                COALESCE(total_linea, 0)    AS total_linea
                            FROM facturas_xml_det
                            WHERE factura_id = ?
                            ORDER BY id
                        """, (int(from_xml_raw),))
                        det_rows = _xml_cur.fetchall()
                        det_cols = [c[0] for c in _xml_cur.description]
                        # Agrupar líneas por indicador (CE/C0) igual que el ingreso manual
                        _grupos = {}  # indicador -> {base, iva, tot}
                        for dr in det_rows:
                            d = {det_cols[i]: dr[i] for i in range(len(det_cols))}
                            base = float(d.get('base_imponible') or 0)
                            iva  = float(d.get('iva') or 0)
                            tot  = float(d.get('total_linea') or 0)
                            if tot == 0:
                                tot = base + iva
                            ind = 'CE' if iva > 0 else 'C0'
                            if ind not in _grupos:
                                _grupos[ind] = {'base': 0.0, 'iva': 0.0, 'tot': 0.0}
                            _grupos[ind]['base'] += base
                            _grupos[ind]['iva']  += iva
                            _grupos[ind]['tot']  += tot
                        detalles_xml = []
                        for ind, g in _grupos.items():
                            base = round(g['base'], 2)
                            iva  = round(g['iva'], 2)
                            tot  = round(g['tot'], 2)
                            detalles_xml.append({
                                'observacion':      '',
                                'descripcion':      '',
                                'motivo':           '',
                                'centro_costo':     '',
                                'indicador':        ind,
                                'con_soporte':      base if ind != 'C0' else 0,
                                'sin_soporte':      0,
                                'subtotal_factura': base,
                                'servicios_10':     0,
                                'subtotal_sin_iva': 0 if ind != 'C0' else base,
                                'iva':              iva,
                                'total_con_iva':    tot,
                            })
                        _xml_cur.close()
                    except Exception as _xml_err:
                        current_app.logger.warning("from_xml preload error id=%s: %s", from_xml_raw, _xml_err)
                        detalles_xml = []
                return render_template(
                    'gastos_form.html',
                    modo='nuevo',
                    proveedores=proveedores,
                    form=form_preload,
                    detalles=detalles_xml if detalles_xml else None,
                    from_xml=bool(form_preload),
                    es_comercial_qp=es_comercial_qp,
                    usuario=session.get('usuario'),
                    tiene_caja_chica=tiene_caja_chica,
                    tipo_caja_chica=tipo_caja_chica,
                    rol=session.get('rol'),
                    active_page='gastos_tarjeta'
                )

            # 3. Manejo POST
            reembolso_vendedor = 1 if es_comercial_qp else (1 if request.form.get('reembolso_vendedor') else 0)

            # recargar catálogo para re-render en errores
            try:
                cur.execute("""
                    SELECT id, nombre
                    FROM terceros
                    WHERE UPPER(LTRIM(RTRIM(tipo))) = 'P'
                    AND COALESCE(activo, 1) = 1
                    ORDER BY nombre
                """)
                proveedores = cur.fetchall()
            except Exception:
                proveedores = []

            # -------- Normalización de entradas --------
            fecha = gh.build_fecha(
                request.form.get('anio'),
                request.form.get('mes'),
                request.form.get('dia')
            )
            try:
                y, m, d = fecha.split('-')
                anio, mes, dia = int(y), int(m), int(d)
            except Exception:
                anio, mes, dia = 2000, 1, 1

            factura_xml_id_raw = (request.form.get('factura_xml_id') or '').strip()
            factura_xml_id = int(factura_xml_id_raw) if factura_xml_id_raw.isdigit() else None

            motivo = (request.form.get('descripcion') or request.form.get('motivo') or '').strip()
            # Fallback: si no hay campo cabecera, tomar el primer det_centro_costo (igual que en editar)
            centro_costo = (request.form.get('centro_costo') or '').strip()
            if not centro_costo:
                _cc_tmp = request.form.getlist('det_centro_costo')
                centro_costo = (_cc_tmp[0] if _cc_tmp else '').strip()
            fecha_autorizacion = (request.form.get('fecha_autorizacion') or '').strip()
            numero_factura = (request.form.get('numero_factura') or '').strip()
            clave_autorizacion = (request.form.get('clave_autorizacion') or '').strip()
            orden_compra = (request.form.get('orden_compra') or '').strip()
            ccb = 1 if str(request.form.get('ccb', '')).lower() in ('on', '1', 'true') else 0
            tarjeta_sin_soporte = 1 if str(request.form.get('tarjeta_sin_soporte', '')).lower() in ('on', '1', 'true') else 0
            boletos_aereos = 1 if str(request.form.get('boletos_aereos', '')).lower() in ('on', '1', 'true') else 0

            # según tu lógica actual
            allow_free_provider = True

            # ── Validación campos requeridos para tipo=tarjeta con CCB ────────
            tipo_registro_raw = (request.form.get('tipo_registro') or '').strip()
            es_tarjeta_ccb = (
                not tipo_registro_raw or tipo_registro_raw == 'tarjeta'
            ) and ccb == 1

            if es_tarjeta_ccb:
                errores_ccb = []
                if not motivo:
                    errores_ccb.append('Detalle / descripción')
                if not numero_factura:
                    errores_ccb.append('N° de factura')
                total_form_ccb = 0.0
                try:
                    total_form_ccb = float(request.form.get('total_con_iva') or 0)
                except (ValueError, TypeError):
                    total_form_ccb = 0.0
                if total_form_ccb <= 0:
                    errores_ccb.append('Total con IVA (debe ser mayor a cero)')
                if errores_ccb:
                    return render_nuevo(
                        dict(request.form),
                        'danger',
                        'Para gastos con cargo a bono (CCB) son obligatorios: '
                        + ', '.join(errores_ccb) + '.'
                    )

            # Duplicado de N° factura
            if numero_factura:
                cur.execute(
                    f"SELECT COUNT(1) AS n FROM {TABLE_GASTOS} WHERE numero_factura = ?",
                    (numero_factura,)
                )
                dup_row = cur.fetchone()
                dup_count = dup_row['n'] if dup_row else 0
                if dup_count > 0:
                    return render_nuevo(dict(request.form), 'danger', 'El N° factura ya existe. Verifique.')

            # ── Validación cruzada factura XML vs formulario ──────────────────
            # Detecta el desacople entre factura_xml_id y numero_factura/total
            # que ocurre cuando el usuario cambia de factura sin que el JS limpie
            # los campos anteriores.
            if factura_xml_id:
                cur.execute("""
                    SELECT
                        COALESCE(estab,'') + '-' + COALESCE(pto_emi,'') + '-' +
                        RIGHT(REPLICATE('0',9) + CAST(CAST(COALESCE(secuencial,'0') AS INT) AS VARCHAR(9)), 9)
                            AS numero_factura,
                        COALESCE(total, 0) AS total_con_iva
                    FROM facturas_xml WHERE id = ?
                """, (factura_xml_id,))
                xml_row = cur.fetchone()
                if xml_row:
                    xml_numero = (xml_row['numero_factura'] or '').strip()
                    xml_total  = float(xml_row['total_con_iva'] or 0)

                    # Verificar que el número de factura del form coincida con el XML
                    if numero_factura and xml_numero and numero_factura != xml_numero:
                        return render_nuevo(
                            dict(request.form), 'danger',
                            f'El N° factura del formulario ({numero_factura}) no coincide '
                            f'con la factura XML seleccionada ({xml_numero}). '
                            'Seleccione la factura correcta.'
                        )

                    # Verificar que el total del form no difiera más de $0.10 del XML
                    total_form = 0.0
                    try:
                        total_form = float(request.form.get('h_total_con_iva') or 0)
                    except (ValueError, TypeError):
                        total_form = 0.0
                    if xml_total > 0 and total_form > 0 and abs(total_form - xml_total) > 0.10:
                        return render_nuevo(
                            dict(request.form), 'danger',
                            f'El total del formulario (${total_form:,.2f}) no coincide '
                            f'con el total de la factura XML seleccionada (${xml_total:,.2f}). '
                            'Verifique que eligió la factura correcta.'
                        )

            # ================== PROVEEDOR: ID / IDENTIFICACIÓN ==================
            proveedor_id_raw = (request.form.get('proveedor_id') or '').strip()
            proveedor_txt = (request.form.get('proveedor') or '').strip()
            proveedor_ident = (request.form.get('proveedor_identificacion') or '').strip()

            # Si viene de factura XML y aún no tenemos id/identificación, intentar autollenar por RUC
            if factura_xml_id and (not proveedor_id_raw and not proveedor_ident):
                cur.execute("""
                    SELECT TOP 1
                        t.id AS prov_id,
                        t.nombre AS prov_nombre,
                        t.identificacion AS prov_ident
                    FROM facturas_xml f
                    JOIN terceros t
                    ON UPPER(LTRIM(RTRIM(t.tipo))) = 'P'
                    AND COALESCE(t.activo, 1) = 1
                    AND LTRIM(RTRIM(COALESCE(t.identificacion, ''))) = LTRIM(RTRIM(COALESCE(f.ruc_emisor, '')))
                    WHERE f.id = ?
                """, (factura_xml_id,))
                vinc = cur.fetchone()

                current_app.logger.info(
                    "[NUEVO_GASTO][AUTO_PROV] factura_xml_id=%s -> prov_id=%s, nombre=%s, ident=%s",
                    factura_xml_id,
                    vinc["prov_id"] if vinc else None,
                    vinc["prov_nombre"] if vinc else None,
                    vinc["prov_ident"] if vinc else None,
                )

                if vinc:
                    proveedor_id_raw = str(vinc["prov_id"])
                    proveedor_ident = (vinc["prov_ident"] or '').strip()
                    if not proveedor_txt:
                        proveedor_txt = vinc["prov_nombre"]

            pid = int(proveedor_id_raw) if proveedor_id_raw.isdigit() else None
            prov_name = None

            # 1) Si ya viene id válido
            if pid:
                cur.execute("""
                    SELECT nombre
                    FROM terceros
                    WHERE id = ?
                    AND UPPER(LTRIM(RTRIM(tipo))) = 'P'
                    AND COALESCE(activo, 1) = 1
                """, (pid,))
                row = cur.fetchone()
                prov_name = row['nombre'] if row else None
                if not prov_name:
                    pid = None

            # 2) Buscar por identificación
            if not pid and proveedor_ident:
                cur.execute("""
                    SELECT TOP 1 id, nombre
                    FROM terceros
                    WHERE UPPER(LTRIM(RTRIM(tipo))) = 'P'
                    AND COALESCE(activo, 1) = 1
                    AND LTRIM(RTRIM(COALESCE(identificacion, ''))) = LTRIM(RTRIM(?))
                """, (proveedor_ident,))
                row = cur.fetchone()
                if row:
                    pid, prov_name = row['id'], row['nombre']

            # 3) Fallback por nombre
            if not pid and proveedor_txt:
                cur.execute("""
                    SELECT TOP 1 id, nombre
                    FROM terceros
                    WHERE UPPER(LTRIM(RTRIM(tipo))) = 'P'
                    AND COALESCE(activo, 1) = 1
                    AND LOWER(LTRIM(RTRIM(COALESCE(nombre, '')))) = LOWER(LTRIM(RTRIM(?)))
                """, (proveedor_txt,))
                row = cur.fetchone()
                if row:
                    pid, prov_name = row['id'], row['nombre']

            current_app.logger.info(
                "[NUEVO_GASTO] factura_xml_id=%r proveedor_id_raw=%r proveedor_ident=%r proveedor_txt=%r pid=%r",
                factura_xml_id_raw, proveedor_id_raw, proveedor_ident, proveedor_txt, pid
            )

            if not pid and not proveedor_txt:
                return render_nuevo(
                    dict(request.form),
                    'danger',
                    'Debe ingresar el nombre del Proveedor.'
                )

            # proveedor libre / autocreación
            if allow_free_provider and not pid:
                if not proveedor_txt or not proveedor_ident:
                    return render_nuevo(
                        dict(request.form),
                        'danger',
                        'Debe ingresar Proveedor y RUC.'
                    )

                cur.execute("""
                    SELECT TOP 1 id, nombre
                    FROM terceros
                    WHERE UPPER(LTRIM(RTRIM(tipo))) = 'P'
                    AND COALESCE(activo, 1) = 1
                    AND LTRIM(RTRIM(COALESCE(identificacion, ''))) = LTRIM(RTRIM(?))
                """, (proveedor_ident,))
                row = cur.fetchone()

                if row:
                    pid = row['id']
                    prov_name = row['nombre']
                else:
                    cur.execute("""
                        INSERT INTO terceros (nombre, identificacion, tipo, activo)
                        OUTPUT INSERTED.id
                        VALUES (?, ?, 'P', 1)
                    """, (proveedor_txt, proveedor_ident))
                    pid_row = cur.fetchone()
                    pid = pid_row[0] if pid_row else None
                    prov_name = proveedor_txt

                    if not pid:
                        raise ValueError("No se pudo recuperar el id del proveedor creado.")

                current_app.logger.info(
                    "[NUEVO_GASTO][AUTO_CREATE_PROV] pid=%r nombre=%r ident=%r",
                    pid, prov_name, proveedor_ident
                )

            proveedor_val = prov_name or proveedor_txt

            # ================== ADJUNTOS MÚLTIPLES ==================
            UPLOAD_DIR = os.path.join(current_app.root_path, "static", "uploads")
            os.makedirs(UPLOAD_DIR, exist_ok=True)

            saved_files = []
            files = []
            try:
                files = request.files.getlist('archivo')
            except Exception:
                files = []

            filename_db = None

            for f in files:
                if not f or not f.filename:
                    continue

                original = secure_filename(f.filename)
                base, ext = os.path.splitext(original)
                ext = (ext or '').lower()

                MAX_BASE = 60
                base = (base or 'archivo')[:MAX_BASE].rstrip(' ._-')
                new_name = f"{base}__{uuid.uuid4().hex}{ext}"
                disk_path = os.path.join(UPLOAD_DIR, new_name)

                try:
                    f.save(disk_path)
                    saved_files.append(new_name)
                    current_app.logger.warning("[NUEVO][ADJ] guardado=%s (orig=%s)", new_name, original)
                except Exception as e:
                    current_app.logger.exception("No se pudo guardar archivo %s: %s", original, e)
                    flash(f"No se pudo guardar el archivo {original}", "danger")

                if ext == ".xml":
                    try:
                        with open(disk_path, "rb") as fh:
                            raw = fh.read()
                        header, detalles = parse_sri_xml(raw)
                        clave = (header.get("clave_acceso") or "").strip()
                        if clave:
                            cur.execute("SELECT id FROM facturas_xml WHERE clave_acceso = ?", (clave,))
                            if cur.fetchone():
                                flash(f"XML duplicado (clave ya existe): {original}", "warning")
                    except Exception as e:
                        current_app.logger.warning("XML no parseable (%s): %s", original, e)

            filename_db = saved_files[0] if saved_files else None

            # ================== OBTENER APROBADOR GA ==================
            gerente = obtener_gerente_real(conn, uid)
            aprobador_ga_id = gerente["id"] if gerente else None
            aprob_ga_user_id = gerente['id'] if gerente else None

            current_app.logger.info("Aprobador GA final asignado: %s", aprobador_ga_id)

            # ================== DETALLE ==================
            def nums(name):
                return [gh.parse_num(x) for x in request.form.getlist(name)]

            descs = request.form.getlist('det_descripcion')
            obss = request.form.getlist('det_observacion')
            centros = request.form.getlist('det_centro_costo')
            motivos = request.form.getlist('det_motivo')
            inds = request.form.getlist('det_indicador')

            current_app.logger.warning(
                "==================== [NUEVO_GASTO] DATOS DEL FORMULARIO (DETALLE) ===================="
            )
            current_app.logger.warning("[NUEVO_GASTO][FORM] archivo adjunto(s): %r", request.files.getlist('archivo') and [f.filename for f in request.files.getlist('archivo') if f and f.filename])
            current_app.logger.warning("[NUEVO_GASTO][FORM] det_descripcion  : %r", descs)
            current_app.logger.warning("[NUEVO_GASTO][FORM] det_observacion  : %r", obss)
            current_app.logger.warning("[NUEVO_GASTO][FORM] det_motivo       : %r", motivos)
            current_app.logger.warning("[NUEVO_GASTO][FORM] det_centro_costo : %r", centros)
            current_app.logger.warning("[NUEVO_GASTO][FORM] det_indicador    : %r", inds)

            cs = nums('det_con_soporte')
            ss = nums('det_sin_soporte')
            sf = nums('det_subtotal_factura')
            s10 = nums('det_servicios_10')
            ssi = nums('det_subtotal_sin_iva')
            ivs = nums('det_iva')
            tci = nums('det_total_con_iva')

            n = max(
                len(descs), len(obss), len(centros), len(motivos), len(inds),
                len(cs), len(ss), len(sf), len(s10), len(ssi), len(ivs), len(tci), 1
            )

            rows = []
            for i in range(n):
                ind = (inds[i].strip().upper() if i < len(inds) and isinstance(inds[i], str) else '')
                csi = cs[i] if i < len(cs) else 0.0
                ssi_i = ssi[i] if i < len(ssi) else 0.0

                base_ref = _q2(csi if (csi or 0) > 0 else ssi_i)
                iva_in_raw = ivs[i] if i < len(ivs) else None
                iva_in_dec = _q2(iva_in_raw)

                tarifa_pct = Decimal("0.00")
                if base_ref > 0 and iva_in_dec > 0:
                    tarifa_pct = (iva_in_dec / base_ref * Decimal("100")).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )

                if not ind:
                    ind = indicador_por_tarifa(tarifa_pct)
                if not ind:
                    ind = IVA_INDICADOR_DEFAULT

                rate = (tarifa_pct / Decimal("100")) if tarifa_pct > 0 else Decimal("0.00")
                iva_in = ivs[i] if i < len(ivs) else None
                iva_calc = float((_q2(csi) * rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                iva_val = float(iva_in) if iva_in is not None else iva_calc

                r = dict(
                    descripcion=(descs[i].strip() if i < len(descs) else ''),
                    observacion=(obss[i].strip() if i < len(obss) else ''),
                    centro_costo=(centros[i].strip() if i < len(centros) else ''),
                    motivo=(motivos[i].strip() if i < len(motivos) else ''),
                    indicador=ind,
                    con_soporte=csi,
                    sin_soporte=(ss[i] if i < len(ss) else 0.0),
                    subtotal_factura=(sf[i] if i < len(sf) else (csi + (ss[i] if i < len(ss) else 0.0))),
                    servicios_10=(s10[i] if i < len(s10) else 0.0),
                    subtotal_sin_iva=ssi_i,
                    iva=iva_val,
                    total_con_iva=(tci[i] if i < len(tci) else 0.0),
                )

                if (
                    not r['descripcion'] and not r['observacion'] and not r['motivo'] and
                    all((r[k] or 0) == 0 for k in (
                        'con_soporte', 'sin_soporte', 'subtotal_factura',
                        'servicios_10', 'subtotal_sin_iva', 'iva', 'total_con_iva'
                    ))
                ):
                    continue

                if not r['total_con_iva']:
                    r['total_con_iva'] = (
                        r['subtotal_factura'] + r['iva'] + r['subtotal_sin_iva'] + r['servicios_10']
                    )

                rows.append(r)

            if not rows:
                rows.append(dict(
                    descripcion='',
                    observacion='',
                    centro_costo='',
                    motivo='',
                    indicador='',
                    con_soporte=0,
                    sin_soporte=0,
                    subtotal_factura=0,
                    servicios_10=0,
                    subtotal_sin_iva=0,
                    iva=0,
                    total_con_iva=0
                ))

            tot = lambda k: sum((r[k] or 0) for r in rows)

            # =========================================
            # Opción A: desglose IVA por tarifa en JSON
            # =========================================
            iva_by_tarifa = {}

            for r in rows:
                base_ref = _q2((r.get("con_soporte") or 0) if (r.get("con_soporte") or 0) > 0 else (r.get("subtotal_sin_iva") or 0))
                iva_val_dec = _q2(r.get("iva"))

                tarifa_pct = Decimal("0.00")
                if base_ref > 0 and iva_val_dec > 0:
                    tarifa_pct = (iva_val_dec / base_ref * Decimal("100")).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )

                key = str(tarifa_pct.quantize(Decimal("0.01")))
                if key not in iva_by_tarifa:
                    iva_by_tarifa[key] = {
                        "tarifa": float(tarifa_pct),
                        "base": Decimal("0.00"),
                        "valor": Decimal("0.00"),
                        "indicador": indicador_por_tarifa(tarifa_pct)
                    }

                iva_by_tarifa[key]["base"] += _q2(r.get("subtotal_sin_iva"))
                iva_by_tarifa[key]["valor"] += iva_val_dec

            impuestos_json = json.dumps({
                "iva": [
                    {
                        "tarifa": v["tarifa"],
                        "base": float(v["base"]),
                        "valor": float(v["valor"]),
                        "indicador": v["indicador"]
                    }
                    for _, v in sorted(iva_by_tarifa.items(), key=lambda kv: Decimal(kv[0]))
                ]
            }, ensure_ascii=False)

            # ================== Tipo de registro ==================
            tipo_registro = request.form.get('tipo_registro')
            es_caja_chica = 0
            es_reembolso_vendedor = 0

            if tipo_registro == 'caja_chica':
                es_caja_chica = 1
                es_reembolso_vendedor = 0
            elif tipo_registro == 'reembolso':
                es_reembolso_vendedor = 1
                es_caja_chica = 0

            if es_caja_chica or es_reembolso_vendedor:
                tarjeta_sin_soporte = 0
                boletos_aereos = 0

            # ── Validación campos requeridos para reembolso y caja chica ─────
            if es_reembolso_vendedor or es_caja_chica:
                tipo_label = 'Reembolso de vendedor' if es_reembolso_vendedor else 'Caja chica'
                errores_tipo = []
                if not motivo:
                    errores_tipo.append('Detalle / descripción')
                try:
                    total_tipo = float(request.form.get('total_con_iva') or 0)
                except (ValueError, TypeError):
                    total_tipo = 0.0
                if total_tipo <= 0:
                    errores_tipo.append('Total con IVA (debe ser mayor a cero)')
                if errores_tipo:
                    return render_nuevo(
                        dict(request.form),
                        'danger',
                        f'Para {tipo_label} son obligatorios: '
                        + ', '.join(errores_tipo) + '.'
                    )

            # ================== ASEGURAR COLUMNAS Y TABLAS EN SQL SERVER ==================
            # gastos_tarjeta_archivos
            try:
                cur.execute("""
                    IF OBJECT_ID(N'dbo.gastos_tarjeta_archivos', N'U') IS NULL
                    BEGIN
                        CREATE TABLE dbo.gastos_tarjeta_archivos(
                            id INT IDENTITY(1,1) NOT NULL PRIMARY KEY,
                            gasto_id INT NOT NULL,
                            filename NVARCHAR(500) NOT NULL,
                            uploaded_at DATETIME NOT NULL DEFAULT GETDATE()
                        );
                    END
                """)
                conn.commit()
            except Exception as _e:
                current_app.logger.warning("[NUEVO_GASTO] No se pudo asegurar gastos_tarjeta_archivos: %s", _e)

            # Columnas centro_costo / motivo / indicador en gastos_tarjeta_detalle
            # La migración SQLite no corre en SQL Server — se aseguran aquí
            for _col, _tipo in [('centro_costo', 'NVARCHAR(200)'),
                                 ('motivo',       'NVARCHAR(300)'),
                                 ('indicador',    'NVARCHAR(10)')]:
                try:
                    cur.execute(f"""
                        IF NOT EXISTS (
                            SELECT 1 FROM sys.columns
                            WHERE object_id = OBJECT_ID(N'dbo.gastos_tarjeta_detalle')
                              AND name = N'{_col}'
                        )
                        ALTER TABLE dbo.gastos_tarjeta_detalle ADD {_col} {_tipo} NULL
                    """)
                    conn.commit()
                except Exception as _e:
                    current_app.logger.warning("[NUEVO_GASTO] No se pudo asegurar columna %s en detalle: %s", _col, _e)

            # Columnas boletos_aereos / tarjeta_sin_soporte en gastos_tarjeta (cabecera)
            for _col, _tipo in [('boletos_aereos',     'BIT'),
                                 ('tarjeta_sin_soporte','BIT')]:
                try:
                    cur.execute(f"""
                        IF NOT EXISTS (
                            SELECT 1 FROM sys.columns
                            WHERE object_id = OBJECT_ID(N'dbo.gastos_tarjeta')
                              AND name = N'{_col}'
                        )
                        ALTER TABLE dbo.gastos_tarjeta ADD {_col} {_tipo} NULL
                    """)
                    conn.commit()
                except Exception as _e:
                    current_app.logger.warning("[NUEVO_GASTO] No se pudo asegurar columna %s en cabecera: %s", _col, _e)

            current_app.logger.warning(
                "==================== [NUEVO_GASTO] FILAS PROCESADAS (pre-INSERT): %d fila(s) ====================",
                len(rows)
            )
            for _i, _r in enumerate(rows):
                current_app.logger.warning(
                    "[NUEVO_GASTO][FILA %d] descripcion='%s' | observacion='%s' | motivo='%s' | centro_costo='%s' | indicador='%s'",
                    _i, _r.get('descripcion',''), _r.get('observacion',''),
                    _r.get('motivo',''), _r.get('centro_costo',''), _r.get('indicador','')
                )

            # ================== ESCRITURA SQL SERVER ==================
            gasto_id = None

            try:
                cur.execute(f"""
                    INSERT INTO {TABLE_GASTOS}
                    (
                        anio, mes, dia, fecha, motivo, proveedor_id, proveedor, centro_costo,
                        con_soporte, sin_soporte, subtotal_factura, servicios_10,
                        subtotal_sin_iva, iva, total_con_iva, archivo, usuario_id,
                        fecha_autorizacion, numero_factura, orden_compra,
                        clave_autorizacion, ccb, factura_xml_id,
                        reembolso_vendedor, es_caja_chica, tarjeta_sin_soporte,
                        impuestos_json, boletos_aereos
                    )
                    OUTPUT INSERTED.id
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    anio, mes, dia, fecha, motivo, pid, proveedor_val, centro_costo,
                    tot('con_soporte'), tot('sin_soporte'), tot('subtotal_factura'), tot('servicios_10'),
                    tot('subtotal_sin_iva'), tot('iva'), tot('total_con_iva'),
                    filename_db, uid,
                    fecha_autorizacion, numero_factura, orden_compra,
                    clave_autorizacion, ccb, factura_xml_id,
                    es_reembolso_vendedor, es_caja_chica, tarjeta_sin_soporte,
                    impuestos_json, boletos_aereos
                ))

                gasto_row = cur.fetchone()
                gasto_id = gasto_row[0] if gasto_row else None
                if not gasto_id:
                    raise ValueError("No se pudo recuperar el id insertado del gasto.")

                if factura_xml_id:
                    cur.execute("""
                        UPDATE facturas_xml
                        SET estado = 'PROCESADO'
                        WHERE id = ?
                        AND estado = 'PENDIENTE'
                    """, (factura_xml_id,))

                for fname in saved_files:
                    cur.execute(
                        "INSERT INTO gastos_tarjeta_archivos(gasto_id, filename) VALUES (?, ?)",
                        (gasto_id, fname)
                    )

                current_app.logger.warning(
                    "==================== [NUEVO_GASTO] ADJUNTOS GUARDADOS: %r ====================",
                    saved_files
                )
                for r in rows:
                    current_app.logger.warning(
                        "==================== [NUEVO_GASTO] INSERT DETALLE (gasto_id=%r) ====================",
                        gasto_id
                    )
                    current_app.logger.warning("[NUEVO_GASTO][DET] descripcion  : '%s'", r.get('descripcion',''))
                    current_app.logger.warning("[NUEVO_GASTO][DET] observacion  : '%s'", r.get('observacion',''))
                    current_app.logger.warning("[NUEVO_GASTO][DET] motivo       : '%s'", r.get('motivo',''))
                    current_app.logger.warning("[NUEVO_GASTO][DET] centro_costo : '%s'", r.get('centro_costo',''))
                    current_app.logger.warning("[NUEVO_GASTO][DET] indicador    : '%s'", r.get('indicador',''))
                    cur.execute("""
                        INSERT INTO gastos_tarjeta_detalle(
                            gasto_id, descripcion, observacion, centro_costo, motivo, indicador,
                            con_soporte, sin_soporte, subtotal_factura, servicios_10,
                            subtotal_sin_iva, iva, total_con_iva
                        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (
                        gasto_id, r['descripcion'], r['observacion'], r['centro_costo'], r['motivo'],
                        r['indicador'], r['con_soporte'], r['sin_soporte'], r['subtotal_factura'],
                        r['servicios_10'], r['subtotal_sin_iva'], r['iva'], r['total_con_iva']
                    ))

                gh.recalc_gasto_totales(conn, gasto_id)
                conn.commit()

            except Exception as e:
                try:
                    conn.rollback()
                except Exception:
                    pass

                current_app.logger.exception("Error insertando gasto: %s", e)
                return render_nuevo(
                    dict(request.form),
                    'danger',
                    'Ocurrió un error al guardar el gasto.'
                )

            # Notificación fuera de transacción
            try:
                if gasto_id:
                    mail.notify_gasto_created(app, gasto_id, uid)
            except Exception:
                app.logger.exception("No se pudo enviar la notificación de creación")

            flash('Gasto creado correctamente.', 'success')
            return redirect(url_for('lista_gastos'))

        finally:
            try:
                conn.close()
            except Exception:
                pass
    
    
    from flask import jsonify, request, session
    from datetime import datetime
    from flask import jsonify, request, session, current_app

    @app.route('/reembolsos/gastos/<int:gasto_id>/rechazar-notificar', methods=['POST'], endpoint='rechazar_gasto_notificar_gerente')
    @require_login
    def rechazar_gasto_notificar_gerente(gasto_id):
        current_app.logger.info("[RECHAZAR_NOTIFICAR] llega gasto_id=%s", gasto_id)

        role = (session.get('rol') or '').strip().lower()
        if role not in ('gerente general', 'admin', 'gerente financiero'):
            return jsonify(ok=False, msg='No autorizado'), 403

        by_uid = session.get('usuario_id') or session.get('user_id')

        data = request.get_json(silent=True) or {}
        comentario = (data.get('comentario') or '').strip()

        if len(comentario) > 1200:
            comentario = comentario[:1200]

        conn = get_db()
        try:
            current_app.logger.warning("DB PATH=%s", current_app.config.get("DATABASE"))

            from modules.scheduler_jobs import enqueue_gasto_rejected_gg

            enqueue_gasto_rejected_gg(
                conn,
                gasto_id=gasto_id,
                by_user_id=by_uid,
                comentario=comentario
            )

            return jsonify(ok=True)

        except Exception as e:
            current_app.logger.exception("[RECHAZAR_NOTIFICAR] error")
            return jsonify(ok=False, msg=str(e)), 500

        finally:
            conn.close()






    @app.route('/reembolsos/facturas-xml/<int:fid>/eliminar',
            methods=['POST'],
            endpoint='factura_xml_eliminar')
    @require_login
    @require_permission('gastos_tarjeta', 'eliminar')
    def factura_xml_eliminar(fid):
        conn = get_db(); cur = conn.cursor()

        # Verificar que exista
        cur.execute("SELECT estado FROM facturas_xml WHERE id=?", (fid,))
        row = cur.fetchone()
        if not row:
            conn.close()
            flash('Factura XML no encontrada.', 'warning')
            return redirect(url_for('facturas_xml_list'))

        estado = (row['estado'] or '').upper()

        # Regla: solo se puede eliminar si NO está procesada
        if estado == 'PROCESADO':
            conn.close()
            flash('No se puede eliminar una factura XML procesada.', 'danger')
            return redirect(url_for('facturas_xml_list'))

        # Extra de seguridad: que no esté vinculada a ningún gasto
        cur.execute(f"SELECT COUNT(1) AS n FROM {TABLE_GASTOS} WHERE factura_xml_id=?", (fid,))
        used = (cur.fetchone() or {'n': 0})['n']
        if used > 0:
            conn.close()
            flash('No se puede eliminar la factura porque está vinculada a uno o más gastos.', 'danger')
            return redirect(url_for('facturas_xml_list'))

        try:
            # Borrar detalle y cabecera
            cur.execute("DELETE FROM facturas_xml_det WHERE factura_id=?", (fid,))
            cur.execute("DELETE FROM facturas_xml WHERE id=?", (fid,))
            conn.commit()
            flash('Factura XML eliminada correctamente.', 'success')
        except Exception as e:
            conn.rollback()
            app.logger.exception("Error eliminando factura_xml %s: %s", fid, e)
            flash('No se pudo eliminar la factura XML.', 'danger')
        finally:
            conn.close()

        return redirect(url_for('facturas_xml_list')) 

 
    
        

     # EDITAR
    @app.route('/reembolsos/gastos/<int:gasto_id>/editar', methods=['GET', 'POST'], endpoint='editar_gasto')
    @require_login
    @require_permission('gastos_tarjeta', 'editar')
    def editar_gasto(gasto_id):

        conn = get_db()
        cur = conn.cursor()

        # --------------------------
        # Rol / alcance (alineado a lista)
        # --------------------------
        role_name = (session.get('rol') or '').lower()
        uid = session.get('usuario_id') or session.get('user_id')

        PRIV_ALL = ('admin', 'coordinador', 'gerente general', 'gerente financiero')
        GERENTE_ROLES = ('gerente', 'gerente de área', 'gerente de area')

        def _allowed_user_ids_for_role():
            if role_name in PRIV_ALL:
                return None
            if role_name in GERENTE_ROLES:
                gerente_id = uid
                subs = obtener_subordinados(conn, gerente_id) or []
                allowed = set(subs)
                if gerente_id:
                    allowed.add(gerente_id)
                return allowed
            return {uid} if uid else set()

        allowed_ids = _allowed_user_ids_for_role()

        # --------------------------
        # Helpers locales
        # --------------------------
        def _parts_fecha(fecha_val):
            if not fecha_val:
                return '', '', ''

            # datetime.date / datetime.datetime
            try:
                if hasattr(fecha_val, 'year') and hasattr(fecha_val, 'month') and hasattr(fecha_val, 'day'):
                    return str(fecha_val.year), int(fecha_val.month), int(fecha_val.day)
            except Exception:
                pass

            # texto
            s = str(fecha_val).strip()
            if not s:
                return '', '', ''

            # si viene con hora: YYYY-MM-DD HH:MM:SS
            if ' ' in s:
                s = s.split(' ')[0].strip()

            # si viene ISO con T
            if 'T' in s:
                s = s.split('T')[0].strip()

            parts = s.split('-')
            if len(parts) == 3:
                y, m, d = parts[0], parts[1], parts[2]
                try:
                    return str(y), int(m), int(d)
                except Exception:
                    return str(y), m, d

            return '', '', ''

        # --------------------------
        # Asegurar tabla de adjuntos y columnas de detalle (SQL Server)
        # --------------------------
        TABLE_GASTOS_ARCH = 'gastos_tarjeta_archivos'
        cur.execute(f"""
            IF OBJECT_ID(N'dbo.{TABLE_GASTOS_ARCH}', N'U') IS NULL
            BEGIN
                CREATE TABLE dbo.{TABLE_GASTOS_ARCH}(
                    id INT IDENTITY(1,1) NOT NULL PRIMARY KEY,
                    gasto_id INT NOT NULL,
                    filename NVARCHAR(500) NOT NULL,
                    uploaded_at DATETIME NOT NULL DEFAULT GETDATE()
                );
            END
        """)
        conn.commit()

        for _col, _tipo in [('centro_costo', 'NVARCHAR(200)'),
                             ('motivo',       'NVARCHAR(300)'),
                             ('indicador',    'NVARCHAR(10)')]:
            try:
                cur.execute(f"""
                    IF NOT EXISTS (
                        SELECT 1 FROM sys.columns
                        WHERE object_id = OBJECT_ID(N'dbo.gastos_tarjeta_detalle')
                          AND name = N'{_col}'
                    )
                    ALTER TABLE dbo.gastos_tarjeta_detalle ADD {_col} {_tipo} NULL
                """)
                conn.commit()
            except Exception as _e:
                current_app.logger.warning("[EDITAR_GASTO] No se pudo asegurar columna %s en detalle: %s", _col, _e)

        # Columnas boletos_aereos / tarjeta_sin_soporte en gastos_tarjeta (cabecera)
        for _col, _tipo in [('boletos_aereos',     'BIT'),
                             ('tarjeta_sin_soporte','BIT')]:
            try:
                cur.execute(f"""
                    IF NOT EXISTS (
                        SELECT 1 FROM sys.columns
                        WHERE object_id = OBJECT_ID(N'dbo.gastos_tarjeta')
                          AND name = N'{_col}'
                    )
                    ALTER TABLE dbo.gastos_tarjeta ADD {_col} {_tipo} NULL
                """)
                conn.commit()
            except Exception as _e:
                current_app.logger.warning("[EDITAR_GASTO] No se pudo asegurar columna %s en cabecera: %s", _col, _e)

        # --------------------------
        # Traer gasto
        # --------------------------
        cur.execute("""
            SELECT
                g.*,
                p.nombre AS proveedor_name,
                p.identificacion AS proveedor_ident
            FROM gastos_tarjeta g
            LEFT JOIN terceros p ON g.proveedor_id = p.id
            WHERE g.id = ?
        """, (gasto_id,))
        g_row = cur.fetchone()

        if not g_row:
            conn.close()
            flash('Gasto no encontrado.', 'warning')
            return redirect(url_for('lista_gastos'))

        g = dict(g_row)

        # Control de acceso
        gasto_uid = g.get('usuario_id')
        if allowed_ids is not None and gasto_uid not in allowed_ids:
            conn.close()
            flash('No tiene acceso para editar este gasto.', 'danger')
            return redirect(url_for('lista_gastos'))

        # ==========================================================
        # POST
        # ==========================================================
        if request.method == 'POST':
            saved_files = []
            try:
                # 1) Captura de datos básicos
                fecha = gh.build_fecha(
                    request.form.get('anio'),
                    request.form.get('mes'),
                    request.form.get('dia')
                )
                motivo = (request.form.get('descripcion') or request.form.get('motivo') or '').strip()

                # cabecera centro_costo: si no existe input cabecera, toma el primero del detalle
                centro_costo = (request.form.get('centro_costo') or '').strip()
                if not centro_costo:
                    centros_tmp = request.form.getlist('det_centro_costo')
                    centro_costo = (centros_tmp[0] if centros_tmp else '') or ''
                    centro_costo = centro_costo.strip()

                # --- LÓGICA DE TIPO DE REGISTRO EXCLUYENTE ---
                tipo_registro = request.form.get('tipo_registro')
                v_es_caja_chica = 1 if tipo_registro == 'caja_chica' else 0
                v_es_reembolso = 1 if tipo_registro == 'reembolso' else 0
                tarjeta_sin_soporte = 1 if str(request.form.get('tarjeta_sin_soporte', '')).lower() in ('on', '1', 'true') else 0
                boletos_aereos = 1 if str(request.form.get('boletos_aereos', '')).lower() in ('on', '1', 'true') else 0

                if v_es_caja_chica or v_es_reembolso:
                    tarjeta_sin_soporte = 0
                if v_es_caja_chica or v_es_reembolso:
                    boletos_aereos = 0

                ccb = 1 if (
                    str(request.form.get('ccb', '')).lower() in ('on', '1', 'true')
                    and not v_es_caja_chica
                    and not v_es_reembolso
                ) else 0

                # ── Validación campos requeridos para tipo=tarjeta con CCB ────
                if ccb == 1 and not v_es_caja_chica and not v_es_reembolso:
                    numero_factura_ed = (request.form.get('numero_factura') or '').strip()
                    errores_ccb = []
                    if not motivo:
                        errores_ccb.append('Detalle / descripción')
                    if not numero_factura_ed:
                        errores_ccb.append('N° de factura')
                    try:
                        total_form_ccb = float(request.form.get('total_con_iva') or 0)
                    except (ValueError, TypeError):
                        total_form_ccb = 0.0
                    if total_form_ccb <= 0:
                        errores_ccb.append('Total con IVA (debe ser mayor a cero)')
                    if errores_ccb:
                        flash(
                            'Para gastos con cargo a bono (CCB) son obligatorios: '
                            + ', '.join(errores_ccb) + '.',
                            'danger'
                        )
                        return redirect(request.url)

                # ── Validación campos requeridos para reembolso y caja chica ─
                if v_es_reembolso or v_es_caja_chica:
                    tipo_label = 'Reembolso de vendedor' if v_es_reembolso else 'Caja chica'
                    errores_tipo = []
                    if not motivo:
                        errores_tipo.append('Detalle / descripción')
                    try:
                        total_tipo = float(request.form.get('total_con_iva') or 0)
                    except (ValueError, TypeError):
                        total_tipo = 0.0
                    if total_tipo <= 0:
                        errores_tipo.append('Total con IVA (debe ser mayor a cero)')
                    if errores_tipo:
                        flash(
                            f'Para {tipo_label} son obligatorios: '
                            + ', '.join(errores_tipo) + '.',
                            'danger'
                        )
                        return redirect(request.url)

                # 2) Resolución de Proveedor
                proveedor_id_raw = (request.form.get('proveedor_id') or '').strip()
                pid = int(proveedor_id_raw) if proveedor_id_raw.isdigit() else g.get('proveedor_id')
                proveedor_val = (request.form.get('proveedor') or g.get('proveedor') or '').strip()

                # 3) Leer detalle y calcular totales
                def nums(name):
                    return [gh.parse_num(x) if gh.parse_num(x) is not None else 0.0 for x in request.form.getlist(name)]

                descs = request.form.getlist('det_descripcion')
                obs = request.form.getlist('det_observacion')
                motivos = request.form.getlist('det_motivo')
                centros = request.form.getlist('det_centro_costo')
                inds = request.form.getlist('det_indicador')

                current_app.logger.warning(
                    "==================== [EDITAR_GASTO %s] DATOS DEL FORMULARIO (DETALLE) ====================",
                    gasto_id
                )
                current_app.logger.warning("[EDITAR_GASTO %s][FORM] archivo adjunto(s): %r", gasto_id, [f.filename for f in (request.files.getlist('archivo') or []) if f and f.filename])
                current_app.logger.warning("[EDITAR_GASTO %s][FORM] det_descripcion  : %r", gasto_id, descs)
                current_app.logger.warning("[EDITAR_GASTO %s][FORM] det_observacion  : %r", gasto_id, obs)
                current_app.logger.warning("[EDITAR_GASTO %s][FORM] det_motivo       : %r", gasto_id, motivos)
                current_app.logger.warning("[EDITAR_GASTO %s][FORM] det_centro_costo : %r", gasto_id, centros)
                current_app.logger.warning("[EDITAR_GASTO %s][FORM] det_indicador    : %r", gasto_id, inds)

                con_sup = nums('det_con_soporte')
                sin_sup = nums('det_sin_soporte')
                sub_fac = nums('det_subtotal_factura')
                serv10 = nums('det_servicios_10')
                sub_sin = nums('det_subtotal_sin_iva')
                ivas = nums('det_iva')
                totals = nums('det_total_con_iva')

                new_rows = []
                for (d, o, m, cc, ind, con, sin, sf, s10, ssi, iva, tot) in zip_longest(
                    descs, obs, motivos, centros, inds,
                    con_sup, sin_sup, sub_fac, serv10, sub_sin, ivas, totals,
                    fillvalue=''
                ):
                    d = (d or '').strip()
                    o = (o or '').strip()
                    m = (m or '').strip()
                    cc = (cc or '').strip()
                    ind = (ind or '').strip()

                    con = float(con) if isinstance(con, (int, float)) else (gh.parse_num(con) or 0.0)
                    sin = float(sin) if isinstance(sin, (int, float)) else (gh.parse_num(sin) or 0.0)
                    sf = float(sf) if isinstance(sf, (int, float)) else (gh.parse_num(sf) or 0.0)
                    s10 = float(s10) if isinstance(s10, (int, float)) else (gh.parse_num(s10) or 0.0)
                    ssi = float(ssi) if isinstance(ssi, (int, float)) else (gh.parse_num(ssi) or 0.0)
                    iva = float(iva) if isinstance(iva, (int, float)) else (gh.parse_num(iva) or 0.0)
                    tot = float(tot) if isinstance(tot, (int, float)) else (gh.parse_num(tot) or 0.0)

                    if not (d or o or m or cc or ind) and (con == 0 and sin == 0 and sf == 0 and s10 == 0 and ssi == 0 and iva == 0 and tot == 0):
                        continue

                    new_rows.append({
                        "descripcion": d,
                        "observacion": o,
                        "motivo": m,
                        "centro_costo": cc,
                        "indicador": ind,
                        "con_soporte": con,
                        "sin_soporte": sin,
                        "subtotal_factura": sf,
                        "servicios_10": s10,
                        "subtotal_sin_iva": ssi,
                        "iva": iva,
                        "total_con_iva": tot,
                    })

                sin_soporte_val = gh.parse_num(request.form.get('sum_sin_soporte'))
                if sin_soporte_val is None:
                    sin_soporte_val = 0.0

                subtotal_factura_val = gh.parse_num(request.form.get('subtotal_factura'))
                if subtotal_factura_val is None:
                    subtotal_factura_val = 0.0

                iva_val = gh.parse_num(request.form.get('iva'))
                if iva_val is None:
                    iva_val = 0.0

                total_con_iva_val = gh.parse_num(request.form.get('total_con_iva'))
                if total_con_iva_val is None:
                    total_con_iva_val = 0.0

                current_app.logger.warning(
                    "==================== [EDITAR_GASTO %s] FILAS PROCESADAS (pre-INSERT): %d fila(s) ====================",
                    gasto_id, len(new_rows)
                )
                for _i, _r in enumerate(new_rows):
                    current_app.logger.warning(
                        "[EDITAR_GASTO %s][FILA %d] descripcion='%s' | observacion='%s' | motivo='%s' | centro_costo='%s' | indicador='%s'",
                        gasto_id, _i, _r.get('descripcion',''), _r.get('observacion',''),
                        _r.get('motivo',''), _r.get('centro_costo',''), _r.get('indicador','')
                    )

                # 4) Manejo de archivos
                UPLOAD_DIR = os.path.join(current_app.root_path, "static", "uploads")
                os.makedirs(UPLOAD_DIR, exist_ok=True)

                try:
                    files = request.files.getlist('archivo')
                except Exception:
                    files = []

                for f in files:
                    if not f or not f.filename:
                        continue

                    original = secure_filename(f.filename)
                    base, ext = os.path.splitext(original)
                    ext = (ext or '').lower()

                    MAX_BASE = 60
                    base = (base or 'archivo')[:MAX_BASE].rstrip(' ._-')
                    new_name = f"{base}__{uuid.uuid4().hex}{ext}"
                    disk_path = os.path.join(UPLOAD_DIR, new_name)

                    try:
                        f.save(disk_path)
                        saved_files.append(new_name)
                        current_app.logger.warning("[EDITAR][ADJ] guardado=%s (orig=%s)", new_name, original)
                    except Exception as e:
                        current_app.logger.exception("No se pudo guardar archivo %s: %s", original, e)
                        flash(f"No se pudo guardar el archivo {original}", "danger")

                filename_db = (g.get('archivo') or '').strip()
                if (not filename_db) and saved_files:
                    filename_db = saved_files[0]
                if not filename_db:
                    filename_db = g.get('archivo') or None

                # 5) TRANSACCIÓN ÚNICA DB
                for fname in saved_files:
                    cur.execute(
                        "INSERT INTO gastos_tarjeta_archivos (gasto_id, filename) VALUES (?, ?)",
                        (gasto_id, fname)
                    )

                current_app.logger.warning(
                    "==================== [EDITAR_GASTO %s] ADJUNTOS GUARDADOS: %r ====================",
                    gasto_id, saved_files
                )
                cur.execute("DELETE FROM gastos_tarjeta_detalle WHERE gasto_id = ?", (gasto_id,))
                for r in new_rows:
                    current_app.logger.warning(
                        "==================== [EDITAR_GASTO %s] INSERT DETALLE ====================", gasto_id
                    )
                    current_app.logger.warning("[EDITAR_GASTO %s][DET] descripcion  : '%s'", gasto_id, r.get('descripcion',''))
                    current_app.logger.warning("[EDITAR_GASTO %s][DET] observacion  : '%s'", gasto_id, r.get('observacion',''))
                    current_app.logger.warning("[EDITAR_GASTO %s][DET] motivo       : '%s'", gasto_id, r.get('motivo',''))
                    current_app.logger.warning("[EDITAR_GASTO %s][DET] centro_costo : '%s'", gasto_id, r.get('centro_costo',''))
                    current_app.logger.warning("[EDITAR_GASTO %s][DET] indicador    : '%s'", gasto_id, r.get('indicador',''))
                    cur.execute("""
                        INSERT INTO gastos_tarjeta_detalle(
                            gasto_id, descripcion, observacion,
                            con_soporte, sin_soporte, subtotal_factura, servicios_10,
                            subtotal_sin_iva, iva, total_con_iva,
                            centro_costo, motivo, indicador
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        gasto_id,
                        r["descripcion"], r["observacion"],
                        r["con_soporte"], r["sin_soporte"], r["subtotal_factura"], r["servicios_10"],
                        r["subtotal_sin_iva"], r["iva"], r["total_con_iva"],
                        r["centro_costo"], r["motivo"], r["indicador"]
                    ))

                cur.execute("""
                    UPDATE gastos_tarjeta SET
                        fecha = ?,
                        motivo = ?,
                        proveedor_id = ?,
                        proveedor = ?,
                        centro_costo = ?,
                        reembolso_vendedor = ?,
                        es_caja_chica = ?,
                        ccb = ?,
                        subtotal_factura = ?,
                        sin_soporte = ?,
                        iva = ?,
                        total_con_iva = ?,
                        archivo = ?,
                        numero_factura = ?,
                        factura_xml_id = ?,
                        tarjeta_sin_soporte = ?,
                        boletos_aereos = ?
                    WHERE id = ?
                """, (
                    fecha,
                    motivo,
                    pid,
                    proveedor_val,
                    centro_costo,
                    v_es_reembolso,
                    v_es_caja_chica,
                    ccb,
                    subtotal_factura_val,
                    sin_soporte_val,
                    iva_val,
                    total_con_iva_val,
                    filename_db,
                    request.form.get('numero_factura'),
                    request.form.get('factura_xml_id'),
                    tarjeta_sin_soporte,
                    boletos_aereos,
                    gasto_id
                ))

                conn.commit()
                conn.close()
                flash('Gasto actualizado correctamente.', 'success')
                return redirect(url_for('lista_gastos'))

            except Exception as e:
                try:
                    conn.rollback()
                except Exception:
                    pass
                current_app.logger.exception("Error editando gasto: %s", e)
                flash(f'Ocurrió un error: {e}', 'danger')
                # cae al GET render

        # ==========================================================
        # GET
        # ==========================================================
        anio, mes, dia = _parts_fecha(g.get('fecha'))

        form_data = {
            'anio': anio,
            'mes': mes,
            'dia': dia,
            'descripcion': g.get('motivo'),
            'proveedor_id': g.get('proveedor_id'),
            'proveedor': g.get('proveedor'),
            'sum_sin_soporte': g.get('sin_soporte'),
            'subtotal_factura': g.get('subtotal_factura'),
            'iva': g.get('iva'),
            'total_con_iva': g.get('total_con_iva'),
            'ccb': g.get('ccb'),
            'es_caja_chica': g.get('es_caja_chica'),
            'reembolso_vendedor': g.get('reembolso_vendedor'),
            'proveedor_identificacion': g.get('proveedor_ident') or '',
            'numero_factura': g.get('numero_factura'),
            'tarjeta_sin_soporte': g.get('tarjeta_sin_soporte', 0),
            'factura_xml_id': g.get('factura_xml_id'),
            'boletos_aereos': g.get('boletos_aereos', 0),
        }

        es_cc = True if int(g.get('es_caja_chica') or 0) == 1 else False
        es_rv = True if int(g.get('reembolso_vendedor') or 0) == 1 else False

        cur.execute("""
            SELECT *
            FROM gastos_tarjeta_detalle
            WHERE gasto_id = ?
            ORDER BY id
        """, (gasto_id,))
        detalles = [dict(r) for r in cur.fetchall()]

        cur.execute("""
            SELECT id, nombre
            FROM terceros
            WHERE UPPER(tipo) = 'P' AND activo = 1
            ORDER BY nombre
        """)
        proveedores = cur.fetchall()

        cur.execute("""
            SELECT filename, uploaded_at
            FROM gastos_tarjeta_archivos
            WHERE gasto_id = ?
            ORDER BY id
        """, (gasto_id,))
        adjuntos = [dict(r) for r in cur.fetchall()]

        legacy = (g.get('archivo') or '').strip()
        if legacy and not any(a.get('filename') == legacy for a in adjuntos):
            adjuntos.insert(0, {"filename": legacy, "uploaded_at": None})

        cur.execute("""
            SELECT COALESCE(tipo_caja_chica, 'NINGUNA') AS tipo_caja_chica
            FROM usuarios
            WHERE id = ?
        """, (g['usuario_id'],))
        u_tipo = cur.fetchone()

        tipo_caja_chica = (
            (u_tipo['tipo_caja_chica'] or 'NINGUNA').strip().upper()
            if u_tipo else 'NINGUNA'
        )

        if not es_cc:
            tipo_caja_chica = 'NINGUNA'

        conn.close()

        return render_template(
            'gastos_form.html',
            modo='editar',
            proveedores=proveedores,
            form=form_data,
            g=g,
            detalles=detalles,
            tiene_caja_chica=es_cc,
            tipo_caja_chica=tipo_caja_chica,
            es_comercial_qp=es_rv,
            usuario=session.get('usuario'),
            rol=session.get('rol'),
            active_page='gastos_tarjeta',
            adjuntos=adjuntos,
        )


    from flask import jsonify, session
    import sqlite3

    @app.route('/api/secuencias/caja-chica/next', methods=['POST'], endpoint='api_next_caja_chica_factura')
    @require_login
    def api_next_caja_chica_factura():
        uid = session.get('usuario_id') or session.get('user_id')
        if not uid:
            return jsonify(ok=False, msg='Sesión inválida'), 401

        conn = get_db()
        cur = conn.cursor()

        seq_key = 'caja_chica_factura'

        try:
            cur.execute("""
                IF NOT EXISTS (
                    SELECT 1
                    FROM user_sequences
                    WHERE usuario_id = ? AND [key] = ?
                )
                INSERT INTO user_sequences (usuario_id, [key], next_value)
                VALUES (?, ?, 1)
            """, (uid, seq_key, uid, seq_key))

            cur.execute("""
                SELECT next_value
                FROM user_sequences
                WHERE usuario_id = ? AND [key] = ?
            """, (uid, seq_key))

            row = cur.fetchone()
            next_val = int(row["next_value"] if row else 1)

            while True:
                candidato = f"CC-{uid}-{next_val:06d}"

                cur.execute("""
                    SELECT TOP 1 1
                    FROM gastos_tarjeta
                    WHERE es_caja_chica = 1
                    AND usuario_id = ?
                    AND LTRIM(RTRIM(COALESCE(numero_factura, ''))) = ?
                """, (uid, candidato))

                if cur.fetchone() is None:
                    break

                next_val += 1

            cur.execute("""
                UPDATE user_sequences
                SET next_value = ?,
                    updated_at = GETDATE()
                WHERE usuario_id = ? AND [key] = ?
            """, (next_val + 1, uid, seq_key))

            conn.commit()
            return jsonify(ok=True, numero=candidato)

        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            return jsonify(ok=False, msg=str(e)), 500

        finally:
            try:
                conn.close()
            except Exception:
                pass
    import sqlite3
    import re
    from flask import request, jsonify, current_app

    @app.route('/api/facturas-xml/search', methods=['GET'])
    @require_login
    def api_facturas_xml_search():
        conn = get_db()
        cur = conn.cursor()

        q = (request.args.get('q') or '').strip()
        try:
            limit = int(request.args.get('limit') or 10)
        except ValueError:
            limit = 10

        limit = max(1, min(limit, 50))

        where = ["1 = 1"]
        params = []

        # Solo pendientes
        where.append("f.estado = 'PENDIENTE'")

        # Excluir facturas ya usadas en gastos_tarjeta
        where.append(f"""
            NOT EXISTS (
                SELECT 1
                FROM {TABLE_GASTOS} g
                WHERE g.factura_xml_id = f.id
                OR LTRIM(RTRIM(COALESCE(g.numero_factura, ''))) = LTRIM(RTRIM(
                        COALESCE(f.estab, '') + '-' +
                        COALESCE(f.pto_emi, '') + '-' +
                        RIGHT(
                            REPLICATE('0', 9) +
                            CAST(CAST(COALESCE(f.secuencial, '0') AS INT) AS VARCHAR(9)),
                            9
                        )
                ))
            )
        """)

        # Misma lógica de búsqueda que la pantalla
        apply_facturas_xml_search(where, params, q, alias="f")

        where_clause = " AND ".join(where)

        sql = f"""
            SELECT TOP ({limit})
                f.*,
                CASE
                    WHEN EXISTS (
                        SELECT 1
                        FROM terceros t
                        WHERE UPPER(LTRIM(RTRIM(COALESCE(t.tipo, '')))) = 'P'
                        AND COALESCE(t.activo, 1) = 1
                        AND LTRIM(RTRIM(COALESCE(t.identificacion, ''))) = LTRIM(RTRIM(COALESCE(f.ruc_emisor, '')))
                    ) THEN 1
                    ELSE 0
                END AS proveedor_ok
            FROM facturas_xml f
            WHERE {where_clause}
            ORDER BY
                TRY_CONVERT(
                    date,
                    SUBSTRING(f.fecha_emision, 7, 4) + '-' +
                    SUBSTRING(f.fecha_emision, 4, 2) + '-' +
                    SUBSTRING(f.fecha_emision, 1, 2)
                ) DESC,
                f.id DESC
        """

        current_app.logger.info("API FACTURAS XML SEARCH SQL: %s | params=%s", sql, params)

        cur.execute(sql, params)
        rows = cur.fetchall()

        data = []
        for r in rows:
            try:
                row = dict(r)
            except Exception:
                cols = [c[0] for c in cur.description]
                row = {cols[i]: r[i] for i in range(len(cols))}

            estab = (row.get('estab') or '').strip()
            pto = (row.get('pto_emi') or '').strip()

            try:
                sec = int(row.get('secuencial') or 0)
                sec_str = f"{sec:09d}"
            except Exception:
                sec_str = str(row.get('secuencial') or '').strip()

            numero = f"{estab}-{pto}-{sec_str}"

            data.append({
                "id": row.get("id"),
                "numero": numero,
                "clave_acceso": row.get("clave_acceso"),
                "fecha_emision": row.get("fecha_emision"),
                "fecha_autorizacion": row.get("fecha_autorizacion"),
                "razon_social_emisor": row.get("razon_social_emisor"),
                "ruc_emisor": row.get("ruc_emisor"),
                "total": row.get("total"),
                "estado": row.get("estado"),
                "proveedor_ok": row.get("proveedor_ok"),
            })

        conn.close()
        return jsonify(data)

    @app.route('/api/gastos/sugerencias-proveedor', methods=['GET'], endpoint='api_sugerencias_proveedor')
    @require_login
    def api_sugerencias_proveedor():
        proveedor = (request.args.get('proveedor') or '').strip()
        proveedor_id_raw = (request.args.get('proveedor_id') or '').strip()

        if not proveedor and not proveedor_id_raw:
            return jsonify([])

        conn = get_db()
        cur = conn.cursor()

        uid = session.get('usuario_id') or session.get('user_id')

        params = []
        where_prov = ""
        if proveedor_id_raw.isdigit():
            where_prov = "AND g.proveedor_id = ?"
            params.append(int(proveedor_id_raw))
        elif proveedor:
            where_prov = "AND LTRIM(RTRIM(LOWER(COALESCE(g.proveedor,'')))) = LTRIM(RTRIM(LOWER(?)))"
            params.append(proveedor)

        where_user = ""
        if uid:
            where_user = "AND g.usuario_id = ?"
            params.append(int(uid))

        # Motivo se guarda en gastos_tarjeta_detalle.motivo (código contable: 6390001001)
        # Hacer JOIN con param_values solo para obtener el nombre legible (label)
        cur.execute(f"""
            SELECT TOP 5
                d.motivo                            AS motivo_cod,
                COALESCE(pv.nombre, d.motivo)       AS motivo_label,
                d.centro_costo,
                COUNT(*)                            AS freq
            FROM gastos_tarjeta_detalle d
            JOIN {TABLE_GASTOS} g ON g.id = d.gasto_id
            LEFT JOIN param_groups pg
                ON LOWER(LTRIM(RTRIM(pg.nombre))) IN ('motivo gasto','motivos','motivo')
            LEFT JOIN param_values pv
                ON pv.group_id = pg.id
                AND LTRIM(RTRIM(LOWER(COALESCE(pv.valor,'')))) = LTRIM(RTRIM(LOWER(COALESCE(d.motivo,''))))
            WHERE d.motivo IS NOT NULL AND LTRIM(RTRIM(d.motivo)) != ''
              AND d.centro_costo IS NOT NULL AND LTRIM(RTRIM(d.centro_costo)) != ''
              {where_prov}
              {where_user}
            GROUP BY d.motivo, COALESCE(pv.nombre, d.motivo), d.centro_costo
            ORDER BY freq DESC
        """, params)

        rows = cur.fetchall()
        cols = [c[0] for c in cur.description]
        conn.close()

        result = []
        for r in rows:
            row = {cols[i]: r[i] for i in range(len(cols))}
            result.append({
                "motivo":       str(row.get("motivo_cod") or ""),
                "motivo_label": str(row.get("motivo_label") or ""),
                "centro_costo": str(row.get("centro_costo") or ""),
                "freq":         int(row.get("freq") or 0),
            })

        return jsonify(result)

    import glob
    import shutil

    @app.post('/reembolsos/xml/procesar-carpeta', endpoint='facturas_xml_procesar_carpeta')
    @require_login
    @require_permission('gastos_tarjeta', 'ver')  # o permiso más restringido
    def facturas_xml_procesar_carpeta():
        conn = get_db()
 
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        folder = current_app.config.get("XML_BULK_FOLDER")
        folder_ok = current_app.config.get("XML_BULK_PROCESADOS")

        os.makedirs(folder, exist_ok=True)
        os.makedirs(folder_ok, exist_ok=True)

        # Procesamos solo un lote por llamada para no tardar demasiado
        MAX_PROC = 300   # por ejemplo 300 XML por clic
        paths = sorted(glob.glob(os.path.join(folder, "*.xml")))[:MAX_PROC]

        if not paths:
            flash("No hay archivos XML pendientes en la carpeta masiva.", "info")
            return redirect(url_for("facturas_xml_list"))

        insertados = duplicados = errores = vacios = 0

        from datetime import datetime
        from werkzeug.utils import secure_filename

        for path in paths:
            fname_orig = os.path.basename(path)

            try:
                with open(path, "rb") as f:
                    raw = f.read()
            except Exception as e:
                current_app.logger.error("No se pudo leer %s: %s", path, e, exc_info=True)
                errores += 1
                continue

            if not raw or not raw.strip():
                vacios += 1
                current_app.logger.error("Archivo XML vacío: %s", fname_orig)
                # mover a procesados para no volver a intentarlo
                shutil.move(path, os.path.join(folder_ok, fname_orig))
                continue

            try:
                header, detalles = parse_sri_xml(raw)
            except Exception as e:
                current_app.logger.error("Error parseando XML %s: %s", fname_orig, e, exc_info=True)
                errores += 1
                # lo movemos a procesados_erroneos si quieres, o lo dejamos
                continue

            # evitar duplicados por clave_acceso
            cur.execute("SELECT id FROM facturas_xml WHERE clave_acceso=?", (header['clave_acceso'],))
            row = cur.fetchone()
            if row:
                duplicados += 1
                # igual lo movemos a procesados para no reintentar
                shutil.move(path, os.path.join(folder_ok, fname_orig))
                continue

            # guardamos el archivo en la misma carpeta de uploads que usas hoy
            safe_name = secure_filename(fname_orig) or f"xml_{int(datetime.now().timestamp())}.xml"
            name, ext = os.path.splitext(safe_name)
            safe_name = f"{name}_{int(datetime.now().timestamp())}{ext}"
            upload_path = os.path.join(
                app.config['UPLOAD_FOLDER'], 'xml_facturas', safe_name
            )
            os.makedirs(os.path.dirname(upload_path), exist_ok=True)
            with open(upload_path, "wb") as out:
                out.write(raw)

            # insertar cabecera
            cur.execute("""
                INSERT INTO facturas_xml (
                    clave_acceso, numero_autorizacion, tipo_comprobante, cod_doc,
                    fecha_emision, fecha_autorizacion,
                    ruc_emisor, razon_social_emisor,
                    ruc_cliente, razon_social_cliente,
                    estab, pto_emi, secuencial,
                    subtotal, descuento, iva, total, moneda,propina,
                    estado, archivo
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                header['clave_acceso'],
                header.get('numero_autorizacion'),
                header.get('tipo_comprobante'),
                header.get('cod_doc'),
                header.get('fecha_emision'),
                header.get('fecha_autorizacion'),
                header.get('ruc_emisor'),
                header.get('razon_social_emisor'),
                header.get('ruc_cliente'),
                header.get('razon_social_cliente'),
                header.get('estab'),
                header.get('pto_emi'),
                header.get('secuencial'),
                header.get('subtotal'),
                header.get('descuento'),
                header.get('iva'),
                header.get('total'),
                header.get('moneda'),
                header.get('propina'),
                'PENDIENTE',
                safe_name
            ))
            factura_id = cur.lastrowid

            # insertar detalle
            for d in detalles:
                cur.execute("""
                    INSERT INTO facturas_xml_det (
                        factura_id, codigo_principal, descripcion,
                        cantidad, precio_unitario, descuento,
                        base_imponible, iva, total_linea
                    ) VALUES (?,?,?,?,?,?,?,?,?)
                """, (
                    factura_id,
                    d.get('codigo_principal'),
                    d.get('descripcion'),
                    d.get('cantidad'),
                    d.get('precio_unitario'),
                    d.get('descuento'),
                    d.get('base_imponible'),
                    d.get('iva'),
                    d.get('total_linea'),
                ))

            insertados += 1
            # mover a carpeta de procesados
            shutil.move(path, os.path.join(folder_ok, fname_orig))

        conn.commit()
        restantes = len(glob.glob(os.path.join(folder, "*.xml")))

        flash(
            f"Lote procesado. Insertados: {insertados} • Duplicados: {duplicados} "
            f"• Errores: {errores} • Vacíos: {vacios} • Restantes en carpeta: {restantes}",
            "success" if insertados else "warning",
        )

        return redirect(url_for("facturas_xml_list"))


    # DASHBOARD y EXPORTS
    # DASHBOARD y EXPORTS
    @app.route('/reembolsos/dashboard', methods=['GET'], endpoint='gastos_dashboard')
    @require_login
    @require_permission('gastos_tarjeta', 'ver')
    def gastos_dashboard():
        # Rango de fechas (yyyy-mm-dd)
        desde, hasta = gh.rango_fechas_desde_request()

        conn = get_db()
        cur = conn.cursor()

        # --- Seguridad/alcance por rol ---
        role_name = (session.get('rol') or '').strip().lower()
        uid = session.get('usuario_id') or session.get('user_id')

        privileged = {'admin', 'gerente general', 'gerente financiero', 'coordinador'}
        gerente_roles = {'gerente', 'gerente de área', 'gerente de area'}

        # -------------------------------
        # Helper: ids visibles por jerarquía
        # -------------------------------
        def _get_subordinates_ids(conn, jefe_id: int) -> list:
            """
            Devuelve lista de IDs (incluye al jefe) con todos sus subordinados
            (directos e indirectos) usando usuarios.jefe_id.
            Compatible con SQL Server.
            """
            if not jefe_id:
                return []

            try:
                rows = conn.execute("""
                    WITH sub AS (
                        SELECT id
                        FROM usuarios
                        WHERE id = ?
                        UNION ALL
                        SELECT u.id
                        FROM usuarios u
                        INNER JOIN sub s ON u.jefe_id = s.id
                    )
                    SELECT id
                    FROM sub
                """, (int(jefe_id),)).fetchall()

                return [r['id'] for r in rows] if rows else [int(jefe_id)]
            except Exception:
                # fallback ultra seguro
                return [int(jefe_id)]

        # WHERE base + único JOIN a usuarios
        where = ["CAST(g.fecha AS date) BETWEEN CAST(? AS date) AND CAST(? AS date)"]
        args = [desde, hasta]
        join_u = "LEFT JOIN usuarios u ON u.id = g.usuario_id"

        # -------------------------------
        # NUEVA LÓGICA DE VISIBILIDAD
        # -------------------------------
        if role_name in privileged:
            # ve todo → no agregamos filtro
            pass

        elif role_name in gerente_roles:
            # gerente → ve gastos de subordinados (y los propios)
            visible_ids = _get_subordinates_ids(conn, uid)
            if not visible_ids:
                # por si acaso
                where.append("g.usuario_id = ?")
                args.append(uid or -1)
            else:
                ph = ",".join(["?"] * len(visible_ids))
                where.append(f"g.usuario_id IN ({ph})")
                args.extend(visible_ids)

        else:
            # usuario normal → solo lo suyo
            where.append("g.usuario_id = ?")
            args.append(uid or -1)

        # -------- KPIs --------
        cur.execute(f"""
            SELECT
                COALESCE(SUM(g.con_soporte),0)       AS con_soporte,
                COALESCE(SUM(g.sin_soporte),0)       AS sin_soporte,
                COALESCE(SUM(g.subtotal_factura),0)  AS subtotal_factura,
                COALESCE(SUM(g.servicios_10),0)      AS servicios_10,
                COALESCE(SUM(g.subtotal_sin_iva),0)  AS subtotal_sin_iva,
                COALESCE(SUM(g.iva),0)               AS iva,
                COALESCE(SUM(g.total_con_iva),0)     AS total_con_iva,
                COUNT(*) AS num_gastos
            FROM {TABLE_GASTOS} g
            {join_u}
            WHERE {" AND ".join(where)}
        """, args)
        kpis = dict(cur.fetchone() or {})

        # -------- Serie por día (línea) --------
        cur.execute(f"""
            SELECT
                CONVERT(varchar(10), CAST(g.fecha AS date), 23) AS d,
                COALESCE(SUM(g.total_con_iva),0) AS total
            FROM {TABLE_GASTOS} g
            {join_u}
            WHERE {" AND ".join(where)}
            GROUP BY CONVERT(varchar(10), CAST(g.fecha AS date), 23)
            ORDER BY d
        """, args)
        serie_dias = cur.fetchall()

        # -------- Top proveedores --------
        cur.execute(f"""
            SELECT TOP 5
                COALESCE(t.nombre, g.proveedor) AS proveedor,
                COALESCE(SUM(g.total_con_iva),0) AS total
            FROM {TABLE_GASTOS} g
            LEFT JOIN terceros t ON t.id = g.proveedor_id
            {join_u}
            WHERE {" AND ".join(where)}
            GROUP BY COALESCE(t.nombre, g.proveedor)
            ORDER BY total DESC
        """, args)
        top_proveedores = cur.fetchall()

        # -------- Top motivos --------
        cur.execute(f"""
            SELECT TOP 5
                COALESCE(NULLIF(LTRIM(RTRIM(g.motivo)), ''), '(sin motivo)') AS motivo,
                COALESCE(SUM(g.total_con_iva),0) AS total
            FROM {TABLE_GASTOS} g
            {join_u}
            WHERE {" AND ".join(where)}
            GROUP BY COALESCE(NULLIF(LTRIM(RTRIM(g.motivo)), ''), '(sin motivo)')
            ORDER BY total DESC
        """, args)
        top_motivos = cur.fetchall()

        # -------- Barras por usuario --------
        cur.execute(f"""
            SELECT TOP 12
                COALESCE(u.username, 'Sin usuario') AS usuario,
                COALESCE(SUM(g.total_con_iva),0) AS total
            FROM {TABLE_GASTOS} g
            {join_u}
            WHERE {" AND ".join(where)}
            GROUP BY COALESCE(u.username, 'Sin usuario')
            ORDER BY total DESC
        """, args)
        serie_usuarios = cur.fetchall()

        # -------- Barras apiladas por mes (Con/Sin soporte) --------
        cur.execute(f"""
            SELECT
                CONVERT(varchar(7), CAST(g.fecha AS date), 23) AS ym,
                COALESCE(SUM(g.con_soporte),0) AS con_sop,
                COALESCE(SUM(g.sin_soporte),0) AS sin_sop
            FROM {TABLE_GASTOS} g
            {join_u}
            WHERE {" AND ".join(where)}
            GROUP BY CONVERT(varchar(7), CAST(g.fecha AS date), 23)
            ORDER BY ym
        """, args)
        por_mes = cur.fetchall()

        # -------- Doughnut CCB vs No CCB --------
        cur.execute(f"""
            SELECT
                COALESCE(SUM(CASE WHEN COALESCE(g.ccb,0)=1 THEN g.total_con_iva ELSE 0 END),0) AS ccb,
                COALESCE(SUM(CASE WHEN COALESCE(g.ccb,0)=0 THEN g.total_con_iva ELSE 0 END),0) AS no_ccb
            FROM {TABLE_GASTOS} g
            {join_u}
            WHERE {" AND ".join(where)}
        """, args)
        ccb_row = dict(cur.fetchone() or {'ccb': 0, 'no_ccb': 0})

        # -------- Embudo de aprobación (conteos) --------
        def _count_extra(extra_where: str) -> int:
            sql = f"""
            SELECT COUNT(*) AS n
            FROM {TABLE_GASTOS} g
            {join_u}
            WHERE {" AND ".join(where + [extra_where])}
            """
            cur.execute(sql, args)
            return int((cur.fetchone() or {'n': 0})['n'])

        cur.execute(f"""
            SELECT COUNT(*) AS n
            FROM {TABLE_GASTOS} g
            {join_u}
            WHERE {" AND ".join(where)}
        """, args)
        n_total = int((cur.fetchone() or {'n': 0})['n'])

        n_sin_sap = _count_extra("(g.sap_contabilizacion IS NULL OR LTRIM(RTRIM(g.sap_contabilizacion))='')")
        n_ga      = _count_extra("COALESCE(g.ga_aprobado,0)=1")
        n_gf      = _count_extra("COALESCE(g.gf_aprobado,0)=1")
        n_gg      = _count_extra("COALESCE(g.gg_aprobado,0)=1")

        # --- Total con IVA por departamento vs mes (solo roles privilegiados) ---
        show_dep_chart = role_name in privileged
        dep_mes_labels, dep_mes_series = [], []

        if show_dep_chart:
            cur.execute(f"""
                SELECT
                    CONVERT(varchar(7), CAST(g.fecha AS date), 23) AS ym,
                    COALESCE(d.nombre, 'Depto ' + CAST(COALESCE(u.departamento_id, 0) AS varchar(20))) AS depto,
                    COALESCE(SUM(g.total_con_iva), 0) AS total
                FROM {TABLE_GASTOS} g
                {join_u}
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                WHERE {" AND ".join(where)}
                GROUP BY
                    CONVERT(varchar(7), CAST(g.fecha AS date), 23),
                    COALESCE(d.nombre, 'Depto ' + CAST(COALESCE(u.departamento_id, 0) AS varchar(20)))
                ORDER BY ym, depto
            """, args)
            rows = [dict(r) for r in cur.fetchall()]

            months = sorted({r['ym'] for r in rows})
            pos = {m: i for i, m in enumerate(months)}
            series_map = {}  # depto -> [vals]

            for r in rows:
                name = r['depto'] or 'Sin depto'
                series_map.setdefault(name, [0.0] * len(months))[pos[r['ym']]] = float(r['total'] or 0)

            dep_mes_labels = months
            dep_mes_series = [{'name': k, 'data': v} for k, v in series_map.items()]

        conn.close()

        return render_template(
            'gastos_dashboard.html',
            desde=desde, hasta=hasta,
            kpis=kpis,
            serie_dias=serie_dias,
            top_proveedores=top_proveedores,
            top_motivos=top_motivos,
            serie_usuarios=serie_usuarios,
            por_mes=por_mes,
            ccb_row=ccb_row,
            n_total=n_total, n_sin_sap=n_sin_sap, n_ga=n_ga, n_gf=n_gf, n_gg=n_gg,
            show_dep_chart=show_dep_chart,
            dep_mes_labels=dep_mes_labels,
            dep_mes_series=dep_mes_series,
            usuario=session.get('usuario'), rol=session.get('rol'),
            active_page='gastos_dashboard'
        )

    @app.route('/reembolsos/xml', methods=['GET', 'POST'], endpoint='facturas_xml_list')
    @require_login
    @require_permission('facturas_xml_list', 'ver')   # o un permiso específico tipo 'facturas_xml'
    def facturas_xml_list():
        from datetime import datetime
        from werkzeug.utils import secure_filename
        from flask import current_app

        conn = get_db()
        cur = conn.cursor()

        # ---------------- POST: CARGA DE XML ----------------
        if request.method == 'POST':
            files = request.files.getlist('xmls') or []
            if not files:
                flash('Seleccione al menos un archivo XML.', 'warning')
                return redirect(url_for('facturas_xml_list'))

            # ⚠️ LIMITE DE ARCHIVOS POR LOTE
            MAX_FILES = 2000
            if len(files) > MAX_FILES:
                flash(
                    f"Has seleccionado {len(files)} archivos. "
                    f"El máximo permitido por carga es {MAX_FILES}. "
                    "Por favor divida la carga en varios grupos.",
                    "warning",
                )
                return redirect(url_for('facturas_xml_list'))

            # ⚠️ LIMITE DE TAMAÑO TOTAL DEL LOTE (por ejemplo 200 MB)
            MAX_BYTES = 200 * 1024 * 1024
            total_bytes = 0
            for f in files:
                if not f or not f.filename:
                    continue
                try:
                    f.stream.seek(0, os.SEEK_END)
                    size = f.stream.tell()
                    f.stream.seek(0)
                except Exception:
                    size = 0
                total_bytes += size
                if total_bytes > MAX_BYTES:
                    flash(
                        "El tamaño total de los XML seleccionados supera el límite permitido "
                        "(200 MB). Por favor súbalos en varios lotes.",
                        "warning",
                    )
                    return redirect(url_for('facturas_xml_list'))

            insertados = duplicados = errores = vacios = 0

            try:
                for f in files:
                    if not f or not f.filename:
                        continue

                    try:
                        f.stream.seek(0)
                    except Exception:
                        pass

                    raw = f.read()

                    current_app.logger.warning(
                        "DEBUG UPLOAD: %s len(raw)=%s", f.filename, len(raw)
                    )

                    if not raw or not raw.strip():
                        vacios += 1
                        current_app.logger.error(
                            "Archivo XML vacío o no leído: %s", f.filename
                        )
                        continue

                    try:
                        header, detalles = parse_sri_xml(raw)
                    except Exception as e:
                        current_app.logger.error(
                            "Error parseando XML %s: %s", f.filename, e, exc_info=True
                        )
                        errores += 1
                        continue

                    # evita cargar 2 veces la misma clave
                    cur.execute(
                        "SELECT id FROM facturas_xml WHERE clave_acceso = ?",
                        (header['clave_acceso'],)
                    )
                    row = cur.fetchone()
                    if row:
                        duplicados += 1
                        continue

                    # guardamos archivo con nombre seguro
                    fname = secure_filename(f.filename)
                    if fname:
                        name, ext = os.path.splitext(fname)
                        fname = f"{name}_{int(datetime.now().timestamp())}{ext}"
                        upload_path = os.path.join(
                            current_app.config['UPLOAD_FOLDER'],
                            'xml_facturas',
                            fname
                        )
                        os.makedirs(os.path.dirname(upload_path), exist_ok=True)
                        with open(upload_path, 'wb') as out:
                            out.write(raw)
                    else:
                        fname = None

                    # valores extra del header
                    base_iva = float(header.get('base_iva') or 0.0)
                    iva_tarifa = float(header.get('iva_tarifa') or 0.0)
                    subtotal_0 = float(header.get('subtotal_0') or 0.0)
                    subtotal_15 = float(header.get('subtotal_15') or 0.0)
                    propina = float(header.get('propina') or 0.0)

                    # INSERT cabecera con OUTPUT INSERTED.id (SQL Server)
                    cur.execute("""
                        INSERT INTO facturas_xml (
                            clave_acceso, numero_autorizacion, tipo_comprobante, cod_doc,
                            fecha_emision, fecha_autorizacion,
                            ruc_emisor, razon_social_emisor,
                            ruc_cliente, razon_social_cliente,
                            estab, pto_emi, secuencial,
                            subtotal, descuento, iva, total, moneda,
                            base_iva, iva_tarifa, subtotal_0, subtotal_15, propina,
                            estado, archivo
                        )
                        OUTPUT INSERTED.id
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (
                        header['clave_acceso'],
                        header.get('numero_autorizacion'),
                        header.get('tipo_comprobante'),
                        header.get('cod_doc'),
                        header.get('fecha_emision'),
                        header.get('fecha_autorizacion'),
                        header.get('ruc_emisor'),
                        header.get('razon_social_emisor'),
                        header.get('ruc_cliente'),
                        header.get('razon_social_cliente'),
                        header.get('estab'),
                        header.get('pto_emi'),
                        header.get('secuencial'),
                        header.get('subtotal'),
                        header.get('descuento'),
                        header.get('iva'),
                        header.get('total'),
                        header.get('moneda'),
                        base_iva,
                        iva_tarifa,
                        subtotal_0,
                        subtotal_15,
                        propina,
                        'PENDIENTE',
                        fname,
                    ))
                    factura_row = cur.fetchone()
                    factura_id = factura_row[0] if factura_row else None

                    if not factura_id:
                        raise ValueError("No se pudo recuperar el id insertado de facturas_xml.")

                    # insertar detalle
                    for d in detalles:
                        cur.execute("""
                            INSERT INTO facturas_xml_det (
                                factura_id, codigo_principal, descripcion,
                                cantidad, precio_unitario, descuento,
                                base_imponible, iva, total_linea
                            ) VALUES (?,?,?,?,?,?,?,?,?)
                        """, (
                            factura_id,
                            d.get('codigo_principal'),
                            d.get('descripcion'),
                            d.get('cantidad'),
                            d.get('precio_unitario'),
                            d.get('descuento'),
                            d.get('base_imponible'),
                            d.get('iva'),
                            d.get('total_linea'),
                        ))

                    insertados += 1

                conn.commit()

            except Exception as e:
                conn.rollback()
                current_app.logger.exception("Error cargando facturas_xml: %s", e)
                flash('Ocurrió un error al cargar los XML.', 'danger')
                conn.close()
                return redirect(url_for('facturas_xml_list'))

            msg = (
                f"XML cargados: {insertados} • "
                f"Duplicados: {duplicados} • "
                f"Errores de parseo: {errores} • "
                f"Archivos vacíos/no leídos: {vacios}"
            )

            flash(msg, 'success' if insertados else 'warning')
            conn.close()
            return redirect(url_for('facturas_xml_list'))

        # ---------------- GET: LISTADO ----------------
        page = request.args.get('page', 1, type=int)
        per_page = 10

        desde = (request.args.get('desde') or '').strip()
        hasta = (request.args.get('hasta') or '').strip()
        emisor = (request.args.get('emisor') or '').strip()
        estado = (request.args.get('estado') or '').strip().upper()

        where_parts = ["1 = 1"]
        params = []

        # fecha_emision viene como dd/mm/yyyy, usamos _fecha_sql() ya adaptado a SQL Server
        if desde:
            where_parts.append(f"{_fecha_sql('f.fecha_emision')} >= CAST(? AS date)")
            params.append(desde)

        if hasta:
            where_parts.append(f"{_fecha_sql('f.fecha_emision')} <= CAST(? AS date)")
            params.append(hasta)

        if emisor:
            where_parts.append("UPPER(COALESCE(f.razon_social_emisor, '')) LIKE UPPER(?)")
            params.append(f"%{emisor}%")

        if estado:
            where_parts.append("f.estado = ?")
            params.append(estado)

        where_clause = " AND ".join(where_parts)

        # Totales
        sql_totales = f"""
            SELECT
                SUM(CASE WHEN f.estado = 'PENDIENTE' THEN 1 ELSE 0 END) AS total_pend,
                SUM(CASE WHEN f.estado = 'PROCESADO' THEN 1 ELSE 0 END) AS total_proc,
                COUNT(*) AS total
            FROM facturas_xml f
            WHERE {where_clause}
        """
        cur.execute(sql_totales, params)
        counts = cur.fetchone()

        if counts is None:
            total_pend = total_proc = total = 0
        else:
            try:
                total_pend = counts["total_pend"] or 0
                total_proc = counts["total_proc"] or 0
                total = counts["total"] or 0
            except Exception:
                total_pend = counts[0] or 0
                total_proc = counts[1] or 0
                total = counts[2] or 0

        total_pages = max(1, (total + per_page - 1) // per_page)
        page = max(1, min(page, total_pages))
        offset = (page - 1) * per_page

        # Lista principal SQL Server: OFFSET/FETCH
        sql_list = f"""
            SELECT
                f.*,
                CASE
                    WHEN EXISTS (
                        SELECT 1
                        FROM terceros t
                        WHERE t.tipo = 'P'
                        AND t.activo = 1
                        AND LTRIM(RTRIM(t.identificacion)) = LTRIM(RTRIM(f.ruc_emisor))
                    ) THEN 1
                    ELSE 0
                END AS proveedor_ok
            FROM facturas_xml f
            WHERE {where_clause}
            ORDER BY {_fecha_sql('f.fecha_emision')} DESC, f.id DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """
        list_params = list(params) + [offset, per_page]

        current_app.logger.info(
            "FACTURAS_XML LIST SQL: %s | params=%s",
            sql_list, list_params
        )

        cur.execute(sql_list, list_params)
        facturas = cur.fetchall()

        conn.close()

        prev_page = page - 1 if page > 1 else None
        next_page = page + 1 if page < total_pages else None

        return render_template(
            'facturas_xml_list.html',
            facturas=facturas,
            usuario=session.get('usuario'),
            rol=session.get('rol'),
            active_page='facturas_xml_list',
            page=page,
            total_pages=total_pages,
            prev_page=prev_page,
            next_page=next_page,
            total=total,
            total_pend=total_pend,
            total_proc=total_proc,
        )
   
   
    
    @app.route('/reembolsos/api/factura_xml/<int:fid>/resumen')
    @require_login
    @require_permission('gastos_tarjeta', 'crear')
    def factura_xml_resumen(fid):
        conn = get_db()
        cur = conn.cursor()

        try:
            cur.execute("""
                SELECT
                    f.id,
                    f.clave_acceso,
                    f.numero_autorizacion,
                    f.tipo_comprobante,
                    f.fecha_emision,
                    f.fecha_autorizacion,
                    f.ruc_emisor,
                    f.razon_social_emisor,
                    f.ruc_cliente,
                    f.razon_social_cliente,
                    f.subtotal,
                    f.descuento,
                    f.iva,
                    f.total,
                    f.base_iva,
                    f.iva_tarifa,
                    f.subtotal_0,
                    f.subtotal_15,
                    f.estado,
                    (
                        COALESCE(f.estab, '') + '-' +
                        COALESCE(f.pto_emi, '') + '-' +
                        RIGHT(
                            REPLICATE('0', 9) +
                            CAST(CAST(COALESCE(f.secuencial, '0') AS INT) AS VARCHAR(9)),
                            9
                        )
                    ) AS numero_factura
                FROM facturas_xml f
                WHERE f.id = ?
            """, (fid,))
            row = cur.fetchone()

            if not row:
                return jsonify({"ok": False, "error": "Factura no encontrada"}), 404

            try:
                row = dict(row)
            except Exception:
                cols = [c[0] for c in cur.description]
                row = {cols[i]: row[i] for i in range(len(cols))}

            payload = {
                "id": row["id"],
                "clave": row["clave_acceso"],
                "numeroAut": row["numero_autorizacion"],
                "tipoComprobante": row["tipo_comprobante"],
                "fechaEmision": row["fecha_emision"],
                "fechaAut": row["fecha_autorizacion"],
                "proveedorRuc": row["ruc_emisor"],
                "proveedorNombre": row["razon_social_emisor"],
                "clienteRuc": row["ruc_cliente"],
                "clienteNombre": row["razon_social_cliente"],

                "subtotal": row["subtotal"] or 0,
                "descuento": row["descuento"] or 0,
                "iva": row["iva"] or 0,
                "iva_total": row["iva"] or 0,
                "total": row["total"] or 0,

                "base_iva": row["base_iva"] or 0,
                "subtotal_0": row["subtotal_0"] or 0,
                "subtotal_15": row["subtotal_15"] or 0,
                "iva_tarifa": row["iva_tarifa"] or 15,

                "serie": row["numero_factura"],
                "estado": row["estado"],
            }

            return jsonify({"ok": True, "payload": payload})

        except Exception as e:
            current_app.logger.exception("Error en factura_xml_resumen(fid=%s): %s", fid, e)
            return jsonify({"ok": False, "error": "Error interno al obtener resumen XML"}), 500

        finally:
            try:
                conn.close()
            except Exception:
                pass

    # EXPORTS
    @app.route('/reembolsos/gastos/export/excel', methods=['GET'], endpoint='export_gastos_excel')
    @require_login
    @require_permission('gastos_tarjeta', 'exportar')
    def _export_excel():
        _log_export_params("EXPORT_EXCEL_IN")
        return gx.export_gastos_excel_response()

    @app.route('/reembolsos/gastos/export/pdf', methods=['GET'], endpoint='export_gastos_pdf')
    @require_login
    @require_permission('gastos_tarjeta', 'exportar')
    def _export_pdf():
        return gx.export_gastos_pdf_response()

    @app.route('/reembolsos/gastos/export/csv', methods=['GET'], endpoint='export_gastos_csv')
    @require_login
    @require_permission('gastos_tarjeta', 'exportar')
    def _export_csv():
        return gx.export_gastos_csv_response()
    
    @app.route('/api/proveedores/search')
    @require_login
    def api_proveedores_search():
        conn = get_db()
        cur = conn.cursor()

        q = (request.args.get('q') or '').strip()
        identificacion = (request.args.get('identificacion') or '').strip()

        try:
            limit = int(request.args.get('limit') or 10)
        except ValueError:
            limit = 10

        limit = max(1, min(limit, 20))

        try:
            if identificacion:
                cur.execute(f"""
                    SELECT TOP ({limit}) id, nombre, identificacion
                    FROM terceros
                    WHERE UPPER(LTRIM(RTRIM(tipo))) = 'P'
                    AND COALESCE(activo, 1) = 1
                    AND REPLACE(COALESCE(identificacion, ''), ' ', '') = REPLACE(?, ' ', '')
                    ORDER BY nombre
                """, (identificacion,))
            elif q:
                like = f"%{q}%"
                cur.execute(f"""
                    SELECT TOP ({limit}) id, nombre, identificacion
                    FROM terceros
                    WHERE UPPER(LTRIM(RTRIM(tipo))) = 'P'
                    AND COALESCE(activo, 1) = 1
                    AND (
                            COALESCE(nombre, '') LIKE ?
                        OR COALESCE(identificacion, '') LIKE ?
                    )
                    ORDER BY nombre
                """, (like, like))
            else:
                cur.execute(f"""
                    SELECT TOP ({limit}) id, nombre, identificacion
                    FROM terceros
                    WHERE UPPER(LTRIM(RTRIM(tipo))) = 'P'
                    AND COALESCE(activo, 1) = 1
                    ORDER BY nombre
                """)

            rows = []
            for r in cur.fetchall():
                try:
                    rows.append(dict(r))
                except Exception:
                    rows.append({
                        "id": r[0],
                        "nombre": r[1],
                        "identificacion": r[2],
                    })

            return jsonify(rows)
        finally:
            conn.close()
    

    @app.route(
        '/reembolsos/gastos/<int:gasto_id>/enviar-sap',
        methods=['POST'],
        endpoint='enviar_gasto_sap'
    )
    @require_login
    @require_permission('gastos_tarjeta', 'ver')
    def enviar_gasto_sap(gasto_id):
        """
        Enviar un gasto individual a SAP.

        Reglas:
        - Solo 'admin' o 'coordinador'.
        - Requiere aprobaciones según el tipo.
        - No permitir si ya tiene sap_contabilizacion.
        - Sociedad SAP se obtiene desde:
            gastos_tarjeta.usuario_id
            -> usuarios.empresa_id
            -> empresas.rep_nacionalidad
        """
        from datetime import datetime
        import json
        import requests
        from decimal import Decimal, ROUND_HALF_UP

        role = (session.get('rol') or '').strip().lower()
        if role not in ('admin', 'coordinador'):
            return jsonify(ok=False, msg='No autorizado'), 403

        conn = get_db()

        try:
            # -----------------------------
            # Config SAP
            # -----------------------------
            cfg = get_sap_config_from_db(conn)

            SAP_URL = cfg["SAP_URL"]
            SAP_URL_QAS = cfg["SAP_URL_QAS"]
            SAP_CLIENT = cfg["SAP_CLIENT"]
            SAP_USER = cfg["SAP_USER"]
            SAP_PASS = cfg["SAP_PASS"]

            if not (SAP_URL or '').startswith(('http://', 'https://')):
                return jsonify(ok=False, msg=f'SAP_URL inválida: {SAP_URL!r}'), 500

            # -----------------------------
            # Helpers Decimal
            # -----------------------------
            Q2 = Decimal("0.01")

            def _D(x):
                try:
                    return Decimal(str(x or "0"))
                except Exception:
                    return Decimal("0")

            def _q2(x: Decimal) -> Decimal:
                return x.quantize(Q2, rounding=ROUND_HALF_UP)

            def _fmt_ddmmyyyy(iso_date):
                try:
                    return datetime.strptime(str(iso_date), "%Y-%m-%d").strftime("%d.%m.%Y")
                except Exception:
                    return ""

            def _distribuir_monto(monto_linea: Decimal, distribucion_cc: list) -> list:
                monto_linea = _q2(monto_linea)
                acumulado = Decimal("0.00")
                out = []
                n = len(distribucion_cc)

                for idx, cc in enumerate(distribucion_cc):
                    if idx == n - 1:
                        monto_cc = _q2(monto_linea - acumulado)
                    else:
                        porc = _D(cc.get("porcentaje")) / Decimal("100")
                        monto_cc = _q2(monto_linea * porc)
                        acumulado += monto_cc

                    out.append(monto_cc)

                return out

            def _row_to_dict(row):
                try:
                    return dict(row)
                except Exception:
                    return None

            cur = conn.cursor()

            # -----------------------------
            # Cabecera
            # -----------------------------
            cur.execute(f"""
                SELECT
                    g.*,

                    u.username AS usuario_username,
                    u.codigo_sap AS usuario_codigo_sap,
                    u.identificacion AS usuario_cedula,
                    u.departamento_id,
                    u.empresa_id AS usuario_empresa_id,
                    COALESCE(u.tiene_caja_chica, 0) AS tiene_caja_chica,
                    COALESCE(u.tipo_caja_chica, 'NINGUNA') AS tipo_caja_chica,

                    t.codigo_sap AS proveedor_codigo_sap,
 
                    e.razon_social AS empresa_razon_social,
                    LTRIM(RTRIM(COALESCE(e.rep_nacionalidad, ''))) AS empresa_sociedad,
                    LTRIM(RTRIM(COALESCE(e.usuario_sap, ''))) AS empresa_usuario_sap

                FROM {TABLE_GASTOS} g
                LEFT JOIN usuarios u
                    ON u.id = g.usuario_id
                LEFT JOIN terceros t
                    ON t.id = g.proveedor_id
                LEFT JOIN empresas e
                    ON e.id = u.empresa_id
                WHERE g.id = ?
            """, (gasto_id,))
            g_row = cur.fetchone()

            if not g_row:
                return jsonify(ok=False, msg='Gasto no encontrado'), 404

            g = _row_to_dict(g_row) or {}

            sociedad_sap = (g.get("empresa_sociedad") or "").strip().upper()
            usuario_sap = (g.get("empresa_usuario_sap") or "").strip().upper()

            if not usuario_sap:
                return jsonify(
                    ok=False,
                    msg=(
                        "No se pudo determinar el Usuario SAP. "
                        "Revise que la empresa tenga configurado el campo usuario_sap."
                    )
                ), 400

            if not sociedad_sap:
                return jsonify(
                    ok=False,
                    msg=(
                        "No se pudo determinar la Sociedad SAP. "
                        "Revise que el usuario tenga empresa asignada y que la empresa tenga "
                        "configurado el campo rep_nacionalidad."
                    )
                ), 400

            es_caja_chica = int(g.get('es_caja_chica') or 0) == 1
            es_reembolso = int(g.get('reembolso_vendedor') or 0) == 1
            es_tarjeta_online = int(g.get('tarjeta_sin_soporte') or 0) == 1
            es_boletos_aereos = int(g.get('boletos_aereos') or 0) == 1

            tipo_caja_chica = (g.get("tipo_caja_chica") or "NINGUNA").strip().upper()
            es_caja_chica_con_detalle = es_caja_chica and tipo_caja_chica == "DETALLE_FACTURA"

            es_tipo_3 = es_caja_chica or es_reembolso or es_tarjeta_online
            es_tipo_4 = es_boletos_aereos

            proveedor_codigo_sap = (g.get('proveedor_codigo_sap') or "").strip()
            acreedor_tipo4 = ""

            if es_tipo_4:
                if not proveedor_codigo_sap:
                    return jsonify(ok=False, msg="Proveedor sin código SAP (terceros.codigo_sap)."), 400
                acreedor_tipo4 = proveedor_codigo_sap

            acreedor_tipo3 = (g.get('usuario_codigo_sap') or "").strip()

            if es_tarjeta_online:
                gerente_id = _get_ultimo_jefe_id(conn, g.get('usuario_id'))

                if gerente_id:
                    cur.execute(
                        "SELECT COALESCE(codigo_sap,'') AS codigo_sap FROM usuarios WHERE id = ?",
                        (gerente_id,)
                    )
                    r = cur.fetchone()

                    codigo_gerente = ""
                    if r:
                        try:
                            codigo_gerente = (r["codigo_sap"] or "").strip()
                        except Exception:
                            codigo_gerente = (r[0] or "").strip()

                    if codigo_gerente:
                        acreedor_tipo3 = codigo_gerente

            ga_ok = int(g.get('ga_aprobado') or 0) == 1
            gg_ok = int(g.get('gg_aprobado') or 0) == 1
            gf_ok = int(g.get('gf_aprobado') or 0) == 1

            def _norm_actor(gasto):
                return (
                    gasto.get("ga_actor")
                    or gasto.get("ga_aprobador")
                    or gasto.get("ga_aprobador_rol")
                    or ""
                ).strip().lower()

            def _ga_effective_ok(gasto, ga_ok_, gg_ok_, gf_ok_, es_restringido_):
                if not es_restringido_:
                    return ga_ok_

                actor = _norm_actor(gasto)

                if actor in ("gg", "gerente general", "gerente_general"):
                    return gg_ok_

                if actor in ("gf", "gerente financiero", "gerente_financiero"):
                    return gf_ok_

                if (not ga_ok_) and (gg_ok_ or gf_ok_):
                    return True

                return ga_ok_

            es_restringido = es_caja_chica or es_reembolso
            ga_eff_ok = _ga_effective_ok(g, ga_ok, gg_ok, gf_ok, es_restringido)

            if es_restringido:
                if not ga_eff_ok:
                    return jsonify(ok=False, msg='Requiere aprobación GA (o su rol sustituto)'), 400
            else:
                if not (ga_ok and gg_ok and gf_ok):
                    return jsonify(ok=False, msg='Requiere aprobación de GA/GG/GF'), 400

            if (g.get('sap_contabilizacion') or '').strip():
                return jsonify(ok=False, msg='Ya fue enviado a SAP'), 400

            cur.execute("SELECT identificacion FROM usuarios WHERE id = ?", (g['usuario_id'],))
            u_creador = cur.fetchone()

            try:
                cedula_creador = (u_creador['identificacion'] if u_creador else "") or ""
            except Exception:
                cedula_creador = (u_creador[0] if u_creador else "") or ""

            # -----------------------------
            # Detalle
            # -----------------------------
            cur.execute("""
                SELECT
                    descripcion,
                    observacion,
                    centro_costo,
                    motivo,
                    indicador,
                    con_soporte,
                    sin_soporte,
                    subtotal_factura,
                    servicios_10,
                    subtotal_sin_iva,
                    iva,
                    total_con_iva
                FROM gastos_tarjeta_detalle
                WHERE gasto_id = ?
                ORDER BY id
            """, (gasto_id,))
            dets = []

            for r in cur.fetchall():
                try:
                    dets.append(dict(r))
                except Exception:
                    dets.append({
                        "descripcion": r[0],
                        "observacion": r[1],
                        "centro_costo": r[2],
                        "motivo": r[3],
                        "indicador": r[4],
                        "con_soporte": r[5],
                        "sin_soporte": r[6],
                        "subtotal_factura": r[7],
                        "servicios_10": r[8],
                        "subtotal_sin_iva": r[9],
                        "iva": r[10],
                        "total_con_iva": r[11],
                    })

            if not dets:
                return jsonify(ok=False, msg='El gasto no tiene detalle'), 400

            # -----------------------------
            # Totales + fechas
            # -----------------------------
            total_doc = sum(_D(d.get('total_con_iva')) for d in dets)
            importe_total_con_iva_str = f"{_q2(total_doc):.2f}"

            fecha_doc_str = _fmt_ddmmyyyy(g.get('fecha'))
            orden_compra = (g.get('orden_compra') or '').strip()
            tiene_oc = bool(orden_compra)

            # -----------------------------
            # Tipo 1/2 base
            # -----------------------------
            header = {
                "Agrupador": "1",
                "Tipo_Ejecucion": "1",
                "Sociedad": sociedad_sap,
                "Acreedor": (g.get('proveedor_codigo_sap') or str(g.get('proveedor_id') or "")),
                "Referencia": g.get('numero_factura') or "",
                "Fecha_Documento": fecha_doc_str,
                "Importe_Total": importe_total_con_iva_str,
                "Moneda": "USD",
                "Clase_Documento": "KR",
                "Lugar": "0001",
                "Texto_Cabecera": g.get('motivo') or "",
                "Num_Tienda": "0001",
                "Clave_Referencia_3": "01",
                "Via_Pago": "T",
                "Condicion_Pago": "Z01D",
            }

            if tiene_oc:
                header["Tipo_Ejecucion"] = "2"

            detalle = []

            for idx, d in enumerate(dets, start=1):
                total_con_iva_linea = float(d.get('total_con_iva') or 0.0)
                base_linea = float(d.get('subtotal_factura') or 0.0)

                if tiene_oc:
                    importe_pos = base_linea if base_linea > 0 else total_con_iva_linea
                else:
                    importe_pos = total_con_iva_linea

                pedido_posicion = str(idx * 10) if tiene_oc else ""

                detalle.append({
                    "Cuenta_Mayor": d.get('motivo') or "",
                    "Importe_Posicion": f"{importe_pos:.2f}",
                    "Texto_Posicion": g.get('motivo') or "",
                    "Division": "CORP",
                    "Centro_Costo": d.get('centro_costo') or "",
                    "Indicador_Iva": d.get('indicador') or "",
                    "Pedido": orden_compra if tiene_oc else "",
                    "Pedido_Posicion": pedido_posicion,
                    "Asignacion": "",
                })

            try:
                secuencia_sap = obtener_siguiente_secuencia(conn, 'reembolsos')
            except Exception as e:
                current_app.logger.error("Error al generar secuencia: %s", e)
                secuencia_sap = "0000020782"

            # ==========================================================
            # TIPO 3 / TIPO 4
            # ==========================================================
            if es_tipo_3 or es_tipo_4:
                fecha_liq = _fmt_ddmmyyyy(g.get('fecha'))

                distribucion_cc = []

                if es_reembolso:
                    id_dueno = g['usuario_id']

                    cur.execute("""
                        SELECT
                            ucc.porcentaje,
                            pv.valor AS codigo_cc_sap,
                            pv.nombre AS nombre_cc
                        FROM usuarios_cc ucc
                        JOIN param_values pv
                            ON pv.id = ucc.centro_costo_id
                        WHERE ucc.usuario_id = ?
                        AND pv.activo = 1
                    """, (id_dueno,))

                    distribucion_cc = []
                    for r in cur.fetchall():
                        try:
                            distribucion_cc.append(dict(r))
                        except Exception:
                            distribucion_cc.append({
                                "porcentaje": r[0],
                                "codigo_cc_sap": r[1],
                                "nombre_cc": r[2],
                            })

                    if not distribucion_cc:
                        return jsonify(
                            ok=False,
                            msg=f"Error: El usuario {g.get('usuario_username')} no tiene centros de costo."
                        ), 400

                def _make_liq_documento():
                    return {
                        "Fecha_Documento": fecha_doc_str,
                        "Tipo_Id": "51",
                        "Id_Fiscal": cedula_creador,
                        "Moneda": "USD",
                        "Tipo_Doc_Liquidacion": "01",
                        "Cod_Doc_Liquidacion": g.get('estab') or "001",
                        "Serie_Doc_Liquidacion": "002",
                        "Numero_Doc_Liquidacion": secuencia_sap,
                        "Numero_Autorizacion": cedula_creador,
                        "Liquidacion_Detalle": []
                    }

                documento = {
                    "Agrupador": "1",
                    "Tipo_Ejecucion": "3",
                    "Sociedad": sociedad_sap,
                    "Acreedor": acreedor_tipo4 if es_tipo_4 else acreedor_tipo3,
                    "Referencia": g.get('numero_factura') or "",
                    "Importe_Total": importe_total_con_iva_str,
                    "Moneda": "USD",
                    "Clase_Documento": "KP" if es_tipo_4 else "KA",
                    "Lugar": "0001",
                    "Via_Pago": "T",
                    "Usuario_SAP": usuario_sap,
                    "Fecha_Liquidacion": fecha_liq,
                    "Motivo_Liquidacion": "REEMBOLSO DE GASTOS",
                    "Liquidacion_Documentos": []
                }

                total_tipo3 = Decimal("0.00")
                motivo_gasto = (g.get("motivo") or "").strip()

                def _indicador_liquidacion(detalle_linea):
                    """
                    Mantiene la lógica actual para caja chica normal/reembolso/tarjeta online:
                    - C0

                    Pero si es Caja Chica con Detalle:
                    - usa el indicador de la línea del XML/detalle.
                    """
                    if es_caja_chica_con_detalle:
                        return (detalle_linea.get("indicador") or "C0").strip().upper()

                    return "C0"

                for d in dets:
                    monto_linea = _D(d.get("total_con_iva"))

                    if monto_linea <= 0:
                        continue

                    cuenta_gasto = d.get("motivo") or ""
                    division = "CORP"
                    indicador_iva = _indicador_liquidacion(d)

                    if es_caja_chica or es_tarjeta_online or es_tipo_4:
                        cc_linea = (d.get('centro_costo') or '').strip()

                        if not cc_linea:
                            return jsonify(
                                ok=False,
                                msg='Caja chica requiere Centro de Costo en el detalle.'
                            ), 400

                        liq_doc = _make_liq_documento()

                        liq_doc["Liquidacion_Detalle"].append({
                            "Cuenta_Gasto": cuenta_gasto,
                            "Division": division,
                            "Indicador_Iva": indicador_iva,
                            "Centro_Costo": cc_linea,
                            "Importe_Posicion": f"{_q2(monto_linea):.2f}",
                            "Motivo_Detalle": motivo_gasto
                        })

                        documento["Liquidacion_Documentos"].append(liq_doc)
                        total_tipo3 += monto_linea

                    else:
                        montos_cc = _distribuir_monto(monto_linea, distribucion_cc)

                        for cc_row, m in zip(distribucion_cc, montos_cc):
                            liq_doc = _make_liq_documento()

                            liq_doc["Liquidacion_Detalle"].append({
                                "Cuenta_Gasto": cuenta_gasto,
                                "Division": division,
                                "Indicador_Iva": indicador_iva,
                                "Centro_Costo": cc_row["codigo_cc_sap"],
                                "Importe_Posicion": f"{m:.2f}",
                                "Motivo_Detalle": motivo_gasto
                            })

                            documento["Liquidacion_Documentos"].append(liq_doc)

                        total_tipo3 += monto_linea

                documento["Importe_Total"] = f"{_q2(total_tipo3):.2f}"
                payload = [documento]

            else:
                documento_std = header.copy()
                documento_std["Detalle"] = detalle
                payload = [documento_std]

            trama_json = json.dumps(payload, indent=2, ensure_ascii=False)
            current_app.logger.info("Payload SAP (gasto_id=%s):\n%s", gasto_id, trama_json)

            # -----------------------------
            # Llamada a SAP
            # -----------------------------
            try:
                r = requests.post(
                    SAP_URL,
                    params={"sap-client": SAP_CLIENT},
                    json=payload,
                    auth=(SAP_USER, SAP_PASS),
                    timeout=30,
                )
            except Exception as e:
                cur.execute(
                    f"UPDATE {TABLE_GASTOS} "
                    "SET sap_error_msg = ?, sap_enviado_at = ? "
                    "WHERE id = ?",
                    (f'EXC: {e}', datetime.utcnow().isoformat(timespec='seconds'), gasto_id)
                )
                conn.commit()
                return jsonify(ok=False, msg='No se pudo conectar a SAP'), 502

            try:
                resp = r.json()
            except Exception:
                resp = {"raw": r.text}

            current_app.logger.info(
                "Respuesta SAP (status=%s, gasto_id=%s): %s",
                r.status_code,
                gasto_id,
                json.dumps(resp, indent=2, ensure_ascii=False)
                if isinstance(resp, (dict, list))
                else str(resp)
            )

            error_sap = _extraer_error_sap(resp)

            if error_sap:
                cur.execute(
                    f"UPDATE {TABLE_GASTOS} "
                    "SET sap_error_msg = ?, sap_response_json = ?, sap_enviado_at = ? "
                    "WHERE id = ?",
                    (
                        error_sap,
                        json.dumps(resp, ensure_ascii=False)[:4000],
                        datetime.utcnow().isoformat(timespec="seconds"),
                        gasto_id
                    )
                )
                conn.commit()
                return jsonify(ok=False, msg=error_sap), 400

            if r.status_code >= 400:
                err = None

                if isinstance(resp, dict):
                    err = resp.get('error') or resp.get('message')

                if not err:
                    err = f'HTTP {r.status_code}'

                cur.execute(
                    f"UPDATE {TABLE_GASTOS} "
                    "SET sap_error_msg = ?, sap_response_json = ?, sap_enviado_at = ? "
                    "WHERE id = ?",
                    (
                        str(err),
                        json.dumps(resp, ensure_ascii=False)[:4000],
                        datetime.utcnow().isoformat(timespec='seconds'),
                        gasto_id
                    )
                )
                conn.commit()
                return jsonify(ok=False, msg=str(err)), 400

            # -----------------------------
            # Extraer doc
            # -----------------------------
            doc = ''
            first = None

            if isinstance(resp, list) and resp:
                first = resp[0]
            elif isinstance(resp, dict):
                first = resp

            if isinstance(first, dict):
                doc = (
                    first.get('doc_number')
                    or first.get('document')
                    or first.get('num_doc')
                    or first.get('DOCUMENTO')
                    or first.get('documento')
                    or ''
                )

            cur.execute(f"""
                UPDATE {TABLE_GASTOS}
                SET sap_contabilizacion = ?,
                    sap_response_json   = ?,
                    sap_enviado_at      = ?,
                    sap_error_msg       = NULL
                WHERE id = ?
            """, (
                doc,
                json.dumps(resp, ensure_ascii=False)[:4000],
                datetime.utcnow().isoformat(timespec='seconds'),
                gasto_id
            ))
            conn.commit()

            return jsonify(ok=True, doc=doc or None)

        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass

            current_app.logger.exception("ERROR enviar_gasto_sap")
            return jsonify(ok=False, msg=str(e)), 500

        finally:
            try:
                conn.close()
            except Exception:
                pass

    @app.route(
        '/reembolsos/gastos/enviar-sap-masivo',
        methods=['POST'],
        endpoint='enviar_gasto_sap_masivo'
    )
    @require_login
    @require_permission('gastos_tarjeta', 'ver')
    def enviar_gasto_sap_masivo():
        """
        Envío masivo a SAP (solo coordinador/admin).

        - Tipo 1 / 2: se envía individual usando enviar_gasto_sap(gid)
        - Tipo 3:
            * Caja chica / Reembolso / Tarjeta online / Boletos aéreos
            * Reembolso: distribución por usuarios_cc
            * Caja chica: CC desde detalle
            * Cada línea va en su propio Liquidacion_Documento

        Sociedad SAP:
            gastos_tarjeta.usuario_id
            -> usuarios.empresa_id
            -> empresas.rep_nacionalidad
        """
        from datetime import datetime
        import json
        import requests
        from collections import defaultdict
        from decimal import Decimal, ROUND_HALF_UP

        role = (session.get('rol') or '').strip().lower()
        if role not in ('admin', 'coordinador'):
            return jsonify(ok=False, msg='No autorizado'), 403

        payload_in = request.get_json(silent=True) or {}
        ids = payload_in.get('ids') or []

        if not isinstance(ids, list) or not ids:
            return jsonify(ok=False, msg='Debe enviar una lista de ids.'), 400

        ids_int = []
        for gid in ids:
            try:
                ids_int.append(int(gid))
            except Exception:
                pass

        ids_int = list(dict.fromkeys(ids_int))

        if not ids_int:
            return jsonify(ok=False, msg='No hay IDs válidos.'), 400

        conn = get_db()

        try:
            cfg = get_sap_config_from_db(conn)

            SAP_URL = cfg["SAP_URL"]
            SAP_URL_QAS = cfg["SAP_URL_QAS"]
            SAP_CLIENT = cfg["SAP_CLIENT"]
            SAP_USER = cfg["SAP_USER"]
            SAP_PASS = cfg["SAP_PASS"]

            if not (SAP_URL or '').startswith(('http://', 'https://')):
                return jsonify(ok=False, msg=f'SAP_URL inválida: {SAP_URL!r}'), 500

            Q2 = Decimal("0.01")

            def _D(x):
                try:
                    return Decimal(str(x or "0"))
                except Exception:
                    return Decimal("0")

            def _q2(x: Decimal) -> Decimal:
                return x.quantize(Q2, rounding=ROUND_HALF_UP)

            def _fmt_ddmmyyyy(iso_date):
                try:
                    return datetime.strptime(str(iso_date), "%Y-%m-%d").strftime("%d.%m.%Y")
                except Exception:
                    return ""

            def _distribuir_monto(monto_linea: Decimal, distribucion_cc: list) -> list:
                monto_linea = _q2(monto_linea)
                acumulado = Decimal("0.00")
                out = []
                n = len(distribucion_cc)

                for idx, cc in enumerate(distribucion_cc):
                    if idx == n - 1:
                        monto_cc = _q2(monto_linea - acumulado)
                    else:
                        porc = _D(cc.get("porcentaje")) / Decimal("100")
                        monto_cc = _q2(monto_linea * porc)
                        acumulado += monto_cc

                    out.append(monto_cc)

                return out

            def _norm_actor(g):
                return (
                    g.get("ga_actor")
                    or g.get("ga_aprobador")
                    or g.get("ga_aprobador_rol")
                    or ""
                ).strip().lower()

            def _ga_effective_ok(g, ga_ok, gg_ok, gf_ok, es_restringido):
                if not es_restringido:
                    return ga_ok

                actor = _norm_actor(g)

                if actor in ("gg", "gerente general", "gerente_general"):
                    return gg_ok

                if actor in ("gf", "gerente financiero", "gerente_financiero"):
                    return gf_ok

                if (not ga_ok) and (gg_ok or gf_ok):
                    return True

                return ga_ok

            def _get_codigo_sap_ultimo_jefe(cur, conn, usuario_id: int) -> str:
                try:
                    jefe_id = _get_ultimo_jefe_id(conn, usuario_id)
                    if not jefe_id:
                        return ""

                    cur.execute(
                        "SELECT COALESCE(codigo_sap,'') AS codigo_sap FROM usuarios WHERE id = ?",
                        (jefe_id,)
                    )
                    r = cur.fetchone()

                    if not r:
                        return ""

                    try:
                        return (r["codigo_sap"] or "").strip()
                    except Exception:
                        return (r[0] or "").strip()

                except Exception:
                    return ""

            def _make_liq_documento(g0, numero_doc_liq, fecha_doc, cod="001", serie="002"):
                return {
                    "Fecha_Documento": fecha_doc,
                    "Tipo_Id": "51",
                    "Id_Fiscal": g0["usuario_cedula"],
                    "Moneda": "USD",
                    "Tipo_Doc_Liquidacion": "01",
                    "Cod_Doc_Liquidacion": cod,
                    "Serie_Doc_Liquidacion": serie,
                    "Numero_Doc_Liquidacion": numero_doc_liq,
                    "Numero_Autorizacion": g0["usuario_cedula"],
                    "Liquidacion_Detalle": []
                }


            def _usuario_sap_from_gasto(g):
                usuario_sap = (g.get("empresa_usuario_sap") or "").strip().upper()

                if not usuario_sap:
                    raise Exception(
                        f"Gasto {g.get('id')}: no se pudo determinar Usuario SAP. "
                        "Revise que la empresa tenga configurado el campo usuario_sap."
                    )

                return usuario_sap
            
            def _sociedad_from_gasto(g):
                sociedad = (g.get("empresa_sociedad") or "").strip().upper()

                if not sociedad:
                    raise Exception(
                        f"Gasto {g.get('id')}: no se pudo determinar Sociedad SAP. "
                        "Revise que el usuario tenga empresa asignada y que la empresa tenga "
                        "configurado el campo rep_nacionalidad."
                    )

                return sociedad

            def _indicador_liquidacion(gasto, detalle_linea):
                """
                Mantiene la lógica actual:
                - Caja chica normal / reembolso / tarjeta online / boletos: C0

                Pero si es Caja Chica con Detalle:
                - usa el indicador real de la línea.
                """
                es_caja_chica = int(gasto.get("es_caja_chica") or 0) == 1
                tipo_cc = (gasto.get("tipo_caja_chica") or "NINGUNA").strip().upper()

                if es_caja_chica and tipo_cc == "DETALLE_FACTURA":
                    return (detalle_linea.get("indicador") or "C0").strip().upper()

                return "C0"

            cur = conn.cursor()
            results, sent, errors = [], 0, 0

            ph = ",".join(["?"] * len(ids_int))

            # ---------------------------------------------------------
            # Cabeceras
            # ---------------------------------------------------------
            cur.execute(f"""
                SELECT
                    g.*,

                    u.username,
                    u.codigo_sap AS usuario_codigo_sap,
                    u.identificacion AS usuario_cedula,
                    u.empresa_id AS usuario_empresa_id,
                    COALESCE(u.tipo_caja_chica, 'NINGUNA') AS tipo_caja_chica,

                    t.codigo_sap AS proveedor_codigo_sap,

                    
                    e.razon_social AS empresa_razon_social,
                    LTRIM(RTRIM(COALESCE(e.rep_nacionalidad, ''))) AS empresa_sociedad,
                    LTRIM(RTRIM(COALESCE(e.usuario_sap, ''))) AS empresa_usuario_sap

                FROM {TABLE_GASTOS} g
                JOIN usuarios u
                    ON u.id = g.usuario_id
                LEFT JOIN terceros t
                    ON t.id = g.proveedor_id
                LEFT JOIN empresas e
                    ON e.id = u.empresa_id
                WHERE g.id IN ({ph})
            """, ids_int)

            gastos = []

            for r in cur.fetchall():
                try:
                    gastos.append(dict(r))
                except Exception:
                    gastos.append({
                        "id": r[0],
                        "usuario_id": r[1],
                    })

            gastos_by_id = {g["id"]: g for g in gastos}

            # ---------------------------------------------------------
            # Detalles
            # ---------------------------------------------------------
            cur.execute(f"""
                SELECT *
                FROM gastos_tarjeta_detalle
                WHERE gasto_id IN ({ph})
                ORDER BY gasto_id, id
            """, ids_int)

            dets_by_gasto = defaultdict(list)

            for r in cur.fetchall():
                try:
                    dets_by_gasto[r["gasto_id"]].append(dict(r))
                except Exception:
                    row = dict(zip([c[0] for c in cur.description], r))
                    dets_by_gasto[row["gasto_id"]].append(row)

            # ---------------------------------------------------------
            # Separar tipo 3 vs estándar + validar aprobaciones
            # ---------------------------------------------------------
            tipo3_ids, standard_ids = [], []

            for gid in ids_int:
                g = gastos_by_id.get(gid)

                if not g:
                    continue

                # Validar sociedad desde empresa del usuario
                _sociedad_from_gasto(g)

                ga_ok = int(g.get("ga_aprobado") or 0) == 1
                gg_ok = int(g.get("gg_aprobado") or 0) == 1
                gf_ok = int(g.get("gf_aprobado") or 0) == 1

                es_caja_chica = int(g.get("es_caja_chica") or 0) == 1
                es_reembolso = int(g.get("reembolso_vendedor") or 0) == 1
                es_boletos_aereos = int(g.get("boletos_aereos") or 0) == 1
                es_tarjeta_online = int(g.get("tarjeta_sin_soporte") or 0) == 1

                es_restringido = es_caja_chica or es_reembolso
                ga_eff_ok = _ga_effective_ok(g, ga_ok, gg_ok, gf_ok, es_restringido)

                if es_restringido:
                    if not ga_eff_ok:
                        actor = _norm_actor(g)

                        if actor.startswith("gg") or "general" in actor:
                            why = "requiere aprobación GG (actúa como GA)"
                        elif actor.startswith("gf") or "financ" in actor:
                            why = "requiere aprobación GF (actúa como GA)"
                        else:
                            why = "requiere aprobación GA (o su rol sustituto)"

                        return jsonify(ok=False, msg=f"Gasto {gid}: {why}"), 400

                else:
                    if not (ga_ok and gg_ok and gf_ok):
                        return jsonify(ok=False, msg=f"Gasto {gid}: requiere aprobación GA, GG y GF"), 400

                if (g.get("sap_contabilizacion") or "").strip():
                    return jsonify(ok=False, msg=f"Gasto {gid}: ya fue enviado a SAP"), 400

                es_tipo3 = bool(es_caja_chica or es_reembolso or es_tarjeta_online or es_boletos_aereos)

                if es_tipo3:
                    tipo3_ids.append(gid)
                else:
                    standard_ids.append(gid)

            # ---------------------------------------------------------
            # Envíos estándar
            # ---------------------------------------------------------
            for gid in standard_ids:
                resp = enviar_gasto_sap(gid)
                data = resp.get_json(silent=True) if hasattr(resp, "get_json") else {}
                ok = bool(data.get("ok"))

                results.append({
                    "id": gid,
                    "ok": ok,
                    "doc": data.get("doc"),
                    "msg": data.get("msg")
                })

                sent += int(ok)
                errors += int(not ok)

            # ---------------------------------------------------------
            # Agrupar tipo 3
            # ---------------------------------------------------------
            groups = defaultdict(list)

            for gid in tipo3_ids:
                g = gastos_by_id[gid]

                if int(g.get("boletos_aereos") or 0) == 1:
                    tipo = "boletos_aereos"
                elif int(g.get("tarjeta_sin_soporte") or 0) == 1:
                    tipo = "tarjeta_online"
                else:
                    tipo = "caja_chica" if int(g.get("es_caja_chica") or 0) == 1 else "reembolso"

                sociedad = _sociedad_from_gasto(g)

                if tipo == "boletos_aereos":
                    
                    groups[(sociedad, usuario_sap, g.get("proveedor_id"), tipo)].append(gid)
                else:
                    usuario_sap = _usuario_sap_from_gasto(g)
                    groups[(sociedad, usuario_sap, g["usuario_id"], tipo)].append(gid)
 

            # ---------------------------------------------------------
            # Envío tipo 3
            # ---------------------------------------------------------
            for group_key, gids in groups.items():
                sociedad_sap, usuario_sap, owner_id, tipo = group_key
                g0 = gastos_by_id[gids[0]]

                if tipo == "boletos_aereos":
                    prov_sap = (g0.get("proveedor_codigo_sap") or "").strip()

                    if not prov_sap:
                        raise Exception("Proveedor sin código SAP (terceros.codigo_sap).")

                    acreedor = prov_sap

                elif tipo == "tarjeta_online":
                    acreedor = (g0.get("usuario_codigo_sap") or "").strip()
                    cod_gerente = _get_codigo_sap_ultimo_jefe(cur, conn, owner_id)

                    if cod_gerente:
                        acreedor = cod_gerente

                else:
                    acreedor = (g0.get("usuario_codigo_sap") or "").strip()

                distribucion_cc = []

                if tipo == "reembolso":
                    cur.execute("""
                        SELECT
                            ucc.porcentaje,
                            pv.valor AS codigo_cc_sap
                        FROM usuarios_cc ucc
                        JOIN param_values pv
                            ON pv.id = ucc.centro_costo_id
                        WHERE ucc.usuario_id = ?
                        AND pv.activo = 1
                    """, (owner_id,))

                    distribucion_cc = []

                    for r in cur.fetchall():
                        try:
                            distribucion_cc.append(dict(r))
                        except Exception:
                            distribucion_cc.append({
                                "porcentaje": r[0],
                                "codigo_cc_sap": r[1],
                            })

                    if not distribucion_cc:
                        raise Exception(f"Usuario {g0.get('username')} sin centros de costo.")

                numero_doc_liq = obtener_siguiente_secuencia(conn, 'reembolsos')
                fecha_liq = _fmt_ddmmyyyy(g0.get("fecha"))
                fecha_doc = _fmt_ddmmyyyy(g0.get("fecha"))

                documento = {
                    "Agrupador": "1",
                    "Tipo_Ejecucion": "3",
                    "Sociedad": sociedad_sap,
                    "Acreedor": acreedor,
                    "Referencia": g0.get("numero_factura") or "",
                    "Importe_Total": "0.00",
                    "Moneda": "USD",
                    "Clase_Documento": "KP" if tipo == "boletos_aereos" else "KA",
                    "Lugar": "0001",
                    "Via_Pago": "T",
                    "Usuario_SAP": usuario_sap,
                    "Fecha_Liquidacion": fecha_liq,
                    "Motivo_Liquidacion": "REEMBOLSO DE GASTOS",
                    "Liquidacion_Documentos": []
                }

                total = Decimal("0.00")

                for gid in gids:
                    g_head = gastos_by_id[gid]
                    motivo_gasto = (g_head.get("motivo") or "").strip()

                    for d in dets_by_gasto.get(gid, []):
                        monto_linea = _D(d.get("total_con_iva"))

                        if monto_linea <= 0:
                            continue

                        cuenta_gasto = d.get("motivo") or ""
                        division = "CORP"
                        indicador_iva = _indicador_liquidacion(g_head, d)

                        if tipo in ("caja_chica", "tarjeta_online", "boletos_aereos"):
                            cc = (d.get("centro_costo") or "").strip()

                            if not cc:
                                raise Exception(f"Gasto {gid}: requiere Centro de Costo en el detalle.")

                            liq_doc = _make_liq_documento(g0, numero_doc_liq, fecha_doc)

                            liq_doc["Liquidacion_Detalle"].append({
                                "Cuenta_Gasto": cuenta_gasto,
                                "Division": division,
                                "Indicador_Iva": indicador_iva,
                                "Centro_Costo": cc,
                                "Importe_Posicion": f"{_q2(monto_linea):.2f}",
                                "Motivo_Detalle": motivo_gasto
                            })

                            documento["Liquidacion_Documentos"].append(liq_doc)
                            total += monto_linea

                        else:
                            montos_cc = _distribuir_monto(monto_linea, distribucion_cc)

                            for cc_row, m in zip(distribucion_cc, montos_cc):
                                liq_doc = _make_liq_documento(g0, numero_doc_liq, fecha_doc)

                                liq_doc["Liquidacion_Detalle"].append({
                                    "Cuenta_Gasto": cuenta_gasto,
                                    "Division": division,
                                    "Indicador_Iva": indicador_iva,
                                    "Centro_Costo": cc_row["codigo_cc_sap"],
                                    "Importe_Posicion": f"{m:.2f}",
                                    "Motivo_Detalle": motivo_gasto
                                })

                                documento["Liquidacion_Documentos"].append(liq_doc)

                            total += monto_linea

                documento["Importe_Total"] = f"{_q2(total):.2f}"
                payload_sap = [documento]

                trama = json.dumps(payload_sap, indent=2, ensure_ascii=False)
                current_app.logger.info("[SAP_MASIVO] Payload grupo %s:\n%s", str(group_key), trama)

                r = requests.post(
                    SAP_URL,
                    params={"sap-client": SAP_CLIENT},
                    json=payload_sap,
                    auth=(SAP_USER, SAP_PASS),
                    timeout=30,
                )

                ct = (r.headers.get("Content-Type") or "").lower()
                text_snip = (r.text or "")[:800]

                try:
                    if "application/json" in ct or (r.text and r.text.lstrip().startswith(("{", "["))):
                        resp = r.json()
                    else:
                        resp = {
                            "_non_json": True,
                            "text": text_snip,
                            "content_type": ct
                        }
                except Exception:
                    resp = {
                        "_json_error": True,
                        "text": text_snip,
                        "content_type": ct
                    }

                if not r.ok:
                    current_app.logger.error(
                        "[SAP_MASIVO] SAP HTTP %s | CT=%s | BODY=%s",
                        r.status_code,
                        ct,
                        text_snip
                    )
                    return jsonify(
                        ok=False,
                        msg=f"SAP respondió {r.status_code}. Revise log para detalle."
                    ), 502

                error_sap = _extraer_error_sap(resp)

                if error_sap:
                    current_app.logger.error(
                        "[SAP_MASIVO] SAP ERROR LOGICO | grupo=%s | error=%s | resp=%s",
                        str(group_key),
                        error_sap,
                        json.dumps(resp, ensure_ascii=False)[:4000]
                    )
                    return jsonify(ok=False, msg=error_sap), 400

                doc = ""

                if isinstance(resp, list) and resp:
                    doc = (
                        resp[0].get("documento")
                        or resp[0].get("DOCUMENTO")
                        or resp[0].get("document")
                        or resp[0].get("doc_number")
                        or ""
                    )
                elif isinstance(resp, dict):
                    doc = (
                        resp.get("documento")
                        or resp.get("DOCUMENTO")
                        or resp.get("document")
                        or resp.get("doc_number")
                        or ""
                    )

                for gid in gids:
                    cur.execute(
                        f"""
                        UPDATE {TABLE_GASTOS}
                        SET sap_contabilizacion = ?,
                            sap_response_json = ?,
                            sap_enviado_at = ?,
                            sap_error_msg = NULL
                        WHERE id = ?
                        """,
                        (
                            doc,
                            json.dumps(resp, ensure_ascii=False)[:4000],
                            datetime.utcnow().isoformat(timespec='seconds'),
                            gid
                        )
                    )

                    results.append({
                        "id": gid,
                        "ok": True,
                        "doc": doc or None,
                        "msg": None
                    })

                    sent += 1

                conn.commit()

            return jsonify(ok=True, sent=sent, errors=errors, results=results)

        except Exception as e:
            conn.rollback()
            current_app.logger.exception("[SAP_MASIVO] ERROR")
            return jsonify(ok=False, msg=str(e)), 500

        finally:
            try:
                conn.close()
            except Exception:
                pass
    @app.route('/reembolsos/xml/<int:fid>/estado', methods=['POST'], endpoint='factura_xml_estado')
    @require_login
    @require_permission('gastos_tarjeta', 'editar')
    def factura_xml_estado(fid):
        nuevo = (request.form.get('estado') or 'PENDIENTE').upper()
        if nuevo not in ('PENDIENTE', 'PROCESADO'):
            nuevo = 'PENDIENTE'

        conn = get_db() 
        cur = conn.cursor()
        cur.execute("UPDATE facturas_xml SET estado=? WHERE id=?", (nuevo, fid))
        conn.commit()
        conn.close()
        flash(f'Factura marcada como {nuevo}.', 'success')
        return redirect(url_for('facturas_xml_list'))

 