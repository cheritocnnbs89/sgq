# modules/gastos_exports.py
from __future__ import annotations

import os
import io
import shutil
import tempfile
import subprocess
from datetime import datetime, date

from flask import Response, current_app, request, session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..db import get_db
from ..config import TABLE_GASTOS
from .gastos_helpers import collect_gastos_filters  # mismo colector de filtros que la lista

NUM_FMT = "#,##0.00"


# ==========================================================
# OpenPyXL helpers (estilo)
# ==========================================================
def _thin_border():
    s = Side(border_style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_border_range(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = _thin_border()


def _obtener_subordinados(conn, jefe_id: int) -> list[int]:
    """
    Retorna TODOS los subordinados (directos e indirectos) de un jefe usando usuarios.jefe_id.
    (misma lógica que en routes_gastos_tarjeta.py)
    """
    if not jefe_id:
        return []

    result = set()
    pendientes = [jefe_id]
    cur = conn.cursor()

    while pendientes:
        actual = pendientes.pop()
        cur.execute("SELECT id FROM usuarios WHERE jefe_id = ?", (actual,))
        hijos = [row["id"] for row in cur.fetchall()]
        for h in hijos:
            if h not in result:
                result.add(h)
                pendientes.append(h)

    return list(result)


def _find_row_contains(ws, text):
    tgt = (text or "").strip().lower()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == tgt:
                return r
    return None


def _unmerge_in_area(ws, r1, c1, r2, c2):
    for cr in list(ws.merged_cells.ranges):
        if not (cr.max_row < r1 or cr.min_row > r2 or cr.max_col < c1 or cr.min_col > c2):
            for rr in range(cr.min_row, cr.max_row + 1):
                for cc in range(cr.min_col, cr.max_col + 1):
                    _ = ws.cell(row=rr, column=cc)
            if cr in ws.merged_cells.ranges:
                ws.unmerge_cells(
                    start_row=cr.min_row,
                    start_column=cr.min_col,
                    end_row=cr.max_row,
                    end_column=cr.max_col,
                )


def _unmerge_row(ws, row, c1=1, c2=None):
    if c2 is None:
        c2 = ws.max_column
    to_unmerge = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= row <= rng.max_row and not (rng.max_col < c1 or rng.min_col > c2):
            to_unmerge.append(str(rng))
    for rs in to_unmerge:
        ws.unmerge_cells(range_string=rs)


def _style_total_row(ws, total_row: int, *, label_col=6, block_start=7, block_end=13):
    thin = Side(style="thin", color="000000")
    med = Side(style="medium", color="000000")

    lab = ws.cell(total_row, label_col)
    lab.alignment = Alignment(horizontal="left", vertical="center")
    lab.font = Font(bold=False)
    lab.border = Border(top=med, left=thin, right=thin, bottom=thin)

    for col in range(block_start, block_end + 1):
        left = med if col == block_start else thin
        right = med if col == block_end else thin
        cell = ws.cell(total_row, col)
        cell.border = Border(top=med, left=left, right=right, bottom=thin)
        cell.alignment = Alignment(horizontal="right", vertical="center")


# ==========================================================
# Header helpers (llenar cabecera plantilla)
# ==========================================================
def _norm_label(s: str) -> str:
    if not s:
        return ""
    s = s.strip().lower()
    # normaliza espacios y quita ":" para empatar plantillas tipo "NOMBRE  :" o "COMPAÑÍA:"
    s = " ".join(s.split())
    s = s.replace(":", "")
    return s


def _find_cell_by_label(ws, label: str):
    """
    Busca una celda cuyo texto sea igual al label (normalizado).
    Retorna (row, col) o None.
    """
    tgt = _norm_label(label)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and _norm_label(v) == tgt:
                return (r, c)
    return None


def _top_left_of_merge(ws, row: int, col: int):
    """
    Si (row,col) cae dentro de un merge, devuelve la celda top-left del merge.
    Si no, devuelve (row,col).
    """
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return (rng.min_row, rng.min_col)
    return (row, col)


def _write_right_of_label(ws, label: str, value: str):
    """
    Escribe el value en la celda a la derecha del label.
    Si esa celda cae dentro de un merge, escribe en el top-left del merge para que Excel lo muestre.
    """
    pos = _find_cell_by_label(ws, label)
    if not pos:
        return False

    r, c = pos
    tr, tc = _top_left_of_merge(ws, r, c + 1)
    cell = ws.cell(tr, tc)
    cell.value = value or ""
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return True


def _get_ultimo_jefe_activo(conn, user_id: int | None) -> int | None:
    """
    Igual que tu regla en routes_gastos_tarjeta.py:
    gerente = último jefe activo en la cadena usuarios.jefe_id
    """
    if not user_id:
        return None

    cur = conn.cursor()
    cur.execute("SELECT jefe_id FROM usuarios WHERE id = ?", (user_id,))
    row = cur.fetchone()
    if not row:
        return None

    jefe_id = row["jefe_id"]
    if not jefe_id:
        return None

    seen = set()
    last_valid = None

    while jefe_id and jefe_id not in seen:
        seen.add(jefe_id)

        cur.execute("""
            SELECT id, jefe_id
            FROM usuarios
            WHERE id = ?
              AND COALESCE(disabled, 0) = 0
        """, (jefe_id,))
        j = cur.fetchone()

        if not j:
            break

        last_valid = j["id"]
        jefe_id = j["jefe_id"]

    return last_valid


def _get_usuario_header_data(conn):
    uid = session.get("usuario_id") or session.get("user_id")
    if not uid:
        return {"compania": "", "nombre": "", "cargo": "", "departamento": "", "aprobado_por": ""}

    uid = int(uid)
    gerente_id = _get_ultimo_jefe_activo(conn, uid)

    cur = conn.cursor()
    cur.execute(
        """
        SELECT
            COALESCE(e.razon_social, '') AS compania,
            COALESCE(u.nombre_completo, u.username, '') AS nombre,
            COALESCE(p.nombre, '') AS cargo,
            COALESCE(d.nombre, '') AS departamento
        FROM usuarios u
        LEFT JOIN empresas e      ON e.id = u.empresa_id
        LEFT JOIN puestos p       ON p.id = u.puesto_id
        LEFT JOIN departamentos d ON d.id = u.departamento_id
        WHERE u.id = ?
        LIMIT 1
        """,
        (uid,),
    )
    row = cur.fetchone()
    rd = dict(row) if row else {}

    aprobado_por = ""
    if gerente_id:
        cur.execute(
            "SELECT COALESCE(nombre_completo, username, '') AS n FROM usuarios WHERE id = ? LIMIT 1",
            (int(gerente_id),),
        )
        j = cur.fetchone()
        aprobado_por = (j["n"] or "") if j else ""

    return {
        "compania": rd.get("compania") or "",
        "nombre": rd.get("nombre") or "",
        "cargo": rd.get("cargo") or "",
        "departamento": rd.get("departamento") or "",
        "aprobado_por": aprobado_por or "",
    }

def _fill_excel_header(ws, header_data: dict, filtros: dict):
    """
    Llena cabecera en la plantilla:
      COMPAÑÍA:
      NOMBRE  :
      CARGO  :
      FECHA DE LIQUIDACIÓN:
      PERIODO DE LIQUIDACIÓN
      DEPARTAMENTO:
      DIVISION:  (vacío)
    """
    desde = (filtros.get("desde") or request.args.get("desde") or "").strip()
    hasta = (filtros.get("hasta") or request.args.get("hasta") or "").strip()

    if desde and hasta:
        periodo = f"{desde} - {hasta}"
    elif desde:
        periodo = f"Desde {desde}"
    elif hasta:
        periodo = f"Hasta {hasta}"
    else:
        periodo = ""

    fecha_liq = date.today().strftime("%Y-%m-%d")

    _write_right_of_label(ws, "COMPAÑÍA:", header_data.get("compania", ""))
    _write_right_of_label(ws, "NOMBRE  :", header_data.get("nombre", ""))
    _write_right_of_label(ws, "CARGO  :", header_data.get("cargo", ""))

    # En la plantilla el texto real es "FECHA DE LIQUIDACIÓN:" (aunque se vea truncado)
    _write_right_of_label(ws, "FECHA DE LIQUIDACIÓN:", fecha_liq)

    _write_right_of_label(ws, "PERIODO DE LIQUIDACIÓN", periodo)
    _write_right_of_label(ws, "DEPARTAMENTO:", header_data.get("departamento", ""))

    # Por requerimiento: dejar vacío
    _write_right_of_label(ws, "DIVISION:", "")
    _write_right_of_label(ws, "Aprobado por:", header_data.get("aprobado_por", ""))
    _write_right_of_label(ws, "Fecha", date.today().strftime("%Y-%m-%d"))
    _write_right_of_label(ws, "Aprobado por:", header_data.get("aprobado_por", ""))
    _write_right_of_label(ws, "Aprobado por:", header_data.get("aprobado_por", ""))
  
    _write_right_of_label(ws, "Fecha", header_data.get("aprobado_fecha", ""))
    _write_right_of_label(ws, "Revisado por    :", header_data.get("revisado_por", ""))
    _write_right_of_label(ws, "Elaborado por:", header_data.get("nombre", ""))

    
    
    





# ==========================================================
# Scope helpers (para export)  -> que coincida con la LISTA
# ==========================================================
def _role_lower(sess) -> str:
    return (sess.get("rol") or "").strip().lower()


def _role_scope_for_exports(conn, session):
    """
    Devuelve (join_u, extra_where, extra_args) alineado con la pantalla.
    - admin/coordinador: ven todo.
    - gerente (GA): subordinados + él mismo (por usuarios.jefe_id).
    - GG/GF: ven TARJETA de todos, pero CAJA_CHICA/REEMBOLSO solo si están en su scope.
    """
    role_name = (session.get("rol") or "").strip().lower()
    uid = session.get("usuario_id") or session.get("user_id")

    join_u = "LEFT JOIN usuarios u ON u.id = g.usuario_id"
    extra_where, extra_args = [], []

    if not uid:
        extra_where.append("1=0")
        return join_u, extra_where, extra_args

    # PRIV_ALL como en tu lista (admin/coordinador siempre)
    if role_name in ("admin", "coordinador") or bool(session.get("is_admin")):
        return join_u, extra_where, extra_args

    # GG / GF: tarjeta todos; restringidos solo scope
    if role_name in ("gerente general", "gerente financiero"):
        scope_ids = set(_obtener_subordinados(conn, int(uid)) or [])
        scope_ids.add(int(uid))

        placeholders = ",".join("?" * len(scope_ids)) if scope_ids else "?"
        extra_where.append(
            f"""
            (
                (COALESCE(g.es_caja_chica,0)=0 AND COALESCE(g.reembolso_vendedor,0)=0)
                OR g.usuario_id IN ({placeholders})
            )
            """.strip()
        )
        extra_args.extend(list(scope_ids) if scope_ids else [int(uid)])
        return join_u, extra_where, extra_args

    # GA: subordinados + él mismo
    if role_name in ("gerente", "gerente de área", "gerente de area"):
        allowed_ids = set(_obtener_subordinados(conn, int(uid)) or [])
        allowed_ids.add(int(uid))

        placeholders = ",".join("?" * len(allowed_ids)) if allowed_ids else "?"
        extra_where.append(f"g.usuario_id IN ({placeholders})")
        extra_args.extend(list(allowed_ids) if allowed_ids else [int(uid)])
        return join_u, extra_where, extra_args

    # Usuario normal: solo sus gastos
    extra_where.append("g.usuario_id = ?")
    extra_args.append(int(uid))
    return join_u, extra_where, extra_args


def _apply_role_scope_for_exports(conn, where: list[str], args: list):
    """
    Modifica where/args para que export tenga el MISMO alcance que la lista.
    """
    role_name = (session.get("rol") or "").strip().lower()
    uid = session.get("usuario_id") or session.get("user_id")

    if not uid:
        where.append("1=0")
        return

    uid = int(uid)

    PRIV_ALL = ("admin", "coordinador")
    GERENTE_ROLES = ("gerente", "gerente de área", "gerente de area")

    if role_name in PRIV_ALL or bool(session.get("is_admin")):
        return

    # GG/GF: Tarjeta todos, pero restringidos solo scope
    if role_name in ("gerente general", "gerente financiero"):
        scope_ids = set(_obtener_subordinados(conn, uid) or [])
        scope_ids.add(uid)

        placeholders = ",".join("?" * len(scope_ids)) if scope_ids else "?"
        where.append(
            f"""(
                (COALESCE(g.es_caja_chica,0)=0 AND COALESCE(g.reembolso_vendedor,0)=0)
                OR g.usuario_id IN ({placeholders})
            )"""
        )
        args.extend(list(scope_ids) if scope_ids else [uid])
        return

    # GA: subordinados + él mismo
    if role_name in GERENTE_ROLES:
        allowed_ids = set(_obtener_subordinados(conn, uid) or [])
        allowed_ids.add(uid)

        placeholders = ",".join("?" * len(allowed_ids)) if allowed_ids else "?"
        where.append(f"g.usuario_id IN ({placeholders})")
        args.extend(list(allowed_ids) if allowed_ids else [uid])
        return

    # Usuario normal
    where.append("g.usuario_id = ?")
    args.append(uid)




def _parse_ids_req() -> list[int]:
    ids_raw = (request.args.get("ids") or "").strip()
    if not ids_raw:
        return []
    out = []
    for x in ids_raw.split(","):
        x = (x or "").strip()
        if x.isdigit():
            out.append(int(x))
    # unique preservando orden
    seen = set()
    uniq = []
    for i in out:
        if i not in seen:
            seen.add(i)
            uniq.append(i)
    return uniq


def _fetch_allowed_ids_in_scope(conn, ids_req: list[int]) -> list[int]:
    """
    Filtra ids_req dejando solo los que el usuario puede ver según el mismo scope de la lista.
    Retorna IDs válidos (en el mismo orden de ids_req).
    """
    if not ids_req:
        return []

    where_scope = []
    args_scope: list = []

    role_name = (session.get("rol") or "").strip().lower()
    uid = session.get("usuario_id") or session.get("user_id")

    if not uid:
        return []

    uid = int(uid)

    # Admin/Coordinador: todo
    if role_name in ("admin", "coordinador") or bool(session.get("is_admin")):
        allowed_set = set(ids_req)
        return [i for i in ids_req if i in allowed_set]

    # Gerentes: subordinados + yo
    if role_name in ("gerente", "gerente de área", "gerente de area"):
        allowed_users = set(_obtener_subordinados(conn, uid) or [])
        allowed_users.add(uid)
        if not allowed_users:
            return []
        placeholders_u = ",".join("?" * len(allowed_users))
        where_scope.append(f"g.usuario_id IN ({placeholders_u})")
        args_scope.extend(list(allowed_users))

    # GG/GF: tarjeta todos; restringidos solo scope
    elif role_name in ("gerente general", "gerente financiero"):
        scope_users = set(_obtener_subordinados(conn, uid) or [])
        scope_users.add(uid)
        placeholders_u = ",".join("?" * len(scope_users)) if scope_users else "?"
        where_scope.append(
            f"""(
                (COALESCE(g.es_caja_chica,0)=0 AND COALESCE(g.reembolso_vendedor,0)=0)
                OR g.usuario_id IN ({placeholders_u})
            )"""
        )
        args_scope.extend(list(scope_users) if scope_users else [uid])

    # Usuario normal: solo sus gastos
    else:
        where_scope.append("g.usuario_id = ?")
        args_scope.append(uid)

    placeholders_ids = ",".join("?" * len(ids_req))
    sql = f"""
        SELECT g.id
        FROM {TABLE_GASTOS} g
        WHERE g.id IN ({placeholders_ids})
    """
    sql_args: list = list(ids_req)

    if where_scope:
        sql += " AND " + " AND ".join(where_scope)
        sql_args.extend(args_scope)

    cur = conn.cursor()
    cur.execute(sql, sql_args)
    allowed = {int(r[0]) for r in cur.fetchall()}

    # mantener el orden original
    return [i for i in ids_req if i in allowed]


def _tipo_label_from_row(rd: dict) -> str:
    """
    Determina el tipo de gasto según flags.
    - Caja chica si es_caja_chica=1
    - Reembolso vendedor si reembolso_vendedor=1
    - Caso contrario: tarjeta corporativa
    """
    if int(rd.get("es_caja_chica") or 0) == 1:
        return "CAJA CHICA"
    if int(rd.get("reembolso_vendedor") or 0) == 1:
        return "REEMBOLSO DE VENDEDOR"
    return "TARJETA CORPORATIVA"


def _build_report_title_from_rows(rows) -> str:
    """
    Construye el título principal en base a los tipos presentes.
    Si hay 1 tipo: "GASTOS DE <TIPO>"
    Si hay varios: "GASTOS DE <TIPO1> / <TIPO2> / ..."
    """
    tipos = []
    seen = set()
    for r in rows or []:
        rd = dict(r)
        t = _tipo_label_from_row(rd)
        if t not in seen:
            seen.add(t)
            tipos.append(t)

    if not tipos:
        # fallback (si no hay filas, deja el título de tarjeta como default)
        return "GASTOS DE TARJETA CORPORATIVA"

    if len(tipos) == 1:
        return f"GASTOS DE {tipos[0]}"

    return "GASTOS DE " + " / ".join(tipos)


def _set_report_title(ws, new_title: str) -> bool:
    """
    Busca la celda que contiene el título actual (ej. "GASTO DE TARJETA CORPORATIVA")
    y la reemplaza por new_title. Maneja merges (escribe en top-left).
    """
    candidates = {
        "gasto de tarjeta corporativa",
        "gastos de tarjeta corporativa",
        "gasto de tarjeta",
        "gastos de tarjeta",
    }

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() in candidates:
                tr, tc = _top_left_of_merge(ws, r, c)
                ws.cell(tr, tc).value = new_title
                ws.cell(tr, tc).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(tr, tc).font = Font(bold=True)  # respeta lo general
                return True
    return False

# ==========================================================
# XLSX builder
# ==========================================================
# ==========================================================
# XLSX builder
# ==========================================================
def _build_gastos_workbook():
    """
    Devuelve un Workbook con el reporte.
    (Plantilla opcional en: <static>/plantillas/Modelo Excel Reporte Tarjeta de Crédito.xlsx)

    ✅ Cambios incluidos SIN perder tu lógica:
    - Si llega ?ids=1,2,3 exporta SOLO esos IDs (prioridad).
    - Valida que esos IDs estén dentro del scope del rol (seguridad).
    - Si no llega ids, exporta por filtros como antes (collect_gastos_filters + scope + tipo + gerente).
    - Mantiene: plantilla, cabecera usuario, título dinámico, CCB, totales, estilos.
    - Parse de fecha blindado para evitar ValueError por strings raros.
    """
    current_app.logger.warning("EXPORT args=%r", dict(request.args))
    filtros, where, args, _ = collect_gastos_filters(request, session)
    current_app.logger.warning("EXPORT where=%r", where)
    current_app.logger.warning("EXPORT sql_args=%r", args)
    DETAIL_START_ROW = 12
    TOTAL_ANCHOR_TEXT = "Total"

    # Bloque numérico G..M (7..13) para sumas. Agregamos CCB como columna 14.
    DETAIL_COL_COUNT = 14
    LABEL_TOTAL_COL = 6  # F
    TOTAL_COL_START = 7  # G (con soporte)
    TOTAL_COL_END = 13   # M (total con IVA)

    conn = get_db()
    cur = conn.cursor()

    try:
        # ======================================================
        # 0) ✅ IDs visibles desde UI (prioridad)
        # ======================================================
        ids_req = _parse_ids_req()  # helper que te pasé

        # ======================================================
        # 1) filtros base que la lista (fechas, proveedor, etc.)
        # ======================================================
        filtros, where, args, _ = collect_gastos_filters(request, session)

        # ======================================================
        # 2) scope igual que la lista
        # ======================================================
        _apply_role_scope_for_exports(conn, where, args)

        # ======================================================
        # 3) refuerzos: gerente y tipo (tu lógica original)
        # ======================================================
        tipo = (request.args.get("tipo") or "").strip().lower()

        gerente_id_raw = (request.args.get("gerente_id") or "").strip()
        if gerente_id_raw.isdigit():
            gerente_id = int(gerente_id_raw)

            # Trae subordinados (y opcionalmente incluyes al gerente)
            sub_ids = set(_obtener_subordinados(conn, gerente_id) or [])
            sub_ids.add(gerente_id)

            placeholders = ",".join("?" * len(sub_ids)) if sub_ids else "?"
            where.append(f"g.usuario_id IN ({placeholders})")
            args.extend(list(sub_ids) if sub_ids else [gerente_id])
        if tipo in ("caja chica", "caja_chica"):
            where.append("COALESCE(g.es_caja_chica,0)=1")
        elif tipo in ("reembolso", "reembolso vendedor", "reembolso_vendedor"):
            where.append("COALESCE(g.reembolso_vendedor,0)=1")
        elif tipo in (
            "tarjeta", "tarjeta corporativa", "tarjeta_corporativa",
            "tarjeta_online", "tarjeta credito", "tarjeta crédito"
        ):
            where.append("COALESCE(g.es_caja_chica,0)=0")
            where.append("COALESCE(g.reembolso_vendedor,0)=0")




        # ======================================================
        # 4) ✅ Query: si hay ids -> SOLO ids (con validación scope)
        #     si NO hay ids -> filtros normales (tu comportamiento)
        # ======================================================
        if ids_req:
            ids_ok = _fetch_allowed_ids_in_scope(conn, ids_req)  # helper que te pasé

            if not ids_ok:
                rows = []
            else:
                placeholders = ",".join("?" * len(ids_ok))

                # Si quieres mantener el ORDEN visible, usa ORDER BY CASE.
                # Si no lo quieres, cambia a "ORDER BY date(g.fecha) DESC, g.id DESC"
                order_case = " ".join([f"WHEN ? THEN {i}" for i, _id in enumerate(ids_ok)])

                sql = f"""
                    SELECT
                        g.id,
                        g.fecha,
                        COALESCE(g.es_caja_chica,0) AS es_caja_chica,
                        COALESCE(g.reembolso_vendedor,0) AS reembolso_vendedor,

                        COALESCE(
                            NULLIF(TRIM(g.motivo), ''),
                            (SELECT d.motivo
                               FROM gastos_tarjeta_detalle d
                              WHERE d.gasto_id=g.id
                                AND TRIM(COALESCE(d.motivo,'')) <> ''
                              ORDER BY d.id
                              LIMIT 1)
                        ) AS motivo_resuelto,

                        COALESCE(
                            NULLIF(TRIM(g.centro_costo), ''),
                            (SELECT d.centro_costo
                               FROM gastos_tarjeta_detalle d
                              WHERE d.gasto_id=g.id
                                AND TRIM(COALESCE(d.centro_costo,'')) <> ''
                              ORDER BY d.id
                              LIMIT 1)
                        ) AS centro_resuelto,

                        g.ccb,
                        g.con_soporte, g.sin_soporte, g.subtotal_factura,
                        g.servicios_10, g.subtotal_sin_iva, g.iva, g.total_con_iva, g.archivo,
                        COALESCE(t.nombre, g.proveedor) AS proveedor_nombre,
                        COALESCE(g.ga_aprobado,0)    AS ga_aprobado,
                        COALESCE(g.ga_aprobado_por,0) AS ga_aprobado_por,
                        g.ga_aprobado_at             AS ga_aprobado_at,

                        COALESCE(g.gg_aprobado,0)    AS gg_aprobado,
                        COALESCE(g.gg_aprobado_por,0) AS gg_aprobado_por,
                        g.gg_aprobado_at             AS gg_aprobado_at,

                        COALESCE(g.gf_aprobado,0)    AS gf_aprobado,
                        COALESCE(g.gf_aprobado_por,0) AS gf_aprobado_por,
                        g.gf_aprobado_at             AS gf_aprobado_at,
                        COALESCE(g.boletos_aereos,0) AS boletos_aereos

                    FROM {TABLE_GASTOS} g
                    LEFT JOIN usuarios u ON u.id = g.usuario_id   -- 🔥 IMPORTANTE

                    LEFT JOIN terceros t ON t.id = g.proveedor_id
                    WHERE g.id IN ({placeholders})
                    ORDER BY CASE g.id {order_case} END
                """
                # args: ids_ok para IN + ids_ok para CASE
                cur.execute(sql, list(ids_ok) + list(ids_ok))
                rows = cur.fetchall()

        else:
            sql = (
                f"""
                SELECT
                    g.id,
                    g.fecha,
                    COALESCE(g.es_caja_chica,0) AS es_caja_chica,
                    COALESCE(g.reembolso_vendedor,0) AS reembolso_vendedor,

                    COALESCE(
                        NULLIF(TRIM(g.motivo), ''),
                        (SELECT d.motivo
                           FROM gastos_tarjeta_detalle d
                          WHERE d.gasto_id=g.id
                            AND TRIM(COALESCE(d.motivo,'')) <> ''
                          ORDER BY d.id
                          LIMIT 1)
                    ) AS motivo_resuelto,

                    COALESCE(
                        NULLIF(TRIM(g.centro_costo), ''),
                        (SELECT d.centro_costo
                           FROM gastos_tarjeta_detalle d
                          WHERE d.gasto_id=g.id
                            AND TRIM(COALESCE(d.centro_costo,'')) <> ''
                          ORDER BY d.id
                          LIMIT 1)
                    ) AS centro_resuelto,

                    g.ccb,
                    g.con_soporte, g.sin_soporte, g.subtotal_factura,
                    g.servicios_10, g.subtotal_sin_iva, g.iva, g.total_con_iva, g.archivo,
                    COALESCE(t.nombre, g.proveedor) AS proveedor_nombre,
                    COALESCE(g.ga_aprobado,0)    AS ga_aprobado,
                    COALESCE(g.ga_aprobado_por,0) AS ga_aprobado_por,
                    g.ga_aprobado_at             AS ga_aprobado_at,

                    COALESCE(g.gg_aprobado,0)    AS gg_aprobado,
                    COALESCE(g.gg_aprobado_por,0) AS gg_aprobado_por,
                    g.gg_aprobado_at             AS gg_aprobado_at,

                    COALESCE(g.gf_aprobado,0)    AS gf_aprobado,
                    COALESCE(g.gf_aprobado_por,0) AS gf_aprobado_por,
                    g.gf_aprobado_at             AS gf_aprobado_at,
                    COALESCE(g.boletos_aereos,0) AS boletos_aereos

                FROM {TABLE_GASTOS} g
                LEFT JOIN usuarios u ON u.id = g.usuario_id   -- 🔥 IMPORTANTE

                LEFT JOIN terceros t ON t.id = g.proveedor_id
                """
                + (" WHERE " + " AND ".join(where) if where else "")
                + """
                ORDER BY date(g.fecha) DESC, g.id DESC
                """
            )
            cur.execute(sql, args)
            rows = cur.fetchall()

    except Exception:
        current_app.logger.exception("[EXPORT] fallo construyendo/ejecutando query")
        rows = []
    finally:
        # NO cierres aún: conn se usa para header si hay plantilla
        pass

    n = len(rows)

    # ======================================================
    # Plantilla (si existe)
    # ======================================================
    template_path = os.path.join(
        current_app.static_folder,
        "plantillas",
        "Modelo Excel Reporte Tarjeta de Crédito.xlsx",
    )

    if os.path.exists(template_path):
        wb = load_workbook(template_path)
        ws = wb.active

        # Llenar cabecera con datos del usuario logeado
        try:
            hdr = _get_usuario_header_data(conn)

            # 🔹 Verificar si existe al menos un gasto aprobado por gerente
            aprobado = False
            if rows:
                for row in rows:
                    rd = dict(row)
                    # Si es caja chica o reembolso -> validar GA
                    if rd.get("es_caja_chica") or rd.get("reembolso_vendedor"):
                        if rd.get("ga_aprobado") == 1:
                            aprobado = True
                    else:
                        # tarjeta normal requiere flujo completo
                        if (
                            rd.get("ga_aprobado") == 1 and
                            rd.get("gg_aprobado") == 1 and
                            rd.get("gf_aprobado") == 1
                        ):
                            aprobado = True

            if not aprobado:
                hdr["aprobado_por"] = ""
            def _parse_dt(s):
                if not s:
                    return None
                s = str(s).strip()
                if not s:
                    return None
                try:
                    return datetime.fromisoformat(s)  # soporta 'YYYY-MM-DD' y 'YYYY-MM-DD HH:MM:SS'
                except Exception:
                    try:
                        return datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
                    except Exception:
                        try:
                            return datetime.strptime(s[:10], "%Y-%m-%d")
                        except Exception:
                            return None

            def _pick_aprobacion_header(rows):
                """
                Retorna (aprobado_por_id, aprobado_at_dt) en base a las filas exportadas.
                Regla:
                - si es caja_chica/reembolso/boletos -> GA
                - caso contrario -> GF
                Si exportas varias filas, toma la fecha MÁS RECIENTE.
                """
                best_dt = None
                best_uid = None

                for r in rows or []:
                    rd = dict(r)
                    es_restringido = (
                        int(rd.get("es_caja_chica") or 0) == 1
                        or int(rd.get("reembolso_vendedor") or 0) == 1
                        or int(rd.get("boletos_aereos") or 0) == 1
                    )

                    if es_restringido:
                        if int(rd.get("ga_aprobado") or 0) == 1:
                            dt = _parse_dt(rd.get("ga_aprobado_at"))
                            if dt and (best_dt is None or dt > best_dt):
                                best_dt = dt
                                best_uid = int(rd.get("ga_aprobado_por") or 0) or None
                    else:
                        if int(rd.get("gf_aprobado") or 0) == 1:
                            dt = _parse_dt(rd.get("gf_aprobado_at"))
                            if dt and (best_dt is None or dt > best_dt):
                                best_dt = dt
                                best_uid = int(rd.get("gf_aprobado_por") or 0) or None

                return best_uid, best_dt





            hdr = _get_usuario_header_data(conn)  # aquí tú ya vas a usar "gerente real" si quieres
            aprob_uid, aprob_dt = _pick_aprobacion_header(rows)

            hdr["aprobado_por"] = ""
            hdr["aprobado_fecha"] = ""
            hdr["revisado_por"] = ""

            if aprob_dt:
                # ✅ Aprobado por = gerente real del usuario que descarga (solo si hay aprobación)
                gerente_id = _get_ultimo_jefe_activo(
                    conn,
                    int(session.get("usuario_id") or session.get("user_id") or 0)
                )

                aprob_name = ""
                if gerente_id:
                    cur2 = conn.cursor()
                    cur2.execute(
                        "SELECT COALESCE(nombre_completo, username, '') AS n FROM usuarios WHERE id=? LIMIT 1",
                        (int(gerente_id),)
                    )
                    rr = cur2.fetchone()
                    aprob_name = (rr["n"] or "") if rr else ""

                hdr["aprobado_por"] = aprob_name or ""
                hdr["aprobado_fecha"] = aprob_dt.strftime("%Y-%m-%d")
                hdr["revisado_por"] = aprob_name or ""

            _fill_excel_header(ws, hdr, filtros)


        except Exception:
            pass

        # ✅ Título dinámico según tipos de gasto presentes
        try:
            titulo = _build_report_title_from_rows(rows)
            _set_report_title(ws, titulo)
        except Exception:
            pass

        # (se quedaba duplicado en tu función; lo mantengo igual para no cambiarte comportamientos)
        try:
            hdr = _get_usuario_header_data(conn)
            #_fill_excel_header(ws, hdr, filtros)
        except Exception:
            pass

        # Asegurar encabezado "CCB" aun si la plantilla no lo tiene
        header_row = DETAIL_START_ROW - 1
        try:
            hdr_cell = ws.cell(header_row, 14)
            if not (hdr_cell.value and str(hdr_cell.value).strip()):
                hdr_cell.value = "CCB"
                hdr_cell.alignment = Alignment(horizontal="center", vertical="center")
                hdr_cell.font = Font(bold=True)
        except Exception:
            pass

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        headers = [
            "Año", "Mes", "Día",
            "Motivo del gasto",
            "Nombre del Proveedor",
            "Centro de costo",
            "Gastos con soporte",
            "Gastos sin soporte",
            "Subtotal factura",
            "10% servicios",
            "Subtotal sin IVA",
            "IVA",
            "Total con IVA",
            "CCB",
        ]
        ws.append(headers)
        DETAIL_START_ROW = 2  # fallback

    # Ya podemos cerrar conexión
    try:
        conn.close()
    except Exception:
        pass

    # ======================================================
    # Ajustes de ancho de columna
    # ======================================================
    try:
        colw = {
            1: 6, 2: 6, 3: 6,
            4: 32, 5: 32, 6: 18,
            7: 14, 8: 14, 9: 14,
            10: 16, 11: 16, 12: 12,
            13: 14, 14: 8,
        }
        for c, w in colw.items():
            ws.column_dimensions[chr(64 + c)].width = w
    except Exception:
        pass

    # ======================================================
    # Ajusta fila Total pegada al detalle
    # ======================================================
    total_row = _find_row_contains(ws, TOTAL_ANCHOR_TEXT)
    if total_row is None:
        total_row = max(DETAIL_START_ROW + n + 1, DETAIL_START_ROW + 3)
        ws.insert_rows(total_row)
        ws.cell(total_row, TOTAL_COL_START).value = TOTAL_ANCHOR_TEXT

    capacidad = total_row - DETAIL_START_ROW
    if capacidad < 0:
        ws.insert_rows(DETAIL_START_ROW, amount=abs(capacidad) + n + 1)
        total_row = DETAIL_START_ROW + n + 1
        capacidad = total_row - DETAIL_START_ROW

    if n > capacidad:
        ws.insert_rows(total_row, amount=n - capacidad)
        total_row += (n - capacidad)
    elif n < capacidad and capacidad > 0:
        ws.delete_rows(DETAIL_START_ROW + n, amount=capacidad - n)
        total_row -= (capacidad - n)

    end_detail_row = DETAIL_START_ROW + max(n - 1, 0)

    # Limpia merges en bloque
    _unmerge_in_area(ws, DETAIL_START_ROW, 1, max(total_row, end_detail_row), DETAIL_COL_COUNT)
    ws.cell(total_row, TOTAL_COL_START).value = TOTAL_ANCHOR_TEXT

    # ======================================================
    # Detalle
    # ======================================================
    def _safe_parse_date(val):
        """
        Evita que el export se caiga si 'fecha' viene corrupta (ej: 'VACUNA ...').
        Retorna datetime o None.
        """
        if val is None:
            return None
        if isinstance(val, datetime):
            return val
        if isinstance(val, date):
            return datetime(val.year, val.month, val.day)

        s = str(val).strip()
        if not s:
            return None

        # 1) ISO (YYYY-MM-DD o YYYY-MM-DDTHH:MM:SS)
        try:
            return datetime.fromisoformat(s)
        except Exception:
            pass

        # 2) primeros 10 chars tipo YYYY-MM-DD
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d")
        except Exception:
            return None

    r = DETAIL_START_ROW
    for row in rows:
        rd = dict(row)

        dt = _safe_parse_date(rd.get("fecha"))
        # si fecha inválida, dejamos blanco en año/mes/día
        y = dt.year if dt else ""
        m = dt.month if dt else ""
        d = dt.day if dt else ""

        data = [
            y,
            m,
            d,
            rd.get("motivo_resuelto") or "",
            rd.get("proveedor_nombre") or "",
            rd.get("centro_resuelto") or "",
            rd.get("con_soporte") or 0,
            rd.get("sin_soporte") or 0,
            rd.get("subtotal_factura") or 0,
            rd.get("servicios_10") or 0,
            rd.get("subtotal_sin_iva") or 0,
            rd.get("iva") or 0,
            rd.get("total_con_iva") or 0,
            ("Sí" if (rd.get("ccb") in (1, True, "1")) else "No"),
        ]

        for c, val in enumerate(data, start=1):
            cell = ws.cell(r, c, val)
            if c in (4, 5, 6):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif 7 <= c <= 13:
                try:
                    float(val)
                    cell.number_format = NUM_FMT
                except Exception:
                    pass
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
        r += 1

    # ======================================================
    # Fila Total sin fórmulas (G..M), etiqueta en F
    # ======================================================
    end_detail_row = DETAIL_START_ROW + max(n - 1, 0)

    def _sum_col(col_idx, r1, r2):
        s = 0.0
        for rr in range(r1, r2 + 1):
            v = ws.cell(rr, col_idx).value
            try:
                s += float(v or 0)
            except Exception:
                pass
        return s

    _unmerge_row(ws, total_row, c1=1, c2=ws.max_column)
    ws.delete_rows(total_row, 1)
    ws.insert_rows(total_row, 1)

    lbl = ws.cell(total_row, LABEL_TOTAL_COL, TOTAL_ANCHOR_TEXT)
    lbl.alignment = Alignment(vertical="center")

    for col in range(TOTAL_COL_START, TOTAL_COL_END + 1):
        total_value = _sum_col(col, DETAIL_START_ROW, end_detail_row)
        cell = ws.cell(total_row, col, round(total_value, 2))
        cell.number_format = NUM_FMT
        cell.alignment = Alignment(horizontal="right", vertical="center")

    _style_total_row(
        ws,
        total_row,
        label_col=LABEL_TOTAL_COL,
        block_start=TOTAL_COL_START,
        block_end=TOTAL_COL_END,
    )

    if n > 0:
        _apply_border_range(ws, DETAIL_START_ROW, 1, end_detail_row, DETAIL_COL_COUNT)

    try:
        ws.freeze_panes = ws.cell(DETAIL_START_ROW, 1)
    except Exception:
        pass

    return wb

# ==========================================================
# PDF conversion
# ==========================================================
def _find_soffice():
    cfg = (current_app.config.get("SOFFICE_BIN") if current_app else None)
    candidates = [
        os.environ.get("SOFFICE_BIN"),
        cfg,
        "soffice",
        "/usr/bin/soffice",
        "/snap/bin/soffice",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for c in candidates:
        if not c:
            continue
        if os.path.isabs(c) and os.path.exists(c):
            return c
        w = shutil.which(c)
        if w:
            return w
    return None


def _convert_with_excel_com(xlsx_path, pdf_path):
    import win32com.client, pythoncom

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
    try:
        xlTypePDF = 0
        wb.ExportAsFixedFormat(xlTypePDF, os.path.abspath(pdf_path))
    finally:
        wb.Close(SaveChanges=False)
        excel.Quit()
        pythoncom.CoUninitialize()


def _xlsx_to_pdf(xlsx_path, pdf_path):
    soffice = _find_soffice()
    if soffice:
        outdir = os.path.dirname(pdf_path)
        base = os.path.splitext(os.path.basename(xlsx_path))[0]
        proc = subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir", outdir, xlsx_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        lo_pdf = os.path.join(outdir, f"{base}.pdf")
        if proc.returncode == 0 and os.path.exists(lo_pdf):
            if os.path.abspath(lo_pdf) != os.path.abspath(pdf_path):
                os.replace(lo_pdf, pdf_path)
            return
        raise RuntimeError(f"LibreOffice falló: {proc.stdout}\n{proc.stderr}")

    if os.name == "nt":
        _convert_with_excel_com(xlsx_path, pdf_path)
        return

    raise RuntimeError("No se encontró LibreOffice ni Excel COM para convertir a PDF.")


# ==========================================================
# Flask responses
# ==========================================================
def export_gastos_excel_response():
    current_app.logger.warning("EXPORT EXCEL args=%r", dict(request.args))

    wb = _build_gastos_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=gastos_tarjeta.xlsx"},
    )

 
def export_gastos_pdf_response():
    wb = _build_gastos_workbook()
    with tempfile.TemporaryDirectory() as tmp:
        xlsx_path = os.path.join(tmp, "gastos_tarjeta.xlsx")
        pdf_path = os.path.join(tmp, "gastos_tarjeta.pdf")
        wb.save(xlsx_path)
        _xlsx_to_pdf(xlsx_path, pdf_path)
        with open(pdf_path, "rb") as f:
            pdf_bytes = f.read()

    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": "attachment; filename=gastos_tarjeta.pdf"},
    )


def export_gastos_csv_response():
    conn = get_db()
    cur = conn.cursor()
    try:
        filtros, where, args, _ = collect_gastos_filters(request, session)

        # aplicar mismo scope de la lista
        _apply_role_scope_for_exports(conn, where, args)

        sql = (
            f"""
            SELECT
                g.id,
                g.fecha,
                COALESCE(
                    NULLIF(TRIM(g.motivo), ''),
                    (SELECT d.motivo
                       FROM gastos_tarjeta_detalle d
                      WHERE d.gasto_id = g.id
                        AND TRIM(COALESCE(d.motivo,'')) <> ''
                      ORDER BY d.id
                      LIMIT 1)
                ) AS motivo_resuelto,
                COALESCE(
                    NULLIF(TRIM(g.centro_costo), ''),
                    (SELECT d.centro_costo
                       FROM gastos_tarjeta_detalle d
                      WHERE d.gasto_id = g.id
                        AND TRIM(COALESCE(d.centro_costo,'')) <> ''
                      ORDER BY d.id
                      LIMIT 1)
                ) AS centro_resuelto,
                g.ccb,
                g.con_soporte, g.sin_soporte, g.subtotal_factura,
                g.servicios_10, g.subtotal_sin_iva, g.iva, g.total_con_iva,
                g.archivo,
                COALESCE(t.nombre, g.proveedor) AS proveedor_nombre
            FROM {TABLE_GASTOS} g
            LEFT JOIN usuarios u ON u.id = g.usuario_id   -- 🔥 IMPORTANTE

            LEFT JOIN terceros t ON t.id = g.proveedor_id
            """
            + (" WHERE " + " AND ".join(where) if where else "")
            + """
            ORDER BY date(g.fecha) DESC, g.id DESC
            """
        )

        cur.execute(sql, args)
        rows = cur.fetchall()
    finally:
        try:
            conn.close()
        except Exception:
            pass

    import csv

    buff = io.StringIO()
    writer = csv.writer(buff)

    headers = [
        "ID",
        "Fecha",
        "Motivo",
        "Proveedor",
        "Centro de costo",
        "Con soporte",
        "Sin soporte",
        "Subtotal factura",
        "Servicios 10%",
        "Subtotal sin IVA",
        "IVA",
        "Total con IVA",
        "Archivo",
        "CCB",
    ]
    writer.writerow(headers)

    for r in rows:
        rd = dict(r)
        writer.writerow(
            [
                rd.get("id"),
                rd.get("fecha"),
                rd.get("motivo_resuelto") or "",
                rd.get("proveedor_nombre") or "",
                rd.get("centro_resuelto") or "",
                rd.get("con_soporte") or 0,
                rd.get("sin_soporte") or 0,
                rd.get("subtotal_factura") or 0,
                rd.get("servicios_10") or 0,
                rd.get("subtotal_sin_iva") or 0,
                rd.get("iva") or 0,
                rd.get("total_con_iva") or 0,
                rd.get("archivo") or "",
                "Sí" if (rd.get("ccb") in (1, True, "1")) else "No",
            ]
        )

    csv_bytes = ("\ufeff" + buff.getvalue()).encode("utf-8")  # BOM para Excel
    return Response(
        csv_bytes,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=gastos_tarjeta.csv"},
    )
