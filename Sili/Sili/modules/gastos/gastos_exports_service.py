# modules/gastos/gastos_exports_service.py

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .gastos_repository import get_detalle_by_gasto_id
from .gastos_repo import fetch_gastos_rows_for_report


def _to_float(value: Any) -> float:
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def _safe_str(value: Any) -> str:
    return "" if value is None else str(value)


def build_reporte_excel_bytes(conn, ids: list[int] | None = None) -> BytesIO:

    filtros, rows = fetch_gastos_rows_for_report(conn, force_ids=ids)
    gastos = [dict(r) for r in rows]

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Gastos"

    headers = [
        "ID Gasto",
        "Fecha",
        "Proveedor",
        "Motivo",
        "Centro",
        "CCB",
        "Subtotal Factura",
        "IVA",
        "Total con IVA",
        "GA Aprobado",
        "GG Aprobado",
        "GF Aprobado",
        "Detalle Item",
        "Detalle Descripción",
        "Detalle Subtotal",
        "Detalle IVA",
        "Detalle Total",
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for g in gastos:
        detalle = get_detalle_by_gasto_id(conn, int(g["id"]))
        if not detalle:
            detalle = [{}]

        for det in detalle:
            ws.append(
                [
                    g.get("id"),
                    g.get("fecha") or "",
                    g.get("proveedor_nombre") or "",
                    g.get("motivo_resuelto") or "",
                    g.get("centro_resuelto") or "",
                    "SI" if int(g.get("ccb") or 0) == 1 else "NO",
                    _to_float(g.get("subtotal_factura")),
                    _to_float(g.get("iva")),
                    _to_float(g.get("total_con_iva")),
                    "SI" if int(g.get("ga_aprobado") or 0) == 1 else "NO",
                    "SI" if int(g.get("gg_aprobado") or 0) == 1 else "NO",
                    "SI" if int(g.get("gf_aprobado") or 0) == 1 else "NO",
                    det.get("id", ""),
                    det.get("descripcion") or det.get("observacion") or "",
                    _to_float(det.get("subtotal_factura") or det.get("subtotal") or det.get("base_imponible")),
                    _to_float(det.get("iva")),
                    _to_float(det.get("total_con_iva") or det.get("total") or det.get("total_linea")),
                ]
            )

    widths = {
        1: 12, 2: 14, 3: 28, 4: 30, 5: 18, 6: 10, 7: 16, 8: 12, 9: 16,
        10: 12, 11: 12, 12: 12, 13: 12, 14: 36, 15: 16, 16: 12, 17: 16,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    for r in range(2, ws.max_row + 1):
        for c in (7, 8, 9, 15, 16, 17):
            ws.cell(r, c).number_format = '#,##0.00'

    ws.auto_filter.ref = f"A1:Q{ws.max_row}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output




# modules/gastos/gastos_exports_service.py

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .gastos_repository import get_detalle_by_gasto_id


def _to_float(value: Any) -> float:
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def _safe_str(value: Any) -> str:
    return "" if value is None else str(value)


def build_reporte_excel_bytes_from_rows(conn, gastos: list[dict]) -> BytesIO:
    """
    Exporta exactamente las filas ya filtradas por la pantalla.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Gastos"

    headers = [
        "ID Gasto",
        "Fecha",
        "Usuario",
        "Departamento",
        "Proveedor",
        "Motivo",
        "Número Factura",
        "Tipo",
        "CCB",
        "Subtotal Factura",
        "IVA",
        "Total con IVA",
        "GA Aprobado",
        "GG Aprobado",
        "GF Aprobado",
        "SAP",
        "Detalle Item",
        "Detalle Descripción",
        "Detalle Subtotal",
        "Detalle IVA",
        "Detalle Total",
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for g in gastos:
        detalle = get_detalle_by_gasto_id(conn, int(g["id"]))
        if not detalle:
            detalle = [{}]

        for det in detalle:
            ws.append(
                [
                    g.get("id"),
                    g.get("fecha") or g.get("created_at") or "",
                    g.get("usuario_username") or "",
                    g.get("departamento_nombre") or "",
                    g.get("proveedor_nombre") or g.get("proveedor") or "",
                    g.get("motivo") or "",
                    g.get("numero_factura") or "",
                    g.get("tipo_gasto") or g.get("tipo") or "",
                    "SI" if int(g.get("ccb") or 0) == 1 else "NO",
                    _to_float(g.get("subtotal_factura")),
                    _to_float(g.get("iva")),
                    _to_float(g.get("total_con_iva")),
                    "SI" if int(g.get("ga_aprobado") or 0) == 1 else "NO",
                    "SI" if int(g.get("gg_aprobado") or 0) == 1 else "NO",
                    "SI" if int(g.get("gf_aprobado") or 0) == 1 else "NO",
                    _safe_str(g.get("sap_contabilizacion")),
                    det.get("id", ""),
                    det.get("descripcion") or det.get("observacion") or "",
                    _to_float(det.get("subtotal_factura") or det.get("subtotal") or det.get("base_imponible")),
                    _to_float(det.get("iva")),
                    _to_float(det.get("total_con_iva") or det.get("total") or det.get("total_linea")),
                ]
            )

    widths = {
        1: 12, 2: 14, 3: 18, 4: 18, 5: 28, 6: 30, 7: 18, 8: 16,
        9: 10, 10: 16, 11: 12, 12: 16, 13: 12, 14: 12, 15: 12,
        16: 18, 17: 12, 18: 36, 19: 16, 20: 12, 21: 16,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    for r in range(2, ws.max_row + 1):
        for c in (10, 11, 12, 19, 20, 21):
            ws.cell(r, c).number_format = '#,##0.00'

    ws.auto_filter.ref = f"A1:U{ws.max_row}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output