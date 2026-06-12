# modules/routes_reclamos.py
# -*- coding: utf-8 -*-
from __future__ import annotations
from flask import app, jsonify, request, session
from .db import get_db
from .security import require_login

from datetime import datetime
from flask import jsonify, session
import os
from openai import OpenAI
 
# routes_reclamos.py
import os
from uuid import uuid4
from werkzeug.utils import secure_filename
from flask import abort
import sqlite3
from datetime import datetime
import smtplib
from email.message import EmailMessage
import io
from flask import request, jsonify, session
from modules.om_chat_service import om_chat_responder
from flask import (
    render_template, request, redirect, url_for,
    flash, session, current_app, jsonify, send_file
)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from flask import (
    render_template, request, redirect, url_for,
    flash, session, current_app, jsonify
)

from .db import get_db, get_config_value
from .security import (
    require_login,
    require_permission,
)
from .routes_reclamos_querys import *
from .routes_reclamos_constants import *

# =========================================================
# Helpers internos
# =========================================================

BOSS_ROLES = (
    'jefe',
    'gerente',
    'gerente general',
    'gerente financiero',
    'coordinador',
    'admin'
)


 
# ----------------------------------------------
# Helpers
# ----------------------------------------------
import sqlite3
def _get_sponsors_by_proceso(conn, proceso_id):
    cur = conn.cursor()
    cur.execute(SQL__GET_SPONSORS_BY_PROCESO_SEL_1, (proceso_id,))

    ids = []
    seen = set()

    for r in cur.fetchall():
        uid = r["usuario_id"]
        if uid and int(uid) not in seen:
            seen.add(int(uid))
            ids.append(str(int(uid)))

    return ids


def _can_upload_carta_cliente(conn, uid: int | None) -> bool:
    role = (session.get("rol") or "").strip().lower()

    if role in ("admin", "administrador"):
        return True

    if not uid:
        return False

    cur = conn.cursor()

    cur.execute(SQL__CAN_UPLOAD_CARTA_CLIENTE_SEL_1, (uid,))

    row = cur.fetchone()

    if not row:
        return False

    departamento = (row["departamento_nombre"] or "").strip().lower()
    puesto = (row["puesto_nombre"] or "").strip().lower()

    if departamento == "servicio al cliente":
        return True

    if "servicio al cliente" in puesto:
        return True

    return False

def _can_delete_om(conn, uid: int | None) -> bool:
    """
    Puede eliminar OM:
    - admin / coordinador
    - usuarios del departamento Servicio al Cliente
    - usuarios con puesto que contenga Servicio al Cliente
    """
    role = (session.get("rol") or "").strip().lower()

    if role in ("admin", "administrador", "coordinador"):
        return True

    if not uid:
        return False

    cur = conn.cursor()

    cur.execute(SQL__CAN_UPLOAD_CARTA_CLIENTE_SEL_1, (uid,))

    row = cur.fetchone()

    if not row:
        return False

    departamento = (row["departamento_nombre"] or "").strip().lower()
    puesto = (row["puesto_nombre"] or "").strip().lower()

    if departamento == "servicio al cliente":
        return True

    if "servicio al cliente" in puesto:
        return True

    return False


def _usuario_es_sponsor_del_proceso(conn, proceso_id: int | None, user_id: int | None) -> bool:
    if not proceso_id or not user_id:
        return False

    cur = conn.cursor()
    cur.execute(SQL__USUARIO_ES_SPONSOR_DEL_PROCESO_SEL_1, (proceso_id, user_id))

    return cur.fetchone() is not None

def _fetch_recl_procesos(conn):
    cur = conn.cursor()
    cur.execute(SQL__FETCH_RECL_PROCESOS_SEL_1)
    return cur.fetchall()

def _get_sponsor_principal_by_proceso(conn, proceso_id):
    cur = conn.cursor()
    cur.execute(SQL__GET_SPONSOR_PRINCIPAL_BY_PROCESO_SEL_1, (proceso_id,))

    row = cur.fetchone()
    return int(row["usuario_id"]) if row and row["usuario_id"] else None

def _raw_conn(conn):
    return getattr(conn, "_conn", conn)

def _is_sqlserver_conn(conn) -> bool:
    raw = _raw_conn(conn)
    mod = getattr(raw.__class__, "__module__", "").lower()
    name = getattr(raw.__class__, "__name__", "").lower()
    text = f"{mod} {name}"
    return "pyodbc" in text or "odbc" in text

def _table_exists(conn, table: str) -> bool:
    cur = conn.cursor()

    if _is_sqlserver_conn(conn):
        cur.execute(SQL__TABLE_EXISTS_SEL_1, (table,))
        return cur.fetchone() is not None

    cur.execute(SQL__TABLE_EXISTS_SEL_2, (table,))
    return cur.fetchone() is not None

def _col_names(conn, table: str) -> set[str]:
    cur = conn.cursor()

    try:
        if _is_sqlserver_conn(conn):
            cur.execute(SQL__COL_NAMES_SEL_1, (table,))
            return {str(r["name"]) for r in cur.fetchall()}

        cur.execute(f"PRAGMA table_info({table})")
        rows = cur.fetchall()
        names = set()
        for r in rows:
            try:
                names.add(str(r["name"]))
            except Exception:
                names.add(str(r[1]))
        return names
    except Exception:
        return set()

def _ensure_column(conn, table, column, decl_sql):
    # En SQL Server no alteramos esquema desde runtime
    if _is_sqlserver_conn(conn):
        return

    cur = conn.execute(f"PRAGMA table_info({table})")
    cols = {row["name"] for row in cur.fetchall()}
    if column not in cols:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {decl_sql}")
        conn.commit()

def _run_reclamos_bootstrap_if_needed(conn):
    if _is_sqlserver_conn(conn):
        return

    #ensure_reclamos_schema(conn)
    #ensure_reclamos_catalogos(conn)
    #ensure_geo_schema(conn)
    #_ensure_reclamo_imputados_extra_cols(conn)
    
def _can_export_all_reclamos(conn, uid: int | None) -> bool:
    """
    Puede exportar TODO:
    - rol admin / coordinador
    - o si pertenece a Servicio al Cliente (por puesto o por departamento)
    """
    role = (session.get("rol") or "").strip().lower()
    if role in {"admin", "coordinador"}:
        return True

    if not uid:
        return False

    cur = conn.cursor()

    if _is_sqlserver_conn(conn):
        cur.execute(SQL__CAN_EXPORT_ALL_RECLAMOS_SEL_1, (uid,))
    else:
        cur.execute(SQL__CAN_EXPORT_ALL_RECLAMOS_SEL_2, (uid,))

    row = cur.fetchone()
    if not row:
        return False

    puesto = (row["puesto_nombre"] or "").strip()

    if "SERVICIO AL CLIENTE" in puesto:
        return True

    return False

def _sponsor_acciones_upload_dir() -> str:
    """
    Carpeta física para evidencias de acciones del sponsor.
    Ajusta aquí si ya manejas otra ruta base de uploads.
    """
    base = os.path.join(current_app.root_path, "uploads", "reclamo_imputado_acciones")
    os.makedirs(base, exist_ok=True)
    return base

def _puede_gestionar_imputado_accion(accion_id: int, user_id: int | None = None):
    if not user_id:
        user_id = _current_user_id()

    if not user_id:
        return False, None

    db = get_db()

    if _is_sqlserver_conn(db):
        row = db.execute(SQL__PUEDE_GESTIONAR_IMPUTADO_ACCION_SEL_1, (accion_id,)).fetchone()
    else:
        row = db.execute(SQL__PUEDE_GESTIONAR_IMPUTADO_ACCION_SEL_1, (accion_id,)).fetchone()

    if not row:
        return False, None

    if _is_admin_like():
        return True, row

    es_duenio = int(row["imputado_id"] or 0) == int(user_id or 0)
    aprobada = (row["estado_asignacion"] or "").strip().lower() == "aprobado"

    return bool(es_duenio and aprobada), row

 
def _can_view_all_reclamos(conn, uid: int | None) -> bool:
    if _is_admin_like():
        return True

    role = (session.get('rol') or '').strip().lower()
    if role in ('gerente', 'gerente financiero'):
        return True

    if not uid:
        return False

    cur = conn.cursor()

    try:
        if _is_sqlserver_conn(conn):
            cur.execute(SQL__CAN_VIEW_ALL_RECLAMOS_SEL_1, (uid,))
        else:
            cur.execute(SQL__CAN_VIEW_ALL_RECLAMOS_SEL_2, (uid,))

        row = cur.fetchone()
        if not row:
            return False

        departamento = (row["departamento_nombre"] or "").strip().lower()
        puesto = (row["puesto_nombre"] or "").strip().upper()

        if departamento == "servicio al cliente":
            return True

        if "SERVICIO AL CLIENTE" in puesto:
            return True

    except Exception:
        pass

    return False

def _can_view_all_reclamos_sn_sponsor(conn, uid: int | None) -> bool:
    """
    Regla para la pestaña "Soy Sponsor":
    - Si el rol es 'gerente', NO ve todo
    - Admin-like sí ve todo, excepto gerente
    - Gerente financiero sí ve todo
    - Servicio al Cliente sí ve todo
    """
    role = (session.get("rol") or "").strip().lower()

    if role == "gerente":
        return False

    if _is_admin_like():
        return True

    if role == "gerente financiero":
        return True

    if not uid:
        return False

    cur = conn.cursor()

    try:
        if _is_sqlserver_conn(conn):
            cur.execute(SQL__CAN_UPLOAD_CARTA_CLIENTE_SEL_1, (uid,))
        else:
            cur.execute(SQL__CAN_VIEW_ALL_RECLAMOS_SN_SPONSOR_SEL_1, (uid,))

        row = cur.fetchone()
        if not row:
            return False

        departamento = (row["departamento_nombre"] or "").strip().lower()
        puesto = (row["puesto_nombre"] or "").strip().upper()

        if departamento == "servicio al cliente":
            return True

        if "SERVICIO AL CLIENTE" in puesto:
            return True

    except Exception:
        pass

    return False


def _es_miembro_equipo_reclamo(reclamo_id, user_id):
    current_app.logger.debug(
        "[equipo] Chequeando miembro equipo: reclamo_id=%s user_id=%s",
        reclamo_id, user_id
    )

    if not user_id or not reclamo_id:
        current_app.logger.debug("[equipo] -> False (falta user_id o reclamo_id)")
        return False

    db = get_db()

    row = db.execute(SQL__ES_MIEMBRO_EQUIPO_RECLAMO_SEL_1, (reclamo_id, user_id)).fetchone()

    current_app.logger.debug(
        "[equipo] Resultado query miembro equipo: row=%r -> %s",
        row, bool(row)
    )

    return row is not None

def _notify_sponsor_respuesta_equipo(conn, imputacion_id: int, miembro_id: int, reclamo_codigo: str):
    cur = conn.cursor()

    # =========================================================
    # 1) Obtener datos base de la imputación / reclamo
    # =========================================================
    cur.execute(SQL__NOTIFY_SPONSOR_RESPUESTA_EQUIPO_SEL_1, (miembro_id, miembro_id, miembro_id, imputacion_id))

    base = cur.fetchone()

    if not base:
        return

    reclamo_id = base["reclamo_id"]
    proceso_id = base["proceso_id"]
    miembro_nombre = (
        base["miembro_nombre"]
        or base["miembro_username"]
        or f"UID {miembro_id}"
    )

    # =========================================================
    # 2) Obtener destinatarios: PRINCIPAL + BACKUP del proceso
    #    No desde reclamo_imputados, porque ahí solo existe principal.
    # =========================================================
    sponsor_rows = []

    if proceso_id:
        cur.execute(SQL__NOTIFY_SPONSOR_RESPUESTA_EQUIPO_SEL_2, (proceso_id,))

        sponsor_rows = cur.fetchall()

    # =========================================================
    # 3) Fallback de seguridad:
    #    Si por algún motivo no encuentra sponsors por proceso,
    #    notifica al imputado principal de reclamo_imputados.
    # =========================================================
    if not sponsor_rows:
        cur.execute(SQL__NOTIFY_SPONSOR_RESPUESTA_EQUIPO_SEL_3, (imputacion_id,))

        sponsor_rows = cur.fetchall()

    if not sponsor_rows:
        return

    try:
        link_sponsor = url_for("reclamos", _external=True) + "?tab=imputado"
    except Exception:
        link_sponsor = "http://bitacoraquimpac.com.ec:5000/reclamos?tab=imputado"

    subject = f"[Oportunidad de Mejora] Respuesta registrada por miembro de equipo en {reclamo_codigo}"

    def _parse_dt(v):
        if not v:
            return None

        if isinstance(v, datetime):
            return v

        s = str(v).strip()

        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%Y-%m-%dT%H:%M:%S"
        ):
            try:
                return datetime.strptime(s[:19], fmt)
            except Exception:
                pass

        return None

    def _dias_respuesta(row):
        f_asig = _parse_dt(row["fecha_asignacion_miembro"])
        f_resp = _parse_dt(row["fecha_respuesta_miembro"])

        if not f_asig and not f_resp:
            return "Sin fechas registradas"

        if not f_asig:
            return "Sin fecha de asignación"

        if not f_resp:
            return "Sin fecha de respuesta"

        dias = (f_resp.date() - f_asig.date()).days

        if dias <= 0:
            return "Mismo día"

        return f"{dias} día(s)"

    def _row_mail(lbl, val):
        val = "" if val is None else str(val)
        val = val.replace("\n", "<br>")

        return (
            "<tr>"
            f"<td style='width:210px;background:#ffedd5;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;color:#374151;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;color:#374151;'>"
            f"{val}</td>"
            "</tr>"
        )

    tiempo_respuesta = _dias_respuesta(base)
    enviados = set()

    # =========================================================
    # 4) Enviar a principal + backup, sin duplicar por correo
    # =========================================================
    for s in sponsor_rows:
        sponsor_email = (s["sponsor_email"] or "").strip().lower()

        if not sponsor_email or sponsor_email in enviados:
            continue

        enviados.add(sponsor_email)

        sponsor_nombre = (
            s["sponsor_nombre"]
            or s["sponsor_username"]
            or "Usuario"
        )

        tipo_sponsor = (s["tipo_sponsor"] or "").strip().upper()

        text_body = f"""Hola {sponsor_nombre},

El miembro de equipo {miembro_nombre} registró su respuesta de apoyo para la Oportunidad de Mejora {reclamo_codigo}.

Resumen:
- Cliente: {base["cliente_nombre"] or ""}
- Tipo de OM: {base["tipo_reclamo"] or ""}
- Tipo de trámite: {base["tipo_tramite"] or ""}
- Proceso: {base["proceso_text"] or ""}
- Rol sponsor: {tipo_sponsor}
- Tiempo de respuesta: {tiempo_respuesta}

Por favor ingresa al sistema y revisa el aporte en la pestaña "Soy Sponsor".

Ir al sistema: {link_sponsor}

Este es un mensaje automático.
"""

        html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">
            <tr>
              <td style="background:#f59e0b;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;letter-spacing:.08em;opacity:.95;">
                  Oportunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Respuesta registrada por miembro de equipo
                </div>
                <div style="font-size:12px;opacity:.95;margin-top:6px;">
                  Hola {sponsor_nombre}, el usuario <strong>{miembro_nombre}</strong> ya registró su aporte para la OM {reclamo_codigo}.
                </div>
              </td>
            </tr>

            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row_mail('Código', base['codigo'])}
                  {_row_mail('Miembro que respondió', miembro_nombre)}
                  {_row_mail('Rol sponsor', tipo_sponsor)}
                  {_row_mail('Cliente', base['cliente_nombre'])}
                  {_row_mail('Tipo de OM', base['tipo_reclamo'])}
                  {_row_mail('Proceso', base['proceso_text'])}
                  {_row_mail('Antecedente', base['antecedente'])}
                  {_row_mail('Observación', base['observacion'])}
                </table>

                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link_sponsor}"
                     style="display:inline-block;background:#f59e0b;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Revisar aporte del equipo
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  Ingresa al módulo de reclamos y revisa la pestaña <strong>“Soy Sponsor”</strong>.
                </div>
              </td>
            </tr>

            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

        _send_mail_safe(
            sponsor_email,
            subject,
            text_body,
            html_body=html_body
        )


def _now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _current_user_id() -> int | None:
    return (
        session.get("user_id")
        or session.get("usuario_id")
        or session.get("id")
    )

def _get_respuesta_equipo_acciones_full(db, respuesta_equipo_id: int):
    rows = db.execute(SQL__GET_RESPUESTA_EQUIPO_ACCIONES_FULL_SEL_1, (respuesta_equipo_id,)).fetchall()

    causas = []
    control = []
    correctiva = []

    for r in rows:
        accion_id = int(r["id"])

        evid_rows = db.execute(SQL__GET_RESPUESTA_EQUIPO_ACCIONES_FULL_SEL_2, (accion_id,)).fetchall()

        evidencias = []
        for e in evid_rows:
            ev_id = e["id"]

            evidencias.append({
                "id": ev_id,
                "filename": e["filename"] or "",
                "original_name": e["original_name"] or "",
                "content_type": e["content_type"] or "",
                "size_bytes": int(e["size_bytes"] or 0),
                "created_at": e["created_at"] or "",
                "download_url": url_for(
                    "reclamo_equipo_accion_evidencia_download",
                    evidencia_id=ev_id
                )
            })

        item = {
            "id": accion_id,
            "tipo": (r["tipo"] or "").upper(),
            "descripcion": r["descripcion"] or "",
            "fecha_compromiso": r["fecha_compromiso"] or "",
            "orden": int(r["orden"] or 1),
            "requiere_evidencia": int(r["requiere_evidencia"] or 0),
            "cumplido": int(r["cumplido"] or 0),
            "fecha_cumplimiento": r["fecha_cumplimiento"] or "",
            "evidencias": evidencias,
            "observacion_cumplimiento": r["observacion_cumplimiento"] or "",
        }

        tipo = item["tipo"]
        if tipo == "CAUSA":
            causas.append(item)
        elif tipo == "CONTROL":
            control.append(item)
        elif tipo == "CORRECTIVA":
            correctiva.append(item)

    return {
        "causas": causas,
        "control": control,
        "correctiva": correctiva,
    }
 
def _is_admin_like() -> bool:
    rol = (session.get("rol") or "").strip().lower()
    return rol in ("admin", "coordinador")


def _notify_reclamo_adjuntos_change(conn, reclamo_id: int, actor_id: int | None,
                                    accion: str, filenames: list[str]):
    """
    accion: 'agregados' o 'eliminados'
    filenames: lista de nombres originales afectados
    """
    cur = conn.cursor()
    cur.execute(SQL__NOTIFY_RECLAMO_ADJUNTOS_CHANGE_SEL_1, (reclamo_id,))
    r = cur.fetchone()
    if not r:
        return

    codigo = r["codigo"]
    creador_id = r["creado_por"]

    creador = _get_user_basic(conn, creador_id)
    actor = _get_user_basic(conn, actor_id) if actor_id else None

    if not creador or ("email" not in creador.keys()) or not creador["email"]:
        return
 
    actor_name = (
        actor["nombre_completo"]
        if actor and "nombre_completo" in actor.keys() and actor["nombre_completo"]
        else (actor["username"] if actor else "Sistema")
    )

    lista = "\n".join(f"- {fn}" for fn in filenames) or "(sin detalle)"

    subject = f"[Oportunidad de Mejora] Adjuntos {accion} en {codigo}"
    text_body = f"""Hola {creador['username']},

Se han {accion} archivos en la Oportunidad de Mejora {codigo}.

Acción realizada por: {actor_name}

Archivos:
{lista}

Este es un mensaje automático.
"""

    _send_mail_safe(creador["email"], subject, text_body)

# =========================================================
#   ESQUEMA DE BD
# =========================================================

def _col_names(conn, table: str) -> set[str]:
    cur = conn.cursor()
    try:
        cur.execute(f"PRAGMA table_info({table})")
        return {r["name"] for r in cur.fetchall()}
    except sqlite3.OperationalError:
        return set()

def ensure_reclamos_schema(conn):
    """
    No-op en SQL Server.
    El esquema de reclamos ya debe existir y mantenerse fuera del backend.
    """
    return

def ensure_reclamo_imputado_acciones_schema(conn: sqlite3.Connection):
    """
    No-op en SQL Server.
    El esquema de reclamos ya debe existir y mantenerse fuera del backend.
    """
    return


def _save_respuesta_equipo_acciones(
    conn: sqlite3.Connection,
    respuesta_id: int,
    reclamo_id: int,
    imputacion_id: int,
    miembro_id: int,
    causas: list,
    control: list,
    correctiva: list,
    user_id: int | None,
):
    cur = conn.cursor()
    now = _now_iso()

    # Inactivar acciones anteriores de esta respuesta
    cur.execute(SQL__SAVE_RESPUESTA_EQUIPO_ACCIONES_UPD_1, (now, user_id, respuesta_id))

    def _normalizar_items(items):
        normalizados = []

        for item in (items or []):
            if isinstance(item, dict):
                desc = (item.get("descripcion") or "").strip()
                fecha = (item.get("fecha_compromiso") or "").strip()
            elif isinstance(item, (list, tuple)) and len(item) >= 2:
                desc = (item[0] or "").strip()
                fecha = (item[1] or "").strip()
            else:
                continue

            if not desc and not fecha:
                continue

            normalizados.append((desc, fecha))

        return normalizados

    def _insert_many(tipo: str, items: list, requiere_evidencia: int = 0):
        items_norm = _normalizar_items(items)

        for idx, (desc, fecha) in enumerate(items_norm, start=1):
            cur.execute(SQL__SAVE_RESPUESTA_EQUIPO_ACCIONES_INS_2, (
                respuesta_id,
                reclamo_id,
                imputacion_id,
                miembro_id,
                tipo,
                desc,
                fecha,
                idx,
                requiere_evidencia,
                now,
                user_id
            ))

    _insert_many("CAUSA", causas, 0)
    _insert_many("CONTROL", control, 0)
    _insert_many("CORRECTIVA", correctiva, 1)
 



def ensure_reclamo_respuesta_equipo_acciones_schema(conn: sqlite3.Connection):
    cur = conn.cursor()

    cur.execute(SQL_ENSURE_RECLAMO_RESPUESTA_EQUIPO_ACCIONES_SCHEMA_DDL_1)

    cur.execute(SQL_ENSURE_RECLAMO_RESPUESTA_EQUIPO_ACCIONES_SCHEMA_DDL_2)

    cur.execute("PRAGMA table_info(reclamo_respuesta_equipo_acciones)")
    cols = {r[1] for r in cur.fetchall()}

    required_cols = {
        "respuesta_equipo_id": "INTEGER NOT NULL DEFAULT 0",
        "reclamo_id": "INTEGER NOT NULL DEFAULT 0",
        "imputacion_id": "INTEGER NOT NULL DEFAULT 0",
        "miembro_id": "INTEGER NOT NULL DEFAULT 0",
        "tipo": "TEXT NOT NULL DEFAULT ''",
        "descripcion": "TEXT NOT NULL DEFAULT ''",
        "fecha_compromiso": "TEXT",
        "orden": "INTEGER NOT NULL DEFAULT 1",
        "requiere_evidencia": "INTEGER NOT NULL DEFAULT 0",
        "cumplido": "INTEGER NOT NULL DEFAULT 0",
        "fecha_cumplimiento": "TEXT",
        "reminder_3d_sent": "INTEGER NOT NULL DEFAULT 0",
        "reminder_2d_sent": "INTEGER NOT NULL DEFAULT 0",
        "reminder_1d_sent": "INTEGER NOT NULL DEFAULT 0",
        "escalado_jefe": "INTEGER NOT NULL DEFAULT 0",
        "activo": "INTEGER NOT NULL DEFAULT 1",
        "created_at": "TEXT",
        "created_by": "INTEGER",
        "updated_at": "TEXT",
        "updated_by": "INTEGER",
    }

    for col, ddl in required_cols.items():
        if col not in cols:
            cur.execute(f"ALTER TABLE reclamo_respuesta_equipo_acciones ADD COLUMN {col} {ddl}")

    cur.execute(SQL_ENSURE_RECLAMO_RESPUESTA_EQUIPO_ACCIONES_SCHEMA_DDL_3)
    cur.execute(SQL_ENSURE_RECLAMO_RESPUESTA_EQUIPO_ACCIONES_SCHEMA_DDL_4)
    cur.execute(SQL_ENSURE_RECLAMO_RESPUESTA_EQUIPO_ACCIONES_SCHEMA_DDL_5)
    cur.execute(SQL_ENSURE_RECLAMO_RESPUESTA_EQUIPO_ACCIONES_SCHEMA_DDL_6)
    cur.execute(SQL_ENSURE_RECLAMO_RESPUESTA_EQUIPO_ACCIONES_SCHEMA_DDL_7)
    cur.execute(SQL_ENSURE_RECLAMO_RESPUESTA_EQUIPO_ACCIONES_SCHEMA_DDL_8)
    cur.execute(SQL_ENSURE_RECLAMO_RESPUESTA_EQUIPO_ACCIONES_SCHEMA_DDL_9)

    cur.execute(SQL_ENSURE_RECLAMO_RESPUESTA_EQUIPO_ACCIONES_SCHEMA_DDL_10)
    cur.execute(SQL_ENSURE_RECLAMO_RESPUESTA_EQUIPO_ACCIONES_SCHEMA_DDL_11)



    cur.execute("PRAGMA table_info(reclamo_respuesta_equipo_acciones)")
    cols = {r[1] for r in cur.fetchall()}

    if "observacion_cumplimiento" not in cols:
        cur.execute("ALTER TABLE reclamo_respuesta_equipo_acciones ADD COLUMN observacion_cumplimiento TEXT")

    if "updated_at" not in cols:
        cur.execute("ALTER TABLE reclamo_respuesta_equipo_acciones ADD COLUMN updated_at TEXT")

    if "updated_by" not in cols:
        cur.execute("ALTER TABLE reclamo_respuesta_equipo_acciones ADD COLUMN updated_by INTEGER")

    conn.commit()
 

def _get_respuesta_equipo_acciones(conn: sqlite3.Connection, respuesta_equipo_id: int):
    cur = conn.cursor()
    rows = cur.execute(SQL__GET_RESPUESTA_EQUIPO_ACCIONES_SEL_1, (respuesta_equipo_id,)).fetchall()

    causas, controles, correctivas = [], [], []

    for r in rows:
        item = {
            "id": r["id"],
            "descripcion": r["descripcion"] or "",
            "fecha_compromiso": r["fecha_compromiso"] or "",
            "orden": r["orden"] or 1,
            "requiere_evidencia": int(r["requiere_evidencia"] or 0),
            "cumplido": int(r["cumplido"] or 0),
            "fecha_cumplimiento": r["fecha_cumplimiento"] or "",
        }
        if r["tipo"] == "CAUSA":
            causas.append(item)
        elif r["tipo"] == "CONTROL":
            controles.append(item)
        elif r["tipo"] == "CORRECTIVA":
            correctivas.append(item)

    return {
        "causas": causas,
        "control": controles,
        "correctiva": correctivas,
    }

def _normalizar_acciones_payload(items, label: str):
    out = []

    if not items:
        return out, None

    if not isinstance(items, list):
        return None, f"{label} debe enviarse como lista."

    for it in items:
        it = it or {}
        if not isinstance(it, dict):
            continue

        desc = (it.get("descripcion") or "").strip()
        fecha = (it.get("fecha_compromiso") or "").strip()

        if not desc and not fecha:
            continue

        if not desc:
            return None, f"Falta descripción en {label}."
        if not fecha:
            return None, f"Falta fecha compromiso en {label}."
        if not _is_date_yyyy_mm_dd(fecha):
            return None, f"Fecha inválida en {label}. Use YYYY-MM-DD."

        out.append((desc, fecha))

    return out, None


def _get_imputado_acciones_full(db, imputacion_id: int):
    rows = db.execute(SQL__GET_IMPUTADO_ACCIONES_FULL_SEL_1, (imputacion_id,)).fetchall()

    causas = []
    control = []
    correctiva = []

    for r in rows:
        accion_id = int(r["id"])

        evid_rows = db.execute(SQL__GET_IMPUTADO_ACCIONES_FULL_SEL_2, (accion_id,)).fetchall()

        evidencias = []
        for e in evid_rows:
            ev_id = e["id"]

            evidencias.append({
                "id": ev_id,
                "filename": e["filename"] or "",
                "original_name": e["original_name"] or "",
                "content_type": e["content_type"] or "",
                "size_bytes": int(e["size_bytes"] or 0),
                "created_at": e["created_at"] or "",
                "download_url": url_for(
                    "reclamo_imputado_accion_evidencia_download",
                    evidencia_id=ev_id
                )
            })

        item = {
            "id": accion_id,
            "tipo": (r["tipo"] or "").upper(),
            "descripcion": r["descripcion"] or "",
            "fecha_compromiso": r["fecha_compromiso"] or "",
            "orden": int(r["orden"] or 1),
            "requiere_evidencia": int(r["requiere_evidencia"] or 0),
            "cumplido": int(r["cumplido"] or 0),
            "fecha_cumplimiento": r["fecha_cumplimiento"] or "",
            "observacion_cumplimiento": r["observacion_cumplimiento"] or "",
            "evidencias": evidencias,
        }

        if item["tipo"] == "CAUSA":
            causas.append(item)
        elif item["tipo"] == "CONTROL":
            control.append(item)
        elif item["tipo"] == "CORRECTIVA":
            correctiva.append(item)

    return {
        "causas": causas,
        "control": control,
        "correctiva": correctiva,
    }


BOSS_ROLES = (
    'jefe',
    'gerente',
    'gerente general',
    'gerente financiero',
    'coordinador',
    'admin','usuario'
)
 

def _puede_gestionar_equipo(reclamo_id: int, user_id: int | None = None) -> bool:
    if not user_id:
        user_id = _current_user_id()
        if not user_id:
            return False

    db = get_db()

    fila = db.execute(SQL__PUEDE_GESTIONAR_EQUIPO_SEL_1, (reclamo_id, user_id, user_id)).fetchone()

    return fila is not None


def _puede_ver_equipo(reclamo_id: int, user_id: int | None = None) -> bool:
    """
    Puede ver:
    - el imputado aprobado (responsable del proceso),
    - los miembros del equipo,
    - el creador del reclamo.
    """
    if not user_id:
        user_id = _current_user_id()
        if not user_id:
            return False

    db = get_db()

    # 1) si puede gestionar, también puede ver
    if _puede_gestionar_equipo(reclamo_id, user_id):
        return True

    # 2) miembro del equipo
    fila = db.execute(
        SQL__PUEDE_VER_EQUIPO_SEL_1,
        (reclamo_id, user_id),
    ).fetchone()

    if fila:
        return True

    # 3) creador del reclamo
    fila = db.execute(
        SQL__PUEDE_VER_EQUIPO_SEL_2,
        (reclamo_id, user_id),
    ).fetchone()

    return fila is not None


def _ensure_column(conn, table, column, decl_sql):
    """
    No-op en SQL Server.
    El esquema ya debe existir y mantenerse fuera del backend.
    """
    return

def ensure_reclamo_imputados_fishbone(conn):
    """
    No-op en SQL Server.
    El esquema ya debe existir y mantenerse fuera del backend.
    """
    return

def ensure_reclamo_respuestas_equipo_schema(conn: sqlite3.Connection):
    """
    No-op en SQL Server.
    El esquema ya debe existir y mantenerse fuera del backend.
    """
    return


def ensure_reclamo_adjuntos_schema(conn: sqlite3.Connection):
    """
    No-op en SQL Server.
    El esquema ya debe existir y mantenerse fuera del backend.
    """
    return

# =========================================================
#   CATÁLOGOS param_groups / param_values
# =========================================================

def fetch_productos(conn):
    cur = conn.cursor()
    cur.execute(SQL_FETCH_PRODUCTOS_SEL_1)
    return cur.fetchall()


def _ensure_param_tables(conn: sqlite3.Connection):
    """
    Adapta los catálogos a tu esquema actual:

        param_groups(id, nombre)
        param_values(id, group_id, nombre, valor)

    y, si hace falta, añade columnas 'activo' y 'orden'
    a param_values (no rompe nada existente).
    """
    cur = conn.cursor()

    cur.execute(SQL__ENSURE_PARAM_TABLES_DDL_1)

    cur.execute(SQL__ENSURE_PARAM_TABLES_DDL_2)

    cur.execute("PRAGMA table_info(param_values)")
    cols = {r[1] for r in cur.fetchall()}
    if "activo" not in cols:
        cur.execute("ALTER TABLE param_values ADD COLUMN activo INTEGER NOT NULL DEFAULT 1")
    if "orden" not in cols:
        cur.execute("ALTER TABLE param_values ADD COLUMN orden  INTEGER NOT NULL DEFAULT 1")

    conn.commit()

 
def _ensure_param_group(conn, nombre: str, descripcion: str | None = None):
    """
    En SQL Server no crea grupos en runtime.
    Solo devuelve el id existente o None si no existe.
    """
    cur = conn.cursor()
    cur.execute(SQL__ENSURE_PARAM_GROUP_SEL_1, (nombre,))
    row = cur.fetchone()
    return row["id"] if row else None


def _ensure_param_value(conn: sqlite3.Connection,
                        group_codigo: str,
                        clave: str,
                        valor: str,
                        orden: int = 1,
                        activo: int = 1):
    """
    Usamos:
        param_values.nombre -> clave interna (OTROS, DEMORA_ENTREGA, etc.)
        param_values.valor  -> etiqueta visible ("Otros", "Demora en entrega"...)
    """
    gid = _ensure_param_group(conn, group_codigo, group_codigo)
    cur = conn.cursor()

    cur.execute(SQL__ENSURE_PARAM_VALUE_SEL_1, (gid, clave))
    row = cur.fetchone()

    if row:
        cur.execute(SQL__ENSURE_PARAM_VALUE_UPD_2, (valor, orden, activo, row["id"]))
    else:
        cur.execute(SQL__ENSURE_PARAM_VALUE_INS_3, (gid, clave, valor, orden, activo))

    conn.commit()


def _fetch_param_values(conn: sqlite3.Connection, group_codigo: str):
    """
    Devuelve lista de rows de un grupo concreto.

    Cada row tendrá:
        id, nombre (clave), valor (etiqueta), orden
    """
    #_ensure_param_tables(conn)
    cur = conn.cursor()
    cur.execute(SQL__FETCH_PARAM_VALUES_SEL_1, (group_codigo,))
    return cur.fetchall()


from datetime import datetime
from flask import request, jsonify, session, abort

ACC_TIPOS_VALIDOS = {"CAUSA", "CONTROL", "CORRECTIVA"}

def _now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def _is_date_yyyy_mm_dd(s: str) -> bool:
    try:
        datetime.strptime(s, "%Y-%m-%d")
        return True
    except Exception:
        return False

def _can_edit_equipo(equipo_id: int) -> bool:
    # superadmin/admin
    if session.get("rol") == "admin" or session.get("is_admin") is True:
        return True

    user_id = session.get("user_id") or session.get("id")
    if not user_id:
        return False

    conn = get_db()
    cur = conn.cursor()
    cur.execute(SQL__CAN_EDIT_EQUIPO_SEL_1, (equipo_id,))
    row = cur.fetchone()
    if not row:
        return False

    responsable_id = row["responsable_id"] if hasattr(row, "keys") else row[0]
    colaborador_id = row["colaborador_id"] if hasattr(row, "keys") else row[1]

    return int(user_id) in {int(responsable_id), int(colaborador_id)}


def ensure_reclamos_catalogos(conn):
    """
    No-op en SQL Server.
    Los catálogos ya deben existir y mantenerse fuera del backend.
    """
    return
# =========================================================
#   CATÁLOGOS GEO
# =========================================================
 


def ensure_geo_schema(conn):
    """
    No-op en SQL Server.
    Las tablas geográficas ya deben existir y mantenerse fuera del backend.
    """
    return


def fetch_regiones(conn: sqlite3.Connection):
    cur = conn.cursor()
    cur.execute(SQL_FETCH_REGIONES_SEL_1)
    return cur.fetchall()


def fetch_provincias(conn: sqlite3.Connection, region_id: int | None):
    cur = conn.cursor()
    if region_id:
        cur.execute(SQL_FETCH_PROVINCIAS_SEL_1, (region_id,))
    else:
        cur.execute(SQL_FETCH_PROVINCIAS_SEL_2)
    return cur.fetchall()


def fetch_cantones(conn: sqlite3.Connection, provincia_id: int | None):
    cur = conn.cursor()
    if provincia_id:
        cur.execute(SQL_FETCH_CANTONES_SEL_1, (provincia_id,))
    else:
        cur.execute(SQL_FETCH_CANTONES_SEL_2)
    return cur.fetchall()


def _fetch_tipo_campos(conn: sqlite3.Connection):
    cur = conn.cursor()
    cur.execute("SELECT * FROM reclamo_tipo_campos")
    rows = cur.fetchall()
    return {r["tipo_codigo"]: r for r in rows}


# =========================================================
#   Helpers de negocio
# =========================================================

def _generate_codigo_reclamo(conn: sqlite3.Connection) -> str:
    cur = conn.cursor()
    cur.execute(SQL__GENERATE_CODIGO_RECLAMO_SEL_1)
    row = cur.fetchone()
    if row and row["codigo"]:
        num_part = "".join(ch for ch in row["codigo"] if ch.isdigit())
        try:
            nxt = int(num_part) + 1
        except ValueError:
            nxt = 1
    else:
        nxt = 1
    return f"RECL{nxt:05d}"


def _guess_aprobador_for_user2(conn: sqlite3.Connection, user_id: int | None) -> int | None:
    if not user_id:
        return None

    cur = conn.cursor()
    cur.execute(SQL__GUESS_APROBADOR_FOR_USER2_SEL_1, (user_id,))
    u = cur.fetchone()
    if not u:
        return None

    depto_id = u["departamento_id"]

    cur.execute(f"""
        SELECT TOP 1 id
        FROM usuarios
        WHERE departamento_id = ?
          AND LOWER(rol) IN ({",".join(["?"]*len(BOSS_ROLES))})
        
    """, (depto_id, *[r.lower() for r in BOSS_ROLES]))
    boss = cur.fetchone()
    return boss["id"] if boss else None


def _guess_aprobador_for_user(conn: sqlite3.Connection, user_id: int | None) -> int | None:
    """
    Devuelve el jefe directo del usuario usando usuarios.jefe_id.
    Si no tiene jefe_id definido o el jefe está deshabilitado, cae
    al esquema viejo de buscar alguien del mismo departamento con
    rol en BOSS_ROLES.
    """
    if not user_id:
        return None

    cur = conn.cursor()
    # Primero intentamos con jefe_id
    cur.execute(SQL__GUESS_APROBADOR_FOR_USER_SEL_1, (user_id,))
    u = cur.fetchone()
    if not u:
        return None

    jefe_id = u["jefe_id"]
    depto_id = u["departamento_id"]

    # Si tiene jefe_id definido, validamos que exista y no esté deshabilitado
    if jefe_id:
        cur.execute(SQL__GUESS_APROBADOR_FOR_USER_SEL_2, (jefe_id,))
        j = cur.fetchone()
        if j:
            return j["id"]

    # Fallback: buscar jefe por rol en el mismo departamento (esquema antiguo)
    if not depto_id:
        return None

    cur.execute(f"""
        SELECT TOP 1 id
        FROM usuarios
        WHERE departamento_id = ?
          AND LOWER(rol) IN ({",".join(["?"] * len(BOSS_ROLES))})
          AND COALESCE(disabled, 0) = 0
        ORDER BY id
        
    """, (depto_id, *[r.lower() for r in BOSS_ROLES]))
    boss = cur.fetchone()
    return boss["id"] if boss else None


# =========================================================
#   EMAIL HELPER UNIFICADO
# =========================================================

def _send_mail_safe(to_email: str, subject: str, text_body: str, html_body: str | None = None):
    """
    Envía correo multipart/alternative (texto + HTML) usando la configuración
    SMTP guardada en base de datos (get_config_value).
    """
    if not to_email:
        return

    host = get_config_value('smtp_host') or ''
    port = int(get_config_value('smtp_port') or "587")
    user = get_config_value('smtp_user') or ''
    pwd = get_config_value('smtp_pass') or ''
    from_addr = get_config_value('smtp_from') or user or ''

    if not host or not from_addr:
        current_app.logger.warning("SMTP no configurado correctamente, se omite envío de correo.")
        return

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["To"] = to_email
    msg["From"] = from_addr

    if html_body:
        msg.set_content(text_body or "(ver versión HTML)")
        msg.add_alternative(html_body, subtype="html")
    else:
        msg.set_content(text_body or "")

    try:
        with smtplib.SMTP(host, port, timeout=20) as s:
            s.starttls()
            if user and pwd:
                s.login(user, pwd)
            s.send_message(msg)
        current_app.logger.info("Correo enviado a %s con asunto '%s'", to_email, subject)
    except Exception as e:
        current_app.logger.exception("Error enviando correo: %s", e)


def fetch_usuarios_imputables2(conn, empresa_ids=None):
    """
    Devuelve usuarios imputables con:
      - id
      - username
      - nombre_completo
      - departamento_nombre
      - jefe_nombre  (primer usuario del mismo dpto con rol en BOSS_ROLES)
    Opcionalmente filtra por empresa_id (lista de ints).
    """
    cur = conn.cursor()

    boss_roles = [r.lower() for r in BOSS_ROLES]
    boss_placeholders = ",".join(["?"] * len(boss_roles))

    sql = f"""
        SELECT
            u.id,
            u.username,
            COALESCE(u.nombre_completo, u.username) AS nombre_completo,
            d.nombre AS departamento_nombre,
            (
                SELECT TOP 1
                COALESCE(j.nombre_completo, j.username)
                FROM usuarios j
                WHERE j.departamento_id = u.departamento_id
                  AND LOWER(j.rol) IN ({boss_placeholders})
                ORDER BY j.id
                
            ) AS jefe_nombre,
            u.empresa_id
        FROM usuarios u
        LEFT JOIN departamentos d ON d.id = u.departamento_id
        WHERE COALESCE(u.disabled, 0) = 0
    """

    params = boss_roles[:]

    if empresa_ids:
        emp_placeholders = ",".join(["?"] * len(empresa_ids))
        sql += f" AND COALESCE(u.empresa_id,0) IN ({emp_placeholders})"
        params.extend(empresa_ids)

    sql += " ORDER BY nombre_completo, u.username"

    cur.execute(sql, params)
    return cur.fetchall()


def fetch_usuarios_imputables(conn, empresa_ids=None):
    """
    Devuelve usuarios imputables con:
      - id
      - username
      - nombre_completo
      - departamento_nombre
      - jefe_nombre  (usando usuarios.jefe_id)
    Opcionalmente filtra por empresa_id (lista de ints).
    """
    cur = conn.cursor()

    sql = """
        SELECT
            u.id,
            u.username,
            COALESCE(u.nombre_completo, u.username) AS nombre_completo,
            d.nombre AS departamento_nombre,
            COALESCE(j.nombre_completo, j.username) AS jefe_nombre,
            u.empresa_id,
            u.jefe_id
        FROM usuarios u
        LEFT JOIN departamentos d ON d.id = u.departamento_id
        LEFT JOIN usuarios j ON j.id = u.jefe_id          -- 👈 jefe desde la misma tabla
        WHERE COALESCE(u.disabled, 0) = 0
        and u.identificacion not in ('40623','0911946630','0923577688','0929626729','1307590834'   ,'40736','0902507805','1714868211')
    """
    params = []

    if empresa_ids:
        emp_placeholders = ",".join(["?"] * len(empresa_ids))
        sql += f" AND COALESCE(u.empresa_id, 0) IN ({emp_placeholders})"
        params.extend(empresa_ids)

    sql += " ORDER BY nombre_completo, u.username"

    cur.execute(sql, params)
    return cur.fetchall()


def _get_user_basic(conn: sqlite3.Connection, uid: int | None):
    if not uid:
        return None
    cur = conn.cursor()
    cur.execute(SQL__GET_USER_BASIC_SEL_1, (uid,))
    return cur.fetchone()


def _notify_colaborador_asignado(
    conn,
    colaborador_id: int,
    reclamo_codigo: str,
    responsable_username: str
):
    u = _get_user_basic(conn, colaborador_id)
    if not u or ("email" not in u.keys()) or not u["email"]:
        return

    # --- Datos del reclamo (por codigo) ---
    cur = conn.cursor()
    cur.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_1, (reclamo_codigo,))
    r = cur.fetchone()

    # Imputados (si aplica)
    imputados = ""
    if r:
        if _is_sqlserver_conn(conn):
            cur.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_2, (r["id"],))
        else:
            if _is_sqlserver_conn(conn):
                cur.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_2, (r["id"],))
            else:
                cur.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_3, (r["id"],))
            row = cur.fetchone()
            if row and row["lista"]:
                imputados = row["lista"]
        row = cur.fetchone()
        if row and row["lista"]:
            imputados = row["lista"]

    # Link directo a tab "Soy Sponsor"
    try:
        link_sponsor = url_for("reclamos", _external=True) + "?tab=sponsor"
    except Exception:
        link_sponsor = "https://tu-sistema/reclamos?tab=sponsor"

    nombre = (u["nombre_completo"] or "").strip() if "nombre_completo" in u.keys() else ""
    if not nombre:
        nombre = u["username"]

    subject = f"[Oportunidad de Mejora] Aporte requerido en {reclamo_codigo}"

    # Texto plano (fallback)
    text_body = f"""Hola {nombre},

El responsable {responsable_username} te ha solicitado apoyo para la
Oportunidad de Mejora {reclamo_codigo}.

Por favor ingresa al sistema, pestaña "Soy Sponsor" y registra tu aporte
(causa, acción preventiva y acción correctiva).

Ir al sistema: {link_sponsor}

Este aporte no es la respuesta oficial, pero ayudará al responsable a
construir la respuesta final.

Este es un mensaje automático.
"""

    # ---------- HTML mejorado (mismo estilo que aprobador) ----------
    def _row(lbl, val):
        val = (val or "")
        val = str(val).replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#fee2e2;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;
               font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">
            <!-- Encabezado -->
            <tr>
              <td style="background:#b91c1c;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;
                            letter-spacing:.08em;opacity:.9;">
                  Oportunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Aporte requerido {reclamo_codigo}
                </div>
                <div style="font-size:12px;opacity:.9;margin-top:6px;">
                  Hola {nombre}, el responsable <strong>{responsable_username}</strong> solicitó tu apoyo.
                </div>
              </td>
            </tr>

            <!-- Cuerpo -->
            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row('Código', reclamo_codigo)}
                  {_row('Responsable', responsable_username)}
                  { _row('Fecha OM', r['fecha_reclamo']) if r else '' }
                  { _row('Tipo de OM', r['tipo_reclamo']) if r else '' }
                  { _row('Tipo de Trámite', r['tipo_tramite']) if r else '' }
                  { _row('Cliente', r['cliente_nombre']) if r else '' }
                  { _row('Proceso', r['proceso_text']) if r else '' }
                  { _row('Material', r['material_desc']) if (r and 'material_desc' in r.keys()) else '' }
                  { _row('Fecha de Pedido', r['fecha_pedido']) if r else '' }
                  { _row('Factura', r['factura']) if r else '' }
                  { _row('Guía Remisión', r['guia_remision']) if r else '' }
                  { _row('Imputados', imputados) if imputados else '' }
                  { _row('Antecedente', r['antecedente']) if r else '' }
                  { _row('Observación', r['observacion']) if r else '' }
                </table>

                <div style="margin-top:14px;font-size:12px;color:#374151;">
                  Por favor ingresa a la pestaña <strong>“Soy Sponsor”</strong> y registra tu aporte:
                  <strong>causa</strong>, <strong>acción preventiva</strong> y <strong>acción correctiva</strong>.
                  <br>
                  <span style="color:#6b7280;">
                    Este aporte no es la respuesta oficial, pero ayudará al responsable a construir la respuesta final.
                  </span>
                </div>

                <!-- Botón CTA -->
                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link_sponsor}"
                     style="display:inline-block;background:#2563eb;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Registrar aporte (Soy Sponsor)
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  También puedes ingresar al módulo de reclamos desde el sistema
                  y abrir la bandeja <strong>“Soy Sponsor”</strong>.
                </div>
              </td>
            </tr>

            <!-- Pie -->
            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>

          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

    _send_mail_safe(u["email"], subject, text_body, html_body=html_body)

def _notify_responsable_aporte_listo(conn, responsable_id: int, reclamo_codigo: str, colaborador_username: str):
    r = _get_user_basic(conn, responsable_id)
    if not r or ("email" not in r.keys()) or not r["email"]:
        return

    nombre = r["nombre_completo"] if r.get("nombre_completo") else r["username"]

    subject = f"[Oportunidad de Mejora] Aporte recibido en {reclamo_codigo}"
    text_body = f"""Hola {nombre},

El usuario {colaborador_username} registró su aporte técnico en la
Oportunidad de Mejora {reclamo_codigo}.

Puedes revisarlo en la sección "Soy Sponsor" y, si lo consideras adecuado,
aprovecharlo para tu respuesta final como responsable.

Este es un mensaje automático.
"""
    _send_mail_safe(r["email"], subject, text_body)

def _notify_colaborador_aporte_rechazado(conn, colaborador_id: int, reclamo_codigo: str, motivo: str):
    """
    Notifica rechazo de aporte de equipo a:
    - miembro/colaborador que registró el aporte
    - sponsor principal y backup del proceso
    - usuarios de Servicio al Cliente

    Reutiliza:
    - _get_user_basic
    - _get_sponsor_emails_by_reclamo
    - _send_mail_safe
    """

    colaborador = _get_user_basic(conn, colaborador_id)

    colaborador_email = ""
    colaborador_nombre = "Miembro de equipo"

    if colaborador:
        colaborador_email = (colaborador["email"] or "").strip()
        colaborador_nombre = (
            colaborador["nombre_completo"]
            if "nombre_completo" in colaborador.keys() and colaborador["nombre_completo"]
            else colaborador["username"]
        )

    motivo_txt = (motivo or "Sin detalle").strip()

    cur = conn.cursor()

    # =========================================================
    # Datos base de la OM
    # =========================================================
    cur.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_1, (reclamo_codigo,))

    r = cur.fetchone()

    try:
        link_sistema = url_for("reclamos", _external=True) + "?tab=sponsor"
    except Exception:
        link_sistema = "https://tu-sistema/reclamos?tab=sponsor"

    subject = f"[Oportunidad de Mejora] Aporte de equipo rechazado {reclamo_codigo}"

    def _row(lbl, val):
        val = "" if val is None else str(val)
        val = val.replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#fef3c7;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    # =========================================================
    # Destinatarios
    # =========================================================
    destinatarios = []

    # 1) Miembro/colaborador
    if colaborador_email:
        destinatarios.append({
            "email": colaborador_email,
            "nombre": colaborador_nombre,
            "rol_notificacion": "MIEMBRO DE EQUIPO"
        })

    # 2) Sponsor principal + backup
    try:
        for s in _get_sponsor_emails_by_reclamo(conn, reclamo_codigo):
            destinatarios.append({
                "email": s["email"],
                "nombre": s["nombre"] or s["username"] or "Usuario",
                "rol_notificacion": s["tipo_sponsor"] or "SPONSOR"
            })
    except Exception:
        current_app.logger.exception(
            "No se pudo obtener sponsor principal/backup para OM %s",
            reclamo_codigo
        )

    # 3) Servicio al Cliente
    cur.execute(SQL__NOTIFY_COLABORADOR_APORTE_RECHAZADO_SEL_1)

    for sc in cur.fetchall():
        destinatarios.append({
            "email": sc["email"],
            "nombre": sc["nombre"] or sc["username"] or "Servicio al Cliente",
            "rol_notificacion": "SERVICIO AL CLIENTE"
        })

    # =========================================================
    # Enviar sin duplicar correos
    # =========================================================
    enviados = set()

    for d in destinatarios:
        email = (d.get("email") or "").strip().lower()

        if not email or email in enviados:
            continue

        enviados.add(email)

        nombre_destinatario = d.get("nombre") or "Usuario"
        rol_notificacion = d.get("rol_notificacion") or ""

        text_body = f"""Hola {nombre_destinatario},

El aporte técnico registrado por {colaborador_nombre} para la Oportunidad de Mejora {reclamo_codigo} fue RECHAZADO.

Rol de notificación:
{rol_notificacion}

Motivo del rechazo:
{motivo_txt}

Por favor ingresa al sistema para revisar el detalle de la OM.

Ir al sistema:
{link_sistema}

Este es un mensaje automático.
"""

        html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;
               font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">

            <tr>
              <td style="background:#b45309;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;
                            letter-spacing:.08em;opacity:.9;">
                  Oportunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Aporte de equipo rechazado {reclamo_codigo}
                </div>
                <div style="font-size:12px;opacity:.9;margin-top:6px;">
                  Hola {nombre_destinatario}, se rechazó un aporte de equipo y requiere seguimiento.
                </div>
              </td>
            </tr>

            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row('Código', reclamo_codigo)}
                  {_row('Miembro de equipo', colaborador_nombre)}
                  {_row('Rol de notificación', rol_notificacion)}
                  { _row('Fecha OM', r['fecha_reclamo']) if r else '' }
                  { _row('Tipo de OM', r['tipo_reclamo']) if r else '' }
                  { _row('Tipo de Trámite', r['tipo_tramite']) if r else '' }
                  { _row('Cliente', r['cliente_nombre']) if r else '' }
                  { _row('Proceso', r['proceso_text']) if r else '' }
                  { _row('Material', r['material_desc']) if (r and 'material_desc' in r.keys()) else '' }
                  { _row('Fecha de Pedido', r['fecha_pedido']) if r else '' }
                  { _row('Factura', r['factura']) if r else '' }
                  { _row('Guía Remisión', r['guia_remision']) if r else '' }
                  { _row('Antecedente', r['antecedente']) if r else '' }
                  { _row('Observación', r['observacion']) if r else '' }
                  {_row('Motivo del rechazo', motivo_txt)}
                </table>

                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link_sistema}"
                     style="display:inline-block;background:#f97316;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Revisar OM en el sistema
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  Este correo fue enviado al miembro de equipo, sponsor principal,
                  backup y Servicio al Cliente.
                </div>
              </td>
            </tr>

            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>

          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

        _send_mail_safe(
            email,
            subject,
            text_body,
            html_body=html_body
        )


MAX_FILES_PER_RECLAMO = 5
MAX_FILE_SIZE_BYTES = 5 * 1024 * 1024  # 5 MB


def _get_reclamos_upload_folder() -> str:
    """
    Carpeta donde se guardan físicamente los adjuntos de reclamos.
    Puedes sobreescribirla con app.config['RECLAMOS_UPLOAD_FOLDER'].
    """
    base = current_app.config.get('RECLAMOS_UPLOAD_FOLDER')
    if not base:
        base = os.path.join(current_app.instance_path, 'reclamos_adjuntos')
    os.makedirs(base, exist_ok=True)
    return base


def _save_adjuntos_for_reclamo(conn: sqlite3.Connection,
                               reclamo_id: int,
                               files,
                               user_id: int | None) -> str | None:
    """
    Guarda adjuntos para un reclamo, respetando:
      - Máx 5 archivos por OM
      - Máx 5 MB por archivo
    Devuelve:
      - None si todo OK
      - Mensaje de error (str) si algo no cumple las reglas
    """
    #ensure_reclamo_adjuntos_schema(conn)
    cur = conn.cursor()

    # Cuántos adjuntos ya tiene esta OM
    cur.execute(SQL__SAVE_ADJUNTOS_FOR_RECLAMO_SEL_1, (reclamo_id,))
    row = cur.fetchone()
    existing = row["c"] if row else 0

    remaining = MAX_FILES_PER_RECLAMO - existing
    if remaining <= 0:
        return "Ya existe el máximo de archivos para esta OM (5)."

    upload_folder = _get_reclamos_upload_folder()
    saved_any = False

    for f in (files or [])[:remaining]:
        if not f or not f.filename:
            continue

        # Medir tamaño del archivo
        f.stream.seek(0, os.SEEK_END)
        size = f.stream.tell()
        f.stream.seek(0)

        if size > MAX_FILE_SIZE_BYTES:
            return f"El archivo '{f.filename}' supera el máximo permitido de 5 MB."

        original_name = f.filename
        safe_name = secure_filename(original_name) or "archivo"

        # Nombre único en disco
        ts = datetime.now().strftime("%Y%m%d%H%M%S%f")
        filename_on_disk = f"{reclamo_id}_{ts}_{safe_name}"
        path = os.path.join(upload_folder, filename_on_disk)

        # Guardar archivo en disco
        f.save(path)

        # Registrar en BD
        cur.execute(SQL__SAVE_ADJUNTOS_FOR_RECLAMO_INS_2, (
            reclamo_id,
            filename_on_disk,
            original_name,
            f.mimetype,
            size,
            user_id,
            _now_iso()
        ))
        saved_any = True

    # No hacemos commit aquí; lo hace el caller
    return None


from flask import url_for

def _notify_aprobador_imputacion(conn, aprobador_id, reclamo_codigo, imputado_username):
    jefe = _get_user_basic(conn, aprobador_id)
    if not jefe or ("email" not in jefe.keys()) or not jefe["email"]:
        return

    # --- Datos del reclamo (por codigo) ---
    cur = conn.cursor()
    cur.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_1, (reclamo_codigo,))
    r = cur.fetchone()

    # Imputados del caso (por si hay más de uno)
    imputados = imputado_username or ""
    if r:
        cur.execute(SQL__NOTIFY_APROBADOR_IMPUTACION_SEL_1, (r["id"],))
        row = cur.fetchone()
        if row and row["lista"]:
            imputados = row["lista"]

    # Link directo al tab "Por aprobar (Jefe)"
    try:
        link_aprobar = url_for("reclamos", _external=True) + "?tab=aprobar"
    except Exception:
        link_aprobar = "https://tu-sistema/reclamos?tab=aprobar"

    nombre = (
        jefe['nombre_completo']
        if 'nombre_completo' in jefe.keys() and jefe['nombre_completo']
        else jefe['username']
    )

    subject = f"[Oportunida de Mejora] Aprobación pendiente {reclamo_codigo}"

    text_body = f"""Hola {nombre},

Hay un reclamo {reclamo_codigo} con imputación pendiente para: {imputados}.
Por favor revísalo y aprueba/rechaza la imputación.

Ir al sistema: {link_aprobar}

Este es un mensaje automático.
"""

    def _row(lbl, val):
        val = (val or "").replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#fee2e2;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;
               font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">
            <tr>
              <td style="background:#b91c1c;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;
                            letter-spacing:.08em;opacity:.9;">
                  Oprotunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Aprobación pendiente {reclamo_codigo}
                </div>
                <div style="font-size:12px;opacity:.9;margin-top:6px;">
                  Hola {nombre}, tienes una imputación pendiente de revisión.
                </div>
              </td>
            </tr>

            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row('Código', reclamo_codigo)}
                  {_row('Fecha OM', r['fecha_reclamo']) if r else ''}
                  {_row('Tipo de OM', r['tipo_reclamo']) if r else ''}
                  {_row('Tipo de Trámite', r['tipo_tramite']) if r else ''}
                  {_row('Cliente', r['cliente_nombre']) if r else ''}
                  {_row('Proceso', r['proceso_text']) if r else ''}
                  {_row('Material', r['material_desc']) if (r and 'material_desc' in r.keys()) else ''}
                  {_row('Fecha de Pedido', r['fecha_pedido']) if r else ''}
                  {_row('Factura', r['factura']) if r else ''}
                  {_row('Guía Remisión', r['guia_remision']) if r else ''}
                  {_row('Imputados', imputados)}
                  {_row('Antecedente', r['antecedente']) if r else ''}
                  {_row('Observación', r['observacion']) if r else ''}
                </table>

                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link_aprobar}"
                     style="display:inline-block;background:#2563eb;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Revisar y aprobar imputación
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  También puedes ingresar al módulo de reclamos desde el sistema
                  y abrir la bandeja <strong>“Por aprobar (Jefe)”</strong>.
                </div>
              </td>
            </tr>

            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>

          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

    _send_mail_safe(jefe["email"], subject, text_body, html_body=html_body)

def _notify_creador_rechazo_asignacion(conn, creador_id, reclamo_codigo, imputado_username, motivo):
    c = _get_user_basic(conn, creador_id)
    if not c:
        return
    subject = f"[Oportunidad de Mejora] Imputación rechazada en {reclamo_codigo}"
    body = (
        f"Hola {c['username']},\n\n"
        f"El reclamo {reclamo_codigo} fue RECHAZADO para el usuario {imputado_username}.\n"
        f"Motivo del rechazo:\n{motivo or 'Sin motivo informado.'}\n\n"
        "Este es un mensaje automático."
    )
    _send_mail_safe(c["email"], subject, body)


def _notify_imputado_aprobado(conn, imputado_id, reclamo_codigo):
    u = _get_user_basic(conn, imputado_id)
    if not u or ("email" not in u.keys()) or not u["email"]:
        return

    # --- Datos del reclamo (por código) ---
    cur = conn.cursor()
    cur.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_1, (reclamo_codigo,))
    r = cur.fetchone()

    # Link directo al tab "Soy responsable"
    try:
        link_responder = url_for("reclamos", _external=True) + "?tab=imputado"
    except Exception:
        link_responder = "https://tu-sistema/reclamos?tab=imputado"

    nombre = (
        u['nombre_completo']
        if 'nombre_completo' in u.keys() and u['nombre_completo']
        else u['username']
    )

    subject = f"[Oportunidad de Mejora] Nueva OM asignada ({reclamo_codigo})"

    # Texto plano (fallback)
    text_body = f"""Hola {nombre},

Se te ha asignado la Oportunidad de Mejora {reclamo_codigo}.
Por favor ingresa al sistema, revisa el detalle y registra:

- Causa raíz
- Acción preventiva
- Acción correctiva

Ir al sistema: {link_responder}

Este es un mensaje automático.
"""

    # ---------- HTML similar al del aprobador ----------
    def _row(lbl, val):
        val = (val or "").replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#dbeafe;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;
               font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">
            <!-- Encabezado -->
            <tr>
              <td style="background:#2563eb;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;
                            letter-spacing:.08em;opacity:.9;">
                  Oportunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Nueva OM asignada {reclamo_codigo}
                </div>
                <div style="font-size:12px;opacity:.9;margin-top:6px;">
                  Hola {nombre}, se te ha asignado esta OM para análisis y respuesta técnica.
                </div>
              </td>
            </tr>

            <!-- Cuerpo -->
            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row('Código', reclamo_codigo)}
                  { _row('Fecha OM', r['fecha_reclamo']) if r else '' }
                  { _row('Tipo de OM', r['tipo_reclamo']) if r else '' }
                  { _row('Tipo de Trámite', r['tipo_tramite']) if r else '' }
                  { _row('Cliente', r['cliente_nombre']) if r else '' }
                  { _row('Proceso', r['proceso_text']) if r else '' }
                  { _row('Material', r['material_desc']) if (r and 'material_desc' in r.keys()) else '' }
                  { _row('Fecha de Pedido', r['fecha_pedido']) if r else '' }
                  { _row('Factura', r['factura']) if r else '' }
                  { _row('Guía Remisión', r['guia_remision']) if r else '' }
                  { _row('Antecedente', r['antecedente']) if r else '' }
                  { _row('Observación', r['observacion']) if r else '' }
                </table>

                <!-- Botón CTA -->
                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link_responder}"
                     style="display:inline-block;background:#2563eb;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Ingresar y responder medidas
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  Una vez completes la causa, acción preventiva y correctiva,
                  tu jefe revisará y aprobará la respuesta técnica.
                </div>
              </td>
            </tr>

            <!-- Pie -->
            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>

          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

    _send_mail_safe(u["email"], subject, text_body, html_body=html_body)

def _notify_jefe_respuesta_listo(conn, aprobador_id, reclamo_codigo, imputado_username):
    jefe = _get_user_basic(conn, aprobador_id)
    if not jefe or ("email" not in jefe.keys()) or not jefe["email"]:
        return

    # --- Datos del reclamo (por código) ---
    cur = conn.cursor()
    if _is_sqlserver_conn(conn):
        cur.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_1, (reclamo_codigo,))
    else:
        cur.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_1, (reclamo_codigo,))
    r = cur.fetchone()
 
    # Imputados del caso (por si hay más de uno)
    imputados = imputado_username or ""
    if r:
        cur.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_3, (r["id"],))
        row = cur.fetchone()
        if row and row["lista"]:
            imputados = row["lista"]

    # Link directo al tab "Por aprobar (Jefe)" para validar la respuesta técnica
    try:
        link_validar = url_for("reclamos", _external=True) + "?tab=aprobar"
    except Exception:
        link_validar = "https://tu-sistema/reclamos?tab=aprobar"

    nombre = (
        jefe['nombre_completo']
        if 'nombre_completo' in jefe.keys() and jefe['nombre_completo']
        else jefe['username']
    )

    subject = f"[Oportunidad de Mejora] Validar respuesta técnica {reclamo_codigo}"

    # Texto plano (fallback)
    text_body = f"""Hola {nombre},

El usuario {imputado_username} ha registrado la respuesta técnica para la
Oportunidad de Mejora {reclamo_codigo}.

Por favor revisa las medidas propuestas (causa, acción preventiva y correctiva)
y aprueba o rechaza la respuesta.

Ir al sistema: {link_validar}

Este es un mensaje automático.
"""

    # ---------- HTML mejorado ----------
    def _row(lbl, val):
        val = (val or "").replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#fee2e2;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;
               font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">
            <!-- Encabezado -->
            <tr>
              <td style="background:#1d4ed8;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;
                            letter-spacing:.08em;opacity:.9;">
                  Oportunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Validar respuesta técnica {reclamo_codigo}
                </div>
                <div style="font-size:12px;opacity:.9;margin-top:6px;">
                  Hola {nombre}, el usuario {imputado_username} ha registrado su respuesta
                  y está pendiente de tu aprobación.
                </div>
              </td>
            </tr>

            <!-- Cuerpo -->
            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row('Código', reclamo_codigo)}
                  { _row('Fecha OM', r['fecha_reclamo']) if r else '' }
                  { _row('Tipo de OM', r['tipo_reclamo']) if r else '' }
                  { _row('Tipo de Trámite', r['tipo_tramite']) if r else '' }
                  { _row('Cliente', r['cliente_nombre']) if r else '' }
                  { _row('Proceso', r['proceso_text']) if r else '' }
                  { _row('Material', r['material_desc']) if (r and 'material_desc' in r.keys()) else '' }
                  { _row('Fecha de Pedido', r['fecha_pedido']) if r else '' }
                  { _row('Factura', r['factura']) if r else '' }
                  { _row('Guía Remisión', r['guia_remision']) if r else '' }
                  {_row('Imputados', imputados)}
                  { _row('Antecedente', r['antecedente']) if r else '' }
                  { _row('Observación', r['observacion']) if r else '' }
                </table>

                <!-- Botón CTA -->
                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link_validar}"
                     style="display:inline-block;background:#2563eb;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Revisar y validar respuesta técnica
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  También puedes ingresar al módulo de reclamos desde el sistema
                  y abrir la bandeja <strong>“Por aprobar (Jefe)”</strong> para validar las respuestas.
                </div>
              </td>
            </tr>

            <!-- Pie -->
            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>

          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

    _send_mail_safe(jefe["email"], subject, text_body, html_body=html_body)

def _notify_imputado_respuesta_rechazada(conn, imputado_id, reclamo_codigo, motivo):
    """
    Notifica rechazo de respuesta técnica a:
    - imputado/responsable que debe corregir
    - sponsor principal y backup del proceso
    - usuarios de Servicio al Cliente

    Reutiliza:
    - _get_user_basic
    - _get_sponsor_emails_by_reclamo
    - _send_mail_safe
    """

    u = _get_user_basic(conn, imputado_id)
    if not u or ("email" not in u.keys()) or not u["email"]:
        imputado_email = None
        imputado_nombre = "Responsable técnico"
    else:
        imputado_email = (u["email"] or "").strip()
        imputado_nombre = (
            u["nombre_completo"]
            if "nombre_completo" in u.keys() and u["nombre_completo"]
            else u["username"]
        )

    cur = conn.cursor()

    # =========================================================
    # Datos del reclamo
    # =========================================================
    cur.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_1, (reclamo_codigo,))

    r = cur.fetchone()

    try:
        link_responder = url_for("reclamos", _external=True) + "?tab=imputado"
    except Exception:
        link_responder = "https://tu-sistema/reclamos?tab=imputado"

    motivo_txt = (motivo or "Sin detalle").strip()

    subject = f"[Oportunidad de Mejora] Ajuste requerido en respuesta {reclamo_codigo}"

    # =========================================================
    # Helper visual para correo HTML
    # =========================================================
    def _row(lbl, val):
        val = "" if val is None else str(val)
        val = val.replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#fef3c7;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    # =========================================================
    # Servicio al Cliente
    # No existe actualmente una función que devuelva la lista.
    # Se reutiliza la misma regla ya usada en _can_view_all_reclamos:
    # departamento Servicio al Cliente o puesto que contenga Servicio al Cliente.
    # =========================================================
    cur.execute(SQL__NOTIFY_COLABORADOR_APORTE_RECHAZADO_SEL_1)

    servicio_cliente_rows = cur.fetchall()

    # =========================================================
    # Destinatarios
    # =========================================================
    destinatarios = []

    # 1) Imputado / responsable que debe corregir
    if imputado_email:
        destinatarios.append({
            "email": imputado_email,
            "nombre": imputado_nombre,
            "rol_notificacion": "RESPONSABLE"
        })

    # 2) Sponsor principal + backup
    # Reutiliza función existente.
    try:
        for s in _get_sponsor_emails_by_reclamo(conn, reclamo_codigo):
            destinatarios.append({
                "email": s["email"],
                "nombre": s["nombre"] or s["username"] or "Usuario",
                "rol_notificacion": s["tipo_sponsor"] or "SPONSOR"
            })
    except Exception:
        current_app.logger.exception(
            "No se pudo obtener sponsor principal/backup para reclamo %s",
            reclamo_codigo
        )

    # 3) Servicio al Cliente
    for sc in servicio_cliente_rows:
        destinatarios.append({
            "email": sc["email"],
            "nombre": sc["nombre"] or sc["username"] or "Servicio al Cliente",
            "rol_notificacion": "SERVICIO AL CLIENTE"
        })

    enviados = set()

    for d in destinatarios:
        email = (d.get("email") or "").strip().lower()
        if not email or email in enviados:
            continue

        enviados.add(email)

        nombre_destinatario = d.get("nombre") or "Usuario"
        rol_notificacion = d.get("rol_notificacion") or ""

        text_body = f"""Hola {nombre_destinatario},

La respuesta técnica de la Oportunidad de Mejora {reclamo_codigo} fue RECHAZADA y requiere ajustes.

Responsable técnico:
{imputado_nombre}

Rol de notificación:
{rol_notificacion}

Motivo del rechazo:
{motivo_txt or 'Sin detalle'}

El responsable debe actualizar la causa, acción de control y acción correctiva en el sistema.

Ir al sistema:
{link_responder}

Este es un mensaje automático.
"""

        html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;
               font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">

            <tr>
              <td style="background:#b45309;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;
                            letter-spacing:.08em;opacity:.9;">
                  Oportunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Ajuste requerido en respuesta {reclamo_codigo}
                </div>
                <div style="font-size:12px;opacity:.9;margin-top:6px;">
                  Hola {nombre_destinatario}, se rechazó la respuesta técnica y requiere seguimiento.
                </div>
              </td>
            </tr>

            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row('Código', reclamo_codigo)}
                  {_row('Responsable técnico', imputado_nombre)}
                  {_row('Rol de notificación', rol_notificacion)}
                  { _row('Fecha OM', r['fecha_reclamo']) if r else '' }
                  { _row('Tipo de OM', r['tipo_reclamo']) if r else '' }
                  { _row('Tipo de Trámite', r['tipo_tramite']) if r else '' }
                  { _row('Cliente', r['cliente_nombre']) if r else '' }
                  { _row('Proceso', r['proceso_text']) if r else '' }
                  { _row('Material', r['material_desc']) if (r and 'material_desc' in r.keys()) else '' }
                  { _row('Fecha de Pedido', r['fecha_pedido']) if r else '' }
                  { _row('Factura', r['factura']) if r else '' }
                  { _row('Guía Remisión', r['guia_remision']) if r else '' }
                  { _row('Antecedente', r['antecedente']) if r else '' }
                  { _row('Observación', r['observacion']) if r else '' }
                  {_row('Motivo del rechazo', motivo_txt or 'Sin detalle')}
                </table>

                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link_responder}"
                     style="display:inline-block;background:#f97316;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Revisar respuesta técnica
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  El responsable debe ingresar al módulo de reclamos, pestaña
                  <strong>“Soy responsable”</strong>, seleccionar la OM y actualizar
                  la causa raíz, acción de control y acción correctiva.
                </div>
              </td>
            </tr>

            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>

          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

        _send_mail_safe(
            email,
            subject,
            text_body,
            html_body=html_body
        )

 
def _get_sponsor_emails_by_reclamo(conn, reclamo_codigo: str):
    cur = conn.cursor()
    cur.execute(SQL__GET_SPONSOR_EMAILS_BY_RECLAMO_SEL_1, (reclamo_codigo,))

    return cur.fetchall()


def _notify_creador_respuesta_aprobada(conn, creador_id, reclamo_codigo, imputado_username):
    c = _get_user_basic(conn, creador_id)
    if not c or ("email" not in c.keys()) or not c["email"]:
        return

    # --- Datos del reclamo (por código) ---
    cur = conn.cursor()
    cur.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_1, (reclamo_codigo,))
    r = cur.fetchone()

    # Link directo a "Mis reclamos"
    try:
        link_mis_reclamos = url_for("reclamos", _external=True) + "?tab=mios"
    except Exception:
        link_mis_reclamos = "https://tu-sistema/reclamos?tab=mios"

    nombre = (
        c['nombre_completo']
        if 'nombre_completo' in c.keys() and c['nombre_completo']
        else c['username']
    )

    responsable = imputado_username or "Responsable técnico"

    subject = f"[Oportunidad de Mejora] Respuesta final aprobada {reclamo_codigo}"

    # Texto plano (fallback)
    text_body = f"""Hola {nombre},

El responsable {responsable} registró y su jefe aprobó la respuesta técnica
de la Oportunidad de Mejora {reclamo_codigo}.

Ya puedes consultarla en el sistema.

Ir al sistema: {link_mis_reclamos}

Este es un mensaje automático.
"""

    # ---------- HTML mejorado ----------
    def _row(lbl, val):
        val = (val or "").replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#dcfce7;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;
               font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">
            <!-- Encabezado -->
            <tr>
              <td style="background:#15803d;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;
                            letter-spacing:.08em;opacity:.9;">
                  Oportunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Respuesta final aprobada {reclamo_codigo}
                </div>
                <div style="font-size:12px;opacity:.9;margin-top:6px;">
                  Hola {nombre}, la respuesta técnica de esta OM fue aprobada por el jefe del responsable.
                </div>
              </td>
            </tr>

            <!-- Cuerpo -->
            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row('Código', reclamo_codigo)}
                  {_row('Responsable', responsable)}
                  { _row('Fecha OM', r['fecha_reclamo']) if r else '' }
                  { _row('Tipo de OM', r['tipo_reclamo']) if r else '' }
                  { _row('Tipo de Trámite', r['tipo_tramite']) if r else '' }
                  { _row('Cliente', r['cliente_nombre']) if r else '' }
                  { _row('Proceso', r['proceso_text']) if r else '' }
                  { _row('Material', r['material_desc']) if (r and 'material_desc' in r.keys()) else '' }
                  { _row('Fecha de Pedido', r['fecha_pedido']) if r else '' }
                  { _row('Factura', r['factura']) if r else '' }
                  { _row('Guía Remisión', r['guia_remision']) if r else '' }
                  { _row('Antecedente', r['antecedente']) if r else '' }
                  { _row('Observación', r['observacion']) if r else '' }
                </table>

                <!-- Botón CTA -->
                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link_mis_reclamos}"
                     style="display:inline-block;background:#16a34a;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Ver respuesta técnica
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  También puedes ingresar al módulo de reclamos y abrir la pestaña
                  <strong>“Mis reclamos”</strong> para revisar el detalle completo de la OM.
                </div>
              </td>
            </tr>

            <!-- Pie -->
            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>

          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

    destinatarios = []

    # creador de la OM
    if c["email"]:
        destinatarios.append({
            "email": c["email"],
            "nombre": nombre
        })

    # principal + backup del proceso
    for s in _get_sponsor_emails_by_reclamo(conn, reclamo_codigo):
        destinatarios.append({
            "email": s["email"],
            "nombre": s["nombre"] or s["username"] or "Usuario"
        })

    enviados = set()

    for d in destinatarios:
        email = (d["email"] or "").strip().lower()
        if not email or email in enviados:
            continue

        enviados.add(email)

        # opcional: personalizar saludo por destinatario
        html_final = html_body.replace(f"Hola {nombre}", f"Hola {d['nombre']}")
        text_final = text_body.replace(f"Hola {nombre}", f"Hola {d['nombre']}")

        _send_mail_safe(email, subject, text_final, html_body=html_final)


def _notify_creador_rechazo_validacion(conn, reclamo_id: int, reclamo_codigo: str,
                                        creador_nombre: str, motivo: str):
    """
    Notifica a sponsors (principal+backup), Servicio al Cliente y miembros de
    equipo cuando el creador de la OM rechaza la respuesta técnica.
    """
    cur = conn.cursor()

    # Datos completos de la OM
    cur.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_1, (reclamo_codigo,))
    r = cur.fetchone()

    try:
        link = url_for("reclamos", _external=True) + "?tab=imputado"
    except Exception:
        link = "http://bitacoraquimpac.com.ec:5000/reclamos?tab=imputado"

    subject = f"[Oportunidad de Mejora] Respuesta rechazada por el creador — {reclamo_codigo}"

    motivo_visible = motivo.strip() if motivo else "Sin motivo especificado."

    def _row(lbl, val):
        val = (val or "").replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#fee2e2;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    def _html(dest_nombre):
        return f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;
               font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">
            <!-- Encabezado -->
            <tr>
              <td style="background:#dc2626;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;
                            letter-spacing:.08em;opacity:.9;">
                  Oportunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Respuesta rechazada — {reclamo_codigo}
                </div>
                <div style="font-size:12px;opacity:.9;margin-top:6px;">
                  Hola {dest_nombre}, el creador <strong>{creador_nombre}</strong>
                  rechazó la respuesta técnica de esta OM.
                </div>
              </td>
            </tr>

            <!-- Cuerpo -->
            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row('Código', reclamo_codigo)}
                  { _row('Fecha OM', r['fecha_reclamo']) if r else '' }
                  { _row('Tipo de OM', r['tipo_reclamo']) if r else '' }
                  { _row('Tipo de Trámite', r['tipo_tramite']) if r else '' }
                  { _row('Cliente', r['cliente_nombre']) if r else '' }
                  { _row('Proceso', r['proceso_text']) if r else '' }
                  { _row('Material', r['material_desc']) if (r and 'material_desc' in r.keys()) else '' }
                  { _row('Fecha de Pedido', r['fecha_pedido']) if r else '' }
                  { _row('Factura', r['factura']) if r else '' }
                  { _row('Guía Remisión', r['guia_remision']) if r else '' }
                  { _row('Observación', r['observacion']) if r else '' }
                  {_row('Rechazado por', creador_nombre)}
                  {_row('Motivo del rechazo', motivo_visible)}
                </table>

                <!-- Botón CTA -->
                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link}"
                     style="display:inline-block;background:#dc2626;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Ingresar y registrar nueva respuesta
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  La OM ha vuelto a estado <strong>Abierto</strong>.
                  Se requiere registrar una nueva respuesta técnica.
                </div>
              </td>
            </tr>

            <!-- Pie -->
            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>

          </table>
        </td>
      </tr>
    </table>
  </body>
</html>"""

    enviados = set()

    def _enviar(email, nombre):
        email = (email or "").strip().lower()
        if not email or email in enviados:
            return
        enviados.add(email)
        txt = (
            f"Hola {nombre},\n\n"
            f"El creador {creador_nombre} rechazó la respuesta de la OM {reclamo_codigo}.\n\n"
            f"Motivo: {motivo_visible}\n\n"
            f"La OM volvió a estado Abierto. Por favor ingresa al sistema y "
            f"registra una nueva respuesta técnica.\n\n"
            f"Ir al sistema: {link}\n\nEste es un mensaje automático."
        )
        _send_mail_safe(email, subject, txt, html_body=_html(nombre))

    # 1. Sponsors PRINCIPAL + BACKUP de cada proceso de la OM
    cur.execute(SQL_VALIDAR_CREADOR_SEL_BASE, (reclamo_id,))
    row_om = cur.fetchone()
    if row_om and row_om["proceso_id"]:
        cur.execute(SQL_VALIDAR_CREADOR_SEL_SPONSORS, (row_om["proceso_id"],))
        for s in cur.fetchall():
            _enviar(s["sponsor_email"], s["sponsor_nombre"])

    # 2. Imputados directos del reclamo
    cur.execute(SQL_VALIDAR_CREADOR_SEL_IMPUTADOS, (reclamo_id,))
    for row in cur.fetchall():
        _enviar(row["imputado_email"], row["imputado_nombre"])

    # 3. Servicio al Cliente
    cur.execute(SQL_VALIDAR_CREADOR_SEL_SAC)
    for row in cur.fetchall():
        _enviar(row["email"], row["nombre"])

    # 4. Miembros de equipo
    cur.execute(SQL_VALIDAR_CREADOR_SEL_EQUIPO, (reclamo_id,))
    for row in cur.fetchall():
        _enviar(row["miembro_email"], row["miembro_nombre"])


import re

def _acciones_v2_start() -> str:
    """
    Fecha desde la cual aplicas el cambio SOLO hacia adelante.
    Puedes setearlo en config: app.config['RECL_ACCIONES_V2_START'] = '2025-12-20'
    """
    return current_app.config.get("RECL_ACCIONES_V2_START", "2025-12-20")

_LINE_RE = re.compile(
    r"^\s*-\s*(.*?)\s*(?:\((\d{4}-\d{2}-\d{2})\))?\s*$"
)

def _parse_acciones_txt(txt: str | None):
    """
    Convierte texto tipo:
      - Hacer algo (2025-12-31)
      - Otra cosa
    en lista [{descripcion, fecha_compromiso}]
    """
    items = []
    for line in (txt or "").splitlines():
        line = (line or "").strip()
        if not line:
            continue
        m = _LINE_RE.match(line)
        if m:
            desc = (m.group(1) or "").strip()
            fecha = (m.group(2) or "").strip()
        else:
            desc = line.lstrip("-").strip()
            fecha = ""
        if desc or fecha:
            items.append({"id": None, "descripcion": desc, "fecha_compromiso": fecha})
    return items

def _equipo_es_v2(conn: sqlite3.Connection, equipo_id: int) -> bool:
    """
    True si:
    - ya tiene acciones en tabla nueva, o
    - su fecha_asignacion es >= a la fecha de activación.
    """
    cur = conn.cursor()

    # 1) Si ya existen acciones, ya es v2
    cur.execute("SELECT TOP 1 FROM reclamo_equipo_acciones WHERE equipo_id = ? ", (equipo_id,))
    if cur.fetchone():
        return True

    # 2) Si es posterior a la fecha de activación, es v2
    cur.execute("SELECT TOP 1 fecha_asignacion FROM reclamo_equipo WHERE id = ? ", (equipo_id,))
    row = cur.fetchone()
    fa = (row["fecha_asignacion"] if row and ("fecha_asignacion" in row.keys()) else None) if row else None
    if not fa:
        return False

    # fa = "YYYY-MM-DD HH:MM:SS" -> comparar solo la fecha
    fa_date = fa[:10]
    return fa_date >= _acciones_v2_start()

# =========================================================
#   REGISTER ROUTES
# =========================================================

def register_reclamos_routes(app):

    # ----------------------------------------------
    # API clientes para autocomplete
    # ----------------------------------------------
    @app.route('/api/reclamos/clientes', methods=['GET'], endpoint='api_reclamos_clientes')
    @require_login
    def api_reclamos_clientes():
        q = (request.args.get('q') or '').strip()
        conn = get_db()
        cur = conn.cursor()

        sql = """
            SELECT TOP 25
                id,
                nombre,
                identificacion,
                email,
                telefono,
                direccion
            FROM terceros
            WHERE tipo = 'C'
        """
        params = []

        if q:
            sql += " AND (nombre LIKE ? OR identificacion LIKE ?)"
            like = f"%{q}%"
            params.extend([like, like])

        sql += " ORDER BY nombre"

        cur.execute(sql, params)

        items = []
        for r in cur.fetchall():
            items.append({
                "id": r["id"],
                "nombre": r["nombre"],
                "identificacion": r["identificacion"],
                "email": r["email"],
                "telefono": r["telefono"],
                "direccion": r["direccion"],
            })

        return jsonify(ok=True, items=items)

    @app.route('/reclamos/api/subtipos', methods=['GET'], endpoint='reclamos_api_subtipos')
    @require_login
    def reclamos_api_subtipos():
        tipo_id = request.args.get('tipo_id', type=int)
        if not tipo_id:
            return jsonify([])

        conn = get_db()
        cur = conn.cursor()

        # No crear tablas ni grupos desde backend.
        # Buscar directamente el grupo de subtipos existente.
        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_1)
        row = cur.fetchone()

        if not row:
            return jsonify([])

        gid = row["id"]

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_2, (gid, tipo_id))

        data = [
            {
                "id": r["id"],
                "nombre": r["nombre"],
                "valor": r["valor"],
            }
            for r in cur.fetchall()
        ]

        return jsonify(data)


    @app.post("/reclamos/equipo/<int:equipo_id>/acciones")
    @require_login
    def api_save_equipo_acciones(equipo_id):
        if not _can_edit_equipo(equipo_id):
            abort(403)

        # 1) Leer JSON
        data = request.get_json(silent=True) or {}

        # 2) Adapter: soportar payload “simple” (strings + fechas sueltas) además del payload v2 (listas)
        #    - causa + fecha_causa
        #    - preventiva/control + fecha_preventiva/fecha_control
        #    - correctiva + fecha_correctiva
        #
        #    Si ya viene como lista (v2), se respeta.
        def _as_list_value(val):
            """Convierte val en lista de dicts si viene como dict; si viene str no aplica aquí."""
            if val is None:
                return None
            if isinstance(val, list):
                return val
            if isinstance(val, dict):
                return [val]
            return val  # puede ser str u otro

        data["causas"] = _as_list_value(data.get("causas"))
        data["control"] = _as_list_value(data.get("control"))
        data["correctiva"] = _as_list_value(data.get("correctiva"))

        # Si NO vienen como listas válidas, pero llegan como simple (string), armamos v2
        # CAUSA
        if not isinstance(data.get("causas"), list):
            causa_txt = (data.get("causa") or "").strip()
            causa_fecha = (data.get("fecha_causa") or data.get("causa_fecha") or "").strip()
            # Si "causas" viene como string accidental, lo tomamos como descripción
            if isinstance(data.get("causas"), str) and not causa_txt:
                causa_txt = (data.get("causas") or "").strip()

            if causa_txt or causa_fecha:
                data["causas"] = [{
                    "descripcion": causa_txt,
                    "fecha_compromiso": causa_fecha,
                }]
            else:
                data["causas"] = []

        # CONTROL / PREVENTIVA
        if not isinstance(data.get("control"), list):
            ctrl_txt = (data.get("control") or "").strip() if isinstance(data.get("control"), str) else ""
            # “preventiva” como alias de control
            if not ctrl_txt:
                ctrl_txt = (data.get("preventiva") or "").strip()

            ctrl_fecha = (data.get("fecha_control") or data.get("fecha_preventiva") or "").strip()

            if ctrl_txt or ctrl_fecha:
                data["control"] = [{
                    "descripcion": ctrl_txt,
                    "fecha_compromiso": ctrl_fecha,
                }]
            else:
                data["control"] = []

        # CORRECTIVA
        if not isinstance(data.get("correctiva"), list):
            corr_txt = (data.get("correctiva") or "").strip() if isinstance(data.get("correctiva"), str) else ""
            # por si viene en clave "correctiva" (string) y/o "correctiva_texto"
            if not corr_txt:
                corr_txt = (data.get("correctiva_texto") or "").strip()

            corr_fecha = (data.get("fecha_correctiva") or "").strip()

            if corr_txt or corr_fecha:
                data["correctiva"] = [{
                    "descripcion": corr_txt,
                    "fecha_compromiso": corr_fecha,
                }]
            else:
                data["correctiva"] = []

        # 3) Ya en formato v2 (listas)
        causas     = data.get("causas") or []
        control    = data.get("control") or []
        correctiva = data.get("correctiva") or []

        def _norm(items, label):
            out = []
            for it in items:
                it = it or {}
                if not isinstance(it, dict):
                    # si por algún motivo llega mal, lo ignoramos
                    continue

                desc = (it.get("descripcion") or "").strip()
                fecha = (it.get("fecha_compromiso") or "").strip()

                # ignorar fila vacía
                if not desc and not fecha:
                    continue

                if not desc:
                    return None, f"Falta descripción en {label}."
                if not fecha:
                    return None, f"Falta fecha compromiso en {label}."
                if not _is_date_yyyy_mm_dd(fecha):
                    return None, f"Fecha compromiso inválida en {label}. Use YYYY-MM-DD."

                out.append((desc, fecha))
            return out, None

        causas_n, err = _norm(causas, "Causa")
        if err:
            return jsonify(ok=False, error=err), 400

        control_n, err = _norm(control, "Acción de Control")
        if err:
            return jsonify(ok=False, error=err), 400

        correctiva_n, err = _norm(correctiva, "Acción Correctiva")
        if err:
            return jsonify(ok=False, error=err), 400
 
        user_id = session.get("user_id") or session.get("id")
        now = _now_str()

        conn = get_db()

        # ✅ CAMBIO CLAVE: asegurar esquema SIEMPRE antes de usar la tabla v2
        #ensure_reclamos_catalogos(conn)

        cur = conn.cursor()

        # Reemplazo total (simple, robusto y sin “histórico”)
        cur.execute("DELETE FROM reclamo_equipo_acciones WHERE equipo_id = ?", (equipo_id,))

        def _insert_many(tipo, items):
            for desc, fecha in items:
                cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_INS_3, (equipo_id, tipo, desc, fecha, now, user_id))

        _insert_many("CAUSA", causas_n)
        _insert_many("CONTROL", control_n)
        _insert_many("CORRECTIVA", correctiva_n)

        # (Opcional recomendado) espejo a reclamo_equipo.respuesta_* para compatibilidad
        def _join(items):
            # "- texto (YYYY-MM-DD)"
            return "\n".join([f"- {d} ({f})" for d, f in items])

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_4, (_join(causas_n), _join(control_n), _join(correctiva_n), now, equipo_id))

        conn.commit()
        return jsonify(ok=True)


    import re
    from flask import jsonify, abort, request, session
    # asumiendo que ya tienes: get_db, require_login, _can_edit_equipo, ensure_reclamos_catalogos

    def _row_get(row, key, idx, default=""):
        """Soporta sqlite Row (dict-like) o tupla."""
        if not row:
            return default
        try:
            if hasattr(row, "keys"):
                return row[key] if key in row.keys() else default
        except Exception:
            pass
        try:
            return row[idx]
        except Exception:
            return default

    def _parse_acciones_txt_v2(txt):
        """
        Parsea formato:
        - descripcion (YYYY-MM-DD)
        y retorna:
        [{"id": None, "descripcion": "...", "fecha_compromiso": "YYYY-MM-DD"}, ...]
        """
        out = []
        for line in (txt or "").splitlines():
            line = (line or "").strip()
            if not line:
                continue
            if line.startswith("-"):
                line = line[1:].strip()

            # Captura fecha al final si existe
            m = re.match(r"^(.*?)(?:\s*\((\d{4}-\d{2}-\d{2})\))?\s*$", line)
            if not m:
                continue

            desc = (m.group(1) or "").strip()
            fecha = (m.group(2) or "").strip()

            # ignora fila vacía
            if not desc and not fecha:
                continue

            out.append({
                "id": None,
                "descripcion": desc,
                "fecha_compromiso": fecha or ""
            })
        return out


    @app.get("/reclamos/equipo/<int:equipo_id>/acciones")
    @require_login
    def api_get_equipo_acciones(equipo_id):

        if not _can_edit_equipo(equipo_id):
            abort(403)

        conn = get_db()

        # ✅ Asegura que exista la tabla v2 antes de consultar
        #ensure_reclamos_catalogos(conn)

        cur = conn.cursor()

        # ✅ V2: intenta leer SIEMPRE desde la tabla nueva.
        # Esto evita el problema de que _equipo_es_v2() falle y no cargue fechas.
        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_5, (equipo_id,))
        rows = cur.fetchall() or []

        if rows:
            causas, control, correctiva = [], [], []

            for r in rows:
                rid   = _row_get(r, "id", 0, None)
                tipo  = (_row_get(r, "tipo", 1, "") or "").strip()
                desc  = _row_get(r, "descripcion", 2, "") or ""
                fecha = _row_get(r, "fecha_compromiso", 3, "") or ""

                item = {"id": rid, "descripcion": desc, "fecha_compromiso": fecha}

                if tipo == "CAUSA":
                    causas.append(item)
                elif tipo == "CONTROL":
                    control.append(item)
                elif tipo == "CORRECTIVA":
                    correctiva.append(item)

            return jsonify(ok=True, causas=causas, control=control, correctiva=correctiva)

        # 🔁 Fallback histórico: lee campos viejos y parsea también la fecha si está en "(YYYY-MM-DD)"
        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_6, (equipo_id,))
        row = cur.fetchone()

        causas_txt     = _row_get(row, "respuesta_causa", 0, "")
        control_txt    = _row_get(row, "respuesta_preventiva", 1, "")
        correctiva_txt = _row_get(row, "respuesta_correctiva", 2, "")

        causas     = _parse_acciones_txt_v2(causas_txt)
        control    = _parse_acciones_txt_v2(control_txt)
        correctiva = _parse_acciones_txt_v2(correctiva_txt)

        return jsonify(ok=True, causas=causas, control=control, correctiva=correctiva)


    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/json")
    @require_login
    def equipo_respuestas_json(reclamo_id):
        uid = _current_user_id()

        db = get_db()
        #ensure_reclamos_schema(db)
        #ensure_reclamo_respuestas_equipo_schema(db)
        

        # ✅ Solo responsable aprobado, miembros, creador o admin/coordinador
        if not (_puede_ver_equipo(reclamo_id, uid) or _is_admin_like()):
            return jsonify({"error": "No autorizado"}), 403

        imputacion_id = request.args.get("imputacion_id", type=int)

        # Si el front no manda imputacion_id, tomamos la última imputación del reclamo
        if not imputacion_id:
            row_imp = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_7, (reclamo_id,)).fetchone()
            imputacion_id = row_imp["id"] if row_imp else None

        # -------------------------
        # Caso 1: SIN imputación
        # -------------------------
        if not imputacion_id:
            rows = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_8, (reclamo_id,)).fetchall()

        # -------------------------
        # Caso 2: CON imputación
        # -------------------------
        else:
            rows = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_9, (reclamo_id,)).fetchall()
        # -------------------------
        # Payload (para ambos casos)
        # -------------------------
        payload = []
        for row in rows:
            r = dict(row)  # convertir Row -> dict (así puedes usar .get)

            # Siempre tendremos equipo_id e id por el SELECT (alias)
            equipo_id = r.get("equipo_id")
            if equipo_id is None:
                # fallback por si en algún momento cambias el SELECT
                equipo_id = r.get("id")

            payload.append({
                # id del registro en reclamo_equipo_respuestas (sirve para Quitar)
                "id": r["equipo_id"],

                # explícitos para que el front no se confunda
                "equipo_id": r["equipo_id"],
                "reclamo_id": r["reclamo_id"],
                "imputacion_id": r.get("imputacion_id"),

                # ESTE es el id del usuario real (miembro)
                "usuario_id": r["usuario_id"],
                "miembro_id": r["usuario_id"],

                "nombre": r["nombre"],
                "username": r["username"],
                "rol": r["rol"],
                "departamento": r["departamento"],
                "puede_responder": bool(r["puede_responder"]),
                "tiene_respuesta": bool(r["tiene_respuesta"]),
                "fecha_respuesta": r["fecha_respuesta"],

                # id del registro en reclamo_respuestas_equipo (si existe)
                "respuesta_id": r.get("respuesta_id"),
            })


        return jsonify(payload)


    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/add", methods=["POST"])
    @require_login
    def equipo_respuestas_add(reclamo_id):
        payload = request.get_json(silent=True) or {}
        usuario_id = payload.get("usuario_id")
        imputacion_id = payload.get("imputacion_id")

        uid = _current_user_id()
        if not uid:
            return jsonify({"error": "Sesión inválida. Vuelve a iniciar sesión."}), 401

        db = get_db()

        # SOLO responsable aprobado o admin/coordinador
        if not (_puede_gestionar_equipo(reclamo_id, uid) or _is_admin_like()):
            return jsonify({"error": "No tienes permiso para modificar el equipo de respuestas."}), 403

        if not usuario_id:
            return jsonify({"error": "Selecciona un usuario para agregar al equipo."}), 400

        row_user = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_10, (usuario_id,)).fetchone()
        if not row_user:
            return jsonify({"error": "El usuario seleccionado no existe."}), 400

        # Si el front no manda imputacion_id, lo deducimos desde el sponsor actual
        if not imputacion_id:
            row_imp = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_11, (reclamo_id, uid)).fetchone()

            if not row_imp:
                return jsonify({"error": "No se pudo determinar imputación_id para este reclamo."}), 400

            imputacion_id = row_imp["id"]

        # Validar que la imputación pertenezca al reclamo
        row_ok = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_12, (imputacion_id, reclamo_id)).fetchone()

        if not row_ok:
            return jsonify({"error": "La imputación no pertenece a este reclamo."}), 400

        # Evita duplicado en el mismo reclamo
        dup = db.execute(SQL__ES_MIEMBRO_EQUIPO_RECLAMO_SEL_1, (reclamo_id, usuario_id)).fetchone()

        if dup:
            return jsonify({"error": "El usuario ya está en el equipo de respuestas."}), 400

        now = _now_iso()

        db.execute(SQL_REGISTER_RECLAMOS_ROUTES_INS_13, (reclamo_id, imputacion_id, usuario_id, uid, now))

        db.commit()

        # Notificar por correo al miembro agregado
        try:
            row_r = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_14, (reclamo_id,)).fetchone()

            reclamo_codigo = row_r["codigo"] if row_r and row_r["codigo"] else f"RECL#{reclamo_id}"

            row_actor = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_15, (uid,)).fetchone()

            if row_actor:
                resp_username = (
                    row_actor["nombre_completo"]
                    if row_actor.get("nombre_completo")
                    else row_actor["username"]
                )
            else:
                resp_username = f"UID {uid}"

            _notify_colaborador_asignado(
                db,
                int(usuario_id),
                reclamo_codigo,
                resp_username
            )
        except Exception:
            current_app.logger.exception(
                "No se pudo enviar correo al miembro agregado. reclamo_id=%s usuario_id=%s",
                reclamo_id, usuario_id
            )

        return jsonify({"ok": True})


    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/guardar", methods=["POST"])
    @require_login
    def equipo_respuestas_guardar(reclamo_id):
        current_app.logger.info("[MAIL_EQUIPO] entra a equipo_respuestas_guardar reclamo_id=%s", reclamo_id)

        payload = request.get_json(silent=True) or {}

        uid = _current_user_id()
        if not uid:
            return jsonify({"error": "Sesión inválida. Vuelve a iniciar sesión."}), 401

        imputacion_id = payload.get("imputacion_id")

        # =========================================================
        # Compatibilidad: payload viejo
        # =========================================================
        causa = (payload.get("causa") or "").strip()
        preventiva = (payload.get("preventiva") or "").strip()
        correctiva = (payload.get("correctiva") or "").strip()

        fecha_causa = (payload.get("fecha_causa") or "").strip()
        fecha_preventiva = (payload.get("fecha_preventiva") or "").strip()
        fecha_correctiva = (payload.get("fecha_correctiva") or "").strip()

        metodo_analisis = (payload.get("metodo_analisis") or "").strip()

        why1 = (payload.get("why1") or "").strip()
        why2 = (payload.get("why2") or "").strip()
        why3 = (payload.get("why3") or "").strip()
        why4 = (payload.get("why4") or "").strip()
        why5 = (payload.get("why5") or "").strip()

        fish_metodo = (payload.get("fish_metodo") or "").strip()
        fish_maquinas = (payload.get("fish_maquinas") or "").strip()
        fish_materiales = (payload.get("fish_materiales") or "").strip()
        fish_personas = (payload.get("fish_personas") or "").strip()
        fish_entorno = (payload.get("fish_entorno") or "").strip()
        fish_medicion = (payload.get("fish_medicion") or "").strip()

        # =========================================================
        # Nuevo payload múltiple
        # =========================================================
        causas_payload = payload.get("causas")
        control_payload = payload.get("control")
        correctiva_payload = payload.get("correctiva")

        if causas_payload is None and control_payload is None and not isinstance(correctiva_payload, list):
            causas_payload = []
            control_payload = []
            correctiva_payload = []

            if causa or fecha_causa:
                causas_payload.append({
                    "descripcion": causa,
                    "fecha_compromiso": fecha_causa
                })

            if preventiva or fecha_preventiva:
                control_payload.append({
                    "descripcion": preventiva,
                    "fecha_compromiso": fecha_preventiva
                })

            if correctiva or fecha_correctiva:
                correctiva_payload.append({
                    "descripcion": correctiva,
                    "fecha_compromiso": fecha_correctiva
                })

        causas_n, err = _normalizar_acciones_payload(causas_payload, "Causa")
        if err:
            return jsonify({"error": err}), 400

        controles_n, err = _normalizar_acciones_payload(control_payload, "Acción de Control")
        if err:
            return jsonify({"error": err}), 400

        correctivas_n, err = _normalizar_acciones_payload(correctiva_payload, "Acción Correctiva")
        if err:
            return jsonify({"error": err}), 400

        if not causas_n or not controles_n or not correctivas_n:
            return jsonify({"error": "Debes ingresar al menos una Causa, una Acción de Control y una Acción Correctiva."}), 400

        db = get_db()

        # Si no viene imputacion_id, deducirla desde la asignación del equipo del usuario actual
        if not imputacion_id:
            row_eq = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_16, (reclamo_id, uid)).fetchone()

            if not row_eq or not row_eq["imputacion_id"]:
                return jsonify({"error": "No se pudo determinar imputación_id para guardar la respuesta."}), 400

            imputacion_id = row_eq["imputacion_id"]

        # Validar permisos
        row_perm = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_17, (reclamo_id, imputacion_id, uid)).fetchone()

        if not row_perm and not _is_admin_like():
            return jsonify({"error": "No tienes permiso para responder en este reclamo."}), 403

        # ¿Existe respuesta previa?
        row_exist = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_18, (reclamo_id, imputacion_id, uid)).fetchone()

        es_nueva_respuesta = row_exist is None
        now = _now_iso()

        # Resumen legacy
        causa_txt = "\n".join([f"- {desc} ({fecha})" for desc, fecha in causas_n])
        preventiva_txt = "\n".join([f"- {desc} ({fecha})" for desc, fecha in controles_n])
        correctiva_txt = "\n".join([f"- {desc} ({fecha})" for desc, fecha in correctivas_n])

        primera_fecha_causa = causas_n[0][1] if causas_n else ""
        primera_fecha_control = controles_n[0][1] if controles_n else ""
        primera_fecha_correctiva = correctivas_n[0][1] if correctivas_n else ""

        if row_exist:
            respuesta_equipo_id = row_exist["id"]

            db.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_19, (
                metodo_analisis,
                causa_txt,
                preventiva_txt,
                correctiva_txt,
                primera_fecha_causa,
                primera_fecha_control,
                primera_fecha_correctiva,
                fish_metodo,
                fish_maquinas,
                fish_materiales,
                fish_personas,
                fish_entorno,
                fish_medicion,
                why1,
                why2,
                why3,
                why4,
                why5,
                now,
                uid,
                respuesta_equipo_id
            ))
        else:
            row_new = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_INS_20, (
                reclamo_id,
                imputacion_id,
                uid,
                metodo_analisis,
                causa_txt,
                preventiva_txt,
                correctiva_txt,
                primera_fecha_causa,
                primera_fecha_control,
                primera_fecha_correctiva,
                fish_metodo,
                fish_maquinas,
                fish_materiales,
                fish_personas,
                fish_entorno,
                fish_medicion,
                why1,
                why2,
                why3,
                why4,
                why5,
                uid,
                now
            )).fetchone()

            if not row_new:
                db.rollback()
                return jsonify({"error": "No se pudo guardar la respuesta del equipo."}), 500

            respuesta_equipo_id = row_new[0] if not isinstance(row_new, dict) else row_new["id"]

        # Guardar detalle múltiple real
        _save_respuesta_equipo_acciones(
            db,
            respuesta_id=respuesta_equipo_id,
            reclamo_id=reclamo_id,
            imputacion_id=imputacion_id,
            miembro_id=uid,
            causas=causas_n,
            control=controles_n,
            correctiva=correctivas_n,
            user_id=uid
        )

        db.commit()

        # Notificar al sponsor SOLO la primera vez
        if es_nueva_respuesta:
            try:
                row_r = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_14, (reclamo_id,)).fetchone()

                reclamo_codigo = row_r["codigo"] if row_r and row_r["codigo"] else f"RECL#{reclamo_id}"

                current_app.logger.info(
                    "[MAIL_EQUIPO] Se envía mail sponsor reclamo_id=%s imputacion_id=%s miembro_id=%s",
                    reclamo_id, imputacion_id, uid
                )

                _notify_sponsor_respuesta_equipo(
                    db,
                    int(imputacion_id),
                    int(uid),
                    reclamo_codigo
                )
            except Exception:
                current_app.logger.exception(
                    "[MAIL_EQUIPO] Error notificando sponsor por respuesta de equipo. reclamo_id=%s imputacion_id=%s miembro_id=%s",
                    reclamo_id, imputacion_id, uid
                )

        return jsonify({
            "ok": True,
            "respuesta_equipo_id": respuesta_equipo_id
        })


        
    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/aporte", methods=["GET"])
    @require_login
    def equipo_respuestas_ver_aporte(reclamo_id):
        db = get_db()
        uid = _current_user_id()

        imputacion_id = request.args.get("imputacion_id", type=int)
        miembro_id = request.args.get("miembro_id", type=int)

        if not uid:
            return jsonify(ok=False, error="Sesión inválida. Vuelve a iniciar sesión."), 401

        # Si el front manda mal miembro_id (id del registro en tabla puente), lo corregimos
        if miembro_id and not db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_21, (miembro_id,)).fetchone():
            row_fix = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_22, (miembro_id, reclamo_id)).fetchone()

            if row_fix:
                miembro_id = row_fix["usuario_id"]

        if not imputacion_id or not miembro_id:
            return jsonify(ok=False, error="Falta imputacion_id o miembro_id"), 400

        # =========================================================
        # Datos base de la OM / imputación
        # =========================================================
        row_base = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_23, (reclamo_id, imputacion_id)).fetchone()

        if not row_base:
            return jsonify(ok=False, error="No se encontró la imputación de la OM."), 404

        can_view_all = _can_view_all_reclamos(db, uid)

        # Sponsor principal: está en reclamo_imputados
        es_sponsor_principal = int(row_base["imputado_id"] or 0) == int(uid or 0)

        # Sponsor backup o principal configurado por proceso
        es_sponsor_del_proceso = _usuario_es_sponsor_del_proceso(
            db,
            row_base["proceso_id"],
            uid
        )

        # El mismo miembro puede ver su propio aporte
        is_same_member = int(uid or 0) == int(miembro_id or 0)

        row_team_member = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_24, (reclamo_id, imputacion_id, uid)).fetchone()

        can_view = bool(
            can_view_all
            or es_sponsor_principal
            or es_sponsor_del_proceso
            or (is_same_member and row_team_member)
            or _is_admin_like()
        )

        if not can_view:
            return jsonify(ok=False, error="No autorizado"), 403

        # =========================================================
        # Buscar respuesta del miembro
        # =========================================================
        row = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_25, (reclamo_id, imputacion_id, miembro_id)).fetchone()

        item = dict(row) if row else {}

        if item.get("id"):
            acciones = _get_respuesta_equipo_acciones_full(db, int(item["id"]))
            item["causas"] = acciones.get("causas", [])
            item["control"] = acciones.get("control", [])
            item["correctiva_items"] = acciones.get("correctiva", [])
        else:
            item["causas"] = []
            item["control"] = []
            item["correctiva_items"] = []

        return jsonify(ok=True, item=item)

    @app.route("/reclamos/imputacion/<int:imputacion_id>/respuestas_equipo/json")
    @require_login
    def reclamo_respuestas_equipo_json(imputacion_id):
        db = get_db()

        rows = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_26, (imputacion_id,)).fetchall()

        items = []
        for r in rows:
            items.append({
                "id": r["id"],
                "miembro_nombre": r["miembro_nombre"] or r["miembro_username"],
                "causa": r["causa"],
                "preventiva": r["preventiva"],
                "correctiva": r["correctiva"],
                "created_at": r["created_at"],
            })

        return jsonify({"ok": True, "items": items})


    @app.route("/reclamos/imputacion/<int:imputacion_id>/responder_equipo", methods=["POST"])
    @require_login
    def reclamo_responder_equipo(imputacion_id):
        db = get_db()

        uid = (
            session.get("user_id")
            or session.get("id")
            or session.get("usuario_id")
            or (session.get("user") or {}).get("id")
        )

        current_app.logger.debug(
            "[equipo_responder] imputacion_id=%s uid=%s", imputacion_id, uid
        )

        # 1) Obtener reclamo_id de la imputación
        imp = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_27, (imputacion_id,)).fetchone()

        current_app.logger.debug(
            "[equipo_responder] imputacion -> %r", dict(imp) if imp else None
        )

        if not imp:
            return jsonify({"ok": False, "error": "Imputación no encontrada."}), 404

        reclamo_id = imp["reclamo_id"]
        current_app.logger.debug(
            "[equipo_responder] reclamo_id derivado=%s", reclamo_id
        )

        # 2) Validar que el usuario sea miembro del equipo para ese reclamo
        es_miembro = _es_miembro_equipo_reclamo(reclamo_id, uid)
        current_app.logger.debug(
            "[equipo_responder] _es_miembro_equipo_reclamo => %s", es_miembro
        )

        if not es_miembro:
            return jsonify({"ok": False, "error": "No tienes permiso para responder como equipo en esta OM."}), 403

        data = request.get_json(silent=True) or {}
        current_app.logger.debug("[equipo_responder] payload recibido=%r", data)

        metodo_analisis = (data.get("metodo_analisis") or "").strip()

        # =========================================================
        # NUEVO FORMATO: listas
        # =========================================================
        causas_items = data.get("causas") or []
        control_items = data.get("control") or []
        correctiva_items = data.get("correctiva") or []

        # =========================================================
        # COMPATIBILIDAD: si aún llega formato viejo, convertirlo
        # =========================================================
        if not causas_items and (data.get("causa") or data.get("fecha_causa")):
            causas_items = [{
                "descripcion": (data.get("causa") or "").strip(),
                "fecha_compromiso": (data.get("fecha_causa") or "").strip(),
            }]

        if not control_items and (data.get("preventiva") or data.get("fecha_preventiva")):
            control_items = [{
                "descripcion": (data.get("preventiva") or "").strip(),
                "fecha_compromiso": (data.get("fecha_preventiva") or "").strip(),
            }]

        if not correctiva_items and (data.get("correctiva") or data.get("fecha_correctiva")):
            correctiva_items = [{
                "descripcion": (data.get("correctiva") or "").strip(),
                "fecha_compromiso": (data.get("fecha_correctiva") or "").strip(),
            }]

        def _norm_items(items, nombre_bloque):
            out = []
            for i, it in enumerate(items, start=1):
                if not isinstance(it, dict):
                    continue

                descripcion = (it.get("descripcion") or "").strip()
                fecha = (it.get("fecha_compromiso") or "").strip()

                if not descripcion and not fecha:
                    continue

                if not descripcion:
                    return None, f"Falta descripción en {nombre_bloque} #{i}."
                if not fecha:
                    return None, f"Falta fecha en {nombre_bloque} #{i}."

                out.append({
                    "descripcion": descripcion,
                    "fecha_compromiso": fecha
                })

            return out, None

        causas_items, err = _norm_items(causas_items, "causa")
        if err:
            return jsonify({"ok": False, "error": err}), 400

        control_items, err = _norm_items(control_items, "acción de control")
        if err:
            return jsonify({"ok": False, "error": err}), 400

        correctiva_items, err = _norm_items(correctiva_items, "acción correctiva")
        if err:
            return jsonify({"ok": False, "error": err}), 400

        if not causas_items or not control_items or not correctiva_items:
            return jsonify({
                "ok": False,
                "error": "Debe ingresar al menos una causa, una acción de control y una acción correctiva."
            }), 400

        # =========================================================
        # Campos de análisis
        # =========================================================
        fish_metodo = (data.get("fish_metodo") or "").strip()
        fish_maquinas = (data.get("fish_maquinas") or "").strip()
        fish_materiales = (data.get("fish_materiales") or "").strip()
        fish_personas = (data.get("fish_personas") or "").strip()
        fish_entorno = (data.get("fish_entorno") or "").strip()
        fish_medicion = (data.get("fish_medicion") or "").strip()

        why1 = (data.get("why1") or "").strip()
        why2 = (data.get("why2") or "").strip()
        why3 = (data.get("why3") or "").strip()
        why4 = (data.get("why4") or "").strip()
        why5 = (data.get("why5") or "").strip()

        # =========================================================
        # Resumen legacy para compatibilidad
        # =========================================================
        def _join_legacy(items):
            return "\n".join(
                f"- {x['descripcion']} ({x['fecha_compromiso']})"
                for x in items
            )

        causa_txt = _join_legacy(causas_items)
        preventiva_txt = _join_legacy(control_items)
        correctiva_txt = _join_legacy(correctiva_items)

        fecha_causa = causas_items[0]["fecha_compromiso"] if causas_items else ""
        fecha_preventiva = control_items[0]["fecha_compromiso"] if control_items else ""
        fecha_correctiva = correctiva_items[0]["fecha_compromiso"] if correctiva_items else ""

        # 3) Verificar si ya existe respuesta activa del miembro
        row = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_18, (reclamo_id, imputacion_id, uid)).fetchone()

        if row:
            respuesta_id = row["id"]

            db.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_28, (
                metodo_analisis,
                causa_txt, preventiva_txt, correctiva_txt,
                fecha_causa, fecha_preventiva, fecha_correctiva,
                fish_metodo, fish_maquinas, fish_materiales,
                fish_personas, fish_entorno, fish_medicion,
                why1, why2, why3, why4, why5,
                respuesta_id
            ))
        else:
            cur = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_INS_29, (
                reclamo_id,
                imputacion_id,
                uid,
                metodo_analisis,
                causa_txt, preventiva_txt, correctiva_txt,
                fecha_causa, fecha_preventiva, fecha_correctiva,
                fish_metodo, fish_maquinas, fish_materiales,
                fish_personas, fish_entorno, fish_medicion,
                why1, why2, why3, why4, why5,
                1,
                uid
            ))

            # SQL Server: obtener id insertado
            try:
                respuesta_id = cur.lastrowid
            except Exception:
                respuesta_id = None

            if not respuesta_id:
                row_new = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_30, (reclamo_id, imputacion_id, uid)).fetchone()
                respuesta_id = row_new["id"] if row_new else None

            if not respuesta_id:
                db.rollback()
                return jsonify({
                    "ok": False,
                    "error": "No se pudo obtener el identificador de la respuesta registrada."
                }), 500

        # =========================================================
        # Guardar acciones detalle
        # =========================================================
        _save_respuesta_equipo_acciones(
            db,
            respuesta_id=respuesta_id,
            reclamo_id=reclamo_id,
            imputacion_id=imputacion_id,
            miembro_id=uid,
            causas=causas_items,
            control=control_items,
            correctiva=correctiva_items,
            user_id=uid
        )

        db.commit()

        try:
            row_r = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_14, (reclamo_id,)).fetchone()

            reclamo_codigo = row_r["codigo"] if row_r and row_r["codigo"] else f"RECL#{reclamo_id}"

            _notify_sponsor_respuesta_equipo(
                db,
                int(imputacion_id),
                int(uid),
                reclamo_codigo
            )
        except Exception:
            current_app.logger.exception(
                "[MAIL_EQUIPO] Error notificando sponsor por respuesta de equipo. reclamo_id=%s imputacion_id=%s miembro_id=%s",
                reclamo_id, imputacion_id, uid
            )

        return jsonify({"ok": True})


    # ----------------------------------------------
    # EQUIPO: RESPONSABLE APRUEBA / RECHAZA APORTE
    # ----------------------------------------------
    @app.route(
        '/reclamos/equipo/<int:eq_id>/aprobar',
        methods=['POST'],
        endpoint='reclamos_equipo_aprobar'
    )
    @require_login
    @require_permission('reclamos', 'aprobar')
    def reclamos_equipo_aprobar(eq_id):
        uid = _current_user_id()
        data = request.get_json(silent=True) or {}

        accion = (data.get("accion") or "").strip().lower()  # aprobar / rechazar
        motivo = (data.get("motivo") or "").strip()

        if not uid:
            return jsonify(ok=False, msg="Sesión inválida. Vuelve a iniciar sesión."), 401

        conn = get_db()
        cur = conn.cursor()

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_31, (eq_id,))

        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="Registro de equipo no encontrado"), 404

        # =========================================================
        # AUTORIZACIÓN
        # =========================================================
        # Puede aprobar/rechazar:
        # - responsable original
        # - admin/coordinador
        # - sponsor principal o backup configurado en RECL_PROCESO_SPONSOR
        # - cualquier usuario que pueda gestionar equipo en esa OM
        # =========================================================
        es_responsable_original = int(row["responsable_id"] or 0) == int(uid or 0)

        es_sponsor_del_proceso = _usuario_es_sponsor_del_proceso(
            conn,
            row["proceso_id"],
            uid
        )

        puede_gestionar_equipo = _puede_gestionar_equipo(
            int(row["reclamo_id"]),
            uid
        )

        if not (
            es_responsable_original
            or es_sponsor_del_proceso
            or puede_gestionar_equipo
            or _is_admin_like()
        ):
            conn.close()
            return jsonify(ok=False, msg="No autorizado"), 403

        if row["estado"] not in ("respondido", "pendiente"):
            conn.close()
            return jsonify(ok=False, msg="No hay aporte pendiente de revisión"), 400

        now = _now_iso()

        if accion == "aprobar":
            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_32, (now, eq_id))

        elif accion == "rechazar":
            if not motivo:
                conn.close()
                return jsonify(ok=False, msg="Motivo obligatorio al rechazar"), 400

            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_33, (now, motivo, eq_id))

            _notify_colaborador_aporte_rechazado(
                conn,
                row["colaborador_id"],
                row["codigo"],
                motivo
            )

        else:
            conn.close()
            return jsonify(ok=False, msg="Acción inválida"), 400

        conn.commit()
        conn.close()

        return jsonify(ok=True)

    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/<int:er_id>/eliminar", methods=["POST"])
    @require_login
    def reclamo_equipo_respuestas_eliminar(reclamo_id, er_id):
        uid = _current_user_id()
        if not (_puede_gestionar_equipo(reclamo_id, uid) or _is_admin_like()):
            return jsonify({"ok": False, "error": "No tienes permiso para modificar el equipo de respuestas."}), 403

        db = get_db()
        #ensure_reclamos_schema(db)

        cur = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_34, (er_id, reclamo_id))
        db.commit()

        if cur.rowcount == 0:
            return jsonify({"ok": False, "error": "No se encontró el registro a eliminar (id incorrecto)."}), 404

        return jsonify({"ok": True})

 
    @app.route('/reclamos/api/usuario_por_cedula/<cedula>', methods=['GET'])
    @require_login
    def reclamos_usuario_por_cedula(cedula):
        print("[DEBUG reclamos] cedula", cedula)

        conn = get_db()
        cur = conn.cursor()

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_35, (cedula,))

        row = cur.fetchone()

        if not row:
            return jsonify(ok=False, msg="No existe usuario con esa cédula"), 404

        return jsonify(ok=True, usuario={
            "id": row["id"],
            "nombre": row["nombre_completo"],
            "username": row["username"],
            "departamento": row["departamento_nombre"],
            "jefe": row["jefe_nombre"]
        })

    # ----------------------------------------------
    # EQUIPO DEL RESPONSABLE: ASIGNAR COLABORADORES
    # ----------------------------------------------
    @app.route(
        '/reclamos/<int:reclamo_id>/equipo/asignar',
        methods=['POST'],
        endpoint='reclamos_equipo_asignar'
    )
    @require_login
    @require_permission('reclamos', 'editar')
    def reclamos_equipo_asignar(reclamo_id):
        uid = _current_user_id()  # responsable actual (imputado principal)
        data = request.get_json(silent=True) or {}
        ids = data.get("colaboradores") or []  # lista de IDs de usuarios

        if not isinstance(ids, list) or not ids:
            return jsonify(ok=False, msg="Debe indicar al menos un colaborador"), 400

        conn = get_db()
        #ensure_reclamos_schema(conn)
        
        cur = conn.cursor()

        # Validar que el usuario actual sea el imputado principal de la OM
        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_36, (reclamo_id,))
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="Oportunidad de Mejora no encontrada"), 404

        imputado_principal = row["imputado_id"]
        codigo = row["codigo"]

        # Solo el imputado/responsable principal o admin puede asignar equipo
        if (imputado_principal != uid) and (not _is_admin_like()):
            conn.close()
            return jsonify(ok=False, msg="No autorizado para asignar equipo"), 403

        # Evitar duplicados
        responsable_id = imputado_principal or uid
        now = _now_iso()

        # Para obtener username del responsable (para el correo)
        resp_basic = _get_user_basic(conn, responsable_id)
        resp_username = resp_basic["username"] if resp_basic else f"UID {responsable_id}"

        creados = 0
        for raw_id in ids:
            try:
                colaborador_id = int(raw_id)
            except Exception:
                continue

            if colaborador_id == responsable_id:
                continue  # no tiene sentido asignarse a sí mismo como colaborador

            # ¿ya existe?
            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_37, (reclamo_id, colaborador_id))
            if cur.fetchone():
                continue

            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_INS_38, (
                reclamo_id,
                responsable_id,
                colaborador_id,
                'pendiente',
                now
            ))
            creados += 1

            # Notificación al colaborador
            _notify_colaborador_asignado(conn, colaborador_id, codigo, resp_username)

        conn.commit()
        conn.close()

        if not creados:
            return jsonify(ok=False, msg="No se creó ningún registro nuevo (todos ya estaban asignados)"), 400

        return jsonify(ok=True, msg=f"{creados} colaborador(es) asignado(s).")

    # ----------------------------------------------
    # LISTA / BANDEJA
    # ----------------------------------------------
  
    @app.get("/reclamos/<int:reclamo_id>/detalle-json")
    @require_login
    def reclamo_detalle_json(reclamo_id):
        db = get_db()

        reclamo = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_39, (reclamo_id,)).fetchone()

        if not reclamo:
            return jsonify(ok=False), 404

        imputados = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_40, (reclamo_id,)).fetchall()

        return jsonify(
            ok=True,
            reclamo=dict(reclamo),
            imputados=[dict(i) for i in imputados]
        )


    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/usuarios", methods=["GET"])
    @require_login
    def equipo_respuestas_usuarios(reclamo_id):
        uid = _current_user_id()
        if not (_puede_gestionar_equipo(reclamo_id, uid) or _is_admin_like()):
            return jsonify({"ok": False, "error": "No autorizado"}), 403

        db = get_db()
        rows = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_41, (reclamo_id,)).fetchall()

        return jsonify({"ok": True, "items": [dict(r) for r in rows]})

    # ----------------------------------------------
    # EQUIPO: COLABORADOR REGISTRA SU APORTE
    # ----------------------------------------------
    @app.route(
        '/reclamos/equipo/<int:eq_id>/responder',
        methods=['POST'],
        endpoint='reclamos_equipo_responder'
    )
    @require_login
    @require_permission('reclamos', 'editar')
    def reclamos_equipo_responder(eq_id):
        uid = _current_user_id()
        data = request.get_json(silent=True) or {}

        causa = (data.get("causa") or "").strip()
        preventiva = (data.get("preventiva") or "").strip()
        correctiva = (data.get("correctiva") or "").strip()

        if not causa or not preventiva or not correctiva:
            return jsonify(ok=False, msg="Todos los campos del aporte son obligatorios"), 400

        conn = get_db()
        #ensure_reclamos_schema(conn)
        
        cur = conn.cursor()

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_42, (eq_id,))
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="Registro de equipo no encontrado"), 404

        if (row["colaborador_id"] != uid) and (not _is_admin_like()):
            conn.close()
            return jsonify(ok=False, msg="No autorizado"), 403

        # Solo permitimos responder si está pendiente o ya respondió (para actualizar)
        if row["estado"] not in ("pendiente", "respondido"):
            conn.close()
            return jsonify(ok=False, msg="No se puede modificar este aporte en el estado actual"), 400

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_43, (
            causa,
            preventiva,
            correctiva,
            _now_iso(),
            eq_id
        ))

        # Notificar al responsable que ya tiene un aporte listo
        _notify_responsable_aporte_listo(
            conn,
            row["responsable_id"],
            row["codigo"],
            row["colaborador_username"] or f"UID {row['colaborador_id']}"
        )

        conn.commit()
        conn.close()
        return jsonify(ok=True)



    # ----------------------------------------------
    # API: respuestas técnicas por reclamo (todas las imputaciones)
    # ----------------------------------------------
    @app.route('/reclamos/api/<int:reclamo_id>/respuestas', methods=['GET'], endpoint='reclamos_api_respuestas')
    @require_login
    def reclamos_api_respuestas(reclamo_id):
        conn = get_db()
        #ensure_reclamos_schema(conn)
        cur = conn.cursor()

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_44, (reclamo_id,))


        items = [dict(r) for r in cur.fetchall()]
        conn.close()
        return jsonify({"items": items})

    @app.route('/reclamos/api/<int:reclamo_id>/respuestas-detalle', methods=['GET'])
    @require_login
    def reclamos_api_respuestas_detalle(reclamo_id):
        db = get_db()
        cur = db.cursor()
        items = []

        # =====================================================
        # RESPUESTAS OFICIALES: SPONSOR / IMPUTADO
        # =====================================================
        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_45, (reclamo_id,))

        imputados = cur.fetchall()

        for r in imputados:
            it = dict(r)

            acciones = _get_imputado_acciones_full(db, int(it["imputacion_id"]))

            it["causas"] = acciones.get("causas", [])
            it["control"] = acciones.get("control", [])
            it["correctiva_items"] = acciones.get("correctiva", [])

            # Fallback legacy: si no hay registros en reclamo_imputado_acciones,
            # usar los campos resumen de reclamo_imputados.
            if not it["causas"] and (it.get("respuesta_causa") or it.get("fecha_causa")):
                it["causas"] = [{
                    "id": None,
                    "tipo": "CAUSA",
                    "descripcion": it.get("respuesta_causa") or "",
                    "fecha_compromiso": it.get("fecha_causa") or "",
                    "orden": 1,
                    "requiere_evidencia": 0,
                    "cumplido": 0,
                    "fecha_cumplimiento": "",
                    "observacion_cumplimiento": "",
                    "evidencias": []
                }]

            if not it["control"] and (it.get("respuesta_preventiva") or it.get("fecha_preventiva")):
                it["control"] = [{
                    "id": None,
                    "tipo": "CONTROL",
                    "descripcion": it.get("respuesta_preventiva") or "",
                    "fecha_compromiso": it.get("fecha_preventiva") or "",
                    "orden": 1,
                    "requiere_evidencia": 0,
                    "cumplido": 0,
                    "fecha_cumplimiento": "",
                    "observacion_cumplimiento": "",
                    "evidencias": []
                }]

            if not it["correctiva_items"] and (it.get("respuesta_correctiva") or it.get("fecha_correctiva")):
                it["correctiva_items"] = [{
                    "id": None,
                    "tipo": "CORRECTIVA",
                    "descripcion": it.get("respuesta_correctiva") or "",
                    "fecha_compromiso": it.get("fecha_correctiva") or "",
                    "orden": 1,
                    "requiere_evidencia": 1,
                    "cumplido": 0,
                    "fecha_cumplimiento": "",
                    "observacion_cumplimiento": "",
                    "evidencias": []
                }]

            # IMPORTANTE:
            # No hacer continue aunque no tenga causa/control/correctiva.
            # Si SORTEGA está como sponsor/imputado, debe aparecer como tarjeta.
            items.append(it)

        # =====================================================
        # RESPUESTAS DEL EQUIPO
        # =====================================================
        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_46, (reclamo_id,))

        equipo_rows = cur.fetchall()

        for r in equipo_rows:
            it = dict(r)

            acciones = _get_respuesta_equipo_acciones_full(
                db,
                int(it["respuesta_equipo_id"])
            )

            it["causas"] = acciones.get("causas", [])
            it["control"] = acciones.get("control", [])
            it["correctiva_items"] = acciones.get("correctiva", [])

            # Fallback legacy equipo
            if not it["causas"] and (it.get("causa") or it.get("fecha_causa")):
                it["causas"] = [{
                    "id": None,
                    "tipo": "CAUSA",
                    "descripcion": it.get("causa") or "",
                    "fecha_compromiso": it.get("fecha_causa") or "",
                    "orden": 1,
                    "requiere_evidencia": 0,
                    "cumplido": 0,
                    "fecha_cumplimiento": "",
                    "observacion_cumplimiento": "",
                    "evidencias": []
                }]

            if not it["control"] and (it.get("preventiva") or it.get("fecha_preventiva")):
                it["control"] = [{
                    "id": None,
                    "tipo": "CONTROL",
                    "descripcion": it.get("preventiva") or "",
                    "fecha_compromiso": it.get("fecha_preventiva") or "",
                    "orden": 1,
                    "requiere_evidencia": 0,
                    "cumplido": 0,
                    "fecha_cumplimiento": "",
                    "observacion_cumplimiento": "",
                    "evidencias": []
                }]

            if not it["correctiva_items"] and (it.get("correctiva") or it.get("fecha_correctiva")):
                it["correctiva_items"] = [{
                    "id": None,
                    "tipo": "CORRECTIVA",
                    "descripcion": it.get("correctiva") or "",
                    "fecha_compromiso": it.get("fecha_correctiva") or "",
                    "orden": 1,
                    "requiere_evidencia": 1,
                    "cumplido": 0,
                    "fecha_cumplimiento": "",
                    "observacion_cumplimiento": "",
                    "evidencias": []
                }]

            items.append(it)

        return jsonify({
            "ok": True,
            "items": items,
            "total": len(items)
        })
    @app.route('/reclamos/equipo-acciones/<int:accion_id>/cumplir', methods=['POST'])
    @require_login
    def reclamo_equipo_accion_cumplir(accion_id):

        db = get_db()
        cur = db.cursor()

        data = request.json or {}
        cumplido = 1 if data.get("cumplido") else 0
        fecha = data.get("fecha_cumplimiento")

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_47, (
            cumplido,
            fecha,
            session.get("usuario_id"),
            accion_id
        ))

        db.commit()

        return jsonify({
            "ok": True,
            "msg": "Cumplimiento actualizado"
        })
   

    @app.route('/reclamos/equipo-acciones/evidencias/<int:evidencia_id>/download', methods=['GET'])
    @require_login
    def reclamo_equipo_accion_evidencia_download(evidencia_id):
        db = get_db()

        current_app.logger.info(
            "DESCARGA EVIDENCIA EQUIPO -> evidencia_id=%s",
            evidencia_id
        )

        row = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_48, (evidencia_id,)).fetchone()

        current_app.logger.info(
            "ROW EVIDENCIA EQUIPO -> %s",
            dict(row) if row else None
        )

        if not row or int(row["activo"] or 0) != 1:
            current_app.logger.warning(
                "EVIDENCIA NO ENCONTRADA O INACTIVA -> evidencia_id=%s",
                evidencia_id
            )
            abort(404)

        folder = os.path.join(current_app.config["UPLOAD_FOLDER"], "om_evidencias")
        path = os.path.join(folder, row["filename"])

        current_app.logger.info(
            "PATH EVIDENCIA EQUIPO -> %s",
            path
        )

        if not os.path.isfile(path):
            current_app.logger.warning(
                "ARCHIVO FISICO NO EXISTE -> %s",
                path
            )
            abort(404)

        return send_file(
            path,
            as_attachment=True,
            download_name=row["original_name"] or row["filename"],
            mimetype=row["content_type"] or "application/octet-stream"
        )

  
    @app.route("/reclamos/equipo-acciones/evidencias/<int:evidencia_id>/download", methods=["GET"])
    @require_login
    def descargar_evidencia_equipo_accion(evidencia_id):
        db = get_db()
        uid = _current_user_id()

        if not uid:
            return redirect(url_for("login"))

        row = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_49, (evidencia_id,)).fetchone()

        if not row:
            abort(404)

        reclamo_id = row["reclamo_id"]
        imputacion_id = row["imputacion_id"]
        miembro_id = row["miembro_id"]

        can_view_all = _can_view_all_reclamos(db, uid)

        row_owner = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_50, (imputacion_id, reclamo_id, uid)).fetchone()

        row_team_member = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_24, (reclamo_id, imputacion_id, uid)).fetchone()

        is_same_member = (uid == miembro_id)

        can_view = bool(can_view_all or row_owner or (is_same_member and row_team_member))
        if not can_view:
            abort(403)

        filename = row["filename"] or ""
        original_name = row["original_name"] or filename
        content_type = row["content_type"] or "application/octet-stream"

        ruta_archivo = os.path.join(
            current_app.config["UPLOAD_FOLDER"],
            "reclamos_equipo_evidencias",
            filename
        )

        if not os.path.exists(ruta_archivo):
            abort(404)

        return send_file(
            ruta_archivo,
            as_attachment=True,
            download_name=original_name,
            mimetype=content_type
        )    
        
    @app.route('/reclamos/equipo-acciones/<int:accion_id>/evidencia', methods=['POST'])
    @require_login
    def reclamo_equipo_accion_subir_evidencia(accion_id):

        db = get_db()
        cur = db.cursor()

        file = request.files.get("file")
        if not file:
            return jsonify({"ok": False, "error": "Archivo requerido"}), 400

        filename = secure_filename(file.filename)

        folder = os.path.join(current_app.config["UPLOAD_FOLDER"], "om_evidencias")
        os.makedirs(folder, exist_ok=True)

        path = os.path.join(folder, filename)
        file.save(path)

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_INS_51, (
            accion_id,
            filename,
            file.filename,
            file.content_type,
            os.path.getsize(path),
            session.get("usuario_id")
        ))

        db.commit()
 
        return jsonify({
            "ok": True,
            "msg": "Evidencia subida"
        })

        
    @app.route('/reclamos/equipo-acciones/<int:respuesta_id>/seguimiento', methods=['GET'])
    @require_login
    def reclamo_equipo_acciones_seguimiento(respuesta_id):

        db = get_db()
        cur = db.cursor()

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_52, (respuesta_id,))

        rows = [dict(r) for r in cur.fetchall()]

        return jsonify({
            "ok": True,
            "items": rows
        })
    
    
    # ----------------------------------------------
    # LISTA / BANDEJA
    # ----------------------------------------------
    @app.route('/reclamos', methods=['GET'], endpoint='reclamos')
    @require_login
    @require_permission('reclamos', 'ver')
    def reclamos_lista():
        uid = _current_user_id()

        conn = get_db()
        # Ya no crear/alterar esquema desde backend en SQL Server
        # _run_reclamos_bootstrap_if_needed(conn)

        productos = fetch_productos(conn)
        cur = conn.cursor()
        is_sqlserver = _is_sqlserver_conn(conn)

        puede_crear_materiales = False

        if _is_admin_like():
            puede_crear_materiales = True
        elif uid:
            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_53, (uid,))
            row_p = cur.fetchone()
            if row_p and row_p["nombre"]:
                nombre_puesto = row_p["nombre"].strip().upper()
                if nombre_puesto == "CORRDINADOR(A) SERVICIO AL CLIENTE":
                    puede_crear_materiales = True

        q_estado = (request.args.get('estado') or '').strip().lower()

        # =========================
        # Filtros (GET)
        # =========================
        f_codigo   = (request.args.get('codigo') or '').strip()
        f_cliente  = (request.args.get('cliente') or '').strip()
        f_proceso  = (request.args.get('proceso') or '').strip()
        f_motivo   = (request.args.get('motivo') or '').strip()
        f_tramite  = (request.args.get('tramite') or '').strip()
        f_desde    = (request.args.get('desde') or '').strip()
        f_hasta    = (request.args.get('hasta') or '').strip()
        f_estado_global = (request.args.get('estado_global') or '').strip()

        filtros = {
            "estado": q_estado,
            "codigo": f_codigo,
            "cliente": f_cliente,
            "proceso": f_proceso,
            "motivo": f_motivo,
            "tramite": f_tramite,
            "desde": f_desde,
            "hasta": f_hasta,
            "estado_global": f_estado_global,
        }

        def _build_where_filtros(alias_r: str = "r"):
            w = ""
            p = []

            if f_codigo:
                w += f" AND {alias_r}.codigo LIKE ?"
                p.append(f"%{f_codigo}%")

            if f_cliente:
                w += f" AND COALESCE({alias_r}.cliente_nombre,'') LIKE ?"
                p.append(f"%{f_cliente}%")

            if f_proceso:
                w += f" AND COALESCE({alias_r}.proceso_text,'') LIKE ?"
                p.append(f"%{f_proceso}%")

            if f_motivo:
                w += f" AND COALESCE({alias_r}.tipo_reclamo,'') = ?"
                p.append(f_motivo)

            if f_tramite:
                w += f" AND COALESCE({alias_r}.tipo_tramite,'') = ?"
                p.append(f_tramite)

            if f_desde:
                if is_sqlserver:
                    w += f" AND TRY_CONVERT(date, {alias_r}.fecha_reclamo) >= CAST(? AS date)"
                else:
                    w += f" AND date(substr({alias_r}.fecha_reclamo,1,10)) >= date(?)"
                p.append(f_desde)

            if f_hasta:
                if is_sqlserver:
                    w += f" AND TRY_CONVERT(date, {alias_r}.fecha_reclamo) <= CAST(? AS date)"
                else:
                    w += f" AND date(substr({alias_r}.fecha_reclamo,1,10)) <= date(?)"
                p.append(f_hasta)

            if f_estado_global:
                w += f" AND COALESCE({alias_r}.estado_global,'') = ?"
                p.append(f_estado_global)

            return w, p

        # -----------------------------------------
        # helpers locales plazo / alertas
        # -----------------------------------------
        from datetime import datetime, timedelta, date

        def _get_param_int_by_id(conn, pv_id: int, default: int) -> int:
            cur2 = conn.cursor()
            if is_sqlserver:
                cur2.execute(SQL__GET_PARAM_INT_BY_ID_SEL_SS, (pv_id,))
            else:
                cur2.execute(SQL__GET_PARAM_INT_BY_ID_SEL_SL, (pv_id,))
            row = cur2.fetchone()
            if not row:
                return default
            try:
                return int(str(row["valor"]).strip())
            except Exception:
                return default

        def _parse_fecha(fecha_val):
            if not fecha_val:
                return None

            if isinstance(fecha_val, datetime):
                return fecha_val

            if isinstance(fecha_val, date):
                return datetime.combine(fecha_val, datetime.min.time())

            s = str(fecha_val).strip()
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M"):
                try:
                    return datetime.strptime(s[:19], fmt)
                except Exception:
                    pass
            return None

        def _enrich_deadline(rows, dias_plazo: int):
            hoy = datetime.now()
            out = []

            for r in rows or []:
                d = dict(r)
                f = _parse_fecha(d.get("fecha_reclamo"))

                if not f:
                    d.update({
                        "dias_plazo": dias_plazo,
                        "fecha_limite": "",
                        "dias_transcurridos": 0,
                        "dias_restantes": dias_plazo,
                        "pct_avance": 0,
                        "deadline_estado": "ok",
                        "alerta_gg": 0,
                        "pct_avance_10": 0,
                        "pct_hito_5": 50,
                        "pct_hito_10": 100,
                        "is_done": 0,
                        "dias_para_d5": 5,
                        "dias_para_d10": 10,
                    })
                    out.append(d)
                    continue

                limite = f + timedelta(days=dias_plazo)
                trans = max(0, (hoy.date() - f.date()).days)

                ventana_alertas = 10
                pct10 = int(min(100, max(0, round((trans / ventana_alertas) * 100)))) if ventana_alertas > 0 else 100

                rest = (limite.date() - hoy.date()).days
                pct = int(min(100, max(0, round((trans / dias_plazo) * 100)))) if dias_plazo > 0 else 100

                estado_txt = (d.get("estado_imputacion") or d.get("estado") or "").lower()
                is_done = ("cerrad" in estado_txt) or ("aprobada" in estado_txt)

                dias_para_d5 = max(0, 5 - trans)
                dias_para_d10 = max(0, 10 - trans)

                if is_done:
                    deadline_estado = "done"
                    pct = 100
                    rest = None
                    trans = None
                elif rest < 0:
                    deadline_estado = "danger"
                    pct = 100
                elif trans >= 5:
                    deadline_estado = "warn"
                else:
                    deadline_estado = "ok"

                pend_resp = "pendiente respuesta del imputado" in (d.get("estado_imputacion") or "").lower()
                alerta_gg = 1 if (pend_resp and trans is not None and trans >= 5) else 0

                d.update({
                    "dias_plazo": dias_plazo,
                    "fecha_limite": limite.strftime("%Y-%m-%d"),
                    "dias_transcurridos": trans,
                    "dias_restantes": rest,
                    "pct_avance": pct,
                    "deadline_estado": deadline_estado,
                    "alerta_gg": alerta_gg,
                    "pct_avance_10": pct10,
                    "pct_hito_5": 50,
                    "pct_hito_10": 100,
                    "is_done": 1 if is_done else 0,
                    "dias_para_d5": dias_para_d5,
                    "dias_para_d10": dias_para_d10,
                })
                out.append(d)

            return out

        # 1) Creados por mí
        created_list = []
        if uid:
            params = []
            can_view_all = _can_view_all_reclamos(conn, uid)

            if can_view_all:
                where = "1=1"
            else:
                where = "r.creado_por = ?"
                params = [uid]

            if q_estado:
                where += " AND (COALESCE(ri.estado_asignacion,'') = ? OR COALESCE(ri.estado_respuesta,'') = ?)"
                params += [q_estado, q_estado]

            f_sql, f_params = _build_where_filtros("r")
            where += f_sql
            params += f_params

            sql_created = f"""
                SELECT
                    r.id,
                    r.codigo,
                    r.fecha_reclamo,
                    r.fecha_creacion,
                    r.cliente_nombre,
                    r.observacion,
                    r.material_desc,
                    r.procede,
                    r.estado_global AS estado,
                    r.proceso_text AS proceso_nombre,
                    c.nombre AS ciudad,
                    r.antecedente AS submotivo,
                    tr.valor AS motivo,
                    tt.valor AS tramite,

                    (
                        SELECT STUFF((
                            SELECT DISTINCT ', ' + COALESCE(ui2.username, '')
                            FROM reclamo_imputados ri2
                            LEFT JOIN usuarios ui2 ON ui2.id = ri2.imputado_id
                            WHERE ri2.reclamo_id = r.id
                            FOR XML PATH(''), TYPE
                        ).value('.', 'NVARCHAR(MAX)'), 1, 2, '')
                    ) AS imputados_resumen,

                    COALESCE(ri.respuesta_causa, '')      AS causa,
                    COALESCE(ri.respuesta_preventiva, '') AS preventiva,
                    COALESCE(ri.respuesta_correctiva, '') AS correctiva,
                    COALESCE(ri.fecha_causa,'')           AS fecha_causa,
                    COALESCE(ri.fecha_preventiva,'')      AS fecha_preventiva,
                    COALESCE(ri.fecha_correctiva,'')      AS fecha_correctiva,

                    COALESCE(ri.metodo_analisis,'') AS metodo_analisis,
                    COALESCE(ri.why1,'')            AS why1,
                    COALESCE(ri.why2,'')            AS why2,
                    COALESCE(ri.why3,'')            AS why3,
                    COALESCE(ri.why4,'')            AS why4,
                    COALESCE(ri.why5,'')            AS why5,
                    COALESCE(ri.fish_metodo,'')     AS fish_metodo,
                    COALESCE(ri.fish_maquinas,'')   AS fish_maquinas,
                    COALESCE(ri.fish_materiales,'') AS fish_materiales,
                    COALESCE(ri.fish_personas,'')   AS fish_personas,
                    COALESCE(ri.fish_entorno,'')    AS fish_entorno,
                    COALESCE(ri.fish_medicion,'')   AS fish_medicion,
                    r.tipo_reclamo,
                    r.antecedente,

                    (
                        SELECT STUFF((
                            SELECT DISTINCT ', ' + COALESCE(u2.nombre_completo, u2.username)
                            FROM reclamo_equipo_respuestas eq2
                            JOIN usuarios u2 ON u2.id = eq2.usuario_id
                            WHERE eq2.reclamo_id = r.id
                            AND eq2.imputacion_id = ri.id
                            AND eq2.activo = 1
                            FOR XML PATH(''), TYPE
                        ).value('.', 'NVARCHAR(MAX)'), 1, 2, '')
                    ) AS equipo_resumen,
                    r.requiere_carta_cliente,
                    r.carta_cliente_notif_at,
                    r.creado_por,
                    COALESCE(r.validacion_creador, '') AS validacion_creador

                FROM reclamos r
                OUTER APPLY (
                    SELECT TOP 1 ri1.*
                    FROM reclamo_imputados ri1
                    WHERE ri1.reclamo_id = r.id
                    ORDER BY ri1.id DESC
                ) ri
                LEFT JOIN cantones c ON c.id = r.canton_id
                LEFT JOIN param_groups gtr ON gtr.nombre = 'RECL_TIPO'
                LEFT JOIN param_values tr ON tr.group_id = gtr.id AND tr.nombre = r.tipo_reclamo
                LEFT JOIN param_groups gtt ON gtt.nombre = 'RECL_TRAMITE'
                LEFT JOIN param_values tt ON tt.group_id = gtt.id AND tt.nombre = r.tipo_tramite
                WHERE {where}
                ORDER BY r.id DESC
            """

            cur.execute(sql_created, params)
            created_list = cur.fetchall()

        # 2) Reclamos donde soy aprobador (jefe)
        approve_list = []
        if uid:
            jefe_id = int(uid)

            f_sql, f_params = _build_where_filtros("r")

            sql_approve = f"""
                SELECT
                    r.id,
                    r.codigo,
                    r.fecha_reclamo,
                    r.fecha_creacion,
                    r.cliente_nombre,
                    r.proceso_text AS proceso_text,
                    r.tipo_tramite AS tipo_tramite,
                    r.material_desc AS material_desc,
                    r.procede AS procede,
                    r.observacion AS observacion,
                    c.nombre AS ciudad,

                    ri.id AS imputacion_id,
                    ri.estado_asignacion,
                    ri.estado_respuesta,

                    CASE
                        WHEN ri.estado_asignacion = 'pend_aprobacion'
                            THEN 'Pendiente aceptación del responsable'
                        WHEN ri.estado_asignacion = 'rechazado'
                            THEN 'Imputación rechazada'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'sin_respuesta'
                            THEN 'Pendiente respuesta del imputado'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'pendiente_jefe'
                            THEN 'Respuesta pendiente de aprobación'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'aprobada'
                            THEN 'Cerrado'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'rechazada'
                            THEN 'Respuesta rechazada'
                        ELSE COALESCE(ri.estado_asignacion,'') + '/' + COALESCE(ri.estado_respuesta,'')
                    END AS estado_imputacion,

                    ri.motivo_rechazo_asignacion AS rechazo_motivo,
                    ri.respuesta_causa           AS causa,
                    ri.respuesta_preventiva      AS preventiva,
                    ri.respuesta_correctiva      AS correctiva,
                    COALESCE(ri.fecha_causa,'')        AS fecha_causa,
                    COALESCE(ri.fecha_preventiva,'')   AS fecha_preventiva,
                    COALESCE(ri.fecha_correctiva,'')   AS fecha_correctiva,

                    ri.metodo_analisis AS metodo_analisis,
                    ri.why1,
                    ri.why2,
                    ri.why3,
                    ri.why4,
                    ri.why5,
                    ri.fish_metodo,
                    ri.fish_maquinas,
                    ri.fish_materiales,
                    ri.fish_personas,
                    ri.fish_entorno,
                    ri.fish_medicion,

                    u.nombre_completo AS imputado_nombre
                FROM reclamo_imputados ri
                JOIN reclamos r ON r.id = ri.reclamo_id
                LEFT JOIN usuarios u ON u.id = ri.imputado_id
                LEFT JOIN cantones c ON c.id = r.canton_id
                WHERE ri.aprobador_id = ?
                AND (
                        TRIM(ri.estado_asignacion) = 'pend_aprobacion'
                        OR (
                            TRIM(ri.estado_asignacion) = 'aprobado'
                            AND TRIM(ri.estado_respuesta) = 'pendiente_jefe'
                        )
                    )
                {f_sql}
                ORDER BY r.id DESC, ri.id DESC
            """

            cur.execute(sql_approve, (jefe_id, *f_params))
            approve_list = cur.fetchall()

        # 3) Donde soy imputado (sponsor)
        imputado_list = []
        if uid:
            params = []
            can_view_all = _can_view_all_reclamos_sn_sponsor(conn, uid)

            if can_view_all:
                where = "ri.estado_asignacion = 'aprobado'"
            else:
                where = """
                    ri.estado_asignacion = 'aprobado'
                    AND (
                        ri.imputado_id = ?
                        OR EXISTS (
                            SELECT 1
                            FROM param_values pvs
                            JOIN param_groups pgs
                            ON pgs.id = pvs.group_id
                            JOIN usuarios uspv
                            ON LTRIM(RTRIM(uspv.identificacion)) = LTRIM(RTRIM(pvs.nombre))
                            WHERE pgs.nombre = 'RECL_PROCESO_SPONSOR'
                            AND COALESCE(pvs.activo, 1) = 1
                            AND pvs.parent_id = r.proceso_id
                            AND uspv.id = ?
                            AND COALESCE(uspv.disabled, 0) = 0
                            AND UPPER(LTRIM(RTRIM(COALESCE(pvs.valor, '')))) IN ('PRINCIPAL', 'BACKUP')
                        )
                    )
                """
                params = [uid, uid]

            f_sql, f_params = _build_where_filtros("r")
            where += f_sql
            params += f_params

            sql_imputado = f"""
                SELECT
                    r.id,
                    r.codigo,
                    r.fecha_reclamo,
                    r.fecha_creacion,
                    r.cliente_nombre,
                    r.observacion,
                    r.proceso_text AS proceso_text,
                    r.material_desc AS material_desc,
                    r.procede AS procede,
                    c.nombre AS ciudad,

                    tr.valor AS motivo,
                    tt.valor AS tramite,
                    r.antecedente AS submotivo,
                    r.tipo_reclamo,
                    r.tipo_tramite,

                    ri.id AS imputacion_id,

                    CASE
                        WHEN ri.estado_asignacion = 'pend_aprobacion'
                            THEN 'Pendiente aceptación del responsable'
                        WHEN ri.estado_asignacion = 'rechazado'
                            THEN 'Imputación rechazada'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'sin_respuesta'
                            THEN 'Pendiente respuesta del imputado'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'pendiente_jefe'
                            THEN 'Respuesta pendiente de aprobación'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'aprobada'
                            THEN 'Cerrado'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'rechazada'
                            THEN 'Respuesta rechazada'
                        ELSE COALESCE(ri.estado_asignacion,'') + '/' + COALESCE(ri.estado_respuesta,'')
                    END AS estado_imputacion,

                    ri.respuesta_causa      AS causa,
                    ri.respuesta_preventiva AS preventiva,
                    ri.respuesta_correctiva AS correctiva,
                    COALESCE(ri.fecha_causa,'')        AS fecha_causa,
                    COALESCE(ri.fecha_preventiva,'')   AS fecha_preventiva,
                    COALESCE(ri.fecha_correctiva,'')   AS fecha_correctiva,

                    ri.metodo_analisis AS metodo_analisis,
                    ri.why1, ri.why2, ri.why3, ri.why4, ri.why5,
                    ri.fish_metodo, ri.fish_maquinas, ri.fish_materiales,
                    ri.fish_personas, ri.fish_entorno, ri.fish_medicion,

                    u_creador.nombre_completo AS creador_nombre,
                    u_creador.username        AS creador_username,

                    u_imp.nombre_completo AS imputado_nombre,
                    u_imp.username        AS imputado_username,

                    (
                        SELECT STUFF((
                            SELECT DISTINCT ', ' + COALESCE(u2.nombre_completo, u2.username)
                            FROM reclamo_equipo_respuestas eq2
                            JOIN usuarios u2 ON u2.id = eq2.usuario_id
                            WHERE eq2.reclamo_id = r.id
                            AND eq2.activo = 1
                            FOR XML PATH(''), TYPE
                        ).value('.', 'NVARCHAR(MAX)'), 1, 2, '')
                    ) AS equipo_resumen,
                r.requiere_carta_cliente,
                r.carta_cliente_notif_at

                FROM reclamo_imputados ri
                JOIN reclamos r ON r.id = ri.reclamo_id
                LEFT JOIN cantones c ON c.id = r.canton_id
                LEFT JOIN usuarios u_creador ON u_creador.id = r.creado_por
                LEFT JOIN usuarios u_imp ON u_imp.id = ri.imputado_id
                LEFT JOIN param_groups gtr ON gtr.nombre = 'RECL_TIPO'
                LEFT JOIN param_values tr ON tr.group_id = gtr.id AND tr.nombre = r.tipo_reclamo
                LEFT JOIN param_groups gtt ON gtt.nombre = 'RECL_TRAMITE'
                LEFT JOIN param_values tt ON tt.group_id = gtt.id AND tt.nombre = r.tipo_tramite
                WHERE {where}
                ORDER BY r.id DESC, ri.id DESC
            """

            cur.execute(sql_imputado, params)
            imputado_list = cur.fetchall()

        tipos_reclamo = _fetch_param_values(conn, "RECL_TIPO")
        tipos_tramite = _fetch_param_values(conn, "RECL_TRAMITE")
        #procesos = _fetch_param_values(conn, "RECL_PROCESO")
        procesos = _fetch_recl_procesos(conn)
        tipos_campos = _fetch_tipo_campos(conn)
        regiones = fetch_regiones(conn)
        materiales = _fetch_param_values(conn, "RECL_MATERIAL")
        usuarios_imputables = fetch_usuarios_imputables(conn)

        # 4) OM donde soy colaborador del responsable (equipo)
        equipo_list = []
        if uid:
            f_sql, f_params = _build_where_filtros("r")

            sql_equipo = f"""
                SELECT
                    eq.id            AS equipo_id,
                    eq.imputacion_id AS imputacion_id,
                    eq.reclamo_id    AS reclamo_id,
                    eq.usuario_id    AS usuario_id,
                    rre.miembro_id   AS miembro_id,

                    rre.fecha_causa,
                    rre.fecha_preventiva,
                    rre.fecha_correctiva,

                    tr.valor AS motivo,
                    r.antecedente AS submotivo,
                    tt.valor AS tramite,

                    r.codigo,
                    r.fecha_reclamo,
                    r.fecha_creacion,

                    COALESCE(r.cliente_nombre, cli.nombre) AS cliente_nombre,
                    r.observacion,
                    r.proceso_text AS proceso_text,
                    COALESCE(r.material_desc, pv.valor, pv.nombre) AS material_desc,

                    COALESCE(rre.causa, '')      AS causa,
                    COALESCE(rre.preventiva, '') AS preventiva,
                    COALESCE(rre.correctiva, '') AS correctiva,

                    COALESCE((
                        SELECT STUFF((
                            SELECT ' | ' +
                                COALESCE(a2.descripcion, '') +
                                CASE
                                    WHEN a2.fecha_compromiso IS NOT NULL
                                    THEN ' (' + CONVERT(VARCHAR(10), a2.fecha_compromiso, 23) + ')'
                                    ELSE ''
                                END
                            FROM reclamo_respuesta_equipo_acciones a2
                            WHERE a2.respuesta_equipo_id = rre.id
                            AND a2.tipo = 'CAUSA'
                            AND a2.activo = 1
                            ORDER BY a2.orden, a2.id
                            FOR XML PATH(''), TYPE
                        ).value('.', 'NVARCHAR(MAX)'), 1, 3, '')
                    ), '') AS causas_detalle,

                    COALESCE((
                        SELECT STUFF((
                            SELECT ' | ' +
                                COALESCE(a2.descripcion, '') +
                                CASE
                                    WHEN a2.fecha_compromiso IS NOT NULL
                                    THEN ' (' + CONVERT(VARCHAR(10), a2.fecha_compromiso, 23) + ')'
                                    ELSE ''
                                END
                            FROM reclamo_respuesta_equipo_acciones a2
                            WHERE a2.respuesta_equipo_id = rre.id
                            AND a2.tipo = 'CONTROL'
                            AND a2.activo = 1
                            ORDER BY a2.orden, a2.id
                            FOR XML PATH(''), TYPE
                        ).value('.', 'NVARCHAR(MAX)'), 1, 3, '')
                    ), '') AS control_detalle,

                    COALESCE((
                        SELECT STUFF((
                            SELECT ' | ' +
                                COALESCE(a2.descripcion, '') +
                                CASE
                                    WHEN a2.fecha_compromiso IS NOT NULL
                                    THEN ' (' + CONVERT(VARCHAR(10), a2.fecha_compromiso, 23) + ')'
                                    ELSE ''
                                END
                            FROM reclamo_respuesta_equipo_acciones a2
                            WHERE a2.respuesta_equipo_id = rre.id
                            AND a2.tipo = 'CORRECTIVA'
                            AND a2.activo = 1
                            ORDER BY a2.orden, a2.id
                            FOR XML PATH(''), TYPE
                        ).value('.', 'NVARCHAR(MAX)'), 1, 3, '')
                    ), '') AS correctiva_detalle,

                    COALESCE(rre.metodo_analisis, '') AS metodo_analisis,
                    COALESCE(rre.why1, '') AS why1,
                    COALESCE(rre.why2, '') AS why2,
                    COALESCE(rre.why3, '') AS why3,
                    COALESCE(rre.why4, '') AS why4,
                    COALESCE(rre.why5, '') AS why5,

                    COALESCE(rre.fish_metodo, '')     AS fish_metodo,
                    COALESCE(rre.fish_maquinas, '')   AS fish_maquinas,
                    COALESCE(rre.fish_materiales, '') AS fish_materiales,
                    COALESCE(rre.fish_personas, '')   AS fish_personas,
                    COALESCE(rre.fish_entorno, '')    AS fish_entorno,
                    COALESCE(rre.fish_medicion, '')   AS fish_medicion,

                    CASE
                        WHEN EXISTS (
                            SELECT 1
                            FROM reclamo_respuesta_equipo_acciones a
                            WHERE a.respuesta_equipo_id = rre.id
                            AND a.activo = 1
                        )
                        OR COALESCE(TRIM(rre.causa), '') <> ''
                        OR COALESCE(TRIM(rre.preventiva), '') <> ''
                        OR COALESCE(TRIM(rre.correctiva), '') <> ''
                        THEN ''
                        ELSE ''
                    END AS estado_equipo,

                    COALESCE(us.nombre_completo, us.username, '') AS sponsor,
                    r.estado_global,
                    COALESCE(ucr.nombre_completo, ucr.username, '') AS creador_nombre

                FROM reclamo_equipo_respuestas eq
                JOIN reclamos r ON r.id = eq.reclamo_id
                LEFT JOIN reclamo_imputados ri ON ri.id = eq.imputacion_id
                LEFT JOIN usuarios us ON us.id = ri.imputado_id
                LEFT JOIN usuarios ucr ON ucr.id = r.creado_por
                LEFT JOIN param_groups gtr ON gtr.nombre = 'RECL_TIPO'
                LEFT JOIN param_values tr ON tr.group_id = gtr.id AND tr.nombre = r.tipo_reclamo
                LEFT JOIN param_groups gtt ON gtt.nombre = 'RECL_TRAMITE'
                LEFT JOIN param_values tt ON tt.group_id = gtt.id AND tt.nombre = r.tipo_tramite
                LEFT JOIN reclamo_respuestas_equipo rre
                    ON rre.id = (
                            SELECT MAX(rre2.id)
                            FROM reclamo_respuestas_equipo rre2
                            WHERE rre2.reclamo_id = eq.reclamo_id
                            AND rre2.imputacion_id = eq.imputacion_id
                            AND rre2.miembro_id = eq.usuario_id
                            AND rre2.activo = 1
                    )
                LEFT JOIN terceros cli
                    ON cli.id = r.cliente_id
                    AND cli.tipo = 'C'
                LEFT JOIN param_values pv
                    ON pv.id = r.material_id
                    AND pv.group_id = (
                            SELECT TOP 1 id
                            FROM param_groups
                            WHERE nombre = 'RECL_MATERIAL'
                    )
                WHERE eq.usuario_id = ?
                AND eq.activo = 1
                {f_sql}
                ORDER BY r.fecha_reclamo DESC, r.id DESC, eq.id DESC
            """

            cur.execute(sql_equipo, (uid, *f_params))
            equipo_list = cur.fetchall()

        def _get_gerente_general_email(conn):
            cur2 = conn.cursor()
            if is_sqlserver:
                cur2.execute(SQL__GET_GERENTE_GENERAL_EMAIL_SEL_SS)
            else:
                cur2.execute(SQL__GET_GERENTE_GENERAL_EMAIL_SEL_SL)
            row = cur2.fetchone()
            return row["email"] if row else None

        def _notify_gg_if_needed(conn, dias_alerta: int):
            gg_email = _get_gerente_general_email(conn)
            if not gg_email:
                return

            cur2 = conn.cursor()

            if is_sqlserver:
                cur2.execute(SQL__NOTIFY_GG_IF_NEEDED_SEL_SS, (dias_alerta,))
            else:
                cur2.execute(SQL__NOTIFY_GG_IF_NEEDED_SEL_SL, (dias_alerta,))

            rows = cur2.fetchall()
            if not rows:
                return

            items = "\n".join([f"- {x['codigo']} | {x['cliente_nombre']} | {x['fecha_reclamo']}" for x in rows])
            subject = "[OM] Alerta: Sponsor sin respuesta (día 5)"
            text_body = f"""Estimado/a,

    Se detectaron Oportunidades de Mejora sin respuesta del sponsor al día {dias_alerta}:

    {items}

    Por favor revisar en la plataforma (tab: Soy Sponsor / Responsable).

    Saludos.
    """

            _send_mail_safe(gg_email, subject, text_body, None)

            ids = [x["id"] for x in rows]
            cur2.execute(
                f"UPDATE reclamos SET gg_notificado=1, gg_notificado_fecha=? WHERE id IN ({','.join(['?']*len(ids))})",
                (_now_iso(), *ids)
            )
            conn.commit()

        dias_plazo = _get_param_int_by_id(conn, 11047, default=5)

        created_list  = [dict(x) for x in (created_list or [])]
        approve_list  = [dict(x) for x in (approve_list or [])]
        imputado_list = [dict(x) for x in (imputado_list or [])]
        equipo_list   = [dict(x) for x in (equipo_list or [])]

        created_list  = _enrich_deadline(created_list, dias_plazo)
        approve_list  = _enrich_deadline(approve_list, dias_plazo)
        imputado_list = _enrich_deadline(imputado_list, dias_plazo)
        equipo_list   = _enrich_deadline(equipo_list, dias_plazo)

        # _notify_gg_if_needed(conn, dias_alerta=5)
        puede_eliminar_om = _can_delete_om(conn, uid)
        puede_subir_carta_cliente = _can_upload_carta_cliente(conn, uid)
        conn.close()

        return render_template(
            'reclamos_lista.html',
            created_list=created_list,
            approve_list=approve_list,
            imputado_list=imputado_list,
            q_estado=q_estado,
            rol=session.get('rol'),
            usuario=session.get('usuario'),
            user_id=uid,
            permissions=session.get('permissions', {}),
            tipos_reclamo=tipos_reclamo,
            tipos_tramite=tipos_tramite,
            materiales=materiales,
            puede_crear_materiales=puede_crear_materiales,
            procesos=procesos,
            tipos_campos=tipos_campos,
            regiones=regiones,
            usuarios_imputables=usuarios_imputables,
            productos=productos,
            equipo_list=equipo_list,
            filtros=filtros,
            puede_eliminar_om=puede_eliminar_om,
            active_page='reclamos',
            puede_subir_carta_cliente=puede_subir_carta_cliente,
            current_user_id=int(uid) if uid else None
        )

 # ----------------------------------------------
# EXPORTAR A EXCEL - RECLAMOS DONDE INTERVIENE EL USUARIO
# (admin/coordinador => exporta TODO)
# + Respeta filtros del front (codigo, cliente, proceso, motivo, estado_global, desde, hasta)
# + Incluye columnas del CREADOR de la OM (creado_por -> usuarios)
# ----------------------------------------------
    @app.route('/reclamos/export/mis', methods=['GET'], endpoint='reclamos_export_mis')
    @require_login
    @require_permission('reclamos', 'ver')
    def reclamos_export_mis():
        uid = _current_user_id()
        if not uid:
            flash("No se pudo determinar el usuario actual.", "danger")
            return redirect(url_for('reclamos'))

        conn = get_db()
        export_all = _can_export_all_reclamos(conn, uid)
        cur = conn.cursor()

        f_codigo = (request.args.get('codigo') or '').strip()
        f_cliente = (request.args.get('cliente') or '').strip()
        f_proceso = (request.args.get('proceso') or '').strip()
        f_motivo = (request.args.get('motivo') or '').strip()
        f_estado_global = (request.args.get('estado_global') or '').strip()
        f_desde = (request.args.get('desde') or '').strip()
        f_hasta = (request.args.get('hasta') or '').strip()

        def _build_where_export(alias_r="r"):
            w = ""
            p = []

            if f_codigo:
                w += f" AND {alias_r}.codigo LIKE ?"
                p.append(f"%{f_codigo}%")

            if f_cliente:
                w += f" AND ISNULL({alias_r}.cliente_nombre,'') LIKE ?"
                p.append(f"%{f_cliente}%")

            if f_proceso:
                w += f" AND ISNULL({alias_r}.proceso_text,'') LIKE ?"
                p.append(f"%{f_proceso}%")

            if f_motivo:
                w += f" AND ISNULL({alias_r}.tipo_reclamo,'') = ?"
                p.append(f_motivo)

            if f_estado_global:
                w += f" AND ISNULL({alias_r}.estado_global,'') = ?"
                p.append(f_estado_global)

            if f_desde:
                w += f" AND TRY_CONVERT(date, {alias_r}.fecha_reclamo) >= TRY_CONVERT(date, ?)"
                p.append(f_desde)

            if f_hasta:
                w += f" AND TRY_CONVERT(date, {alias_r}.fecha_reclamo) <= TRY_CONVERT(date, ?)"
                p.append(f_hasta)

            return w, p

        f_sql, f_params = _build_where_export("r")

        sql_ctes = """
            WITH
            stats_respuesta_sponsor AS (
                SELECT
                    ri.reclamo_id,
                    ri.id AS imputacion_id,
                    AVG(
                        CASE
                            WHEN TRY_CONVERT(date, ri.fecha_respuesta_imputado) IS NOT NULL
                            AND TRY_CONVERT(date, r.fecha_reclamo) IS NOT NULL
                            THEN DATEDIFF(
                                DAY,
                                TRY_CONVERT(date, r.fecha_reclamo),
                                TRY_CONVERT(date, ri.fecha_respuesta_imputado)
                            )
                        END
                    ) AS dias_promedio_respuesta_sponsor
                FROM reclamo_imputados ri
                JOIN reclamos r ON r.id = ri.reclamo_id
                GROUP BY ri.reclamo_id, ri.id
            ),

            equipo_asignacion_detalle AS (
                SELECT
                    eq.reclamo_id,
                    eq.imputacion_id,
                    eq.usuario_id AS miembro_id,
                    ISNULL(u.nombre_completo, u.username) AS miembro_nombre,
                    eq.creado_at AS fecha_asignacion_miembro,
                    CASE
                        WHEN TRY_CONVERT(date, eq.creado_at) IS NOT NULL
                        AND TRY_CONVERT(date, r.fecha_reclamo) IS NOT NULL
                        THEN DATEDIFF(
                            DAY,
                            TRY_CONVERT(date, r.fecha_reclamo),
                            TRY_CONVERT(date, eq.creado_at)
                        )
                    END AS dias_asignacion_miembro
                FROM reclamo_equipo_respuestas eq
                JOIN reclamos r ON r.id = eq.reclamo_id
                JOIN usuarios u ON u.id = eq.usuario_id
                WHERE ISNULL(eq.activo, 1) = 1
            ),

            equipo_respuesta_detalle AS (
                SELECT
                    ead.reclamo_id,
                    ead.imputacion_id,
                    ead.miembro_id,
                    ead.miembro_nombre,
                    ead.fecha_asignacion_miembro,
                    ead.dias_asignacion_miembro,

                    ISNULL(rre.revision_at, rre.created_at) AS fecha_respuesta_miembro,

                    CASE
                        WHEN TRY_CONVERT(date, ead.fecha_asignacion_miembro) IS NOT NULL
                        AND TRY_CONVERT(date, ISNULL(rre.revision_at, rre.created_at)) IS NOT NULL
                        THEN DATEDIFF(
                            DAY,
                            TRY_CONVERT(date, ead.fecha_asignacion_miembro),
                            TRY_CONVERT(date, ISNULL(rre.revision_at, rre.created_at))
                        )
                    END AS dias_respuesta_miembro,

                    CASE
                        WHEN TRY_CONVERT(date, ead.fecha_asignacion_miembro) IS NOT NULL
                        AND ISNULL(rre.revision_at, rre.created_at) IS NULL
                        THEN DATEDIFF(
                            DAY,
                            TRY_CONVERT(date, ead.fecha_asignacion_miembro),
                            CAST(GETDATE() AS date)
                        )
                        ELSE 0
                    END AS dias_sin_respuesta_miembro
                FROM equipo_asignacion_detalle ead
                LEFT JOIN reclamo_respuestas_equipo rre
                    ON rre.id = (
                        SELECT MAX(rre2.id)
                        FROM reclamo_respuestas_equipo rre2
                        WHERE rre2.reclamo_id = ead.reclamo_id
                        AND rre2.imputacion_id = ead.imputacion_id
                        AND rre2.miembro_id = ead.miembro_id
                        AND ISNULL(rre2.activo, 1) = 1
                    )
            ),

            stats_asignacion_equipo AS (
                SELECT
                    x.reclamo_id,
                    x.imputacion_id,
                    STRING_AGG(x.miembro_nombre, ', ') AS miembros_equipo,
                    COUNT(DISTINCT x.miembro_id) AS total_miembros_equipo,
                    MIN(x.fecha_asignacion_miembro) AS fecha_primera_asignacion_equipo,
                    AVG(CAST(x.dias_asignacion_miembro AS decimal(18,2))) AS dias_promedio_asignacion_equipo
                FROM (
                    SELECT DISTINCT
                        reclamo_id,
                        imputacion_id,
                        miembro_id,
                        miembro_nombre,
                        fecha_asignacion_miembro,
                        dias_asignacion_miembro
                    FROM equipo_asignacion_detalle
                ) x
                GROUP BY x.reclamo_id, x.imputacion_id
            ),

            stats_respuesta_equipo AS (
                SELECT
                    erd.reclamo_id,
                    erd.imputacion_id,

                    STRING_AGG(
                        CASE
                            WHEN erd.fecha_respuesta_miembro IS NOT NULL
                            THEN erd.miembro_nombre
                        END,
                        ', '
                    ) AS miembros_equipo_respondieron,

                    STRING_AGG(
                        CASE
                            WHEN erd.fecha_respuesta_miembro IS NULL
                            THEN erd.miembro_nombre
                        END,
                        ', '
                    ) AS miembros_equipo_pendientes,

                    MIN(erd.fecha_respuesta_miembro) AS fecha_primera_respuesta_equipo,
                    AVG(CAST(erd.dias_respuesta_miembro AS decimal(18,2))) AS dias_promedio_respuesta_equipo,

                    AVG(
                        CAST(
                            CASE
                                WHEN erd.fecha_respuesta_miembro IS NULL
                                THEN erd.dias_sin_respuesta_miembro
                            END AS decimal(18,2)
                        )
                    ) AS dias_promedio_sin_respuesta_equipo,

                    MAX(erd.dias_sin_respuesta_miembro) AS dias_max_sin_respuesta_equipo
                FROM equipo_respuesta_detalle erd
                GROUP BY erd.reclamo_id, erd.imputacion_id
            )
        """

        sql_select = """
            SELECT
                r.codigo AS codigo_om,
                r.fecha_reclamo,
                r.fecha_creacion,

                ucr.username AS creador_username,
                ISNULL(ucr.nombre_completo, ucr.username) AS creador_nombre,

                tr.valor AS tipo_om,
                tt.valor AS tipo_tramite,
                r.proceso_text,
                r.cliente_nombre,
                r.cliente_identificacion,
                r.cliente_contacto,
                r.cliente_email,
                r.cliente_telefono,
                r.material_desc,
                c.nombre AS ciudad,
                r.observacion,
                r.procede,
                r.estado_global,

                ri.id AS imputacion_id,
                ui.username AS imputado_username,
                ISNULL(ui.nombre_completo, ui.username) AS imputado_nombre,
                uj.username AS jefe_username,
                ISNULL(uj.nombre_completo, uj.username) AS jefe_nombre,

                ri.estado_asignacion,
                ri.fecha_aprobacion_asignacion,
                ri.fecha_rechazo_asignacion,
                ri.motivo_rechazo_asignacion,

                ri.respuesta_causa,
                ri.respuesta_preventiva,
                ri.respuesta_correctiva,
                ISNULL(ri.fecha_causa, '') AS fecha_causa,
                ISNULL(ri.fecha_preventiva, '') AS fecha_preventiva,
                ISNULL(ri.fecha_correctiva, '') AS fecha_correctiva,
                ri.fecha_respuesta_imputado,

                CASE
                    WHEN TRY_CONVERT(date, ri.fecha_respuesta_imputado) IS NOT NULL
                    AND TRY_CONVERT(date, r.fecha_reclamo) IS NOT NULL
                    THEN DATEDIFF(
                        DAY,
                        TRY_CONVERT(date, r.fecha_reclamo),
                        TRY_CONVERT(date, ri.fecha_respuesta_imputado)
                    )
                END AS dias_respuesta_sponsor,

                CASE
                    WHEN ri.fecha_respuesta_imputado IS NULL
                    AND TRY_CONVERT(date, r.fecha_reclamo) IS NOT NULL
                    THEN DATEDIFF(
                        DAY,
                        TRY_CONVERT(date, r.fecha_reclamo),
                        CAST(GETDATE() AS date)
                    )
                    ELSE 0
                END AS dias_sin_respuesta_sponsor,

                ISNULL(sae.miembros_equipo, '') AS miembros_equipo,
                ISNULL(sae.total_miembros_equipo, 0) AS total_miembros_equipo,
                ISNULL(CONVERT(varchar(19), sae.fecha_primera_asignacion_equipo, 120), '') AS fecha_primera_asignacion_equipo,
                ISNULL(sae.dias_promedio_asignacion_equipo, 0) AS dias_promedio_asignacion_equipo,

                ISNULL(sre.miembros_equipo_respondieron, '') AS miembros_equipo_respondieron,
                ISNULL(sre.miembros_equipo_pendientes, '') AS miembros_equipo_pendientes,
                ISNULL(CONVERT(varchar(19), sre.fecha_primera_respuesta_equipo, 120), '') AS fecha_primera_respuesta_equipo,
                ISNULL(sre.dias_promedio_respuesta_equipo, 0) AS dias_promedio_respuesta_equipo,
                ISNULL(sre.dias_promedio_sin_respuesta_equipo, 0) AS dias_promedio_sin_respuesta_equipo,
                ISNULL(sre.dias_max_sin_respuesta_equipo, 0) AS dias_max_sin_respuesta_equipo,

                ri.estado_respuesta,
                ri.fecha_aprobacion_respuesta,
                ri.fecha_rechazo_respuesta,
                ri.motivo_rechazo_respuesta

            FROM reclamos r
            LEFT JOIN reclamo_imputados ri ON ri.reclamo_id = r.id
            LEFT JOIN usuarios ui ON ui.id = ri.imputado_id
            LEFT JOIN usuarios uj ON uj.id = ri.aprobador_id
            LEFT JOIN usuarios ucr ON ucr.id = r.creado_por
            LEFT JOIN cantones c ON c.id = r.canton_id

            LEFT JOIN stats_respuesta_sponsor srs
                ON srs.reclamo_id = r.id
            AND srs.imputacion_id = ri.id

            LEFT JOIN stats_asignacion_equipo sae
                ON sae.reclamo_id = r.id
            AND sae.imputacion_id = ri.id

            LEFT JOIN stats_respuesta_equipo sre
                ON sre.reclamo_id = r.id
            AND sre.imputacion_id = ri.id

            LEFT JOIN param_groups gtr ON gtr.nombre = 'RECL_TIPO'
            LEFT JOIN param_values tr ON tr.group_id = gtr.id AND tr.nombre = r.tipo_reclamo

            LEFT JOIN param_groups gtt ON gtt.nombre = 'RECL_TRAMITE'
            LEFT JOIN param_values tt ON tt.group_id = gtt.id AND tt.nombre = r.tipo_tramite

            {where_clause}
            ORDER BY r.id DESC, ri.id DESC
        """

        if export_all:
            where_clause = f"""
                WHERE 1=1
                {f_sql}
            """
            sql = sql_ctes + sql_select.format(where_clause=where_clause)
            cur.execute(sql, f_params)
        else:
            where_clause = f"""
                WHERE r.id IN (SELECT id FROM mis_reclamos)
                {f_sql}
            """
            sql = sql_ctes + """
                , mis_reclamos AS (
                    SELECT DISTINCT r.id
                    FROM reclamos r
                    LEFT JOIN reclamo_imputados ri ON ri.reclamo_id = r.id
                    LEFT JOIN reclamo_equipo_respuestas eq
                        ON eq.reclamo_id = r.id
                    AND ISNULL(eq.activo, 1) = 1
                    LEFT JOIN reclamo_respuestas_equipo rre
                        ON rre.reclamo_id = r.id
                    AND ISNULL(rre.activo, 1) = 1
                    WHERE r.creado_por = ?
                    OR ri.imputado_id = ?
                    OR ri.aprobador_id = ?
                    OR eq.usuario_id = ?
                    OR rre.miembro_id = ?
                )
            """ + sql_select.format(where_clause=where_clause)

            cur.execute(sql, (uid, uid, uid, uid, uid, *f_params))

        rows = cur.fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Reclamos (Todos)" if export_all else "Mis reclamos"

        headers = [
            "Código OM",
            "Fecha OM",
            "Fecha creación (sistema)",
            "Creador OM (usuario)",
            "Creador OM (nombre)",
            "Tipo OM",
            "Tipo trámite",
            "Proceso",
            "Cliente",
            "Identificación cliente",
            "Contacto cliente",
            "Correo cliente",
            "Teléfono cliente",
            "Ciudad",
            "Material",
            "Observación",
            "Procede",
            "Estado global",
            "ID imputación",
            "Sponsor / Responsable (usuario)",
            "Sponsor / Responsable (nombre)",
            "Jefe/aprobador (usuario)",
            "Jefe/aprobador (nombre)",
            "Estado imputación",
            "Fecha aprobación imputación",
            "Fecha rechazo imputación",
            "Motivo rechazo imputación",
            "Respuesta sponsor - Causa",
            "Respuesta sponsor - Acción preventiva",
            "Respuesta sponsor - Acción correctiva",
            "Fecha causa sponsor",
            "Fecha preventiva sponsor",
            "Fecha correctiva sponsor",
            "Fecha respuesta sponsor",
            "Días respuesta sponsor",
            "Días sin respuesta sponsor",
            "Miembros equipo designados",
            "Total miembros equipo",
            "Fecha primera asignación equipo",
            "Días promedio asignación equipo",
            "Miembros equipo que respondieron",
            "Miembros equipo pendientes",
            "Fecha primera respuesta equipo",
            "Días promedio respuesta equipo",
            "Días promedio sin respuesta equipo",
            "Días máximo sin respuesta equipo",
            "Estado respuesta sponsor",
            "Fecha aprobación respuesta sponsor",
            "Fecha rechazo respuesta sponsor",
            "Motivo rechazo respuesta sponsor",
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F2937")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="D1D5DB")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        for r in rows:
            ws.append([
                r["codigo_om"],
                r["fecha_reclamo"],
                r["fecha_creacion"],
                r["creador_username"],
                r["creador_nombre"],
                r["tipo_om"],
                r["tipo_tramite"],
                r["proceso_text"],
                r["cliente_nombre"],
                r["cliente_identificacion"],
                r["cliente_contacto"],
                r["cliente_email"],
                r["cliente_telefono"],
                r["ciudad"],
                r["material_desc"],
                r["observacion"],
                r["procede"],
                r["estado_global"],
                r["imputacion_id"],
                r["imputado_username"],
                r["imputado_nombre"],
                r["jefe_username"],
                r["jefe_nombre"],
                r["estado_asignacion"],
                r["fecha_aprobacion_asignacion"],
                r["fecha_rechazo_asignacion"],
                r["motivo_rechazo_asignacion"],
                r["respuesta_causa"],
                r["respuesta_preventiva"],
                r["respuesta_correctiva"],
                r["fecha_causa"],
                r["fecha_preventiva"],
                r["fecha_correctiva"],
                r["fecha_respuesta_imputado"],
                r["dias_respuesta_sponsor"],
                r["dias_sin_respuesta_sponsor"],
                r["miembros_equipo"],
                r["total_miembros_equipo"],
                r["fecha_primera_asignacion_equipo"],
                r["dias_promedio_asignacion_equipo"],
                r["miembros_equipo_respondieron"],
                r["miembros_equipo_pendientes"],
                r["fecha_primera_respuesta_equipo"],
                r["dias_promedio_respuesta_equipo"],
                r["dias_promedio_sin_respuesta_equipo"],
                r["dias_max_sin_respuesta_equipo"],
                r["estado_respuesta"],
                r["fecha_aprobacion_respuesta"],
                r["fecha_rechazo_respuesta"],
                r["motivo_rechazo_respuesta"],
            ])

        max_row = ws.max_row
        max_col = ws.max_column
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(val))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 70)

        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "reclamos_todos.xlsx" if export_all else f"mis_reclamos_{uid}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )    
    
    # ----------------------------------------------
    # APIs geo
    # ----------------------------------------------
    @app.route('/reclamos/api/provincias', methods=['GET'], endpoint='reclamos_api_provincias')
    @require_login
    def reclamos_api_provincias():
        region_id = request.args.get('region_id', type=int)

        conn = get_db()
        #ensure_geo_schema(conn)
        rows = fetch_provincias(conn, region_id)
        conn.close()

    #secure core

        data = [dict(id=r['id'], nombre=r['nombre']) for r in rows]
        return jsonify(data)

    @app.route('/reclamos/api/materiales/nuevo', methods=['POST'], endpoint='reclamos_api_materiales_nuevo')
    @require_login
    @require_permission('reclamos', 'editar')
    def reclamos_api_materiales_nuevo():
        nombre = (request.form.get('nombre') or '').strip()
        if not nombre:
            return jsonify(ok=False, msg="El nombre del material es obligatorio"), 400

        uid = _current_user_id()
        conn = get_db()
        #ensure_reclamos_schema(conn)
        #ensure_reclamos_catalogos(conn)
        cur = conn.cursor()

        # Permisos:
        if not _is_admin_like():
            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_53, (uid,))
            row_p = cur.fetchone()
            nombre_puesto = (row_p["nombre"] if row_p and row_p["nombre"] else "").strip().upper()
            if nombre_puesto != "CORRDINADOR(A) SERVICIO AL CLIENTE":
                conn.close()
                return jsonify(ok=False, msg="No está autorizado para crear materiales."), 403

        base = ''.join(
            ch for ch in nombre.upper()
            if (ch.isalnum() or ch in (" ", "_", "-"))
        ).replace(" ", "_")
        if not base:
            base = "MAT"

        codigo = base
        contador = 1
        while True:
            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_54, (codigo,))
            if not cur.fetchone():
                break
            contador += 1
            codigo = f"{base}_{contador}"

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_55)
        row = cur.fetchone()
        next_ord = row["next_ord"] if row else 1

        gid = _ensure_param_group(conn, "RECL_MATERIAL", "Materiales de reclamos")

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_INS_56, (gid, codigo, nombre, next_ord))
        mid = cur.lastrowid

        conn.commit()
        conn.close()

        return jsonify(ok=True, item={"id": mid, "nombre": nombre})

    @app.route('/reclamos/api/cantones', methods=['GET'], endpoint='reclamos_api_cantones')
    @require_login
    def reclamos_api_cantones():
        provincia_id = request.args.get('provincia_id', type=int)

        conn = get_db()
        #ensure_geo_schema(conn)
        rows = fetch_cantones(conn, provincia_id)
        conn.close()

        data = [dict(id=r['id'], nombre=r['nombre']) for r in rows]
        return jsonify(data)

    # ----------------------------------------------
    # CREAR RECLAMO
    # ----------------------------------------------
    @app.route('/reclamos/nuevo', methods=['POST'], endpoint='reclamos_nuevo')
    @require_login
    @require_permission('reclamos', 'crear')
    def reclamos_nuevo():
        print("DEBUG NUEVO RECLAMO:", dict(request.form))

        uid = _current_user_id()
        if not uid:
            flash("No se pudo determinar el usuario actual.", "danger")
            return redirect(url_for('reclamos'))

        conn = get_db()
        # ensure_reclamos_schema(conn)
        cur = conn.cursor()

        fecha_reclamo = (request.form.get('fecha_reclamo') or '').strip()
        cliente_id = (request.form.get('cliente_id') or '').strip() or None
        cliente_nombre = (request.form.get('cliente_nombre') or '').strip()
        cliente_identificacion = (request.form.get('cliente_identificacion') or '').strip()
        cliente_direccion = (request.form.get('cliente_direccion') or '').strip()
        cliente_contacto = (request.form.get('cliente_contacto') or '').strip()
        cliente_email = (request.form.get('cliente_email') or '').strip()
        cliente_telefono = (request.form.get('cliente_telefono') or '').strip()

        region_id = (request.form.get('region_id') or '').strip() or None
        provincia_id = (request.form.get('provincia_id') or '').strip() or None
        canton_id = (request.form.get('canton_id') or '').strip() or None

        tipo_tramite = (request.form.get('tipo_tramite') or '').strip()
        tipo_reclamo = (request.form.get('tipo_reclamo') or '').strip()

        proceso_ids_raw = request.form.getlist('proceso_id')
        proceso_ids = [int(v) for v in proceso_ids_raw if v.strip().isdigit()]
        proceso_id = proceso_ids[0] if proceso_ids else None

        proceso_text = (request.form.get('proceso_text') or '').strip()

        if proceso_ids:
            textos = []
            for pid in proceso_ids:
                cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_57, (pid,))
                row_proc = cur.fetchone()
                if row_proc and row_proc["valor"]:
                    textos.append(row_proc["valor"].strip())
            if textos:
                proceso_text = ', '.join(textos)

        antecedente = (request.form.get('antecedente') or '').strip()
        antecedente_raw = (request.form.get('antecedente') or '').strip()
        antecedente = antecedente_raw

        # ✅ Si viene un ID (del select), guardamos el texto (pv.valor)
        if antecedente_raw.isdigit():
            try:
                cur.execute("SELECT valor FROM param_values WHERE id = ?", (int(antecedente_raw),))
                row_sub = cur.fetchone()
                if row_sub and row_sub["valor"]:
                    antecedente = row_sub["valor"].strip()
            except Exception:
                pass

        fecha_pedido = (request.form.get('fecha_pedido') or '').strip()
        factura = (request.form.get('factura') or '').strip()
        guia_remision = (request.form.get('guia_remision') or '').strip()

        raw_material_id = (request.form.get('material_id') or '').strip()
        material_id = int(raw_material_id) if raw_material_id else None
        material_desc = ''

        if material_id:
            cur.execute("SELECT valor FROM param_values WHERE id = ?", (material_id,))
            row_mat = cur.fetchone()
            if row_mat:
                material_desc = row_mat["valor"] or ""

        persona_atendio = (request.form.get('persona_atendio') or '').strip()
        persona_atendio_cedula = (request.form.get('persona_atendio_cedula') or '').strip()
        fecha_ofrec_entrega = (request.form.get('fecha_ofrec_entrega') or '').strip()
        fecha_entrega = (request.form.get('fecha_entrega') or '').strip()

        observacion = (request.form.get('observacion') or '').strip()
        procede = (request.form.get('procede') or '').strip()
        requiere_carta_cliente = 1 if (request.form.get('requiere_carta_cliente') or '').strip() == '1' else 0
        codigo = _generate_codigo_reclamo(conn)
        archivos = request.files.getlist('adjuntos')

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_INS_58, (
            codigo,
            _now_iso(),
            _now_iso(),
            cliente_id, cliente_nombre, cliente_identificacion,
            cliente_direccion, cliente_contacto, cliente_email, cliente_telefono,
            region_id, provincia_id, canton_id,
            tipo_tramite, tipo_reclamo,
            proceso_id, proceso_text, antecedente,
            fecha_pedido, factura, guia_remision,
            material_id, material_desc,
            persona_atendio, persona_atendio_cedula,
            fecha_ofrec_entrega, fecha_entrega,
            observacion,
            procede,
            requiere_carta_cliente,
            None,
            uid,
            'abierto'
        ))
        row_new = cur.fetchone()
        reclamo_id = row_new["id"] if row_new and row_new["id"] is not None else None

        if reclamo_id is None:
            conn.rollback()
            flash("No se pudo recuperar el ID del reclamo creado.", "danger")
            return redirect(url_for('reclamos'))

        # =========================================================
        # IMPUTADOS / SPONSOR RESPONSABLE
        # =========================================================
        imputados_ids = request.form.getlist('imputados[]')

        # Lista separada SOLO para notificaciones.
        # Esto evita crear dos líneas, pero permite notificar principal + backup.
        sponsor_notify_ids = []

        if proceso_ids:
            seen_principals = set()
            seen_notify = set()
            collected_principals = []
            collected_notify = []

            for pid in proceso_ids:
                sp = _get_sponsor_principal_by_proceso(conn, pid)
                if sp and sp not in seen_principals:
                    seen_principals.add(sp)
                    collected_principals.append(str(sp))
                for nid in _get_sponsors_by_proceso(conn, pid):
                    if nid not in seen_notify:
                        seen_notify.add(nid)
                        collected_notify.append(nid)

            imputados_ids = collected_principals
            sponsor_notify_ids = collected_notify

        if not imputados_ids:
            conn.rollback()
            flash("Los procesos seleccionados no tienen sponsor principal configurado.", "danger")
            return redirect(url_for('reclamos'))

        responsable_username = None
        sponsor_principal_id_creado = None

        for raw_uid in imputados_ids:
            try:
                imputado_id = int(raw_uid)
            except Exception:
                continue

            # ✅ Ahora el aprobador ES el mismo imputado
            aprobador_id = imputado_id

            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_INS_59, (
                reclamo_id,
                imputado_id,
                aprobador_id,
                'aprobado',
                'sin_respuesta'
            ))

            imp_user = _get_user_basic(conn, imputado_id)

            if responsable_username is None:
                responsable_username = imp_user["username"] if imp_user else f"UID {imputado_id}"

            if sponsor_principal_id_creado is None:
                sponsor_principal_id_creado = imputado_id

            # IMPORTANTE:
            # Ya no enviamos el correo aquí, porque aquí solo está el principal.
            # El envío se hace abajo usando sponsor_notify_ids para incluir backup.

        # =========================================================
        # NOTIFICACIÓN PRINCIPAL + BACKUP
        # =========================================================
        # Si hay proceso, notificamos a los sponsors configurados:
        # PRINCIPAL + BACKUP.
        #
        # Si no hay proceso, conservamos el comportamiento anterior:
        # notificar a los imputados manuales.
        # =========================================================
        if proceso_id:
            ids_para_notificar = sponsor_notify_ids
        else:
            ids_para_notificar = imputados_ids

        # Datos completos de la OM para el payload del correo
        try:
            from modules.scheduler_jobs import enqueue_om_nueva_registro, ensure_om_evento_templates
            ensure_om_evento_templates(conn)

            _cur_om = conn.cursor()
            _cur_om.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_1, (codigo,))
            _r_om = _cur_om.fetchone()

            try:
                _cta_url = url_for("reclamos", _external=True) + "?tab=aprobar"
            except Exception:
                _cta_url = "http://bitacoraquimpac.com.ec:5000/reclamos?tab=aprobar"

            # Resumen de imputados para incluir en el correo
            _cur_om.execute(SQL__NOTIFY_APROBADOR_IMPUTACION_SEL_1, (reclamo_id,))
            _row_imp = _cur_om.fetchone()
            _imputados_txt = (_row_imp["lista"] if _row_imp and _row_imp["lista"]
                              else responsable_username or "")

            _base_payload = {
                "codigo":      codigo,
                "fecha_om":    str(_r_om["fecha_reclamo"]) if _r_om else "",
                "tipo_om":     str(_r_om["tipo_reclamo"])  if _r_om else "",
                "tipo_tramite":str(_r_om["tipo_tramite"])  if _r_om else "",
                "cliente":     str(_r_om["cliente_nombre"])if _r_om else "",
                "proceso":     str(_r_om["proceso_text"])  if _r_om else "",
                "material":    str(_r_om["material_desc"]) if (_r_om and "material_desc" in _r_om.keys()) else "",
                "factura":     str(_r_om["factura"])       if _r_om else "",
                "observacion": str(_r_om["observacion"])   if _r_om else "",
                "imputados":   _imputados_txt,
                "cta_url":     _cta_url,
            }
            _use_queue = True
        except Exception as _eq_err:
            current_app.logger.warning("[nueva_om] enqueue setup error: %s", _eq_err)
            _use_queue = False

        enviados = set()

        for raw_notify_id in ids_para_notificar:
            try:
                notify_user_id = int(raw_notify_id)
            except Exception:
                continue

            if notify_user_id in enviados:
                continue

            enviados.add(notify_user_id)

            if _use_queue:
                try:
                    dest = _get_user_basic(conn, notify_user_id)
                    dest_nombre = (
                        (dest.get("nombre_completo") or dest.get("username") or "")
                        if dest else ""
                    )
                    p = dict(_base_payload)
                    p["destinatario_nombre"] = dest_nombre
                    enqueue_om_nueva_registro(conn,
                        user_id=notify_user_id,
                        reclamo_id=reclamo_id,
                        payload=p,
                    )
                except Exception as _eq2:
                    current_app.logger.warning("[nueva_om] enqueue fallback uid=%s: %s", notify_user_id, _eq2)
                    _notify_aprobador_imputacion(conn, notify_user_id, codigo,
                        responsable_username or f"UID {sponsor_principal_id_creado or notify_user_id}")
            else:
                _notify_aprobador_imputacion(conn, notify_user_id, codigo,
                    responsable_username or f"UID {sponsor_principal_id_creado or notify_user_id}")

        # Guardar adjuntos (si los hay)
        error_adj = _save_adjuntos_for_reclamo(conn, reclamo_id, archivos, uid)
        if error_adj:
            # No bloqueamos la creación de la OM, solo avisamos
            flash(error_adj, "warning")

        conn.commit()
        flash("Reclamo creado correctamente.", "success")
        return redirect(url_for('reclamos'))
    
    
    
    # ----------------------------------------------
    # APROBACIÓN / RECHAZO IMPUTACIÓN
    # ----------------------------------------------
    @app.route('/reclamos/imputacion/<int:imp_id>/aprobar', methods=['POST'], endpoint='reclamos_aprobar_imputacion')
    @require_login
    @require_permission('reclamos', 'aprobar')
    def reclamos_aprobar_imputacion(imp_id):
        uid = _current_user_id()
        data = request.get_json(silent=True) or {}
        accion = (data.get("accion") or "").strip().lower()
        motivo = (data.get("motivo") or "").strip()

        conn = get_db()
        #ensure_reclamos_schema(conn)
        cur = conn.cursor()

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_60, (imp_id,))
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="Imputación no encontrada"), 404

        if (row["aprobador_id"] != uid) and (not _is_admin_like()):
            conn.close()
            return jsonify(ok=False, msg="No autorizado"), 403

        if row["estado_asignacion"] in ("aprobado", "rechazado"):
            conn.close()
            return jsonify(ok=False, msg="La imputación ya fue gestionada"), 400

        if accion == "aprobar":
            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_61, (_now_iso(), imp_id))

            try:
                from modules.scheduler_jobs import enqueue_om_nueva_asignado, ensure_om_evento_templates
                ensure_om_evento_templates(conn)

                dest = _get_user_basic(conn, row["imputado_uid"])
                dest_nombre = (
                    (dest.get("nombre_completo") or dest.get("username") or "")
                    if dest else ""
                )
                try:
                    cta_url = url_for("reclamos", _external=True) + "?tab=imputado"
                except Exception:
                    cta_url = "http://bitacoraquimpac.com.ec:5000/reclamos?tab=imputado"

                cur2 = conn.cursor()
                cur2.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_1, (row["codigo"],))
                r_om = cur2.fetchone()

                enqueue_om_nueva_asignado(conn,
                    user_id=row["imputado_uid"],
                    reclamo_id=row["reclamo_id"],
                    payload={
                        "codigo": row["codigo"],
                        "destinatario_nombre": dest_nombre,
                        "fecha_om":    str(r_om["fecha_reclamo"]) if r_om else "",
                        "tipo_om":     str(r_om["tipo_reclamo"])  if r_om else "",
                        "tipo_tramite":str(r_om["tipo_tramite"])  if r_om else "",
                        "cliente":     str(r_om["cliente_nombre"])if r_om else "",
                        "proceso":     str(r_om["proceso_text"])  if r_om else "",
                        "material":    str(r_om["material_desc"]) if (r_om and "material_desc" in r_om.keys()) else "",
                        "factura":     str(r_om["factura"])       if r_om else "",
                        "observacion": str(r_om["observacion"])   if r_om else "",
                        "cta_url": cta_url,
                    }
                )
            except Exception as _e:
                current_app.logger.warning("[aprobar_imputacion] enqueue fallback: %s", _e)
                _notify_imputado_aprobado(conn, row["imputado_uid"], row["codigo"])

        elif accion == "rechazar":
            if not motivo:
                conn.close()
                return jsonify(ok=False, msg="Motivo es obligatorio al rechazar"), 400

            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_62, (_now_iso(), motivo, imp_id))

            _notify_creador_rechazo_asignacion(
                conn,
                row["creado_por"],
                row["codigo"],
                row["imputado_username"],
                motivo
            )
        else:
            conn.close()
            return jsonify(ok=False, msg="Acción inválida"), 400

        conn.commit()
        conn.close()
        return jsonify(ok=True)



    @app.route('/reclamos/<int:reclamo_id>/adjuntos', methods=['POST'], endpoint='reclamos_add_adjuntos')
    @require_login
    @require_permission('reclamos', 'editar')
    def reclamos_add_adjuntos(reclamo_id):
        uid = _current_user_id()
        archivos = request.files.getlist('adjuntos')

        if not archivos:
            return jsonify(ok=False, msg="No se cargó ningún archivo."), 400

        conn = get_db()
        #ensure_reclamos_schema(conn)
        #ensure_reclamo_adjuntos_schema(conn)
        cur = conn.cursor()

        # Validar que exista la OM
        cur.execute("SELECT id FROM reclamos WHERE id = ?", (reclamo_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify(ok=False, msg="Oportunidad de Mejora no encontrada."), 404

        # Guardar adjuntos
        error_adj = _save_adjuntos_for_reclamo(conn, reclamo_id, archivos, uid)
        if error_adj:
            conn.rollback()
            conn.close()
            return jsonify(ok=False, msg=error_adj), 400

        # Obtener nombres originales de los archivos que se están subiendo,
        # limitado a lo que efectivamente se pudo guardar (mismo orden)
        filenames = [f.filename for f in archivos if f and f.filename]

        conn.commit()

        # Notificar al creador de la OM
        try:
            _notify_reclamo_adjuntos_change(conn, reclamo_id, uid, "agregados", filenames)
        except Exception:
            # No romper por fallo de correo
            current_app.logger.exception("Error enviando correo de adjuntos agregados")

        conn.close()
        return jsonify(ok=True, msg="Archivos adjuntados correctamente.")


    @app.route('/reclamos/adjunto/<int:adj_id>/delete',
            methods=['POST'],
            endpoint='reclamos_delete_adjunto')
    @require_login
    @require_permission('reclamos', 'editar')
    def reclamos_delete_adjunto(adj_id):
        uid = _current_user_id()

        conn = get_db()
        #ensure_reclamo_adjuntos_schema(conn)
        cur = conn.cursor()

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_63, (adj_id,))
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="Adjunto no encontrado"), 404

        upload_folder = _get_reclamos_upload_folder()
        path = os.path.join(upload_folder, row["filename"])

        # Borra registro de BD
        cur.execute("DELETE FROM reclamo_adjuntos WHERE id = ?", (adj_id,))
        conn.commit()
        conn.close()

        # Intenta borrar archivo físico (si existe)
        try:
            if os.path.exists(path):
                os.remove(path)
        except Exception as e:
            current_app.logger.warning(
                "No se pudo eliminar archivo físico %s: %s", path, e
            )

        return jsonify(ok=True, msg="Adjunto eliminado correctamente.")

 
    @app.route('/reclamos/api/<int:reclamo_id>/adjuntos', methods=['GET'], endpoint='reclamos_api_adjuntos')
    @require_login
    def reclamos_api_adjuntos(reclamo_id):
        conn = get_db()
        cur = conn.cursor()

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_64, (reclamo_id,))

        rows = []
        for r in cur.fetchall():
            rows.append({
                "id": r["id"],
                "original_name": r["original_name"] or "",
                "content_type": r["content_type"] or "",
                "size_bytes": int(r["size_bytes"] or 0),
                "created_at": r["created_at"] or "",
                "creado_por": r["creado_por"],
                "cargado_por": r["cargado_por"] or ""
            })

        return jsonify(ok=True, items=rows)   
        
    
    @app.route('/reclamos/adjunto/<int:adj_id>/download', methods=['GET'], endpoint='reclamos_download_adjunto')
    @require_login
    def reclamos_download_adjunto(adj_id):
        conn = get_db()
        #ensure_reclamo_adjuntos_schema(conn)
        cur = conn.cursor()

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_65, (adj_id,))
        row = cur.fetchone()
        conn.close()

        if not row:
            abort(404)

        upload_folder = _get_reclamos_upload_folder()
        path = os.path.join(upload_folder, row["filename"])

        if not os.path.exists(path):
            abort(404)

        return send_file(
            path,
            as_attachment=True,
            download_name=row["original_name"],
            mimetype=row["content_type"] or "application/octet-stream"
        )

  

    def _ensure_reclamo_imputados_extra_cols(conn):
        """
        No-op en SQL Server.
        Las columnas extra de reclamo_imputados ya deben existir.
        """
        return
 
    # ----------------------------------------------
    # RESPUESTA TÉCNICA IMPUTADO (AUTO-APROBADA + AUTO-CIERRE)
    # ----------------------------------------------
    @app.route(
        '/reclamos/imputacion/<int:imp_id>/responder',
        methods=['POST'],
        endpoint='reclamos_responder_imputado'
    )
    @require_login
    @require_permission('reclamos', 'editar')
    def reclamos_responder_imputado(imp_id):
        uid = _current_user_id()
        data = request.get_json(silent=True) or {}

        def _safe_str(v):
            return (v or "").strip() if isinstance(v, str) else ""

        metodo = _safe_str(data.get("metodo_analisis") or "FISHBONE").upper()

        # ===== NUEVO FORMATO =====
        causas_items = data.get("causas") or []
        control_items = data.get("control") or []
        correctiva_items = data.get("correctiva") or []

        if not isinstance(causas_items, list):
            causas_items = []
        if not isinstance(control_items, list):
            control_items = []
        if not isinstance(correctiva_items, list):
            correctiva_items = []

        # ===== COMPATIBILIDAD FORMATO VIEJO =====
        causa_legacy = _safe_str(data.get("causa"))
        preventiva_legacy = _safe_str(data.get("preventiva"))
        correctiva_legacy = _safe_str(data.get("correctiva"))

        fecha_causa_legacy = _safe_str(data.get("fecha_causa"))
        fecha_preventiva_legacy = _safe_str(data.get("fecha_preventiva"))
        fecha_correctiva_legacy = _safe_str(data.get("fecha_correctiva"))

        if not causas_items and causa_legacy:
            causas_items = [{
                "descripcion": causa_legacy,
                "fecha_compromiso": fecha_causa_legacy
            }]

        if not control_items and preventiva_legacy:
            control_items = [{
                "descripcion": preventiva_legacy,
                "fecha_compromiso": fecha_preventiva_legacy
            }]

        if not correctiva_items and correctiva_legacy:
            correctiva_items = [{
                "descripcion": correctiva_legacy,
                "fecha_compromiso": fecha_correctiva_legacy
            }]

        # ===== 5 porqués =====
        why1 = _safe_str(data.get("why1"))
        why2 = _safe_str(data.get("why2"))
        why3 = _safe_str(data.get("why3"))
        why4 = _safe_str(data.get("why4"))
        why5 = _safe_str(data.get("why5"))

        # ===== Fishbone =====
        fish_metodo     = _safe_str(data.get("fish_metodo"))
        fish_maquinas   = _safe_str(data.get("fish_maquinas"))
        fish_materiales = _safe_str(data.get("fish_materiales"))
        fish_personas   = _safe_str(data.get("fish_personas"))
        fish_entorno    = _safe_str(data.get("fish_entorno"))
        fish_medicion   = _safe_str(data.get("fish_medicion"))

        conn = get_db()
        #ensure_reclamos_schema(conn)
        _ensure_reclamo_imputados_extra_cols(conn)
        #ensure_reclamo_imputado_acciones_schema(conn)

        cur = conn.cursor()

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_66, (imp_id,))
        
        
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="Imputación no encontrada"), 404

        es_imputado_principal = int(row["imputado_id"] or 0) == int(uid or 0)

        es_sponsor_backup_o_principal = _usuario_es_sponsor_del_proceso(
            conn,
            row["proceso_id"],
            uid
        )

        if not (es_imputado_principal or es_sponsor_backup_o_principal or _is_admin_like()):
            conn.close()
            return jsonify(ok=False, msg="No autorizado"), 403

        if row["estado_asignacion"] != "aprobado":
            conn.close()
            return jsonify(ok=False, msg="No puedes responder: imputación aún no aprobada"), 400

        # ===== VALIDACIONES =====
        if not causas_items or not control_items or not correctiva_items:
            conn.close()
            return jsonify(
                ok=False,
                msg="Debe ingresar al menos una causa, una acción de control y una correctiva"
            ), 400

        if metodo == "5WHYS":
            if not (why1 and why2 and why3 and why4 and why5):
                conn.close()
                return jsonify(
                    ok=False,
                    msg="Debe completar los 5 porqués cuando el método es '5 porqués'."
                ), 400

        now_iso = _now_iso()
        reclamo_id = int(row["reclamo_id"])

        # ===== ARMAR RESUMEN LEGACY =====
        def _join_items(items):
            vals = []
            for x in items:
                desc = _safe_str((x or {}).get("descripcion"))
                fecha = _safe_str((x or {}).get("fecha_compromiso"))
                if desc and fecha:
                    vals.append(f"{desc} ({fecha})")
                elif desc:
                    vals.append(desc)
                elif fecha:
                    vals.append(f"({fecha})")
            return " | ".join(vals)

        respuesta_causa = _join_items(causas_items)
        respuesta_preventiva = _join_items(control_items)
        respuesta_correctiva = _join_items(correctiva_items)

        # primera fecha visible legacy
        fecha_causa = _safe_str((causas_items[0] or {}).get("fecha_compromiso")) if causas_items else ""
        fecha_preventiva = _safe_str((control_items[0] or {}).get("fecha_compromiso")) if control_items else ""
        fecha_correctiva = _safe_str((correctiva_items[0] or {}).get("fecha_compromiso")) if correctiva_items else ""

        # ===== GUARDAR CABECERA OFICIAL =====
        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_67, (
            metodo,
            respuesta_causa, fecha_causa,
            respuesta_preventiva, fecha_preventiva,
            respuesta_correctiva, fecha_correctiva,
            why1, why2, why3, why4, why5,
            fish_metodo, fish_maquinas, fish_materiales, fish_personas, fish_entorno, fish_medicion,
            now_iso,
            now_iso,
            imp_id
        ))

        # ===== GUARDAR DETALLE EN TABLA HIJA =====
        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_68, (now_iso, uid, imp_id, reclamo_id))

        def _insert_items(tipo, items, requiere_evidencia_default=0):
            orden = 1
            for x in items:
                descripcion = _safe_str((x or {}).get("descripcion"))
                fecha_compromiso = _safe_str((x or {}).get("fecha_compromiso"))

                if not descripcion and not fecha_compromiso:
                    continue

                cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_INS_69, (
                    imp_id,
                    reclamo_id,
                    tipo,
                    descripcion,
                    fecha_compromiso,
                    orden,
                    int((x or {}).get("requiere_evidencia", requiere_evidencia_default) or 0),
                    now_iso, uid,
                    now_iso, uid
                ))
                orden += 1

        _insert_items("CAUSA", causas_items, 0)
        _insert_items("CONTROL", control_items, 0)
        _insert_items("CORRECTIVA", correctiva_items, 1)

        # ===== CIERRE DE OM =====
        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_70, (reclamo_id,))
        rowp = cur.fetchone()
        pend = int((rowp["pend"] if rowp else 0) or 0)

        if pend == 0:
            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_71, (reclamo_id,))

        try:
            _notify_creador_respuesta_aprobada(
                conn,
                row["creado_por"],
                row["codigo"],
                row["imputado_username"]
            )
        except Exception as e:
            current_app.logger.warning("[responder_imputado] no se pudo notificar al creador: %s", e)

        conn.commit()
        conn.close()
        return jsonify(ok=True)
    
    
    @app.route('/reclamos/<int:reclamo_id>/eliminar', methods=['POST'], endpoint='reclamos_eliminar')
    @require_login
    def reclamos_eliminar(reclamo_id):
        conn = get_db()
        #ensure_reclamos_schema(conn)
        #ensure_reclamo_adjuntos_schema(conn)
        #ensure_reclamo_respuestas_equipo_schema(conn)

        uid = session.get("usuario_id") or session.get("user_id") or session.get("id")
        role = (session.get('rol') or '').strip().lower()

        # ✅ Permitido:
        # - admin
        # - personal de Servicio al Cliente
        autorizado = (role == 'admin') or _can_view_all_reclamos(conn, uid)

        if not autorizado:
            conn.close()
            return jsonify(ok=False, msg='No autorizado'), 403

        cur = conn.cursor()

        # Validar existencia + estado global
        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_72, (reclamo_id,))
        r = cur.fetchone()

        if not r:
            conn.close()
            return jsonify(ok=False, msg='OM no encontrada'), 404

        # ✅ No permitir eliminar si está cerrada
        if (r["estado_global"] or "") == "cerrado":
            conn.close()
            return jsonify(ok=False, msg='No se puede eliminar una OM con estado global Cerrado'), 400

        # Adjuntos físicos (si existen)
        cur.execute("SELECT filename FROM reclamo_adjuntos WHERE reclamo_id = ?", (reclamo_id,))
        files = [row["filename"] for row in cur.fetchall() if row and row["filename"]]

        # Borrado en cascada (manual)
        try:
            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_DEL_73, (reclamo_id,))
        except sqlite3.OperationalError:
            pass

        cur.execute("DELETE FROM reclamo_respuestas_equipo WHERE reclamo_id = ?", (reclamo_id,))
        cur.execute("DELETE FROM reclamo_equipo_respuestas WHERE reclamo_id = ?", (reclamo_id,))
        cur.execute("DELETE FROM reclamo_imputados WHERE reclamo_id = ?", (reclamo_id,))
        cur.execute("DELETE FROM reclamo_adjuntos WHERE reclamo_id = ?", (reclamo_id,))
        cur.execute("DELETE FROM reclamos WHERE id = ?", (reclamo_id,))

        conn.commit()
        conn.close()

        # Borrar archivos físicos
        upload_folder = _get_reclamos_upload_folder()
        for fn in files:
            try:
                path = os.path.join(upload_folder, fn)
                if os.path.exists(path):
                    os.remove(path)
            except Exception:
                current_app.logger.exception("No se pudo eliminar adjunto físico: %s", fn)

        return jsonify(ok=True, msg=f"OM {r['codigo']} eliminada correctamente.")
        
        
    # ----------------------------------------------
    # VALIDAR RESPUESTA (JEFE)
    # ----------------------------------------------
    @app.route('/reclamos/imputacion/<int:imp_id>/validar_respuesta', methods=['POST'], endpoint='reclamos_validar_respuesta')
    @require_login
    @require_permission('reclamos', 'aprobar')
    def reclamos_validar_respuesta(imp_id):
        uid = _current_user_id()
        data = request.get_json(silent=True) or {}
        accion = (data.get("accion") or "").strip().lower()
        motivo = (data.get("motivo") or "").strip()

        conn = get_db()
        #ensure_reclamos_schema(conn)
        cur = conn.cursor()

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_60, (imp_id,))
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="Registro no encontrado"), 404

        if (row["aprobador_id"] != uid) and (not _is_admin_like()):
            conn.close()
            return jsonify(ok=False, msg="No autorizado"), 403

        if row["estado_asignacion"] != "aprobado":
            conn.close()
            return jsonify(ok=False, msg="Imputación aún no aprobada, no se valida respuesta"), 400

        if row["estado_respuesta"] not in ("pendiente_jefe", "rechazada"):
            conn.close()
            return jsonify(ok=False, msg="No hay respuesta pendiente de validar"), 400

        if accion == "aprobar":
            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_74, (_now_iso(), imp_id))

            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_71, (row["reclamo_id"],))
            
            _notify_creador_respuesta_aprobada(
                conn,
                row["creado_por"],
                row["codigo"],
                row["imputado_username"]
            )

        elif accion == "rechazar":
            if not motivo:
                conn.close()
                return jsonify(ok=False, msg="Motivo obligatorio al rechazar"), 400

            cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_75, (_now_iso(), motivo, imp_id))

            _notify_imputado_respuesta_rechazada(
                conn,
                row["imputado_uid"],
                row["codigo"],
                motivo
            )
        else:
            conn.close()
            return jsonify(ok=False, msg="Acción inválida"), 400

        conn.commit()
        conn.close()
        return jsonify(ok=True)



    from flask import jsonify
    from datetime import datetime

    # =========================
    # 1) Resumen Ejecutivo
    # =========================
    @app.route("/api/dashboard/resumen")
    @require_login
    def reclamos_api_dashboard_resumen():
        db = get_db()
        row = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_76).fetchone()

        # OM creadas por mes (últimos 6 meses)
        rows_mes = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_77).fetchall()

        return jsonify({
            "kpi": dict(row or {}),
            "serie_mes": [dict(r) for r in rows_mes]
        })


    # =========================
    # 2) Flujo de estados
    # =========================
    @app.route("/api/dashboard/estados")
    @require_login
    def reclamos_api_dashboard_estados():
        db = get_db()
        rows = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_78).fetchall()
        return jsonify({"items": [dict(r) for r in rows]})


    # =========================
    # 3) SLA / tiempos de ciclo
    #    (ajusta 'fecha_cierre' al nombre real)
    # =========================
 
    @app.route("/api/dashboard/sla")
    @require_login
    def reclamos_api_dashboard_sla():
        db = get_db()

        if _is_sqlserver_conn(db):
            rows = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_79).fetchall()
        else:
            rows = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_80).fetchall()

        return jsonify({"items": [dict(r) for r in rows]})


 
    # =========================
    # 4) OM por sponsor / imputado
    # =========================
    @app.route("/api/dashboard/imputados")
    @require_login
    def reclamos_api_dashboard_imputados():
        db = get_db()

        if _is_sqlserver_conn(db):
            rows = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_81).fetchall()
        else:
            rows = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_82).fetchall()

        return jsonify({"items": [dict(r) for r in rows]})

    # =========================
    # 5) Clientes / procesos
    # =========================
   

    @app.route("/api/dashboard/clientes_procesos")
    @require_login
    def reclamos_api_dashboard_clientes_procesos():
        db = get_db()

        if _is_sqlserver_conn(db):
            top_clientes = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_83).fetchall()
        else:
            top_clientes = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_84).fetchall()

        por_proceso = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_85).fetchall()

        return jsonify({
            "clientes": [dict(r) for r in top_clientes],
            "procesos": [dict(r) for r in por_proceso]
        })




    # =========================
    # 6) OM por Departamento  (ya lo tenías, lo dejo completo)
    # =========================
    @app.route("/api/dashboard/departamentos")
    @require_login
    def reclamos_api_dashboard_departamentos():
        db = get_db()
        rows = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_86).fetchall()

        return jsonify({"items": [dict(row) for row in rows]})



    # routes_reclamos.py (solo la parte de la ruta completa)
    @app.route('/reclamos/dashboard', methods=['GET'], endpoint='reclamos_dashboard')
    @require_login
    @require_permission('reclamos', 'ver')
    def reclamos_dashboard():
        conn = get_db()
        cur = conn.cursor()
        is_sqlserver = _is_sqlserver_conn(conn)

        fecha_desde = (request.args.get("desde") or "").strip()
        fecha_hasta = (request.args.get("hasta") or "").strip()
        depto_sel   = (request.args.get("depto") or "").strip()
        proceso_sel = (request.args.get("proceso") or "").strip()

        where = []
        params = []

        if fecha_desde:
            if is_sqlserver:
                where.append(
                    "CAST(TRY_CONVERT(date, COALESCE(r.fecha_reclamo, r.fecha_creacion)) AS date) >= CAST(? AS date)"
                )
            else:
                where.append("date(COALESCE(r.fecha_reclamo, r.fecha_creacion)) >= date(?)")
            params.append(fecha_desde)

        if fecha_hasta:
            if is_sqlserver:
                where.append(
                    "CAST(TRY_CONVERT(date, COALESCE(r.fecha_reclamo, r.fecha_creacion)) AS date) <= CAST(? AS date)"
                )
            else:
                where.append("date(COALESCE(r.fecha_reclamo, r.fecha_creacion)) <= date(?)")
            params.append(fecha_hasta)

        if depto_sel:
            where.append("COALESCE(d.nombre, 'SIN DEPARTAMENTO') = ?")
            params.append(depto_sel)

        if proceso_sel:
            if is_sqlserver:
                where.append("""
                    (
                        ? = 'SIN PROCESO'
                        AND COALESCE(NULLIF(LTRIM(RTRIM(r.proceso_text)), ''), 'SIN PROCESO') = 'SIN PROCESO'
                    )
                    OR
                    (
                        ? <> 'SIN PROCESO'
                        AND (
                            '|' + REPLACE(REPLACE(REPLACE(UPPER(COALESCE(r.proceso_text, '')), ',', '|'), ';', '|'), ' | ', '|') + '|'
                        ) LIKE ?
                    )
                """)
            else:
                where.append("""
                    (
                        ? = 'SIN PROCESO'
                        AND COALESCE(NULLIF(TRIM(r.proceso_text), ''), 'SIN PROCESO') = 'SIN PROCESO'
                    )
                    OR
                    (
                        ? <> 'SIN PROCESO'
                        AND (
                            '|' || REPLACE(REPLACE(REPLACE(UPPER(COALESCE(r.proceso_text, '')), ',', '|'), ';', '|'), ' | ', '|') || '|'
                        ) LIKE ?
                    )
                """)

            params.extend([
                proceso_sel,
                proceso_sel,
                f"%|{proceso_sel.upper()}|%"
            ])

        where_sql = "WHERE " + " AND ".join(where) if where else ""

        # ========= KPIs =========
        cur.execute(f"""
            SELECT
                COUNT(DISTINCT r.id) AS total,

                SUM(
                    CASE
                        WHEN LOWER(COALESCE(r.estado_global,'')) = 'cerrado'
                            OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                        THEN 1 ELSE 0
                    END
                ) AS cerradas,

                SUM(
                    CASE
                        WHEN LOWER(COALESCE(r.estado_global,'')) LIKE 'pendiente respuesta del imputado%'
                        THEN 1 ELSE 0
                    END
                ) AS pend_imputados,

                SUM(
                    CASE
                        WHEN LOWER(COALESCE(r.estado_global,'')) LIKE 'respuesta pendiente de aprobación%'
                        OR LOWER(COALESCE(r.estado_global,'')) LIKE 'pendiente respuesta del sponsor%'
                        OR LOWER(COALESCE(r.estado_global,'')) LIKE 'pendiente respuesta del responsable%'
                        THEN 1 ELSE 0
                    END
                ) AS pend_sponsor
            FROM reclamos r
            LEFT JOIN usuarios u ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            {where_sql}
        """, params)

        row = cur.fetchone()
        total = int(row["total"] or 0) if row else 0
        cerradas = int(row["cerradas"] or 0) if row else 0
        pend_imputados = int(row["pend_imputados"] or 0) if row else 0
        pend_sponsor = int(row["pend_sponsor"] or 0) if row else 0

        kpis = {
            "total": total,
            "pend_jefe": pend_sponsor,
            "pend_imputados": pend_imputados,
            "cerradas": cerradas,
            "abiertas": max(total - cerradas, 0),
        }

        # ========= CHART: estados =========
        cur.execute(f"""
            SELECT
                COALESCE(r.estado_global, 'SIN ESTADO') AS estado,
                COUNT(*) AS total
            FROM reclamos r
            LEFT JOIN usuarios u ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            {where_sql}
            GROUP BY COALESCE(r.estado_global, 'SIN ESTADO')
            ORDER BY total DESC
        """, params)
        rows = cur.fetchall()
        chart_estados = {
            "labels": [r["estado"] for r in rows],
            "total": [int(r["total"] or 0) for r in rows]
        }

        # ========= CHART: meses =========
        if is_sqlserver:
            cur.execute(f"""
                SELECT
                    CONVERT(VARCHAR(7), TRY_CONVERT(datetime, r.fecha_creacion), 120) AS ym,
                    COUNT(*) AS total
                FROM reclamos r
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                GROUP BY CONVERT(VARCHAR(7), TRY_CONVERT(datetime, r.fecha_creacion), 120)
                ORDER BY ym
            """, params)
        else:
            cur.execute(f"""
                SELECT
                    strftime('%Y-%m', r.fecha_creacion) AS ym,
                    COUNT(*) AS total
                FROM reclamos r
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                GROUP BY strftime('%Y-%m', r.fecha_creacion)
                ORDER BY ym
            """, params)

        rows = cur.fetchall()
        chart_meses = {
            "labels": [r["ym"] for r in rows if r["ym"]],
            "total": [int(r["total"] or 0) for r in rows if r["ym"]]
        }

        # ========= CHART: días =========
        if is_sqlserver:
            cur.execute(f"""
                SELECT
                    CAST(TRY_CONVERT(date, r.fecha_creacion) AS date) AS dia,
                    COUNT(*) AS total
                FROM reclamos r
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                GROUP BY CAST(TRY_CONVERT(date, r.fecha_creacion) AS date)
                ORDER BY CAST(TRY_CONVERT(date, r.fecha_creacion) AS date)
            """, params)
        else:
            cur.execute(f"""
                SELECT
                    date(r.fecha_creacion) AS dia,
                    COUNT(*) AS total
                FROM reclamos r
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                GROUP BY date(r.fecha_creacion)
                ORDER BY date(r.fecha_creacion)
            """, params)

        rows = cur.fetchall()
        chart_dias = {
            "labels": [str(r["dia"]) for r in rows if r["dia"]],
            "total": [int(r["total"] or 0) for r in rows if r["dia"]],
        }

        # ========= CHART: imputados (creador como proxy) =========
        if is_sqlserver:
            cur.execute(f"""
                SELECT TOP 10
                    COALESCE(u.nombre_completo, u.username, 'SIN USUARIO') AS imputado,
                    SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado'
                            OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 1 ELSE 0 END) AS cerradas,
                    SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado'
                            OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 0 ELSE 1 END) AS abiertas
                FROM reclamos r
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                GROUP BY COALESCE(u.nombre_completo, u.username, 'SIN USUARIO')
                ORDER BY (
                    SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado'
                            OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 0 ELSE 1 END)
                    +
                    SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado'
                            OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 1 ELSE 0 END)
                ) DESC
            """, params)
        else:
            cur.execute(f"""
                SELECT TOP 10
                    COALESCE(u.nombre_completo, u.username, 'SIN USUARIO') AS imputado,
                    SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado' OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 1 ELSE 0 END) AS cerradas,
                    SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado' OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 0 ELSE 1 END) AS abiertas
                FROM reclamos r
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                GROUP BY COALESCE(u.nombre_completo, u.username, 'SIN USUARIO')
                ORDER BY (abiertas + cerradas) DESC
                 
            """, params)

        rows = cur.fetchall()
        chart_imputados = {
            "labels":   [r["imputado"] for r in rows],
            "abiertas": [int(r["abiertas"] or 0) for r in rows],
            "cerradas": [int(r["cerradas"] or 0) for r in rows],
        }
        # ========= CHART: procesos involucrados =========
        # Cuenta participación por proceso. Si una OM tiene varios procesos en proceso_text,
        # la OM cuenta una vez en cada proceso involucrado.
        from collections import defaultdict
        import re

        cur.execute(f"""
            SELECT
                r.id,
                COALESCE(r.estado_global, '') AS estado_global,
                COALESCE(r.proceso_text, '') AS proceso_text
            FROM reclamos r
            LEFT JOIN usuarios u ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            {where_sql}
        """, params)

        rows_part = cur.fetchall()

        conteo_proc = defaultdict(lambda: {"abiertas": 0, "cerradas": 0, "total": 0})

        def _split_procesos_involucrados(txt):
            txt = (txt or "").strip()
            if not txt:
                return ["SIN PROCESO"]

            # Soporta separadores usados normalmente cuando una OM comparte procesos:
            # COMERCIAL | LOGISTICA
            # COMERCIAL, LOGISTICA
            # COMERCIAL; LOGISTICA
            parts = re.split(r"\s*(?:\||;|,)\s*", txt)

            # Evita duplicar el mismo proceso dentro de la misma OM
            clean = []
            seen = set()

            for p in parts:
                p = (p or "").strip()
                if not p:
                    continue

                key = p.upper()
                if key not in seen:
                    seen.add(key)
                    clean.append(p)

            return clean or ["SIN PROCESO"]

        def _es_cerrada(estado):
            estado = (estado or "").strip().lower()
            return estado == "cerrado" or estado.startswith("cerr")

        for r in rows_part:
            estado = r["estado_global"] if hasattr(r, "keys") else r[1]
            procesos_txt = r["proceso_text"] if hasattr(r, "keys") else r[2]

            cerrada = _es_cerrada(estado)

            for proceso in _split_procesos_involucrados(procesos_txt):
                conteo_proc[proceso]["total"] += 1

                if cerrada:
                    conteo_proc[proceso]["cerradas"] += 1
                else:
                    conteo_proc[proceso]["abiertas"] += 1

        items_proc = sorted(
            conteo_proc.items(),
            key=lambda x: x[1]["total"],
            reverse=True
        )[:15]

        chart_procesos_involucrados = {
            "labels": [p for p, v in items_proc],
            "abiertas": [int(v["abiertas"] or 0) for p, v in items_proc],
            "cerradas": [int(v["cerradas"] or 0) for p, v in items_proc],
            "total": [int(v["total"] or 0) for p, v in items_proc],

            # OM únicas después de aplicar los filtros actuales.
            # Sirve para diferenciar OM reales vs participaciones por proceso.
            "om_unicas": len(rows_part),
        }

        # ========= CHART: procesos =========
        proceso_expr = "COALESCE(NULLIF(LTRIM(RTRIM(r.proceso_text)), ''), 'SIN PROCESO')" if is_sqlserver else "COALESCE(NULLIF(TRIM(r.proceso_text), ''), 'SIN PROCESO')"

        if is_sqlserver:
            cur.execute(f"""
                SELECT TOP 10
                    {proceso_expr} AS proceso,
                    SUM(CASE
                            WHEN LOWER(COALESCE(r.estado_global,''))='cerrado'
                                OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 1 ELSE 0
                        END) AS cerradas,
                    SUM(CASE
                            WHEN LOWER(COALESCE(r.estado_global,''))='cerrado'
                                OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 0 ELSE 1
                        END) AS abiertas,
                    COUNT(*) AS total
                FROM reclamos r
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                GROUP BY {proceso_expr}
                ORDER BY COUNT(*) DESC
            """, params)
        else:
            cur.execute(f"""
                SELECT TOP 10
                    {proceso_expr} AS proceso,
                    SUM(CASE
                            WHEN LOWER(COALESCE(r.estado_global,''))='cerrado'
                                OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 1 ELSE 0
                        END) AS cerradas,
                    SUM(CASE
                            WHEN LOWER(COALESCE(r.estado_global,''))='cerrado'
                                OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 0 ELSE 1
                        END) AS abiertas,
                    COUNT(*) AS total
                FROM reclamos r
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                GROUP BY {proceso_expr}
                ORDER BY total DESC
                
            """, params)

        rows = cur.fetchall()
        chart_procesos = {
            "labels":   [r["proceso"] for r in rows],
            "abiertas": [int(r["abiertas"] or 0) for r in rows],
            "cerradas": [int(r["cerradas"] or 0) for r in rows],
            "total":    [int(r["total"] or 0) for r in rows],
        }

        # ========= CHART: tipos =========
        if is_sqlserver:
            cur.execute(f"""
                SELECT TOP 10
                    COALESCE(r.tipo_reclamo, 'SIN TIPO') AS tipo,
                    SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado'
                            OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 1 ELSE 0 END) AS cerradas,
                    SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado'
                            OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 0 ELSE 1 END) AS abiertas
                FROM reclamos r
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                GROUP BY COALESCE(r.tipo_reclamo, 'SIN TIPO')
                ORDER BY (
                    SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado'
                            OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 0 ELSE 1 END)
                    +
                    SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado'
                            OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 1 ELSE 0 END)
                ) DESC
            """, params)
        else:
            cur.execute(f"""
                SELECT TOP 10
                    COALESCE(r.tipo_reclamo, 'SIN TIPO') AS tipo,
                    SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado' OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 1 ELSE 0 END) AS cerradas,
                    SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado' OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                            THEN 0 ELSE 1 END) AS abiertas
                FROM reclamos r
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                GROUP BY COALESCE(r.tipo_reclamo, 'SIN TIPO')
                ORDER BY (abiertas + cerradas) DESC
                
            """, params)

        rows = cur.fetchall()
        chart_tipos = {
            "labels":   [r["tipo"] for r in rows],
            "abiertas": [int(r["abiertas"] or 0) for r in rows],
            "cerradas": [int(r["cerradas"] or 0) for r in rows],
        }

        # ========= CHART: tiempo promedio de respuesta por proceso =========
        if is_sqlserver:
            cur.execute(f"""
                SELECT TOP 12
                    {proceso_expr} AS proceso,
                    AVG(
                        CASE
                            WHEN ri.fecha_respuesta_imputado IS NOT NULL
                            AND TRY_CONVERT(datetime, COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion)) IS NOT NULL
                            AND TRY_CONVERT(datetime, ri.fecha_respuesta_imputado) IS NOT NULL
                            THEN CAST(DATEDIFF(
                                DAY,
                                TRY_CONVERT(datetime, COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion)),
                                TRY_CONVERT(datetime, ri.fecha_respuesta_imputado)
                            ) AS FLOAT)
                        END
                    ) AS dias_promedio_respuesta,
                    COUNT(DISTINCT r.id) AS total_om
                FROM reclamos r
                LEFT JOIN reclamo_imputados ri ON ri.reclamo_id = r.id
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                GROUP BY {proceso_expr}
                HAVING AVG(
                    CASE
                        WHEN ri.fecha_respuesta_imputado IS NOT NULL
                        AND TRY_CONVERT(datetime, COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion)) IS NOT NULL
                        AND TRY_CONVERT(datetime, ri.fecha_respuesta_imputado) IS NOT NULL
                        THEN CAST(DATEDIFF(
                            DAY,
                            TRY_CONVERT(datetime, COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion)),
                            TRY_CONVERT(datetime, ri.fecha_respuesta_imputado)
                        ) AS FLOAT)
                    END
                ) IS NOT NULL
                ORDER BY dias_promedio_respuesta DESC
            """, params)
        else:
            cur.execute(f"""
                SELECT TOP 12
                    {proceso_expr} AS proceso,
                    AVG(
                        CASE
                            WHEN ri.fecha_respuesta_imputado IS NOT NULL
                            THEN julianday(ri.fecha_respuesta_imputado)
                                - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion))
                        END
                    ) AS dias_promedio_respuesta,
                    COUNT(DISTINCT r.id) AS total_om
                FROM reclamos r
                LEFT JOIN reclamo_imputados ri ON ri.reclamo_id = r.id
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                GROUP BY {proceso_expr}
                HAVING AVG(
                    CASE
                        WHEN ri.fecha_respuesta_imputado IS NOT NULL
                        THEN julianday(ri.fecha_respuesta_imputado)
                            - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion))
                    END
                ) IS NOT NULL
                ORDER BY dias_promedio_respuesta DESC
             """, params)

        rows = cur.fetchall()
        chart_tiempos = {
            "labels": [r["proceso"] for r in rows],
            "promedio": [round(float(r["dias_promedio_respuesta"] or 0), 2) for r in rows],
            "total_om": [int(r["total_om"] or 0) for r in rows],
        }

        
        # ========= CHART: top procesos con OM vencidas =========
        extra_where = (" AND " if where_sql else "WHERE ")
        if is_sqlserver:
            cur.execute(f"""
                WITH base AS (
                    SELECT
                        r.id,
                        r.codigo,
                        COALESCE(NULLIF(LTRIM(RTRIM(r.proceso_text)), ''), 'SIN PROCESO') AS proceso,
                        ri.id AS imputacion_id,
                        ri.estado_asignacion,
                        ri.estado_respuesta,
                        ri.fecha_aprobacion_asignacion,
                        r.fecha_creacion,
                        CASE
                            WHEN ri.estado_asignacion = 'pend_aprobacion' THEN 'Pendiente aceptación del responsable'
                            WHEN ri.estado_asignacion = 'rechazado' THEN 'Imputación rechazada'
                            WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'sin_respuesta' THEN 'Pendiente respuesta del imputado'
                            WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'pendiente_jefe' THEN 'Respuesta pendiente de aprobación'
                            WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'aprobada' THEN 'Cerrado'
                            WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'rechazada' THEN 'Respuesta rechazada'
                            ELSE COALESCE(ri.estado_asignacion, '') + '/' + COALESCE(ri.estado_respuesta, '')
                        END AS estado_imputacion
                    FROM reclamos r
                    INNER JOIN reclamo_imputados ri
                        ON ri.reclamo_id = r.id
                    LEFT JOIN usuarios u ON u.id = r.creado_por
                    LEFT JOIN departamentos d ON d.id = u.departamento_id
                    {where_sql}
                )
                SELECT TOP 10
                    proceso,
                    COUNT(DISTINCT id) AS vencidas
                FROM base
                WHERE estado_imputacion = 'Pendiente respuesta del imputado'
                AND DATEDIFF(
                        DAY,
                        COALESCE(
                            TRY_CONVERT(datetime, fecha_aprobacion_asignacion, 120),
                            TRY_CONVERT(datetime, fecha_creacion, 120)
                        ),
                        GETDATE()
                    ) > 5
                GROUP BY proceso
                ORDER BY vencidas DESC, proceso ASC
            """, params)
        else:
            cur.execute(f"""
                SELECT TOP 10
                    {proceso_expr} AS proceso,
                    COUNT(DISTINCT r.id) AS vencidas
                FROM reclamos r
                INNER JOIN reclamo_imputados ri
                    ON ri.reclamo_id = r.id
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                {extra_where}
                TRIM(COALESCE(ri.estado_asignacion, '')) = 'aprobado'
                AND TRIM(COALESCE(ri.estado_respuesta, '')) = 'sin_respuesta'
                AND (
                    julianday('now') - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion))
                ) > 5
                GROUP BY {proceso_expr}
                ORDER BY vencidas DESC, proceso ASC
            """, params)

        rows = cur.fetchall()
        chart_vencidas = {
            "labels": [r["proceso"] for r in rows],
            "total":  [int(r["vencidas"] or 0) for r in rows],
        }         

        # ========= CHART: responsables con OM vencidas =========
        responsable_expr = """COALESCE(
            NULLIF(LTRIM(RTRIM(uimp.nombre_completo)), ''),
            NULLIF(LTRIM(RTRIM(uimp.username)), ''),
            'SIN RESPONSABLE'
        )""" if is_sqlserver else """COALESCE(
            NULLIF(TRIM(uimp.nombre_completo), ''),
            NULLIF(TRIM(uimp.username), ''),
            'SIN RESPONSABLE'
        )"""

        if is_sqlserver:
            cur.execute(f"""
                SELECT TOP 10
                    {responsable_expr} AS responsable,
                    COUNT(DISTINCT ri.id) AS vencidas
                FROM reclamos r
                INNER JOIN reclamo_imputados ri ON ri.reclamo_id = r.id
                LEFT JOIN usuarios uimp ON uimp.id = ri.imputado_id
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                {extra_where}
                ri.fecha_aprobacion_asignacion IS NOT NULL
                AND ri.fecha_respuesta_imputado IS NULL
                AND DATEDIFF(
                    DAY,
                    TRY_CONVERT(datetime, ri.fecha_aprobacion_asignacion),
                    GETDATE()
                ) > 5
                GROUP BY {responsable_expr}
                ORDER BY COUNT(DISTINCT ri.id) DESC, {responsable_expr} ASC
            """, params)
        else:
            cur.execute(f"""
                SELECT TOP 10
                    {responsable_expr} AS responsable,
                    COUNT(DISTINCT ri.id) AS vencidas
                FROM reclamos r
                INNER JOIN reclamo_imputados ri ON ri.reclamo_id = r.id
                LEFT JOIN usuarios uimp ON uimp.id = ri.imputado_id
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                {extra_where}
                ri.fecha_aprobacion_asignacion IS NOT NULL
                AND ri.fecha_respuesta_imputado IS NULL
                AND (julianday('now') - julianday(ri.fecha_aprobacion_asignacion)) > 5
                GROUP BY {responsable_expr}
                ORDER BY vencidas DESC, responsable ASC
             """, params)

        rows = cur.fetchall()
        chart_vencidas_responsables = {
            "labels": [r["responsable"] for r in rows],
            "total":  [int(r["vencidas"] or 0) for r in rows],
        }

        # ========= CHART: miembros de equipo con mayor atraso =========
        miembro_expr = """COALESCE(
            NULLIF(LTRIM(RTRIM(ume.nombre_completo)), ''),
            NULLIF(LTRIM(RTRIM(ume.username)), ''),
            'SIN MIEMBRO'
        )""" if is_sqlserver else """COALESCE(
            NULLIF(TRIM(ume.nombre_completo), ''),
            NULLIF(TRIM(ume.username), ''),
            'SIN MIEMBRO'
        )"""

        sponsor_expr = """COALESCE(
            NULLIF(LTRIM(RTRIM(us.nombre_completo)), ''),
            NULLIF(LTRIM(RTRIM(us.username)), ''),
            'SIN SPONSOR'
        )""" if is_sqlserver else """COALESCE(
            NULLIF(TRIM(us.nombre_completo), ''),
            NULLIF(TRIM(us.username), ''),
            'SIN SPONSOR'
        )"""

        if is_sqlserver:
            cur.execute(f"""
                WITH base AS (
                    SELECT
                        r.id,
                        r.codigo,
                        COALESCE(NULLIF(LTRIM(RTRIM(r.proceso_text)), ''), 'SIN PROCESO') AS proceso,
                        ri.id AS imputacion_id,
                        ri.imputado_id,
                        ri.estado_asignacion,
                        ri.estado_respuesta,
                        ri.fecha_aprobacion_asignacion,
                        r.fecha_creacion,
                        CASE
                            WHEN ri.estado_asignacion = 'pend_aprobacion' THEN 'Pendiente aceptación del responsable'
                            WHEN ri.estado_asignacion = 'rechazado' THEN 'Imputación rechazada'
                            WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'sin_respuesta' THEN 'Pendiente respuesta del imputado'
                            WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'pendiente_jefe' THEN 'Respuesta pendiente de aprobación'
                            WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'aprobada' THEN 'Cerrado'
                            WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'rechazada' THEN 'Respuesta rechazada'
                            ELSE COALESCE(ri.estado_asignacion, '') + '/' + COALESCE(ri.estado_respuesta, '')
                        END AS estado_imputacion
                    FROM reclamos r
                    INNER JOIN reclamo_imputados ri
                        ON ri.reclamo_id = r.id
                    LEFT JOIN usuarios u ON u.id = r.creado_por
                    LEFT JOIN departamentos d ON d.id = u.departamento_id
                    {where_sql}
                )
                SELECT TOP 10
                    {miembro_expr} AS miembro_equipo,
                    {sponsor_expr} AS sponsor,
                    b.proceso,
                    COUNT(DISTINCT eq.id) AS vencidas,
                    CAST(AVG(CAST(DATEDIFF(
                        DAY,
                        COALESCE(
                            TRY_CONVERT(datetime, b.fecha_aprobacion_asignacion, 120),
                            TRY_CONVERT(datetime, b.fecha_creacion, 120)
                        ),
                        GETDATE()
                    ) AS FLOAT)) AS DECIMAL(10,2)) AS atraso_promedio,
                    MAX(CAST(DATEDIFF(
                        DAY,
                        COALESCE(
                            TRY_CONVERT(datetime, b.fecha_aprobacion_asignacion, 120),
                            TRY_CONVERT(datetime, b.fecha_creacion, 120)
                        ),
                        GETDATE()
                    ) AS FLOAT)) AS atraso_maximo
                FROM base b
                INNER JOIN reclamo_equipo_respuestas eq
                    ON eq.reclamo_id = b.id
                AND eq.imputacion_id = b.imputacion_id
                AND eq.activo = 1
                LEFT JOIN usuarios ume ON ume.id = eq.usuario_id
                LEFT JOIN usuarios us ON us.id = b.imputado_id
                WHERE b.estado_imputacion = 'Pendiente respuesta del imputado'
                AND DATEDIFF(
                        DAY,
                        COALESCE(
                            TRY_CONVERT(datetime, b.fecha_aprobacion_asignacion, 120),
                            TRY_CONVERT(datetime, b.fecha_creacion, 120)
                        ),
                        GETDATE()
                    ) > 5
                GROUP BY eq.usuario_id, {miembro_expr}, {sponsor_expr}, b.proceso
                ORDER BY atraso_maximo DESC, vencidas DESC, {miembro_expr} ASC
            """, params)
        else:
            cur.execute(f"""
                SELECT TOP 10
                    {miembro_expr} AS miembro_equipo,
                    {sponsor_expr} AS sponsor,
                    {proceso_expr} AS proceso,
                    COUNT(DISTINCT eq.id) AS vencidas,
                    ROUND(
                        AVG(
                            julianday('now') - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion))
                        ),
                        2
                    ) AS atraso_promedio,
                    MAX(
                        ROUND(
                            julianday('now') - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion)),
                            2
                        )
                    ) AS atraso_maximo
                FROM reclamos r
                INNER JOIN reclamo_imputados ri
                    ON ri.reclamo_id = r.id
                INNER JOIN reclamo_equipo_respuestas eq
                    ON eq.reclamo_id = r.id
                AND eq.imputacion_id = ri.id
                AND eq.activo = 1
                LEFT JOIN usuarios ume ON ume.id = eq.usuario_id
                LEFT JOIN usuarios us ON us.id = ri.imputado_id
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_sql}
                {extra_where}
                TRIM(COALESCE(ri.estado_asignacion, '')) = 'aprobado'
                AND TRIM(COALESCE(ri.estado_respuesta, '')) = 'sin_respuesta'
                AND (julianday('now') - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion))) > 5
                GROUP BY eq.usuario_id, {sponsor_expr}, {proceso_expr}, {miembro_expr}
                ORDER BY atraso_maximo DESC, vencidas DESC, miembro_equipo ASC
            """, params)

        rows = cur.fetchall()
        chart_vencidas_equipo = {
            "labels": [f'{r["miembro_equipo"]} | {r["sponsor"]}' for r in rows],
            "miembro": [r["miembro_equipo"] for r in rows],
            "sponsor": [r["sponsor"] for r in rows],
            "proceso": [r["proceso"] for r in rows],
            "total": [int(r["vencidas"] or 0) for r in rows],
            "atraso_promedio": [float(r["atraso_promedio"] or 0) for r in rows],
            "atraso_maximo": [float(r["atraso_maximo"] or 0) for r in rows],
        }       



        # ========= CHART: línea de tiempo de OM abiertas y cuello de botella =========
        from collections import defaultdict
        from datetime import datetime

        def _parse_dt_dashboard(v):
            if not v:
                return None

            if isinstance(v, datetime):
                return v

            s = str(v).strip()
            if not s:
                return None

            for fmt in (
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d %H:%M",
                "%Y-%m-%d",
                "%Y-%m-%dT%H:%M:%S",
            ):
                try:
                    return datetime.strptime(s[:19], fmt)
                except Exception:
                    pass

            return None


        def _dias_desde(v):
            dt = _parse_dt_dashboard(v)
            if not dt:
                return 0
            return max((datetime.now().date() - dt.date()).days, 0)


        def _fmt_fecha_dashboard(v):
            dt = _parse_dt_dashboard(v)
            return dt.strftime("%Y-%m-%d") if dt else ""


        def _append_open_condition(base_where_sql):
            open_condition = """
                (
                    LOWER(COALESCE(r.estado_global,'')) <> 'cerrado'
                    AND UPPER(COALESCE(r.estado_global,'')) NOT LIKE 'CERR%'
                )
            """

            if base_where_sql:
                return base_where_sql + " AND " + open_condition

            return "WHERE " + open_condition


        where_abiertas_sql = _append_open_condition(where_sql)
        params_linea = list(params)

        if is_sqlserver:
            cur.execute(f"""
                SELECT
                    r.id,
                    r.codigo,
                    COALESCE(r.proceso_text, 'SIN PROCESO') AS proceso,
                    COALESCE(r.cliente_nombre, '') AS cliente,
                    COALESCE(r.tipo_reclamo, '') AS tipo_reclamo,
                    r.fecha_creacion,
                    r.fecha_reclamo,
                    COALESCE(r.estado_global, '') AS estado_global,

                    (
                        SELECT COUNT(*)
                        FROM reclamo_imputados ri
                        WHERE ri.reclamo_id = r.id
                    ) AS total_imputaciones,

                    (
                        SELECT COUNT(*)
                        FROM reclamo_imputados ri
                        WHERE ri.reclamo_id = r.id
                        AND COALESCE(ri.estado_asignacion, '') = 'pend_aprobacion'
                    ) AS imputaciones_pend_aprobacion,

                    (
                        SELECT COUNT(*)
                        FROM reclamo_imputados ri
                        WHERE ri.reclamo_id = r.id
                        AND COALESCE(ri.estado_asignacion, '') = 'aprobado'
                        AND COALESCE(ri.estado_respuesta, '') = 'sin_respuesta'
                    ) AS sponsors_sin_respuesta,

                    (
                        SELECT COUNT(*)
                        FROM reclamo_imputados ri
                        WHERE ri.reclamo_id = r.id
                        AND COALESCE(ri.estado_asignacion, '') = 'aprobado'
                        AND COALESCE(ri.estado_respuesta, '') = 'pendiente_jefe'
                    ) AS pendientes_jefe,

                    (
                        SELECT COUNT(*)
                        FROM reclamo_imputados ri
                        WHERE ri.reclamo_id = r.id
                        AND COALESCE(ri.estado_asignacion, '') = 'aprobado'
                        AND COALESCE(ri.estado_respuesta, '') = 'rechazada'
                    ) AS respuestas_rechazadas,

                    (
                        SELECT COUNT(*)
                        FROM reclamo_equipo_respuestas er
                        WHERE er.reclamo_id = r.id
                        AND COALESCE(er.activo, 1) = 1
                    ) AS equipo_total,

                    (
                        SELECT COUNT(*)
                        FROM reclamo_equipo_respuestas er
                        WHERE er.reclamo_id = r.id
                        AND COALESCE(er.activo, 1) = 1
                        AND NOT EXISTS (
                            SELECT 1
                            FROM reclamo_respuestas_equipo rre
                            WHERE rre.reclamo_id = er.reclamo_id
                            AND rre.imputacion_id = er.imputacion_id
                            AND rre.miembro_id = er.usuario_id
                            AND COALESCE(rre.activo, 1) = 1
                        )
                    ) AS equipo_pendiente,

                    (
                        SELECT COUNT(DISTINCT rre.miembro_id)
                        FROM reclamo_respuestas_equipo rre
                        WHERE rre.reclamo_id = r.id
                        AND COALESCE(rre.activo, 1) = 1
                    ) AS equipo_respondido,

                    (
                        SELECT MAX(er.creado_at)
                        FROM reclamo_equipo_respuestas er
                        WHERE er.reclamo_id = r.id
                        AND COALESCE(er.activo, 1) = 1
                    ) AS fecha_ultimo_equipo_asignado,

                    (
                        SELECT MAX(rre.created_at)
                        FROM reclamo_respuestas_equipo rre
                        WHERE rre.reclamo_id = r.id
                        AND COALESCE(rre.activo, 1) = 1
                    ) AS fecha_ultima_respuesta_equipo,

                    (
                        SELECT MAX(ri.fecha_respuesta_imputado)
                        FROM reclamo_imputados ri
                        WHERE ri.reclamo_id = r.id
                    ) AS fecha_respuesta_sponsor,

                    (
                        SELECT MAX(ri.fecha_aprobacion_asignacion)
                        FROM reclamo_imputados ri
                        WHERE ri.reclamo_id = r.id
                    ) AS fecha_aprobacion_asignacion,

                    (
                        SELECT STRING_AGG(CAST(COALESCE(u.nombre_completo, u.username) AS VARCHAR(MAX)), ', ')
                        FROM reclamo_imputados ri
                        LEFT JOIN usuarios u ON u.id = ri.imputado_id
                        WHERE ri.reclamo_id = r.id
                    ) AS sponsors,

                    (
                        SELECT STRING_AGG(CAST(COALESCE(u.nombre_completo, u.username) AS VARCHAR(MAX)), ', ')
                        FROM reclamo_equipo_respuestas er
                        LEFT JOIN usuarios u ON u.id = er.usuario_id
                        WHERE er.reclamo_id = r.id
                        AND COALESCE(er.activo, 1) = 1
                    ) AS miembros_equipo

                FROM reclamos r
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_abiertas_sql}
                ORDER BY COALESCE(r.fecha_creacion, r.fecha_reclamo) ASC
            """, params_linea)
        else:
            cur.execute(f"""
                SELECT
                    r.id,
                    r.codigo,
                    COALESCE(r.proceso_text, 'SIN PROCESO') AS proceso,
                    COALESCE(r.cliente_nombre, '') AS cliente,
                    COALESCE(r.tipo_reclamo, '') AS tipo_reclamo,
                    r.fecha_creacion,
                    r.fecha_reclamo,
                    COALESCE(r.estado_global, '') AS estado_global,

                    (
                        SELECT COUNT(*)
                        FROM reclamo_imputados ri
                        WHERE ri.reclamo_id = r.id
                    ) AS total_imputaciones,

                    (
                        SELECT COUNT(*)
                        FROM reclamo_imputados ri
                        WHERE ri.reclamo_id = r.id
                        AND COALESCE(ri.estado_asignacion, '') = 'pend_aprobacion'
                    ) AS imputaciones_pend_aprobacion,

                    (
                        SELECT COUNT(*)
                        FROM reclamo_imputados ri
                        WHERE ri.reclamo_id = r.id
                        AND COALESCE(ri.estado_asignacion, '') = 'aprobado'
                        AND COALESCE(ri.estado_respuesta, '') = 'sin_respuesta'
                    ) AS sponsors_sin_respuesta,

                    (
                        SELECT COUNT(*)
                        FROM reclamo_imputados ri
                        WHERE ri.reclamo_id = r.id
                        AND COALESCE(ri.estado_asignacion, '') = 'aprobado'
                        AND COALESCE(ri.estado_respuesta, '') = 'pendiente_jefe'
                    ) AS pendientes_jefe,

                    (
                        SELECT COUNT(*)
                        FROM reclamo_imputados ri
                        WHERE ri.reclamo_id = r.id
                        AND COALESCE(ri.estado_asignacion, '') = 'aprobado'
                        AND COALESCE(ri.estado_respuesta, '') = 'rechazada'
                    ) AS respuestas_rechazadas,

                    (
                        SELECT COUNT(*)
                        FROM reclamo_equipo_respuestas er
                        WHERE er.reclamo_id = r.id
                        AND COALESCE(er.activo, 1) = 1
                    ) AS equipo_total,

                    (
                        SELECT COUNT(*)
                        FROM reclamo_equipo_respuestas er
                        WHERE er.reclamo_id = r.id
                        AND COALESCE(er.activo, 1) = 1
                        AND NOT EXISTS (
                            SELECT 1
                            FROM reclamo_respuestas_equipo rre
                            WHERE rre.reclamo_id = er.reclamo_id
                            AND rre.imputacion_id = er.imputacion_id
                            AND rre.miembro_id = er.usuario_id
                            AND COALESCE(rre.activo, 1) = 1
                        )
                    ) AS equipo_pendiente,

                    (
                        SELECT COUNT(DISTINCT rre.miembro_id)
                        FROM reclamo_respuestas_equipo rre
                        WHERE rre.reclamo_id = r.id
                        AND COALESCE(rre.activo, 1) = 1
                    ) AS equipo_respondido,

                    (
                        SELECT MAX(er.creado_at)
                        FROM reclamo_equipo_respuestas er
                        WHERE er.reclamo_id = r.id
                        AND COALESCE(er.activo, 1) = 1
                    ) AS fecha_ultimo_equipo_asignado,

                    (
                        SELECT MAX(rre.created_at)
                        FROM reclamo_respuestas_equipo rre
                        WHERE rre.reclamo_id = r.id
                        AND COALESCE(rre.activo, 1) = 1
                    ) AS fecha_ultima_respuesta_equipo,

                    (
                        SELECT MAX(ri.fecha_respuesta_imputado)
                        FROM reclamo_imputados ri
                        WHERE ri.reclamo_id = r.id
                    ) AS fecha_respuesta_sponsor,

                    (
                        SELECT MAX(ri.fecha_aprobacion_asignacion)
                        FROM reclamo_imputados ri
                        WHERE ri.reclamo_id = r.id
                    ) AS fecha_aprobacion_asignacion,

                    (
                        SELECT GROUP_CONCAT(COALESCE(u.nombre_completo, u.username), ', ')
                        FROM reclamo_imputados ri
                        LEFT JOIN usuarios u ON u.id = ri.imputado_id
                        WHERE ri.reclamo_id = r.id
                    ) AS sponsors,

                    (
                        SELECT GROUP_CONCAT(COALESCE(u.nombre_completo, u.username), ', ')
                        FROM reclamo_equipo_respuestas er
                        LEFT JOIN usuarios u ON u.id = er.usuario_id
                        WHERE er.reclamo_id = r.id
                        AND COALESCE(er.activo, 1) = 1
                    ) AS miembros_equipo

                FROM reclamos r
                LEFT JOIN usuarios u ON u.id = r.creado_por
                LEFT JOIN departamentos d ON d.id = u.departamento_id
                {where_abiertas_sql}
                ORDER BY COALESCE(r.fecha_creacion, r.fecha_reclamo) ASC
            """, params_linea)

        rows_linea = cur.fetchall()

        timeline_items = []
        cuellos_count = defaultdict(int)
        cuellos_dias = defaultdict(list)
        etapas_dias = defaultdict(list)

        mayor_atraso = 0
        total_dias_abiertas = 0

        for r in rows_linea:
            row = dict(r)

            fecha_base = row.get("fecha_creacion") or row.get("fecha_reclamo")
            dias_abierta = _dias_desde(fecha_base)
            total_dias_abiertas += dias_abierta
            mayor_atraso = max(mayor_atraso, dias_abierta)

            total_imputaciones = int(row.get("total_imputaciones") or 0)
            imputaciones_pend_aprobacion = int(row.get("imputaciones_pend_aprobacion") or 0)
            sponsors_sin_respuesta = int(row.get("sponsors_sin_respuesta") or 0)
            pendientes_jefe = int(row.get("pendientes_jefe") or 0)
            respuestas_rechazadas = int(row.get("respuestas_rechazadas") or 0)

            equipo_total = int(row.get("equipo_total") or 0)
            equipo_pendiente = int(row.get("equipo_pendiente") or 0)
            equipo_respondido = int(row.get("equipo_respondido") or 0)

            cuello = "Revisión"
            etapa_actual = "Abierta sin etapa clara"
            responsable_actual = "Por revisar"
            motivo = "No se pudo determinar una etapa específica con los datos actuales."
            fecha_etapa = fecha_base

            if total_imputaciones == 0:
                cuello = "Sin sponsor"
                etapa_actual = "Pendiente asignación"
                responsable_actual = "Servicio al Cliente / Coordinador"
                motivo = "La OM está abierta y aún no tiene sponsor/responsable asignado."
                fecha_etapa = fecha_base

            elif imputaciones_pend_aprobacion > 0:
                cuello = "Aprobador/Jefe"
                etapa_actual = "Pendiente aprobación de asignación"
                responsable_actual = row.get("sponsors") or "Jefe / aprobador"
                motivo = "La asignación del responsable aún está pendiente de aprobación."
                fecha_etapa = fecha_base

            elif equipo_total > 0 and equipo_pendiente > 0:
                cuello = "Miembro del equipo"
                etapa_actual = "Esperando aporte del equipo"
                responsable_actual = row.get("miembros_equipo") or "Miembro del equipo"
                motivo = f"Hay {equipo_pendiente} miembro(s) de equipo sin registrar aporte."
                fecha_etapa = row.get("fecha_ultimo_equipo_asignado") or fecha_base

            elif equipo_total > 0 and equipo_pendiente == 0 and sponsors_sin_respuesta > 0:
                cuello = "Sponsor"
                etapa_actual = "Equipo respondió; falta respuesta sponsor"
                responsable_actual = row.get("sponsors") or "Sponsor"
                motivo = "El equipo ya registró sus aportes, pero falta la respuesta final del sponsor."
                fecha_etapa = row.get("fecha_ultima_respuesta_equipo") or row.get("fecha_aprobacion_asignacion") or fecha_base

            elif sponsors_sin_respuesta > 0:
                cuello = "Sponsor"
                etapa_actual = "Esperando respuesta sponsor"
                responsable_actual = row.get("sponsors") or "Sponsor"
                motivo = "El sponsor/responsable tiene la OM asignada, pero aún no registra respuesta técnica."
                fecha_etapa = row.get("fecha_aprobacion_asignacion") or fecha_base

            elif pendientes_jefe > 0:
                cuello = "Aprobador/Jefe"
                etapa_actual = "Pendiente aprobación de respuesta"
                responsable_actual = "Jefe / aprobador"
                motivo = "La respuesta técnica fue registrada y está pendiente de aprobación."
                fecha_etapa = row.get("fecha_respuesta_sponsor") or fecha_base

            elif respuestas_rechazadas > 0:
                cuello = "Sponsor"
                etapa_actual = "Respuesta rechazada / ajuste pendiente"
                responsable_actual = row.get("sponsors") or "Sponsor"
                motivo = "La respuesta fue rechazada y requiere ajuste por parte del sponsor/responsable."
                fecha_etapa = row.get("fecha_respuesta_sponsor") or fecha_base

            dias_en_etapa = _dias_desde(fecha_etapa)

            cuellos_count[cuello] += 1
            cuellos_dias[cuello].append(dias_en_etapa)
            etapas_dias[etapa_actual].append(dias_en_etapa)

            timeline_items.append({
                "id": row.get("id"),
                "codigo": row.get("codigo") or "",
                "cliente": row.get("cliente") or "",
                "proceso": row.get("proceso") or "SIN PROCESO",
                "tipo_reclamo": row.get("tipo_reclamo") or "",
                "fecha_inicio": _fmt_fecha_dashboard(fecha_base),
                "dias_abierta": int(dias_abierta),
                "dias_en_etapa": int(dias_en_etapa),
                "etapa_actual": etapa_actual,
                "cuello": cuello,
                "responsable_actual": responsable_actual or "",
                "motivo": motivo,
                "sponsors": row.get("sponsors") or "",
                "miembros_equipo": row.get("miembros_equipo") or "",
                "equipo_total": equipo_total,
                "equipo_pendiente": equipo_pendiente,
                "equipo_respondido": equipo_respondido,
                "pasos": {
                    "creada": True,
                    "sponsor": total_imputaciones > 0,
                    "equipo": equipo_total > 0,
                    "respuesta_equipo": equipo_total > 0 and equipo_pendiente == 0,
                    "respuesta_sponsor": sponsors_sin_respuesta == 0 and total_imputaciones > 0,
                    "aprobacion": pendientes_jefe > 0,
                }
            })

        timeline_items = sorted(
            timeline_items,
            key=lambda x: (x["dias_en_etapa"], x["dias_abierta"]),
            reverse=True
        )

        cuellos_labels = sorted(cuellos_count.keys(), key=lambda k: cuellos_count[k], reverse=True)

        chart_cuellos = {
            "labels": cuellos_labels,
            "total": [int(cuellos_count[k]) for k in cuellos_labels],
            "dias_promedio": [
                round(sum(cuellos_dias[k]) / len(cuellos_dias[k]), 2) if cuellos_dias[k] else 0
                for k in cuellos_labels
            ],
        }

        etapas_labels = sorted(
            etapas_dias.keys(),
            key=lambda k: (sum(etapas_dias[k]) / len(etapas_dias[k])) if etapas_dias[k] else 0,
            reverse=True
        )

        chart_aging_etapas = {
            "labels": etapas_labels,
            "dias_promedio": [
                round(sum(etapas_dias[k]) / len(etapas_dias[k]), 2) if etapas_dias[k] else 0
                for k in etapas_labels
            ],
            "total": [len(etapas_dias[k]) for k in etapas_labels],
        }

        cuello_principal = cuellos_labels[0] if cuellos_labels else "Sin datos"

        chart_linea_abiertas = {
            "kpis": {
                "total_abiertas": len(timeline_items),
                "promedio_dias_abierta": round(total_dias_abiertas / len(timeline_items), 2) if timeline_items else 0,
                "mayor_atraso": int(mayor_atraso),
                "cuello_principal": cuello_principal,
            },
            "cuellos": chart_cuellos,
            "aging_etapas": chart_aging_etapas,
            "items": timeline_items[:80],
        }


        chart = {
            "estados": chart_estados,
            "meses": chart_meses,
            "dias": chart_dias,
            "imputados": chart_imputados,
            "procesos": chart_procesos,
            "tipos": chart_tipos,
            "tiempos": chart_tiempos,
            "vencidas_responsables": chart_vencidas_responsables,
            "vencidas": chart_vencidas,
            "vencidas_equipo": chart_vencidas_equipo,
            "procesos_involucrados": chart_procesos_involucrados,
            "linea_abiertas": chart_linea_abiertas,

        }

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_87)
        deptos = [r["depto"] for r in cur.fetchall()]

        # ========= FILTRO: procesos únicos =========
        # El combo muestra procesos individuales aunque en la OM estén combinados.
        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_88)

        procesos_set = set()

        for r in cur.fetchall():
            txt = r["proceso_text"] if hasattr(r, "keys") else r[0]
            txt = (txt or "").strip()

            if not txt:
                procesos_set.add("SIN PROCESO")
                continue

            # Normalizar separadores comunes
            partes = re.split(r"\s*(?:\||;|,)\s*", txt)

            for p in partes:
                p = (p or "").strip()
                if p:
                    procesos_set.add(p)

        procesos = sorted(procesos_set, key=lambda x: x.upper())

        return render_template(
            "reclamos_dashboard.html",
            kpis=kpis,
            chart=chart,
            deptos=deptos,
            depto_sel=depto_sel,
            procesos=procesos,
            proceso_sel=proceso_sel,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            active_page="reclamos_dashboard",
        ) 

    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/aprobar", methods=["POST"])
    @require_login
    def equipo_respuestas_aprobar(reclamo_id):
        uid = _current_user_id()
        data = request.get_json(silent=True) or {}

        imputacion_id = data.get("imputacion_id")
        miembro_id = data.get("miembro_id")

        if not uid:
            return jsonify(ok=False, error="Sesión inválida. Vuelve a iniciar sesión."), 401

        if not imputacion_id or not miembro_id:
            return jsonify(ok=False, error="Falta imputacion_id o miembro_id"), 400

        try:
            imputacion_id = int(imputacion_id)
            miembro_id = int(miembro_id)
        except (TypeError, ValueError):
            return jsonify(ok=False, error="imputacion_id o miembro_id inválido"), 400

        db = get_db()

        # =========================================================
        # Datos base de la imputación / OM
        # =========================================================
        row_base = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_89, (reclamo_id, imputacion_id)).fetchone()

        if not row_base:
            return jsonify(ok=False, error="No se encontró la imputación de la OM."), 404

        # =========================================================
        # Permiso:
        # - sponsor principal: ri.imputado_id
        # - sponsor backup/principal por proceso: RECL_PROCESO_SPONSOR
        # - admin/coordinador
        # =========================================================
        es_sponsor_principal = int(row_base["imputado_id"] or 0) == int(uid or 0)

        es_sponsor_del_proceso = _usuario_es_sponsor_del_proceso(
            db,
            row_base["proceso_id"],
            uid
        )

        puede_gestionar_equipo = _puede_gestionar_equipo(
            reclamo_id,
            uid
        )

        if not (
            es_sponsor_principal
            or es_sponsor_del_proceso
            or puede_gestionar_equipo
            or _is_admin_like()
        ):
            return jsonify(ok=False, error="No autorizado"), 403

        # Validar que el miembro pertenece al equipo de esa imputación
        row_equipo = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_90, (reclamo_id, imputacion_id, miembro_id)).fetchone()

        if not row_equipo:
            return jsonify(ok=False, error="El miembro no pertenece al equipo de esta OM."), 404

        # Validar que exista respuesta activa
        row_resp = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_91, (reclamo_id, imputacion_id, miembro_id)).fetchone()

        if not row_resp:
            return jsonify(ok=False, error="El miembro aún no tiene respuesta registrada."), 400

        # Registrar aprobación si existen columnas de revisión.
        # Se deja protegido para no romper si todavía no existen.
        try:
            db.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_92, (uid, _now_iso(), reclamo_id, imputacion_id, miembro_id))
            db.commit()
        except Exception:
            db.rollback()
            # No rompemos el flujo si esas columnas no existen
            pass

        return jsonify(ok=True)



    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/rechazar", methods=["POST"])
    @require_login
    def equipo_respuestas_rechazar(reclamo_id):
        uid = _current_user_id()
        data = request.get_json(silent=True) or {}

        imputacion_id = data.get("imputacion_id")
        miembro_id = data.get("miembro_id")
        motivo = (data.get("motivo") or "").strip()

        if not uid:
            return jsonify(ok=False, error="Sesión inválida. Vuelve a iniciar sesión."), 401

        if not imputacion_id or not miembro_id or not motivo:
            return jsonify(ok=False, error="Falta imputacion_id, miembro_id o motivo"), 400

        try:
            imputacion_id = int(imputacion_id)
            miembro_id = int(miembro_id)
        except (TypeError, ValueError):
            return jsonify(ok=False, error="imputacion_id o miembro_id inválido"), 400

        db = get_db()

        # =========================================================
        # Datos base de la imputación / OM / miembro
        # =========================================================
        row_base = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_93, (miembro_id, reclamo_id, imputacion_id)).fetchone()

        if not row_base:
            return jsonify(ok=False, error="No se encontró la imputación de la OM."), 404

        # =========================================================
        # Permiso:
        # - sponsor principal: ri.imputado_id
        # - sponsor backup/principal por proceso: RECL_PROCESO_SPONSOR
        # - admin/coordinador
        # =========================================================
        es_sponsor_principal = int(row_base["imputado_id"] or 0) == int(uid or 0)

        es_sponsor_del_proceso = _usuario_es_sponsor_del_proceso(
            db,
            row_base["proceso_id"],
            uid
        )

        puede_gestionar_equipo = _puede_gestionar_equipo(
            reclamo_id,
            uid
        )

        if not (
            es_sponsor_principal
            or es_sponsor_del_proceso
            or puede_gestionar_equipo
            or _is_admin_like()
        ):
            return jsonify(ok=False, error="No autorizado"), 403

        # Validar que el miembro pertenece al equipo de esa imputación
        row_equipo = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_90, (reclamo_id, imputacion_id, miembro_id)).fetchone()

        if not row_equipo:
            return jsonify(ok=False, error="El miembro no pertenece al equipo de esta OM."), 404

        # Validar que exista respuesta activa
        row_resp = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_91, (reclamo_id, imputacion_id, miembro_id)).fetchone()

        if not row_resp:
            return jsonify(ok=False, error="El miembro aún no tiene respuesta registrada."), 400

        # Registrar rechazo si existen columnas de revisión.
        # Se protege para no romper si todavía no existen.
        try:
            db.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_94, (
                motivo,
                uid,
                _now_iso(),
                reclamo_id,
                imputacion_id,
                miembro_id
            ))
            db.commit()
        except Exception:
            db.rollback()
            # No rompemos el flujo si alguna columna todavía no existe.
            pass

        # Notificar al miembro
        try:
            _notify_colaborador_aporte_rechazado(
                db,
                miembro_id,
                row_base["codigo"],
                motivo
            )
        except Exception:
            current_app.logger.exception(
                "No se pudo notificar rechazo de aporte. reclamo_id=%s imputacion_id=%s miembro_id=%s",
                reclamo_id,
                imputacion_id,
                miembro_id
            )

        return jsonify(ok=True)
    @app.route('/reclamos/equipo-acciones/<int:accion_id>/observacion', methods=['POST'])
    @require_login
    def reclamo_equipo_accion_guardar_observacion(accion_id):
        db = get_db()
        uid = _current_user_id()

        #ensure_reclamo_respuesta_equipo_acciones_schema(db)

        data = request.get_json(silent=True) or {}
        observacion = (data.get("observacion") or "").strip()

        row = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_95, (accion_id,)).fetchone()

        if not row:
            return jsonify(ok=False, error="Acción no encontrada"), 404

        if int(row["cumplido"] or 0) == 1:
            return jsonify(ok=False, error="La acción ya está cumplida y no permite editar observación"), 400

        db.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_96, (observacion, uid, accion_id))

        db.commit()

        return jsonify(ok=True, msg="Observación guardada")    
        

    @app.route('/reclamos/equipo-acciones/evidencias/<int:evidencia_id>/eliminar', methods=['POST'])
    @require_login
    def reclamo_equipo_accion_evidencia_eliminar(evidencia_id):
        db = get_db()
        uid = _current_user_id()

        row = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_97, (evidencia_id,)).fetchone()

        if not row:
            return jsonify(ok=False, error="Evidencia no encontrada"), 404

        if int(row["evidencia_activa"] or 0) != 1:
            return jsonify(ok=False, error="La evidencia ya fue eliminada"), 400

        if int(row["accion_cumplida"] or 0) == 1:
            return jsonify(ok=False, error="No se puede eliminar evidencia de una acción ya cumplida"), 400

        db.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_98, (evidencia_id,))

        db.commit()

        return jsonify(ok=True, msg="Evidencia eliminada")


    @app.route('/reclamos/imputacion/<int:imp_id>/respuesta-detalle', methods=['GET'])
    @require_login
    def reclamo_imputacion_respuesta_detalle(imp_id):
        conn = get_db()
        #ensure_reclamos_schema(conn)
        #ensure_reclamo_imputado_acciones_schema(conn)

        cur = conn.cursor()
        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_99, (imp_id,))
        row = cur.fetchone()

        if not row:
            return jsonify(ok=False, error="Imputación no encontrada"), 404

        item = dict(row)

        acciones = _get_imputado_acciones_full(conn, imp_id)
        item["causas"] = acciones.get("causas", [])
        item["control"] = acciones.get("control", [])
        item["correctiva_items"] = acciones.get("correctiva", [])

        return jsonify(ok=True, item=item)
    


    @app.route('/reclamos/imputado-acciones/<int:accion_id>/observacion', methods=['POST'])
    @require_login
    def reclamo_imputado_accion_observacion(accion_id):
        db = get_db()
        #ensure_reclamos_schema(db)
        #ensure_reclamo_imputado_acciones_schema(db)

        uid = _current_user_id()
        permitido, row = _puede_gestionar_imputado_accion(accion_id, uid)
        if not permitido:
            return jsonify(ok=False, error="No autorizado para editar esta acción"), 403

        if not row:
            return jsonify(ok=False, error="Acción no encontrada"), 404

        if int(row["accion_activa"] or 0) != 1:
            return jsonify(ok=False, error="La acción no está activa"), 400

        if int(row["cumplido"] or 0) == 1:
            return jsonify(ok=False, error="La acción ya está cumplida y no permite editar observación"), 400

        data = request.get_json(silent=True) or {}
        observacion = (data.get("observacion") or "").strip()

        db.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_100, (observacion, uid, accion_id))

        db.commit()
        return jsonify(ok=True, msg="Observación guardada")


    @app.route('/reclamos/imputado-acciones/<int:accion_id>/cumplir', methods=['POST'])
    @require_login
    def reclamo_imputado_accion_cumplir(accion_id):
        db = get_db()
        #ensure_reclamos_schema(db)
        #ensure_reclamo_imputado_acciones_schema(db)

        uid = _current_user_id()
        permitido, row = _puede_gestionar_imputado_accion(accion_id, uid)
        if not permitido:
            return jsonify(ok=False, error="No autorizado para editar esta acción"), 403

        if not row:
            return jsonify(ok=False, error="Acción no encontrada"), 404

        if int(row["accion_activa"] or 0) != 1:
            return jsonify(ok=False, error="La acción no está activa"), 400

        if int(row["cumplido"] or 0) == 1:
            return jsonify(ok=False, error="La acción ya está cumplida"), 400

        data = request.get_json(silent=True) or {}
        cumplido = data.get("cumplido", True)
        fecha_cumplimiento = (data.get("fecha_cumplimiento") or "").strip()

        if not bool(cumplido):
            return jsonify(ok=False, error="Acción inválida"), 400

        if not fecha_cumplimiento:
            fecha_cumplimiento = datetime.now().strftime("%Y-%m-%d")

        if not _is_date_yyyy_mm_dd(fecha_cumplimiento):
            return jsonify(ok=False, error="Fecha inválida. Use YYYY-MM-DD"), 400

        db.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_101, (fecha_cumplimiento, uid, accion_id))

        db.commit()
        return jsonify(ok=True, msg="Acción marcada como cumplida")


    @app.route('/reclamos/imputado-acciones/<int:accion_id>/evidencia', methods=['POST'])
    @require_login
    def reclamo_imputado_accion_evidencia(accion_id):
        db = get_db()

        uid = _current_user_id()
        permitido, row = _puede_gestionar_imputado_accion(accion_id, uid)
        if not permitido:
            return jsonify(ok=False, error="No autorizado para cargar evidencia"), 403

        if not row:
            return jsonify(ok=False, error="Acción no encontrada"), 404

        if int(row["accion_activa"] or 0) != 1:
            return jsonify(ok=False, error="La acción no está activa"), 400

        if int(row["cumplido"] or 0) == 1:
            return jsonify(ok=False, error="No se puede cargar evidencia en una acción ya cumplida"), 400

        f = request.files.get("file")
        if not f:
            return jsonify(ok=False, error="No se recibió archivo"), 400

        original_name = (f.filename or "").strip()
        if not original_name:
            return jsonify(ok=False, error="El archivo no tiene nombre válido"), 400

        safe_name = secure_filename(original_name)
        ext = os.path.splitext(safe_name)[1].lower()
        physical_name = f"{accion_id}_{uuid4().hex}{ext}"

        folder = _sponsor_acciones_upload_dir()
        full_path = os.path.join(folder, physical_name)
        f.save(full_path)

        try:
            size_bytes = os.path.getsize(full_path)
        except Exception:
            size_bytes = 0

        cur = db.cursor()
        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_INS_102, (
            accion_id,
            physical_name,
            original_name,
            (f.mimetype or ""),
            int(size_bytes or 0),
            uid
        ))

        row_id = cur.fetchone()
        evidencia_id = row_id["id"] if row_id and row_id["id"] is not None else None

        if evidencia_id is None:
            db.rollback()
            return jsonify(ok=False, error="No se pudo recuperar el ID de la evidencia"), 500

        db.commit()

        return jsonify(ok=True, item={
            "id": evidencia_id,
            "accion_id": accion_id,
            "filename": physical_name,
            "original_name": original_name,
            "content_type": (f.mimetype or ""),
            "size_bytes": int(size_bytes or 0),
            "download_url": url_for("reclamo_imputado_accion_evidencia_download", evidencia_id=evidencia_id)
        })

    @app.route('/reclamos/imputado-acciones/evidencias/<int:evidencia_id>/eliminar', methods=['POST'])
    @require_login
    def reclamo_imputado_accion_evidencia_eliminar(evidencia_id):
        db = get_db()
        uid = _current_user_id()

        row = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_103, (evidencia_id,)).fetchone()

        if not row:
            return jsonify(ok=False, error="Evidencia no encontrada"), 404

        permitido, _accion = _puede_gestionar_imputado_accion(int(row["accion_id"]), uid)
        if not permitido:
            return jsonify(ok=False, error="No autorizado para eliminar esta evidencia"), 403

        if int(row["evidencia_activa"] or 0) != 1:
            return jsonify(ok=False, error="La evidencia ya fue eliminada"), 400

        if int(row["accion_cumplida"] or 0) == 1:
            return jsonify(ok=False, error="No se puede eliminar evidencia de una acción cumplida"), 400

        filename = (row["filename"] or "").strip()
        if filename:
            try:
                path = os.path.join(_sponsor_acciones_upload_dir(), filename)
                if os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass

        db.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_104, (evidencia_id,))
        db.commit()

        return jsonify(ok=True, msg="Evidencia eliminada")

    @app.route('/reclamos/imputado-acciones/evidencias/<int:evidencia_id>/download', methods=['GET'])
    @require_login
    def reclamo_imputado_accion_evidencia_download(evidencia_id):
        db = get_db()
        uid = _current_user_id()

        row = db.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_105, (evidencia_id,)).fetchone()

        if not row:
            return abort(404)

        permitido, _accion = _puede_gestionar_imputado_accion(int(row["accion_id"]), uid)
        if not permitido and not _is_admin_like():
            return abort(403)

        if int(row["evidencia_activa"] or 0) != 1:
            return abort(404)

        filename = (row["filename"] or "").strip()
        if not filename:
            return abort(404)

        path = os.path.join(_sponsor_acciones_upload_dir(), filename)
        if not os.path.exists(path):
            return abort(404)

        return send_file(
            path,
            as_attachment=True,
            download_name=(row["original_name"] or filename),
            mimetype=(row["content_type"] or "application/octet-stream")
        )




    @app.route('/reclamos/ia/mejorar_respuesta', methods=['POST'])
    @require_login
    def reclamos_ia_mejorar_respuesta():
        import json

        current_app.logger.info(">>> IA endpoint llamado")

        data = request.get_json(silent=True)
        current_app.logger.info(">>> JSON recibido: %s", data)
        current_app.logger.info(">>> OPENAI_API_KEY presente: %s", bool(os.getenv("OPENAI_API_KEY")))

        if not data:
            return jsonify(ok=False, msg="No llegó JSON al backend"), 400

        try:
            client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        except Exception as e:
            current_app.logger.exception("Error cargando cliente OpenAI")
            return jsonify(ok=False, msg=f"Error configurando OpenAI: {e}"), 500

        observacion = (data.get("observacion") or "").strip()
        metodo = (data.get("metodo_analisis") or "FISHBONE").strip().upper()

        fish_metodo = (data.get("fish_metodo") or "").strip()
        fish_maquinas = (data.get("fish_maquinas") or "").strip()
        fish_materiales = (data.get("fish_materiales") or "").strip()
        fish_personas = (data.get("fish_personas") or "").strip()
        fish_entorno = (data.get("fish_entorno") or "").strip()
        fish_medicion = (data.get("fish_medicion") or "").strip()

        why1 = (data.get("why1") or "").strip()
        why2 = (data.get("why2") or "").strip()
        why3 = (data.get("why3") or "").strip()
        why4 = (data.get("why4") or "").strip()
        why5 = (data.get("why5") or "").strip()

        causas = data.get("causas") or []
        control = data.get("control") or []
        correctiva = data.get("correctiva") or []

        def _norm_items(items):
            out = []
            if not isinstance(items, list):
                return out

            for it in items:
                if not isinstance(it, dict):
                    continue

                desc = (it.get("descripcion") or "").strip()
                fecha = (it.get("fecha_compromiso") or "").strip()

                if desc or fecha:
                    out.append({
                        "descripcion": desc,
                        "fecha_compromiso": fecha
                    })
            return out

        causas_n = _norm_items(causas)
        control_n = _norm_items(control)
        correctiva_n = _norm_items(correctiva)

        if (
            not observacion and
            not fish_metodo and not fish_maquinas and not fish_materiales and
            not fish_personas and not fish_entorno and not fish_medicion and
            not why1 and not why2 and not why3 and not why4 and not why5 and
            not causas_n and not control_n and not correctiva_n
        ):
            return jsonify(ok=False, msg="No hay información suficiente para mejorar."), 400

        # =========================
        # Prompt + schema por método
        # =========================
        if metodo == "5WHYS":
            prompt = f"""
    Eres especialista en calidad, reclamos y oportunidades de mejora.

    Mejora la redacción técnica del análisis 5 porqués y de las acciones propuestas.
    Reglas:
    - redacta en tono formal y claro
    - evita culpar personas
    - enfócate en procesos y mejora
    - no inventes hechos
    - debes devolver obligatoriamente why1, why2, why3, why4 y why5
    - si ya existen causas, acciones de control o correctivas, mejóralas
    - si no existen, propón una redacción inicial coherente con la observación y el análisis
    - conserva fecha_compromiso solo si ya viene informada

    OBSERVACIÓN:
    {observacion}

    5 POR QUÉS ACTUALES:
    1. {why1}
    2. {why2}
    3. {why3}
    4. {why4}
    5. {why5}

    CAUSAS ACTUALES:
    {json.dumps(causas_n, ensure_ascii=False)}

    ACCIONES DE CONTROL ACTUALES:
    {json.dumps(control_n, ensure_ascii=False)}

    ACCIONES CORRECTIVAS ACTUALES:
    {json.dumps(correctiva_n, ensure_ascii=False)}
    """

            schema = {
                "name": "respuesta_om_mejorada_5whys",
                "strict": True,
                "schema": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "why1": {"type": "string"},
                        "why2": {"type": "string"},
                        "why3": {"type": "string"},
                        "why4": {"type": "string"},
                        "why5": {"type": "string"},
                        "causas": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "descripcion": {"type": "string"},
                                    "fecha_compromiso": {"type": "string"}
                                },
                                "required": ["descripcion", "fecha_compromiso"]
                            }
                        },
                        "control": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "descripcion": {"type": "string"},
                                    "fecha_compromiso": {"type": "string"}
                                },
                                "required": ["descripcion", "fecha_compromiso"]
                            }
                        },
                        "correctiva": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "descripcion": {"type": "string"},
                                    "fecha_compromiso": {"type": "string"}
                                },
                                "required": ["descripcion", "fecha_compromiso"]
                            }
                        },
                        "resumen": {"type": "string"}
                    },
                    "required": [
                        "why1", "why2", "why3", "why4", "why5",
                        "causas", "control", "correctiva", "resumen"
                    ]
                }
            }

        else:
            prompt = f"""
    Eres especialista en calidad, reclamos y oportunidades de mejora.

    Mejora la redacción técnica del análisis fishbone y de las acciones propuestas.
    Reglas:
    - redacta en tono formal y claro
    - evita culpar personas
    - enfócate en procesos y mejora
    - no inventes hechos
    - debes devolver obligatoriamente fish_metodo, fish_maquinas, fish_materiales, fish_personas, fish_entorno y fish_medicion
    - si ya existen causas, acciones de control o correctivas, mejóralas
    - si no existen, propón una redacción inicial coherente con la observación y el análisis
    - conserva fecha_compromiso solo si ya viene informada

    OBSERVACIÓN:
    {observacion}

    FISHBONE ACTUAL:
    - Método/Proceso: {fish_metodo}
    - Máquinas/Equipos: {fish_maquinas}
    - Materiales/Insumos: {fish_materiales}
    - Personas: {fish_personas}
    - Entorno: {fish_entorno}
    - Medición/Información: {fish_medicion}

    CAUSAS ACTUALES:
    {json.dumps(causas_n, ensure_ascii=False)}

    ACCIONES DE CONTROL ACTUALES:
    {json.dumps(control_n, ensure_ascii=False)}

    ACCIONES CORRECTIVAS ACTUALES:
    {json.dumps(correctiva_n, ensure_ascii=False)}
    """

            schema = {
                "name": "respuesta_om_mejorada_fishbone",
                "strict": True,
                "schema": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "fish_metodo": {"type": "string"},
                        "fish_maquinas": {"type": "string"},
                        "fish_materiales": {"type": "string"},
                        "fish_personas": {"type": "string"},
                        "fish_entorno": {"type": "string"},
                        "fish_medicion": {"type": "string"},
                        "causas": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "descripcion": {"type": "string"},
                                    "fecha_compromiso": {"type": "string"}
                                },
                                "required": ["descripcion", "fecha_compromiso"]
                            }
                        },
                        "control": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "descripcion": {"type": "string"},
                                    "fecha_compromiso": {"type": "string"}
                                },
                                "required": ["descripcion", "fecha_compromiso"]
                            }
                        },
                        "correctiva": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "descripcion": {"type": "string"},
                                    "fecha_compromiso": {"type": "string"}
                                },
                                "required": ["descripcion", "fecha_compromiso"]
                            }
                        },
                        "resumen": {"type": "string"}
                    },
                    "required": [
                        "fish_metodo", "fish_maquinas", "fish_materiales",
                        "fish_personas", "fish_entorno", "fish_medicion",
                        "causas", "control", "correctiva", "resumen"
                    ]
                }
            }

        try:
            completion = client.chat.completions.create(
                model="gpt-5-mini",
                messages=[
                    {
                        "role": "developer",
                        "content": "Responde únicamente con JSON válido que cumpla exactamente el esquema indicado."
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                response_format={
                    "type": "json_schema",
                    "json_schema": schema
                }
            )

            raw = (completion.choices[0].message.content or "").strip()
            current_app.logger.info(">>> RAW OPENAI: %s", raw)

            result = json.loads(raw)

            causas_out = _norm_items(result.get("causas") or [])
            control_out = _norm_items(result.get("control") or [])
            correctiva_out = _norm_items(result.get("correctiva") or [])
            resumen_out = (result.get("resumen") or "").strip()

            payload = {
                "ok": True,
                "causas": causas_out,
                "control": control_out,
                "correctiva": correctiva_out,
                "resumen": resumen_out,
                "raw": raw
            }

            if metodo == "5WHYS":
                payload.update({
                    "why1": (result.get("why1") or ""),
                    "why2": (result.get("why2") or ""),
                    "why3": (result.get("why3") or ""),
                    "why4": (result.get("why4") or ""),
                    "why5": (result.get("why5") or "")
                })
            else:
                payload.update({
                    "fish_metodo": (result.get("fish_metodo") or ""),
                    "fish_maquinas": (result.get("fish_maquinas") or ""),
                    "fish_materiales": (result.get("fish_materiales") or ""),
                    "fish_personas": (result.get("fish_personas") or ""),
                    "fish_entorno": (result.get("fish_entorno") or ""),
                    "fish_medicion": (result.get("fish_medicion") or "")
                })

            current_app.logger.info(">>> IA payload final: %s", payload)
            return jsonify(**payload)

        except Exception as e:
            msg = str(e)

            if "insufficient_quota" in msg:
                return jsonify(
                    ok=False,
                    msg="La IA no está disponible actualmente porque la cuenta API no tiene saldo suficiente."
                ), 400

            current_app.logger.exception("Error IA mejorar respuesta")
            return jsonify(ok=False, msg=f"No se pudo mejorar el texto con IA: {e}"), 500
    
    @app.route("/api/om-chat", methods=["POST"])
    @require_login
    def api_om_chat():
        data = request.get_json(silent=True) or {}
        pregunta = (data.get("pregunta") or "").strip()

        if not pregunta:
            return jsonify(ok=False, error="Pregunta vacía."), 400

        user_id = session.get("user_id") or session.get("usuario_id")

        # Recuperar historial de la sesión (máx. 20 mensajes = 10 turnos)
        historial = session.get("om_chat_historial", [])

        try:
            result = om_chat_responder(pregunta, user_id, historial)

            # Actualizar historial si la respuesta fue exitosa
            if result.get("ok") and result.get("respuesta"):
                historial.append({"role": "user",      "content": pregunta})
                historial.append({"role": "assistant",  "content": result["respuesta"]})
                session["om_chat_historial"] = historial[-20:]   # mantener últimos 10 turnos

            status = 200 if result.get("ok") else 400
            return jsonify(result), status

        except Exception as e:
            current_app.logger.exception("Error en asistente OM")
            return jsonify(
                ok=False,
                error=f"No se pudo procesar la pregunta: {e}"
            ), 500

    @app.route("/api/om/analizar-descripcion", methods=["POST"])
    @require_login
    def api_om_analizar_descripcion():
        """Analiza la descripción de una OM con IA y devuelve feedback por criterio."""
        data      = request.get_json(silent=True) or {}
        texto     = (data.get("texto") or "").strip()
        motivo    = (data.get("motivo") or "").strip()
        submotivo = (data.get("submotivo") or "").strip()

        if not texto or len(texto) < 10:
            return jsonify(ok=False, error="Escribe una descripción antes de analizar."), 400

        try:
            from openai import OpenAI
            import os, json as _json
            client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

            contexto_motivo = ""
            if motivo:
                contexto_motivo = f"\nMotivo seleccionado: {motivo}"
                if submotivo:
                    contexto_motivo += f" / {submotivo}"

            prompt = f"""Eres un experto en gestión de calidad ISO 9001. Analiza la siguiente descripción de una Oportunidad de Mejora (OM / reclamo de cliente) y evalúa cada criterio.{contexto_motivo}

Descripción del usuario:
\"\"\"{texto}\"\"\"

Evalúa estos 7 criterios y responde SOLO con JSON válido:
{{
  "puntaje": <número del 1 al 7 de cuántos criterios se cumplen>,
  "criterios": [
    {{"id": 1, "texto": "¿Describe claramente qué ocurrió?",              "estado": "ok|warn|fail", "nota": "...breve..."}},
    {{"id": 2, "texto": "¿Se enfoca en el proceso, no en culpar personas?","estado": "ok|warn|fail", "nota": "..."}},
    {{"id": 3, "texto": "¿Tiene fecha, contexto o evidencia mínima?",      "estado": "ok|warn|fail", "nota": "..."}},
    {{"id": 4, "texto": "¿Indica el impacto causado?",                     "estado": "ok|warn|fail", "nota": "..."}},
    {{"id": 5, "texto": "¿Permite identificar el proceso responsable?",    "estado": "ok|warn|fail", "nota": "..."}},
    {{"id": 6, "texto": "¿El motivo/submotivo parece correcto?",           "estado": "ok|warn|fail", "nota": "..."}},
    {{"id": 7, "texto": "¿La redacción es profesional y objetiva?",        "estado": "ok|warn|fail", "nota": "..."}}
  ],
  "sugerencia": "Una sugerencia concreta de cómo mejorar la descripción (máx 2 oraciones). Si está bien, di que está lista."
}}

Reglas:
- "ok" = cumple bien el criterio
- "warn" = cumple parcialmente o podría mejorar
- "fail" = no cumple el criterio
- La nota debe ser muy breve (máx 10 palabras)
- No inventes información que no esté en el texto"""

            resp = client.chat.completions.create(
                model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
                messages=[
                    {"role": "system", "content": "Responde únicamente con JSON válido. Sin explicaciones adicionales."},
                    {"role": "user",   "content": prompt},
                ],
                max_tokens=600,
                temperature=0.2,
            )
            raw = resp.choices[0].message.content.strip()
            # Limpiar posibles bloques ```json
            if raw.startswith("```"):
                raw = raw.split("```")[1]
                if raw.startswith("json"):
                    raw = raw[4:]
            result = _json.loads(raw.strip())
            return jsonify(ok=True, **result)

        except Exception as exc:
            current_app.logger.error("api_om_analizar_descripcion error: %s", exc)
            return jsonify(ok=False, error="No se pudo analizar en este momento."), 500

    @app.route("/api/om-chat/debug-sponsor", methods=["GET"])
    @require_login
    def api_om_chat_debug_sponsor():
        """Diagnóstico temporal: muestra qué encuentra en param_groups/param_values."""
        from modules.db import get_db as _get_db
        conn = _get_db()
        cur = conn.cursor()
        out = {}
        try:
            cur.execute("SELECT TOP 5 id, nombre FROM param_groups WHERE nombre LIKE '%SPONSOR%' OR nombre LIKE '%PROCESO%'")
            out["grupos"] = [dict(r) for r in cur.fetchall()]
        except Exception as e:
            out["grupos_error"] = str(e)
        try:
            cur.execute("""
                SELECT TOP 10 pv.id, pv.nombre, pv.valor, pv.parent_id, pv.activo, pg.nombre AS grupo
                FROM param_values pv JOIN param_groups pg ON pg.id = pv.group_id
                WHERE pg.nombre = 'RECL_PROCESO_SPONSOR'
                ORDER BY pv.parent_id, pv.id
            """)
            out["sponsor_values"] = [dict(r) for r in cur.fetchall()]
        except Exception as e:
            out["sponsor_values_error"] = str(e)
        return jsonify(out)

    @app.route("/api/om-chat/reset", methods=["POST"])
    @require_login
    def api_om_chat_reset():
        """Limpia el historial de conversación del asistente OM."""
        session.pop("om_chat_historial", None)
        return jsonify(ok=True, msg="Conversación reiniciada.")
    


    @app.route('/reclamos/api/proceso/<int:proceso_id>/sponsor-principal', methods=['GET'])
    @require_login
    def reclamos_api_proceso_sponsor_principal(proceso_id):
        conn = get_db()
        cur = conn.cursor()
 

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_106, (proceso_id,))

        row = cur.fetchone()
        conn.close()

        if not row:
            return jsonify(ok=False, msg="Este proceso no tiene sponsor principal configurado."), 404

        return jsonify(ok=True, usuario={
            "id": row["id"],
            "nombre": row["nombre"],
            "username": row["username"],
            "departamento": row["departamento"],
            "jefe": row["jefe"],
        })
    

    @app.route('/reclamos/api/proceso/<int:proceso_id>/sponsors')
    @require_login
    def reclamos_api_proceso_sponsors(proceso_id):
        conn = get_db()
        cur = conn.cursor()

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_107, (proceso_id,))

        return jsonify({
            "ok": True,
            "items": [dict(r) for r in cur.fetchall()]
        })

    @app.route('/reclamos/<int:reclamo_id>/carta-cliente', methods=['POST'], endpoint='reclamos_subir_carta_cliente')
    @require_login
    def reclamos_subir_carta_cliente(reclamo_id):
        uid = _current_user_id()

        if not uid:
            return jsonify(ok=False, msg="Sesión inválida."), 401

        conn = get_db()

        if not _can_upload_carta_cliente(conn, uid):
            conn.close()
            return jsonify(ok=False, msg="No autorizado para subir carta al cliente."), 403

        cur = conn.cursor()

        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_SEL_108, (reclamo_id,))

        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="OM no encontrada."), 404

        if int(row["requiere_carta_cliente"] or 0) != 1:
            conn.close()
            return jsonify(ok=False, msg="Esta OM no fue marcada como OM con carta al cliente."), 400

        if "cerrad" not in (row["estado_global"] or ""):
            conn.close()
            return jsonify(ok=False, msg="La carta final solo se puede subir cuando la OM esté cerrada."), 400

        file = request.files.get("carta_cliente")

        if not file or not file.filename:
            conn.close()
            return jsonify(ok=False, msg="Debe seleccionar un archivo."), 400

        filename_original = secure_filename(file.filename or "")
        if not filename_original:
            conn.close()
            return jsonify(ok=False, msg="Nombre de archivo inválido."), 400

        ext = filename_original.rsplit(".", 1)[-1].lower() if "." in filename_original else ""
        if ext not in ("pdf", "doc", "docx"):
            conn.close()
            return jsonify(ok=False, msg="Solo se permite PDF, DOC o DOCX."), 400

        # =========================================================
        # Guarda la carta como adjunto normal de la OM.
        # Usa la misma tabla/carpeta que los demás adjuntos:
        # reclamo_adjuntos + _get_reclamos_upload_folder()
        # =========================================================
        error_adj = _save_adjuntos_for_reclamo(conn, reclamo_id, [file], uid)

        if error_adj:
            conn.rollback()
            conn.close()
            return jsonify(ok=False, msg=error_adj), 400

        # =========================================================
        # SQL Server: usar GETDATE() para evitar error de conversión
        # nvarchar -> datetime por formato regional.
        # =========================================================
        cur.execute(SQL_REGISTER_RECLAMOS_ROUTES_UPD_109, (reclamo_id,))

        conn.commit()
        conn.close()

        return jsonify(ok=True, msg="Carta final al cliente cargada correctamente.")

    # =========================================================
    # GENERAR CARTA CLIENTE PDF (OpenAI + WeasyPrint)
    # =========================================================
    @app.route('/reclamos/<int:reclamo_id>/generar-carta-pdf', methods=['POST'],
               endpoint='reclamos_generar_carta_pdf')
    @require_login
    def reclamos_generar_carta_pdf(reclamo_id):
        from flask import send_file
        import io

        uid = _current_user_id()
        if not uid:
            return jsonify(ok=False, msg="No autenticado"), 401

        conn = get_db()
        if not _can_upload_carta_cliente(conn, uid):
            conn.close()
            return jsonify(ok=False, msg="No autorizado para generar la carta al cliente."), 403

        try:
            from modules.routes_reclamos_pdf import generar_carta_cliente_pdf
            pdf_bytes, filename = generar_carta_cliente_pdf(conn, reclamo_id, uid)
            return send_file(
                io.BytesIO(pdf_bytes),
                mimetype="application/pdf",
                as_attachment=True,
                download_name=filename,
            )
        except ValueError as e:
            return jsonify(ok=False, msg=str(e)), 400
        except Exception as e:
            current_app.logger.error("[generar_carta_pdf] error: %s", e)
            return jsonify(ok=False, msg=f"Error al generar el PDF: {e}"), 500
        finally:
            try:
                conn.close()
            except Exception:
                pass

    # =========================================================
    # VALIDACIÓN DEL CREADOR — aceptar / rechazar respuesta
    # =========================================================
    @app.route('/reclamos/<int:reclamo_id>/validar-creador', methods=['POST'],
               endpoint='reclamos_validar_creador')
    @require_login
    def reclamos_validar_creador(reclamo_id):
        uid = session.get("usuario_id") or session.get("user_id") or session.get("id")
        if not uid:
            return jsonify(ok=False, msg="No autenticado"), 401

        data    = request.get_json(silent=True) or {}
        accion  = (data.get("accion") or "").strip().lower()   # 'aceptar' | 'rechazar'
        motivo  = (data.get("motivo") or "").strip()

        if accion not in ("aceptar", "rechazar"):
            return jsonify(ok=False, msg="Acción inválida."), 400

        conn = get_db()
        cur  = conn.cursor()

        cur.execute(SQL_VALIDAR_CREADOR_SEL_BASE, (reclamo_id,))
        om = cur.fetchone()

        if not om:
            conn.close()
            return jsonify(ok=False, msg="OM no encontrada."), 404

        if int(om["creado_por"] or 0) != int(uid):
            conn.close()
            return jsonify(ok=False, msg="Solo el creador puede validar esta OM."), 403

        estado_global_actual = (om["estado_global"] or "").lower()
        if estado_global_actual != "cerrado":
            conn.close()
            return jsonify(ok=False, msg="Solo se puede validar una OM en estado Cerrado."), 400

        if accion == "aceptar":
            cur.execute(SQL_VALIDAR_CREADOR_UPD_ESTADO, ("aprobado", "cerrado", reclamo_id))
            conn.commit()
            conn.close()
            return jsonify(ok=True, msg="Respuesta aceptada. La OM permanece cerrada.")

        # — rechazar —
        if not motivo:
            conn.close()
            return jsonify(ok=False, msg="Debe indicar el motivo del rechazo."), 400

        cur.execute(SQL_VALIDAR_CREADOR_UPD_ESTADO, ("rechazado", "abierto", reclamo_id))
        cur.execute(SQL_VALIDAR_CREADOR_UPD_IMPUTACION, (reclamo_id,))
        conn.commit()

        creador_row = _get_user_basic(conn, uid)
        creador_nombre = (
            (creador_row.get("nombre_completo") or creador_row.get("username") or f"UID {uid}")
            if creador_row else f"UID {uid}"
        )

        try:
            from modules.scheduler_jobs import enqueue_om_rechazo_creador, ensure_om_evento_templates
            ensure_om_evento_templates(conn)

            try:
                cta_url = url_for("reclamos", _external=True) + "?tab=imputado"
            except Exception:
                cta_url = "http://bitacoraquimpac.com.ec:5000/reclamos?tab=imputado"

            cur2 = conn.cursor()
            cur2.execute(SQL__NOTIFY_COLABORADOR_ASIGNADO_SEL_1, (om["codigo"],))
            r_om = cur2.fetchone()

            base_payload = {
                "codigo":         om["codigo"],
                "rechazado_por":  creador_nombre,
                "motivo_rechazo": motivo,
                "fecha_om":       str(r_om["fecha_reclamo"]) if r_om else "",
                "tipo_om":        str(r_om["tipo_reclamo"])  if r_om else "",
                "tipo_tramite":   str(r_om["tipo_tramite"])  if r_om else "",
                "cliente":        str(r_om["cliente_nombre"])if r_om else "",
                "proceso":        str(r_om["proceso_text"])  if r_om else "",
                "material":       str(r_om["material_desc"]) if (r_om and "material_desc" in r_om.keys()) else "",
                "factura":        str(r_om["factura"])       if r_om else "",
                "observacion":    str(r_om["observacion"])   if r_om else "",
                "cta_url": cta_url,
            }

            enqueued = set()

            def _encolar(dest_uid, dest_nombre):
                if dest_uid in enqueued:
                    return
                enqueued.add(dest_uid)
                p = dict(base_payload)
                p["destinatario_nombre"] = dest_nombre
                enqueue_om_rechazo_creador(conn,
                    user_id=dest_uid,
                    reclamo_id=reclamo_id,
                    payload=p,
                )

            # 1. Sponsors PRINCIPAL + BACKUP
            if om.get("proceso_id"):
                cur2.execute(SQL_VALIDAR_CREADOR_SEL_SPONSORS, (om["proceso_id"],))
                for s in cur2.fetchall():
                    if s["sponsor_id"]:
                        _encolar(s["sponsor_id"], s["sponsor_nombre"] or "")

            # 2. Imputados directos
            cur2.execute(SQL_VALIDAR_CREADOR_SEL_IMPUTADOS, (reclamo_id,))
            for row_i in cur2.fetchall():
                if row_i["imputado_id"]:
                    _encolar(row_i["imputado_id"], row_i["imputado_nombre"] or "")

            # 3. Servicio al Cliente
            cur2.execute(SQL_VALIDAR_CREADOR_SEL_SAC)
            for row_s in cur2.fetchall():
                if row_s["usuario_id"]:
                    _encolar(row_s["usuario_id"], row_s["nombre"] or "")

            # 4. Miembros de equipo
            cur2.execute(SQL_VALIDAR_CREADOR_SEL_EQUIPO, (reclamo_id,))
            for row_e in cur2.fetchall():
                if row_e["usuario_id"]:
                    _encolar(row_e["usuario_id"], row_e["miembro_nombre"] or "")

        except Exception as e:
            current_app.logger.warning(
                "[validar_creador] Error al encolar notificación: %s", e
            )

        conn.close()
        return jsonify(ok=True, msg="Respuesta rechazada. La OM volvió a estado Abierto.")