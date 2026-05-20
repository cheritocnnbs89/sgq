# modules/routes_reclamos.py
# -*- coding: utf-8 -*-
from __future__ import annotations
from flask import app, jsonify, request, session
from .db import get_db
from .security import require_login

from datetime import datetime
from flask import jsonify, session
import os
from openai import OpenAI
 
# routes_reclamos.py
import os
from uuid import uuid4
from werkzeug.utils import secure_filename
from flask import abort
import sqlite3
from datetime import datetime
import smtplib
from email.message import EmailMessage
import io
from flask import (
    render_template, request, redirect, url_for,
    flash, session, current_app, jsonify, send_file
)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from flask import (
    render_template, request, redirect, url_for,
    flash, session, current_app, jsonify
)

from .db import get_db, get_config_value
from .security import (
    require_login,
    require_permission,
)

# =========================================================
# Helpers internos
# =========================================================

BOSS_ROLES = (
    'jefe',
    'gerente',
    'gerente general',
    'gerente financiero',
    'coordinador',
    'admin'
)


 
# ----------------------------------------------
# Helpers
# ----------------------------------------------
import sqlite3

def _raw_conn(conn):
    return getattr(conn, "_conn", conn)

def _is_sqlserver_conn(conn) -> bool:
    raw = _raw_conn(conn)
    mod = getattr(raw.__class__, "__module__", "").lower()
    name = getattr(raw.__class__, "__name__", "").lower()
    text = f"{mod} {name}"
    return "pyodbc" in text or "odbc" in text

def _table_exists(conn, table: str) -> bool:
    cur = conn.cursor()

    if _is_sqlserver_conn(conn):
        cur.execute("""
            SELECT 1
            FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_NAME = ?
        """, (table,))
        return cur.fetchone() is not None

    cur.execute("""
        SELECT name
        FROM sqlite_master
        WHERE type = 'table' AND name = ?
    """, (table,))
    return cur.fetchone() is not None

def _col_names(conn, table: str) -> set[str]:
    cur = conn.cursor()

    try:
        if _is_sqlserver_conn(conn):
            cur.execute("""
                SELECT COLUMN_NAME AS name
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_NAME = ?
            """, (table,))
            return {str(r["name"]) for r in cur.fetchall()}

        cur.execute(f"PRAGMA table_info({table})")
        rows = cur.fetchall()
        names = set()
        for r in rows:
            try:
                names.add(str(r["name"]))
            except Exception:
                names.add(str(r[1]))
        return names
    except Exception:
        return set()

def _ensure_column(conn, table, column, decl_sql):
    # En SQL Server no alteramos esquema desde runtime
    if _is_sqlserver_conn(conn):
        return

    cur = conn.execute(f"PRAGMA table_info({table})")
    cols = {row["name"] for row in cur.fetchall()}
    if column not in cols:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {decl_sql}")
        conn.commit()

def _run_reclamos_bootstrap_if_needed(conn):
    if _is_sqlserver_conn(conn):
        return

    ensure_reclamos_schema(conn)
    ensure_reclamos_catalogos(conn)
    ensure_geo_schema(conn)
    _ensure_reclamo_imputados_extra_cols(conn)
    
def _can_export_all_reclamos(conn, uid: int | None) -> bool:
    """
    Puede exportar TODO:
    - rol admin / coordinador
    - o si pertenece a Servicio al Cliente (por puesto o por departamento)
    """
    role = (session.get("rol") or "").strip().lower()
    if role in {"admin", "coordinador"}:
        return True

    if not uid:
        return False

    cur = conn.cursor()

    # Traer puesto y departamento del usuario
    # Ajusta nombres de columnas según tu esquema:
    # - u.departamento_nombre (si existe)
    # - o dep.nombre si tienes tabla departamentos
    cur.execute("""
        SELECT
            UPPER(TRIM(COALESCE(p.nombre,''))) AS puesto_nombre
           
        FROM usuarios u
        LEFT JOIN puestos p ON p.id = u.puesto_id
        WHERE u.id = ?
        LIMIT 1
    """, (uid,))
    row = cur.fetchone()
    if not row:
        return False
    
    

    puesto = (row["puesto_nombre"] or "").strip()
 
    # ✅ Regla Servicio al Cliente:
    # - por departamento
    # - o por puesto (coordinador servicio al cliente)
    if "SERVICIO AL CLIENTE" in puesto:
        return True

    return False


def _sponsor_acciones_upload_dir() -> str:
    """
    Carpeta física para evidencias de acciones del sponsor.
    Ajusta aquí si ya manejas otra ruta base de uploads.
    """
    base = os.path.join(current_app.root_path, "uploads", "reclamo_imputado_acciones")
    os.makedirs(base, exist_ok=True)
    return base

def _puede_gestionar_imputado_accion(accion_id: int, user_id: int | None = None):
    if not user_id:
        user_id = _current_user_id()

    if not user_id:
        return False, None

    db = get_db()
    ensure_reclamos_schema(db)
    ensure_reclamo_imputado_acciones_schema(db)

    row = db.execute("""
        SELECT
            a.id,
            a.imputacion_id,
            a.reclamo_id,
            COALESCE(a.activo, 1) AS accion_activa,
            COALESCE(a.cumplido, 0) AS cumplido,
            ri.imputado_id,
            COALESCE(ri.estado_asignacion, '') AS estado_asignacion
        FROM reclamo_imputado_acciones a
        JOIN reclamo_imputados ri
          ON ri.id = a.imputacion_id
        WHERE a.id = ?
        LIMIT 1
    """, (accion_id,)).fetchone()

    if not row:
        return False, None

    if _is_admin_like():
        return True, row

    es_duenio = int(row["imputado_id"] or 0) == int(user_id or 0)
    aprobada = (row["estado_asignacion"] or "").strip().lower() == "aprobado"

    return bool(es_duenio and aprobada), row
 

def _can_view_all_reclamos(conn, uid: int | None) -> bool:
    # 1) Admin-like => todo
    if _is_admin_like():
        return True

    # 2) Roles futuros (y puedes incluir otros si quieres)
    role = (session.get('rol') or '').strip().lower()
    if role in ('gerente', 'gerente financiero'):
        return True

    # 3) Por departamento: Servicio al Cliente
    if not uid:
        return False

    cur = conn.cursor()

    # Detecta columnas disponibles en usuarios (para no romper si cambia tu esquema)
    cur.execute("PRAGMA table_info(usuarios)")
    ucols = {r["name"] for r in cur.fetchall()}

    # A) Si existe departamento_id + tabla departamentos
    if "departamento_id" in ucols:
        try:
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='departamentos'")
            if cur.fetchone():
                cur.execute("""
                    SELECT d.nombre
                    FROM usuarios u
                    JOIN departamentos d ON d.id = u.departamento_id
                    WHERE u.id = ?
                    LIMIT 1
                """, (uid,))
                row = cur.fetchone()
                if row and (row["nombre"] or "").strip().lower() == "servicio al cliente":
                    return True
        except Exception:
            pass  # fallback abajo

    # B) Si en usuarios tienes un texto directo tipo usuarios.departamento
    if "departamento" in ucols:
        try:
            cur.execute("SELECT departamento FROM usuarios WHERE id = ? LIMIT 1", (uid,))
            row = cur.fetchone()
            if row and (row["departamento"] or "").strip().lower() == "servicio al cliente":
                return True
        except Exception:
            pass

    # C) Fallback por puesto (tú ya lo usas): si el puesto contiene “SERVICIO AL CLIENTE”
    if "puesto_id" in ucols:
        try:
            cur.execute("""
                SELECT p.nombre
                FROM usuarios u
                LEFT JOIN puestos p ON p.id = u.puesto_id
                WHERE u.id = ?
                LIMIT 1
            """, (uid,))
            row = cur.fetchone()
            if row and "SERVICIO AL CLIENTE" in ((row["nombre"] or "").strip().upper()):
                return True
        except Exception:
            pass

    return False

def _can_view_all_reclamos_sn_sponsor(conn, uid: int | None) -> bool:
    """
    Regla para la pestaña "Soy Sponsor":
    - Admin-like NO aplica si el rol es "gerente" (para que NO vea todos como sponsor).
    - Gerente financiero sí puede ver todo (si así lo decides).
    - Servicio al Cliente sí puede ver todo (por depto/puesto).
    """
    role = (session.get('rol') or '').strip().lower()

    # 🚫 Bloqueo explícito: un gerente NO debe ver todo en "Soy Sponsor"
    # (aunque _is_admin_like() lo considere "admin-like")
    if role == "gerente":
        return False

    # ✅ Admin-like (pero ya excluimos gerente arriba)
    if _is_admin_like():
        return True

    # ✅ Roles permitidos a ver todo en "Soy Sponsor"
    if role in ('gerente financiero',):
        return True

    if not uid:
        return False

    cur = conn.cursor()

    # Detecta columnas disponibles en usuarios (para no romper si cambia tu esquema)
    cur.execute("PRAGMA table_info(usuarios)")
    ucols = {r["name"] for r in cur.fetchall()}

    # A) Si existe departamento_id + tabla departamentos
    if "departamento_id" in ucols:
        try:
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='departamentos'")
            if cur.fetchone():
                cur.execute("""
                    SELECT d.nombre
                    FROM usuarios u
                    JOIN departamentos d ON d.id = u.departamento_id
                    WHERE u.id = ?
                    LIMIT 1
                """, (uid,))
                row = cur.fetchone()
                if row and (row["nombre"] or "").strip().lower() == "servicio al cliente":
                    return True
        except Exception:
            pass

    # B) Si en usuarios tienes un texto directo tipo usuarios.departamento
    if "departamento" in ucols:
        try:
            cur.execute("SELECT departamento FROM usuarios WHERE id = ? LIMIT 1", (uid,))
            row = cur.fetchone()
            if row and (row["departamento"] or "").strip().lower() == "servicio al cliente":
                return True
        except Exception:
            pass

    # C) Fallback por puesto
    if "puesto_id" in ucols:
        try:
            cur.execute("""
                SELECT p.nombre
                FROM usuarios u
                LEFT JOIN puestos p ON p.id = u.puesto_id
                WHERE u.id = ?
                LIMIT 1
            """, (uid,))
            row = cur.fetchone()
            if row and "SERVICIO AL CLIENTE" in ((row["nombre"] or "").strip().upper()):
                return True
        except Exception:
            pass

    return False
def _es_miembro_equipo_reclamo(reclamo_id, user_id):
    # 🔎 DEBUG: ver si está llegando bien
    current_app.logger.debug(
        "[equipo] Chequeando miembro equipo: reclamo_id=%s user_id=%s",
        reclamo_id, user_id
    )

    if not user_id or not reclamo_id:
        current_app.logger.debug("[equipo] -> False (falta user_id o reclamo_id)")
        return False

    db = get_db()
    row = db.execute("""
        SELECT 1
        FROM reclamo_equipo_respuestas
        WHERE reclamo_id = ?
          AND usuario_id = ?
          AND activo = 1
        LIMIT 1
    """, (reclamo_id, user_id)).fetchone()

    current_app.logger.debug(
        "[equipo] Resultado query miembro equipo: row=%r -> %s",
        row, bool(row)
    )

    return row is not None
def _notify_sponsor_respuesta_equipo(conn, imputacion_id: int, miembro_id: int, reclamo_codigo: str):
    """
    Notifica al sponsor/imputado de la OM que un miembro del equipo ya respondió.
    """

    cur = conn.cursor()

    # Sponsor / imputado dueño de la imputación
    cur.execute("""
        SELECT
            ri.imputado_id,
            COALESCE(us.nombre_completo, us.username) AS sponsor_nombre,
            us.email AS sponsor_email,
            COALESCE(um.nombre_completo, um.username) AS miembro_nombre,
            um.username AS miembro_username,
            r.id AS reclamo_id,
            r.codigo,
            r.tipo_reclamo,
            r.tipo_tramite,
            r.cliente_nombre,
            r.proceso_text,
            r.observacion,
            r.antecedente
        FROM reclamo_imputados ri
        JOIN reclamos r
          ON r.id = ri.reclamo_id
        LEFT JOIN usuarios us
          ON us.id = ri.imputado_id
        LEFT JOIN usuarios um
          ON um.id = ?
        WHERE ri.id = ?
        LIMIT 1
    """, (miembro_id, imputacion_id))
    row = cur.fetchone()

    if not row:
        current_app.logger.warning(
            "[MAIL_EQUIPO] No se encontró imputación/sponsor. imputacion_id=%s miembro_id=%s",
            imputacion_id, miembro_id
        )
        return

    sponsor_email = row["sponsor_email"]
    if not sponsor_email:
        current_app.logger.warning(
            "[MAIL_EQUIPO] Sponsor sin correo. imputacion_id=%s sponsor_id=%s",
            imputacion_id, row["imputado_id"]
        )
        return

    sponsor_nombre = row["sponsor_nombre"] or "Usuario"
    miembro_nombre = row["miembro_nombre"] or row["miembro_username"] or f"UID {miembro_id}"

    try:
        link_sponsor = url_for("reclamos", _external=True) + "?tab=sponsor"
    except Exception:
        link_sponsor = "https://tu-sistema/reclamos?tab=sponsor"

    subject = f"[Oportunidad de Mejora] Respuesta registrada por miembro de equipo en {reclamo_codigo}"

    text_body = f"""Hola {sponsor_nombre},

El miembro de equipo {miembro_nombre} registró su respuesta de apoyo para la Oportunidad de Mejora {reclamo_codigo}.

Resumen:
- Cliente: {row["cliente_nombre"] or ""}
- Tipo de OM: {row["tipo_reclamo"] or ""}
- Tipo de trámite: {row["tipo_tramite"] or ""}
- Proceso: {row["proceso_text"] or ""}

Por favor ingresa al sistema y revisa el aporte en la pestaña "Soy Sponsor".

Ir al sistema: {link_sponsor}

Este es un mensaje automático.
"""


    def _row_html(lbl, val):
        val = "" if val is None else str(val)
        val = val.replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#dbeafe;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    html_body = f"""\
    <!doctype html>
    <html>
    <body style="margin:0;padding:0;background:#f3f4f6;font-family:Segoe UI,Arial,sans-serif;">
        <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
            style="background:#f3f4f6;padding:24px 0;">
        <tr>
            <td align="center">
            <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                    style="max-width:720px;background:#ffffff;border-radius:8px;
                            border:1px solid #e5e7eb;overflow:hidden;">
                <tr>
                <td style="background:#2563eb;padding:16px 20px;color:#ffffff;">
                    <div style="font-size:12px;text-transform:uppercase;letter-spacing:.08em;opacity:.9;">
                    Oportunidad de Mejora
                    </div>
                    <div style="font-size:18px;font-weight:700;margin-top:4px;">
                    Respuesta registrada por miembro de equipo
                    </div>
                    <div style="font-size:12px;opacity:.9;margin-top:6px;">
                    Hola {sponsor_nombre}, el usuario <strong>{miembro_nombre}</strong> ya registró su aporte para la OM {reclamo_codigo}.
                    </div>
                </td>
                </tr>

                <tr>
                <td style="padding:18px 20px 10px 20px;">
                    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                        style="border-collapse:collapse;">
                    {_row_html('Código', row['codigo'])}
                    {_row_html('Miembro que respondió', miembro_nombre)}
                    {_row_html('Cliente', row['cliente_nombre'])}
                    {_row_html('Tipo de OM', row['tipo_reclamo'])}
                    {_row_html('Tipo de trámite', row['tipo_tramite'])}
                    {_row_html('Proceso', row['proceso_text'])}
                    {_row_html('Antecedente', row['antecedente'])}
                    {_row_html('Observación', row['observacion'])}
                    </table>

                    <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                    <a href="{link_sponsor}"
                        style="display:inline-block;background:#2563eb;color:#ffffff;
                                text-decoration:none;padding:10px 18px;border-radius:6px;
                                font-weight:600;font-size:13px;">
                        Revisar aporte del equipo
                    </a>
                    </div>

                    <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                    Ingresa al módulo de reclamos y revisa la pestaña <strong>“Soy Sponsor”</strong>.
                    </div>
                </td>
                </tr>

                <tr>
                <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                            font-size:11px;color:#9ca3af;">
                    Este es un mensaje automático. No responda a este correo.
                </td>
                </tr>
            </table>
            </td>
        </tr>
        </table>
    </body>
    </html>
    """

    _send_mail_safe(sponsor_email, subject, text_body, html_body=html_body)

def _now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _current_user_id() -> int | None:
    return (
        session.get("user_id")
        or session.get("usuario_id")
        or session.get("id")
    )

def _get_respuesta_equipo_acciones_full(db, respuesta_equipo_id: int):
    rows = db.execute("""
         
        SELECT
            a.id,
            a.tipo,
            a.descripcion,
            a.fecha_compromiso,
            COALESCE(a.orden, 1) AS orden,
            COALESCE(a.requiere_evidencia, 0) AS requiere_evidencia,
            COALESCE(a.cumplido, 0) AS cumplido,
            COALESCE(a.fecha_cumplimiento, '') AS fecha_cumplimiento,
            COALESCE(a.observacion_cumplimiento, '') AS observacion_cumplimiento
        FROM reclamo_respuesta_equipo_acciones a
        WHERE a.respuesta_equipo_id = ?
          AND COALESCE(a.activo, 1) = 1
        ORDER BY
            CASE a.tipo
                WHEN 'CAUSA' THEN 1
                WHEN 'CONTROL' THEN 2
                WHEN 'CORRECTIVA' THEN 3
                ELSE 9
            END,
            a.orden,
            a.id
    """, (respuesta_equipo_id,)).fetchall()

    causas = []
    control = []
    correctiva = []

    for r in rows:
        accion_id = int(r["id"])

        evid_rows = db.execute("""
            SELECT
                e.id,
                e.filename,
                e.original_name,
                COALESCE(e.content_type, '') AS content_type,
                COALESCE(e.size_bytes, 0) AS size_bytes,
                COALESCE(e.created_at, '') AS created_at
            FROM reclamo_respuesta_equipo_accion_evidencias e
            WHERE e.accion_id = ?
              AND COALESCE(e.activo, 1) = 1
            ORDER BY e.id
        """, (accion_id,)).fetchall()

        item = {
            "id": accion_id,
            "tipo": (r["tipo"] or "").upper(),
            "descripcion": r["descripcion"] or "",
            "fecha_compromiso": r["fecha_compromiso"] or "",
            "orden": int(r["orden"] or 1),
            "requiere_evidencia": int(r["requiere_evidencia"] or 0),
            "cumplido": int(r["cumplido"] or 0),
            "fecha_cumplimiento": r["fecha_cumplimiento"] or "",
            "evidencias": [dict(x) for x in evid_rows],
            "observacion_cumplimiento": r["observacion_cumplimiento"] or "",        }

        tipo = item["tipo"]
        if tipo == "CAUSA":
            causas.append(item)
        elif tipo == "CONTROL":
            control.append(item)
        elif tipo == "CORRECTIVA":
            correctiva.append(item)

    return {
        "causas": causas,
        "control": control,
        "correctiva": correctiva,
    }
def _is_admin_like() -> bool:
    rol = (session.get("rol") or "").strip().lower()
    return rol in ("admin", "coordinador")


def _notify_reclamo_adjuntos_change(conn, reclamo_id: int, actor_id: int | None,
                                    accion: str, filenames: list[str]):
    """
    accion: 'agregados' o 'eliminados'
    filenames: lista de nombres originales afectados
    """
    cur = conn.cursor()
    cur.execute("""
        SELECT codigo, creado_por
        FROM reclamos
        WHERE id = ?
    """, (reclamo_id,))
    r = cur.fetchone()
    if not r:
        return

    codigo = r["codigo"]
    creador_id = r["creado_por"]

    creador = _get_user_basic(conn, creador_id)
    actor = _get_user_basic(conn, actor_id) if actor_id else None

    if not creador or ("email" not in creador.keys()) or not creador["email"]:
        return
 
    actor_name = (
        actor["nombre_completo"]
        if actor and "nombre_completo" in actor.keys() and actor["nombre_completo"]
        else (actor["username"] if actor else "Sistema")
    )

    lista = "\n".join(f"- {fn}" for fn in filenames) or "(sin detalle)"

    subject = f"[Oportunidad de Mejora] Adjuntos {accion} en {codigo}"
    text_body = f"""Hola {creador['username']},

Se han {accion} archivos en la Oportunidad de Mejora {codigo}.

Acción realizada por: {actor_name}

Archivos:
{lista}

Este es un mensaje automático.
"""

    _send_mail_safe(creador["email"], subject, text_body)

# =========================================================
#   ESQUEMA DE BD
# =========================================================

def _col_names(conn, table: str) -> set[str]:
    cur = conn.cursor()
    try:
        cur.execute(f"PRAGMA table_info({table})")
        return {r["name"] for r in cur.fetchall()}
    except sqlite3.OperationalError:
        return set()


def ensure_reclamos_schema(conn: sqlite3.Connection):
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS productos (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo    TEXT UNIQUE,
            nombre    TEXT NOT NULL,
            activo    INTEGER NOT NULL DEFAULT 1
        );
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS reclamo_equipo_respuestas (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            reclamo_id      INTEGER NOT NULL,
            imputacion_id   INTEGER,
            usuario_id      INTEGER NOT NULL,
            puede_responder INTEGER NOT NULL DEFAULT 1,
            activo          INTEGER NOT NULL DEFAULT 1,
            creado_por      INTEGER,
            creado_at       TEXT,
            FOREIGN KEY (reclamo_id)    REFERENCES reclamos(id),
            FOREIGN KEY (imputacion_id) REFERENCES reclamo_imputados(id),
            FOREIGN KEY (usuario_id)    REFERENCES usuarios(id)
        );
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS reclamos(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE,
            fecha_reclamo TEXT,
            fecha_creacion TEXT,
            cliente_id INTEGER,
            cliente_nombre TEXT,
            cliente_identificacion TEXT,
            cliente_direccion TEXT,
            cliente_contacto TEXT,
            cliente_email TEXT,
            cliente_telefono TEXT,

            region_id INTEGER,
            provincia_id INTEGER,
            canton_id INTEGER,

            tipo_tramite TEXT,
            tipo_reclamo TEXT,

            proceso_text TEXT,
            antecedente TEXT,
            fecha_pedido TEXT,
            factura TEXT,
            guia_remision TEXT,

            material_id INTEGER,
            material_desc TEXT,

            persona_atendio TEXT,
            persona_atendio_cedula TEXT,

            fecha_ofrec_entrega TEXT,
            fecha_entrega TEXT,

            observacion TEXT,

            procede TEXT,
            creado_por INTEGER,
            estado_global TEXT
        )
    """)

    cols_rec = _col_names(conn, "reclamos")
    if "procede" not in cols_rec:
        cur.execute("ALTER TABLE reclamos ADD COLUMN procede TEXT")
    if "estado_global" not in cols_rec:
        cur.execute("ALTER TABLE reclamos ADD COLUMN estado_global TEXT")

    cur.execute("""
        CREATE TABLE IF NOT EXISTS reclamo_imputados(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reclamo_id INTEGER NOT NULL,
            imputado_id INTEGER NOT NULL,
            aprobador_id INTEGER,

            estado_asignacion TEXT,
            fecha_aprobacion_asignacion TEXT,
            fecha_rechazo_asignacion TEXT,
            motivo_rechazo_asignacion TEXT,

            respuesta_causa TEXT,
            respuesta_preventiva TEXT,
            respuesta_correctiva TEXT,
            fecha_respuesta_imputado TEXT,

            estado_respuesta TEXT,
            fecha_aprobacion_respuesta TEXT,
            fecha_rechazo_respuesta TEXT,
            motivo_rechazo_respuesta TEXT,

            visible_creador INTEGER DEFAULT 0
        )
    """)

    cols_imp = _col_names(conn, "reclamo_imputados")
    if "visible_creador" not in cols_imp:
        cur.execute("ALTER TABLE reclamo_imputados ADD COLUMN visible_creador INTEGER DEFAULT 0")

    # =========================================================
    # NUEVO: acciones múltiples por imputación principal
    # =========================================================
    cur.execute("""
        CREATE TABLE IF NOT EXISTS reclamo_imputado_acciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            imputacion_id INTEGER NOT NULL,
            reclamo_id INTEGER NOT NULL,

            tipo TEXT NOT NULL,                 -- CAUSA | CONTROL | CORRECTIVA
            descripcion TEXT NOT NULL,
            fecha_compromiso TEXT,              -- YYYY-MM-DD
            orden INTEGER NOT NULL DEFAULT 1,

            requiere_evidencia INTEGER NOT NULL DEFAULT 0,
            cumplido INTEGER NOT NULL DEFAULT 0,
            fecha_cumplimiento TEXT,

            reminder_3d_sent INTEGER NOT NULL DEFAULT 0,
            reminder_2d_sent INTEGER NOT NULL DEFAULT 0,
            reminder_1d_sent INTEGER NOT NULL DEFAULT 0,
            escalado_jefe INTEGER NOT NULL DEFAULT 0,

            activo INTEGER NOT NULL DEFAULT 1,
            created_at TEXT,
            created_by INTEGER,
            updated_at TEXT,
            updated_by INTEGER,

            FOREIGN KEY (imputacion_id) REFERENCES reclamo_imputados(id),
            FOREIGN KEY (reclamo_id) REFERENCES reclamos(id)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS reclamo_accion_evidencias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            accion_id INTEGER NOT NULL,

            filename TEXT NOT NULL,
            original_name TEXT NOT NULL,
            content_type TEXT,
            size_bytes INTEGER,

            creado_por INTEGER,
            created_at TEXT,
            activo INTEGER NOT NULL DEFAULT 1,

            FOREIGN KEY (accion_id) REFERENCES reclamo_imputado_acciones(id)
        )
    """)

    # Índices tabla de acciones
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_ria_imputacion
        ON reclamo_imputado_acciones(imputacion_id)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_ria_reclamo
        ON reclamo_imputado_acciones(reclamo_id)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_ria_tipo
        ON reclamo_imputado_acciones(tipo)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_ria_fecha_compromiso
        ON reclamo_imputado_acciones(fecha_compromiso)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_ria_cumplido
        ON reclamo_imputado_acciones(cumplido, activo)
    """)

    # Índices tabla de evidencias
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_rae_accion
        ON reclamo_accion_evidencias(accion_id)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_rae_activo
        ON reclamo_accion_evidencias(activo)
    """)

    ensure_reclamo_imputados_fishbone(conn)
    ensure_reclamo_respuestas_equipo_schema(conn)
    ensure_reclamo_imputado_acciones_schema(conn)
    ensure_reclamo_respuesta_equipo_acciones_schema(conn)
    conn.commit()


def ensure_reclamo_imputado_acciones_schema(conn: sqlite3.Connection):
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS reclamo_imputado_acciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            imputacion_id INTEGER NOT NULL,
            reclamo_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            descripcion TEXT NOT NULL,
            fecha_compromiso TEXT,
            orden INTEGER NOT NULL DEFAULT 1,
            requiere_evidencia INTEGER NOT NULL DEFAULT 0,
            cumplido INTEGER NOT NULL DEFAULT 0,
            fecha_cumplimiento TEXT,
            reminder_3d_sent INTEGER NOT NULL DEFAULT 0,
            reminder_2d_sent INTEGER NOT NULL DEFAULT 0,
            reminder_1d_sent INTEGER NOT NULL DEFAULT 0,
            escalado_jefe INTEGER NOT NULL DEFAULT 0,
            activo INTEGER NOT NULL DEFAULT 1,
            created_at TEXT,
            created_by INTEGER,
            updated_at TEXT,
            updated_by INTEGER,
            observacion_cumplimiento TEXT,
            FOREIGN KEY (imputacion_id) REFERENCES reclamo_imputados(id),
            FOREIGN KEY (reclamo_id) REFERENCES reclamos(id)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS reclamo_accion_evidencias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            accion_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            original_name TEXT NOT NULL,
            content_type TEXT,
            size_bytes INTEGER,
            creado_por INTEGER,
            created_at TEXT,
            activo INTEGER NOT NULL DEFAULT 1,
            FOREIGN KEY (accion_id) REFERENCES reclamo_imputado_acciones(id)
        )
    """)

    cur.execute("PRAGMA table_info(reclamo_imputado_acciones)")
    cols = {r[1] for r in cur.fetchall()}

    required_cols = {
        "imputacion_id": "INTEGER NOT NULL DEFAULT 0",
        "reclamo_id": "INTEGER NOT NULL DEFAULT 0",
        "tipo": "TEXT NOT NULL DEFAULT ''",
        "descripcion": "TEXT NOT NULL DEFAULT ''",
        "fecha_compromiso": "TEXT",
        "orden": "INTEGER NOT NULL DEFAULT 1",
        "requiere_evidencia": "INTEGER NOT NULL DEFAULT 0",
        "cumplido": "INTEGER NOT NULL DEFAULT 0",
        "fecha_cumplimiento": "TEXT",
        "reminder_3d_sent": "INTEGER NOT NULL DEFAULT 0",
        "reminder_2d_sent": "INTEGER NOT NULL DEFAULT 0",
        "reminder_1d_sent": "INTEGER NOT NULL DEFAULT 0",
        "escalado_jefe": "INTEGER NOT NULL DEFAULT 0",
        "activo": "INTEGER NOT NULL DEFAULT 1",
        "created_at": "TEXT",
        "created_by": "INTEGER",
        "updated_at": "TEXT",
        "updated_by": "INTEGER",

        # NUEVA COLUMNA
        "observacion_cumplimiento": "TEXT"
    }

    for col, ddl in required_cols.items():
        if col not in cols:
            cur.execute(f"ALTER TABLE reclamo_imputado_acciones ADD COLUMN {col} {ddl}")

    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_ria_imputacion
        ON reclamo_imputado_acciones(imputacion_id)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_ria_reclamo
        ON reclamo_imputado_acciones(reclamo_id)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_ria_tipo
        ON reclamo_imputado_acciones(tipo)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_ria_fecha_compromiso
        ON reclamo_imputado_acciones(fecha_compromiso)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_ria_cumplido
        ON reclamo_imputado_acciones(cumplido, activo)
    """)

    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_rae_accion
        ON reclamo_accion_evidencias(accion_id)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_rae_activo
        ON reclamo_accion_evidencias(activo)
    """)

    conn.commit()
def _save_respuesta_equipo_acciones(
    conn: sqlite3.Connection,
    respuesta_id: int,
    reclamo_id: int,
    imputacion_id: int,
    miembro_id: int,
    causas: list,
    control: list,
    correctiva: list,
    user_id: int | None,
):
    cur = conn.cursor()
    now = _now_iso()

    # Inactivar acciones anteriores de esta respuesta
    cur.execute("""
        UPDATE reclamo_respuesta_equipo_acciones
        SET activo = 0,
            updated_at = ?,
            updated_by = ?
        WHERE respuesta_equipo_id = ?
          AND activo = 1
    """, (now, user_id, respuesta_id))

    def _normalizar_items(items):
        normalizados = []

        for item in (items or []):
            if isinstance(item, dict):
                desc = (item.get("descripcion") or "").strip()
                fecha = (item.get("fecha_compromiso") or "").strip()
            elif isinstance(item, (list, tuple)) and len(item) >= 2:
                desc = (item[0] or "").strip()
                fecha = (item[1] or "").strip()
            else:
                continue

            if not desc and not fecha:
                continue

            normalizados.append((desc, fecha))

        return normalizados

    def _insert_many(tipo: str, items: list, requiere_evidencia: int = 0):
        items_norm = _normalizar_items(items)

        for idx, (desc, fecha) in enumerate(items_norm, start=1):
            cur.execute("""
                INSERT INTO reclamo_respuesta_equipo_acciones (
                    respuesta_equipo_id,
                    reclamo_id,
                    imputacion_id,
                    miembro_id,
                    tipo,
                    descripcion,
                    fecha_compromiso,
                    orden,
                    requiere_evidencia,
                    activo,
                    created_at,
                    created_by
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
            """, (
                respuesta_id,
                reclamo_id,
                imputacion_id,
                miembro_id,
                tipo,
                desc,
                fecha,
                idx,
                requiere_evidencia,
                now,
                user_id
            ))

    _insert_many("CAUSA", causas, 0)
    _insert_many("CONTROL", control, 0)
    _insert_many("CORRECTIVA", correctiva, 1)
 



def ensure_reclamo_respuesta_equipo_acciones_schema(conn: sqlite3.Connection):
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS reclamo_respuesta_equipo_acciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,

            respuesta_equipo_id INTEGER NOT NULL,
            reclamo_id INTEGER NOT NULL,
            imputacion_id INTEGER NOT NULL,
            miembro_id INTEGER NOT NULL,

            tipo TEXT NOT NULL,                 -- CAUSA | CONTROL | CORRECTIVA
            descripcion TEXT NOT NULL,
            fecha_compromiso TEXT,              -- YYYY-MM-DD
            orden INTEGER NOT NULL DEFAULT 1,

            requiere_evidencia INTEGER NOT NULL DEFAULT 0,
            cumplido INTEGER NOT NULL DEFAULT 0,
            fecha_cumplimiento TEXT,

            reminder_3d_sent INTEGER NOT NULL DEFAULT 0,
            reminder_2d_sent INTEGER NOT NULL DEFAULT 0,
            reminder_1d_sent INTEGER NOT NULL DEFAULT 0,
            escalado_jefe INTEGER NOT NULL DEFAULT 0,

            activo INTEGER NOT NULL DEFAULT 1,
            created_at TEXT,
            created_by INTEGER,
            updated_at TEXT,
            updated_by INTEGER,

            FOREIGN KEY (respuesta_equipo_id) REFERENCES reclamo_respuestas_equipo(id),
            FOREIGN KEY (reclamo_id) REFERENCES reclamos(id),
            FOREIGN KEY (imputacion_id) REFERENCES reclamo_imputados(id),
            FOREIGN KEY (miembro_id) REFERENCES usuarios(id)
        );
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS reclamo_respuesta_equipo_accion_evidencias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            accion_id INTEGER NOT NULL,

            filename TEXT NOT NULL,
            original_name TEXT NOT NULL,
            content_type TEXT,
            size_bytes INTEGER,

            creado_por INTEGER,
            created_at TEXT,
            activo INTEGER NOT NULL DEFAULT 1,

            FOREIGN KEY (accion_id) REFERENCES reclamo_respuesta_equipo_acciones(id)
        );
    """)

    cur.execute("PRAGMA table_info(reclamo_respuesta_equipo_acciones)")
    cols = {r[1] for r in cur.fetchall()}

    required_cols = {
        "respuesta_equipo_id": "INTEGER NOT NULL DEFAULT 0",
        "reclamo_id": "INTEGER NOT NULL DEFAULT 0",
        "imputacion_id": "INTEGER NOT NULL DEFAULT 0",
        "miembro_id": "INTEGER NOT NULL DEFAULT 0",
        "tipo": "TEXT NOT NULL DEFAULT ''",
        "descripcion": "TEXT NOT NULL DEFAULT ''",
        "fecha_compromiso": "TEXT",
        "orden": "INTEGER NOT NULL DEFAULT 1",
        "requiere_evidencia": "INTEGER NOT NULL DEFAULT 0",
        "cumplido": "INTEGER NOT NULL DEFAULT 0",
        "fecha_cumplimiento": "TEXT",
        "reminder_3d_sent": "INTEGER NOT NULL DEFAULT 0",
        "reminder_2d_sent": "INTEGER NOT NULL DEFAULT 0",
        "reminder_1d_sent": "INTEGER NOT NULL DEFAULT 0",
        "escalado_jefe": "INTEGER NOT NULL DEFAULT 0",
        "activo": "INTEGER NOT NULL DEFAULT 1",
        "created_at": "TEXT",
        "created_by": "INTEGER",
        "updated_at": "TEXT",
        "updated_by": "INTEGER",
    }

    for col, ddl in required_cols.items():
        if col not in cols:
            cur.execute(f"ALTER TABLE reclamo_respuesta_equipo_acciones ADD COLUMN {col} {ddl}")

    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_rrea_respuesta
        ON reclamo_respuesta_equipo_acciones(respuesta_equipo_id)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_rrea_reclamo
        ON reclamo_respuesta_equipo_acciones(reclamo_id)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_rrea_imputacion
        ON reclamo_respuesta_equipo_acciones(imputacion_id)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_rrea_miembro
        ON reclamo_respuesta_equipo_acciones(miembro_id)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_rrea_tipo
        ON reclamo_respuesta_equipo_acciones(tipo)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_rrea_fecha_compromiso
        ON reclamo_respuesta_equipo_acciones(fecha_compromiso)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_rrea_cumplido
        ON reclamo_respuesta_equipo_acciones(cumplido, activo)
    """)

    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_rree_accion
        ON reclamo_respuesta_equipo_accion_evidencias(accion_id)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_rree_activo
        ON reclamo_respuesta_equipo_accion_evidencias(activo)
    """)



    cur.execute("PRAGMA table_info(reclamo_respuesta_equipo_acciones)")
    cols = {r[1] for r in cur.fetchall()}

    if "observacion_cumplimiento" not in cols:
        cur.execute("ALTER TABLE reclamo_respuesta_equipo_acciones ADD COLUMN observacion_cumplimiento TEXT")

    if "updated_at" not in cols:
        cur.execute("ALTER TABLE reclamo_respuesta_equipo_acciones ADD COLUMN updated_at TEXT")

    if "updated_by" not in cols:
        cur.execute("ALTER TABLE reclamo_respuesta_equipo_acciones ADD COLUMN updated_by INTEGER")

    conn.commit()
 

def _get_respuesta_equipo_acciones(conn: sqlite3.Connection, respuesta_equipo_id: int):
    cur = conn.cursor()
    rows = cur.execute("""
        SELECT id, tipo, descripcion, fecha_compromiso, orden,
               requiere_evidencia, cumplido, fecha_cumplimiento
        FROM reclamo_respuesta_equipo_acciones
        WHERE respuesta_equipo_id = ?
          AND activo = 1
        ORDER BY tipo, orden, id
    """, (respuesta_equipo_id,)).fetchall()

    causas, controles, correctivas = [], [], []

    for r in rows:
        item = {
            "id": r["id"],
            "descripcion": r["descripcion"] or "",
            "fecha_compromiso": r["fecha_compromiso"] or "",
            "orden": r["orden"] or 1,
            "requiere_evidencia": int(r["requiere_evidencia"] or 0),
            "cumplido": int(r["cumplido"] or 0),
            "fecha_cumplimiento": r["fecha_cumplimiento"] or "",
        }
        if r["tipo"] == "CAUSA":
            causas.append(item)
        elif r["tipo"] == "CONTROL":
            controles.append(item)
        elif r["tipo"] == "CORRECTIVA":
            correctivas.append(item)

    return {
        "causas": causas,
        "control": controles,
        "correctiva": correctivas,
    }

def _normalizar_acciones_payload(items, label: str):
    out = []

    if not items:
        return out, None

    if not isinstance(items, list):
        return None, f"{label} debe enviarse como lista."

    for it in items:
        it = it or {}
        if not isinstance(it, dict):
            continue

        desc = (it.get("descripcion") or "").strip()
        fecha = (it.get("fecha_compromiso") or "").strip()

        if not desc and not fecha:
            continue

        if not desc:
            return None, f"Falta descripción en {label}."
        if not fecha:
            return None, f"Falta fecha compromiso en {label}."
        if not _is_date_yyyy_mm_dd(fecha):
            return None, f"Fecha inválida en {label}. Use YYYY-MM-DD."

        out.append((desc, fecha))

    return out, None


def _get_imputado_acciones_full(db, imputacion_id: int):
    rows = db.execute("""
        SELECT
            a.id,
            a.tipo,
            a.descripcion,
            a.fecha_compromiso,
            COALESCE(a.orden, 1) AS orden,
            COALESCE(a.requiere_evidencia, 0) AS requiere_evidencia,
            COALESCE(a.cumplido, 0) AS cumplido,
            COALESCE(a.fecha_cumplimiento, '') AS fecha_cumplimiento,
            COALESCE(a.observacion_cumplimiento, '') AS observacion_cumplimiento
        FROM reclamo_imputado_acciones a
        WHERE a.imputacion_id = ?
          AND COALESCE(a.activo, 1) = 1
        ORDER BY
            CASE a.tipo
                WHEN 'CAUSA' THEN 1
                WHEN 'CONTROL' THEN 2
                WHEN 'CORRECTIVA' THEN 3
                ELSE 9
            END,
            a.orden,
            a.id
    """, (imputacion_id,)).fetchall()

    causas = []
    control = []
    correctiva = []

    for r in rows:
        accion_id = int(r["id"])

        evid_rows = db.execute("""
            SELECT
                e.id,
                e.filename,
                e.original_name,
                COALESCE(e.content_type, '') AS content_type,
                COALESCE(e.size_bytes, 0) AS size_bytes,
                COALESCE(e.created_at, '') AS created_at
            FROM reclamo_accion_evidencias e
            WHERE e.accion_id = ?
              AND COALESCE(e.activo, 1) = 1
            ORDER BY e.id
        """, (accion_id,)).fetchall()

        item = {
            "id": accion_id,
            "tipo": (r["tipo"] or "").upper(),
            "descripcion": r["descripcion"] or "",
            "fecha_compromiso": r["fecha_compromiso"] or "",
            "orden": int(r["orden"] or 1),
            "requiere_evidencia": int(r["requiere_evidencia"] or 0),
            "cumplido": int(r["cumplido"] or 0),
            "fecha_cumplimiento": r["fecha_cumplimiento"] or "",
            "observacion_cumplimiento": r["observacion_cumplimiento"] or "",
            "evidencias": [dict(x) for x in evid_rows],
        }

        if item["tipo"] == "CAUSA":
            causas.append(item)
        elif item["tipo"] == "CONTROL":
            control.append(item)
        elif item["tipo"] == "CORRECTIVA":
            correctiva.append(item)

    return {
        "causas": causas,
        "control": control,
        "correctiva": correctiva,
    }

BOSS_ROLES = (
    'jefe',
    'gerente',
    'gerente general',
    'gerente financiero',
    'coordinador',
    'admin','usuario'
)
 

def _puede_gestionar_equipo(reclamo_id: int, user_id: int | None = None) -> bool:
    """
    Solo el IMPUTADO APROBADO del reclamo puede gestionar el equipo.
    """
    if not user_id:
        user_id = _current_user_id()
        if not user_id:
            return False

    db = get_db()
    fila = db.execute(
        """
        SELECT 1
        FROM reclamo_imputados ri
        WHERE ri.reclamo_id = ?
          AND ri.imputado_id = ?
          AND ri.estado_asignacion = 'aprobado'
        LIMIT 1
        """,
        (reclamo_id, user_id),
    ).fetchone()
    return fila is not None


def _puede_ver_equipo(reclamo_id: int, user_id: int | None = None) -> bool:
    """
    Puede ver:
    - el imputado aprobado (responsable del proceso),
    - los miembros del equipo,
    - el creador del reclamo.
    """
    if not user_id:
        user_id = _current_user_id()
        if not user_id:
            return False

    db = get_db()

    # 1) si puede gestionar, también puede ver
    if _puede_gestionar_equipo(reclamo_id, user_id):
        return True

    # 2) miembro del equipo
    fila = db.execute(
        """
        SELECT 1
        FROM reclamo_equipo_respuestas er
        WHERE er.reclamo_id = ?
          AND er.usuario_id = ?
        LIMIT 1
        """,
        (reclamo_id, user_id),
    ).fetchone()
    if fila:
        return True

    # 3) creador del reclamo
    fila = db.execute(
        """
        SELECT 1
        FROM reclamos r
        WHERE r.id = ?
          AND r.creado_por = ?
        LIMIT 1
        """,
        (reclamo_id, user_id),
    ).fetchone()
    return fila is not None

def _ensure_column(conn, table, column, decl_sql):
    cur = conn.execute(f"PRAGMA table_info({table})")
    cols = {row["name"] for row in cur.fetchall()}
    if column not in cols:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {decl_sql}")
        conn.commit()

def ensure_reclamo_imputados_fishbone(conn):
    _ensure_column(conn, "reclamo_imputados", "fish_metodo",    "TEXT")
    _ensure_column(conn, "reclamo_imputados", "fish_maquinas",  "TEXT")
    _ensure_column(conn, "reclamo_imputados", "fish_materiales","TEXT")
    _ensure_column(conn, "reclamo_imputados", "fish_personas",  "TEXT")
    _ensure_column(conn, "reclamo_imputados", "fish_entorno",   "TEXT")
    _ensure_column(conn, "reclamo_imputados", "fish_medicion",  "TEXT")
import sqlite3

def ensure_reclamo_respuestas_equipo_schema(conn: sqlite3.Connection):
    cur = conn.cursor()

    # 1) Crear tabla si no existe (esquema base)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS reclamo_respuestas_equipo(
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            reclamo_id     INTEGER NOT NULL,
            imputacion_id  INTEGER NOT NULL,
            miembro_id     INTEGER NOT NULL,

            metodo_analisis TEXT,

            causa          TEXT NOT NULL,
            preventiva     TEXT NOT NULL,
            correctiva     TEXT NOT NULL,

            fish_metodo     TEXT,
            fish_maquinas   TEXT,
            fish_materiales TEXT,
            fish_personas   TEXT,
            fish_entorno    TEXT,
            fish_medicion   TEXT,

            why1 TEXT, why2 TEXT, why3 TEXT, why4 TEXT, why5 TEXT,

            created_at     TEXT,
            created_by     INTEGER,

            activo         INTEGER NOT NULL DEFAULT 1,

            FOREIGN KEY (reclamo_id)    REFERENCES reclamos(id),
            FOREIGN KEY (imputacion_id) REFERENCES reclamo_imputados(id),
            FOREIGN KEY (miembro_id)    REFERENCES usuarios(id)
        );
    """)

    # 2) Detectar columnas existentes
    cur.execute("PRAGMA table_info(reclamo_respuestas_equipo);")
    existing_cols = {row[1] for row in cur.fetchall()}  # row[1] = name

    # 3) Definir columnas requeridas (para ALTER si faltan)
    # Nota: en ALTER TABLE de SQLite, si agregas NOT NULL debe tener DEFAULT.
    required_cols = {
        "reclamo_id":     "INTEGER NOT NULL DEFAULT 0",
        "imputacion_id":  "INTEGER NOT NULL DEFAULT 0",
        "miembro_id":     "INTEGER NOT NULL DEFAULT 0",

        "metodo_analisis":"TEXT",

        "causa":          "TEXT NOT NULL DEFAULT ''",
        "preventiva":     "TEXT NOT NULL DEFAULT ''",
        "correctiva":     "TEXT NOT NULL DEFAULT ''",

        "fish_metodo":     "TEXT",
        "fish_maquinas":   "TEXT",
        "fish_materiales": "TEXT",
        "fish_personas":   "TEXT",
        "fish_entorno":    "TEXT",
        "fish_medicion":   "TEXT",

        "why1": "TEXT",
        "why2": "TEXT",
        "why3": "TEXT",
        "why4": "TEXT",
        "why5": "TEXT",

        "created_at": "TEXT",
        "created_by": "INTEGER",

        "activo": "INTEGER NOT NULL DEFAULT 1",

         "fecha_causa": "TEXT",
        "fecha_preventiva": "TEXT",
        "fecha_correctiva": "TEXT",

        "estado_revision": "TEXT DEFAULT ''",
        "revision_at": "TEXT",
        "revision_by": "INTEGER",
        "revision_msg": "TEXT",
    }

    for col, ddl in required_cols.items():
        if col not in existing_cols:
            cur.execute(f"ALTER TABLE reclamo_respuestas_equipo ADD COLUMN {col} {ddl};")

    # 4) Índices
    cur.execute("CREATE INDEX IF NOT EXISTS idx_rre_reclamo ON reclamo_respuestas_equipo(reclamo_id);")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_rre_imputacion ON reclamo_respuestas_equipo(imputacion_id);")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_rre_miembro ON reclamo_respuestas_equipo(miembro_id);")

    conn.commit()


def ensure_reclamo_adjuntos_schema(conn: sqlite3.Connection):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS reclamo_adjuntos(
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            reclamo_id   INTEGER NOT NULL,
            filename     TEXT NOT NULL,          -- nombre físico en disco
            original_name TEXT NOT NULL,         -- nombre original
            content_type TEXT,
            size_bytes   INTEGER,
            creado_por   INTEGER,
            created_at   TEXT,
            FOREIGN KEY(reclamo_id) REFERENCES reclamos(id)
        )
    """)
    conn.commit()

# =========================================================
#   CATÁLOGOS param_groups / param_values
# =========================================================

def fetch_productos(conn):
    cur = conn.cursor()
    cur.execute("""
        SELECT id, nombre
        FROM productos
        WHERE COALESCE(activo, 1) = 1
        ORDER BY nombre
    """)
    return cur.fetchall()


def _ensure_param_tables(conn: sqlite3.Connection):
    """
    Adapta los catálogos a tu esquema actual:

        param_groups(id, nombre)
        param_values(id, group_id, nombre, valor)

    y, si hace falta, añade columnas 'activo' y 'orden'
    a param_values (no rompe nada existente).
    """
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS param_groups(
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre  TEXT UNIQUE NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS param_values(
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id INTEGER NOT NULL,
            nombre   TEXT NOT NULL,
            valor    TEXT,
            activo   INTEGER NOT NULL DEFAULT 1,
            orden    INTEGER NOT NULL DEFAULT 1,
            FOREIGN KEY(group_id) REFERENCES param_groups(id)
        )
    """)

    cur.execute("PRAGMA table_info(param_values)")
    cols = {r[1] for r in cur.fetchall()}
    if "activo" not in cols:
        cur.execute("ALTER TABLE param_values ADD COLUMN activo INTEGER NOT NULL DEFAULT 1")
    if "orden" not in cols:
        cur.execute("ALTER TABLE param_values ADD COLUMN orden  INTEGER NOT NULL DEFAULT 1")

    conn.commit()


def _ensure_param_group(conn: sqlite3.Connection, codigo: str, descripcion: str) -> int:
    _ensure_param_tables(conn)
    cur = conn.cursor()
    cur.execute("SELECT id FROM param_groups WHERE nombre = ?", (codigo,))
    row = cur.fetchone()
    if row:
        return row["id"]

    cur.execute("INSERT INTO param_groups(nombre) VALUES (?)", (codigo,))
    conn.commit()
    return cur.lastrowid


def _ensure_param_value(conn: sqlite3.Connection,
                        group_codigo: str,
                        clave: str,
                        valor: str,
                        orden: int = 1,
                        activo: int = 1):
    """
    Usamos:
        param_values.nombre -> clave interna (OTROS, DEMORA_ENTREGA, etc.)
        param_values.valor  -> etiqueta visible ("Otros", "Demora en entrega"...)
    """
    gid = _ensure_param_group(conn, group_codigo, group_codigo)
    cur = conn.cursor()

    cur.execute("""
        SELECT id
        FROM param_values
        WHERE group_id = ? AND nombre = ?
    """, (gid, clave))
    row = cur.fetchone()

    if row:
        cur.execute("""
            UPDATE param_values
               SET valor = ?,
                   orden = ?,
                   activo = ?
             WHERE id = ?
        """, (valor, orden, activo, row["id"]))
    else:
        cur.execute("""
            INSERT INTO param_values(group_id, nombre, valor, orden, activo)
            VALUES (?, ?, ?, ?, ?)
        """, (gid, clave, valor, orden, activo))

    conn.commit()


def _fetch_param_values(conn: sqlite3.Connection, group_codigo: str):
    """
    Devuelve lista de rows de un grupo concreto.

    Cada row tendrá:
        id, nombre (clave), valor (etiqueta), orden
    """
    _ensure_param_tables(conn)
    cur = conn.cursor()
    cur.execute("""
        SELECT pv.id, pv.nombre, pv.valor, pv.orden
        FROM param_values pv
        JOIN param_groups pg ON pg.id = pv.group_id
        WHERE pg.nombre = ?
          AND COALESCE(pv.activo, 1) = 1
        ORDER BY pv.orden, pv.valor, pv.nombre
    """, (group_codigo,))
    return cur.fetchall()


from datetime import datetime
from flask import request, jsonify, session, abort

ACC_TIPOS_VALIDOS = {"CAUSA", "CONTROL", "CORRECTIVA"}

def _now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def _is_date_yyyy_mm_dd(s: str) -> bool:
    try:
        datetime.strptime(s, "%Y-%m-%d")
        return True
    except Exception:
        return False

def _can_edit_equipo(equipo_id: int) -> bool:
    # superadmin/admin
    if session.get("rol") == "admin" or session.get("is_admin") is True:
        return True

    user_id = session.get("user_id") or session.get("id")
    if not user_id:
        return False

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT responsable_id, colaborador_id
        FROM reclamo_equipo
        WHERE id = ?
    """, (equipo_id,))
    row = cur.fetchone()
    if not row:
        return False

    responsable_id = row["responsable_id"] if hasattr(row, "keys") else row[0]
    colaborador_id = row["colaborador_id"] if hasattr(row, "keys") else row[1]

    return int(user_id) in {int(responsable_id), int(colaborador_id)}


def ensure_reclamos_catalogos(conn: sqlite3.Connection):
    # Tipos de reclamo
    tipos = [
        ("OTROS",             "Otros"),
        ("FUERA_LINEA",       "Fuera de línea"),
        ("DANIO_ENVASES",     "Daño de envases"),
        ("ETIQUETADO",        "Etiquetado"),
        ("DEMORA_ENTREGA",    "Demora en entrega"),
        ("CALIDAD_PRODUCTO",  "Calidad de producto"),
        ("CALIDAD_SERVICIO",  "Calidad de servicio"),
        ("DIFERENCIA_ENTREGA",  "Diferencia en Entrega"),
    ]
    for idx, (clave, label) in enumerate(tipos, start=1):
        _ensure_param_value(conn, "RECL_TIPO", clave, label, orden=idx)

    # Materiales (catálogo RECL_MATERIAL)
    materiales_seed = [
         
        ("0927060624", "SISTEMAS"),
        
    ]
    for idx, (clave, label) in enumerate(materiales_seed, start=1):
        _ensure_param_value(conn, "RECL_MATERIAL", clave, label, orden=idx)

    # Tipo de trámite
    tramites = [
        ("INTERNO", "Interno"),
        ("EXTERNO", "Externo"),
    ]
    for idx, (clave, label) in enumerate(tramites, start=1):
        _ensure_param_value(conn, "RECL_TRAMITE", clave, label, orden=idx)

    # Grupo de procesos con valores por defecto
    procesos_defecto = [
        ("0927060624",      "SISTEMAS"),
        
    ]
    for idx, (clave, label) in enumerate(procesos_defecto, start=1):
        _ensure_param_value(conn, "RECL_PROCESO", clave, label, orden=idx)

    # Config de campos por tipo
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS reclamo_equipo (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            reclamo_id       INTEGER NOT NULL,
            responsable_id   INTEGER NOT NULL,
            colaborador_id   INTEGER NOT NULL,
            estado           TEXT,
            fecha_asignacion TEXT,
            fecha_respuesta  TEXT,
            fecha_aprobacion TEXT,
            fecha_rechazo    TEXT,
            motivo_rechazo   TEXT,
            respuesta_causa        TEXT,
            respuesta_preventiva   TEXT,
            respuesta_correctiva   TEXT
        )
    """)
    


    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS reclamo_equipo_acciones (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            equipo_id        INTEGER NOT NULL,      -- FK lógico a reclamo_equipo.id
            tipo             TEXT NOT NULL,         -- 'CAUSA' | 'CONTROL' | 'CORRECTIVA'
            descripcion      TEXT NOT NULL,
            fecha_compromiso TEXT NOT NULL,         -- 'YYYY-MM-DD'
            created_at       TEXT,
            created_by       INTEGER
        )
    """)

    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_req_acc_equipo
        ON reclamo_equipo_acciones (equipo_id)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_req_acc_tipo
        ON reclamo_equipo_acciones (tipo)
    """)
 

    cur.execute("""
        CREATE TABLE IF NOT EXISTS reclamo_tipo_campos(
            tipo_codigo TEXT PRIMARY KEY,
            etiqueta    TEXT,

            usa_region           INTEGER DEFAULT 0,
            usa_tramite          INTEGER DEFAULT 0,
            usa_proceso          INTEGER DEFAULT 0,
            usa_fecha_reclamo    INTEGER DEFAULT 1,
            usa_fecha_pedido     INTEGER DEFAULT 0,
            usa_fecha_ofrec      INTEGER DEFAULT 0,
            usa_fecha_entrega    INTEGER DEFAULT 0,
            usa_material         INTEGER DEFAULT 0,
            usa_persona_atendio  INTEGER DEFAULT 0,
            usa_colaborador      INTEGER DEFAULT 0,
            usa_factura_guia     INTEGER DEFAULT 0,
            usa_antecedente      INTEGER DEFAULT 1,
            usa_observacion      INTEGER DEFAULT 1
        )
    """)

    seeds = {
        "OTROS": dict(
            usa_region=1, usa_tramite=1, usa_proceso=1,
            usa_fecha_reclamo=1, usa_fecha_pedido=1,
            usa_factura_guia=1, usa_material=1,
            usa_antecedente=1, usa_observacion=1
        ),
        "FUERA_LINEA": dict(
            usa_region=1, usa_tramite=1, usa_proceso=1,
            usa_fecha_reclamo=1, usa_fecha_pedido=1,
            usa_factura_guia=1, usa_material=1,
            usa_antecedente=1, usa_observacion=1
        ),
        "DANIO_ENVASES": dict(
            usa_region=1, usa_tramite=1, usa_proceso=1,
            usa_fecha_reclamo=1, usa_fecha_pedido=1,
            usa_factura_guia=1, usa_material=1,
            usa_antecedente=1, usa_observacion=1
        ),
        "ETIQUETADO": dict(
            usa_region=1, usa_tramite=1, usa_proceso=1,
            usa_fecha_reclamo=1, usa_fecha_pedido=1,
            usa_factura_guia=1, usa_material=1,
            usa_antecedente=1, usa_observacion=1
        ),
        "DEMORA_ENTREGA": dict(
            usa_region=0, usa_tramite=0,
            usa_proceso=1,
            usa_fecha_reclamo=1,
            usa_fecha_pedido=0,
            usa_fecha_ofrec=1,
            usa_fecha_entrega=1,
            usa_material=1,
            usa_persona_atendio=1,
            usa_factura_guia=1,
            usa_antecedente=0,
            usa_observacion=1
        ),
        "CALIDAD_PRODUCTO": dict(
            usa_region=0, usa_tramite=0,
            usa_proceso=1,
            usa_fecha_reclamo=0,
            usa_fecha_pedido=1,
            usa_material=1,
            usa_persona_atendio=1,
            usa_factura_guia=1,
            usa_antecedente=0,
            usa_observacion=1
        ),
        "CALIDAD_SERVICIO": dict(
            usa_region=0, usa_tramite=0,
            usa_proceso=1,
            usa_fecha_reclamo=0,
            usa_fecha_pedido=1,
            usa_material=0,
            usa_persona_atendio=1,
            usa_colaborador=1,
            usa_factura_guia=1,
            usa_antecedente=0,
            usa_observacion=1
        ),
        "DIFERENCIA_ENTREGA": dict(
            usa_region=0, usa_tramite=0,
            usa_proceso=1,
            usa_fecha_reclamo=0,
            usa_fecha_pedido=1,
            usa_material=1,
            usa_persona_atendio=1,
            usa_colaborador=1,
            usa_factura_guia=1,
            usa_antecedente=0,
            usa_observacion=1
        ),
         "INC_PLAN": dict(
            usa_region=0, usa_tramite=0,
            usa_proceso=1,
            usa_fecha_reclamo=0,
            usa_fecha_pedido=1,
            usa_material=1,
            usa_persona_atendio=1,
            usa_colaborador=1,
            usa_factura_guia=1,
            usa_antecedente=0,
            usa_observacion=1
        ),
        
         "GES_PEDIDO": dict(
            usa_region=0, usa_tramite=0,
            usa_proceso=1,
            usa_fecha_reclamo=0,
            usa_fecha_pedido=1,
            usa_material=1,
            usa_persona_atendio=1,
            usa_colaborador=1,
            usa_factura_guia=1,
            usa_antecedente=0,
            usa_observacion=1
        ),
    }

    for tipo_codigo, cfg in seeds.items():
        cur.execute("""
            SELECT pv.valor
            FROM param_values pv
            JOIN param_groups pg ON pg.id = pv.group_id
            WHERE pg.nombre = 'RECL_TIPO' AND pv.nombre = ?
        """, (tipo_codigo,))

        row = cur.fetchone()
        etiqueta = row["valor"] if row else tipo_codigo

        cur.execute(
            "SELECT tipo_codigo FROM reclamo_tipo_campos WHERE tipo_codigo=?",
            (tipo_codigo,)
        )
        if cur.fetchone():
            cur.execute("""
                UPDATE reclamo_tipo_campos
                   SET etiqueta=?,
                       usa_region=?,
                       usa_tramite=?,
                       usa_proceso=?,
                       usa_fecha_reclamo=?,
                       usa_fecha_pedido=?,
                       usa_fecha_ofrec=?,
                       usa_fecha_entrega=?,
                       usa_material=?,
                       usa_persona_atendio=?,
                       usa_colaborador=?,
                       usa_factura_guia=?,
                       usa_antecedente=?,
                       usa_observacion=?
                 WHERE tipo_codigo=?
            """, (
                etiqueta,
                cfg.get("usa_region", 0),
                cfg.get("usa_tramite", 0),
                cfg.get("usa_proceso", 0),
                cfg.get("usa_fecha_reclamo", 1),
                cfg.get("usa_fecha_pedido", 0),
                cfg.get("usa_fecha_ofrec", 0),
                cfg.get("usa_fecha_entrega", 0),
                cfg.get("usa_material", 0),
                cfg.get("usa_persona_atendio", 0),
                cfg.get("usa_colaborador", 0),
                cfg.get("usa_factura_guia", 0),
                cfg.get("usa_antecedente", 1),
                cfg.get("usa_observacion", 1),
                tipo_codigo
            ))
        else:
            cur.execute("""
                INSERT INTO reclamo_tipo_campos(
                    tipo_codigo, etiqueta,
                    usa_region, usa_tramite, usa_proceso,
                    usa_fecha_reclamo, usa_fecha_pedido,
                    usa_fecha_ofrec, usa_fecha_entrega,
                    usa_material, usa_persona_atendio, usa_colaborador,
                    usa_factura_guia, usa_antecedente, usa_observacion
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                tipo_codigo, etiqueta,
                cfg.get("usa_region", 0),
                cfg.get("usa_tramite", 0),
                cfg.get("usa_proceso", 0),
                cfg.get("usa_fecha_reclamo", 1),
                cfg.get("usa_fecha_pedido", 0),
                cfg.get("usa_fecha_ofrec", 0),
                cfg.get("usa_fecha_entrega", 0),
                cfg.get("usa_material", 0),
                cfg.get("usa_persona_atendio", 0),
                cfg.get("usa_colaborador", 0),
                cfg.get("usa_factura_guia", 0),
                cfg.get("usa_antecedente", 1),
                cfg.get("usa_observacion", 1),
            ))
    conn.commit()


# =========================================================
#   CATÁLOGOS GEO
# =========================================================

def ensure_geo_schema(conn: sqlite3.Connection):
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS regiones(
            id     INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT UNIQUE NOT NULL,
            activo INTEGER NOT NULL DEFAULT 1,
            orden  INTEGER NOT NULL DEFAULT 1
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS provincias(
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            region_id INTEGER NOT NULL,
            nombre    TEXT NOT NULL,
            activo    INTEGER NOT NULL DEFAULT 1,
            orden     INTEGER NOT NULL DEFAULT 1,
            FOREIGN KEY(region_id) REFERENCES regiones(id)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS cantones(
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            provincia_id INTEGER NOT NULL,
            nombre       TEXT NOT NULL,
            activo       INTEGER NOT NULL DEFAULT 1,
            orden        INTEGER NOT NULL DEFAULT 1,
            FOREIGN KEY(provincia_id) REFERENCES provincias(id)
        )
    """)

    conn.commit()


def fetch_regiones(conn: sqlite3.Connection):
    cur = conn.cursor()
    cur.execute("""
        SELECT id, nombre
        FROM regiones
        WHERE COALESCE(activo,1) = 1
        ORDER BY orden, nombre
    """)
    return cur.fetchall()


def fetch_provincias(conn: sqlite3.Connection, region_id: int | None):
    cur = conn.cursor()
    if region_id:
        cur.execute("""
            SELECT id, nombre
            FROM provincias
            WHERE region_id = ? AND COALESCE(activo,1)=1
            ORDER BY orden, nombre
        """, (region_id,))
    else:
        cur.execute("""
            SELECT id, nombre
            FROM provincias
            WHERE COALESCE(activo,1)=1
            ORDER BY orden, nombre
        """)
    return cur.fetchall()


def fetch_cantones(conn: sqlite3.Connection, provincia_id: int | None):
    cur = conn.cursor()
    if provincia_id:
        cur.execute("""
            SELECT id, nombre
            FROM cantones
            WHERE provincia_id = ? AND COALESCE(activo,1)=1
            ORDER BY orden, nombre
        """, (provincia_id,))
    else:
        cur.execute("""
            SELECT id, nombre
            FROM cantones
            WHERE COALESCE(activo,1)=1
            ORDER BY orden, nombre
        """)
    return cur.fetchall()


def _fetch_tipo_campos(conn: sqlite3.Connection):
    cur = conn.cursor()
    cur.execute("SELECT * FROM reclamo_tipo_campos")
    rows = cur.fetchall()
    return {r["tipo_codigo"]: r for r in rows}


# =========================================================
#   Helpers de negocio
# =========================================================

def _generate_codigo_reclamo(conn: sqlite3.Connection) -> str:
    cur = conn.cursor()
    cur.execute("""
        SELECT codigo FROM reclamos
        WHERE codigo LIKE 'RECL%'
        ORDER BY id DESC
        LIMIT 1
    """)
    row = cur.fetchone()
    if row and row["codigo"]:
        num_part = "".join(ch for ch in row["codigo"] if ch.isdigit())
        try:
            nxt = int(num_part) + 1
        except ValueError:
            nxt = 1
    else:
        nxt = 1
    return f"RECL{nxt:05d}"


def _guess_aprobador_for_user2(conn: sqlite3.Connection, user_id: int | None) -> int | None:
    if not user_id:
        return None

    cur = conn.cursor()
    cur.execute("""
        SELECT departamento_id, LOWER(rol) AS rol
        FROM usuarios
        WHERE id = ?
    """, (user_id,))
    u = cur.fetchone()
    if not u:
        return None

    depto_id = u["departamento_id"]

    cur.execute(f"""
        SELECT id
        FROM usuarios
        WHERE departamento_id = ?
          AND LOWER(rol) IN ({",".join(["?"]*len(BOSS_ROLES))})
        LIMIT 1
    """, (depto_id, *[r.lower() for r in BOSS_ROLES]))
    boss = cur.fetchone()
    return boss["id"] if boss else None


def _guess_aprobador_for_user(conn: sqlite3.Connection, user_id: int | None) -> int | None:
    """
    Devuelve el jefe directo del usuario usando usuarios.jefe_id.
    Si no tiene jefe_id definido o el jefe está deshabilitado, cae
    al esquema viejo de buscar alguien del mismo departamento con
    rol en BOSS_ROLES.
    """
    if not user_id:
        return None

    cur = conn.cursor()
    # Primero intentamos con jefe_id
    cur.execute("""
        SELECT jefe_id, departamento_id
        FROM usuarios
        WHERE id = ?
    """, (user_id,))
    u = cur.fetchone()
    if not u:
        return None

    jefe_id = u["jefe_id"]
    depto_id = u["departamento_id"]

    # Si tiene jefe_id definido, validamos que exista y no esté deshabilitado
    if jefe_id:
        cur.execute("""
            SELECT id
            FROM usuarios
            WHERE id = ?
              AND COALESCE(disabled, 0) = 0
        """, (jefe_id,))
        j = cur.fetchone()
        if j:
            return j["id"]

    # Fallback: buscar jefe por rol en el mismo departamento (esquema antiguo)
    if not depto_id:
        return None

    cur.execute(f"""
        SELECT id
        FROM usuarios
        WHERE departamento_id = ?
          AND LOWER(rol) IN ({",".join(["?"] * len(BOSS_ROLES))})
          AND COALESCE(disabled, 0) = 0
        ORDER BY id
        LIMIT 1
    """, (depto_id, *[r.lower() for r in BOSS_ROLES]))
    boss = cur.fetchone()
    return boss["id"] if boss else None


# =========================================================
#   EMAIL HELPER UNIFICADO
# =========================================================

def _send_mail_safe(to_email: str, subject: str, text_body: str, html_body: str | None = None):
    """
    Envía correo multipart/alternative (texto + HTML) usando la configuración
    SMTP guardada en base de datos (get_config_value).
    """
    if not to_email:
        return

    host = get_config_value('smtp_host') or ''
    port = int(get_config_value('smtp_port') or "587")
    user = get_config_value('smtp_user') or ''
    pwd = get_config_value('smtp_pass') or ''
    from_addr = get_config_value('smtp_from') or user or ''

    if not host or not from_addr:
        current_app.logger.warning("SMTP no configurado correctamente, se omite envío de correo.")
        return

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["To"] = to_email
    msg["From"] = from_addr

    if html_body:
        msg.set_content(text_body or "(ver versión HTML)")
        msg.add_alternative(html_body, subtype="html")
    else:
        msg.set_content(text_body or "")

    try:
        with smtplib.SMTP(host, port, timeout=20) as s:
            s.starttls()
            if user and pwd:
                s.login(user, pwd)
            s.send_message(msg)
        current_app.logger.info("Correo enviado a %s con asunto '%s'", to_email, subject)
    except Exception as e:
        current_app.logger.exception("Error enviando correo: %s", e)


def fetch_usuarios_imputables2(conn, empresa_ids=None):
    """
    Devuelve usuarios imputables con:
      - id
      - username
      - nombre_completo
      - departamento_nombre
      - jefe_nombre  (primer usuario del mismo dpto con rol en BOSS_ROLES)
    Opcionalmente filtra por empresa_id (lista de ints).
    """
    cur = conn.cursor()

    boss_roles = [r.lower() for r in BOSS_ROLES]
    boss_placeholders = ",".join(["?"] * len(boss_roles))

    sql = f"""
        SELECT
            u.id,
            u.username,
            COALESCE(u.nombre_completo, u.username) AS nombre_completo,
            d.nombre AS departamento_nombre,
            (
                SELECT COALESCE(j.nombre_completo, j.username)
                FROM usuarios j
                WHERE j.departamento_id = u.departamento_id
                  AND LOWER(j.rol) IN ({boss_placeholders})
                ORDER BY j.id
                LIMIT 1
            ) AS jefe_nombre,
            u.empresa_id
        FROM usuarios u
        LEFT JOIN departamentos d ON d.id = u.departamento_id
        WHERE COALESCE(u.disabled, 0) = 0
    """

    params = boss_roles[:]

    if empresa_ids:
        emp_placeholders = ",".join(["?"] * len(empresa_ids))
        sql += f" AND COALESCE(u.empresa_id,0) IN ({emp_placeholders})"
        params.extend(empresa_ids)

    sql += " ORDER BY nombre_completo, u.username"

    cur.execute(sql, params)
    return cur.fetchall()


def fetch_usuarios_imputables(conn, empresa_ids=None):
    """
    Devuelve usuarios imputables con:
      - id
      - username
      - nombre_completo
      - departamento_nombre
      - jefe_nombre  (usando usuarios.jefe_id)
    Opcionalmente filtra por empresa_id (lista de ints).
    """
    cur = conn.cursor()

    sql = """
        SELECT
            u.id,
            u.username,
            COALESCE(u.nombre_completo, u.username) AS nombre_completo,
            d.nombre AS departamento_nombre,
            COALESCE(j.nombre_completo, j.username) AS jefe_nombre,
            u.empresa_id,
            u.jefe_id
        FROM usuarios u
        LEFT JOIN departamentos d ON d.id = u.departamento_id
        LEFT JOIN usuarios j ON j.id = u.jefe_id          -- 👈 jefe desde la misma tabla
        WHERE COALESCE(u.disabled, 0) = 0
        and u.identificacion not in ('40623','0911946630','0923577688','0929626729','1307590834','0917013252'
        ,'40736','0902507805','1714868211')
    """
    params = []

    if empresa_ids:
        emp_placeholders = ",".join(["?"] * len(empresa_ids))
        sql += f" AND COALESCE(u.empresa_id, 0) IN ({emp_placeholders})"
        params.extend(empresa_ids)

    sql += " ORDER BY nombre_completo, u.username"

    cur.execute(sql, params)
    return cur.fetchall()


def _get_user_basic(conn: sqlite3.Connection, uid: int | None):
    if not uid:
        return None
    cur = conn.cursor()
    cur.execute("""
        SELECT id, username, email, rol, departamento_id, nombre_completo
        FROM usuarios
        WHERE id = ?
    """, (uid,))
    return cur.fetchone()


def _notify_colaborador_asignado(
    conn,
    colaborador_id: int,
    reclamo_codigo: str,
    responsable_username: str
):
    u = _get_user_basic(conn, colaborador_id)
    if not u or ("email" not in u.keys()) or not u["email"]:
        return

    # --- Datos del reclamo (por codigo) ---
    cur = conn.cursor()
    cur.execute("""
        SELECT id, fecha_reclamo, tipo_reclamo, tipo_tramite,
               cliente_nombre, proceso_text, material_desc,
               fecha_pedido, factura, guia_remision,
               antecedente, observacion
        FROM reclamos
        WHERE codigo = ?
        LIMIT 1
    """, (reclamo_codigo,))
    r = cur.fetchone()

    # Imputados (si aplica)
    imputados = ""
    if r:
        cur.execute("""
            SELECT GROUP_CONCAT(u.username, ', ') AS lista
            FROM reclamo_imputados ri
            JOIN usuarios u ON u.id = ri.imputado_id
            WHERE ri.reclamo_id = ?
        """, (r["id"],))
        row = cur.fetchone()
        if row and row["lista"]:
            imputados = row["lista"]

    # Link directo a tab "Soy Sponsor"
    try:
        link_sponsor = url_for("reclamos", _external=True) + "?tab=sponsor"
    except Exception:
        link_sponsor = "https://tu-sistema/reclamos?tab=sponsor"

    nombre = (u["nombre_completo"] or "").strip() if "nombre_completo" in u.keys() else ""
    if not nombre:
        nombre = u["username"]

    subject = f"[Oportunidad de Mejora] Aporte requerido en {reclamo_codigo}"

    # Texto plano (fallback)
    text_body = f"""Hola {nombre},

El responsable {responsable_username} te ha solicitado apoyo para la
Oportunidad de Mejora {reclamo_codigo}.

Por favor ingresa al sistema, pestaña "Soy Sponsor" y registra tu aporte
(causa, acción preventiva y acción correctiva).

Ir al sistema: {link_sponsor}

Este aporte no es la respuesta oficial, pero ayudará al responsable a
construir la respuesta final.

Este es un mensaje automático.
"""

    # ---------- HTML mejorado (mismo estilo que aprobador) ----------
    def _row(lbl, val):
        val = (val or "")
        val = str(val).replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#fee2e2;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;
               font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">
            <!-- Encabezado -->
            <tr>
              <td style="background:#b91c1c;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;
                            letter-spacing:.08em;opacity:.9;">
                  Oportunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Aporte requerido {reclamo_codigo}
                </div>
                <div style="font-size:12px;opacity:.9;margin-top:6px;">
                  Hola {nombre}, el responsable <strong>{responsable_username}</strong> solicitó tu apoyo.
                </div>
              </td>
            </tr>

            <!-- Cuerpo -->
            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row('Código', reclamo_codigo)}
                  {_row('Responsable', responsable_username)}
                  { _row('Fecha OM', r['fecha_reclamo']) if r else '' }
                  { _row('Tipo de OM', r['tipo_reclamo']) if r else '' }
                  { _row('Tipo de Trámite', r['tipo_tramite']) if r else '' }
                  { _row('Cliente', r['cliente_nombre']) if r else '' }
                  { _row('Proceso', r['proceso_text']) if r else '' }
                  { _row('Material', r['material_desc']) if (r and 'material_desc' in r.keys()) else '' }
                  { _row('Fecha de Pedido', r['fecha_pedido']) if r else '' }
                  { _row('Factura', r['factura']) if r else '' }
                  { _row('Guía Remisión', r['guia_remision']) if r else '' }
                  { _row('Imputados', imputados) if imputados else '' }
                  { _row('Antecedente', r['antecedente']) if r else '' }
                  { _row('Observación', r['observacion']) if r else '' }
                </table>

                <div style="margin-top:14px;font-size:12px;color:#374151;">
                  Por favor ingresa a la pestaña <strong>“Soy Sponsor”</strong> y registra tu aporte:
                  <strong>causa</strong>, <strong>acción preventiva</strong> y <strong>acción correctiva</strong>.
                  <br>
                  <span style="color:#6b7280;">
                    Este aporte no es la respuesta oficial, pero ayudará al responsable a construir la respuesta final.
                  </span>
                </div>

                <!-- Botón CTA -->
                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link_sponsor}"
                     style="display:inline-block;background:#2563eb;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Registrar aporte (Soy Sponsor)
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  También puedes ingresar al módulo de reclamos desde el sistema
                  y abrir la bandeja <strong>“Soy Sponsor”</strong>.
                </div>
              </td>
            </tr>

            <!-- Pie -->
            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>

          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

    _send_mail_safe(u["email"], subject, text_body, html_body=html_body)

def _notify_responsable_aporte_listo(conn, responsable_id: int, reclamo_codigo: str, colaborador_username: str):
    r = _get_user_basic(conn, responsable_id)
    if not r or ("email" not in r.keys()) or not r["email"]:
        return

    nombre = r["nombre_completo"] if r.get("nombre_completo") else r["username"]

    subject = f"[Oportunidad de Mejora] Aporte recibido en {reclamo_codigo}"
    text_body = f"""Hola {nombre},

El usuario {colaborador_username} registró su aporte técnico en la
Oportunidad de Mejora {reclamo_codigo}.

Puedes revisarlo en la sección "Soy Sponsor" y, si lo consideras adecuado,
aprovecharlo para tu respuesta final como responsable.

Este es un mensaje automático.
"""
    _send_mail_safe(r["email"], subject, text_body)


def _notify_colaborador_aporte_rechazado(conn, colaborador_id: int, reclamo_codigo: str, motivo: str):
    u = _get_user_basic(conn, colaborador_id)
    if not u or ("email" not in u.keys()) or not u["email"]:
        return

    nombre = u["nombre_completo"] if u.get("nombre_completo") else u["username"]
    motivo_txt = (motivo or "Sin detalle").strip()

    subject = f"[Oportunidad de Mejora] Aporte rechazado en {reclamo_codigo}"
    text_body = f"""Hola {nombre},

Tu aporte técnico para la Oportunidad de Mejora {reclamo_codigo} fue marcado
como rechazado por el responsable.

Motivo:
{motivo_txt}

Puedes coordinar con tu responsable si necesitas más detalles.

Este es un mensaje automático.
"""
    _send_mail_safe(u["email"], subject, text_body)


MAX_FILES_PER_RECLAMO = 5
MAX_FILE_SIZE_BYTES = 5 * 1024 * 1024  # 5 MB


def _get_reclamos_upload_folder() -> str:
    """
    Carpeta donde se guardan físicamente los adjuntos de reclamos.
    Puedes sobreescribirla con app.config['RECLAMOS_UPLOAD_FOLDER'].
    """
    base = current_app.config.get('RECLAMOS_UPLOAD_FOLDER')
    if not base:
        base = os.path.join(current_app.instance_path, 'reclamos_adjuntos')
    os.makedirs(base, exist_ok=True)
    return base


def _save_adjuntos_for_reclamo(conn: sqlite3.Connection,
                               reclamo_id: int,
                               files,
                               user_id: int | None) -> str | None:
    """
    Guarda adjuntos para un reclamo, respetando:
      - Máx 5 archivos por OM
      - Máx 5 MB por archivo
    Devuelve:
      - None si todo OK
      - Mensaje de error (str) si algo no cumple las reglas
    """
    ensure_reclamo_adjuntos_schema(conn)
    cur = conn.cursor()

    # Cuántos adjuntos ya tiene esta OM
    cur.execute("""
        SELECT COUNT(*) AS c
        FROM reclamo_adjuntos
        WHERE reclamo_id = ?
    """, (reclamo_id,))
    row = cur.fetchone()
    existing = row["c"] if row else 0

    remaining = MAX_FILES_PER_RECLAMO - existing
    if remaining <= 0:
        return "Ya existe el máximo de archivos para esta OM (5)."

    upload_folder = _get_reclamos_upload_folder()
    saved_any = False

    for f in (files or [])[:remaining]:
        if not f or not f.filename:
            continue

        # Medir tamaño del archivo
        f.stream.seek(0, os.SEEK_END)
        size = f.stream.tell()
        f.stream.seek(0)

        if size > MAX_FILE_SIZE_BYTES:
            return f"El archivo '{f.filename}' supera el máximo permitido de 5 MB."

        original_name = f.filename
        safe_name = secure_filename(original_name) or "archivo"

        # Nombre único en disco
        ts = datetime.now().strftime("%Y%m%d%H%M%S%f")
        filename_on_disk = f"{reclamo_id}_{ts}_{safe_name}"
        path = os.path.join(upload_folder, filename_on_disk)

        # Guardar archivo en disco
        f.save(path)

        # Registrar en BD
        cur.execute("""
            INSERT INTO reclamo_adjuntos(
                reclamo_id, filename, original_name,
                content_type, size_bytes,
                creado_por, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            reclamo_id,
            filename_on_disk,
            original_name,
            f.mimetype,
            size,
            user_id,
            _now_iso()
        ))
        saved_any = True

    # No hacemos commit aquí; lo hace el caller
    return None


from flask import url_for

def _notify_aprobador_imputacion(conn, aprobador_id, reclamo_codigo, imputado_username):
    jefe = _get_user_basic(conn, aprobador_id)
    if not jefe or ("email" not in jefe.keys()) or not jefe["email"]:
        return

    # --- Datos del reclamo (por codigo) ---
    cur = conn.cursor()
    cur.execute("""
        SELECT id, fecha_reclamo, tipo_reclamo, tipo_tramite,
               cliente_nombre, proceso_text, material_desc,
               fecha_pedido, factura, guia_remision,
               antecedente, observacion
        FROM reclamos
        WHERE codigo = ?
        LIMIT 1
    """, (reclamo_codigo,))
    r = cur.fetchone()

    # Imputados del caso (por si hay más de uno)
    imputados = imputado_username or ""
    if r:
        cur.execute("""
            SELECT GROUP_CONCAT(u.username, ', ') AS lista
            FROM reclamo_imputados ri
            JOIN usuarios u ON u.id = ri.imputado_id
            WHERE ri.reclamo_id = ?
        """, (r["id"],))
        row = cur.fetchone()
        if row and row["lista"]:
            imputados = row["lista"]

    # Link directo al tab "Por aprobar (Jefe)"
    try:
        link_aprobar = url_for("reclamos", _external=True) + "?tab=aprobar"
    except Exception:
        link_aprobar = "https://tu-sistema/reclamos?tab=aprobar"

    nombre = (
        jefe['nombre_completo']
        if 'nombre_completo' in jefe.keys() and jefe['nombre_completo']
        else jefe['username']
    )

    subject = f"[Oportunida de Mejora] Aprobación pendiente {reclamo_codigo}"

    # Texto plano (fallback)
    text_body = f"""Hola {nombre},

Hay un reclamo {reclamo_codigo} con imputación pendiente para: {imputados}.
Por favor revísalo y aprueba/rechaza la imputación.

Ir al sistema: {link_aprobar}

Este es un mensaje automático.
"""

    # ---------- HTML mejorado ----------
    def _row(lbl, val):
        val = (val or "").replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#fee2e2;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;
               font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">
            <!-- Encabezado -->
            <tr>
              <td style="background:#b91c1c;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;
                            letter-spacing:.08em;opacity:.9;">
                  Oprotunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Aprobación pendiente {reclamo_codigo}
                </div>
                <div style="font-size:12px;opacity:.9;margin-top:6px;">
                  Hola {nombre}, tienes una imputación pendiente de revisión.
                </div>
              </td>
            </tr>

            <!-- Cuerpo -->
            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row('Código', reclamo_codigo)}
                  { _row('Fecha OM', r['fecha_reclamo']) if r else '' }
                  { _row('Tipo de OM', r['tipo_reclamo']) if r else '' }
                  { _row('Tipo de Trámite', r['tipo_tramite']) if r else '' }
                  { _row('Cliente', r['cliente_nombre']) if r else '' }
                  { _row('Proceso', r['proceso_text']) if r else '' }
                  { _row('Material', r['material_desc']) if (r and 'material_desc' in r.keys()) else '' }
                  { _row('Fecha de Pedido', r['fecha_pedido']) if r else '' }
                  { _row('Factura', r['factura']) if r else '' }
                  { _row('Guía Remisión', r['guia_remision']) if r else '' }
                  {_row('Imputados', imputados)}
                  { _row('Antecedente', r['antecedente']) if r else '' }
                  { _row('Observación', r['observacion']) if r else '' }
                </table>

                <!-- Botón CTA -->
                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link_aprobar}"
                     style="display:inline-block;background:#2563eb;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Revisar y aprobar imputación
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  También puedes ingresar al módulo de reclamos desde el sistema
                  y abrir la bandeja <strong>“Por aprobar (Jefe)”</strong>.
                </div>
              </td>
            </tr>

            <!-- Pie -->
            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>

          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

    _send_mail_safe(jefe["email"], subject, text_body, html_body=html_body)

def _notify_creador_rechazo_asignacion(conn, creador_id, reclamo_codigo, imputado_username, motivo):
    c = _get_user_basic(conn, creador_id)
    if not c:
        return
    subject = f"[Oportunidad de Mejora] Imputación rechazada en {reclamo_codigo}"
    body = (
        f"Hola {c['username']},\n\n"
        f"El reclamo {reclamo_codigo} fue RECHAZADO para el usuario {imputado_username}.\n"
        f"Motivo del rechazo:\n{motivo or 'Sin motivo informado.'}\n\n"
        "Este es un mensaje automático."
    )
    _send_mail_safe(c["email"], subject, body)


def _notify_imputado_aprobado(conn, imputado_id, reclamo_codigo):
    u = _get_user_basic(conn, imputado_id)
    if not u or ("email" not in u.keys()) or not u["email"]:
        return

    # --- Datos del reclamo (por código) ---
    cur = conn.cursor()
    cur.execute("""
        SELECT id, fecha_reclamo, tipo_reclamo, tipo_tramite,
               cliente_nombre, proceso_text, material_desc,
               fecha_pedido, factura, guia_remision,
               antecedente, observacion
        FROM reclamos
        WHERE codigo = ?
        LIMIT 1
    """, (reclamo_codigo,))
    r = cur.fetchone()

    # Link directo al tab "Soy responsable"
    try:
        link_responder = url_for("reclamos", _external=True) + "?tab=imputado"
    except Exception:
        link_responder = "https://tu-sistema/reclamos?tab=imputado"

    nombre = (
        u['nombre_completo']
        if 'nombre_completo' in u.keys() and u['nombre_completo']
        else u['username']
    )

    subject = f"[Oportunidad de Mejora] Nueva OM asignada ({reclamo_codigo})"

    # Texto plano (fallback)
    text_body = f"""Hola {nombre},

Se te ha asignado la Oportunidad de Mejora {reclamo_codigo}.
Por favor ingresa al sistema, revisa el detalle y registra:

- Causa raíz
- Acción preventiva
- Acción correctiva

Ir al sistema: {link_responder}

Este es un mensaje automático.
"""

    # ---------- HTML similar al del aprobador ----------
    def _row(lbl, val):
        val = (val or "").replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#dbeafe;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;
               font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">
            <!-- Encabezado -->
            <tr>
              <td style="background:#2563eb;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;
                            letter-spacing:.08em;opacity:.9;">
                  Oportunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Nueva OM asignada {reclamo_codigo}
                </div>
                <div style="font-size:12px;opacity:.9;margin-top:6px;">
                  Hola {nombre}, se te ha asignado esta OM para análisis y respuesta técnica.
                </div>
              </td>
            </tr>

            <!-- Cuerpo -->
            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row('Código', reclamo_codigo)}
                  { _row('Fecha OM', r['fecha_reclamo']) if r else '' }
                  { _row('Tipo de OM', r['tipo_reclamo']) if r else '' }
                  { _row('Tipo de Trámite', r['tipo_tramite']) if r else '' }
                  { _row('Cliente', r['cliente_nombre']) if r else '' }
                  { _row('Proceso', r['proceso_text']) if r else '' }
                  { _row('Material', r['material_desc']) if (r and 'material_desc' in r.keys()) else '' }
                  { _row('Fecha de Pedido', r['fecha_pedido']) if r else '' }
                  { _row('Factura', r['factura']) if r else '' }
                  { _row('Guía Remisión', r['guia_remision']) if r else '' }
                  { _row('Antecedente', r['antecedente']) if r else '' }
                  { _row('Observación', r['observacion']) if r else '' }
                </table>

                <!-- Botón CTA -->
                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link_responder}"
                     style="display:inline-block;background:#2563eb;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Ingresar y responder medidas
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  Una vez completes la causa, acción preventiva y correctiva,
                  tu jefe revisará y aprobará la respuesta técnica.
                </div>
              </td>
            </tr>

            <!-- Pie -->
            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>

          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

    _send_mail_safe(u["email"], subject, text_body, html_body=html_body)

def _notify_jefe_respuesta_listo(conn, aprobador_id, reclamo_codigo, imputado_username):
    jefe = _get_user_basic(conn, aprobador_id)
    if not jefe or ("email" not in jefe.keys()) or not jefe["email"]:
        return

    # --- Datos del reclamo (por código) ---
    cur = conn.cursor()
    cur.execute("""
        SELECT id, fecha_reclamo, tipo_reclamo, tipo_tramite,
               cliente_nombre, proceso_text, material_desc,
               fecha_pedido, factura, guia_remision,
               antecedente, observacion
        FROM reclamos
        WHERE codigo = ?
        LIMIT 1
    """, (reclamo_codigo,))
    r = cur.fetchone()

    # Imputados del caso (por si hay más de uno)
    imputados = imputado_username or ""
    if r:
        cur.execute("""
            SELECT GROUP_CONCAT(u.username, ', ') AS lista
            FROM reclamo_imputados ri
            JOIN usuarios u ON u.id = ri.imputado_id
            WHERE ri.reclamo_id = ?
        """, (r["id"],))
        row = cur.fetchone()
        if row and row["lista"]:
            imputados = row["lista"]

    # Link directo al tab "Por aprobar (Jefe)" para validar la respuesta técnica
    try:
        link_validar = url_for("reclamos", _external=True) + "?tab=aprobar"
    except Exception:
        link_validar = "https://tu-sistema/reclamos?tab=aprobar"

    nombre = (
        jefe['nombre_completo']
        if 'nombre_completo' in jefe.keys() and jefe['nombre_completo']
        else jefe['username']
    )

    subject = f"[Oportunidad de Mejora] Validar respuesta técnica {reclamo_codigo}"

    # Texto plano (fallback)
    text_body = f"""Hola {nombre},

El usuario {imputado_username} ha registrado la respuesta técnica para la
Oportunidad de Mejora {reclamo_codigo}.

Por favor revisa las medidas propuestas (causa, acción preventiva y correctiva)
y aprueba o rechaza la respuesta.

Ir al sistema: {link_validar}

Este es un mensaje automático.
"""

    # ---------- HTML mejorado ----------
    def _row(lbl, val):
        val = (val or "").replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#fee2e2;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;
               font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">
            <!-- Encabezado -->
            <tr>
              <td style="background:#1d4ed8;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;
                            letter-spacing:.08em;opacity:.9;">
                  Oportunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Validar respuesta técnica {reclamo_codigo}
                </div>
                <div style="font-size:12px;opacity:.9;margin-top:6px;">
                  Hola {nombre}, el usuario {imputado_username} ha registrado su respuesta
                  y está pendiente de tu aprobación.
                </div>
              </td>
            </tr>

            <!-- Cuerpo -->
            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row('Código', reclamo_codigo)}
                  { _row('Fecha OM', r['fecha_reclamo']) if r else '' }
                  { _row('Tipo de OM', r['tipo_reclamo']) if r else '' }
                  { _row('Tipo de Trámite', r['tipo_tramite']) if r else '' }
                  { _row('Cliente', r['cliente_nombre']) if r else '' }
                  { _row('Proceso', r['proceso_text']) if r else '' }
                  { _row('Material', r['material_desc']) if (r and 'material_desc' in r.keys()) else '' }
                  { _row('Fecha de Pedido', r['fecha_pedido']) if r else '' }
                  { _row('Factura', r['factura']) if r else '' }
                  { _row('Guía Remisión', r['guia_remision']) if r else '' }
                  {_row('Imputados', imputados)}
                  { _row('Antecedente', r['antecedente']) if r else '' }
                  { _row('Observación', r['observacion']) if r else '' }
                </table>

                <!-- Botón CTA -->
                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link_validar}"
                     style="display:inline-block;background:#2563eb;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Revisar y validar respuesta técnica
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  También puedes ingresar al módulo de reclamos desde el sistema
                  y abrir la bandeja <strong>“Por aprobar (Jefe)”</strong> para validar las respuestas.
                </div>
              </td>
            </tr>

            <!-- Pie -->
            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>

          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

    _send_mail_safe(jefe["email"], subject, text_body, html_body=html_body)



def _notify_imputado_respuesta_rechazada(conn, imputado_id, reclamo_codigo, motivo):
    u = _get_user_basic(conn, imputado_id)
    if not u or ("email" not in u.keys()) or not u["email"]:
        return

    # --- Datos del reclamo (por código) ---
    cur = conn.cursor()
    cur.execute("""
        SELECT id, fecha_reclamo, tipo_reclamo, tipo_tramite,
               cliente_nombre, proceso_text, material_desc,
               fecha_pedido, factura, guia_remision,
               antecedente, observacion
        FROM reclamos
        WHERE codigo = ?
        LIMIT 1
    """, (reclamo_codigo,))
    r = cur.fetchone()

    # Link directo al tab "Soy responsable" para que ajuste la respuesta
    try:
        link_responder = url_for("reclamos", _external=True) + "?tab=imputado"
    except Exception:
        link_responder = "https://tu-sistema/reclamos?tab=imputado"

    nombre = (
        u['nombre_completo']
        if 'nombre_completo' in u.keys() and u['nombre_completo']
        else u['username']
    )

    motivo_txt = (motivo or "Sin detalle").strip()

    subject = f"[Oportunidad de Mejora] Ajuste requerido en respuesta {reclamo_codigo}"

    # Texto plano (fallback)
    text_body = f"""Hola {nombre},

Tu respuesta técnica para la Oportunidad de Mejora {reclamo_codigo} fue RECHAZADA
y requiere ajustes.

Motivo del rechazo:
{motivo_txt or 'Sin detalle'}

Por favor actualiza la causa, acción preventiva y acción correctiva.

Ir al sistema: {link_responder}

Este es un mensaje automático.
"""

    # ---------- HTML mejorado ----------
    def _row(lbl, val):
        val = (val or "").replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#fef3c7;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;
               font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">
            <!-- Encabezado -->
            <tr>
              <td style="background:#b45309;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;
                            letter-spacing:.08em;opacity:.9;">
                  Oportunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Ajuste requerido en respuesta {reclamo_codigo}
                </div>
                <div style="font-size:12px;opacity:.9;margin-top:6px;">
                  Hola {nombre}, tu respuesta técnica fue rechazada y requiere ajustes.
                </div>
              </td>
            </tr>

            <!-- Cuerpo -->
            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row('Código', reclamo_codigo)}
                  { _row('Fecha OM', r['fecha_reclamo']) if r else '' }
                  { _row('Tipo de OM', r['tipo_reclamo']) if r else '' }
                  { _row('Tipo de Trámite', r['tipo_tramite']) if r else '' }
                  { _row('Cliente', r['cliente_nombre']) if r else '' }
                  { _row('Proceso', r['proceso_text']) if r else '' }
                  { _row('Material', r['material_desc']) if (r and 'material_desc' in r.keys()) else '' }
                  { _row('Fecha de Pedido', r['fecha_pedido']) if r else '' }
                  { _row('Factura', r['factura']) if r else '' }
                  { _row('Guía Remisión', r['guia_remision']) if r else '' }
                  { _row('Antecedente', r['antecedente']) if r else '' }
                  { _row('Observación', r['observacion']) if r else '' }
                  {_row('Motivo del rechazo', motivo_txt or 'Sin detalle')}
                </table>

                <!-- Botón CTA -->
                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link_responder}"
                     style="display:inline-block;background:#f97316;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Ajustar respuesta técnica
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  Ingresa al módulo de reclamos, pestaña
                  <strong>“Soy responsable”</strong>, selecciona la OM y actualiza
                  la causa raíz, acción preventiva y acción correctiva.
                </div>
              </td>
            </tr>

            <!-- Pie -->
            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>

          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

    _send_mail_safe(u["email"], subject, text_body, html_body=html_body)


def _notify_creador_respuesta_aprobada(conn, creador_id, reclamo_codigo, imputado_username):
    c = _get_user_basic(conn, creador_id)
    if not c or ("email" not in c.keys()) or not c["email"]:
        return

    # --- Datos del reclamo (por código) ---
    cur = conn.cursor()
    cur.execute("""
        SELECT id, fecha_reclamo, tipo_reclamo, tipo_tramite,
               cliente_nombre, proceso_text, material_desc,
               fecha_pedido, factura, guia_remision,
               antecedente, observacion
        FROM reclamos
        WHERE codigo = ?
        LIMIT 1
    """, (reclamo_codigo,))
    r = cur.fetchone()

    # Link directo a "Mis reclamos"
    try:
        link_mis_reclamos = url_for("reclamos", _external=True) + "?tab=mios"
    except Exception:
        link_mis_reclamos = "https://tu-sistema/reclamos?tab=mios"

    nombre = (
        c['nombre_completo']
        if 'nombre_completo' in c.keys() and c['nombre_completo']
        else c['username']
    )

    responsable = imputado_username or "Responsable técnico"

    subject = f"[Oportunidad de Mejora] Respuesta final aprobada {reclamo_codigo}"

    # Texto plano (fallback)
    text_body = f"""Hola {nombre},

El responsable {responsable} registró y su jefe aprobó la respuesta técnica
de la Oportunidad de Mejora {reclamo_codigo}.

Ya puedes consultarla en el sistema.

Ir al sistema: {link_mis_reclamos}

Este es un mensaje automático.
"""

    # ---------- HTML mejorado ----------
    def _row(lbl, val):
        val = (val or "").replace("\n", "<br>")
        return (
            "<tr>"
            f"<td style='width:210px;background:#dcfce7;font-weight:600;"
            "padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{lbl}</td>"
            f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:13px;'>"
            f"{val}</td>"
            "</tr>"
        )

    html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f4f6;
               font-family:Segoe UI,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
           style="background:#f3f4f6;padding:24px 0;">
      <tr>
        <td align="center">
          <table role="presentation" width="720" cellpadding="0" cellspacing="0"
                 style="max-width:720px;background:#ffffff;border-radius:8px;
                        border:1px solid #e5e7eb;overflow:hidden;">
            <!-- Encabezado -->
            <tr>
              <td style="background:#15803d;padding:16px 20px;color:#ffffff;">
                <div style="font-size:12px;text-transform:uppercase;
                            letter-spacing:.08em;opacity:.9;">
                  Oportunidad de Mejora
                </div>
                <div style="font-size:18px;font-weight:700;margin-top:4px;">
                  Respuesta final aprobada {reclamo_codigo}
                </div>
                <div style="font-size:12px;opacity:.9;margin-top:6px;">
                  Hola {nombre}, la respuesta técnica de esta OM fue aprobada por el jefe del responsable.
                </div>
              </td>
            </tr>

            <!-- Cuerpo -->
            <tr>
              <td style="padding:18px 20px 10px 20px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                       style="border-collapse:collapse;">
                  {_row('Código', reclamo_codigo)}
                  {_row('Responsable', responsable)}
                  { _row('Fecha OM', r['fecha_reclamo']) if r else '' }
                  { _row('Tipo de OM', r['tipo_reclamo']) if r else '' }
                  { _row('Tipo de Trámite', r['tipo_tramite']) if r else '' }
                  { _row('Cliente', r['cliente_nombre']) if r else '' }
                  { _row('Proceso', r['proceso_text']) if r else '' }
                  { _row('Material', r['material_desc']) if (r and 'material_desc' in r.keys()) else '' }
                  { _row('Fecha de Pedido', r['fecha_pedido']) if r else '' }
                  { _row('Factura', r['factura']) if r else '' }
                  { _row('Guía Remisión', r['guia_remision']) if r else '' }
                  { _row('Antecedente', r['antecedente']) if r else '' }
                  { _row('Observación', r['observacion']) if r else '' }
                </table>

                <!-- Botón CTA -->
                <div style="margin-top:18px;margin-bottom:6px;text-align:left;">
                  <a href="{link_mis_reclamos}"
                     style="display:inline-block;background:#16a34a;color:#ffffff;
                            text-decoration:none;padding:10px 18px;border-radius:6px;
                            font-weight:600;font-size:13px;">
                    Ver respuesta técnica
                  </a>
                </div>

                <div style="font-size:11px;color:#6b7280;margin-top:8px;">
                  También puedes ingresar al módulo de reclamos y abrir la pestaña
                  <strong>“Mis reclamos”</strong> para revisar el detalle completo de la OM.
                </div>
              </td>
            </tr>

            <!-- Pie -->
            <tr>
              <td style="padding:10px 20px 14px 20px;border-top:1px solid #e5e7eb;
                         font-size:11px;color:#9ca3af;">
                Este es un mensaje automático. No responda a este correo.
              </td>
            </tr>

          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

    _send_mail_safe(c["email"], subject, text_body, html_body=html_body)


import re

def _acciones_v2_start() -> str:
    """
    Fecha desde la cual aplicas el cambio SOLO hacia adelante.
    Puedes setearlo en config: app.config['RECL_ACCIONES_V2_START'] = '2025-12-20'
    """
    return current_app.config.get("RECL_ACCIONES_V2_START", "2025-12-20")

_LINE_RE = re.compile(
    r"^\s*-\s*(.*?)\s*(?:\((\d{4}-\d{2}-\d{2})\))?\s*$"
)

def _parse_acciones_txt(txt: str | None):
    """
    Convierte texto tipo:
      - Hacer algo (2025-12-31)
      - Otra cosa
    en lista [{descripcion, fecha_compromiso}]
    """
    items = []
    for line in (txt or "").splitlines():
        line = (line or "").strip()
        if not line:
            continue
        m = _LINE_RE.match(line)
        if m:
            desc = (m.group(1) or "").strip()
            fecha = (m.group(2) or "").strip()
        else:
            desc = line.lstrip("-").strip()
            fecha = ""
        if desc or fecha:
            items.append({"id": None, "descripcion": desc, "fecha_compromiso": fecha})
    return items

def _equipo_es_v2(conn: sqlite3.Connection, equipo_id: int) -> bool:
    """
    True si:
    - ya tiene acciones en tabla nueva, o
    - su fecha_asignacion es >= a la fecha de activación.
    """
    cur = conn.cursor()

    # 1) Si ya existen acciones, ya es v2
    cur.execute("SELECT 1 FROM reclamo_equipo_acciones WHERE equipo_id = ? LIMIT 1", (equipo_id,))
    if cur.fetchone():
        return True

    # 2) Si es posterior a la fecha de activación, es v2
    cur.execute("SELECT fecha_asignacion FROM reclamo_equipo WHERE id = ? LIMIT 1", (equipo_id,))
    row = cur.fetchone()
    fa = (row["fecha_asignacion"] if row and ("fecha_asignacion" in row.keys()) else None) if row else None
    if not fa:
        return False

    # fa = "YYYY-MM-DD HH:MM:SS" -> comparar solo la fecha
    fa_date = fa[:10]
    return fa_date >= _acciones_v2_start()

# =========================================================
#   REGISTER ROUTES
# =========================================================

def register_reclamos_routes(app):

    # ----------------------------------------------
    # API clientes para autocomplete
    # ----------------------------------------------
    @app.route('/api/reclamos/clientes', methods=['GET'], endpoint='api_reclamos_clientes')
    @require_login
    def api_reclamos_clientes():
        q = (request.args.get('q') or '').strip()
        conn = get_db()
        cur = conn.cursor()

        sql = """
            SELECT id, nombre, identificacion, email, telefono, direccion
            FROM terceros
            WHERE tipo = 'C'
        """
        params = []
        if q:
            sql += " AND (nombre LIKE ? OR identificacion LIKE ?)"
            like = f"%{q}%"
            params.extend([like, like])

        sql += " ORDER BY nombre LIMIT 25"
        cur.execute(sql, params)

        items = []
        for r in cur.fetchall():
            items.append({
                "id": r["id"],
                "nombre": r["nombre"],
                "identificacion": r["identificacion"],
                "email": r["email"],
                "telefono": r["telefono"],
                "direccion": r["direccion"],
            })

        conn.close()
        return jsonify(ok=True, items=items)



    @app.route('/reclamos/api/subtipos', methods=['GET'], endpoint='reclamos_api_subtipos')
    @require_login
    def reclamos_api_subtipos():
        tipo_id = request.args.get('tipo_id', type=int)
        if not tipo_id:
            return jsonify([])

        conn = get_db()
        _ensure_param_tables(conn)
        cur = conn.cursor()

        # Intentamos usar el grupo por nombre; si no existe, caemos a 4520
        cur.execute("SELECT id FROM param_groups WHERE nombre = 'RECL_SUBTIPO' LIMIT 1")
        row = cur.fetchone()
        gid = row["id"] if row else 4520

        cur.execute("""
            SELECT id, nombre, valor, orden
            FROM param_values
            WHERE group_id = ?
              AND COALESCE(activo, 1) = 1
              AND COALESCE(parent_id, 0) = ?
            ORDER BY orden, valor
        """, (gid, tipo_id))

        data = [{"id": r["id"], "nombre": r["nombre"], "valor": r["valor"]} for r in cur.fetchall()]
        conn.close()
        return jsonify(data)




    @app.post("/reclamos/equipo/<int:equipo_id>/acciones")
    @require_login
    def api_save_equipo_acciones(equipo_id):
        if not _can_edit_equipo(equipo_id):
            abort(403)

        # 1) Leer JSON
        data = request.get_json(silent=True) or {}

        # 2) Adapter: soportar payload “simple” (strings + fechas sueltas) además del payload v2 (listas)
        #    - causa + fecha_causa
        #    - preventiva/control + fecha_preventiva/fecha_control
        #    - correctiva + fecha_correctiva
        #
        #    Si ya viene como lista (v2), se respeta.
        def _as_list_value(val):
            """Convierte val en lista de dicts si viene como dict; si viene str no aplica aquí."""
            if val is None:
                return None
            if isinstance(val, list):
                return val
            if isinstance(val, dict):
                return [val]
            return val  # puede ser str u otro

        data["causas"] = _as_list_value(data.get("causas"))
        data["control"] = _as_list_value(data.get("control"))
        data["correctiva"] = _as_list_value(data.get("correctiva"))

        # Si NO vienen como listas válidas, pero llegan como simple (string), armamos v2
        # CAUSA
        if not isinstance(data.get("causas"), list):
            causa_txt = (data.get("causa") or "").strip()
            causa_fecha = (data.get("fecha_causa") or data.get("causa_fecha") or "").strip()
            # Si "causas" viene como string accidental, lo tomamos como descripción
            if isinstance(data.get("causas"), str) and not causa_txt:
                causa_txt = (data.get("causas") or "").strip()

            if causa_txt or causa_fecha:
                data["causas"] = [{
                    "descripcion": causa_txt,
                    "fecha_compromiso": causa_fecha,
                }]
            else:
                data["causas"] = []

        # CONTROL / PREVENTIVA
        if not isinstance(data.get("control"), list):
            ctrl_txt = (data.get("control") or "").strip() if isinstance(data.get("control"), str) else ""
            # “preventiva” como alias de control
            if not ctrl_txt:
                ctrl_txt = (data.get("preventiva") or "").strip()

            ctrl_fecha = (data.get("fecha_control") or data.get("fecha_preventiva") or "").strip()

            if ctrl_txt or ctrl_fecha:
                data["control"] = [{
                    "descripcion": ctrl_txt,
                    "fecha_compromiso": ctrl_fecha,
                }]
            else:
                data["control"] = []

        # CORRECTIVA
        if not isinstance(data.get("correctiva"), list):
            corr_txt = (data.get("correctiva") or "").strip() if isinstance(data.get("correctiva"), str) else ""
            # por si viene en clave "correctiva" (string) y/o "correctiva_texto"
            if not corr_txt:
                corr_txt = (data.get("correctiva_texto") or "").strip()

            corr_fecha = (data.get("fecha_correctiva") or "").strip()

            if corr_txt or corr_fecha:
                data["correctiva"] = [{
                    "descripcion": corr_txt,
                    "fecha_compromiso": corr_fecha,
                }]
            else:
                data["correctiva"] = []

        # 3) Ya en formato v2 (listas)
        causas     = data.get("causas") or []
        control    = data.get("control") or []
        correctiva = data.get("correctiva") or []

        def _norm(items, label):
            out = []
            for it in items:
                it = it or {}
                if not isinstance(it, dict):
                    # si por algún motivo llega mal, lo ignoramos
                    continue

                desc = (it.get("descripcion") or "").strip()
                fecha = (it.get("fecha_compromiso") or "").strip()

                # ignorar fila vacía
                if not desc and not fecha:
                    continue

                if not desc:
                    return None, f"Falta descripción en {label}."
                if not fecha:
                    return None, f"Falta fecha compromiso en {label}."
                if not _is_date_yyyy_mm_dd(fecha):
                    return None, f"Fecha compromiso inválida en {label}. Use YYYY-MM-DD."

                out.append((desc, fecha))
            return out, None

        causas_n, err = _norm(causas, "Causa")
        if err:
            return jsonify(ok=False, error=err), 400

        control_n, err = _norm(control, "Acción de Control")
        if err:
            return jsonify(ok=False, error=err), 400

        correctiva_n, err = _norm(correctiva, "Acción Correctiva")
        if err:
            return jsonify(ok=False, error=err), 400
 
        user_id = session.get("user_id") or session.get("id")
        now = _now_str()

        conn = get_db()

        # ✅ CAMBIO CLAVE: asegurar esquema SIEMPRE antes de usar la tabla v2
        ensure_reclamos_catalogos(conn)

        cur = conn.cursor()

        # Reemplazo total (simple, robusto y sin “histórico”)
        cur.execute("DELETE FROM reclamo_equipo_acciones WHERE equipo_id = ?", (equipo_id,))

        def _insert_many(tipo, items):
            for desc, fecha in items:
                cur.execute("""
                    INSERT INTO reclamo_equipo_acciones (
                        equipo_id, tipo, descripcion, fecha_compromiso, created_at, created_by
                    ) VALUES (?, ?, ?, ?, ?, ?)
                """, (equipo_id, tipo, desc, fecha, now, user_id))

        _insert_many("CAUSA", causas_n)
        _insert_many("CONTROL", control_n)
        _insert_many("CORRECTIVA", correctiva_n)

        # (Opcional recomendado) espejo a reclamo_equipo.respuesta_* para compatibilidad
        def _join(items):
            # "- texto (YYYY-MM-DD)"
            return "\n".join([f"- {d} ({f})" for d, f in items])

        cur.execute("""
            UPDATE reclamo_equipo
            SET
                respuesta_causa      = ?,
                respuesta_preventiva = ?,
                respuesta_correctiva = ?,
                fecha_respuesta      = COALESCE(fecha_respuesta, ?)
            WHERE id = ?
        """, (_join(causas_n), _join(control_n), _join(correctiva_n), now, equipo_id))

        conn.commit()
        return jsonify(ok=True)


    import re
    from flask import jsonify, abort, request, session
    # asumiendo que ya tienes: get_db, require_login, _can_edit_equipo, ensure_reclamos_catalogos

    def _row_get(row, key, idx, default=""):
        """Soporta sqlite Row (dict-like) o tupla."""
        if not row:
            return default
        try:
            if hasattr(row, "keys"):
                return row[key] if key in row.keys() else default
        except Exception:
            pass
        try:
            return row[idx]
        except Exception:
            return default

    def _parse_acciones_txt_v2(txt):
        """
        Parsea formato:
        - descripcion (YYYY-MM-DD)
        y retorna:
        [{"id": None, "descripcion": "...", "fecha_compromiso": "YYYY-MM-DD"}, ...]
        """
        out = []
        for line in (txt or "").splitlines():
            line = (line or "").strip()
            if not line:
                continue
            if line.startswith("-"):
                line = line[1:].strip()

            # Captura fecha al final si existe
            m = re.match(r"^(.*?)(?:\s*\((\d{4}-\d{2}-\d{2})\))?\s*$", line)
            if not m:
                continue

            desc = (m.group(1) or "").strip()
            fecha = (m.group(2) or "").strip()

            # ignora fila vacía
            if not desc and not fecha:
                continue

            out.append({
                "id": None,
                "descripcion": desc,
                "fecha_compromiso": fecha or ""
            })
        return out


    @app.get("/reclamos/equipo/<int:equipo_id>/acciones")
    @require_login
    def api_get_equipo_acciones(equipo_id):

        if not _can_edit_equipo(equipo_id):
            abort(403)

        conn = get_db()

        # ✅ Asegura que exista la tabla v2 antes de consultar
        ensure_reclamos_catalogos(conn)

        cur = conn.cursor()

        # ✅ V2: intenta leer SIEMPRE desde la tabla nueva.
        # Esto evita el problema de que _equipo_es_v2() falle y no cargue fechas.
        cur.execute("""
            SELECT id, tipo, descripcion, fecha_compromiso
            FROM reclamo_equipo_acciones
            WHERE equipo_id = ?
            ORDER BY id ASC
        """, (equipo_id,))
        rows = cur.fetchall() or []

        if rows:
            causas, control, correctiva = [], [], []

            for r in rows:
                rid   = _row_get(r, "id", 0, None)
                tipo  = (_row_get(r, "tipo", 1, "") or "").strip()
                desc  = _row_get(r, "descripcion", 2, "") or ""
                fecha = _row_get(r, "fecha_compromiso", 3, "") or ""

                item = {"id": rid, "descripcion": desc, "fecha_compromiso": fecha}

                if tipo == "CAUSA":
                    causas.append(item)
                elif tipo == "CONTROL":
                    control.append(item)
                elif tipo == "CORRECTIVA":
                    correctiva.append(item)

            return jsonify(ok=True, causas=causas, control=control, correctiva=correctiva)

        # 🔁 Fallback histórico: lee campos viejos y parsea también la fecha si está en "(YYYY-MM-DD)"
        cur.execute("""
            SELECT respuesta_causa, respuesta_preventiva, respuesta_correctiva
            FROM reclamo_equipo
            WHERE id = ?
            LIMIT 1
        """, (equipo_id,))
        row = cur.fetchone()

        causas_txt     = _row_get(row, "respuesta_causa", 0, "")
        control_txt    = _row_get(row, "respuesta_preventiva", 1, "")
        correctiva_txt = _row_get(row, "respuesta_correctiva", 2, "")

        causas     = _parse_acciones_txt_v2(causas_txt)
        control    = _parse_acciones_txt_v2(control_txt)
        correctiva = _parse_acciones_txt_v2(correctiva_txt)

        return jsonify(ok=True, causas=causas, control=control, correctiva=correctiva)


    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/json")
    @require_login
    def equipo_respuestas_json(reclamo_id):
        uid = _current_user_id()

        db = get_db()
        ensure_reclamos_schema(db)
        ensure_reclamo_respuestas_equipo_schema(db)
        

        # ✅ Solo responsable aprobado, miembros, creador o admin/coordinador
        if not (_puede_ver_equipo(reclamo_id, uid) or _is_admin_like()):
            return jsonify({"error": "No autorizado"}), 403

        imputacion_id = request.args.get("imputacion_id", type=int)

        # Si el front no manda imputacion_id, tomamos la última imputación del reclamo
        if not imputacion_id:
            row_imp = db.execute("""
                SELECT id
                FROM reclamo_imputados
                WHERE reclamo_id = ?
                ORDER BY id DESC
                LIMIT 1
            """, (reclamo_id,)).fetchone()
            imputacion_id = row_imp["id"] if row_imp else None

        # -------------------------
        # Caso 1: SIN imputación
        # -------------------------
        if not imputacion_id:
            rows = db.execute("""
                SELECT
                    er.id AS equipo_id,
                    er.id AS id,                 -- por compatibilidad si tu front usa "id"
                    er.reclamo_id,
                    NULL AS imputacion_id,
                    er.usuario_id,
                    er.puede_responder,
                    er.activo,
                    er.creado_por,
                    er.creado_at,
                    u.username,
                    COALESCE(u.nombre_completo, u.username) AS nombre,
                    u.rol AS rol,
                    d.nombre AS departamento,
                    0 AS tiene_respuesta,
                    NULL AS fecha_respuesta,
                    NULL AS respuesta_id
                FROM reclamo_equipo_respuestas er
                JOIN usuarios u ON er.usuario_id = u.id
                LEFT JOIN departamentos d ON u.departamento_id = d.id
                WHERE er.reclamo_id = ?
                AND er.activo = 1
                ORDER BY nombre
            """, (reclamo_id,)).fetchall()

        # -------------------------
        # Caso 2: CON imputación
        # -------------------------
        else:
            rows = db.execute("""
                SELECT
                    er.id AS equipo_id,
                    er.id AS id,                 -- por compatibilidad
                    er.reclamo_id,
                    er.imputacion_id,
                    er.usuario_id,
                    u.username,
                    COALESCE(u.nombre_completo, u.username) AS nombre,
                    er.puede_responder,
                    er.activo,
                    u.rol AS rol,
                    d.nombre AS departamento,

                    CASE WHEN rrmax.id IS NULL THEN 0 ELSE 1 END AS tiene_respuesta,
                    rre.created_at AS fecha_respuesta,
                    rrmax.id AS respuesta_id

                FROM reclamo_equipo_respuestas er
                JOIN usuarios u ON u.id = er.usuario_id
                LEFT JOIN departamentos d ON d.id = u.departamento_id

                LEFT JOIN (
                    SELECT MAX(id) AS id, reclamo_id, imputacion_id, miembro_id
                    FROM reclamo_respuestas_equipo
                    WHERE activo = 1
                    GROUP BY reclamo_id, imputacion_id, miembro_id
                ) rrmax
                ON rrmax.reclamo_id   = er.reclamo_id
                AND rrmax.imputacion_id = er.imputacion_id
                AND rrmax.miembro_id    = er.usuario_id

                LEFT JOIN reclamo_respuestas_equipo rre
                ON rre.id = rrmax.id

                WHERE er.reclamo_id = ?
                AND er.imputacion_id = ?
                AND er.activo = 1

                ORDER BY nombre
            """, (reclamo_id, imputacion_id)).fetchall()

        # -------------------------
        # Payload (para ambos casos)
        # -------------------------
        payload = []
        for row in rows:
            r = dict(row)  # convertir Row -> dict (así puedes usar .get)

            # Siempre tendremos equipo_id e id por el SELECT (alias)
            equipo_id = r.get("equipo_id")
            if equipo_id is None:
                # fallback por si en algún momento cambias el SELECT
                equipo_id = r.get("id")

            payload.append({
                # id del registro en reclamo_equipo_respuestas (sirve para Quitar)
                "id": r["equipo_id"],

                # explícitos para que el front no se confunda
                "equipo_id": r["equipo_id"],
                "reclamo_id": r["reclamo_id"],
                "imputacion_id": r.get("imputacion_id"),

                # ESTE es el id del usuario real (miembro)
                "usuario_id": r["usuario_id"],
                "miembro_id": r["usuario_id"],

                "nombre": r["nombre"],
                "username": r["username"],
                "rol": r["rol"],
                "departamento": r["departamento"],
                "puede_responder": bool(r["puede_responder"]),
                "tiene_respuesta": bool(r["tiene_respuesta"]),
                "fecha_respuesta": r["fecha_respuesta"],

                # id del registro en reclamo_respuestas_equipo (si existe)
                "respuesta_id": r.get("respuesta_id"),
            })


        return jsonify(payload)

        
    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/add", methods=["POST"])
    @require_login
    def equipo_respuestas_add(reclamo_id):
        payload = request.get_json(silent=True) or {}
        usuario_id = payload.get("usuario_id")
        imputacion_id = payload.get("imputacion_id")  # 👈 NUEVO

        uid = _current_user_id()
        if not uid:
            return jsonify({"error": "Sesión inválida. Vuelve a iniciar sesión."}), 401

        db = get_db()
        cur = db.cursor()

        ensure_reclamos_schema(db)

        # ✅ SOLO responsable aprobado o admin/coordinador
        if not (_puede_gestionar_equipo(reclamo_id, uid) or _is_admin_like()):
            return jsonify({"error": "No tienes permiso para modificar el equipo de respuestas."}), 403

        if not usuario_id:
            return jsonify({"error": "Selecciona un usuario para agregar al equipo."}), 400

        row_user = db.execute("SELECT id FROM usuarios WHERE id = ?", (usuario_id,)).fetchone()
        if not row_user:
            return jsonify({"error": "El usuario seleccionado no existe."}), 400

        # 👇 Si el front NO manda imputacion_id, lo intentamos deducir (sponsor = uid)
        if not imputacion_id:
            row_imp = db.execute("""
                SELECT id
                FROM reclamo_imputados
                WHERE reclamo_id = ?
                AND imputado_id = ?
                ORDER BY id DESC
                LIMIT 1
            """, (reclamo_id, uid)).fetchone()
            if not row_imp:
                return jsonify({"error": "No se pudo determinar imputación_id para este reclamo."}), 400
            imputacion_id = row_imp["id"]

        # ✅ Validar que esa imputación pertenece a este reclamo
        row_ok = db.execute("""
            SELECT 1
            FROM reclamo_imputados
            WHERE id = ?
            AND reclamo_id = ?
            LIMIT 1
        """, (imputacion_id, reclamo_id)).fetchone()
        if not row_ok:
            return jsonify({"error": "La imputación no pertenece a este reclamo."}), 400

        # (opcional) evita duplicado en el mismo reclamo (aunque cambie imputación)
        dup = db.execute("""
            SELECT 1
            FROM reclamo_equipo_respuestas
            WHERE reclamo_id = ?
            AND usuario_id = ?
            AND activo = 1
            LIMIT 1
        """, (reclamo_id, usuario_id)).fetchone()
        if dup:
            return jsonify({"error": "El usuario ya está en el equipo de respuestas."}), 400

        # ✅ INSERT CORRECTO (con imputacion_id)
        db.execute("""
            INSERT INTO reclamo_equipo_respuestas (
                reclamo_id, imputacion_id, usuario_id,
                puede_responder, activo, creado_por, creado_at
            ) VALUES (?, ?, ?, 1, 1, ?, ?)
        """, (reclamo_id, imputacion_id, usuario_id, uid, _now_iso()))
        db.commit()
        # ✅ Notificar por correo al miembro agregado
        try:
            row_r = db.execute(
                "SELECT codigo FROM reclamos WHERE id = ? LIMIT 1",
                (reclamo_id,)
            ).fetchone()
            reclamo_codigo = row_r["codigo"] if row_r and row_r["codigo"] else f"RECL#{reclamo_id}"

            resp_row = cur.fetchone()

            # sqlite3.Row -> dict (seguro)
            resp = dict(zip(resp_row.keys(), resp_row)) if resp_row else None

            resp_username = resp.get("username") if resp and resp.get("username") else f"UID {uid}"

            _notify_colaborador_asignado(
                db,
                int(usuario_id),
                reclamo_codigo,
                resp_username
            )
        except Exception:
            current_app.logger.exception(
                "No se pudo enviar correo al miembro agregado. reclamo_id=%s usuario_id=%s",
                reclamo_id, usuario_id
            )

        return jsonify({"ok": True})

    # Guarda Respuesta Miembros de Equipo
    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/guardar", methods=["POST"])
    @require_login
    def equipo_respuestas_guardar(reclamo_id):
        current_app.logger.info("[MAIL_EQUIPO] entra a equipo_respuestas_guardar reclamo_id=%s", reclamo_id)

        payload = request.get_json(silent=True) or {}

        uid = _current_user_id()
        if not uid:
            return jsonify({"error": "Sesión inválida. Vuelve a iniciar sesión."}), 401

        imputacion_id = payload.get("imputacion_id")

        # =========================================================
        # Compatibilidad: payload viejo (campos simples)
        # =========================================================
        causa = (payload.get("causa") or "").strip()
        preventiva = (payload.get("preventiva") or "").strip()
        correctiva = (payload.get("correctiva") or "").strip()

        fecha_causa = (payload.get("fecha_causa") or "").strip()
        fecha_preventiva = (payload.get("fecha_preventiva") or "").strip()
        fecha_correctiva = (payload.get("fecha_correctiva") or "").strip()

        metodo_analisis = (payload.get("metodo_analisis") or "").strip()

        why1 = (payload.get("why1") or "").strip()
        why2 = (payload.get("why2") or "").strip()
        why3 = (payload.get("why3") or "").strip()
        why4 = (payload.get("why4") or "").strip()
        why5 = (payload.get("why5") or "").strip()

        fish_metodo = (payload.get("fish_metodo") or "").strip()
        fish_maquinas = (payload.get("fish_maquinas") or "").strip()
        fish_materiales = (payload.get("fish_materiales") or "").strip()
        fish_personas = (payload.get("fish_personas") or "").strip()
        fish_entorno = (payload.get("fish_entorno") or "").strip()
        fish_medicion = (payload.get("fish_medicion") or "").strip()

        # =========================================================
        # Nuevo payload múltiple
        # =========================================================
        causas_payload = payload.get("causas")
        control_payload = payload.get("control")
        correctiva_payload = payload.get("correctiva")

        # Si no viene el nuevo formato, armamos listas con el formato viejo
        if causas_payload is None and control_payload is None and not isinstance(correctiva_payload, list):
            causas_payload = []
            control_payload = []
            correctiva_payload = []

            if causa or fecha_causa:
                causas_payload.append({
                    "descripcion": causa,
                    "fecha_compromiso": fecha_causa
                })

            if preventiva or fecha_preventiva:
                control_payload.append({
                    "descripcion": preventiva,
                    "fecha_compromiso": fecha_preventiva
                })

            if correctiva or fecha_correctiva:
                correctiva_payload.append({
                    "descripcion": correctiva,
                    "fecha_compromiso": fecha_correctiva
                })

        causas_n, err = _normalizar_acciones_payload(causas_payload, "Causa")
        if err:
            return jsonify({"error": err}), 400

        controles_n, err = _normalizar_acciones_payload(control_payload, "Acción de Control")
        if err:
            return jsonify({"error": err}), 400

        correctivas_n, err = _normalizar_acciones_payload(correctiva_payload, "Acción Correctiva")
        if err:
            return jsonify({"error": err}), 400

        # Validación mínima
        if not causas_n or not controles_n or not correctivas_n:
            return jsonify({"error": "Debes ingresar al menos una Causa, una Acción de Control y una Acción Correctiva."}), 400

        db = get_db()
        ensure_reclamos_schema(db)
        ensure_reclamo_respuestas_equipo_schema(db)
        ensure_reclamo_respuesta_equipo_acciones_schema(db)

        # Si no viene imputacion_id, deducirla desde la asignación del equipo del usuario actual
        if not imputacion_id:
            row_eq = db.execute("""
                SELECT imputacion_id
                FROM reclamo_equipo_respuestas
                WHERE reclamo_id = ?
                AND usuario_id = ?
                AND activo = 1
                ORDER BY id DESC
                LIMIT 1
            """, (reclamo_id, uid)).fetchone()

            if not row_eq or not row_eq["imputacion_id"]:
                return jsonify({"error": "No se pudo determinar imputación_id para guardar la respuesta."}), 400

            imputacion_id = row_eq["imputacion_id"]

        # Validar permisos
        row_perm = db.execute("""
            SELECT 1
            FROM reclamo_equipo_respuestas
            WHERE reclamo_id = ?
            AND imputacion_id = ?
            AND usuario_id = ?
            AND activo = 1
            AND puede_responder = 1
            LIMIT 1
        """, (reclamo_id, imputacion_id, uid)).fetchone()

        if not row_perm and not _is_admin_like():
            return jsonify({"error": "No tienes permiso para responder en este reclamo."}), 403

        # ¿Existe respuesta previa?
        row_exist = db.execute("""
            SELECT id
            FROM reclamo_respuestas_equipo
            WHERE reclamo_id = ?
            AND imputacion_id = ?
            AND miembro_id = ?
            AND activo = 1
            ORDER BY id DESC
            LIMIT 1
        """, (reclamo_id, imputacion_id, uid)).fetchone()

        es_nueva_respuesta = row_exist is None
        now = _now_iso()

        # Resumen legacy para compatibilidad
        causa_txt = "\n".join([f"- {desc} ({fecha})" for desc, fecha in causas_n])
        preventiva_txt = "\n".join([f"- {desc} ({fecha})" for desc, fecha in controles_n])
        correctiva_txt = "\n".join([f"- {desc} ({fecha})" for desc, fecha in correctivas_n])

        primera_fecha_causa = causas_n[0][1] if causas_n else ""
        primera_fecha_control = controles_n[0][1] if controles_n else ""
        primera_fecha_correctiva = correctivas_n[0][1] if correctivas_n else ""

        if row_exist:
            respuesta_equipo_id = row_exist["id"]

            db.execute("""
                UPDATE reclamo_respuestas_equipo
                SET metodo_analisis=?,
                    causa=?, preventiva=?, correctiva=?,
                    fecha_causa=?, fecha_preventiva=?, fecha_correctiva=?,
                    fish_metodo=?, fish_maquinas=?, fish_materiales=?,
                    fish_personas=?, fish_entorno=?, fish_medicion=?,
                    why1=?, why2=?, why3=?, why4=?, why5=?,
                    created_at=?, created_by=?
                WHERE id = ?
            """, (
                metodo_analisis,
                causa_txt, preventiva_txt, correctiva_txt,
                primera_fecha_causa, primera_fecha_control, primera_fecha_correctiva,
                fish_metodo, fish_maquinas, fish_materiales, fish_personas, fish_entorno, fish_medicion,
                why1, why2, why3, why4, why5,
                now, uid,
                respuesta_equipo_id
            ))
        else:
            cur = db.execute("""
                INSERT INTO reclamo_respuestas_equipo (
                    reclamo_id, imputacion_id, miembro_id,
                    metodo_analisis,
                    causa, preventiva, correctiva,
                    fecha_causa, fecha_preventiva, fecha_correctiva,
                    fish_metodo, fish_maquinas, fish_materiales,
                    fish_personas, fish_entorno, fish_medicion,
                    why1, why2, why3, why4, why5,
                    activo, created_by, created_at
                )
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1,?,?,?)
            """, (
                reclamo_id, imputacion_id, uid,
                metodo_analisis,
                causa_txt, preventiva_txt, correctiva_txt,
                primera_fecha_causa, primera_fecha_control, primera_fecha_correctiva,
                fish_metodo, fish_maquinas, fish_materiales,
                fish_personas, fish_entorno, fish_medicion,
                why1, why2, why3, why4, why5,
                uid, now
            ))
            respuesta_equipo_id = cur.lastrowid

        # Guardar detalle múltiple real
        _save_respuesta_equipo_acciones(
            db,
            respuesta_id=respuesta_id,
            reclamo_id=reclamo_id,
            imputacion_id=imputacion_id,
            miembro_id=uid,
            causas=causas_items,
            control=control_items,
            correctiva=correctiva_items,
            user_id=uid
        )

        db.commit()

        # Notificar al sponsor SOLO la primera vez
        if es_nueva_respuesta:
            try:
                row_r = db.execute("""
                    SELECT codigo
                    FROM reclamos
                    WHERE id = ?
                    LIMIT 1
                """, (reclamo_id,)).fetchone()

                reclamo_codigo = row_r["codigo"] if row_r and row_r["codigo"] else f"RECL#{reclamo_id}"

                current_app.logger.info(
                    "[MAIL_EQUIPO] Se envía mail sponsor reclamo_id=%s imputacion_id=%s miembro_id=%s",
                    reclamo_id, imputacion_id, uid
                )

                _notify_sponsor_respuesta_equipo(
                    db,
                    int(imputacion_id),
                    int(uid),
                    reclamo_codigo
                )
            except Exception:
                current_app.logger.exception(
                    "[MAIL_EQUIPO] Error notificando sponsor por respuesta de equipo. reclamo_id=%s imputacion_id=%s miembro_id=%s",
                    reclamo_id, imputacion_id, uid
                )

        return jsonify({
            "ok": True,
            "respuesta_equipo_id": respuesta_equipo_id
        })



    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/aporte", methods=["GET"])
    @require_login
    def equipo_respuestas_ver_aporte(reclamo_id):
        db = get_db()
        uid = _current_user_id()

        imputacion_id = request.args.get("imputacion_id", type=int)
        miembro_id = request.args.get("miembro_id", type=int)

        # Si el front manda mal miembro_id (id del registro en tabla puente),
        # lo corregimos al usuario real.
        if miembro_id and not db.execute(
            "SELECT 1 FROM usuarios WHERE id = ?",
            (miembro_id,)
        ).fetchone():
            row_fix = db.execute("""
                SELECT usuario_id
                FROM reclamo_equipo_respuestas
                WHERE id = ?
                AND reclamo_id = ?
            """, (miembro_id, reclamo_id)).fetchone()
            if row_fix:
                miembro_id = row_fix["usuario_id"]

        if not imputacion_id or not miembro_id:
            return jsonify(ok=False, error="Falta imputacion_id o miembro_id"), 400

        ensure_reclamos_schema(db)
        ensure_reclamo_respuestas_equipo_schema(db)
        ensure_reclamo_respuesta_equipo_acciones_schema(db)

        can_view_all = _can_view_all_reclamos(db, uid)

        # Sponsor/imputado dueño de la imputación
        row_owner = db.execute("""
            SELECT 1
            FROM reclamo_imputados
            WHERE id = ?
            AND reclamo_id = ?
            AND imputado_id = ?
            LIMIT 1
        """, (imputacion_id, reclamo_id, uid)).fetchone()

        # Miembro del equipo viendo SU propio aporte
        is_same_member = (uid == miembro_id)

        # Opcional: validar además que realmente pertenezca al equipo de esa imputación
        row_team_member = db.execute("""
            SELECT 1
            FROM reclamo_equipo_respuestas
            WHERE reclamo_id = ?
            AND imputacion_id = ?
            AND usuario_id = ?
            LIMIT 1
        """, (reclamo_id, imputacion_id, uid)).fetchone()

        can_view = bool(can_view_all or row_owner or (is_same_member and row_team_member))

        if not can_view:
            return jsonify(ok=False, error="No autorizado"), 403

        row = db.execute("""
            SELECT
                rre.*,
                COALESCE(u.nombre_completo, u.username) AS miembro_nombre,
                u.username AS miembro_username,
                rre.metodo_analisis,
                rre.why1, rre.why2, rre.why3, rre.why4, rre.why5,
                rre.fish_metodo, rre.fish_maquinas, rre.fish_materiales,
                rre.fish_personas, rre.fish_entorno, rre.fish_medicion
            FROM reclamo_respuestas_equipo rre
            JOIN usuarios u
            ON u.id = rre.miembro_id
            WHERE rre.reclamo_id = ?
            AND rre.imputacion_id = ?
            AND rre.miembro_id = ?
            AND rre.activo = 1
            ORDER BY rre.id DESC
            LIMIT 1
        """, (reclamo_id, imputacion_id, miembro_id)).fetchone()

        item = dict(row) if row else {}

        if item.get("id"):
            acciones = _get_respuesta_equipo_acciones_full(db, int(item["id"]))
            item["causas"] = acciones.get("causas", [])
            item["control"] = acciones.get("control", [])
            item["correctiva_items"] = acciones.get("correctiva", [])
        else:
            item["causas"] = []
            item["control"] = []
            item["correctiva_items"] = []

        return jsonify(ok=True, item=item)



    @app.route("/reclamos/imputacion/<int:imputacion_id>/respuestas_equipo/json")
    @require_login
    def reclamo_respuestas_equipo_json(imputacion_id):
        db = get_db()
        rows = db.execute("""
            SELECT
                re.id,
                u.nombre_completo AS miembro_nombre,
                u.username AS miembro_username,
                re.causa,
                re.preventiva,
                re.correctiva,
                re.created_at
            FROM reclamo_respuestas_equipo re
            JOIN usuarios u ON u.id = re.miembro_id
            WHERE re.imputacion_id = ?
            ORDER BY re.created_at DESC
        """, (imputacion_id,)).fetchall()

        items = []
        for r in rows:
            items.append({
                "id": r["id"],
                "miembro_nombre": r["miembro_nombre"] or r["miembro_username"],
                "causa": r["causa"],
                "preventiva": r["preventiva"],
                "correctiva": r["correctiva"],
                "created_at": r["created_at"],
            })

        return jsonify({"ok": True, "items": items})

    @app.route("/reclamos/imputacion/<int:imputacion_id>/responder_equipo", methods=["POST"])
    @require_login
    def reclamo_responder_equipo(imputacion_id):
        db = get_db()
        ensure_reclamos_schema(db)
        ensure_reclamo_respuestas_equipo_schema(db)
        ensure_reclamo_respuesta_equipo_acciones_schema(db)

        uid = (
            session.get("user_id")
            or session.get("id")
            or session.get("usuario_id")
            or (session.get("user") or {}).get("id")
        )

        current_app.logger.debug(
            "[equipo_responder] imputacion_id=%s uid=%s", imputacion_id, uid
        )

        # 1) Obtener reclamo_id de la imputación
        imp = db.execute("""
            SELECT reclamo_id
            FROM reclamo_imputados
            WHERE id = ?
        """, (imputacion_id,)).fetchone()

        current_app.logger.debug(
            "[equipo_responder] imputacion -> %r", dict(imp) if imp else None
        )

        if not imp:
            return jsonify({"ok": False, "error": "Imputación no encontrada."}), 404

        reclamo_id = imp["reclamo_id"]
        current_app.logger.debug(
            "[equipo_responder] reclamo_id derivado=%s", reclamo_id
        )

        # 2) Validar que el usuario sea miembro del equipo para ese reclamo
        es_miembro = _es_miembro_equipo_reclamo(reclamo_id, uid)
        current_app.logger.debug(
            "[equipo_responder] _es_miembro_equipo_reclamo => %s", es_miembro
        )

        if not es_miembro:
            return jsonify({"ok": False, "error": "No tienes permiso para responder como equipo en esta OM."}), 403

        data = request.get_json(silent=True) or {}
        current_app.logger.debug("[equipo_responder] payload recibido=%r", data)

        metodo_analisis = (data.get("metodo_analisis") or "").strip()

        # =========================================================
        # NUEVO FORMATO: listas
        # =========================================================
        causas_items = data.get("causas") or []
        control_items = data.get("control") or []
        correctiva_items = data.get("correctiva") or []

        # =========================================================
        # COMPATIBILIDAD: si aún llega formato viejo, convertirlo
        # =========================================================
        if not causas_items and (data.get("causa") or data.get("fecha_causa")):
            causas_items = [{
                "descripcion": (data.get("causa") or "").strip(),
                "fecha_compromiso": (data.get("fecha_causa") or "").strip(),
            }]

        if not control_items and (data.get("preventiva") or data.get("fecha_preventiva")):
            control_items = [{
                "descripcion": (data.get("preventiva") or "").strip(),
                "fecha_compromiso": (data.get("fecha_preventiva") or "").strip(),
            }]

        if not correctiva_items and (data.get("correctiva") or data.get("fecha_correctiva")):
            correctiva_items = [{
                "descripcion": (data.get("correctiva") or "").strip(),
                "fecha_compromiso": (data.get("fecha_correctiva") or "").strip(),
            }]

        def _norm_items(items, nombre_bloque):
            out = []
            for i, it in enumerate(items, start=1):
                if not isinstance(it, dict):
                    continue

                descripcion = (it.get("descripcion") or "").strip()
                fecha = (it.get("fecha_compromiso") or "").strip()

                if not descripcion and not fecha:
                    continue

                if not descripcion:
                    return None, f"Falta descripción en {nombre_bloque} #{i}."
                if not fecha:
                    return None, f"Falta fecha en {nombre_bloque} #{i}."

                out.append({
                    "descripcion": descripcion,
                    "fecha_compromiso": fecha
                })

            return out, None

        causas_items, err = _norm_items(causas_items, "causa")
        if err:
            return jsonify({"ok": False, "error": err}), 400

        control_items, err = _norm_items(control_items, "acción de control")
        if err:
            return jsonify({"ok": False, "error": err}), 400

        correctiva_items, err = _norm_items(correctiva_items, "acción correctiva")
        if err:
            return jsonify({"ok": False, "error": err}), 400

        if not causas_items or not control_items or not correctiva_items:
            return jsonify({
                "ok": False,
                "error": "Debe ingresar al menos una causa, una acción de control y una acción correctiva."
            }), 400

        # =========================================================
        # Campos de análisis
        # =========================================================
        fish_metodo = (data.get("fish_metodo") or "").strip()
        fish_maquinas = (data.get("fish_maquinas") or "").strip()
        fish_materiales = (data.get("fish_materiales") or "").strip()
        fish_personas = (data.get("fish_personas") or "").strip()
        fish_entorno = (data.get("fish_entorno") or "").strip()
        fish_medicion = (data.get("fish_medicion") or "").strip()

        why1 = (data.get("why1") or "").strip()
        why2 = (data.get("why2") or "").strip()
        why3 = (data.get("why3") or "").strip()
        why4 = (data.get("why4") or "").strip()
        why5 = (data.get("why5") or "").strip()

        # =========================================================
        # Resumen legacy para compatibilidad
        # =========================================================
        def _join_legacy(items):
            return "\n".join(
                f"- {x['descripcion']} ({x['fecha_compromiso']})"
                for x in items
            )

        causa_txt = _join_legacy(causas_items)
        preventiva_txt = _join_legacy(control_items)
        correctiva_txt = _join_legacy(correctiva_items)

        fecha_causa = causas_items[0]["fecha_compromiso"] if causas_items else ""
        fecha_preventiva = control_items[0]["fecha_compromiso"] if control_items else ""
        fecha_correctiva = correctiva_items[0]["fecha_compromiso"] if correctiva_items else ""

        # 3) Verificar si ya existe respuesta activa del miembro
        row = db.execute("""
            SELECT id
            FROM reclamo_respuestas_equipo
            WHERE reclamo_id = ?
            AND imputacion_id = ?
            AND miembro_id = ?
            AND activo = 1
            LIMIT 1
        """, (reclamo_id, imputacion_id, uid)).fetchone()

        if row:
            respuesta_id = row["id"]

            db.execute("""
                UPDATE reclamo_respuestas_equipo
                SET
                    metodo_analisis = ?,
                    causa = ?,
                    preventiva = ?,
                    correctiva = ?,
                    fecha_causa = ?,
                    fecha_preventiva = ?,
                    fecha_correctiva = ?,
                    fish_metodo = ?,
                    fish_maquinas = ?,
                    fish_materiales = ?,
                    fish_personas = ?,
                    fish_entorno = ?,
                    fish_medicion = ?,
                    why1 = ?, why2 = ?, why3 = ?, why4 = ?, why5 = ?
                WHERE id = ?
            """, (
                metodo_analisis,
                causa_txt, preventiva_txt, correctiva_txt,
                fecha_causa, fecha_preventiva, fecha_correctiva,
                fish_metodo, fish_maquinas, fish_materiales,
                fish_personas, fish_entorno, fish_medicion,
                why1, why2, why3, why4, why5,
                respuesta_id
            ))
        else:
            cur = db.execute("""
                INSERT INTO reclamo_respuestas_equipo (
                    reclamo_id, imputacion_id, miembro_id,
                    metodo_analisis,
                    causa, preventiva, correctiva,
                    fecha_causa, fecha_preventiva, fecha_correctiva,
                    fish_metodo, fish_maquinas, fish_materiales,
                    fish_personas, fish_entorno, fish_medicion,
                    why1, why2, why3, why4, why5,
                    activo, created_by
                )
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                reclamo_id,
                imputacion_id,
                uid,
                metodo_analisis,
                causa_txt, preventiva_txt, correctiva_txt,
                fecha_causa, fecha_preventiva, fecha_correctiva,
                fish_metodo, fish_maquinas, fish_materiales,
                fish_personas, fish_entorno, fish_medicion,
                why1, why2, why3, why4, why5,
                1,
                uid
            ))
            respuesta_id = cur.lastrowid

        # =========================================================
        # Guardar acciones detalle
        # =========================================================
        _save_respuesta_equipo_acciones(
            db,
            respuesta_id=respuesta_id,
            reclamo_id=reclamo_id,
            imputacion_id=imputacion_id,
            miembro_id=uid,
            causas=causas_items,
            control=control_items,
            correctiva=correctiva_items,
            user_id=uid
        )

        db.commit()

        try:
            row_r = db.execute("""
                SELECT codigo
                FROM reclamos
                WHERE id = ?
                LIMIT 1
            """, (reclamo_id,)).fetchone()

            reclamo_codigo = row_r["codigo"] if row_r and row_r["codigo"] else f"RECL#{reclamo_id}"

            _notify_sponsor_respuesta_equipo(
                db,
                int(imputacion_id),
                int(uid),
                reclamo_codigo
            )
        except Exception:
            current_app.logger.exception(
                "[MAIL_EQUIPO] Error notificando sponsor por respuesta de equipo. reclamo_id=%s imputacion_id=%s miembro_id=%s",
                reclamo_id, imputacion_id, uid
            )

        return jsonify({"ok": True})


    # ----------------------------------------------
    # EQUIPO: RESPONSABLE APRUEBA / RECHAZA APORTE
    # ----------------------------------------------
    @app.route(
        '/reclamos/equipo/<int:eq_id>/aprobar',
        methods=['POST'],
        endpoint='reclamos_equipo_aprobar'
    )
    @require_login
    @require_permission('reclamos', 'aprobar')
    def reclamos_equipo_aprobar(eq_id):
        uid = _current_user_id()
        data = request.get_json(silent=True) or {}
        accion = (data.get("accion") or "").strip().lower()  # aprobar / rechazar
        motivo = (data.get("motivo") or "").strip()

        conn = get_db()
   
        cur = conn.cursor()

        cur.execute("""
            SELECT re.*, r.codigo,
                   u_col.username AS colaborador_username
            FROM reclamo_equipo re
            JOIN reclamos r ON r.id = re.reclamo_id
            LEFT JOIN usuarios u_col ON u_col.id = re.colaborador_id
            WHERE re.id = ?
        """, (eq_id,))
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="Registro de equipo no encontrado"), 404

        # Solo el responsable original o admin pueden aprobar/rechazar aportes
        if (row["responsable_id"] != uid) and (not _is_admin_like()):
            conn.close()
            return jsonify(ok=False, msg="No autorizado"), 403

        if row["estado"] not in ("respondido", "pendiente"):
            conn.close()
            return jsonify(ok=False, msg="No hay aporte pendiente de revisión"), 400

        now = _now_iso()

        if accion == "aprobar":
            cur.execute("""
                UPDATE reclamo_equipo
                SET estado = 'aprobado',
                    fecha_aprobacion = ?,
                    motivo_rechazo = NULL,
                    fecha_rechazo = NULL
                WHERE id = ?
            """, (now, eq_id))

        elif accion == "rechazar":
            if not motivo:
                conn.close()
                return jsonify(ok=False, msg="Motivo obligatorio al rechazar"), 400

            cur.execute("""
                UPDATE reclamo_equipo
                SET estado = 'rechazado',
                    fecha_rechazo = ?,
                    motivo_rechazo = ?,
                    fecha_aprobacion = NULL
                WHERE id = ?
            """, (now, motivo, eq_id))

            _notify_colaborador_aporte_rechazado(
                conn,
                row["colaborador_id"],
                row["codigo"],
                motivo
            )
        else:
            conn.close()
            return jsonify(ok=False, msg="Acción inválida"), 400

        conn.commit()
        conn.close()
        return jsonify(ok=True)


    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/<int:er_id>/eliminar", methods=["POST"])
    @require_login
    def reclamo_equipo_respuestas_eliminar(reclamo_id, er_id):
        uid = _current_user_id()
        if not (_puede_gestionar_equipo(reclamo_id, uid) or _is_admin_like()):
            return jsonify({"ok": False, "error": "No tienes permiso para modificar el equipo de respuestas."}), 403

        db = get_db()
        ensure_reclamos_schema(db)

        cur = db.execute("""
            UPDATE reclamo_equipo_respuestas
            SET activo = 0
            WHERE id = ? AND reclamo_id = ?
        """, (er_id, reclamo_id))
        db.commit()

        if cur.rowcount == 0:
            return jsonify({"ok": False, "error": "No se encontró el registro a eliminar (id incorrecto)."}), 404

        return jsonify({"ok": True})

   
    @app.route('/reclamos/api/usuario_por_cedula/<cedula>', methods=['GET'])
    @require_login
    def reclamos_usuario_por_cedula(cedula):
        print("[DEBUG reclamos] cedula", cedula)

        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
            SELECT u.id,
                u.nombre_completo,
                u.username,
                d.nombre AS departamento_nombre,
                j.nombre_completo AS jefe_nombre
            FROM usuarios u
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            LEFT JOIN usuarios j ON j.id = u.jefe_id
            WHERE u.identificacion = ?
            
            LIMIT 1;
        """, (cedula,))

        row = cur.fetchone()

        if not row:
            return jsonify(ok=False, msg="No existe usuario con esa cédula"), 404
 

        print(row["nombre_completo"])
        return jsonify(ok=True, usuario={
            "id": row["id"],
            "nombre": row["nombre_completo"],
            "username": row["username"],
            "departamento": row["departamento_nombre"],
            "jefe": row["jefe_nombre"]
        })






    # ----------------------------------------------
    # EQUIPO DEL RESPONSABLE: ASIGNAR COLABORADORES
    # ----------------------------------------------
    @app.route(
        '/reclamos/<int:reclamo_id>/equipo/asignar',
        methods=['POST'],
        endpoint='reclamos_equipo_asignar'
    )
    @require_login
    @require_permission('reclamos', 'editar')
    def reclamos_equipo_asignar(reclamo_id):
        uid = _current_user_id()  # responsable actual (imputado principal)
        data = request.get_json(silent=True) or {}
        ids = data.get("colaboradores") or []  # lista de IDs de usuarios

        if not isinstance(ids, list) or not ids:
            return jsonify(ok=False, msg="Debe indicar al menos un colaborador"), 400

        conn = get_db()
        ensure_reclamos_schema(conn)
        
        cur = conn.cursor()

        # Validar que el usuario actual sea el imputado principal de la OM
        cur.execute("""
            SELECT ri.imputado_id, r.codigo
            FROM reclamos r
            LEFT JOIN reclamo_imputados ri ON ri.reclamo_id = r.id
            WHERE r.id = ?
            LIMIT 1
        """, (reclamo_id,))
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="Oportunidad de Mejora no encontrada"), 404

        imputado_principal = row["imputado_id"]
        codigo = row["codigo"]

        # Solo el imputado/responsable principal o admin puede asignar equipo
        if (imputado_principal != uid) and (not _is_admin_like()):
            conn.close()
            return jsonify(ok=False, msg="No autorizado para asignar equipo"), 403

        # Evitar duplicados
        responsable_id = imputado_principal or uid
        now = _now_iso()

        # Para obtener username del responsable (para el correo)
        resp_basic = _get_user_basic(conn, responsable_id)
        resp_username = resp_basic["username"] if resp_basic else f"UID {responsable_id}"

        creados = 0
        for raw_id in ids:
            try:
                colaborador_id = int(raw_id)
            except Exception:
                continue

            if colaborador_id == responsable_id:
                continue  # no tiene sentido asignarse a sí mismo como colaborador

            # ¿ya existe?
            cur.execute("""
                SELECT 1
                FROM reclamo_equipo
                WHERE reclamo_id = ? AND colaborador_id = ?
            """, (reclamo_id, colaborador_id))
            if cur.fetchone():
                continue

            cur.execute("""
                INSERT INTO reclamo_equipo(
                    reclamo_id, responsable_id, colaborador_id,
                    estado, fecha_asignacion
                ) VALUES (?,?,?,?,?)
            """, (
                reclamo_id,
                responsable_id,
                colaborador_id,
                'pendiente',
                now
            ))
            creados += 1

            # Notificación al colaborador
            _notify_colaborador_asignado(conn, colaborador_id, codigo, resp_username)

        conn.commit()
        conn.close()

        if not creados:
            return jsonify(ok=False, msg="No se creó ningún registro nuevo (todos ya estaban asignados)"), 400

        return jsonify(ok=True, msg=f"{creados} colaborador(es) asignado(s).")

    # ----------------------------------------------
    # LISTA / BANDEJA
    # ----------------------------------------------
  
    @app.get("/reclamos/<int:reclamo_id>/detalle-json")
    @require_login
    def reclamo_detalle_json(reclamo_id):
        db = get_db()

        reclamo = db.execute("""
            SELECT
                r.id,
                r.codigo,
                r.fecha_creacion,
                r.cliente_nombre,
                r.proceso_text,
                r.material_desc,
                r.observacion,
                r.procede,
                r.estado_global
            FROM reclamos r
            WHERE r.id = ?
        """, (reclamo_id,)).fetchone()

        if not reclamo:
            return jsonify(ok=False), 404

        imputados = db.execute("""
            SELECT
                u.username,
                ri.respuesta_causa,
                ri.respuesta_preventiva,
                ri.respuesta_correctiva
            FROM reclamo_imputados ri
            JOIN usuarios u ON u.id = ri.imputado_id
            WHERE ri.reclamo_id = ?
        """, (reclamo_id,)).fetchall()

        return jsonify(
            ok=True,
            reclamo=dict(reclamo),
            imputados=[dict(i) for i in imputados]
        )


    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/usuarios", methods=["GET"])
    @require_login
    def equipo_respuestas_usuarios(reclamo_id):
        uid = _current_user_id()
        if not (_puede_gestionar_equipo(reclamo_id, uid) or _is_admin_like()):
            return jsonify({"ok": False, "error": "No autorizado"}), 403

        db = get_db()
        rows = db.execute("""
            SELECT
                u.id,
                COALESCE(u.nombre_completo, u.username) AS nombre,
                u.username,
                u.rol,
                d.nombre AS departamento
            FROM usuarios u
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            WHERE COALESCE(u.disabled,0)=0
            AND u.id NOT IN (
                SELECT usuario_id
                FROM reclamo_equipo_respuestas
                WHERE reclamo_id = ? AND activo = 1
            )
            ORDER BY nombre
        """, (reclamo_id,)).fetchall()

        return jsonify({"ok": True, "items": [dict(r) for r in rows]})

    # ----------------------------------------------
    # EQUIPO: COLABORADOR REGISTRA SU APORTE
    # ----------------------------------------------
    @app.route(
        '/reclamos/equipo/<int:eq_id>/responder',
        methods=['POST'],
        endpoint='reclamos_equipo_responder'
    )
    @require_login
    @require_permission('reclamos', 'editar')
    def reclamos_equipo_responder(eq_id):
        uid = _current_user_id()
        data = request.get_json(silent=True) or {}

        causa = (data.get("causa") or "").strip()
        preventiva = (data.get("preventiva") or "").strip()
        correctiva = (data.get("correctiva") or "").strip()

        if not causa or not preventiva or not correctiva:
            return jsonify(ok=False, msg="Todos los campos del aporte son obligatorios"), 400

        conn = get_db()
        ensure_reclamos_schema(conn)
        
        cur = conn.cursor()

        cur.execute("""
            SELECT re.*, r.codigo,
                   u_resp.username AS responsable_username,
                   u_col.username  AS colaborador_username
            FROM reclamo_equipo re
            JOIN reclamos r       ON r.id = re.reclamo_id
            LEFT JOIN usuarios u_resp ON u_resp.id = re.responsable_id
            LEFT JOIN usuarios u_col  ON u_col.id  = re.colaborador_id
            WHERE re.id = ?
        """, (eq_id,))
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="Registro de equipo no encontrado"), 404

        if (row["colaborador_id"] != uid) and (not _is_admin_like()):
            conn.close()
            return jsonify(ok=False, msg="No autorizado"), 403

        # Solo permitimos responder si está pendiente o ya respondió (para actualizar)
        if row["estado"] not in ("pendiente", "respondido"):
            conn.close()
            return jsonify(ok=False, msg="No se puede modificar este aporte en el estado actual"), 400

        cur.execute("""
            UPDATE reclamo_equipo
            SET respuesta_causa      = ?,
                respuesta_preventiva = ?,
                respuesta_correctiva = ?,
                fecha_respuesta      = ?,
                estado               = 'respondido'
            WHERE id = ?
        """, (
            causa,
            preventiva,
            correctiva,
            _now_iso(),
            eq_id
        ))

        # Notificar al responsable que ya tiene un aporte listo
        _notify_responsable_aporte_listo(
            conn,
            row["responsable_id"],
            row["codigo"],
            row["colaborador_username"] or f"UID {row['colaborador_id']}"
        )

        conn.commit()
        conn.close()
        return jsonify(ok=True)



    # ----------------------------------------------
    # API: respuestas técnicas por reclamo (todas las imputaciones)
    # ----------------------------------------------
    @app.route('/reclamos/api/<int:reclamo_id>/respuestas', methods=['GET'], endpoint='reclamos_api_respuestas')
    @require_login
    def reclamos_api_respuestas(reclamo_id):
        conn = get_db()
        ensure_reclamos_schema(conn)
        cur = conn.cursor()

        cur.execute("""
        SELECT
            ri.id AS imputacion_id,
            COALESCE(u.nombre_completo, u.username) AS imputado_nombre,
            u.username AS imputado_username,

            CASE
                WHEN ri.estado_asignacion = 'pend_aprobacion' THEN 'Pendiente aceptación del responsable'
                WHEN ri.estado_asignacion = 'rechazado' THEN 'Imputación rechazada'
                WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'sin_respuesta' THEN 'Pendiente respuesta del imputado'
                WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'pendiente_jefe' THEN 'Respuesta pendiente de aprobación'
                WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'aprobada' THEN 'Cerrado'
                WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'rechazada' THEN 'Respuesta rechazada'
                ELSE COALESCE(ri.estado_asignacion,'') || '/' || COALESCE(ri.estado_respuesta,'')
            END AS estado_imputacion,

            COALESCE(ri.respuesta_causa, '')      AS causa,
            COALESCE(ri.respuesta_preventiva, '') AS preventiva,
            COALESCE(ri.respuesta_correctiva, '') AS correctiva,

            -- ✅ FECHAS (las que ya agregaste)
            COALESCE(ri.fecha_causa,'')           AS fecha_causa,
            COALESCE(ri.fecha_preventiva,'')      AS fecha_preventiva,
            COALESCE(ri.fecha_correctiva,'')      AS fecha_correctiva,

            -- ✅ PARA EL OJO (diagrama)
            COALESCE(ri.metodo_analisis,'')       AS metodo_analisis,
            COALESCE(ri.why1,'')                  AS why1,
            COALESCE(ri.why2,'')                  AS why2,
            COALESCE(ri.why3,'')                  AS why3,
            COALESCE(ri.why4,'')                  AS why4,
            COALESCE(ri.why5,'')                  AS why5,
            COALESCE(ri.fish_metodo,'')           AS fish_metodo,
            COALESCE(ri.fish_maquinas,'')         AS fish_maquinas,
            COALESCE(ri.fish_materiales,'')       AS fish_materiales,
            COALESCE(ri.fish_personas,'')         AS fish_personas,
            COALESCE(ri.fish_entorno,'')          AS fish_entorno,
            COALESCE(ri.fish_medicion,'')         AS fish_medicion

        FROM reclamo_imputados ri
        LEFT JOIN usuarios u ON u.id = ri.imputado_id
        WHERE ri.reclamo_id = ?
        ORDER BY imputado_nombre, imputado_username
    """, (reclamo_id,))


        items = [dict(r) for r in cur.fetchall()]
        conn.close()
        return jsonify({"items": items})

    @app.route('/reclamos/api/<int:reclamo_id>/respuestas-detalle', methods=['GET'])
    @require_login
    def reclamos_api_respuestas_detalle(reclamo_id):
        db = get_db()
        ensure_reclamos_schema(db)
        ensure_reclamo_respuestas_equipo_schema(db)
        ensure_reclamo_respuesta_equipo_acciones_schema(db)
        ensure_reclamo_imputado_acciones_schema(db)

        cur = db.cursor()
        items = []

        # =========================
        # RESPUESTAS OFICIALES (sponsor / imputado)
        # =========================
        cur.execute("""
            SELECT
                'imputado' AS origen,
                ri.id AS imputacion_id,
                NULL AS miembro_id,
                COALESCE(u.nombre_completo, u.username) AS nombre,
                u.username AS username,

                CASE
                    WHEN ri.estado_asignacion = 'pend_aprobacion' THEN 'Pendiente aceptación del responsable'
                    WHEN ri.estado_asignacion = 'rechazado' THEN 'Imputación rechazada'
                    WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'sin_respuesta' THEN 'Pendiente respuesta del imputado'
                    WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'pendiente_jefe' THEN 'Respuesta pendiente de aprobación'
                    WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'aprobada' THEN 'Cerrado'
                    WHEN ri.estado_asignacion = 'aprobado' AND ri.estado_respuesta = 'rechazada' THEN 'Respuesta rechazada'
                    ELSE COALESCE(ri.estado_asignacion,'') || '/' || COALESCE(ri.estado_respuesta,'')
                END AS estado,

                COALESCE(ri.metodo_analisis,'') AS metodo_analisis,
                COALESCE(ri.why1,'') AS why1,
                COALESCE(ri.why2,'') AS why2,
                COALESCE(ri.why3,'') AS why3,
                COALESCE(ri.why4,'') AS why4,
                COALESCE(ri.why5,'') AS why5,
                COALESCE(ri.fish_metodo,'') AS fish_metodo,
                COALESCE(ri.fish_maquinas,'') AS fish_maquinas,
                COALESCE(ri.fish_materiales,'') AS fish_materiales,
                COALESCE(ri.fish_personas,'') AS fish_personas,
                COALESCE(ri.fish_entorno,'') AS fish_entorno,
                COALESCE(ri.fish_medicion,'') AS fish_medicion,

                COALESCE(ri.respuesta_causa, '') AS respuesta_causa,
                COALESCE(ri.respuesta_preventiva, '') AS respuesta_preventiva,
                COALESCE(ri.respuesta_correctiva, '') AS respuesta_correctiva,
                COALESCE(ri.fecha_causa, '') AS fecha_causa,
                COALESCE(ri.fecha_preventiva, '') AS fecha_preventiva,
                COALESCE(ri.fecha_correctiva, '') AS fecha_correctiva
            FROM reclamo_imputados ri
            LEFT JOIN usuarios u ON u.id = ri.imputado_id
            WHERE ri.reclamo_id = ?
            ORDER BY nombre, username
        """, (reclamo_id,))
        imputados = cur.fetchall()

        for r in imputados:
            it = dict(r)

            acciones = _get_imputado_acciones_full(db, int(r["imputacion_id"]))
            it["causas"] = acciones.get("causas", [])
            it["control"] = acciones.get("control", [])
            it["correctiva_items"] = acciones.get("correctiva", [])

            # fallback legacy si aún no existen acciones hijas
            if not it["causas"] and (it.get("respuesta_causa") or it.get("fecha_causa")):
                it["causas"] = [{
                    "id": None,
                    "tipo": "CAUSA",
                    "descripcion": it.get("respuesta_causa") or "",
                    "fecha_compromiso": it.get("fecha_causa") or "",
                    "orden": 1,
                    "requiere_evidencia": 0,
                    "cumplido": 0,
                    "fecha_cumplimiento": "",
                    "observacion_cumplimiento": "",
                    "evidencias": []
                }]

            if not it["control"] and (it.get("respuesta_preventiva") or it.get("fecha_preventiva")):
                it["control"] = [{
                    "id": None,
                    "tipo": "CONTROL",
                    "descripcion": it.get("respuesta_preventiva") or "",
                    "fecha_compromiso": it.get("fecha_preventiva") or "",
                    "orden": 1,
                    "requiere_evidencia": 0,
                    "cumplido": 0,
                    "fecha_cumplimiento": "",
                    "observacion_cumplimiento": "",
                    "evidencias": []
                }]

            if not it["correctiva_items"] and (it.get("respuesta_correctiva") or it.get("fecha_correctiva")):
                it["correctiva_items"] = [{
                    "id": None,
                    "tipo": "CORRECTIVA",
                    "descripcion": it.get("respuesta_correctiva") or "",
                    "fecha_compromiso": it.get("fecha_correctiva") or "",
                    "orden": 1,
                    "requiere_evidencia": 1,
                    "cumplido": 0,
                    "fecha_cumplimiento": "",
                    "observacion_cumplimiento": "",
                    "evidencias": []
                }]

            items.append(it)

        # =========================
        # RESPUESTAS DEL EQUIPO
        # =========================
        cur.execute("""
            SELECT
                'equipo' AS origen,
                rre.id AS respuesta_equipo_id,
                rre.imputacion_id,
                rre.miembro_id,
                COALESCE(u.nombre_completo, u.username) AS nombre,
                u.username AS username,
                'Aporte de equipo' AS estado,

                COALESCE(rre.metodo_analisis,'') AS metodo_analisis,
                COALESCE(rre.why1,'') AS why1,
                COALESCE(rre.why2,'') AS why2,
                COALESCE(rre.why3,'') AS why3,
                COALESCE(rre.why4,'') AS why4,
                COALESCE(rre.why5,'') AS why5,
                COALESCE(rre.fish_metodo,'') AS fish_metodo,
                COALESCE(rre.fish_maquinas,'') AS fish_maquinas,
                COALESCE(rre.fish_materiales,'') AS fish_materiales,
                COALESCE(rre.fish_personas,'') AS fish_personas,
                COALESCE(rre.fish_entorno,'') AS fish_entorno,
                COALESCE(rre.fish_medicion,'') AS fish_medicion,

                COALESCE(rre.causa, '') AS causa,
                COALESCE(rre.preventiva, '') AS preventiva,
                COALESCE(rre.correctiva, '') AS correctiva,
                COALESCE(rre.fecha_causa, '') AS fecha_causa,
                COALESCE(rre.fecha_preventiva, '') AS fecha_preventiva,
                COALESCE(rre.fecha_correctiva, '') AS fecha_correctiva
            FROM reclamo_respuestas_equipo rre
            LEFT JOIN usuarios u ON u.id = rre.miembro_id
            WHERE rre.reclamo_id = ?
            AND COALESCE(rre.activo,1) = 1
            ORDER BY nombre, username
        """, (reclamo_id,))
        equipo_rows = cur.fetchall()

        for r in equipo_rows:
            it = dict(r)

            acciones = _get_respuesta_equipo_acciones_full(db, int(r["respuesta_equipo_id"]))
            it["causas"] = acciones.get("causas", [])
            it["control"] = acciones.get("control", [])
            it["correctiva_items"] = acciones.get("correctiva", [])

            # fallback legacy si aún no existen acciones hijas
            if not it["causas"] and (it.get("causa") or it.get("fecha_causa")):
                it["causas"] = [{
                    "id": None,
                    "tipo": "CAUSA",
                    "descripcion": it.get("causa") or "",
                    "fecha_compromiso": it.get("fecha_causa") or "",
                    "orden": 1,
                    "requiere_evidencia": 0,
                    "cumplido": 0,
                    "fecha_cumplimiento": "",
                    "observacion_cumplimiento": "",
                    "evidencias": []
                }]

            if not it["control"] and (it.get("preventiva") or it.get("fecha_preventiva")):
                it["control"] = [{
                    "id": None,
                    "tipo": "CONTROL",
                    "descripcion": it.get("preventiva") or "",
                    "fecha_compromiso": it.get("fecha_preventiva") or "",
                    "orden": 1,
                    "requiere_evidencia": 0,
                    "cumplido": 0,
                    "fecha_cumplimiento": "",
                    "observacion_cumplimiento": "",
                    "evidencias": []
                }]

            if not it["correctiva_items"] and (it.get("correctiva") or it.get("fecha_correctiva")):
                it["correctiva_items"] = [{
                    "id": None,
                    "tipo": "CORRECTIVA",
                    "descripcion": it.get("correctiva") or "",
                    "fecha_compromiso": it.get("fecha_correctiva") or "",
                    "orden": 1,
                    "requiere_evidencia": 1,
                    "cumplido": 0,
                    "fecha_cumplimiento": "",
                    "observacion_cumplimiento": "",
                    "evidencias": []
                }]

            items.append(it)

        return jsonify({
            "ok": True,
            "items": items
        })
   
    @app.route('/reclamos/equipo-acciones/<int:accion_id>/cumplir', methods=['POST'])
    @require_login
    def reclamo_equipo_accion_cumplir(accion_id):

        db = get_db()
        cur = db.cursor()

        data = request.json or {}
        cumplido = 1 if data.get("cumplido") else 0
        fecha = data.get("fecha_cumplimiento")

        cur.execute("""
            UPDATE reclamo_respuesta_equipo_acciones
            SET
                cumplido = ?,
                fecha_cumplimiento = ?,
                updated_at = CURRENT_TIMESTAMP,
                updated_by = ?
            WHERE id = ?
        """, (
            cumplido,
            fecha,
            session.get("usuario_id"),
            accion_id
        ))

        db.commit()

        return jsonify({
            "ok": True,
            "msg": "Cumplimiento actualizado"
        })
    
    @app.route('/reclamos/equipo-acciones/evidencias/<int:evidencia_id>/download', methods=['GET'])
    @require_login
    def reclamo_equipo_accion_evidencia_download(evidencia_id):
        db = get_db()
        row = db.execute("""
            SELECT
                e.id,
                e.accion_id,
                e.filename,
                e.original_name,
                COALESCE(e.activo, 1) AS activo
            FROM reclamo_respuesta_equipo_accion_evidencias e
            WHERE e.id = ?
            LIMIT 1
        """, (evidencia_id,)).fetchone()

        if not row or int(row["activo"] or 0) != 1:
            abort(404)

        folder = os.path.join(current_app.config["UPLOAD_FOLDER"], "om_evidencias")
        path = os.path.join(folder, row["filename"])

        if not os.path.isfile(path):
            abort(404)

        return send_file(
            path,
            as_attachment=True,
            download_name=row["original_name"] or row["filename"]
        )
    
    
    @app.route('/reclamos/equipo-acciones/<int:accion_id>/evidencia', methods=['POST'])
    @require_login
    def reclamo_equipo_accion_subir_evidencia(accion_id):

        db = get_db()
        cur = db.cursor()

        file = request.files.get("file")
        if not file:
            return jsonify({"ok": False, "error": "Archivo requerido"}), 400

        filename = secure_filename(file.filename)

        folder = os.path.join(current_app.config["UPLOAD_FOLDER"], "om_evidencias")
        os.makedirs(folder, exist_ok=True)

        path = os.path.join(folder, filename)
        file.save(path)

        cur.execute("""
            INSERT INTO reclamo_respuesta_equipo_accion_evidencias (
                accion_id,
                filename,
                original_name,
                content_type,
                size_bytes,
                creado_por,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """, (
            accion_id,
            filename,
            file.filename,
            file.content_type,
            os.path.getsize(path),
            session.get("usuario_id")
        ))

        db.commit()
 
        return jsonify({
            "ok": True,
            "msg": "Evidencia subida"
        })

        
    @app.route('/reclamos/equipo-acciones/<int:respuesta_id>/seguimiento', methods=['GET'])
    @require_login
    def reclamo_equipo_acciones_seguimiento(respuesta_id):

        db = get_db()
        cur = db.cursor()

        cur.execute("""
            SELECT
                a.id,
                a.tipo,
                a.descripcion,
                a.fecha_compromiso,
                a.cumplido,
                a.fecha_cumplimiento,
                a.requiere_evidencia,

                COUNT(e.id) AS evidencias
            FROM reclamo_respuesta_equipo_acciones a
            LEFT JOIN reclamo_respuesta_equipo_accion_evidencias e
                ON e.accion_id = a.id
                AND e.activo = 1
            WHERE a.respuesta_equipo_id = ?
            AND a.activo = 1
            GROUP BY a.id
            ORDER BY a.orden
        """, (respuesta_id,))

        rows = [dict(r) for r in cur.fetchall()]

        return jsonify({
            "ok": True,
            "items": rows
        })
    
    
    # ----------------------------------------------
    # LISTA / BANDEJA
    # ----------------------------------------------
    @app.route('/reclamos', methods=['GET'], endpoint='reclamos')
    @require_login
    @require_permission('reclamos', 'ver')
    def reclamos_lista():
        uid = _current_user_id()

        conn = get_db()
        ensure_reclamos_schema(conn)
        ensure_reclamos_catalogos(conn)
        ensure_geo_schema(conn)
        # NUEVO: asegurar columnas extra (metodo_analisis, why1..why5, fish_*)
        _ensure_reclamo_imputados_extra_cols(conn)

        productos = fetch_productos(conn)

        cur = conn.cursor()
        puede_crear_materiales = False

        if _is_admin_like():
            puede_crear_materiales = True
        elif uid:
            cur.execute("""
                SELECT p.nombre
                FROM usuarios u
                LEFT JOIN puestos p ON p.id = u.puesto_id
                WHERE u.id = ?
            """, (uid,))
            row_p = cur.fetchone()
            if row_p and row_p["nombre"]:
                nombre_puesto = row_p["nombre"].strip().upper()
                if nombre_puesto == "CORRDINADOR(A) SERVICIO AL CLIENTE":
                    puede_crear_materiales = True

        q_estado = (request.args.get('estado') or '').strip().lower()

        # =========================
        # Filtros (GET)
        # =========================
        f_codigo   = (request.args.get('codigo') or '').strip()
        f_cliente  = (request.args.get('cliente') or '').strip()
        f_proceso  = (request.args.get('proceso') or '').strip()      # viene de RECL_PROCESO (valor)
        f_motivo   = (request.args.get('motivo') or '').strip()       # tipo_reclamo (nombre param)
        f_tramite  = (request.args.get('tramite') or '').strip()      # tipo_tramite (nombre param)
        f_desde    = (request.args.get('desde') or '').strip()        # YYYY-MM-DD
        f_hasta    = (request.args.get('hasta') or '').strip()        # YYYY-MM-DD
        f_estado_global = (request.args.get('estado_global') or '').strip()

        filtros = {
            "estado": q_estado,
            "codigo": f_codigo,
            "cliente": f_cliente,
            "proceso": f_proceso,
            "motivo": f_motivo,
            "tramite": f_tramite,
            "desde": f_desde,
            "hasta": f_hasta,
            "estado_global": f_estado_global,

        }

        def _build_where_filtros(alias_r: str = "r"):
            """
            Devuelve (sql, params) para filtros comunes sobre reclamos.
            Nota: proceso_text suele venir como texto (a veces con comas) => usamos LIKE.
            """
            w = ""
            p = []

            if f_codigo:
                w += f" AND {alias_r}.codigo LIKE ?"
                p.append(f"%{f_codigo}%")

            if f_cliente:
                w += f" AND COALESCE({alias_r}.cliente_nombre,'') LIKE ?"
                p.append(f"%{f_cliente}%")

            if f_proceso:
                w += f" AND COALESCE({alias_r}.proceso_text,'') LIKE ?"
                p.append(f"%{f_proceso}%")

            if f_motivo:
                w += f" AND COALESCE({alias_r}.tipo_reclamo,'') = ?"
                p.append(f_motivo)

            if f_tramite:
                w += f" AND COALESCE({alias_r}.tipo_tramite,'') = ?"
                p.append(f_tramite)

            # Fechas: compara por día con substr(,1,10) (por si viene con hora)
            if f_desde:
                w += f" AND date(substr({alias_r}.fecha_reclamo,1,10)) >= date(?)"
                p.append(f_desde)

            if f_hasta:
                w += f" AND date(substr({alias_r}.fecha_reclamo,1,10)) <= date(?)"
                p.append(f_hasta)
            if f_estado_global:
                w += f" AND COALESCE({alias_r}.estado_global,'') = ?"
                p.append(f_estado_global)

            return w, p

        # 1) Creados por mí (admin => todos)
        created_list = []
        if uid:
            params = []
            can_view_all = _can_view_all_reclamos(conn, uid)

            if can_view_all:
                where = "1=1"                  # ✅ admin / servicio al cliente / gerente / gf ven todo
            else:
                where = "r.creado_por = ?"     # usuario normal: solo lo suyo
                params = [uid]

            if q_estado:
                where += " AND (ri.estado_asignacion = ? OR ri.estado_respuesta = ?)"
                params += [q_estado, q_estado]

            # ✅ filtros comunes
            f_sql, f_params = _build_where_filtros("r")
            where += f_sql
            params += f_params

            cur.execute(f"""
                SELECT
                    r.id,
                    r.codigo,
                    r.fecha_reclamo,
                    r.fecha_creacion,
                    r.cliente_nombre,
                    r.observacion,
                    r.material_desc,
                    r.procede,
                    r.estado_global           AS estado,
                    r.proceso_text            AS proceso_nombre,
                    c.nombre                  AS ciudad,
                    r.antecedente AS submotivo,
                    tr.valor                  AS motivo,
                    tt.valor                  AS tramite,

                    GROUP_CONCAT(DISTINCT ui.username) AS imputados_resumen,

                    COALESCE(ri.respuesta_causa, '')        AS causa,
                    COALESCE(ri.respuesta_preventiva, '')   AS preventiva,
                    COALESCE(ri.respuesta_correctiva, '')   AS correctiva,
                    COALESCE(ri.fecha_causa,'')        AS fecha_causa,
                    COALESCE(ri.fecha_preventiva,'')   AS fecha_preventiva,
                    COALESCE(ri.fecha_correctiva,'')   AS fecha_correctiva,

                    COALESCE(ri.metodo_analisis,'')         AS metodo_analisis,
                    COALESCE(ri.why1,'')                    AS why1,
                    COALESCE(ri.why2,'')                    AS why2,
                    COALESCE(ri.why3,'')                    AS why3,
                    COALESCE(ri.why4,'')                    AS why4,
                    COALESCE(ri.why5,'')                    AS why5,
                    COALESCE(ri.fish_metodo,'')             AS fish_metodo,
                    COALESCE(ri.fish_maquinas,'')           AS fish_maquinas,
                    COALESCE(ri.fish_materiales,'')         AS fish_materiales,
                    COALESCE(ri.fish_personas,'')           AS fish_personas,
                    COALESCE(ri.fish_entorno,'')            AS fish_entorno,
                    COALESCE(ri.fish_medicion,'')           AS fish_medicion,
                    r.tipo_reclamo,
                    r.antecedente,
                    (
                    SELECT GROUP_CONCAT(DISTINCT COALESCE(u2.nombre_completo, u2.username))
                    FROM reclamo_equipo_respuestas eq2
                    JOIN usuarios u2 ON u2.id = eq2.usuario_id
                    WHERE eq2.reclamo_id = r.id
                        AND eq2.activo = 1
                    ) AS equipo_resumen
                FROM reclamos r
                LEFT JOIN reclamo_imputados ri ON ri.reclamo_id = r.id
                LEFT JOIN usuarios ui ON ui.id = ri.imputado_id
                LEFT JOIN cantones c ON c.id = r.canton_id
                LEFT JOIN param_groups gtr ON gtr.nombre = 'RECL_TIPO'
                LEFT JOIN param_values tr ON tr.group_id = gtr.id AND tr.nombre = r.tipo_reclamo
                LEFT JOIN param_groups gtt ON gtt.nombre = 'RECL_TRAMITE'
                LEFT JOIN param_values tt ON tt.group_id = gtt.id AND tt.nombre = r.tipo_tramite
                WHERE {where}
                GROUP BY r.id
                ORDER BY r.id DESC
            """, params)

            created_list = cur.fetchall()

        # 2) Reclamos donde soy aprobador (jefe)
        approve_list = []
        if uid:
            jefe_id = int(uid)

            print("[DEBUG reclamos] jefe_id =", jefe_id)

            cur.execute("""
                SELECT id, reclamo_id, imputado_id, aprobador_id,
                    estado_asignacion, estado_respuesta
                FROM reclamo_imputados
                WHERE aprobador_id = ?
            """, (jefe_id,))
            debug_rows = cur.fetchall()
            print("[DEBUG reclamos] filas en reclamo_imputados para este jefe:", len(debug_rows))
            for dr in debug_rows:
                print("[DEBUG reclamos]  fila:", dict(dr))

            # ✅ filtros comunes para approve
            f_sql, f_params = _build_where_filtros("r")

            sql_approve = f"""
                SELECT
                    r.id,
                    r.codigo,
                    r.fecha_reclamo,
                    r.fecha_creacion,
                    r.cliente_nombre,
                    r.proceso_text          AS proceso_text,
                    r.tipo_tramite          AS tipo_tramite,
                    r.material_desc         AS material_desc,
                    r.procede               AS procede,
                    r.observacion           AS observacion,
                    c.nombre                AS ciudad,

                    ri.id                   AS imputacion_id,
                    ri.estado_asignacion,
                    ri.estado_respuesta,

                    CASE
                        WHEN ri.estado_asignacion = 'pend_aprobacion'
                            THEN 'Pendiente aceptación del responsable'
                        WHEN ri.estado_asignacion = 'rechazado'
                            THEN 'Imputación rechazada'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'sin_respuesta'
                            THEN 'Pendiente respuesta del imputado'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'pendiente_jefe'
                            THEN 'Respuesta pendiente de aprobación'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'aprobada'
                            THEN 'Cerrado'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'rechazada'
                            THEN 'Respuesta rechazada'
                        ELSE COALESCE(ri.estado_asignacion,'') || '/' ||
                            COALESCE(ri.estado_respuesta,'')
                    END AS estado_imputacion,

                    ri.motivo_rechazo_asignacion AS rechazo_motivo,
                    ri.respuesta_causa           AS causa,
                    ri.respuesta_preventiva      AS preventiva,
                    ri.respuesta_correctiva      AS correctiva,
                    COALESCE(ri.fecha_causa,'')        AS fecha_causa,
                    COALESCE(ri.fecha_preventiva,'')   AS fecha_preventiva,
                    COALESCE(ri.fecha_correctiva,'')   AS fecha_correctiva,

                    -- NUEVO: método, 5 porqués y espina para que el jefe vea todo
                    ri.metodo_analisis           AS metodo_analisis,
                    ri.why1,
                    ri.why2,
                    ri.why3,
                    ri.why4,
                    ri.why5,
                    ri.fish_metodo,
                    ri.fish_maquinas,
                    ri.fish_materiales,
                    ri.fish_personas,
                    ri.fish_entorno,
                    ri.fish_medicion,

                    u.nombre_completo            AS imputado_nombre
                FROM reclamo_imputados ri
                JOIN reclamos r   ON r.id = ri.reclamo_id
                LEFT JOIN usuarios u ON u.id = ri.imputado_id
                LEFT JOIN cantones c ON c.id = r.canton_id
                WHERE ri.aprobador_id = ?
                AND (
                        TRIM(ri.estado_asignacion) = 'pend_aprobacion'
                        OR (
                            TRIM(ri.estado_asignacion) = 'aprobado'
                            AND TRIM(ri.estado_respuesta) = 'pendiente_jefe'
                        )
                    )
                {f_sql}
                ORDER BY r.id DESC, ri.id DESC
            """

            cur.execute(sql_approve, (jefe_id, *f_params))
            approve_list = cur.fetchall()
            print("[DEBUG reclamos] approve_list len =", len(approve_list))

        # 3) Donde soy imputado (admin => ver todo como sponsor)
        imputado_list = []
        if uid:
            params = []
            can_view_all = _can_view_all_reclamos_sn_sponsor(conn, uid)  # (si ya lo calculaste arriba, reutilízalo)

            if can_view_all:
                where = "ri.estado_asignacion = 'aprobado'"   # ✅ ven todas las imputaciones aprobadas
            else:
                where = "ri.imputado_id = ? AND ri.estado_asignacion = 'aprobado'"
                params = [uid]

            # ✅ filtros comunes
            f_sql, f_params = _build_where_filtros("r")
            where += f_sql
            params += f_params

            cur.execute(f"""
                SELECT
                    r.id,
                    r.codigo,
                    r.fecha_reclamo,
                    r.fecha_creacion,
                    r.cliente_nombre,
                    r.observacion,
                    r.proceso_text       AS proceso_text,
                    r.material_desc      AS material_desc,
                    r.procede            AS procede,
                    c.nombre             AS ciudad,

                    -- ✅ Motivo / Trámite / Submotivo (como en Mis OM)
                    tr.valor             AS motivo,
                    tt.valor             AS tramite,
                    r.antecedente        AS submotivo,
                    r.tipo_reclamo,
                    r.tipo_tramite,

                    ri.id                AS imputacion_id,

                    CASE
                        WHEN ri.estado_asignacion = 'pend_aprobacion'
                                THEN 'Pendiente aceptación del responsable'
                        WHEN ri.estado_asignacion = 'rechazado'
                            THEN 'Imputación rechazada'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'sin_respuesta'
                            THEN 'Pendiente respuesta del imputado'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'pendiente_jefe'
                            THEN 'Respuesta pendiente de aprobación'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'aprobada'
                            THEN 'Cerrado'
                        WHEN ri.estado_asignacion = 'aprobado'
                            AND ri.estado_respuesta = 'rechazada'
                            THEN 'Respuesta rechazada'
                        ELSE COALESCE(ri.estado_asignacion,'') || '/' ||
                            COALESCE(ri.estado_respuesta,'')
                    END AS estado_imputacion,

                    ri.respuesta_causa      AS causa,
                    ri.respuesta_preventiva AS preventiva,
                    ri.respuesta_correctiva AS correctiva,
                    COALESCE(ri.fecha_causa,'')        AS fecha_causa,
                    COALESCE(ri.fecha_preventiva,'')   AS fecha_preventiva,
                    COALESCE(ri.fecha_correctiva,'')   AS fecha_correctiva,

                    ri.metodo_analisis      AS metodo_analisis,
                    ri.why1, ri.why2, ri.why3, ri.why4, ri.why5,
                    ri.fish_metodo, ri.fish_maquinas, ri.fish_materiales,
                    ri.fish_personas, ri.fish_entorno, ri.fish_medicion,

                    u_creador.nombre_completo AS creador_nombre,
                    u_creador.username        AS creador_username,

                    -- ✅ opcional: ver quién es el imputado real cuando admin ve todo
                    u_imp.nombre_completo     AS imputado_nombre,
                    u_imp.username            AS imputado_username,
                    (
                        SELECT GROUP_CONCAT(DISTINCT COALESCE(u2.nombre_completo, u2.username))
                        FROM reclamo_equipo_respuestas eq2
                        JOIN usuarios u2 ON u2.id = eq2.usuario_id
                        WHERE eq2.reclamo_id = r.id
                            AND eq2.imputacion_id = ri.id
                            AND eq2.activo = 1
                        ) AS equipo_resumen

                FROM reclamo_imputados ri
                JOIN reclamos r ON r.id = ri.reclamo_id
                LEFT JOIN cantones c ON c.id = r.canton_id
                LEFT JOIN usuarios u_creador ON u_creador.id = r.creado_por
                LEFT JOIN usuarios u_imp     ON u_imp.id = ri.imputado_id

                -- ✅ joins para motivo/trámite (igual que el otro bloque)
                LEFT JOIN param_groups gtr ON gtr.nombre = 'RECL_TIPO'
                LEFT JOIN param_values tr ON tr.group_id = gtr.id AND tr.nombre = r.tipo_reclamo

                LEFT JOIN param_groups gtt ON gtt.nombre = 'RECL_TRAMITE'
                LEFT JOIN param_values tt ON tt.group_id = gtt.id AND tt.nombre = r.tipo_tramite

                WHERE {where}
                ORDER BY r.id DESC, ri.id DESC
            """, params)
            
            imputado_list = cur.fetchall()

        tipos_reclamo = _fetch_param_values(conn, "RECL_TIPO")
        tipos_tramite = _fetch_param_values(conn, "RECL_TRAMITE")
        procesos = _fetch_param_values(conn, "RECL_PROCESO")
        tipos_campos = _fetch_tipo_campos(conn)
        regiones = fetch_regiones(conn)
        materiales = _fetch_param_values(conn, "RECL_MATERIAL")
        usuarios_imputables = fetch_usuarios_imputables(conn)
 
        # 4) OM donde soy colaborador del responsable (equipo)
        equipo_list = []
        if uid:
            ensure_reclamo_respuestas_equipo_schema(conn)
            ensure_reclamo_respuesta_equipo_acciones_schema(conn)

            # ✅ filtros comunes
            f_sql, f_params = _build_where_filtros("r")

            cur.execute(f"""
                SELECT
                    eq.id            AS equipo_id,
                    eq.imputacion_id AS imputacion_id,
                    eq.reclamo_id    AS reclamo_id,
                    eq.usuario_id    AS usuario_id,
                    rre.miembro_id   AS miembro_id,

                    rre.fecha_causa,
                    rre.fecha_preventiva,
                    rre.fecha_correctiva,

                    tr.valor AS motivo,
                    r.antecedente AS submotivo,
                    tt.valor AS tramite,

                    r.codigo,
                    r.fecha_reclamo,
                    r.fecha_creacion,

                    COALESCE(r.cliente_nombre, cli.nombre) AS cliente_nombre,
                    r.observacion,
                    r.proceso_text AS proceso_text,
                    COALESCE(r.material_desc, pv.valor, pv.nombre) AS material_desc,

                    -- ✅ Resumen LEGACY (compatibilidad)
                    COALESCE(rre.causa, '')      AS causa,
                    COALESCE(rre.preventiva, '') AS preventiva,
                    COALESCE(rre.correctiva, '') AS correctiva,

                    -- ✅ NUEVO: detalle resumido desde tabla hija
                    COALESCE((
                        SELECT GROUP_CONCAT(a.descripcion || ' (' || a.fecha_compromiso || ')', ' | ')
                        FROM reclamo_respuesta_equipo_acciones a
                        WHERE a.respuesta_equipo_id = rre.id
                        AND a.tipo = 'CAUSA'
                        AND a.activo = 1
                        ORDER BY a.orden, a.id
                    ), '') AS causas_detalle,

                    COALESCE((
                        SELECT GROUP_CONCAT(a.descripcion || ' (' || a.fecha_compromiso || ')', ' | ')
                        FROM reclamo_respuesta_equipo_acciones a
                        WHERE a.respuesta_equipo_id = rre.id
                        AND a.tipo = 'CONTROL'
                        AND a.activo = 1
                        ORDER BY a.orden, a.id
                    ), '') AS control_detalle,

                    COALESCE((
                        SELECT GROUP_CONCAT(a.descripcion || ' (' || a.fecha_compromiso || ')', ' | ')
                        FROM reclamo_respuesta_equipo_acciones a
                        WHERE a.respuesta_equipo_id = rre.id
                        AND a.tipo = 'CORRECTIVA'
                        AND a.activo = 1
                        ORDER BY a.orden, a.id
                    ), '') AS correctiva_detalle,

                    COALESCE(rre.metodo_analisis, '') AS metodo_analisis,
                    COALESCE(rre.why1, '') AS why1,
                    COALESCE(rre.why2, '') AS why2,
                    COALESCE(rre.why3, '') AS why3,
                    COALESCE(rre.why4, '') AS why4,
                    COALESCE(rre.why5, '') AS why5,

                    COALESCE(rre.fish_metodo, '')     AS fish_metodo,
                    COALESCE(rre.fish_maquinas, '')   AS fish_maquinas,
                    COALESCE(rre.fish_materiales, '') AS fish_materiales,
                    COALESCE(rre.fish_personas, '')   AS fish_personas,
                    COALESCE(rre.fish_entorno, '')    AS fish_entorno,
                    COALESCE(rre.fish_medicion, '')   AS fish_medicion,

                    CASE
                        WHEN EXISTS (
                            SELECT 1
                            FROM reclamo_respuesta_equipo_acciones a
                            WHERE a.respuesta_equipo_id = rre.id
                            AND a.activo = 1
                        )
                        OR COALESCE(TRIM(rre.causa), '') <> ''
                        OR COALESCE(TRIM(rre.preventiva), '') <> ''
                        OR COALESCE(TRIM(rre.correctiva), '') <> ''
                        THEN ''
                        ELSE ''
                    END AS estado_equipo,

                    -- ✅ sponsor real = imputado de la imputación
                    COALESCE(us.nombre_completo, us.username, '') AS sponsor,
                    r.estado_global

                FROM reclamo_equipo_respuestas eq
                JOIN reclamos r
                ON r.id = eq.reclamo_id

                -- ✅ sponsor real de esa imputación
                LEFT JOIN reclamo_imputados ri
                ON ri.id = eq.imputacion_id

                LEFT JOIN usuarios us
                ON us.id = ri.imputado_id

                LEFT JOIN param_groups gtr
                ON gtr.nombre = 'RECL_TIPO'
                LEFT JOIN param_values tr
                ON tr.group_id = gtr.id
                AND tr.nombre = r.tipo_reclamo

                LEFT JOIN param_groups gtt
                ON gtt.nombre = 'RECL_TRAMITE'
                LEFT JOIN param_values tt
                ON tt.group_id = gtt.id
                AND tt.nombre = r.tipo_tramite

                -- ✅ última respuesta activa del miembro para esa imputación
                LEFT JOIN reclamo_respuestas_equipo rre
                ON rre.id = (
                        SELECT MAX(rre2.id)
                        FROM reclamo_respuestas_equipo rre2
                        WHERE rre2.reclamo_id = eq.reclamo_id
                        AND rre2.imputacion_id = eq.imputacion_id
                        AND rre2.miembro_id = eq.usuario_id
                        AND rre2.activo = 1
                )

                LEFT JOIN terceros cli
                ON cli.id = r.cliente_id
                AND cli.tipo = 'C'

                LEFT JOIN param_values pv
                ON pv.id = r.material_id
                AND pv.group_id = (
                        SELECT id
                        FROM param_groups
                        WHERE nombre = 'RECL_MATERIAL'
                )

                WHERE eq.usuario_id = ?
                AND eq.activo = 1
                {f_sql}

                ORDER BY r.fecha_reclamo DESC, r.id DESC, eq.id DESC
            """, (uid, *f_params))

            equipo_list = cur.fetchall()

            if equipo_list:
                print("[DEBUG equipo_list primera fila]", dict(equipo_list[0]))

        from datetime import datetime, timedelta

        def _ensure_reclamos_deadline_schema(conn: sqlite3.Connection):
            cur = conn.cursor()
            cur.execute("PRAGMA table_info(reclamos)")
            cols = {r["name"] for r in cur.fetchall()}

            # para evitar notificar 100 veces
            if "gg_notificado" not in cols:
                cur.execute("ALTER TABLE reclamos ADD COLUMN gg_notificado INTEGER NOT NULL DEFAULT 0")
            if "gg_notificado_fecha" not in cols:
                cur.execute("ALTER TABLE reclamos ADD COLUMN gg_notificado_fecha TEXT")
            conn.commit()

        def _get_param_int_by_id(conn: sqlite3.Connection, pv_id: int, default: int) -> int:
            cur = conn.cursor()
            cur.execute("SELECT valor FROM param_values WHERE id = ? AND activo = 1 LIMIT 1", (pv_id,))
            row = cur.fetchone()
            if not row:
                return default
            try:
                return int(str(row["valor"]).strip())
            except Exception:
                return default

        def _parse_fecha(fecha_txt: str | None) -> datetime | None:
            if not fecha_txt:
                return None
            s = str(fecha_txt).strip()
            # tus fechas vienen tipo "2026-01-29 15:05:19" (por lo que ya muestras)
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    return datetime.strptime(s[:19], fmt)
                except Exception:
                    pass
            return None

        def _enrich_deadline(rows, dias_plazo: int, only_pending_check: bool = False):
            """
            Agrega:
            - dias_plazo
            - fecha_limite
            - dias_transcurridos
            - dias_restantes
            - pct_avance (0-100)
            - deadline_estado: ok|warn|danger|done
            - alerta_gg: 1/0 (>=5 días y pendiente respuesta)
            """
            hoy = datetime.now()
            out = []

            for r in rows or []:
                d = dict(r)
                f = _parse_fecha(d.get("fecha_reclamo"))
                if not f:
                    d.update({
                        "dias_plazo": dias_plazo,
                        "fecha_limite": "",
                        "dias_transcurridos": 0,
                        "dias_restantes": dias_plazo,
                        "pct_avance": 0,
                        "deadline_estado": "ok",
                        "alerta_gg": 0,
                    })
                    out.append(d)
                    continue

                limite = f + timedelta(days=dias_plazo)
                trans = (hoy.date() - f.date()).days
                trans = max(0, trans)
                ventana_alertas = 10  # 5 = recordatorio, 10 = escalamiento (medido desde fecha_reclamo)
                pct10 = int(min(100, max(0, round((trans / ventana_alertas) * 100)))) if ventana_alertas > 0 else 100

                rest = (limite.date() - hoy.date()).days
                pct = int(min(100, max(0, round((trans / dias_plazo) * 100)))) if dias_plazo > 0 else 100

                estado_txt = (d.get("estado_imputacion") or d.get("estado") or "").lower()
                # si ya está cerrado, pinta como listo
                is_done = ("cerrad" in estado_txt) or ("aprobada" in estado_txt)

                dias_para_d5  = max(0, 5  - trans)
                dias_para_d10 = max(0, 10 - trans)
                if is_done:
                    deadline_estado = "done"
                    pct = 100
                    # no contar plazo
                    rest = None
                    # ✅ NO toques trans
                    trans = None

                elif rest < 0:
                    deadline_estado = "danger"
                    pct = 100
                elif trans >= 5:
                    deadline_estado = "warn"
                else:
                    deadline_estado = "ok"

                # alerta GG SOLO si sigue pendiente respuesta del imputado
                pend_resp = "pendiente respuesta del imputado" in (d.get("estado_imputacion") or "").lower()
                alerta_gg = 1 if (pend_resp and trans is not None and trans >= 5) else 0

                d.update({
                    "dias_plazo": dias_plazo,
                    "fecha_limite": limite.strftime("%Y-%m-%d"),
                    "dias_transcurridos": trans,
                    "dias_restantes": rest,
                    "pct_avance": pct,
                    "deadline_estado": deadline_estado,
                    "alerta_gg": alerta_gg,
                    "pct_avance_10": pct10,
                    "pct_hito_5": 50,    # 5/10
                    "pct_hito_10": 100,  # 10/10
                    "is_done": 1 if is_done else 0,
                    "dias_para_d5": dias_para_d5,
                    "dias_para_d10": dias_para_d10,
                })
                out.append(d)
            return out

        def _get_gerente_general_email(conn: sqlite3.Connection) -> str | None:
            cur = conn.cursor()
            cur.execute("""
                SELECT email
                FROM usuarios
                WHERE LOWER(TRIM(rol)) = 'gerente general'
                AND email IS NOT NULL AND TRIM(email) <> ''
                LIMIT 1
            """)
            row = cur.fetchone()
            return row["email"] if row else None

        def _notify_gg_if_needed(conn: sqlite3.Connection, dias_alerta: int):
            """
            Notifica al GG cuando:
            - existe al menos 1 imputación aprobada con estado_respuesta='sin_respuesta'
            - han pasado >= dias_alerta desde fecha_reclamo
            - y NO se ha notificado antes (reclamos.gg_notificado=0)
            """
            gg_email = _get_gerente_general_email(conn)
            if not gg_email:
                return

            cur = conn.cursor()

            # OMs candidatas
            cur.execute(f"""
                SELECT r.id, r.codigo, r.fecha_reclamo, COALESCE(r.cliente_nombre,'') AS cliente_nombre
                FROM reclamos r
                WHERE COALESCE(r.gg_notificado,0) = 0
                AND r.id IN (
                    SELECT ri.reclamo_id
                    FROM reclamo_imputados ri
                    WHERE ri.estado_asignacion = 'aprobado'
                    AND ri.estado_respuesta = 'sin_respuesta'
                )
                AND (
                    CAST(julianday('now') - julianday(substr(r.fecha_reclamo,1,10)) AS INTEGER) >= ?
                )
                ORDER BY r.id DESC
            """, (dias_alerta,))
            rows = cur.fetchall()
            if not rows:
                return

            # Arma correo
            items = "\n".join([f"- {x['codigo']} | {x['cliente_nombre']} | {x['fecha_reclamo']}" for x in rows])
            subject = "[OM] Alerta: Sponsor sin respuesta (día 5)"
            text_body = f"""Estimado/a,

                Se detectaron Oportunidades de Mejora sin respuesta del sponsor al día {dias_alerta}:

                {items}

                Por favor revisar en la plataforma (tab: Soy Sponsor / Responsable).

                Saludos.
                """

            _send_mail_safe(gg_email, subject, text_body, None)

            # Marca como notificado
            ids = [x["id"] for x in rows]
            cur.execute(
                f"UPDATE reclamos SET gg_notificado=1, gg_notificado_fecha=? WHERE id IN ({','.join(['?']*len(ids))})",
                (_now_iso(), *ids)
            )
            conn.commit()
            _ensure_reclamos_deadline_schema(conn)

        # PARAM: id 11047 = DIAS_CONTESTACION (ej: 7)
        dias_plazo = _get_param_int_by_id(conn, 11047, default=5)

        # 1) convierte Row -> dict (IMPORTANTE)
        created_list  = [dict(x) for x in (created_list or [])]
        approve_list  = [dict(x) for x in (approve_list or [])]
        imputado_list = [dict(x) for x in (imputado_list or [])]
        equipo_list   = [dict(x) for x in (equipo_list or [])]

        # 2) calcula campos de plazo (dias_restantes, fecha_limite, etc.)
        created_list  = _enrich_deadline(created_list, dias_plazo)
        approve_list  = _enrich_deadline(approve_list, dias_plazo)
        imputado_list = _enrich_deadline(imputado_list, dias_plazo)
        equipo_list   = _enrich_deadline(equipo_list, dias_plazo)

        # Notificar GG si al día 5 sigue sin respuesta
        # _notify_gg_if_needed(conn, dias_alerta=5)
        print("[DEBUG reclamos] type(imputado_list[0]) =", type(imputado_list[0]) if imputado_list else None)

        conn.close()

        return render_template(
            'reclamos_lista.html',
            created_list=created_list,
            approve_list=approve_list,
            imputado_list=imputado_list,
            q_estado=q_estado,
            rol=session.get('rol'),
            usuario=session.get('usuario'),
            user_id=uid,
            permissions=session.get('permissions', {}),
            tipos_reclamo=tipos_reclamo,
            tipos_tramite=tipos_tramite,
            materiales=materiales,
            puede_crear_materiales=puede_crear_materiales,
            procesos=procesos,
            tipos_campos=tipos_campos,
            regiones=regiones,
            usuarios_imputables=usuarios_imputables,
            productos=productos,
            equipo_list=equipo_list,  # 👈 nuevo
            filtros=filtros,          # 👈 NUEVO
            active_page='reclamos'
        )
 

 # ----------------------------------------------
# EXPORTAR A EXCEL - RECLAMOS DONDE INTERVIENE EL USUARIO
# (admin/coordinador => exporta TODO)
# + Respeta filtros del front (codigo, cliente, proceso, motivo, estado_global, desde, hasta)
# + Incluye columnas del CREADOR de la OM (creado_por -> usuarios)
# ----------------------------------------------
    @app.route('/reclamos/export/mis', methods=['GET'], endpoint='reclamos_export_mis')
    @require_login
    @require_permission('reclamos', 'ver')
    def reclamos_export_mis():
        uid = _current_user_id()
        if not uid:
            flash("No se pudo determinar el usuario actual.", "danger")
            return redirect(url_for('reclamos'))

        conn = get_db()
        export_all = _can_export_all_reclamos(conn, uid)

        ensure_reclamos_schema(conn)
        ensure_reclamos_catalogos(conn)
        ensure_geo_schema(conn)
        cur = conn.cursor()

        # =========================
        # Filtros (GET) - mismos que en la lista
        # =========================
        f_codigo  = (request.args.get('codigo') or '').strip()
        f_cliente = (request.args.get('cliente') or '').strip()
        f_proceso = (request.args.get('proceso') or '').strip()
        f_motivo  = (request.args.get('motivo') or '').strip()
        f_estado_global = (request.args.get('estado_global') or '').strip()
        f_desde   = (request.args.get('desde') or '').strip()   # YYYY-MM-DD
        f_hasta   = (request.args.get('hasta') or '').strip()   # YYYY-MM-DD

        def _build_where_export(alias_r: str = "r"):
            w = ""
            p = []

            if f_codigo:
                w += f" AND {alias_r}.codigo LIKE ?"
                p.append(f"%{f_codigo}%")

            if f_cliente:
                w += f" AND COALESCE({alias_r}.cliente_nombre,'') LIKE ?"
                p.append(f"%{f_cliente}%")

            # proceso_text puede venir como texto con comas => LIKE
            if f_proceso:
                w += f" AND COALESCE({alias_r}.proceso_text,'') LIKE ?"
                p.append(f"%{f_proceso}%")

            if f_motivo:
                w += f" AND COALESCE({alias_r}.tipo_reclamo,'') = ?"
                p.append(f_motivo)

            if f_estado_global:
                w += f" AND COALESCE({alias_r}.estado_global,'') = ?"
                p.append(f_estado_global)

            # Fechas por día (por si fecha_reclamo viene con hora)
            if f_desde:
                w += f" AND date(substr({alias_r}.fecha_reclamo,1,10)) >= date(?)"
                p.append(f_desde)

            if f_hasta:
                w += f" AND date(substr({alias_r}.fecha_reclamo,1,10)) <= date(?)"
                p.append(f_hasta)

            return w, p

        f_sql, f_params = _build_where_export("r")

        # ----------------- Data -----------------
        if export_all:
            # ✅ ADMIN/COORDINADOR: TODO el sistema (pero filtrado)
            cur.execute(f"""
                SELECT
                    r.codigo                           AS codigo_om,
                    r.fecha_reclamo,
                    r.fecha_creacion,

                    -- ✅ Creador OM
                    ucr.username                       AS creador_username,
                    COALESCE(ucr.nombre_completo, ucr.username) AS creador_nombre,

                    tr.valor                           AS tipo_om,
                    tt.valor                           AS tipo_tramite,
                    r.proceso_text,
                    r.cliente_nombre,
                    r.cliente_identificacion,
                    r.cliente_contacto,
                    r.cliente_email,
                    r.cliente_telefono,
                    r.material_desc,
                    c.nombre                           AS ciudad,
                    r.observacion,
                    r.procede,
                    r.estado_global,

                    ri.id                              AS imputacion_id,
                    ui.username                        AS imputado_username,
                    COALESCE(ui.nombre_completo, ui.username) AS imputado_nombre,
                    uj.username                        AS jefe_username,
                    COALESCE(uj.nombre_completo, uj.username) AS jefe_nombre,

                    ri.estado_asignacion,
                    ri.fecha_aprobacion_asignacion,
                    ri.fecha_rechazo_asignacion,
                    ri.motivo_rechazo_asignacion,

                    ri.respuesta_causa,
                    ri.respuesta_preventiva,
                    ri.respuesta_correctiva,
                    COALESCE(ri.fecha_causa,'')        AS fecha_causa,
                    COALESCE(ri.fecha_preventiva,'')   AS fecha_preventiva,
                    COALESCE(ri.fecha_correctiva,'')   AS fecha_correctiva,
                    ri.fecha_respuesta_imputado,

                    ri.estado_respuesta,
                    ri.fecha_aprobacion_respuesta,
                    ri.fecha_rechazo_respuesta,
                    ri.motivo_rechazo_respuesta

                FROM reclamos r
                LEFT JOIN reclamo_imputados ri ON ri.reclamo_id = r.id
                LEFT JOIN usuarios ui ON ui.id = ri.imputado_id
                LEFT JOIN usuarios uj ON uj.id = ri.aprobador_id
                LEFT JOIN usuarios ucr ON ucr.id = r.creado_por
                LEFT JOIN cantones c ON c.id = r.canton_id

                LEFT JOIN param_groups gtr ON gtr.nombre = 'RECL_TIPO'
                LEFT JOIN param_values tr ON tr.group_id = gtr.id AND tr.nombre = r.tipo_reclamo

                LEFT JOIN param_groups gtt ON gtt.nombre = 'RECL_TRAMITE'
                LEFT JOIN param_values tt ON tt.group_id = gtt.id AND tt.nombre = r.tipo_tramite

                WHERE 1=1
                {f_sql}
                ORDER BY r.id DESC, ri.id DESC
            """, f_params)
        else:
            # ✅ Usuario normal: donde interviene (creador / imputado / aprobador) + filtros
            cur.execute(f"""
                WITH mis_reclamos AS (
                    SELECT DISTINCT r.id
                    FROM reclamos r
                    LEFT JOIN reclamo_imputados ri ON ri.reclamo_id = r.id
                    WHERE r.creado_por = ?
                    OR ri.imputado_id = ?
                    OR ri.aprobador_id = ?
                )
                SELECT
                    r.codigo                           AS codigo_om,
                    r.fecha_reclamo,
                    r.fecha_creacion,

                    -- ✅ Creador OM
                    ucr.username                       AS creador_username,
                    COALESCE(ucr.nombre_completo, ucr.username) AS creador_nombre,

                    tr.valor                           AS tipo_om,
                    tt.valor                           AS tipo_tramite,
                    r.proceso_text,
                    r.cliente_nombre,
                    r.cliente_identificacion,
                    r.cliente_contacto,
                    r.cliente_email,
                    r.cliente_telefono,
                    r.material_desc,
                    c.nombre                           AS ciudad,
                    r.observacion,
                    r.procede,
                    r.estado_global,

                    ri.id                              AS imputacion_id,
                    ui.username                        AS imputado_username,
                    COALESCE(ui.nombre_completo, ui.username) AS imputado_nombre,
                    uj.username                        AS jefe_username,
                    COALESCE(uj.nombre_completo, uj.username) AS jefe_nombre,

                    ri.estado_asignacion,
                    ri.fecha_aprobacion_asignacion,
                    ri.fecha_rechazo_asignacion,
                    ri.motivo_rechazo_asignacion,

                    ri.respuesta_causa,
                    ri.respuesta_preventiva,
                    ri.respuesta_correctiva,
                    COALESCE(ri.fecha_causa,'')        AS fecha_causa,
                    COALESCE(ri.fecha_preventiva,'')   AS fecha_preventiva,
                    COALESCE(ri.fecha_correctiva,'')   AS fecha_correctiva,
                    ri.fecha_respuesta_imputado,

                    ri.estado_respuesta,
                    ri.fecha_aprobacion_respuesta,
                    ri.fecha_rechazo_respuesta,
                    ri.motivo_rechazo_respuesta

                FROM reclamos r
                LEFT JOIN reclamo_imputados ri ON ri.reclamo_id = r.id
                LEFT JOIN usuarios ui ON ui.id = ri.imputado_id
                LEFT JOIN usuarios uj ON uj.id = ri.aprobador_id
                LEFT JOIN usuarios ucr ON ucr.id = r.creado_por
                LEFT JOIN cantones c ON c.id = r.canton_id

                LEFT JOIN param_groups gtr ON gtr.nombre = 'RECL_TIPO'
                LEFT JOIN param_values tr ON tr.group_id = gtr.id AND tr.nombre = r.tipo_reclamo

                LEFT JOIN param_groups gtt ON gtt.nombre = 'RECL_TRAMITE'
                LEFT JOIN param_values tt ON tt.group_id = gtt.id AND tt.nombre = r.tipo_tramite

                WHERE r.id IN (SELECT id FROM mis_reclamos)
                {f_sql}
                ORDER BY r.id DESC, ri.id DESC
            """, (uid, uid, uid, *f_params))

        rows = cur.fetchall()
        conn.close()

        # ----------------- Construir Excel -----------------
        wb = Workbook()
        ws = wb.active
        ws.title = "Reclamos (Todos)" if export_all else "Mis reclamos"

        headers = [
            "Código OM",
            "Fecha OM",
            "Fecha creación (sistema)",

            # ✅ NUEVO: creador OM
            "Creador OM (usuario)",
            "Creador OM (nombre)",

            "Tipo OM",
            "Tipo trámite",
            "Proceso",
            "Cliente",
            "Identificación cliente",
            "Contacto cliente",
            "Correo cliente",
            "Teléfono cliente",
            "Ciudad",
            "Material",
            "Observación",
            "Procede",
            "Estado global",

            "ID imputación",
            "Imputado (usuario)",
            "Imputado (nombre)",
            "Jefe/aprobador (usuario)",
            "Jefe/aprobador (nombre)",

            "Estado imputación",
            "Fecha aprobación imputación",
            "Fecha rechazo imputación",
            "Motivo rechazo imputación",

            "Respuesta - Causa",
            "Respuesta - Acción preventiva",
            "Respuesta - Acción correctiva",
            "Fecha respuesta imputado",

            "Estado respuesta",
            "Fecha aprobación respuesta",
            "Fecha rechazo respuesta",
            "Motivo rechazo respuesta",
        ]

        # Estilos de encabezado
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F2937")  # gris oscuro
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="D1D5DB")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Escribir encabezados
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # Filas de datos
        for r in rows:
            ws.append([
                r["codigo_om"],
                r["fecha_reclamo"],
                r["fecha_creacion"],

                # ✅ creador OM
                r["creador_username"],
                r["creador_nombre"],

                r["tipo_om"],
                r["tipo_tramite"],
                r["proceso_text"],
                r["cliente_nombre"],
                r["cliente_identificacion"],
                r["cliente_contacto"],
                r["cliente_email"],
                r["cliente_telefono"],
                r["ciudad"],
                r["material_desc"],
                r["observacion"],
                r["procede"],
                r["estado_global"],

                r["imputacion_id"],
                r["imputado_username"],
                r["imputado_nombre"],
                r["jefe_username"],
                r["jefe_nombre"],

                r["estado_asignacion"],
                r["fecha_aprobacion_asignacion"],
                r["fecha_rechazo_asignacion"],
                r["motivo_rechazo_asignacion"],

                r["respuesta_causa"],
                r["respuesta_preventiva"],
                r["respuesta_correctiva"],
                r["fecha_respuesta_imputado"],

                r["estado_respuesta"],
                r["fecha_aprobacion_respuesta"],
                r["fecha_rechazo_respuesta"],
                r["motivo_rechazo_respuesta"],
            ])

        # Bordes y ajuste de texto para todo
        max_row = ws.max_row
        max_col = ws.max_column
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Ajuste simple de ancho de columnas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(val))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

        ws.freeze_panes = "A2"

        # Enviar archivo
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "reclamos_todos.xlsx" if export_all else f"mis_reclamos_{uid}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    # ----------------------------------------------
    # APIs geo
    # ----------------------------------------------
    @app.route('/reclamos/api/provincias', methods=['GET'], endpoint='reclamos_api_provincias')
    @require_login
    def reclamos_api_provincias():
        region_id = request.args.get('region_id', type=int)

        conn = get_db()
        ensure_geo_schema(conn)
        rows = fetch_provincias(conn, region_id)
        conn.close()

        data = [dict(id=r['id'], nombre=r['nombre']) for r in rows]
        return jsonify(data)

    @app.route('/reclamos/api/materiales/nuevo', methods=['POST'], endpoint='reclamos_api_materiales_nuevo')
    @require_login
    @require_permission('reclamos', 'editar')
    def reclamos_api_materiales_nuevo():
        nombre = (request.form.get('nombre') or '').strip()
        if not nombre:
            return jsonify(ok=False, msg="El nombre del material es obligatorio"), 400

        uid = _current_user_id()
        conn = get_db()
        ensure_reclamos_schema(conn)
        ensure_reclamos_catalogos(conn)
        cur = conn.cursor()

        # Permisos:
        if not _is_admin_like():
            cur.execute("""
                SELECT p.nombre
                FROM usuarios u
                LEFT JOIN puestos p ON p.id = u.puesto_id
                WHERE u.id = ?
            """, (uid,))
            row_p = cur.fetchone()
            nombre_puesto = (row_p["nombre"] if row_p and row_p["nombre"] else "").strip().upper()
            if nombre_puesto != "CORRDINADOR(A) SERVICIO AL CLIENTE":
                conn.close()
                return jsonify(ok=False, msg="No está autorizado para crear materiales."), 403

        base = ''.join(
            ch for ch in nombre.upper()
            if (ch.isalnum() or ch in (" ", "_", "-"))
        ).replace(" ", "_")
        if not base:
            base = "MAT"

        codigo = base
        contador = 1
        while True:
            cur.execute("""
                SELECT 1
                FROM param_values pv
                JOIN param_groups pg ON pg.id = pv.group_id
                WHERE pg.nombre = 'RECL_MATERIAL'
                  AND pv.nombre = ?
            """, (codigo,))
            if not cur.fetchone():
                break
            contador += 1
            codigo = f"{base}_{contador}"

        cur.execute("""
            SELECT COALESCE(MAX(pv.orden), 0) + 1 AS next_ord
            FROM param_values pv
            JOIN param_groups pg ON pg.id = pv.group_id
            WHERE pg.nombre = 'RECL_MATERIAL'
        """)
        row = cur.fetchone()
        next_ord = row["next_ord"] if row else 1

        gid = _ensure_param_group(conn, "RECL_MATERIAL", "Materiales de reclamos")

        cur.execute("""
            INSERT INTO param_values (group_id, nombre, valor, activo, orden)
            VALUES (?, ?, ?, 1, ?)
        """, (gid, codigo, nombre, next_ord))
        mid = cur.lastrowid

        conn.commit()
        conn.close()

        return jsonify(ok=True, item={"id": mid, "nombre": nombre})

    @app.route('/reclamos/api/cantones', methods=['GET'], endpoint='reclamos_api_cantones')
    @require_login
    def reclamos_api_cantones():
        provincia_id = request.args.get('provincia_id', type=int)

        conn = get_db()
        ensure_geo_schema(conn)
        rows = fetch_cantones(conn, provincia_id)
        conn.close()

        data = [dict(id=r['id'], nombre=r['nombre']) for r in rows]
        return jsonify(data)

    # ----------------------------------------------
    # CREAR RECLAMO
    # ----------------------------------------------
    @app.route('/reclamos/nuevo', methods=['POST'], endpoint='reclamos_nuevo')
    @require_login
    @require_permission('reclamos', 'crear')
    def reclamos_nuevo():
        print("DEBUG NUEVO RECLAMO:", dict(request.form))   # 👈 agrega esta línea

        uid = _current_user_id()
        if not uid:
            flash("No se pudo determinar el usuario actual.", "danger")
            return redirect(url_for('reclamos'))

        conn = get_db()
        ensure_reclamos_schema(conn)
        cur = conn.cursor()

        fecha_reclamo = (request.form.get('fecha_reclamo') or '').strip()
        cliente_id = (request.form.get('cliente_id') or '').strip() or None
        cliente_nombre = (request.form.get('cliente_nombre') or '').strip()
        cliente_identificacion = (request.form.get('cliente_identificacion') or '').strip()
        cliente_direccion = (request.form.get('cliente_direccion') or '').strip()
        cliente_contacto = (request.form.get('cliente_contacto') or '').strip()
        cliente_email = (request.form.get('cliente_email') or '').strip()
        cliente_telefono = (request.form.get('cliente_telefono') or '').strip()

        region_id = (request.form.get('region_id') or '').strip() or None
        provincia_id = (request.form.get('provincia_id') or '').strip() or None
        canton_id = (request.form.get('canton_id') or '').strip() or None

        tipo_tramite = (request.form.get('tipo_tramite') or '').strip()
        tipo_reclamo = (request.form.get('tipo_reclamo') or '').strip()

        proceso_text = (request.form.get('proceso_text') or '').strip()
        antecedente = (request.form.get('antecedente') or '').strip()
        antecedente_raw = (request.form.get('antecedente') or '').strip()
        antecedente = antecedente_raw

        # ✅ Si viene un ID (del select), guardamos el texto (pv.valor)
        if antecedente_raw.isdigit():
            try:
                cur.execute("SELECT valor FROM param_values WHERE id = ?", (int(antecedente_raw),))
                row_sub = cur.fetchone()
                if row_sub and row_sub["valor"]:
                    antecedente = row_sub["valor"].strip()
            except Exception:
                pass
        fecha_pedido = (request.form.get('fecha_pedido') or '').strip()
        factura = (request.form.get('factura') or '').strip()
        guia_remision = (request.form.get('guia_remision') or '').strip()

        raw_material_id = (request.form.get('material_id') or '').strip()
        material_id = int(raw_material_id) if raw_material_id else None
        material_desc = ''

        if material_id:
            cur.execute("SELECT valor FROM param_values WHERE id = ?", (material_id,))
            row_mat = cur.fetchone()
            if row_mat:
                material_desc = row_mat["valor"] or ""

        persona_atendio = (request.form.get('persona_atendio') or '').strip()
        persona_atendio_cedula = (request.form.get('persona_atendio_cedula') or '').strip()
        fecha_ofrec_entrega = (request.form.get('fecha_ofrec_entrega') or '').strip()
        fecha_entrega = (request.form.get('fecha_entrega') or '').strip()

        observacion = (request.form.get('observacion') or '').strip()
        procede = (request.form.get('procede') or '').strip()

        codigo = _generate_codigo_reclamo(conn)
        archivos = request.files.getlist('adjuntos')

        cur.execute("""
            INSERT INTO reclamos(
                codigo, fecha_reclamo, fecha_creacion,
                cliente_id, cliente_nombre, cliente_identificacion,
                cliente_direccion, cliente_contacto, cliente_email, cliente_telefono,
                region_id, provincia_id, canton_id,
                tipo_tramite, tipo_reclamo,
                proceso_text, antecedente,
                fecha_pedido, factura, guia_remision,
                material_id, material_desc,
                persona_atendio, persona_atendio_cedula,
                fecha_ofrec_entrega, fecha_entrega,
                observacion,
                procede,
                creado_por,
                estado_global
            )
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            codigo,
            #fecha_reclamo,
            _now_iso(),
            _now_iso(),
            cliente_id, cliente_nombre, cliente_identificacion,
            cliente_direccion, cliente_contacto, cliente_email, cliente_telefono,
            region_id, provincia_id, canton_id,
            tipo_tramite, tipo_reclamo,
            proceso_text, antecedente,
            fecha_pedido, factura, guia_remision,
            material_id, material_desc,
            persona_atendio, persona_atendio_cedula,
            fecha_ofrec_entrega, fecha_entrega,
            observacion,
            procede,
            uid,
            'abierto'
        ))
        reclamo_id = cur.lastrowid

        imputados_ids = request.form.getlist('imputados[]')
        for raw_uid in imputados_ids:
            try:
                imputado_id = int(raw_uid)
            except Exception:
                continue

            # ✅ Ahora el aprobador ES el mismo imputado
            aprobador_id = imputado_id

            cur.execute("""
                INSERT INTO reclamo_imputados(
                    reclamo_id,
                    imputado_id,
                    aprobador_id,
                    estado_asignacion,
                    estado_respuesta
                )
                VALUES(?,?,?,?,?)
            """, (
                    reclamo_id,
                    imputado_id,
                    aprobador_id,
                    'aprobado',
                    'sin_respuesta'
            ))

            imp_user = _get_user_basic(conn, imputado_id)
            # Sigue notificando al aprobador (que ahora es el propio imputado)
            _notify_aprobador_imputacion(
                conn,
                aprobador_id,
                codigo,
                imp_user["username"] if imp_user else f"UID {imputado_id}"
            )
    # Guardar adjuntos (si los hay)
        error_adj = _save_adjuntos_for_reclamo(conn, reclamo_id, archivos, uid)
        if error_adj:
            # No bloqueamos la creación de la OM, solo avisamos
            flash(error_adj, "warning")

        conn.commit()
        conn.close()

        flash("Reclamo creado correctamente.", "success")
        return redirect(url_for('reclamos'))

    # ----------------------------------------------
    # APROBACIÓN / RECHAZO IMPUTACIÓN
    # ----------------------------------------------
    @app.route('/reclamos/imputacion/<int:imp_id>/aprobar', methods=['POST'], endpoint='reclamos_aprobar_imputacion')
    @require_login
    @require_permission('reclamos', 'aprobar')
    def reclamos_aprobar_imputacion(imp_id):
        uid = _current_user_id()
        data = request.get_json(silent=True) or {}
        accion = (data.get("accion") or "").strip().lower()
        motivo = (data.get("motivo") or "").strip()

        conn = get_db()
        ensure_reclamos_schema(conn)
        cur = conn.cursor()

        cur.execute("""
            SELECT ri.*, r.codigo, r.creado_por,
                   u_imp.username AS imputado_username,
                   u_imp.id AS imputado_uid
            FROM reclamo_imputados ri
            JOIN reclamos r ON r.id = ri.reclamo_id
            LEFT JOIN usuarios u_imp ON u_imp.id = ri.imputado_id
            WHERE ri.id = ?
        """, (imp_id,))
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="Imputación no encontrada"), 404

        if (row["aprobador_id"] != uid) and (not _is_admin_like()):
            conn.close()
            return jsonify(ok=False, msg="No autorizado"), 403

        if row["estado_asignacion"] in ("aprobado", "rechazado"):
            conn.close()
            return jsonify(ok=False, msg="La imputación ya fue gestionada"), 400

        if accion == "aprobar":
            cur.execute("""
                UPDATE reclamo_imputados
                SET estado_asignacion='aprobado',
                    fecha_aprobacion_asignacion=?,
                    motivo_rechazo_asignacion=NULL,
                    fecha_rechazo_asignacion=NULL
                WHERE id=?
            """, (_now_iso(), imp_id))

            _notify_imputado_aprobado(conn, row["imputado_uid"], row["codigo"])

        elif accion == "rechazar":
            if not motivo:
                conn.close()
                return jsonify(ok=False, msg="Motivo es obligatorio al rechazar"), 400

            cur.execute("""
                UPDATE reclamo_imputados
                SET estado_asignacion='rechazado',
                    fecha_rechazo_asignacion=?,
                    motivo_rechazo_asignacion=?
                WHERE id=?
            """, (_now_iso(), motivo, imp_id))

            _notify_creador_rechazo_asignacion(
                conn,
                row["creado_por"],
                row["codigo"],
                row["imputado_username"],
                motivo
            )
        else:
            conn.close()
            return jsonify(ok=False, msg="Acción inválida"), 400

        conn.commit()
        conn.close()
        return jsonify(ok=True)



    @app.route('/reclamos/<int:reclamo_id>/adjuntos', methods=['POST'], endpoint='reclamos_add_adjuntos')
    @require_login
    @require_permission('reclamos', 'editar')
    def reclamos_add_adjuntos(reclamo_id):
        uid = _current_user_id()
        archivos = request.files.getlist('adjuntos')

        if not archivos:
            return jsonify(ok=False, msg="No se cargó ningún archivo."), 400

        conn = get_db()
        ensure_reclamos_schema(conn)
        ensure_reclamo_adjuntos_schema(conn)
        cur = conn.cursor()

        # Validar que exista la OM
        cur.execute("SELECT id FROM reclamos WHERE id = ?", (reclamo_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify(ok=False, msg="Oportunidad de Mejora no encontrada."), 404

        # Guardar adjuntos
        error_adj = _save_adjuntos_for_reclamo(conn, reclamo_id, archivos, uid)
        if error_adj:
            conn.rollback()
            conn.close()
            return jsonify(ok=False, msg=error_adj), 400

        # Obtener nombres originales de los archivos que se están subiendo,
        # limitado a lo que efectivamente se pudo guardar (mismo orden)
        filenames = [f.filename for f in archivos if f and f.filename]

        conn.commit()

        # Notificar al creador de la OM
        try:
            _notify_reclamo_adjuntos_change(conn, reclamo_id, uid, "agregados", filenames)
        except Exception:
            # No romper por fallo de correo
            current_app.logger.exception("Error enviando correo de adjuntos agregados")

        conn.close()
        return jsonify(ok=True, msg="Archivos adjuntados correctamente.")


    @app.route('/reclamos/adjunto/<int:adj_id>/delete',
            methods=['POST'],
            endpoint='reclamos_delete_adjunto')
    @require_login
    @require_permission('reclamos', 'editar')
    def reclamos_delete_adjunto(adj_id):
        uid = _current_user_id()

        conn = get_db()
        ensure_reclamo_adjuntos_schema(conn)
        cur = conn.cursor()

        cur.execute("""
            SELECT a.id,
                a.reclamo_id,
                a.filename,
                a.original_name,
                r.codigo,
                r.creado_por
            FROM reclamo_adjuntos a
            JOIN reclamos r ON r.id = a.reclamo_id
            WHERE a.id = ?
        """, (adj_id,))
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="Adjunto no encontrado"), 404

        upload_folder = _get_reclamos_upload_folder()
        path = os.path.join(upload_folder, row["filename"])

        # Borra registro de BD
        cur.execute("DELETE FROM reclamo_adjuntos WHERE id = ?", (adj_id,))
        conn.commit()
        conn.close()

        # Intenta borrar archivo físico (si existe)
        try:
            if os.path.exists(path):
                os.remove(path)
        except Exception as e:
            current_app.logger.warning(
                "No se pudo eliminar archivo físico %s: %s", path, e
            )

        return jsonify(ok=True, msg="Adjunto eliminado correctamente.")

    @app.route('/reclamos/api/<int:reclamo_id>/adjuntos', methods=['GET'], endpoint='reclamos_api_adjuntos')
    @require_login
    def reclamos_api_adjuntos(reclamo_id):
        conn = get_db()
        ensure_reclamos_schema(conn)
        ensure_reclamo_adjuntos_schema(conn)
        cur = conn.cursor()

        cur.execute("""
            SELECT
                id,
                original_name,
                content_type,
                size_bytes,
                created_at,
                creado_por
            FROM reclamo_adjuntos
            WHERE reclamo_id = ?
            ORDER BY id
        """, (reclamo_id,))

        rows = [dict(r) for r in cur.fetchall()]


        conn = get_db()
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(reclamo_imputados)")
        cols = {r["name"] for r in cur.fetchall()}

        if "metodo_analisis" not in cols:
            cur.execute("ALTER TABLE reclamo_imputados ADD COLUMN metodo_analisis TEXT")
        for n in range(1, 6):
            col = f"why{n}"
            if col not in cols:
                cur.execute(f"ALTER TABLE reclamo_imputados ADD COLUMN {col} TEXT")
        conn.commit()

        conn.close()
        return jsonify(items=rows)
    @app.route('/reclamos/adjunto/<int:adj_id>/download', methods=['GET'], endpoint='reclamos_download_adjunto')
    @require_login
    def reclamos_download_adjunto(adj_id):
        conn = get_db()
        ensure_reclamo_adjuntos_schema(conn)
        cur = conn.cursor()

        cur.execute("""
            SELECT
                id, reclamo_id, filename, original_name,
                content_type
            FROM reclamo_adjuntos
            WHERE id = ?
        """, (adj_id,))
        row = cur.fetchone()
        conn.close()

        if not row:
            abort(404)

        upload_folder = _get_reclamos_upload_folder()
        path = os.path.join(upload_folder, row["filename"])

        if not os.path.exists(path):
            abort(404)

        return send_file(
            path,
            as_attachment=True,
            download_name=row["original_name"],
            mimetype=row["content_type"] or "application/octet-stream"
        )

  
    # --- helper opcional: asegura columnas extra en reclamo_imputados ---
    def _ensure_reclamo_imputados_extra_cols(conn):
        """
        Asegura que existan las columnas para:
        - metodo_analisis
        - why1..why5
        - fish_metodo, fish_maquinas, fish_materiales, fish_personas, fish_entorno, fish_medicion
        No falla si ya existen.
        """
        cur = conn.execute("PRAGMA table_info(reclamo_imputados)")
        cols = {row["name"] for row in cur.fetchall()}

        def _add(col, decl):
            if col not in cols:
                conn.execute(f"ALTER TABLE reclamo_imputados ADD COLUMN {col} {decl}")
                cols.add(col)

        _add("metodo_analisis", "TEXT")
        _add("why1", "TEXT")
        _add("why2", "TEXT")
        _add("why3", "TEXT")
        _add("why4", "TEXT")
        _add("why5", "TEXT")

        _add("fish_metodo", "TEXT")
        _add("fish_maquinas", "TEXT")
        _add("fish_materiales", "TEXT")
        _add("fish_personas", "TEXT")
        _add("fish_entorno", "TEXT")
        _add("fish_medicion", "TEXT")
        _add("fecha_causa", "TEXT")
        _add("fecha_preventiva", "TEXT")
        _add("fecha_correctiva", "TEXT")

        # si tu tabla no tiene estas, puedes comentarlas:
        # _add("fecha_respuesta_imputado", "TEXT")
        # _add("estado_respuesta", "TEXT")


 
    # ----------------------------------------------
    # RESPUESTA TÉCNICA IMPUTADO (AUTO-APROBADA + AUTO-CIERRE)
    # ----------------------------------------------
    @app.route(
        '/reclamos/imputacion/<int:imp_id>/responder',
        methods=['POST'],
        endpoint='reclamos_responder_imputado'
    )
    @require_login
    @require_permission('reclamos', 'editar')
    def reclamos_responder_imputado(imp_id):
        uid = _current_user_id()
        data = request.get_json(silent=True) or {}

        def _safe_str(v):
            return (v or "").strip() if isinstance(v, str) else ""

        metodo = _safe_str(data.get("metodo_analisis") or "FISHBONE").upper()

        # ===== NUEVO FORMATO =====
        causas_items = data.get("causas") or []
        control_items = data.get("control") or []
        correctiva_items = data.get("correctiva") or []

        if not isinstance(causas_items, list):
            causas_items = []
        if not isinstance(control_items, list):
            control_items = []
        if not isinstance(correctiva_items, list):
            correctiva_items = []

        # ===== COMPATIBILIDAD FORMATO VIEJO =====
        causa_legacy = _safe_str(data.get("causa"))
        preventiva_legacy = _safe_str(data.get("preventiva"))
        correctiva_legacy = _safe_str(data.get("correctiva"))

        fecha_causa_legacy = _safe_str(data.get("fecha_causa"))
        fecha_preventiva_legacy = _safe_str(data.get("fecha_preventiva"))
        fecha_correctiva_legacy = _safe_str(data.get("fecha_correctiva"))

        if not causas_items and causa_legacy:
            causas_items = [{
                "descripcion": causa_legacy,
                "fecha_compromiso": fecha_causa_legacy
            }]

        if not control_items and preventiva_legacy:
            control_items = [{
                "descripcion": preventiva_legacy,
                "fecha_compromiso": fecha_preventiva_legacy
            }]

        if not correctiva_items and correctiva_legacy:
            correctiva_items = [{
                "descripcion": correctiva_legacy,
                "fecha_compromiso": fecha_correctiva_legacy
            }]

        # ===== 5 porqués =====
        why1 = _safe_str(data.get("why1"))
        why2 = _safe_str(data.get("why2"))
        why3 = _safe_str(data.get("why3"))
        why4 = _safe_str(data.get("why4"))
        why5 = _safe_str(data.get("why5"))

        # ===== Fishbone =====
        fish_metodo     = _safe_str(data.get("fish_metodo"))
        fish_maquinas   = _safe_str(data.get("fish_maquinas"))
        fish_materiales = _safe_str(data.get("fish_materiales"))
        fish_personas   = _safe_str(data.get("fish_personas"))
        fish_entorno    = _safe_str(data.get("fish_entorno"))
        fish_medicion   = _safe_str(data.get("fish_medicion"))

        conn = get_db()
        ensure_reclamos_schema(conn)
        _ensure_reclamo_imputados_extra_cols(conn)
        ensure_reclamo_imputado_acciones_schema(conn)

        cur = conn.cursor()

        cur.execute("""
            SELECT
                ri.*,
                r.id        AS reclamo_id,
                r.codigo    AS codigo,
                r.creado_por AS creado_por,

                u_apr.username AS aprobador_username,
                u_apr.id       AS aprobador_uid,

                u_imp.username AS imputado_username,

                u_cre.email    AS creador_email,
                u_cre.username AS creador_username,
                u_cre.nombre_completo AS creador_nombre

            FROM reclamo_imputados ri
            JOIN reclamos r          ON r.id = ri.reclamo_id
            LEFT JOIN usuarios u_imp ON u_imp.id = ri.imputado_id
            LEFT JOIN usuarios u_apr ON u_apr.id = ri.aprobador_id
            LEFT JOIN usuarios u_cre ON u_cre.id = r.creado_por
            WHERE ri.id = ?
        """, (imp_id,))
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="Imputación no encontrada"), 404

        if row["imputado_id"] != uid and (not _is_admin_like()):
            conn.close()
            return jsonify(ok=False, msg="No autorizado"), 403

        if row["estado_asignacion"] != "aprobado":
            conn.close()
            return jsonify(ok=False, msg="No puedes responder: imputación aún no aprobada"), 400

        # ===== VALIDACIONES =====
        if not causas_items or not control_items or not correctiva_items:
            conn.close()
            return jsonify(
                ok=False,
                msg="Debe ingresar al menos una causa, una acción de control y una correctiva"
            ), 400

        if metodo == "5WHYS":
            if not (why1 and why2 and why3 and why4 and why5):
                conn.close()
                return jsonify(
                    ok=False,
                    msg="Debe completar los 5 porqués cuando el método es '5 porqués'."
                ), 400

        now_iso = _now_iso()
        reclamo_id = int(row["reclamo_id"])

        # ===== ARMAR RESUMEN LEGACY =====
        def _join_items(items):
            vals = []
            for x in items:
                desc = _safe_str((x or {}).get("descripcion"))
                fecha = _safe_str((x or {}).get("fecha_compromiso"))
                if desc and fecha:
                    vals.append(f"{desc} ({fecha})")
                elif desc:
                    vals.append(desc)
                elif fecha:
                    vals.append(f"({fecha})")
            return " | ".join(vals)

        respuesta_causa = _join_items(causas_items)
        respuesta_preventiva = _join_items(control_items)
        respuesta_correctiva = _join_items(correctiva_items)

        # primera fecha visible legacy
        fecha_causa = _safe_str((causas_items[0] or {}).get("fecha_compromiso")) if causas_items else ""
        fecha_preventiva = _safe_str((control_items[0] or {}).get("fecha_compromiso")) if control_items else ""
        fecha_correctiva = _safe_str((correctiva_items[0] or {}).get("fecha_compromiso")) if correctiva_items else ""

        # ===== GUARDAR CABECERA OFICIAL =====
        cur.execute("""
            UPDATE reclamo_imputados
            SET metodo_analisis            = ?,
                respuesta_causa            = ?,
                fecha_causa                = ?,
                respuesta_preventiva       = ?,
                fecha_preventiva           = ?,
                respuesta_correctiva       = ?,
                fecha_correctiva           = ?,
                why1                       = ?,
                why2                       = ?,
                why3                       = ?,
                why4                       = ?,
                why5                       = ?,
                fish_metodo                = ?,
                fish_maquinas              = ?,
                fish_materiales            = ?,
                fish_personas              = ?,
                fish_entorno               = ?,
                fish_medicion              = ?,
                fecha_respuesta_imputado   = ?,
                estado_respuesta           = 'aprobada',
                fecha_aprobacion_respuesta = ?,
                motivo_rechazo_respuesta   = NULL,
                fecha_rechazo_respuesta    = NULL,
                visible_creador            = 1
            WHERE id = ?
        """, (
            metodo,
            respuesta_causa, fecha_causa,
            respuesta_preventiva, fecha_preventiva,
            respuesta_correctiva, fecha_correctiva,
            why1, why2, why3, why4, why5,
            fish_metodo, fish_maquinas, fish_materiales, fish_personas, fish_entorno, fish_medicion,
            now_iso,
            now_iso,
            imp_id
        ))

        # ===== GUARDAR DETALLE EN TABLA HIJA =====
        cur.execute("""
            UPDATE reclamo_imputado_acciones
            SET activo = 0,
                updated_at = ?,
                updated_by = ?
            WHERE imputacion_id = ?
            AND reclamo_id = ?
            AND COALESCE(activo,1) = 1
        """, (now_iso, uid, imp_id, reclamo_id))

        def _insert_items(tipo, items, requiere_evidencia_default=0):
            orden = 1
            for x in items:
                descripcion = _safe_str((x or {}).get("descripcion"))
                fecha_compromiso = _safe_str((x or {}).get("fecha_compromiso"))

                if not descripcion and not fecha_compromiso:
                    continue

                cur.execute("""
                    INSERT INTO reclamo_imputado_acciones (
                        imputacion_id,
                        reclamo_id,
                        tipo,
                        descripcion,
                        fecha_compromiso,
                        orden,
                        requiere_evidencia,
                        cumplido,
                        fecha_cumplimiento,
                        created_at,
                        created_by,
                        updated_at,
                        updated_by,
                        activo,
                        observacion_cumplimiento
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, 0, NULL, ?, ?, ?, ?, 1, '')
                """, (
                    imp_id,
                    reclamo_id,
                    tipo,
                    descripcion,
                    fecha_compromiso,
                    orden,
                    int((x or {}).get("requiere_evidencia", requiere_evidencia_default) or 0),
                    now_iso, uid,
                    now_iso, uid
                ))
                orden += 1

        _insert_items("CAUSA", causas_items, 0)
        _insert_items("CONTROL", control_items, 0)
        _insert_items("CORRECTIVA", correctiva_items, 1)

        # ===== CIERRE DE OM =====
        cur.execute("""
            SELECT COUNT(*) AS pend
            FROM reclamo_imputados
            WHERE reclamo_id = ?
            AND estado_asignacion = 'aprobado'
            AND COALESCE(TRIM(estado_respuesta),'') <> 'aprobada'
        """, (reclamo_id,))
        rowp = cur.fetchone()
        pend = int((rowp["pend"] if rowp else 0) or 0)

        if pend == 0:
            cur.execute("""
                UPDATE reclamos
                SET estado_global = 'cerrado'
                WHERE id = ?
            """, (reclamo_id,))

        try:
            _notify_creador_respuesta_aprobada(
                conn,
                row["creado_por"],
                row["codigo"],
                row["imputado_username"]
            )
        except Exception as e:
            current_app.logger.warning("[responder_imputado] no se pudo notificar al creador: %s", e)

        conn.commit()
        conn.close()
        return jsonify(ok=True)
    
    
    @app.route('/reclamos/<int:reclamo_id>/eliminar', methods=['POST'], endpoint='reclamos_eliminar')
    @require_login
    def reclamos_eliminar(reclamo_id):
        conn = get_db()
        ensure_reclamos_schema(conn)
        ensure_reclamo_adjuntos_schema(conn)
        ensure_reclamo_respuestas_equipo_schema(conn)

        uid = session.get("usuario_id") or session.get("user_id") or session.get("id")
        role = (session.get('rol') or '').strip().lower()

        # ✅ Permitido:
        # - admin
        # - personal de Servicio al Cliente
        autorizado = (role == 'admin') or _can_view_all_reclamos(conn, uid)

        if not autorizado:
            conn.close()
            return jsonify(ok=False, msg='No autorizado'), 403

        cur = conn.cursor()

        # Validar existencia + estado global
        cur.execute("""
            SELECT id, codigo, COALESCE(LOWER(TRIM(estado_global)), '') AS estado_global
            FROM reclamos
            WHERE id = ?
        """, (reclamo_id,))
        r = cur.fetchone()

        if not r:
            conn.close()
            return jsonify(ok=False, msg='OM no encontrada'), 404

        # ✅ No permitir eliminar si está cerrada
        if (r["estado_global"] or "") == "cerrado":
            conn.close()
            return jsonify(ok=False, msg='No se puede eliminar una OM con estado global Cerrado'), 400

        # Adjuntos físicos (si existen)
        cur.execute("SELECT filename FROM reclamo_adjuntos WHERE reclamo_id = ?", (reclamo_id,))
        files = [row["filename"] for row in cur.fetchall() if row and row["filename"]]

        # Borrado en cascada (manual)
        try:
            cur.execute("""
                DELETE FROM reclamo_equipo_acciones
                WHERE equipo_id IN (
                    SELECT id FROM reclamo_equipo_respuestas WHERE reclamo_id = ?
                )
            """, (reclamo_id,))
        except sqlite3.OperationalError:
            pass

        cur.execute("DELETE FROM reclamo_respuestas_equipo WHERE reclamo_id = ?", (reclamo_id,))
        cur.execute("DELETE FROM reclamo_equipo_respuestas WHERE reclamo_id = ?", (reclamo_id,))
        cur.execute("DELETE FROM reclamo_imputados WHERE reclamo_id = ?", (reclamo_id,))
        cur.execute("DELETE FROM reclamo_adjuntos WHERE reclamo_id = ?", (reclamo_id,))
        cur.execute("DELETE FROM reclamos WHERE id = ?", (reclamo_id,))

        conn.commit()
        conn.close()

        # Borrar archivos físicos
        upload_folder = _get_reclamos_upload_folder()
        for fn in files:
            try:
                path = os.path.join(upload_folder, fn)
                if os.path.exists(path):
                    os.remove(path)
            except Exception:
                current_app.logger.exception("No se pudo eliminar adjunto físico: %s", fn)

        return jsonify(ok=True, msg=f"OM {r['codigo']} eliminada correctamente.")
        
        
    # ----------------------------------------------
    # VALIDAR RESPUESTA (JEFE)
    # ----------------------------------------------
    @app.route('/reclamos/imputacion/<int:imp_id>/validar_respuesta', methods=['POST'], endpoint='reclamos_validar_respuesta')
    @require_login
    @require_permission('reclamos', 'aprobar')
    def reclamos_validar_respuesta(imp_id):
        uid = _current_user_id()
        data = request.get_json(silent=True) or {}
        accion = (data.get("accion") or "").strip().lower()
        motivo = (data.get("motivo") or "").strip()

        conn = get_db()
        ensure_reclamos_schema(conn)
        cur = conn.cursor()

        cur.execute("""
            SELECT ri.*, r.codigo, r.creado_por,
                   u_imp.username AS imputado_username,
                   u_imp.id AS imputado_uid
            FROM reclamo_imputados ri
            JOIN reclamos r ON r.id = ri.reclamo_id
            LEFT JOIN usuarios u_imp ON u_imp.id = ri.imputado_id
            WHERE ri.id = ?
        """, (imp_id,))
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify(ok=False, msg="Registro no encontrado"), 404

        if (row["aprobador_id"] != uid) and (not _is_admin_like()):
            conn.close()
            return jsonify(ok=False, msg="No autorizado"), 403

        if row["estado_asignacion"] != "aprobado":
            conn.close()
            return jsonify(ok=False, msg="Imputación aún no aprobada, no se valida respuesta"), 400

        if row["estado_respuesta"] not in ("pendiente_jefe", "rechazada"):
            conn.close()
            return jsonify(ok=False, msg="No hay respuesta pendiente de validar"), 400

        if accion == "aprobar":
            cur.execute("""
                UPDATE reclamo_imputados
                SET estado_respuesta='aprobada',
                    fecha_aprobacion_respuesta=?,
                    motivo_rechazo_respuesta=NULL,
                    fecha_rechazo_respuesta=NULL,
                    visible_creador=1
                WHERE id=?
            """, (_now_iso(), imp_id))

            cur.execute("""
                UPDATE reclamos
                SET estado_global = 'cerrado'
                WHERE id = ?
            """, (row["reclamo_id"],))
            
            _notify_creador_respuesta_aprobada(
                conn,
                row["creado_por"],
                row["codigo"],
                row["imputado_username"]
            )

        elif accion == "rechazar":
            if not motivo:
                conn.close()
                return jsonify(ok=False, msg="Motivo obligatorio al rechazar"), 400

            cur.execute("""
                UPDATE reclamo_imputados
                SET estado_respuesta='rechazada',
                    fecha_rechazo_respuesta=?,
                    motivo_rechazo_respuesta=?,
                    fecha_aprobacion_respuesta=NULL
                WHERE id=?
            """, (_now_iso(), motivo, imp_id))

            _notify_imputado_respuesta_rechazada(
                conn,
                row["imputado_uid"],
                row["codigo"],
                motivo
            )
        else:
            conn.close()
            return jsonify(ok=False, msg="Acción inválida"), 400

        conn.commit()
        conn.close()
        return jsonify(ok=True)



    from flask import jsonify
    from datetime import datetime

    # =========================
    # 1) Resumen Ejecutivo
    # =========================
    @app.route("/api/dashboard/resumen")
    @require_login
    def reclamos_api_dashboard_resumen():
        db = get_db()
        row = db.execute("""
            SELECT
                COUNT(*) AS total_om,
                SUM(CASE WHEN estado_global LIKE 'CERR%' THEN 1 ELSE 0 END) AS cerradas,
                SUM(CASE WHEN estado_global LIKE 'RECHA%' THEN 1 ELSE 0 END) AS rechazadas,
                SUM(CASE WHEN estado_global LIKE 'PEND%' THEN 1 ELSE 0 END) AS pendientes,
                SUM(CASE WHEN estado_global LIKE 'EN RESP%' THEN 1 ELSE 0 END) AS en_respuesta
            FROM reclamos
        """).fetchone()

        # OM creadas por mes (últimos 6 meses)
        rows_mes = db.execute("""
            SELECT
                strftime('%Y-%m', fecha_reclamo) AS ym,
                COUNT(*) AS total_mes
            FROM reclamos
            WHERE fecha_reclamo >= date('now', '-6 months')
            GROUP BY ym
            ORDER BY ym
        """).fetchall()

        return jsonify({
            "kpi": dict(row or {}),
            "serie_mes": [dict(r) for r in rows_mes]
        })


    # =========================
    # 2) Flujo de estados
    # =========================
    @app.route("/api/dashboard/estados")
    @require_login
    def reclamos_api_dashboard_estados():
        db = get_db()
        rows = db.execute("""
            SELECT
                estado_global AS estado,
                COUNT(*) AS total
            FROM reclamos
            GROUP BY estado_global
            ORDER BY total DESC
        """).fetchall()
        return jsonify({"items": [dict(r) for r in rows]})


    # =========================
    # 3) SLA / tiempos de ciclo
    #    (ajusta 'fecha_cierre' al nombre real)
    # =========================
    @app.route("/api/dashboard/sla")
    @require_login
    def reclamos_api_dashboard_sla():
        db = get_db()
        rows = db.execute("""
            SELECT
                tipo_reclamo,
                AVG(
                    JULIANDAY(COALESCE(fecha_cierre, DATE('now')))
                    - JULIANDAY(fecha_reclamo)
                ) AS dias_promedio,
                COUNT(*) AS total
            FROM reclamos
            WHERE fecha_reclamo IS NOT NULL
            GROUP BY tipo_reclamo
            HAVING total >= 3           -- opcional: solo tipos con algo de muestras
            ORDER BY dias_promedio DESC
        """).fetchall()
        return jsonify({"items": [dict(r) for r in rows]})


    # =========================
    # 4) OM por sponsor / imputado
    # =========================
    @app.route("/api/dashboard/imputados")
    @require_login
    def reclamos_api_dashboard_imputados():
        db = get_db()
        rows = db.execute("""
            SELECT
                u.nombre_completo AS imputado,
                COUNT(*) AS total_om,
                SUM(CASE WHEN ri.estado_imputacion LIKE 'CERR%' THEN 1 ELSE 0 END) AS cerradas,
                SUM(CASE WHEN ri.estado_imputacion LIKE 'APROB%' THEN 1 ELSE 0 END) AS aprobadas,
                SUM(CASE WHEN ri.estado_imputacion LIKE 'RECHA%' THEN 1 ELSE 0 END) AS rechazadas
            FROM reclamo_imputados ri
            JOIN usuarios u ON u.id = ri.imputado_id
            GROUP BY u.id, u.nombre_completo
            ORDER BY total_om DESC
            LIMIT 10
        """).fetchall()
        return jsonify({"items": [dict(r) for r in rows]})


    # =========================
    # 5) Clientes / procesos
    # =========================
    @app.route("/api/dashboard/clientes_procesos")
    @require_login
    def reclamos_api_dashboard_clientes_procesos():
        db = get_db()

        top_clientes = db.execute("""
            SELECT
                cliente_nombre AS cliente,
                COUNT(*) AS total_om
            FROM reclamos
            GROUP BY cliente_nombre
            ORDER BY total_om DESC
            LIMIT 10
        """).fetchall()

        por_proceso = db.execute("""
            SELECT
                proceso_text AS proceso,
                COUNT(*) AS total_om
            FROM reclamos
            GROUP BY proceso_text
            ORDER BY total_om DESC
        """).fetchall()

        return jsonify({
            "clientes": [dict(r) for r in top_clientes],
            "procesos": [dict(r) for r in por_proceso]
        })


    # =========================
    # 6) OM por Departamento  (ya lo tenías, lo dejo completo)
    # =========================
    @app.route("/api/dashboard/departamentos")
    @require_login
    def reclamos_api_dashboard_departamentos():
        db = get_db()
        rows = db.execute("""
            SELECT
                COALESCE(d.nombre, 'SIN DEPARTAMENTO') AS departamento,
                COUNT(DISTINCT r.id) AS total_om,
                SUM(CASE WHEN r.estado_global LIKE 'CERR%' THEN 1 ELSE 0 END) AS cerradas,
                SUM(CASE WHEN r.estado_global LIKE 'CERR%' THEN 0 ELSE 1 END) AS abiertas
            FROM reclamos r
            LEFT JOIN usuarios u   ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            GROUP BY d.nombre
            ORDER BY total_om DESC
        """).fetchall()

        return jsonify({"items": [dict(row) for row in rows]})



 # routes_reclamos.py (solo la parte de la ruta completa)
    @app.route('/reclamos/dashboard', methods=['GET'], endpoint='reclamos_dashboard')
    @require_login
    @require_permission('reclamos', 'ver')
    def reclamos_dashboard():
        conn = get_db()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        fecha_desde = (request.args.get("desde") or "").strip()
        fecha_hasta = (request.args.get("hasta") or "").strip()
        depto_sel   = (request.args.get("depto") or "").strip()
        proceso_sel = (request.args.get("proceso") or "").strip()

        where = []
        params = []

        if fecha_desde:
            where.append("date(COALESCE(r.fecha_reclamo, r.fecha_creacion)) >= date(?)")
            params.append(fecha_desde)

        if fecha_hasta:
            where.append("date(COALESCE(r.fecha_reclamo, r.fecha_creacion)) <= date(?)")
            params.append(fecha_hasta)

        # NOTA: si quieres que "Departamento" filtre también este dashboard,
        # define qué significa: depto del creador. Aquí lo aplico al creador:
        if depto_sel:
            where.append("COALESCE(d.nombre, 'SIN DEPARTAMENTO') = ?")
            params.append(depto_sel)
        if proceso_sel:
            where.append("COALESCE(NULLIF(TRIM(r.proceso_text), ''), 'SIN PROCESO') = ?")
            params.append(proceso_sel)

        where_sql = "WHERE " + " AND ".join(where) if where else ""

        # ========= KPIs =========
        cur.execute(f"""
            SELECT
                COUNT(DISTINCT r.id) AS total,

                SUM(
                    CASE
                        WHEN LOWER(COALESCE(r.estado_global,'')) = 'cerrado'
                            OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                        THEN 1 ELSE 0
                    END
                ) AS cerradas,

                SUM(
                    CASE
                        WHEN LOWER(COALESCE(r.estado_global,'')) LIKE 'pendiente respuesta del imputado%'
                        THEN 1 ELSE 0
                    END
                ) AS pend_imputados,

                SUM(
                    CASE
                        WHEN LOWER(COALESCE(r.estado_global,'')) LIKE 'respuesta pendiente de aprobación%'
                        OR LOWER(COALESCE(r.estado_global,'')) LIKE 'pendiente respuesta del sponsor%'
                        OR LOWER(COALESCE(r.estado_global,'')) LIKE 'pendiente respuesta del responsable%'
                        THEN 1 ELSE 0
                    END
                ) AS pend_sponsor

            FROM reclamos r
            LEFT JOIN usuarios u ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            {where_sql}
        """, params)

        row = cur.fetchone()
        total = int(row["total"] or 0) if row else 0
        cerradas = int(row["cerradas"] or 0) if row else 0
        pend_imputados = int(row["pend_imputados"] or 0) if row else 0
        pend_sponsor = int(row["pend_sponsor"] or 0) if row else 0

        kpis = {
            "total": total,
            "pend_jefe": pend_sponsor,   # compatibilidad visual
            "pend_imputados": pend_imputados,
            "cerradas": cerradas,
            "abiertas": max(total - cerradas, 0),
            "abiertas": max(total - cerradas, 0),
        }

        # ========= CHART: estados =========
        cur.execute(f"""
            SELECT
            COALESCE(r.estado_global, 'SIN ESTADO') AS estado,
            COUNT(*) AS total
            FROM reclamos r
            LEFT JOIN usuarios u ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            {where_sql}
            GROUP BY estado
            ORDER BY total DESC
        """, params)
        rows = cur.fetchall()
        chart_estados = {"labels": [r["estado"] for r in rows], "total": [r["total"] for r in rows]}

        # ========= CHART: meses =========
        cur.execute(f"""
            SELECT
            strftime('%Y-%m', r.fecha_creacion) AS ym,
            COUNT(*) AS total
            FROM reclamos r
            LEFT JOIN usuarios u ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            {where_sql}
            GROUP BY ym
            ORDER BY ym
        """, params)
        rows = cur.fetchall()
        chart_meses = {"labels": [r["ym"] for r in rows], "total": [r["total"] for r in rows]}
        # ========= CHART: días =========
        cur.execute(f"""
            SELECT
                date(r.fecha_creacion) AS dia,
                COUNT(*) AS total
            FROM reclamos r
            LEFT JOIN usuarios u ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            {where_sql}
            GROUP BY date(r.fecha_creacion)
            ORDER BY date(r.fecha_creacion)
        """, params)
        rows = cur.fetchall()
        chart_dias = {
            "labels": [r["dia"] for r in rows],
            "total":  [r["total"] for r in rows],
        }
        # ========= CHART: imputados (creador como proxy) =========
        cur.execute(f"""
            SELECT
            COALESCE(u.nombre_completo, u.username, 'SIN USUARIO') AS imputado,
            SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado' OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                    THEN 1 ELSE 0 END) AS cerradas,
            SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado' OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                    THEN 0 ELSE 1 END) AS abiertas
            FROM reclamos r
            LEFT JOIN usuarios u ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            {where_sql}
            GROUP BY imputado
            ORDER BY (abiertas + cerradas) DESC
            LIMIT 10
        """, params)
        rows = cur.fetchall()
        chart_imputados = {
            "labels":   [r["imputado"] for r in rows],
            "abiertas": [r["abiertas"] for r in rows],
            "cerradas": [r["cerradas"] for r in rows],
        }

        # ========= CHART: procesos (CREADAS) =========
        # ========= CHART: procesos (abiertas/cerradas + total) =========
        cur.execute(f"""
            SELECT
                COALESCE(NULLIF(TRIM(r.proceso_text), ''), 'SIN PROCESO') AS proceso,
                SUM(
                    CASE
                        WHEN LOWER(COALESCE(r.estado_global,''))='cerrado'
                             OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                        THEN 1 ELSE 0
                    END
                ) AS cerradas,
                SUM(
                    CASE
                        WHEN LOWER(COALESCE(r.estado_global,''))='cerrado'
                             OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                        THEN 0 ELSE 1
                    END
                ) AS abiertas,
                COUNT(*) AS total
            FROM reclamos r
            LEFT JOIN usuarios u ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            {where_sql}
            GROUP BY COALESCE(NULLIF(TRIM(r.proceso_text), ''), 'SIN PROCESO')
            ORDER BY total DESC
            LIMIT 10
        """, params)

        rows = cur.fetchall()
        chart_procesos = {
            "labels":   [r["proceso"] for r in rows],
            "abiertas": [int(r["abiertas"] or 0) for r in rows],
            "cerradas": [int(r["cerradas"] or 0) for r in rows],
            "total":    [int(r["total"] or 0) for r in rows],
        }

        # ========= CHART: tipos =========
        cur.execute(f"""
            SELECT
            COALESCE(r.tipo_reclamo, 'SIN TIPO') AS tipo,
            SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado' OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                    THEN 1 ELSE 0 END) AS cerradas,
            SUM(CASE WHEN LOWER(COALESCE(r.estado_global,''))='cerrado' OR UPPER(COALESCE(r.estado_global,'')) LIKE 'CERR%'
                    THEN 0 ELSE 1 END) AS abiertas
            FROM reclamos r
            LEFT JOIN usuarios u ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            {where_sql}
            GROUP BY tipo
            ORDER BY (abiertas + cerradas) DESC
            LIMIT 10
        """, params)
        rows = cur.fetchall()
        chart_tipos = {
            "labels":   [r["tipo"] for r in rows],
            "abiertas": [r["abiertas"] for r in rows],
            "cerradas": [r["cerradas"] for r in rows],
        }

         # ========= CHART: tiempo promedio de respuesta por proceso =========
        cur.execute(f"""
            SELECT
                COALESCE(NULLIF(TRIM(r.proceso_text), ''), 'SIN PROCESO') AS proceso,
                AVG(
                    CASE
                        WHEN ri.fecha_respuesta_imputado IS NOT NULL
                        THEN julianday(ri.fecha_respuesta_imputado)
                             - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion))
                    END
                ) AS dias_promedio_respuesta,
                COUNT(DISTINCT r.id) AS total_om
            FROM reclamos r
            LEFT JOIN reclamo_imputados ri ON ri.reclamo_id = r.id
            LEFT JOIN usuarios u ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            {where_sql}
            GROUP BY COALESCE(NULLIF(TRIM(r.proceso_text), ''), 'SIN PROCESO')
            HAVING AVG(
                CASE
                    WHEN ri.fecha_respuesta_imputado IS NOT NULL
                    THEN julianday(ri.fecha_respuesta_imputado)
                         - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion))
                END
            ) IS NOT NULL
            ORDER BY dias_promedio_respuesta DESC
            LIMIT 12
        """, params)

        rows = cur.fetchall()
        chart_tiempos = {
            "labels": [r["proceso"] for r in rows],
            "promedio": [round(r["dias_promedio_respuesta"] or 0, 2) for r in rows],
            "total_om": [int(r["total_om"] or 0) for r in rows],
        }


        # ========= CHART: top procesos con OM vencidas (>5 días sin respuesta) =========
        cur.execute(f"""
            SELECT
                COALESCE(NULLIF(TRIM(r.proceso_text), ''), 'SIN PROCESO') AS proceso,
                COUNT(DISTINCT r.id) AS vencidas
            FROM reclamos r
            LEFT JOIN reclamo_imputados ri ON ri.reclamo_id = r.id
            LEFT JOIN usuarios u ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            {where_sql}
              {"AND" if where_sql else "WHERE"}
              ri.fecha_respuesta_imputado IS NULL
              AND (
                    julianday('now') - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion))
                  ) > 5
            GROUP BY COALESCE(NULLIF(TRIM(r.proceso_text), ''), 'SIN PROCESO')
            ORDER BY vencidas DESC, proceso ASC
            LIMIT 10
        """, params)

        rows = cur.fetchall()
        chart_vencidas = {
            "labels": [r["proceso"] for r in rows],
            "total":  [int(r["vencidas"] or 0) for r in rows],
        }


        # ========= CHART: top sponsor / miembros de equipo con OM vencidas (>5 días sin respuesta) =========
        cur.execute(f"""
            SELECT
                COALESCE(
                    NULLIF(TRIM(uimp.nombre_completo), ''),
                    NULLIF(TRIM(uimp.username), ''),
                    'SIN RESPONSABLE'
                ) AS responsable,
                COUNT(DISTINCT ri.id) AS vencidas
            FROM reclamos r
            INNER JOIN reclamo_imputados ri ON ri.reclamo_id = r.id
            LEFT JOIN usuarios uimp ON uimp.id = ri.imputado_id
            LEFT JOIN usuarios u ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            {where_sql}
              {"AND" if where_sql else "WHERE"}
              ri.fecha_aprobacion_asignacion IS NOT NULL
              AND ri.fecha_respuesta_imputado IS NULL
              AND (julianday('now') - julianday(ri.fecha_aprobacion_asignacion)) > 5
            GROUP BY COALESCE(
                NULLIF(TRIM(uimp.nombre_completo), ''),
                NULLIF(TRIM(uimp.username), ''),
                'SIN RESPONSABLE'
            )
            ORDER BY vencidas DESC, responsable ASC
            LIMIT 10
        """, params)

        rows = cur.fetchall()
        chart_vencidas_responsables = {
            "labels": [r["responsable"] for r in rows],
            "total":  [int(r["vencidas"] or 0) for r in rows],
        }



        # ========= CHART: top miembros de equipo con OM vencidas (>5 días sin respuesta) =========
        cur.execute(f"""
            SELECT
                COALESCE(
                    NULLIF(TRIM(ume.nombre_completo), ''),
                    NULLIF(TRIM(ume.username), ''),
                    'SIN MIEMBRO'
                ) AS miembro_equipo,

                COALESCE(
                    NULLIF(TRIM(us.nombre_completo), ''),
                    NULLIF(TRIM(us.username), ''),
                    'SIN SPONSOR'
                ) AS sponsor,

                COALESCE(NULLIF(TRIM(r.proceso_text), ''), 'SIN PROCESO') AS proceso,

                COUNT(DISTINCT eq.id) AS vencidas,

                ROUND(
                    AVG(
                        julianday('now') - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion))
                    ),
                    2
                ) AS atraso_promedio,

                MAX(
                    ROUND(
                        julianday('now') - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion)),
                        2
                    )
                ) AS atraso_maximo

            FROM reclamos r
            INNER JOIN reclamo_imputados ri
                ON ri.reclamo_id = r.id

            INNER JOIN reclamo_equipo_respuestas eq
                ON eq.reclamo_id = r.id
            AND eq.imputacion_id = ri.id
            AND eq.activo = 1

            LEFT JOIN usuarios ume
                ON ume.id = eq.usuario_id

            LEFT JOIN usuarios us
                ON us.id = ri.imputado_id

            LEFT JOIN reclamo_respuestas_equipo rre
                ON rre.reclamo_id = eq.reclamo_id
            AND rre.imputacion_id = eq.imputacion_id
            AND rre.miembro_id = eq.usuario_id
            AND rre.activo = 1

            LEFT JOIN usuarios u
                ON u.id = r.creado_por

            LEFT JOIN departamentos d
                ON d.id = u.departamento_id

            {where_sql}
            {"AND" if where_sql else "WHERE"}
            (
                rre.id IS NULL
                OR (
                    COALESCE(TRIM(rre.causa), '') = ''
                    AND COALESCE(TRIM(rre.preventiva), '') = ''
                    AND COALESCE(TRIM(rre.correctiva), '') = ''
                )
            )
            AND (julianday('now') - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion))) > 5

            GROUP BY eq.usuario_id, sponsor, proceso
            ORDER BY atraso_maximo DESC, vencidas DESC, miembro_equipo ASC
            LIMIT 10
        """, params)

        rows = cur.fetchall()

        chart_vencidas_equipo = {
            "labels": [f'{r["miembro_equipo"]} | {r["sponsor"]}' for r in rows],
            "miembro": [r["miembro_equipo"] for r in rows],
            "sponsor": [r["sponsor"] for r in rows],
            "proceso": [r["proceso"] for r in rows],
            "total": [int(r["vencidas"] or 0) for r in rows],
            "atraso_promedio": [float(r["atraso_promedio"] or 0) for r in rows],
            "atraso_maximo": [float(r["atraso_maximo"] or 0) for r in rows],
        }
            

        # ========= CHART: top miembros de equipo con mayor atraso en responder =========
        # ========= CHART: top miembros de equipo con mayor atraso en responder =========
        cur.execute(f"""
            SELECT
                COALESCE(
                    NULLIF(TRIM(ume.nombre_completo), ''),
                    NULLIF(TRIM(ume.username), ''),
                    'SIN MIEMBRO'
                ) AS miembro_equipo,

                COALESCE(
                    NULLIF(TRIM(us.nombre_completo), ''),
                    NULLIF(TRIM(us.username), ''),
                    'SIN SPONSOR'
                ) AS sponsor,

                COALESCE(NULLIF(TRIM(r.proceso_text), ''), 'SIN PROCESO') AS proceso,

                COUNT(DISTINCT eq.id) AS vencidas,

                ROUND(
                    AVG(
                        julianday('now') - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion))
                    ),
                    2
                ) AS atraso_promedio,

                MAX(
                    ROUND(
                        julianday('now') - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion)),
                        2
                    )
                ) AS atraso_maximo

            FROM reclamos r
            INNER JOIN reclamo_imputados ri
                ON ri.reclamo_id = r.id

            INNER JOIN reclamo_equipo_respuestas eq
                ON eq.reclamo_id = r.id
               AND eq.imputacion_id = ri.id
               AND eq.activo = 1

            LEFT JOIN usuarios ume
                ON ume.id = eq.usuario_id

            LEFT JOIN usuarios us
                ON us.id = ri.imputado_id

            LEFT JOIN reclamo_respuestas_equipo rre
                ON rre.reclamo_id = eq.reclamo_id
               AND rre.imputacion_id = eq.imputacion_id
               AND rre.miembro_id = eq.usuario_id
               AND rre.activo = 1

            LEFT JOIN usuarios u
                ON u.id = r.creado_por

            LEFT JOIN departamentos d
                ON d.id = u.departamento_id

            {where_sql}
              {"AND" if where_sql else "WHERE"}
              (
                  rre.id IS NULL
                  OR (
                      COALESCE(TRIM(rre.causa), '') = ''
                      AND COALESCE(TRIM(rre.preventiva), '') = ''
                      AND COALESCE(TRIM(rre.correctiva), '') = ''
                  )
              )
              AND (julianday('now') - julianday(COALESCE(ri.fecha_aprobacion_asignacion, r.fecha_creacion))) > 5

            GROUP BY eq.usuario_id, sponsor, proceso
            ORDER BY atraso_maximo DESC, vencidas DESC, miembro_equipo ASC
            LIMIT 10
        """, params)

        rows = cur.fetchall()

        chart_vencidas_equipo = {
            "labels": [
                f'{r["miembro_equipo"]} | {r["sponsor"]}'
                for r in rows
            ],
            "miembro": [r["miembro_equipo"] for r in rows],
            "sponsor": [r["sponsor"] for r in rows],
            "proceso": [r["proceso"] for r in rows],
            "total": [int(r["vencidas"] or 0) for r in rows],
            "atraso_promedio": [float(r["atraso_promedio"] or 0) for r in rows],
            "atraso_maximo": [float(r["atraso_maximo"] or 0) for r in rows],
        }        
        ##Envío a pantalla 
        chart = {
            "estados": chart_estados,
            "meses": chart_meses,
            "dias": chart_dias,
            "imputados": chart_imputados,
            "procesos": chart_procesos,
            "tipos": chart_tipos,
            "tiempos": chart_tiempos,
            "vencidas_responsables": chart_vencidas_responsables,
            "vencidas": chart_vencidas,
            "vencidas_equipo": chart_vencidas_equipo,
            

        }

        # combo deptos (del creador)
        cur.execute("""
            SELECT DISTINCT COALESCE(d.nombre, 'SIN DEPARTAMENTO') AS depto
            FROM reclamos r
            LEFT JOIN usuarios u ON u.id = r.creado_por
            LEFT JOIN departamentos d ON d.id = u.departamento_id
            ORDER BY depto
        """)
        deptos = [r["depto"] for r in cur.fetchall()]



        cur.execute("""
            SELECT DISTINCT COALESCE(NULLIF(TRIM(r.proceso_text), ''), 'SIN PROCESO') AS proceso
            FROM reclamos r
            ORDER BY proceso
        """)
        procesos = [r["proceso"] for r in cur.fetchall()]

        return render_template(
            "reclamos_dashboard.html",
            kpis=kpis,
            chart=chart,
            deptos=deptos,
            depto_sel=depto_sel,
            procesos=procesos,
            proceso_sel=proceso_sel,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            active_page="reclamos_dashboard",
        )




    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/aprobar", methods=["POST"])
    @require_login
    def equipo_respuestas_aprobar(reclamo_id):
        uid = _current_user_id()
        data = request.get_json(silent=True) or {}
        imputacion_id = data.get("imputacion_id")
        miembro_id = data.get("miembro_id")

        if not imputacion_id or not miembro_id:
            return jsonify(ok=False, error="Falta imputacion_id o miembro_id"), 400

        db = get_db()
        ensure_reclamos_schema(db)
        ensure_reclamo_respuestas_equipo_schema(db)

        # ✅ Solo el responsable (imputado) de esa imputación puede aprobar
        row = db.execute("""
            SELECT 1
            FROM reclamo_imputados
            WHERE id = ?
            AND reclamo_id = ?
            AND imputado_id = ?
            LIMIT 1
        """, (imputacion_id, reclamo_id, uid)).fetchone()

        if not row and not _is_admin_like():
            return jsonify(ok=False, error="No autorizado"), 403

        # (Opcional) Registrar marca simple (si no tienes columnas, puedes omitir este update)
        try:
            db.execute("""
                UPDATE reclamo_respuestas_equipo
                SET estado_revision='APROBADA', revision_by=?, revision_at=?
                WHERE reclamo_id=? AND imputacion_id=? AND miembro_id=? AND activo=1
            """, (uid, _now_iso(), reclamo_id, imputacion_id, miembro_id))
            db.commit()
        except Exception:
            # si no existen columnas, no rompe
            pass

        return jsonify(ok=True)


    @app.route("/reclamos/<int:reclamo_id>/equipo-respuestas/rechazar", methods=["POST"])
    @require_login
    def equipo_respuestas_rechazar(reclamo_id):
        uid = _current_user_id()
        data = request.get_json(silent=True) or {}
        imputacion_id = data.get("imputacion_id")
        miembro_id = data.get("miembro_id")
        motivo = (data.get("motivo") or "").strip()

        if not imputacion_id or not miembro_id or not motivo:
            return jsonify(ok=False, error="Falta imputacion_id, miembro_id o motivo"), 400

        db = get_db()
        ensure_reclamos_schema(db)
        ensure_reclamo_respuestas_equipo_schema(db)

        # ✅ Solo el responsable (imputado) puede rechazar
        row = db.execute("""
            SELECT r.codigo, u_imp.nombre_completo AS imputado_nombre
            FROM reclamo_imputados ri
            JOIN reclamos r ON r.id = ri.reclamo_id
            LEFT JOIN usuarios u_imp ON u_imp.id = ri.imputado_id
            WHERE ri.id = ?
            AND ri.reclamo_id = ?
            AND ri.imputado_id = ?
            LIMIT 1
        """, (imputacion_id, reclamo_id, uid)).fetchone()

        if not row and not _is_admin_like():
            return jsonify(ok=False, error="No autorizado"), 403

        # Email al miembro
        u = db.execute("""
            SELECT email, nombre_completo, username
            FROM usuarios
            WHERE id = ?
            LIMIT 1
        """, (miembro_id,)).fetchone()

        if not u or not (u["email"] or "").strip():
            return jsonify(ok=False, error="El miembro no tiene email registrado."), 400

        codigo = (row["codigo"] if row else f"OM #{reclamo_id}")
        sponsor_nombre = (row["imputado_nombre"] if row else "Sponsor")

        subject = f"[OM {codigo}] Corrección solicitada en tu aporte"
        body = (
            f"Buenas tardes,\n\n"
            f"Tu aporte para la OM {codigo} fue revisado por {sponsor_nombre}.\n"
            f"Se solicita ajustar la respuesta con estas observaciones:\n\n"
            f"- {motivo}\n\n"
            f"Por favor ingresa a Bitácora y actualiza tu aporte.\n\n"
            f"Saludos.\n"
        )

        _send_mail_safe(u["email"], subject, body)

        # (Opcional) marca en DB
        try:
            db.execute("""
                UPDATE reclamo_respuestas_equipo
                SET estado_revision='RECHAZADA', revision_by=?, revision_at=?, revision_msg=?
                WHERE reclamo_id=? AND imputacion_id=? AND miembro_id=? AND activo=1
            """, (uid, _now_iso(), motivo, reclamo_id, imputacion_id, miembro_id))
            db.commit()
        except Exception:
            pass

        return jsonify(ok=True)
    

    @app.route('/reclamos/equipo-acciones/<int:accion_id>/observacion', methods=['POST'])
    @require_login
    def reclamo_equipo_accion_guardar_observacion(accion_id):
        db = get_db()
        uid = _current_user_id()

        ensure_reclamo_respuesta_equipo_acciones_schema(db)

        data = request.get_json(silent=True) or {}
        observacion = (data.get("observacion") or "").strip()

        row = db.execute("""
            SELECT
                id,
                COALESCE(cumplido, 0) AS cumplido
            FROM reclamo_respuesta_equipo_acciones
            WHERE id = ?
            AND COALESCE(activo, 1) = 1
            LIMIT 1
        """, (accion_id,)).fetchone()

        if not row:
            return jsonify(ok=False, error="Acción no encontrada"), 404

        if int(row["cumplido"] or 0) == 1:
            return jsonify(ok=False, error="La acción ya está cumplida y no permite editar observación"), 400

        db.execute("""
            UPDATE reclamo_respuesta_equipo_acciones
            SET
                observacion_cumplimiento = ?,
                updated_at = CURRENT_TIMESTAMP,
                updated_by = ?
            WHERE id = ?
        """, (observacion, uid, accion_id))

        db.commit()

        return jsonify(ok=True, msg="Observación guardada")    
        

    @app.route('/reclamos/equipo-acciones/evidencias/<int:evidencia_id>/eliminar', methods=['POST'])
    @require_login
    def reclamo_equipo_accion_evidencia_eliminar(evidencia_id):
        db = get_db()
        uid = _current_user_id()

        row = db.execute("""
            SELECT
                e.id,
                e.accion_id,
                e.filename,
                COALESCE(e.activo, 1) AS evidencia_activa,
                COALESCE(a.cumplido, 0) AS accion_cumplida
            FROM reclamo_respuesta_equipo_accion_evidencias e
            JOIN reclamo_respuesta_equipo_acciones a
            ON a.id = e.accion_id
            WHERE e.id = ?
            LIMIT 1
        """, (evidencia_id,)).fetchone()

        if not row:
            return jsonify(ok=False, error="Evidencia no encontrada"), 404

        if int(row["evidencia_activa"] or 0) != 1:
            return jsonify(ok=False, error="La evidencia ya fue eliminada"), 400

        if int(row["accion_cumplida"] or 0) == 1:
            return jsonify(ok=False, error="No se puede eliminar evidencia de una acción ya cumplida"), 400

        db.execute("""
            UPDATE reclamo_respuesta_equipo_accion_evidencias
            SET activo = 0
            WHERE id = ?
        """, (evidencia_id,))

        db.commit()

        return jsonify(ok=True, msg="Evidencia eliminada")


    @app.route('/reclamos/imputacion/<int:imp_id>/respuesta-detalle', methods=['GET'])
    @require_login
    def reclamo_imputacion_respuesta_detalle(imp_id):
        conn = get_db()
        ensure_reclamos_schema(conn)
        ensure_reclamo_imputado_acciones_schema(conn)

        cur = conn.cursor()
        cur.execute("""
            SELECT
                ri.id,
                ri.reclamo_id,
                ri.imputado_id,
                ri.metodo_analisis,
                ri.why1, ri.why2, ri.why3, ri.why4, ri.why5,
                ri.fish_metodo, ri.fish_maquinas, ri.fish_materiales,
                ri.fish_personas, ri.fish_entorno, ri.fish_medicion
            FROM reclamo_imputados ri
            WHERE ri.id = ?
            LIMIT 1
        """, (imp_id,))
        row = cur.fetchone()

        if not row:
            return jsonify(ok=False, error="Imputación no encontrada"), 404

        item = dict(row)

        acciones = _get_imputado_acciones_full(conn, imp_id)
        item["causas"] = acciones.get("causas", [])
        item["control"] = acciones.get("control", [])
        item["correctiva_items"] = acciones.get("correctiva", [])

        return jsonify(ok=True, item=item)
    


    @app.route('/reclamos/imputado-acciones/<int:accion_id>/observacion', methods=['POST'])
    @require_login
    def reclamo_imputado_accion_observacion(accion_id):
        db = get_db()
        ensure_reclamos_schema(db)
        ensure_reclamo_imputado_acciones_schema(db)

        uid = _current_user_id()
        permitido, row = _puede_gestionar_imputado_accion(accion_id, uid)
        if not permitido:
            return jsonify(ok=False, error="No autorizado para editar esta acción"), 403

        if not row:
            return jsonify(ok=False, error="Acción no encontrada"), 404

        if int(row["accion_activa"] or 0) != 1:
            return jsonify(ok=False, error="La acción no está activa"), 400

        if int(row["cumplido"] or 0) == 1:
            return jsonify(ok=False, error="La acción ya está cumplida y no permite editar observación"), 400

        data = request.get_json(silent=True) or {}
        observacion = (data.get("observacion") or "").strip()

        db.execute("""
            UPDATE reclamo_imputado_acciones
            SET
                observacion_cumplimiento = ?,
                updated_at = CURRENT_TIMESTAMP,
                updated_by = ?
            WHERE id = ?
        """, (observacion, uid, accion_id))

        db.commit()
        return jsonify(ok=True, msg="Observación guardada")


    @app.route('/reclamos/imputado-acciones/<int:accion_id>/cumplir', methods=['POST'])
    @require_login
    def reclamo_imputado_accion_cumplir(accion_id):
        db = get_db()
        ensure_reclamos_schema(db)
        ensure_reclamo_imputado_acciones_schema(db)

        uid = _current_user_id()
        permitido, row = _puede_gestionar_imputado_accion(accion_id, uid)
        if not permitido:
            return jsonify(ok=False, error="No autorizado para editar esta acción"), 403

        if not row:
            return jsonify(ok=False, error="Acción no encontrada"), 404

        if int(row["accion_activa"] or 0) != 1:
            return jsonify(ok=False, error="La acción no está activa"), 400

        if int(row["cumplido"] or 0) == 1:
            return jsonify(ok=False, error="La acción ya está cumplida"), 400

        data = request.get_json(silent=True) or {}
        cumplido = data.get("cumplido", True)
        fecha_cumplimiento = (data.get("fecha_cumplimiento") or "").strip()

        if not bool(cumplido):
            return jsonify(ok=False, error="Acción inválida"), 400

        if not fecha_cumplimiento:
            fecha_cumplimiento = datetime.now().strftime("%Y-%m-%d")

        if not _is_date_yyyy_mm_dd(fecha_cumplimiento):
            return jsonify(ok=False, error="Fecha inválida. Use YYYY-MM-DD"), 400

        db.execute("""
            UPDATE reclamo_imputado_acciones
            SET
                cumplido = 1,
                fecha_cumplimiento = ?,
                updated_at = CURRENT_TIMESTAMP,
                updated_by = ?
            WHERE id = ?
        """, (fecha_cumplimiento, uid, accion_id))

        db.commit()
        return jsonify(ok=True, msg="Acción marcada como cumplida")


    @app.route('/reclamos/imputado-acciones/<int:accion_id>/evidencia', methods=['POST'])
    @require_login
    def reclamo_imputado_accion_evidencia(accion_id):
        db = get_db()
        ensure_reclamos_schema(db)
        ensure_reclamo_imputado_acciones_schema(db)

        uid = _current_user_id()
        permitido, row = _puede_gestionar_imputado_accion(accion_id, uid)
        if not permitido:
            return jsonify(ok=False, error="No autorizado para cargar evidencia"), 403

        if not row:
            return jsonify(ok=False, error="Acción no encontrada"), 404

        if int(row["accion_activa"] or 0) != 1:
            return jsonify(ok=False, error="La acción no está activa"), 400

        if int(row["cumplido"] or 0) == 1:
            return jsonify(ok=False, error="No se puede cargar evidencia en una acción ya cumplida"), 400

        f = request.files.get("file")
        if not f:
            return jsonify(ok=False, error="No se recibió archivo"), 400

        original_name = (f.filename or "").strip()
        if not original_name:
            return jsonify(ok=False, error="El archivo no tiene nombre válido"), 400

        safe_name = secure_filename(original_name)
        ext = os.path.splitext(safe_name)[1].lower()
        physical_name = f"{accion_id}_{uuid4().hex}{ext}"

        folder = _sponsor_acciones_upload_dir()
        full_path = os.path.join(folder, physical_name)
        f.save(full_path)

        try:
            size_bytes = os.path.getsize(full_path)
        except Exception:
            size_bytes = 0

        db.execute("""
            INSERT INTO reclamo_accion_evidencias (
                accion_id,
                filename,
                original_name,
                content_type,
                size_bytes,
                creado_por,
                created_at,
                activo
            )
            VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, 1)
        """, (
            accion_id,
            physical_name,
            original_name,
            (f.mimetype or ""),
            int(size_bytes or 0),
            uid
        ))
        evidencia_id = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]

        db.commit()

        return jsonify(ok=True, item={
            "id": evidencia_id,
            "accion_id": accion_id,
            "filename": physical_name,
            "original_name": original_name,
            "content_type": (f.mimetype or ""),
            "size_bytes": int(size_bytes or 0),
            "download_url": url_for("reclamo_imputado_accion_evidencia_download", evidencia_id=evidencia_id)
        })


    @app.route('/reclamos/imputado-acciones/evidencias/<int:evidencia_id>/eliminar', methods=['POST'])
    @require_login
    def reclamo_imputado_accion_evidencia_eliminar(evidencia_id):
        db = get_db()
        ensure_reclamos_schema(db)
        ensure_reclamo_imputado_acciones_schema(db)

        uid = _current_user_id()

        row = db.execute("""
            SELECT
                e.id,
                e.accion_id,
                e.filename,
                COALESCE(e.activo, 1) AS evidencia_activa,
                COALESCE(a.cumplido, 0) AS accion_cumplida,
                a.imputacion_id,
                a.reclamo_id
            FROM reclamo_accion_evidencias e
            JOIN reclamo_imputado_acciones a
            ON a.id = e.accion_id
            WHERE e.id = ?
            LIMIT 1
        """, (evidencia_id,)).fetchone()

        if not row:
            return jsonify(ok=False, error="Evidencia no encontrada"), 404

        permitido, _ = _puede_gestionar_imputado_accion(int(row["accion_id"]), uid)
        if not permitido:
            return jsonify(ok=False, error="No autorizado para eliminar esta evidencia"), 403

        if int(row["evidencia_activa"] or 0) != 1:
            return jsonify(ok=False, error="La evidencia ya fue eliminada"), 400

        if int(row["accion_cumplida"] or 0) == 1:
            return jsonify(ok=False, error="No se puede eliminar evidencia de una acción ya cumplida"), 400

        db.execute("""
            UPDATE reclamo_accion_evidencias
            SET activo = 0
            WHERE id = ?
        """, (evidencia_id,))

        db.commit()
        return jsonify(ok=True, msg="Evidencia eliminada")


    @app.route('/reclamos/imputado-acciones/evidencias/<int:evidencia_id>/download', methods=['GET'])
    @require_login
    def reclamo_imputado_accion_evidencia_download(evidencia_id):
        db = get_db()
        ensure_reclamos_schema(db)
        ensure_reclamo_imputado_acciones_schema(db)

        uid = _current_user_id()

        row = db.execute("""
            SELECT
                e.id,
                e.accion_id,
                e.filename,
                e.original_name,
                COALESCE(e.content_type, 'application/octet-stream') AS content_type,
                COALESCE(e.activo, 1) AS activo,
                a.imputacion_id,
                a.reclamo_id
            FROM reclamo_accion_evidencias e
            JOIN reclamo_imputado_acciones a
            ON a.id = e.accion_id
            WHERE e.id = ?
            LIMIT 1
        """, (evidencia_id,)).fetchone()

        if not row:
            return jsonify(ok=False, error="Evidencia no encontrada"), 404

        permitido = False

        if _is_admin_like():
            permitido = True
        else:
            permitido_accion, _ = _puede_gestionar_imputado_accion(int(row["accion_id"]), uid)
            if permitido_accion:
                permitido = True
            elif _puede_ver_equipo(int(row["reclamo_id"]), uid):
                permitido = True

        if not permitido:
            return jsonify(ok=False, error="No autorizado para descargar esta evidencia"), 403

        if int(row["activo"] or 0) != 1:
            return jsonify(ok=False, error="La evidencia no está disponible"), 404

        full_path = os.path.join(_sponsor_acciones_upload_dir(), row["filename"])
        if not os.path.exists(full_path):
            return jsonify(ok=False, error="Archivo no encontrado en disco"), 404

        return send_file(
            full_path,
            as_attachment=True,
            download_name=row["original_name"],
            mimetype=row["content_type"] or "application/octet-stream"
        )





    @app.route('/reclamos/ia/mejorar_respuesta', methods=['POST'])
    @require_login
    def reclamos_ia_mejorar_respuesta():
        import json

        current_app.logger.info(">>> IA endpoint llamado")

        data = request.get_json(silent=True)
        current_app.logger.info(">>> JSON recibido: %s", data)
        current_app.logger.info(">>> OPENAI_API_KEY presente: %s", bool(os.getenv("OPENAI_API_KEY")))

        if not data:
            return jsonify(ok=False, msg="No llegó JSON al backend"), 400

        try:
            client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        except Exception as e:
            current_app.logger.exception("Error cargando cliente OpenAI")
            return jsonify(ok=False, msg=f"Error configurando OpenAI: {e}"), 500

        observacion = (data.get("observacion") or "").strip()
        metodo = (data.get("metodo_analisis") or "FISHBONE").strip().upper()

        fish_metodo = (data.get("fish_metodo") or "").strip()
        fish_maquinas = (data.get("fish_maquinas") or "").strip()
        fish_materiales = (data.get("fish_materiales") or "").strip()
        fish_personas = (data.get("fish_personas") or "").strip()
        fish_entorno = (data.get("fish_entorno") or "").strip()
        fish_medicion = (data.get("fish_medicion") or "").strip()

        why1 = (data.get("why1") or "").strip()
        why2 = (data.get("why2") or "").strip()
        why3 = (data.get("why3") or "").strip()
        why4 = (data.get("why4") or "").strip()
        why5 = (data.get("why5") or "").strip()

        causas = data.get("causas") or []
        control = data.get("control") or []
        correctiva = data.get("correctiva") or []

        def _norm_items(items):
            out = []
            if not isinstance(items, list):
                return out

            for it in items:
                if not isinstance(it, dict):
                    continue

                desc = (it.get("descripcion") or "").strip()
                fecha = (it.get("fecha_compromiso") or "").strip()

                if desc or fecha:
                    out.append({
                        "descripcion": desc,
                        "fecha_compromiso": fecha
                    })
            return out

        causas_n = _norm_items(causas)
        control_n = _norm_items(control)
        correctiva_n = _norm_items(correctiva)

        if (
            not observacion and
            not fish_metodo and not fish_maquinas and not fish_materiales and
            not fish_personas and not fish_entorno and not fish_medicion and
            not why1 and not why2 and not why3 and not why4 and not why5 and
            not causas_n and not control_n and not correctiva_n
        ):
            return jsonify(ok=False, msg="No hay información suficiente para mejorar."), 400

        # =========================
        # Prompt + schema por método
        # =========================
        if metodo == "5WHYS":
            prompt = f"""
    Eres especialista en calidad, reclamos y oportunidades de mejora.

    Mejora la redacción técnica del análisis 5 porqués y de las acciones propuestas.
    Reglas:
    - redacta en tono formal y claro
    - evita culpar personas
    - enfócate en procesos y mejora
    - no inventes hechos
    - debes devolver obligatoriamente why1, why2, why3, why4 y why5
    - si ya existen causas, acciones de control o correctivas, mejóralas
    - si no existen, propón una redacción inicial coherente con la observación y el análisis
    - conserva fecha_compromiso solo si ya viene informada

    OBSERVACIÓN:
    {observacion}

    5 POR QUÉS ACTUALES:
    1. {why1}
    2. {why2}
    3. {why3}
    4. {why4}
    5. {why5}

    CAUSAS ACTUALES:
    {json.dumps(causas_n, ensure_ascii=False)}

    ACCIONES DE CONTROL ACTUALES:
    {json.dumps(control_n, ensure_ascii=False)}

    ACCIONES CORRECTIVAS ACTUALES:
    {json.dumps(correctiva_n, ensure_ascii=False)}
    """

            schema = {
                "name": "respuesta_om_mejorada_5whys",
                "strict": True,
                "schema": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "why1": {"type": "string"},
                        "why2": {"type": "string"},
                        "why3": {"type": "string"},
                        "why4": {"type": "string"},
                        "why5": {"type": "string"},
                        "causas": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "descripcion": {"type": "string"},
                                    "fecha_compromiso": {"type": "string"}
                                },
                                "required": ["descripcion", "fecha_compromiso"]
                            }
                        },
                        "control": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "descripcion": {"type": "string"},
                                    "fecha_compromiso": {"type": "string"}
                                },
                                "required": ["descripcion", "fecha_compromiso"]
                            }
                        },
                        "correctiva": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "descripcion": {"type": "string"},
                                    "fecha_compromiso": {"type": "string"}
                                },
                                "required": ["descripcion", "fecha_compromiso"]
                            }
                        },
                        "resumen": {"type": "string"}
                    },
                    "required": [
                        "why1", "why2", "why3", "why4", "why5",
                        "causas", "control", "correctiva", "resumen"
                    ]
                }
            }

        else:
            prompt = f"""
    Eres especialista en calidad, reclamos y oportunidades de mejora.

    Mejora la redacción técnica del análisis fishbone y de las acciones propuestas.
    Reglas:
    - redacta en tono formal y claro
    - evita culpar personas
    - enfócate en procesos y mejora
    - no inventes hechos
    - debes devolver obligatoriamente fish_metodo, fish_maquinas, fish_materiales, fish_personas, fish_entorno y fish_medicion
    - si ya existen causas, acciones de control o correctivas, mejóralas
    - si no existen, propón una redacción inicial coherente con la observación y el análisis
    - conserva fecha_compromiso solo si ya viene informada

    OBSERVACIÓN:
    {observacion}

    FISHBONE ACTUAL:
    - Método/Proceso: {fish_metodo}
    - Máquinas/Equipos: {fish_maquinas}
    - Materiales/Insumos: {fish_materiales}
    - Personas: {fish_personas}
    - Entorno: {fish_entorno}
    - Medición/Información: {fish_medicion}

    CAUSAS ACTUALES:
    {json.dumps(causas_n, ensure_ascii=False)}

    ACCIONES DE CONTROL ACTUALES:
    {json.dumps(control_n, ensure_ascii=False)}

    ACCIONES CORRECTIVAS ACTUALES:
    {json.dumps(correctiva_n, ensure_ascii=False)}
    """

            schema = {
                "name": "respuesta_om_mejorada_fishbone",
                "strict": True,
                "schema": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "fish_metodo": {"type": "string"},
                        "fish_maquinas": {"type": "string"},
                        "fish_materiales": {"type": "string"},
                        "fish_personas": {"type": "string"},
                        "fish_entorno": {"type": "string"},
                        "fish_medicion": {"type": "string"},
                        "causas": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "descripcion": {"type": "string"},
                                    "fecha_compromiso": {"type": "string"}
                                },
                                "required": ["descripcion", "fecha_compromiso"]
                            }
                        },
                        "control": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "descripcion": {"type": "string"},
                                    "fecha_compromiso": {"type": "string"}
                                },
                                "required": ["descripcion", "fecha_compromiso"]
                            }
                        },
                        "correctiva": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "descripcion": {"type": "string"},
                                    "fecha_compromiso": {"type": "string"}
                                },
                                "required": ["descripcion", "fecha_compromiso"]
                            }
                        },
                        "resumen": {"type": "string"}
                    },
                    "required": [
                        "fish_metodo", "fish_maquinas", "fish_materiales",
                        "fish_personas", "fish_entorno", "fish_medicion",
                        "causas", "control", "correctiva", "resumen"
                    ]
                }
            }

        try:
            completion = client.chat.completions.create(
                model="gpt-5-mini",
                messages=[
                    {
                        "role": "developer",
                        "content": "Responde únicamente con JSON válido que cumpla exactamente el esquema indicado."
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                response_format={
                    "type": "json_schema",
                    "json_schema": schema
                }
            )

            raw = (completion.choices[0].message.content or "").strip()
            current_app.logger.info(">>> RAW OPENAI: %s", raw)

            result = json.loads(raw)

            causas_out = _norm_items(result.get("causas") or [])
            control_out = _norm_items(result.get("control") or [])
            correctiva_out = _norm_items(result.get("correctiva") or [])
            resumen_out = (result.get("resumen") or "").strip()

            payload = {
                "ok": True,
                "causas": causas_out,
                "control": control_out,
                "correctiva": correctiva_out,
                "resumen": resumen_out,
                "raw": raw
            }

            if metodo == "5WHYS":
                payload.update({
                    "why1": (result.get("why1") or ""),
                    "why2": (result.get("why2") or ""),
                    "why3": (result.get("why3") or ""),
                    "why4": (result.get("why4") or ""),
                    "why5": (result.get("why5") or "")
                })
            else:
                payload.update({
                    "fish_metodo": (result.get("fish_metodo") or ""),
                    "fish_maquinas": (result.get("fish_maquinas") or ""),
                    "fish_materiales": (result.get("fish_materiales") or ""),
                    "fish_personas": (result.get("fish_personas") or ""),
                    "fish_entorno": (result.get("fish_entorno") or ""),
                    "fish_medicion": (result.get("fish_medicion") or "")
                })

            current_app.logger.info(">>> IA payload final: %s", payload)
            return jsonify(**payload)

        except Exception as e:
            msg = str(e)

            if "insufficient_quota" in msg:
                return jsonify(
                    ok=False,
                    msg="La IA no está disponible actualmente porque la cuenta API no tiene saldo suficiente."
                ), 400

            current_app.logger.exception("Error IA mejorar respuesta")
            return jsonify(ok=False, msg=f"No se pudo mejorar el texto con IA: {e}"), 500
    
 
 