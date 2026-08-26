# modules/planificador/routes_planificador.py
# -*- coding: utf-8 -*-

import csv
import io
import urllib.request
import urllib.parse
from datetime import date, datetime, timedelta
from io import BytesIO

from flask import (
    Blueprint, render_template, request, redirect, url_for,
    flash, session, abort, jsonify, current_app,
)

from modules.auth.routes_auth import require_login, require_permission
from . import planificador_repository as repo
from . import planificador_services as svc
from . import planificador_notifications as notif
from .planificador_constants import (
    ACTIVE_KEY, PERM_SOLICITUDES, PERM_CONFIG, PERM_PRESUPUESTO, PERM_INDICADORES,
    ESTADOS, PRIORIDADES,
    ROL_COORDINADOR, ROL_APROBADOR, ROL_MOTORIZADO, ROL_GERENTE_PRESUPUESTO,
    ESTADOS_RESERVADAS, ESTADOS_COORDINADAS, ESTADOS_POR_COMPLETAR,
    ESTADOS_CONFIRMACION_VOUCHER, ESTADOS_ATENDIDAS,
    MOTIVO_VUELO_OTROS, ROLES_CANDIDATOS_AUTOAPROBAR_VUELO,
    SEMAFORO_AMARILLO_PCT,
)
from flask import Response

planificador_bp = Blueprint("planificador", __name__, url_prefix="/planificador")


def _current_user():
    return {
        "id":     session.get("usuario_id"),
        "nombre": session.get("usuario", ""),
        "rol":    session.get("rol", "usuario"),
    }


# ─────────────────────────────────────────────────────────────
# Pantalla principal: Solicitudes
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes", endpoint="planificador_solicitudes")
@require_login
@require_permission(PERM_SOLICITUDES, "ver")
def solicitudes():
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    current_app.logger.info("[PLANIF] usuario_id=%s nombre=%s", u["id"], u.get("nombre"))
    tipos_solicitud = repo.get_tipos_solicitud()

    filters = {
        "estado":      request.args.get("estado", ""),
        "tipo":        request.args.get("tipo", ""),
        "area":        request.args.get("area", ""),
        "fecha_desde": request.args.get("fecha_desde", ""),
        "fecha_hasta": request.args.get("fecha_hasta", ""),
    }

    solicitudes_list = svc.get_solicitudes_for_user(u["id"], u["rol"], filters)
    current_app.logger.info(
        "[PLANIF] total=%d solicitudes: %s",
        len(solicitudes_list),
        [(r.get("id"), r.get("tipo"), r.get("estado")) for r in solicitudes_list],
    )
    departamentos    = repo.get_departamentos()
    usuario_dept     = repo.get_usuario_departamento(u["id"])

    # Enriquecer cada fila con banderas de acciones
    rows = []
    for s in solicitudes_list:
        d = dict(s)
        d["puede_coordinar"]          = svc.puede_coordinar(s, u["id"], ctx)
        d["puede_aprobar"]            = svc.puede_aprobar(s, u["id"], ctx)
        d["puede_aprobar_gerente"]    = svc.puede_aprobar_gerente(s, u["id"], ctx)
        d["puede_completar"]          = svc.puede_completar(s, u["id"], ctx)
        d["puede_reagendar"]          = svc.puede_reagendar(s, u["id"], ctx)
        d["puede_eliminar"]           = svc.puede_eliminar(s, u["id"], ctx)
        d["puede_aprobar_jefe_vuelo"] = svc.puede_aprobar_jefe_vuelo(s, u["id"], ctx)
        d["puede_cotizar_vuelo"]      = svc.puede_cotizar_vuelo(s, u["id"], ctx)
        d["puede_aprobar_gg_vuelo"]   = svc.puede_aprobar_gg_vuelo(s, u["id"], ctx)
        d["puede_completar_vuelo"]    = svc.puede_completar_vuelo(s, u["id"], ctx)
        d["puede_aprobar_jefe_voucher"] = svc.puede_aprobar_jefe_voucher(s, u["id"], ctx)
        d["puede_entregar_voucher"]     = svc.puede_entregar_voucher(s, u["id"], ctx)
        d["estado_label"]    = svc.estado_label(s["estado"])
        d["estado_class"]    = svc.estado_badge_class(s["estado"])
        d["fecha_str"]       = str(s["fecha"]) if s["fecha"] else ""
        rows.append(d)

    # Dividir en secciones
    reservadas, coordinadas, por_completar, confirmacion_voucher, atendidas = svc.agrupar_por_seccion(rows)
    current_app.logger.info(
        "[PLANIF] secciones reservadas=%d coordinadas=%d por_completar=%d "
        "confirmacion_voucher=%d atendidas=%d",
        len(reservadas), len(coordinadas), len(por_completar),
        len(confirmacion_voucher), len(atendidas),
    )
    current_app.logger.info(
        "[PLANIF] coordinadas detalle=%s",
        [(r.get("id"), r.get("tipo"), r.get("estado")) for r in coordinadas],
    )
    por_aprobar = [r for r in rows if r.get("puede_aprobar_gerente")
                   or r.get("puede_aprobar_jefe_vuelo")
                   or r.get("puede_aprobar_gg_vuelo")
                   or r.get("puede_aprobar_jefe_voucher")]

    # Enriquecer con jefe_nombre para todas las filas, cualquier tipo
    con_jefe_sol_ids = list({
        r["solicitante_id"] for r in rows if r.get("solicitante_id")
    })
    jefe_map = repo.get_jefe_nombre_batch(con_jefe_sol_ids) if con_jefe_sol_ids else {}
    for r in rows:
        r["jefe_nombre"] = jefe_map.get(r.get("solicitante_id"), "")

    # Datos de calendario: semana actual ±2 semanas
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    cal_desde = week_start - timedelta(weeks=4)
    cal_hasta = week_start + timedelta(weeks=8)
    cal_rows = repo.get_calendar_solicitudes(str(cal_desde), str(cal_hasta))

    puede_ver_detalle = svc.puede_ver_detalle_completo(ctx)

    cal_events = []
    for c in cal_rows:
        ev = {
            "id":     c["id"],
            "tipo":   c["tipo"],
            "fecha":  str(c["fecha"]) if c["fecha"] else "",
            "hi":     c["hora_inicio"] or "",
            "hf":     c["hora_fin"] or "",
            "estado": c["estado"],
        }
        if puede_ver_detalle:
            ev["area"]         = c["area_solicitante"]
            ev["descripcion"]  = c["descripcion"]
            ev["lugar"]        = c["lugar_destino"]
            ev["solicitante"]  = c["solicitante_nombre"]
        else:
            ev["area"]        = "Ocupado"
            ev["descripcion"] = ""
            ev["lugar"]       = ""
            ev["solicitante"] = ""
        cal_events.append(ev)

    return render_template(
        "planificador/solicitudes.html",
        active_page=ACTIVE_KEY,
        rows=rows,
        reservadas=reservadas,
        coordinadas=coordinadas,
        por_completar=por_completar,
        confirmacion_voucher=confirmacion_voucher,
        atendidas=atendidas,
        por_aprobar=por_aprobar,
        filters=filters,
        tipos=tipos_solicitud,
        motivos_vuelo=repo.get_motivos_vuelo(),
        estados=ESTADOS,
        prioridades=PRIORIDADES,
        ctx=ctx,
        puede_ver_detalle=puede_ver_detalle,
        cal_events=cal_events,
        today=str(today),
        departamentos=departamentos,
        usuario_dept=usuario_dept,
    )


# ─────────────────────────────────────────────────────────────
# AJAX: Detalle de solicitud
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/detalle", endpoint="planificador_detalle")
@require_login
def detalle(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s:
        abort(404)

    # Solo puede ver detalle completo quien tenga acceso
    es_jefe_de_solicitud = s.get("gerente_id") == u["id"]
    if (not svc.puede_ver_detalle_completo(ctx)
            and s["solicitante_id"] != u["id"]
            and not es_jefe_de_solicitud):
        abort(403)

    d = dict(s)
    d["puede_coordinar"]         = svc.puede_coordinar(s, u["id"], ctx)
    d["puede_aprobar"]           = svc.puede_aprobar(s, u["id"], ctx)
    d["puede_aprobar_gerente"]   = svc.puede_aprobar_gerente(s, u["id"], ctx)
    d["puede_completar"]         = svc.puede_completar(s, u["id"], ctx)
    d["puede_reagendar"]         = svc.puede_reagendar(s, u["id"], ctx)
    d["puede_eliminar"]          = svc.puede_eliminar(s, u["id"], ctx)
    d["puede_aprobar_jefe_vuelo"] = svc.puede_aprobar_jefe_vuelo(s, u["id"], ctx)
    d["puede_cotizar_vuelo"]      = svc.puede_cotizar_vuelo(s, u["id"], ctx)
    d["puede_aprobar_gg_vuelo"]   = svc.puede_aprobar_gg_vuelo(s, u["id"], ctx)
    d["puede_completar_vuelo"]         = svc.puede_completar_vuelo(s, u["id"], ctx)
    d["puede_marcar_realizado_vuelo"]  = svc.puede_marcar_realizado_vuelo(s, u["id"], ctx)
    d["puede_liquidar_vuelo"]          = svc.puede_liquidar_vuelo(s, u["id"], ctx)
    d["puede_aprobar_jefe_voucher"]    = svc.puede_aprobar_jefe_voucher(s, u["id"], ctx)
    d["puede_entregar_voucher"]        = svc.puede_entregar_voucher(s, u["id"], ctx)
    d["estado_label"]    = svc.estado_label(s["estado"])
    d["estado_class"]    = svc.estado_badge_class(s["estado"])
    d["fecha_str"]       = str(s["fecha"]) if s["fecha"] else ""
    logs = repo.get_solicitud_logs(sid)

    grupo_solicitudes = []
    if d.get("grupo_id"):
        todas = repo.get_solicitudes_del_grupo(d["grupo_id"])
        grupo_solicitudes = [dict(g) for g in todas if g["id"] != sid]

    adjuntos = repo.get_adjuntos(sid)
    _es_solicitante = s["solicitante_id"] == u["id"]
    _estado_bloqueado_general = s["estado"] in ("COMPLETADA", "RECHAZADA",
                                                "PENDIENTE_APROBACION_JEFE",
                                                "PENDIENTE_APROBACION_GG_VUELO")
    # Para Vuelo en PENDIENTE_INFO_VUELO el adjunto va dentro del form de completar vuelo
    _es_vuelo_pendiente_coord = (s["tipo"] == "Vuelo" and s["estado"] == "PENDIENTE_INFO_VUELO")
    # Voucher en PENDIENTE_LIQUIDACION_VOUCHER: ya pasó de manos del solicitante
    # al coordinador (que registra los costos); el solicitante no debe poder
    # seguir subiendo adjuntos en esta etapa (admin y coordinador sí, por las
    # otras ramas de este OR).
    _es_voucher_pendiente_liquidacion = (
        s["tipo"] == "Voucher" and s["estado"] == "PENDIENTE_LIQUIDACION_VOUCHER"
    )
    puede_subir_adjunto = (
        not _es_vuelo_pendiente_coord
        and (
            ctx["es_admin"]
            or (
                not _estado_bloqueado_general
                and (ctx["es_gerente"] or ctx["tipos_coordinador"] or ctx["tipos_aprobador"])
            )
            or (
                _es_solicitante
                and s["estado"] not in ("COMPLETADA", "RECHAZADA")
                and not _es_voucher_pendiente_liquidacion
            )
        )
    )
    puede_eliminar_adjunto = ctx["es_admin"] or s["estado"] not in ("APROBADA", "COMPLETADA")

    cc_nombre   = repo.get_cc_nombre(d.get("centro_costo_id")) if d.get("centro_costo_id") else None
    tipos_gasto = repo.get_tipos_gasto() if d.get("puede_liquidar_vuelo") else []

    # Al liquidar, sugerir el valor del "Ticket aéreo" con lo que cotizó el coordinador
    costo_ticket_sugerido = None
    if d.get("puede_liquidar_vuelo") and d.get("datos_ticket"):
        import re
        m = re.search(r"\d+(?:[.,]\d+)?", str(d["datos_ticket"]))
        if m:
            costo_ticket_sugerido = m.group(0).replace(",", ".")

    # Aeropuerto: vuelo_completar() lo antepone como "Aeropuerto: XXX" en observacion_coordinador
    aeropuerto_display = None
    if d.get("tipo") == "Vuelo" and d.get("observacion_coordinador"):
        import re
        m = re.search(r"Aeropuerto:\s*(\S+)", str(d["observacion_coordinador"]))
        if m:
            aeropuerto_display = m.group(1)

    voucher_items = []
    if d.get("tipo") == "Voucher":
        voucher_items = repo.get_voucher_items(sid)
        for item in voucher_items:
            item["puede_confirmar"] = svc.puede_confirmar_voucher_item(s, item, u["id"], ctx)
            item["puede_liquidar"]  = svc.puede_liquidar_voucher_item(s, item, u["id"], ctx)
            # El coordinador puede confirmar/marcar-no-utilizado en nombre del
            # usuario (p. ej. si no está disponible); si el que está viendo
            # el formulario no es el propio solicitante, se lo dejamos claro.
            item["accion_por_coordinador"] = (
                item["puede_confirmar"] and s.get("solicitante_id") != u["id"]
            )

    # Presupuesto anual del CC del solicitante (Ticket aéreo), para las
    # pantallas de cotización (coordinador) y aprobación (Gerente General)
    # de Vuelo — les da contexto para decidir si conviene aprobar o no.
    presupuesto_cc = None
    monto_gg = None
    pasaje_gg = None
    hospedaje_gg = None
    datos_ticket_detalle = None
    fecha_cotizacion_fmt = None

    if d.get("puede_aprobar_gg_vuelo") and d.get("datos_ticket"):
        import re
        ticket_txt = str(d["datos_ticket"])
        m = re.search(r"\d+(?:[.,]\d+)?", ticket_txt)
        pasaje_gg = float(m.group(0).replace(",", ".")) if m else 0.0
        hospedaje_gg = float(d.get("cotizacion_hospedaje") or 0)
        monto_gg = round(pasaje_gg + hospedaje_gg, 2)
        # "Pasaje aéreo" en el desglose ya muestra el monto limpio; el texto
        # libre completo (aerolínea, detalle del vuelo) solo se repite si
        # de verdad trae algo más que el número — evita mostrar "$12.00"
        # dos veces cuando el coordinador solo tecleó el monto.
        resto = re.sub(r"[\d.,\s$]", "", ticket_txt)
        if resto:
            datos_ticket_detalle = ticket_txt

    if d.get("puede_cotizar_vuelo") or d.get("puede_aprobar_gg_vuelo"):
        cc_info = repo.get_cc_usuario(s["solicitante_id"])
        if cc_info:
            anio_vuelo = s["fecha"].year if s.get("fecha") else date.today().year
            saldo = repo.get_saldo_anual_presupuesto(
                cc_info["empresa_id"], cc_info["cc_id"], "Ticket aéreo", anio_vuelo
            )
            if saldo["presupuestado"] > 0:
                if monto_gg is not None:
                    # Vista GG: proyecta el presupuesto como quedaría tras
                    # aprobar este monto (no el saldo actual, sin este vuelo).
                    ejec_proy = saldo["ejecutado"] + monto_gg
                    pct_proy = round(ejec_proy / saldo["presupuestado"] * 100, 1)
                    semaforo_proy = ("rojo" if pct_proy >= 100
                                     else ("amarillo" if pct_proy >= SEMAFORO_AMARILLO_PCT else "verde"))
                    presupuesto_cc = {
                        "presupuestado": saldo["presupuestado"], "ejecutado": ejec_proy,
                        "pct": pct_proy, "semaforo": semaforo_proy,
                    }
                else:
                    presupuesto_cc = saldo

    if d.get("puede_aprobar_gg_vuelo") and d.get("fecha_actualizacion"):
        try:
            _MESES_ABR = ["ene", "feb", "mar", "abr", "may", "jun",
                          "jul", "ago", "sep", "oct", "nov", "dic"]
            _f = d["fecha_actualizacion"]
            if isinstance(_f, str):
                _f = datetime.strptime(_f[:16], "%Y-%m-%d %H:%M")
            fecha_cotizacion_fmt = f"{_f.day} {_MESES_ABR[_f.month - 1]} {_f.year}, {_f.strftime('%H:%M')}"
        except Exception:
            fecha_cotizacion_fmt = None

    return render_template(
        "planificador/_detalle_modal_body.html",
        s=d,
        logs=logs,
        grupo_solicitudes=grupo_solicitudes,
        active_page=ACTIVE_KEY,
        today_iso=date.today().isoformat(),
        adjuntos=adjuntos,
        puede_subir_adjunto=puede_subir_adjunto,
        puede_eliminar_adjunto=puede_eliminar_adjunto,
        cc_nombre=cc_nombre,
        tipos_gasto=tipos_gasto,
        aeropuerto_display=aeropuerto_display,
        costo_ticket_sugerido=costo_ticket_sugerido,
        voucher_items=voucher_items,
        voucher_es_coordinador_view=(d.get("tipo") == "Voucher" and not _es_solicitante),
        presupuesto_cc=presupuesto_cc,
        monto_gg=monto_gg,
        pasaje_gg=pasaje_gg,
        hospedaje_gg=hospedaje_gg,
        datos_ticket_detalle=datos_ticket_detalle,
        fecha_cotizacion_fmt=fecha_cotizacion_fmt,
    )


# ─────────────────────────────────────────────────────────────
# AJAX: Verificar duplicado antes de guardar
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/check-duplicado", methods=["GET"], endpoint="planificador_check_duplicado")
@require_login
def check_duplicado():
    from flask import jsonify
    u = _current_user()
    tipo          = request.args.get("tipo", "").strip()
    fecha         = request.args.get("fecha", "").strip()
    fecha_retorno = request.args.get("fecha_retorno", "").strip() or None
    if not tipo or not fecha:
        return jsonify({"duplicado": False})
    conflicto = repo.check_solicitud_duplicada(u["id"], tipo, fecha, fecha_retorno)
    if conflicto:
        return jsonify({"duplicado": True, "solicitud_id": conflicto["id"],
                        "estado": conflicto["estado"], "fecha": conflicto["fecha_str"],
                        "fecha_retorno": conflicto["fecha_retorno_str"]})
    return jsonify({"duplicado": False})


# ─────────────────────────────────────────────────────────────
# POST: Crear solicitud
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/crear", methods=["POST"], endpoint="planificador_crear")
@require_login
@require_permission(PERM_SOLICITUDES, "crear")
def crear():
    u = _current_user()
    tipo  = request.form.get("tipo", "").strip()
    area  = request.form.get("area_solicitante", "").strip()
    desc  = request.form.get("descripcion", "").strip()
    lugar = request.form.get("lugar_destino", "").strip()
    fecha = request.form.get("fecha", "").strip()

    if tipo == "Vuelo":
        campos_base = [tipo, area, desc, fecha]
    elif tipo == "Voucher":
        campos_base = [tipo, area, desc, fecha]
    else:
        campos_base = [tipo, area, desc, lugar, fecha]
    if not all(campos_base):
        flash("Todos los campos obligatorios deben completarse.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))
    if tipo not in repo.get_tipos_solicitud():
        flash("Tipo de solicitud no válido.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    fecha_retorno    = None
    punto_salida     = None
    punto_destino    = None
    requiere_hosp    = 0
    orden_servicio   = None
    cc_id            = None
    requiere_aprov   = 0
    motivo_vuelo     = None
    numero_vouchers  = None
    voucher_items_data = []

    MAX_VOUCHERS = 6

    if tipo == "Voucher":
        numero_vouchers_raw = request.form.get("numero_vouchers", "").strip()
        try:
            numero_vouchers = int(numero_vouchers_raw)
        except (TypeError, ValueError):
            numero_vouchers = 0
        if numero_vouchers < 1:
            flash("Debe indicar el número de vouchers solicitados (mínimo 1).", "warning")
            return redirect(url_for("planificador.planificador_solicitudes"))
        # Tope de 6 vouchers por solicitud: se recorta en silencio, sin error,
        # ya que el formulario ya limita el campo a 6.
        numero_vouchers = min(numero_vouchers, MAX_VOUCHERS)

        for i in range(1, numero_vouchers + 1):
            v_origen  = request.form.get(f"voucher_origen_{i}", "").strip()
            v_destino = request.form.get(f"voucher_destino_{i}", "").strip()
            if not v_origen or not v_destino:
                flash(f"Debe indicar origen y destino del voucher #{i}.", "warning")
                return redirect(url_for("planificador.planificador_solicitudes"))
            voucher_items_data.append({"origen": v_origen, "destino": v_destino})

        # "lugar_destino" ya no se captura como campo único para Voucher; se
        # deja un resumen legible para listados/notificaciones que lo muestran.
        if len(voucher_items_data) == 1:
            lugar = voucher_items_data[0]["destino"] or voucher_items_data[0]["origen"]
        else:
            lugar = f"{len(voucher_items_data)} vouchers — ver detalle"

        try:
            fecha_dt = date.fromisoformat(fecha)
        except ValueError:
            flash("Fecha inválida.", "warning")
            return redirect(url_for("planificador.planificador_solicitudes"))
        if fecha_dt < date.today():
            flash("La fecha de la solicitud de Voucher no puede ser anterior a hoy.", "warning")
            return redirect(url_for("planificador.planificador_solicitudes"))

    if tipo == "Vuelo":
        motivo_vuelo = request.form.get("motivo_vuelo", "").strip()
        if not motivo_vuelo:
            flash("El motivo de la solicitud de Vuelo es obligatorio.", "warning")
            return redirect(url_for("planificador.planificador_solicitudes"))
        if motivo_vuelo == MOTIVO_VUELO_OTROS and not desc:
            flash("Debe indicar el motivo en el campo Observación.", "warning")
            return redirect(url_for("planificador.planificador_solicitudes"))

        fecha_retorno  = request.form.get("fecha_retorno", "").strip() or None
        punto_salida   = request.form.get("punto_salida", "").strip() or None
        punto_destino  = request.form.get("punto_destino", "").strip() or None
        requiere_hosp  = 1 if request.form.get("requiere_hospedaje") else 0
        orden_servicio = request.form.get("orden_servicio", "").strip() or None

        # Detectar CC del usuario y validar saldo anual de presupuesto (Ticket aéreo)
        import datetime
        cc_info = repo.get_cc_usuario(u["id"])
        if not cc_info:
            flash("No tienes un centro de costo configurado, por lo que no puedes crear "
                  "solicitudes de tipo Vuelo. Selecciona otro tipo de solicitud.", "danger")
            return redirect(url_for("planificador.planificador_solicitudes"))
        cc_id = cc_info["cc_id"]
        anio_actual = datetime.date.today().year
        saldo = repo.get_saldo_anual_presupuesto(
            cc_info["empresa_id"], cc_id, "Ticket aéreo", anio_actual
        )
        if saldo["semaforo"] == "rojo":
            requiere_aprov = 1

    ciudad = repo.get_ciudad_usuario(u["id"])

    # Para Vuelo: el flujo empieza con aprobación del jefe directo, salvo que
    # el rol del solicitante esté configurado para auto-aprobar ese paso
    # (ej. gerentes que hoy no tienen jefe directo, pero podrían tenerlo en
    # el futuro sin que eso deba forzar el flujo de aprobación).
    estado_inicial = "PENDIENTE_COORDINACION"
    jefe_id_vuelo   = None
    jefe_nombre_vuelo = None
    autoaprobado_por_rol = False
    if tipo == "Vuelo":
        roles_autoaprobar = repo.get_roles_autoaprobar_jefe_vuelo()
        autoaprobado_por_rol = svc.debe_autoaprobar_jefe_vuelo(u["rol"], roles_autoaprobar)
        if autoaprobado_por_rol:
            current_app.logger.info(
                "[VUELO] usuario_id=%s rol=%s → autoaprobado por configuración de rol",
                u["id"], u.get("rol")
            )
            estado_inicial = "PENDIENTE_COORDINACION"
        else:
            jefe = repo.get_gerente_del_usuario(u["id"])
            current_app.logger.info(
                "[VUELO] usuario_id=%s nombre=%s → jefe=%s",
                u["id"], u.get("nombre"), jefe
            )
            if jefe:
                jefe_id_vuelo     = jefe["id"]
                jefe_nombre_vuelo = jefe["nombre"]
                estado_inicial    = "PENDIENTE_APROBACION_JEFE"
            else:
                current_app.logger.warning(
                    "[VUELO] Sin jefe_id para usuario %s → cae a PENDIENTE_COORDINACION", u["id"]
                )
                estado_inicial = "PENDIENTE_COORDINACION"

    jefe_id_voucher = None
    jefe_nombre_voucher = None
    autoaprobado_por_rol_voucher = False
    if tipo == "Voucher":
        roles_autoaprobar = repo.get_roles_autoaprobar_jefe_vuelo()
        autoaprobado_por_rol_voucher = svc.debe_autoaprobar_jefe_vuelo(u["rol"], roles_autoaprobar)
        if autoaprobado_por_rol_voucher:
            current_app.logger.info(
                "[VOUCHER] usuario_id=%s rol=%s → autoaprobado por configuración de rol",
                u["id"], u.get("rol")
            )
            estado_inicial = "PENDIENTE_ENTREGA_VOUCHER"
        else:
            jefe = repo.get_gerente_del_usuario(u["id"])
            if jefe:
                jefe_id_voucher     = jefe["id"]
                jefe_nombre_voucher = jefe["nombre"]
                estado_inicial      = "PENDIENTE_APROBACION_JEFE"
            else:
                current_app.logger.warning(
                    "[VOUCHER] Sin jefe_id para usuario %s → cae a PENDIENTE_ENTREGA_VOUCHER", u["id"]
                )
                estado_inicial = "PENDIENTE_ENTREGA_VOUCHER"

    # Prevenir duplicados: solo para Vuelo (rango de fechas)
    if tipo == "Vuelo":
        if repo.check_solicitud_duplicada(u["id"], tipo, fecha, fecha_retorno):
            flash("Ya tienes una solicitud activa de Vuelo en la fecha seleccionada. "
                  "No puedes crear una nueva hasta que la anterior sea completada o rechazada.", "warning")
            return redirect(url_for("planificador.planificador_solicitudes"))

    sid = repo.crear_solicitud({
        "tipo":                           tipo,
        "area_solicitante":               area,
        "descripcion":                    desc,
        "lugar_destino":                  lugar,
        "detalle_direccion":              request.form.get("detalle_direccion", "").strip(),
        "contacto":                       request.form.get("contacto", "").strip(),
        "prioridad":                      request.form.get("prioridad", "Normal"),
        "fecha":                          fecha,
        "estado":                         estado_inicial,
        "solicitante_id":                 u["id"],
        "solicitante_nombre":             u["nombre"],
        "ciudad":                         ciudad,
        "fecha_retorno":                  fecha_retorno,
        "punto_salida":                   punto_salida,
        "punto_destino":                  punto_destino,
        "requiere_hospedaje":             requiere_hosp,
        "orden_servicio":                 orden_servicio,
        "centro_costo_id":                cc_id,
        "requiere_aprobacion_presupuesto": requiere_aprov,
        "gerente_id":                     jefe_id_vuelo or jefe_id_voucher,
        "gerente_nombre":                 jefe_nombre_vuelo or jefe_nombre_voucher,
        "motivo_vuelo":                   motivo_vuelo,
        "numero_vouchers":                numero_vouchers,
    })

    if tipo == "Voucher" and sid:
        try:
            repo.crear_voucher_items(sid, voucher_items_data)
        except Exception:
            current_app.logger.exception("[VOUCHER] Error creando voucher_items sid=%s", sid)

    if tipo == "Vuelo":
        if estado_inicial == "PENDIENTE_APROBACION_JEFE":
            # Notificar al jefe directo para que apruebe
            try:
                notif.notif_vuelo_pendiente_jefe(
                    sid, area, fecha, desc, motivo_vuelo or "—",
                    u["nombre"], jefe_id_vuelo, jefe_nombre_vuelo or "—",
                )
            except Exception:
                pass
            msg = "Solicitud de Vuelo creada. Pendiente de aprobación de su jefe directo."
        else:
            # Autoaprobada por rol, o sin jefe directo configurado: pasa
            # directo a coordinación y hay que notificar al coordinador.
            try:
                aprobador_txt = (
                    f"Auto-aprobado (rol {u.get('rol') or '—'})"
                    if autoaprobado_por_rol
                    else "Auto-aprobado (sin jefe directo configurado)"
                )
                notif.notif_vuelo_aprobada_coordinacion(
                    sid, area, fecha, desc, u["nombre"], aprobador_txt,
                )
            except Exception:
                pass
            if autoaprobado_por_rol:
                msg = "Solicitud de Vuelo creada y auto-aprobada según tu rol. Pasa al coordinador para cotizar."
            else:
                msg = "Solicitud de Vuelo creada. No tiene jefe configurado; pasó a coordinación."
    elif tipo == "Voucher":
        if estado_inicial == "PENDIENTE_APROBACION_JEFE":
            try:
                notif.notif_voucher_pendiente_jefe(
                    sid, area, fecha, desc, u["nombre"],
                    jefe_id_voucher, jefe_nombre_voucher or "—",
                )
            except Exception:
                pass
            msg = "Solicitud de Voucher creada. Pendiente de aprobación de su jefe directo."
        else:
            try:
                aprobador_txt = (
                    f"Auto-aprobado (rol {u.get('rol') or '—'})"
                    if autoaprobado_por_rol_voucher
                    else "Auto-aprobado (sin jefe directo configurado)"
                )
                notif.notif_voucher_aprobada_solicitante(
                    sid, area, fecha, desc, u["id"], u["nombre"], aprobador_txt,
                )
            except Exception:
                pass
            try:
                notif.notif_voucher_pendiente_entrega(
                    sid, area, fecha, desc, u["nombre"], aprobador_txt,
                )
            except Exception:
                pass
            msg = "Solicitud de Voucher creada y aprobada. El coordinador debe entregarte los vouchers."
    else:
        try:
            notif.notif_nueva_solicitud(sid, tipo, area, fecha, u["nombre"], solicitante_id=u["id"])
        except Exception:
            pass
        msg = "Solicitud creada. Queda pendiente de coordinación."

    flash(msg, "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# AJAX: Saldo de presupuesto del usuario (para tipo Vuelo)
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/presupuesto/saldo-usuario", methods=["GET"],
                       endpoint="planificador_saldo_usuario")
@require_login
def presupuesto_saldo_usuario():
    from datetime import date as _date
    u = _current_user()
    cc_info = repo.get_cc_usuario(u["id"])
    if not cc_info:
        return jsonify({"ok": False, "msg": "Sin centro de costo asignado al usuario."})
    hoy = _date.today()
    tipo_gasto = "Ticket aéreo"
    saldo_data = repo.get_saldo_anual_presupuesto(
        cc_info["empresa_id"], cc_info["cc_id"], tipo_gasto, hoy.year
    )
    saldo = saldo_data["presupuestado"] - saldo_data["ejecutado"]
    return jsonify({
        "ok":            True,
        "semaforo":      saldo_data["semaforo"],
        "pct":           saldo_data["pct"],
        "anio":          hoy.year,
        "presupuestado": saldo_data["presupuestado"],
        "ejecutado":     saldo_data["ejecutado"],
        "saldo":         round(saldo, 2),
    })


# ─────────────────────────────────────────────────────────────
# AJAX: Otras solicitudes pendientes del mismo tipo (para agrupar)
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/pendientes-mismo-tipo",
                        endpoint="planificador_pendientes_mismo_tipo")
@require_login
def pendientes_mismo_tipo(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_coordinar(s, u["id"], ctx):
        return jsonify([])
    otros = repo.get_solicitudes_pendientes_mismo_tipo(s["tipo"], sid)
    return jsonify([{
        "id":          r["id"],
        "area":        r["area_solicitante"] or "",
        "lugar":       r["lugar_destino"] or "",
        "fecha":       str(r["fecha"]) if r["fecha"] else "",
        "solicitante": r["solicitante_nombre"] or "",
        "descripcion": (r["descripcion"] or "")[:80],
    } for r in otros])


# ─────────────────────────────────────────────────────────────
# POST: Coordinar
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/coordinar", methods=["POST"],
                       endpoint="planificador_coordinar")
@require_login
def coordinar(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_coordinar(s, u["id"], ctx):
        abort(403)

    hi  = request.form.get("hora_inicio", "").strip()
    hf  = request.form.get("hora_fin", "").strip()
    obs = request.form.get("observacion_coordinador", "").strip()

    if not hi or not hf:
        flash("Debe asignar hora de inicio y hora fin.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    # Fecha pasada → no se puede coordinar, solo reagendar
    if s["fecha"] and str(s["fecha"])[:10] < str(date.today()):
        flash(
            "La fecha de esta solicitud ya pasó. Use la opción 'Reagendar' "
            "para asignarle una nueva fecha antes de coordinar.",
            "warning"
        )
        return redirect(url_for("planificador.planificador_solicitudes"))

    if hf <= hi:
        flash("La hora fin debe ser mayor que la hora de inicio.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    # Horario laboral: 08:00 – 17:00
    if hi < "08:00" or hf > "17:00":
        flash("El horario debe estar dentro del rango laboral: 08:00 – 17:00.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    # Si la fecha es hoy y la hora actual ya pasó las 17:00, no se puede planificar para hoy
    today = date.today()
    if str(s["fecha"]) == str(today) and datetime.now().hour >= 17:
        flash(
            "Ya pasaron las 17:00. No se puede planificar para hoy. "
            "Cambia la fecha de la solicitud al siguiente día hábil.",
            "warning"
        )
        return redirect(url_for("planificador.planificador_solicitudes"))

    # Validar que no haya otro registro del MISMO tipo en ese horario
    if repo.check_horario_ocupado(s["tipo"], str(s["fecha"]), hi, hf, exclude_id=sid):
        flash(
            f"Ya existe una solicitud de tipo «{s['tipo']}» planificada en ese horario. "
            "Elija otro horario o cancele la solicitud existente.",
            "warning"
        )
        return redirect(url_for("planificador.planificador_solicitudes"))

    # ¿Se agrupan otras solicitudes del mismo tipo?
    grupo_ids_raw = request.form.getlist("grupo_ids")
    grupo_ids = []
    for gid_str in grupo_ids_raw:
        try:
            gid = int(gid_str)
            gs = repo.get_solicitud_by_id(gid)
            if gs and gs["tipo"] == s["tipo"] and gs["estado"] == "PENDIENTE_COORDINACION":
                grupo_ids.append(gid)
        except (ValueError, TypeError):
            pass

    if grupo_ids:
        all_ids = [sid] + grupo_ids
        grupo_id = repo.crear_grupo_coordinacion(
            s["tipo"], str(s["fecha"]), hi, hf, u["id"], u["nombre"], obs
        )
        repo.coordinar_solicitudes_grupo(all_ids, grupo_id, u["id"], u["nombre"], hi, hf, obs)
        for gid in all_ids:
            gs = repo.get_solicitud_by_id(gid)
            if gs:
                try:
                    notif.notif_coordinada(gid, gs["tipo"], gs["area_solicitante"],
                                           str(gs["fecha"]), hi, hf, u["nombre"])
                except Exception:
                    pass
        flash(
            f"Se coordinaron {len(all_ids)} solicitudes juntas "
            f"en el horario {hi} – {hf}.",
            "success"
        )
    else:
        repo.coordinar_solicitud(sid, u["id"], u["nombre"], hi, hf, obs)
        try:
            notif.notif_coordinada(sid, s["tipo"], s["area_solicitante"],
                                   str(s["fecha"]), hi, hf, u["nombre"])
        except Exception:
            pass
        flash("Solicitud enviada a aprobación.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# POST: Aprobar
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/aprobar", methods=["POST"],
                       endpoint="planificador_aprobar")
@require_login
def aprobar(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_aprobar(s, u["id"], ctx):
        abort(403)

    obs = request.form.get("observacion_aprobador", "").strip()
    repo.aprobar_solicitud(sid, u["id"], u["nombre"], obs)

    # ¿Requiere aprobación gerencial?
    flags = repo.get_tipo_flags(s["tipo"])
    if flags.get("requiere_aprobacion_gerente"):
        gerente = repo.get_gerente_del_usuario(s["solicitante_id"])
        if gerente:
            repo.poner_pendiente_gerente(sid, gerente["id"], gerente["nombre"])
            try:
                notif.notif_pendiente_gerente(
                    sid, s["tipo"], s["area_solicitante"],
                    str(s["fecha"]), s["hora_inicio"] or "", s["hora_fin"] or "",
                    s["lugar_destino"], s["descripcion"],
                    gerente["id"], gerente["nombre"], u["nombre"],
                )
            except Exception:
                pass
            flash("Solicitud aprobada. Pendiente de aprobación gerencial.", "info")
            return redirect(url_for("planificador.planificador_solicitudes"))

    try:
        notif.notif_aprobada(
            sid, s["tipo"], s["area_solicitante"],
            str(s["fecha"]), s["hora_inicio"] or "", s["hora_fin"] or "",
            s["lugar_destino"], s["descripcion"],
            s["solicitante_id"], u["nombre"],
        )
    except Exception:
        pass
    flash("Solicitud aprobada. Aparecerá en el calendario.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# POST: Aprobar todas las solicitudes de un grupo
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/grupo/<int:grupo_id>/aprobar", methods=["POST"],
                        endpoint="planificador_aprobar_grupo")
@require_login
def aprobar_grupo(grupo_id):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])

    miembros = repo.get_solicitudes_del_grupo(grupo_id)
    if not miembros:
        flash("Grupo no encontrado o sin solicitudes.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    obs = request.form.get("observacion_aprobador", "").strip()
    aprobadas = 0

    for m in miembros:
        if not svc.puede_aprobar(m, u["id"], ctx):
            continue
        repo.aprobar_solicitud(m["id"], u["id"], u["nombre"], obs)
        aprobadas += 1
        try:
            notif.notif_aprobada(
                m["id"], m["tipo"], m["area_solicitante"],
                str(m["fecha"]), m["hora_inicio"] or "", m["hora_fin"] or "",
                m["lugar_destino"], "",
                m["solicitante_id"], u["nombre"],
            )
        except Exception:
            pass

    if aprobadas == 0:
        flash("No tienes permiso para aprobar ninguna solicitud de este grupo.", "warning")
    else:
        flash(f"Se aprobaron {aprobadas} solicitudes del grupo #{grupo_id}.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# POST: Aprobar como gerente
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/aprobar-gerente", methods=["POST"],
                       endpoint="planificador_aprobar_gerente")
@require_login
def aprobar_gerente(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_aprobar_gerente(s, u["id"], ctx):
        abort(403)

    obs = request.form.get("observacion_aprobador", "").strip()
    repo.aprobar_por_gerente(sid, u["id"], u["nombre"], obs)
    try:
        notif.notif_aprobada(
            sid, s["tipo"], s["area_solicitante"],
            str(s["fecha"]), s["hora_inicio"] or "", s["hora_fin"] or "",
            s["lugar_destino"], s["descripcion"],
            s["solicitante_id"], u["nombre"],
        )
    except Exception:
        pass
    flash("Solicitud aprobada gerencialmente. Aparecerá en el calendario.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


@planificador_bp.route("/solicitudes/<int:sid>/rechazar-gerente", methods=["POST"],
                       endpoint="planificador_rechazar_gerente")
@require_login
def rechazar_gerente(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_aprobar_gerente(s, u["id"], ctx):
        abort(403)

    obs = request.form.get("observacion_aprobador", "").strip()
    if not obs:
        flash("Para rechazar debe ingresar una observación.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    repo.rechazar_por_gerente(sid, u["id"], u["nombre"], obs)
    try:
        notif.notif_rechazada(sid, s["tipo"], str(s["fecha"]),
                              obs, s["solicitante_id"], u["nombre"])
    except Exception:
        pass
    flash("Solicitud rechazada.", "info")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# POST: Rechazar
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/rechazar", methods=["POST"],
                       endpoint="planificador_rechazar")
@require_login
def rechazar(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_aprobar(s, u["id"], ctx):
        abort(403)

    obs = request.form.get("observacion_aprobador", "").strip()
    if not obs:
        flash("Para rechazar debe ingresar una observación.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    repo.rechazar_solicitud(sid, u["id"], u["nombre"], obs)
    try:
        notif.notif_rechazada(sid, s["tipo"], str(s["fecha"]),
                              obs, s["solicitante_id"], u["nombre"])
    except Exception:
        pass
    flash("Solicitud rechazada.", "info")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# POST: Completar
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/completar", methods=["POST"],
                       endpoint="planificador_completar")
@require_login
def completar(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_completar(s, u["id"], ctx):
        abort(403)

    repo.completar_solicitud(sid, u["id"], u["nombre"])
    flash("Actividad marcada como completada.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# POST: Reagendar (solo coordinadores / admin)
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/reagendar", methods=["POST"],
                       endpoint="planificador_reagendar")
@require_login
def reagendar(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_reagendar(s, u["id"], ctx):
        abort(403)

    nueva_fecha = request.form.get("nueva_fecha", "").strip()
    motivo      = request.form.get("motivo_reagenda", "").strip()
    es_vuelo    = s["tipo"] == "Vuelo"
    nueva_fecha_retorno = request.form.get("nueva_fecha_retorno", "").strip() or None

    if not nueva_fecha:
        flash("Debe indicar la nueva fecha.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))
    if es_vuelo and not nueva_fecha_retorno:
        flash("Debe indicar la nueva fecha de regreso.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    try:
        if date.fromisoformat(nueva_fecha) < date.today():
            flash("La nueva fecha no puede ser anterior al día de hoy.", "warning")
            return redirect(url_for("planificador.planificador_solicitudes"))
        if es_vuelo and date.fromisoformat(nueva_fecha_retorno) < date.fromisoformat(nueva_fecha):
            flash("La fecha de regreso no puede ser anterior a la de salida.", "warning")
            return redirect(url_for("planificador.planificador_solicitudes"))
    except ValueError:
        flash("Fecha inválida.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    # Guardar fecha anterior para la notificación
    fecha_anterior = str(s["fecha"]) if s["fecha"] else "—"

    penalizacion_raw = request.form.get("penalizacion", "0").strip()
    try:
        penalizacion = float(penalizacion_raw) if penalizacion_raw else 0.0
    except ValueError:
        penalizacion = 0.0

    if es_vuelo:
        repo.reagendar_vuelo_a_jefe(sid, nueva_fecha, nueva_fecha_retorno, u["id"], u["nombre"], motivo)
        if penalizacion > 0:
            repo.set_penalizacion(sid, penalizacion)
        # Verificar si el solicitante tiene rol que auto-aprueba el paso de jefe
        roles_autoaprobar = repo.get_roles_autoaprobar_jefe_vuelo()
        sol_rol = repo.get_rol_usuario(s["solicitante_id"])
        if svc.debe_autoaprobar_jefe_vuelo(sol_rol, roles_autoaprobar):
            repo.aprobar_jefe_vuelo(sid, s["solicitante_id"], s["solicitante_nombre"],
                                    "Auto-aprobado (solicitante es gerente)")
            flash(f"Vuelo reagendado para el {nueva_fecha}. Auto-aprobado por rol del solicitante.", "success")
        else:
            # Renotificar al jefe para que vuelva a aprobar con la nueva fecha
            try:
                jefe_id   = s.get("gerente_id")
                jefe_nom  = s.get("gerente_nombre", "—")
                notif.notif_vuelo_pendiente_jefe(
                    sid, s["area_solicitante"], nueva_fecha,
                    s.get("descripcion", ""),
                    s.get("motivo_vuelo") or "—",
                    s["solicitante_nombre"], jefe_id, jefe_nom,
                    es_reagenda=True,
                )
            except Exception:
                pass
            flash(f"Vuelo reagendado para el {nueva_fecha}. Vuelve a aprobación del jefe directo.", "success")
    else:
        repo.reagendar_solicitud(sid, nueva_fecha, u["id"], u["nombre"], motivo, "PENDIENTE_COORDINACION")
        try:
            notif.notif_reagendada(
                sid, s["tipo"], s["area_solicitante"],
                fecha_anterior, nueva_fecha, motivo,
                u["nombre"], s["solicitante_id"],
            )
        except Exception:
            pass
        flash(f"Solicitud reagendada para el {nueva_fecha}. El solicitante fue notificado.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# POST: Eliminar (soft-delete con notificación por email)
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/eliminar", methods=["POST"],
                       endpoint="planificador_eliminar")
@require_login
def eliminar(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_eliminar(s, u["id"], ctx):
        abort(403)

    es_solicitante = (s["solicitante_id"] == u["id"] and
                      not ctx["es_admin"] and
                      not ctx["tipos_coordinador"] and
                      not ctx["tipos_aprobador"])

    repo.insert_solicitud_log(sid, "ELIMINADA", u["id"], u["nombre"],
                              "Solicitud eliminada del sistema.")
    repo.delete_solicitud(sid, eliminado_por_id=u["id"],
                          eliminado_por_nombre=u["nombre"])

    try:
        notif.notif_eliminada(
            sid, s["tipo"], s["area_solicitante"],
            str(s["fecha"]) if s["fecha"] else "—",
            u["nombre"], s["solicitante_id"],
            eliminado_por_es_solicitante=es_solicitante,
        )
    except Exception:
        pass

    flash("Solicitud eliminada.", "info")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# Vuelo: aprobación jefe directo
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/vuelo/aprobar-jefe", methods=["POST"],
                       endpoint="planificador_vuelo_aprobar_jefe")
@require_login
def vuelo_aprobar_jefe(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_aprobar_jefe_vuelo(s, u["id"], ctx):
        abort(403)
    obs = request.form.get("observacion", "").strip()
    repo.aprobar_jefe_vuelo(sid, u["id"], u["nombre"], obs)
    try:
        notif.notif_vuelo_aprobada_coordinacion(
            sid, s["area_solicitante"], str(s["fecha"]),
            s.get("descripcion", ""), s["solicitante_nombre"], u["nombre"],
        )
    except Exception:
        pass
    flash("Vuelo aprobado. Pasa al coordinador para cotizar el pasaje.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


@planificador_bp.route("/solicitudes/<int:sid>/vuelo/rechazar-jefe", methods=["POST"],
                       endpoint="planificador_vuelo_rechazar_jefe")
@require_login
def vuelo_rechazar_jefe(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_aprobar_jefe_vuelo(s, u["id"], ctx):
        abort(403)
    obs = request.form.get("observacion", "").strip()
    if not obs:
        flash("Debe indicar el motivo del rechazo.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))
    repo.rechazar_vuelo(sid, u["id"], u["nombre"], obs)
    try:
        notif.notif_vuelo_rechazada(
            sid, s["area_solicitante"], str(s["fecha"]),
            obs, s["solicitante_nombre"], s["solicitante_id"], u["nombre"],
        )
    except Exception:
        pass
    flash("Solicitud de Vuelo rechazada.", "warning")
    return redirect(url_for("planificador.planificador_solicitudes"))


@planificador_bp.route("/solicitudes/vuelo/aprobar-jefe-masivo", methods=["POST"],
                       endpoint="planificador_vuelo_aprobar_jefe_masivo")
@require_login
def vuelo_aprobar_jefe_masivo():
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    ids_raw = request.form.getlist("ids[]")
    try:
        sids = [int(x) for x in ids_raw if x.strip().isdigit()]
    except Exception:
        sids = []
    if not sids:
        flash("No se seleccionaron solicitudes.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))
    aprobadas = 0
    for sid in sids:
        s = repo.get_solicitud_by_id(sid)
        if not s or not svc.puede_aprobar_jefe_vuelo(s, u["id"], ctx):
            continue
        repo.aprobar_jefe_vuelo(sid, u["id"], u["nombre"], "")
        try:
            notif.notif_vuelo_aprobada_coordinacion(
                sid, s["area_solicitante"], str(s["fecha"]),
                s.get("descripcion", ""), s["solicitante_nombre"], u["nombre"],
            )
        except Exception:
            pass
        aprobadas += 1
    flash(f"{aprobadas} solicitud(es) aprobada(s).", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# Vuelo: coordinador ingresa el valor cotizado del pasaje
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/vuelo/cotizar", methods=["POST"],
                       endpoint="planificador_vuelo_cotizar")
@require_login
def vuelo_cotizar(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_cotizar_vuelo(s, u["id"], ctx):
        abort(403)
    valor = request.form.get("valor_cotizado", "").strip()
    obs   = request.form.get("observacion", "").strip()
    hosp_str = request.form.get("cotizacion_hospedaje", "").strip().replace(",", ".")
    if not valor:
        flash("Debe ingresar el valor cotizado del pasaje.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))
    repo.cotizar_vuelo(sid, u["id"], u["nombre"], valor, obs)
    try:
        hosp_val = float(hosp_str) if hosp_str else 0.0
        if hosp_val > 0:
            repo.set_cotizacion_hospedaje(sid, hosp_val)
    except (ValueError, Exception):
        pass
    try:
        gg_lista = repo.get_gerentes_presupuesto_para_tipo("Vuelo")
        for gg in gg_lista:
            notif.notif_vuelo_pendiente_gg(
                sid, s["area_solicitante"], str(s["fecha"]),
                s.get("descripcion", ""), valor,
                s["solicitante_nombre"], u["nombre"],
                gg.get("id"), gg.get("nombre", "—"),
            )
    except Exception:
        pass
    flash("Cotización registrada. Pasa a aprobación del Gerente General.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# Vuelo: aprobación Gerente de Presupuesto (revisa la cotización)
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/vuelo/aprobar-gg", methods=["POST"],
                       endpoint="planificador_vuelo_aprobar_gg")
@require_login
def vuelo_aprobar_gg(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_aprobar_gg_vuelo(s, u["id"], ctx):
        abort(403)
    obs = request.form.get("observacion", "").strip()
    repo.aprobar_gg_vuelo(sid, u["id"], u["nombre"], obs)
    try:
        notif.notif_vuelo_gg_aprobo_pendiente_info(
            sid, s["area_solicitante"], str(s["fecha"]),
            s.get("descripcion", ""), s["solicitante_nombre"], u["nombre"],
        )
    except Exception:
        pass
    flash("Cotización aprobada. Pasa al coordinador para ingresar la información del vuelo.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


@planificador_bp.route("/solicitudes/<int:sid>/vuelo/rechazar-gg", methods=["POST"],
                       endpoint="planificador_vuelo_rechazar_gg")
@require_login
def vuelo_rechazar_gg(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_aprobar_gg_vuelo(s, u["id"], ctx):
        abort(403)
    obs = request.form.get("observacion", "").strip()
    if not obs:
        flash("Debe indicar el motivo del rechazo.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))
    repo.rechazar_gg_vuelo(sid, u["id"], u["nombre"], obs)
    try:
        notif.notif_vuelo_gg_rechazo_coordinador(
            sid, s["area_solicitante"], str(s["fecha"]),
            obs, s["solicitante_nombre"], u["nombre"],
        )
    except Exception:
        pass
    flash("Cotización rechazada. Vuelve al coordinador para recotizar.", "warning")
    return redirect(url_for("planificador.planificador_solicitudes"))


@planificador_bp.route("/solicitudes/vuelo/aprobar-gg-masivo", methods=["POST"],
                       endpoint="planificador_vuelo_aprobar_gg_masivo")
@require_login
def vuelo_aprobar_gg_masivo():
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    ids_raw = request.form.getlist("ids[]")
    try:
        sids = [int(x) for x in ids_raw if x.strip().isdigit()]
    except Exception:
        sids = []
    if not sids:
        flash("No se seleccionaron solicitudes.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))
    aprobadas = 0
    for sid in sids:
        s = repo.get_solicitud_by_id(sid)
        if not s or not svc.puede_aprobar_gg_vuelo(s, u["id"], ctx):
            continue
        repo.aprobar_gg_vuelo(sid, u["id"], u["nombre"], "")
        try:
            notif.notif_vuelo_gg_aprobo_pendiente_info(
                sid, s["area_solicitante"], str(s["fecha"]),
                s.get("descripcion", ""), s["solicitante_nombre"], u["nombre"],
            )
        except Exception:
            pass
        aprobadas += 1
    flash(f"{aprobadas} solicitud(es) aprobada(s) por GG.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# Vuelo: coordinador registra datos de reserva
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/vuelo/coordinar", methods=["POST"],
                       endpoint="planificador_vuelo_coordinar")
@require_login
def vuelo_coordinar(sid):
    import os, uuid
    from werkzeug.utils import secure_filename
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_coordinar(s, u["id"], ctx):
        abort(403)

    action = request.form.get("action", "coordinar")

    # ── Reprogramar: nueva fecha → vuelve a PENDIENTE_APROBACION_JEFE ──
    if action == "reprogramar":
        nueva_fecha = request.form.get("nueva_fecha", "").strip()
        motivo      = request.form.get("motivo_reprogramacion", "").strip()
        if not nueva_fecha:
            flash("Debe indicar la nueva fecha.", "warning")
            return redirect(url_for("planificador.planificador_solicitudes"))
        if not motivo:
            flash("Debe indicar el motivo de la reprogramación.", "warning")
            return redirect(url_for("planificador.planificador_solicitudes"))
        pen_raw = request.form.get("penalizacion", "0").strip()
        try:
            pen_val = float(pen_raw) if pen_raw else 0.0
        except ValueError:
            pen_val = 0.0
        repo.reagendar_vuelo_a_jefe(sid, nueva_fecha, u["id"], u["nombre"], motivo)
        if pen_val > 0:
            repo.set_penalizacion(sid, pen_val)
        try:
            notif.notif_vuelo_pendiente_jefe(
                sid, s["area_solicitante"], nueva_fecha,
                s.get("descripcion", ""), s.get("motivo_vuelo") or "—",
                s["solicitante_nombre"],
                s.get("gerente_id"), s.get("gerente_nombre", "—"),
            )
        except Exception:
            pass
        flash("Vuelo reprogramado. Vuelve al proceso de aprobación del jefe.", "info")
        return redirect(url_for("planificador.planificador_solicitudes"))

    # ── Coordinar: registrar datos de reserva ──
    hi             = request.form.get("hora_inicio", "").strip()
    hf             = request.form.get("hora_fin", "").strip()
    datos_ticket   = request.form.get("datos_ticket", "").strip()
    datos_hotel    = request.form.get("datos_hotel", "").strip()
    obs            = request.form.get("observacion_coordinador", "").strip()

    if not hi or not hf:
        flash("Debe indicar hora de inicio y hora fin del vuelo.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))
    if not datos_ticket:
        flash("Debe ingresar los datos de la reservación del ticket aéreo.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    # Guardar adjunto si se subió
    attachment_paths = []
    archivo = request.files.get("adjunto_vuelo")
    nombre_original = None
    if archivo and archivo.filename:
        archivo.seek(0, 2)
        tamano = archivo.tell()
        archivo.seek(0)
        if tamano > 10 * 1024 * 1024:
            flash("El archivo supera el límite de 10 MB.", "warning")
            return redirect(url_for("planificador.planificador_solicitudes"))
        nombre_original = archivo.filename
        nombre_seguro   = secure_filename(nombre_original)
        ext             = os.path.splitext(nombre_seguro)[1]
        nombre_guardado = f"{uuid.uuid4().hex}{ext}"
        carpeta = os.path.join(current_app.config["UPLOAD_FOLDER"], "planificador", str(sid))
        os.makedirs(carpeta, exist_ok=True)
        ruta_completa = os.path.join(carpeta, nombre_guardado)
        archivo.save(ruta_completa)
        repo.insert_adjunto(sid, nombre_original, nombre_guardado, tamano, u["id"], u["nombre"])
        attachment_paths = [(ruta_completa, nombre_seguro)]

    repo.coordinar_vuelo(sid, u["id"], u["nombre"], hi, hf, datos_ticket, datos_hotel, obs)
    try:
        notif.notif_vuelo_coordinada_solicitante(
            sid, s["area_solicitante"], str(s["fecha"]),
            s.get("descripcion", ""), hi, hf,
            datos_ticket, datos_hotel,
            s["solicitante_nombre"], s["solicitante_id"],
            u["nombre"], attachment_paths or None,
        )
    except Exception:
        pass
    flash("Datos de reserva registrados. Se notificó al solicitante.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# Vuelo: coordinador registra gestión y completa
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/vuelo/completar", methods=["POST"],
                       endpoint="planificador_vuelo_completar")
@require_login
def vuelo_completar(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_completar_vuelo(s, u["id"], ctx):
        abort(403)
    hora_inicio = request.form.get("hora_inicio", "").strip() or None
    hora_fin    = request.form.get("hora_fin", "").strip() or None
    aeropuerto  = request.form.get("aeropuerto", "").strip()

    partes = []
    if aeropuerto:
        partes.append(f"Aeropuerto: {aeropuerto}")
    obs_base = request.form.get("observacion_coordinador", "").strip()
    if obs_base:
        partes.append(obs_base)
    obs = "\n".join(partes)

    if not obs_base and not aeropuerto:
        flash("Debe ingresar la información de la reservación.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))
    repo.completar_vuelo(sid, u["id"], u["nombre"], obs, hora_inicio, hora_fin)

    # Guardar adjunto del ticket/boleto si se subió
    import os, uuid
    from werkzeug.utils import secure_filename
    archivo = request.files.get("adjunto_vuelo")
    ruta_adjunto = None
    nombre_adjunto = None
    if archivo and archivo.filename:
        archivo.seek(0, 2); tam = archivo.tell(); archivo.seek(0)
        if tam <= 5 * 1024 * 1024:
            nombre_original = secure_filename(archivo.filename)
            ext = os.path.splitext(nombre_original)[1]
            nombre_guardado = f"{uuid.uuid4().hex}{ext}"
            carpeta = os.path.join(current_app.config["UPLOAD_FOLDER"], "planificador", str(sid))
            os.makedirs(carpeta, exist_ok=True)
            ruta_adjunto = os.path.join(carpeta, nombre_guardado)
            archivo.save(ruta_adjunto)
            repo.insert_adjunto(sid, nombre_original, nombre_guardado, tam, u["id"], u["nombre"])
            nombre_adjunto = nombre_original

    # Notificar al solicitante con detalles y archivo adjunto
    try:
        adjuntos_email = [(ruta_adjunto, nombre_adjunto)] if ruta_adjunto else None
        notif.notif_vuelo_completada(
            sid, s["area_solicitante"], str(s["fecha"]),
            s.get("descripcion", ""), obs,
            s["solicitante_nombre"], s["solicitante_id"],
            u["nombre"], adjuntos_email,
        )
    except Exception:
        pass
    flash("Vuelo coordinado. El solicitante fue notificado.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# Vuelo: solicitante confirma que realizó el vuelo → PENDIENTE_LIQUIDACION
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/vuelo/marcar-realizado", methods=["POST"],
                       endpoint="planificador_vuelo_marcar_realizado")
@require_login
def vuelo_marcar_realizado(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_marcar_realizado_vuelo(s, u["id"], ctx):
        abort(403)
    repo.marcar_realizado_vuelo(sid, u["id"], u["nombre"])
    try:
        notif.notif_vuelo_pendiente_liquidacion(
            sid, s["area_solicitante"], str(s["fecha"]),
            s["solicitante_nombre"], s["coordinador_id"],
            s["coordinador_nombre"],
        )
    except Exception:
        pass
    flash("Vuelo marcado como realizado. El coordinador registrará los costos.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# Vuelo: coordinador ingresa costos por tipo → COMPLETADA + deducción presupuesto
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/vuelo/liquidar", methods=["POST"],
                       endpoint="planificador_vuelo_liquidar")
@require_login
def vuelo_liquidar(sid):
    from datetime import date as _date
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_liquidar_vuelo(s, u["id"], ctx):
        abort(403)

    notas = request.form.get("notas_liquidacion", "").strip()

    # Leer costos por tipo de gasto (campos tipo_costo[NombreTipo])
    tipos_gasto = repo.get_tipos_gasto()
    costos_por_tipo = {}
    costo_real = 0.0
    for tipo in tipos_gasto:
        val_str = request.form.get(f"tipo_costo_{tipo}", "").strip().replace(",", ".")
        try:
            val = float(val_str) if val_str else 0.0
        except ValueError:
            val = 0.0
        if val > 0:
            costos_por_tipo[tipo] = val
            costo_real += val

    if costo_real <= 0:
        flash("Debe ingresar al menos un costo mayor a cero.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    # Incluir penalización de reagenda en el costo total
    penalizacion_extra = 0.0
    try:
        pen_raw = request.form.get("penalizacion_extra", "").strip()
        penalizacion_extra = float(pen_raw) if pen_raw else 0.0
    except ValueError:
        penalizacion_extra = 0.0
    # Si no vino del form (campo readonly puede omitirse), usar el valor guardado en DB
    if penalizacion_extra <= 0:
        penalizacion_extra = float(s.get("penalizacion") or 0)

    if penalizacion_extra > 0:
        costos_por_tipo["Penalización reagenda"] = penalizacion_extra
        costo_real += penalizacion_extra

    desglose = ", ".join(f"{t}: ${v:,.2f}" for t, v in costos_por_tipo.items())
    notas_full = f"{desglose}\n{notas}".strip() if notas else desglose
    repo.liquidar_vuelo(sid, u["id"], u["nombre"], costo_real, notas_full)

    # Deducir del presupuesto por tipo de gasto
    if s.get("centro_costo_id"):
        try:
            empresa_id = repo.get_empresa_by_usuario(s["solicitante_id"])
            if empresa_id:
                hoy = _date.today()
                for tipo, costo in costos_por_tipo.items():
                    repo.deducir_presupuesto_vuelo(
                        empresa_id, s["centro_costo_id"], tipo,
                        hoy.year, hoy.month, costo
                    )
        except Exception as exc:
            current_app.logger.warning("[VUELO] Error al deducir presupuesto sid=%s: %s", sid, exc)

    # Notificar al coordinador y al solicitante
    try:
        notif.notif_vuelo_liquidada(
            sid, s["area_solicitante"], str(s["fecha"]),
            s["solicitante_nombre"], s["solicitante_id"],
            u["nombre"], costo_real, notas,
        )
    except Exception:
        pass

    flash("Vuelo completado y costos registrados. El presupuesto fue actualizado.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# Voucher: aprobación jefe directo
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/voucher/aprobar-jefe", methods=["POST"],
                       endpoint="planificador_voucher_aprobar_jefe")
@require_login
def voucher_aprobar_jefe(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_aprobar_jefe_voucher(s, u["id"], ctx):
        abort(403)
    obs = request.form.get("observacion", "").strip()
    repo.aprobar_jefe_voucher(sid, u["id"], u["nombre"], obs)
    try:
        notif.notif_voucher_aprobada_solicitante(
            sid, s["area_solicitante"], str(s["fecha"]),
            s.get("descripcion", ""), s["solicitante_id"], s["solicitante_nombre"], u["nombre"],
        )
    except Exception:
        pass
    try:
        notif.notif_voucher_pendiente_entrega(
            sid, s["area_solicitante"], str(s["fecha"]),
            s.get("descripcion", ""), s["solicitante_nombre"], u["nombre"],
        )
    except Exception:
        pass
    flash("Voucher aprobado. El solicitante y el coordinador fueron notificados.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


@planificador_bp.route("/solicitudes/<int:sid>/voucher/rechazar-jefe", methods=["POST"],
                       endpoint="planificador_voucher_rechazar_jefe")
@require_login
def voucher_rechazar_jefe(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_aprobar_jefe_voucher(s, u["id"], ctx):
        abort(403)
    obs = request.form.get("observacion", "").strip()
    if not obs:
        flash("Debe indicar el motivo del rechazo.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))
    repo.rechazar_jefe_voucher(sid, u["id"], u["nombre"], obs)
    try:
        notif.notif_voucher_rechazada(
            sid, s["area_solicitante"], str(s["fecha"]),
            obs, s["solicitante_nombre"], s["solicitante_id"], u["nombre"],
        )
    except Exception:
        pass
    flash("Solicitud de Voucher rechazada.", "warning")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# Voucher: coordinador entrega los vouchers, registra el secuencial de cada uno
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/voucher/entregar", methods=["POST"],
                       endpoint="planificador_voucher_entregar")
@require_login
def voucher_entregar(sid):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    if not s or not svc.puede_entregar_voucher(s, u["id"], ctx):
        abort(403)

    items = repo.get_voucher_items(sid)
    if not items:
        flash("Esta solicitud no tiene vouchers registrados.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    secuenciales = {}
    for item in items:
        val = request.form.get(f"secuencial_{item['id']}", "").strip()
        if not val:
            flash("Debe indicar el secuencial de todos los vouchers.", "warning")
            return redirect(url_for("planificador.planificador_solicitudes"))
        secuenciales[item["id"]] = val

    repo.entregar_voucher_items(sid, secuenciales, u["id"], u["nombre"])
    try:
        notif.notif_voucher_entregado_usuario(
            sid, s["area_solicitante"], str(s["fecha"]),
            s["solicitante_id"], s["solicitante_nombre"], len(secuenciales),
        )
    except Exception:
        pass
    flash("Vouchers entregados con sus secuenciales. Se notificó al solicitante.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# Voucher: el solicitante confirma UN voucher (adjunto + observación)
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/voucher/item/<int:item_id>/confirmar", methods=["POST"],
                       endpoint="planificador_voucher_confirmar_item")
@require_login
def voucher_confirmar_item(sid, item_id):
    import os, uuid
    from werkzeug.utils import secure_filename

    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    item = repo.get_voucher_item_by_id(item_id)
    if not s or not item or item.get("solicitud_id") != sid:
        abort(404)
    if not svc.puede_confirmar_voucher_item(s, item, u["id"], ctx):
        abort(403)

    obs = request.form.get("observacion", "").strip()

    archivo = request.files.get("adjunto_voucher")
    if not archivo or not archivo.filename:
        flash("Debe subir el respaldo de este voucher.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    archivo.seek(0, 2)
    tamano = archivo.tell()
    archivo.seek(0)
    if tamano > 10 * 1024 * 1024:
        flash("El archivo supera el límite de 10 MB.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    nombre_original = secure_filename(archivo.filename)
    ext             = os.path.splitext(nombre_original)[1]
    nombre_guardado = f"{uuid.uuid4().hex}{ext}"
    carpeta = os.path.join(current_app.config["UPLOAD_FOLDER"], "planificador", str(sid), "vouchers")
    os.makedirs(carpeta, exist_ok=True)
    ruta_completa = os.path.join(carpeta, nombre_guardado)
    archivo.save(ruta_completa)

    todos_confirmados = repo.confirmar_voucher_item(
        item_id, sid, u["id"], u["nombre"],
        nombre_original, nombre_guardado, tamano, obs,
    )

    if todos_confirmados:
        try:
            notif.notif_voucher_pendiente_liquidacion(
                sid, s["area_solicitante"], str(s["fecha"]), s["solicitante_nombre"],
            )
        except Exception:
            pass
        msg_ok = ("Voucher confirmado. Todos tus vouchers quedaron confirmados — "
                  "se notificó al coordinador para registrar los costos.")
    else:
        msg_ok = "Voucher confirmado. Aún tienes vouchers pendientes de confirmar."

    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if is_ajax:
        from flask import jsonify
        return jsonify({"ok": True, "todos_confirmados": todos_confirmados,
                        "adjunto_nombre": nombre_original, "msg": msg_ok})
    flash(msg_ok, "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# Voucher: el solicitante marca UN voucher como no utilizado
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/voucher/item/<int:item_id>/no-utilizado", methods=["POST"],
                       endpoint="planificador_voucher_no_utilizado_item")
@require_login
def voucher_no_utilizado_item(sid, item_id):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    item = repo.get_voucher_item_by_id(item_id)
    if not s or not item or item.get("solicitud_id") != sid:
        abort(404)
    # Mismo permiso que confirmar: es la misma etapa/actor (el solicitante).
    if not svc.puede_confirmar_voucher_item(s, item, u["id"], ctx):
        abort(403)

    obs = request.form.get("observacion", "").strip()
    if not obs:
        msg = "Debe indicar el motivo por el que no utilizó este voucher."
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
        if is_ajax:
            from flask import jsonify
            return jsonify({"ok": False, "msg": msg}), 400
        flash(msg, "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    resultado = repo.marcar_voucher_no_utilizado(item_id, sid, u["id"], u["nombre"], obs)

    if resultado["completada"]:
        try:
            notif.notif_voucher_liquidada(
                sid, s["area_solicitante"], str(s["fecha"]), s["solicitante_nombre"],
                s["solicitante_id"], u["nombre"], resultado["costo_total"],
            )
        except Exception:
            pass
        msg_ok = "Voucher marcado como no utilizado. La solicitud quedó completada."
    elif resultado["todos_confirmados"]:
        try:
            notif.notif_voucher_pendiente_liquidacion(
                sid, s["area_solicitante"], str(s["fecha"]), s["solicitante_nombre"],
            )
        except Exception:
            pass
        msg_ok = ("Voucher marcado como no utilizado. Todos tus vouchers quedaron confirmados — "
                  "se notificó al coordinador para registrar los costos.")
    else:
        msg_ok = "Voucher marcado como no utilizado. Aún tienes vouchers pendientes de confirmar."

    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if is_ajax:
        from flask import jsonify
        return jsonify({"ok": True, "completada": resultado["completada"],
                        "todos_confirmados": resultado["todos_confirmados"], "msg": msg_ok})
    flash(msg_ok, "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# Voucher: coordinador liquida UN voucher (sin validar presupuesto)
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/voucher/item/<int:item_id>/liquidar", methods=["POST"],
                       endpoint="planificador_voucher_liquidar_item")
@require_login
def voucher_liquidar_item(sid, item_id):
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    s = repo.get_solicitud_by_id(sid)
    item = repo.get_voucher_item_by_id(item_id)
    if not s or not item or item.get("solicitud_id") != sid:
        abort(404)
    if not svc.puede_liquidar_voucher_item(s, item, u["id"], ctx):
        abort(403)

    costo_str = request.form.get("costo", "").strip().replace(",", ".")
    try:
        costo = float(costo_str) if costo_str else 0.0
    except ValueError:
        costo = 0.0
    if costo <= 0:
        flash("Debe ingresar un costo mayor a cero.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    todos_liquidados, costo_total = repo.liquidar_voucher_item(item_id, sid, u["id"], u["nombre"], costo)

    if todos_liquidados:
        try:
            notif.notif_voucher_liquidada(
                sid, s["area_solicitante"], str(s["fecha"]),
                s["solicitante_nombre"], s["solicitante_id"],
                u["nombre"], costo_total,
            )
        except Exception:
            pass
        flash(f"Voucher liquidado. Todos los vouchers quedaron completados — costo total: ${costo_total:,.2f}.", "success")
    else:
        flash("Voucher liquidado. Aún hay vouchers pendientes de liquidar.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


# ─────────────────────────────────────────────────────────────
# Voucher: carga masiva de costos (Excel del proveedor de taxis)
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/vouchers/carga-masiva-costos", methods=["POST"],
                       endpoint="planificador_voucher_carga_masiva")
@require_login
def voucher_carga_masiva():
    from flask import jsonify

    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    es_coordinador_voucher = "Voucher" in (ctx.get("tipos_coordinador") or [])
    if not (ctx["es_admin"] or es_coordinador_voucher):
        return jsonify({"ok": False, "msg": "No tiene permiso para cargar costos de vouchers."}), 403

    archivo = request.files.get("archivo_excel")
    if not archivo or not archivo.filename:
        return jsonify({"ok": False, "msg": "Debe seleccionar un archivo Excel (.xlsx)."}), 400
    if not archivo.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"ok": False, "msg": "El archivo debe ser .xlsx o .xlsm."}), 400

    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({"ok": False, "msg": "openpyxl no está instalado en el servidor."}), 500

    try:
        wb = load_workbook(archivo, data_only=True, read_only=True)
    except Exception:
        return jsonify({"ok": False,
                        "msg": "No se pudo leer el archivo. Verifica que sea un Excel válido (.xlsx)."}), 400

    columnas_requeridas = {"voucher", "valor"}
    hoja = None
    col_idx = {}
    header_row_num = None

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=5):
            valores = {
                (str(c.value).strip().lower() if c.value is not None else ""): idx
                for idx, c in enumerate(row)
            }
            if columnas_requeridas.issubset(valores.keys()):
                hoja = ws
                col_idx = valores
                header_row_num = row[0].row
                break
        if hoja is not None:
            break

    if hoja is None:
        return jsonify({
            "ok": False,
            "msg": ("No se encontró en ninguna hoja del archivo una fila de encabezados con las "
                    "columnas 'VOUCHER' y 'VALOR'. Revisa el formato del archivo."),
        }), 400

    idx_voucher = col_idx["voucher"]
    idx_valor   = col_idx["valor"]
    idx_origen  = col_idx.get("origen")
    idx_destino = col_idx.get("destino")

    exitosos = 0
    errores = []
    fila_num = header_row_num

    for row in hoja.iter_rows(min_row=header_row_num + 1):
        fila_num += 1
        celdas = list(row)
        if idx_voucher >= len(celdas):
            continue

        v_voucher = celdas[idx_voucher].value
        if v_voucher is None or str(v_voucher).strip() == "":
            continue  # fila en blanco, no cuenta como error

        secuencial = str(v_voucher).strip()
        if secuencial.endswith(".0"):
            secuencial = secuencial[:-2]

        v_valor = celdas[idx_valor].value if idx_valor < len(celdas) else None
        try:
            costo = float(v_valor)
        except (TypeError, ValueError):
            errores.append({"fila": fila_num, "voucher": secuencial,
                            "motivo": f"Costo inválido: {v_valor!r}"})
            continue

        origen_real = (
            str(celdas[idx_origen].value).strip()
            if idx_origen is not None and idx_origen < len(celdas) and celdas[idx_origen].value
            else ""
        )
        destino_real = (
            str(celdas[idx_destino].value).strip()
            if idx_destino is not None and idx_destino < len(celdas) and celdas[idx_destino].value
            else ""
        )

        item = repo.get_voucher_item_by_secuencial(secuencial)
        if not item:
            errores.append({"fila": fila_num, "voucher": secuencial,
                            "motivo": "No existe ningún voucher con ese secuencial en el sistema."})
            continue

        if item.get("costo") is not None:
            errores.append({"fila": fila_num, "voucher": secuencial,
                            "motivo": "Este voucher ya tenía un costo registrado; se omitió para no duplicar."})
            continue

        s = repo.get_solicitud_by_id(item["solicitud_id"])
        if not s or not svc.puede_liquidar_voucher_item(s, item, u["id"], ctx):
            errores.append({
                "fila": fila_num, "voucher": secuencial,
                "motivo": ("La solicitud de este voucher no está lista para liquidar todavía "
                           "(el solicitante aún no confirma todos sus vouchers)."),
            })
            continue

        try:
            todos_liquidados, costo_total = repo.liquidar_voucher_item(
                item["id"], item["solicitud_id"], u["id"], u["nombre"], costo,
            )
            # El origen/destino real es un dato adicional (para el futuro
            # indicador de coincidencia): si falla al guardarlo no debe
            # invalidar el costo, que ya quedó liquidado arriba.
            if origen_real or destino_real:
                try:
                    repo.set_voucher_item_datos_reales(item["id"], origen_real, destino_real)
                except Exception:
                    current_app.logger.exception(
                        "[VOUCHER-CARGA-MASIVA] No se pudo guardar origen/destino real item_id=%s",
                        item["id"],
                    )
            if todos_liquidados:
                try:
                    notif.notif_voucher_liquidada(
                        item["solicitud_id"], s["area_solicitante"], str(s["fecha"]),
                        s["solicitante_nombre"], s["solicitante_id"], u["nombre"], costo_total,
                    )
                except Exception:
                    pass
            exitosos += 1
        except Exception as exc:
            errores.append({"fila": fila_num, "voucher": secuencial,
                            "motivo": f"Error al guardar: {exc}"})

    return jsonify({
        "ok": True,
        "hoja": hoja.title,
        "procesados": exitosos + len(errores),
        "exitosos": exitosos,
        "errores": errores,
    })


# ─────────────────────────────────────────────────────────────
# Adjuntos de solicitudes
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/<int:sid>/adjuntos/subir", methods=["POST"],
                       endpoint="planificador_adjunto_subir")
@require_login
def adjunto_subir(sid):
    import os, uuid
    from werkzeug.utils import secure_filename

    u = _current_user()
    s = repo.get_solicitud_by_id(sid)
    if not s:
        abort(404)

    # Solo se puede subir archivos si la solicitud ya fue coordinada (no PENDIENTE_COORDINACION)
    if s["estado"] == "PENDIENTE_COORDINACION":
        flash("Solo se pueden adjuntar archivos a solicitudes ya coordinadas.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    # Solo el solicitante, coordinadores, aprobadores, gerentes y admin
    ctx = svc.get_user_context(u["id"], u["rol"])
    es_involucrado = (
        ctx["es_admin"]
        or ctx["es_gerente"]
        or ctx["tipos_coordinador"]
        or ctx["tipos_aprobador"]
        or s["solicitante_id"] == u["id"]
    )
    if not es_involucrado:
        abort(403)

    archivo = request.files.get("adjunto")
    if not archivo or not archivo.filename:
        flash("No se seleccionó ningún archivo.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    archivo.seek(0, 2)
    tamano = archivo.tell()
    archivo.seek(0)
    if tamano > 5 * 1024 * 1024:
        flash("El archivo supera el límite de 5 MB.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    nombre_original = secure_filename(archivo.filename)
    ext = os.path.splitext(nombre_original)[1]
    nombre_guardado = f"{uuid.uuid4().hex}{ext}"

    carpeta = os.path.join(current_app.config["UPLOAD_FOLDER"], "planificador", str(sid))
    os.makedirs(carpeta, exist_ok=True)
    archivo.save(os.path.join(carpeta, nombre_guardado))

    repo.insert_adjunto(sid, nombre_original, nombre_guardado, tamano, u["id"], u["nombre"])
    flash("Archivo adjuntado correctamente.", "success")
    return redirect(url_for("planificador.planificador_solicitudes"))


@planificador_bp.route("/adjuntos/<int:aid>/eliminar", methods=["POST"],
                       endpoint="planificador_adjunto_eliminar")
@require_login
def adjunto_eliminar(aid):
    import os

    u = _current_user()
    adj = repo.get_adjunto_by_id(aid)
    if not adj:
        abort(404)

    ctx = svc.get_user_context(u["id"], u["rol"])
    estado = adj["estado"]
    puede = ctx["es_admin"] or estado not in ("APROBADA", "COMPLETADA")
    if not puede:
        flash("No se puede eliminar archivos de solicitudes aprobadas o completadas.", "warning")
        return redirect(url_for("planificador.planificador_solicitudes"))

    ruta = os.path.join(
        current_app.config["UPLOAD_FOLDER"], "planificador",
        str(adj["solicitud_id"]), adj["nombre_guardado"]
    )
    try:
        os.remove(ruta)
    except OSError:
        pass

    repo.delete_adjunto(aid)
    flash("Archivo eliminado.", "info")
    return redirect(url_for("planificador.planificador_solicitudes"))


@planificador_bp.route("/solicitudes/<int:sid>/adjuntos/<nombre>",
                       endpoint="planificador_adjunto_descargar")
@require_login
def adjunto_descargar(sid, nombre):
    import os
    from flask import send_from_directory

    u = _current_user()
    s = repo.get_solicitud_by_id(sid)
    if not s:
        abort(404)

    ctx = svc.get_user_context(u["id"], u["rol"])
    es_involucrado = (
        ctx["es_admin"] or ctx["es_gerente"]
        or ctx["tipos_coordinador"] or ctx["tipos_aprobador"]
        or s["solicitante_id"] == u["id"]
    )
    if not es_involucrado:
        abort(403)

    carpeta = os.path.join(current_app.config["UPLOAD_FOLDER"], "planificador", str(sid))
    return send_from_directory(carpeta, nombre, as_attachment=True)


# ─────────────────────────────────────────────────────────────
# Configuración
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/configuracion", methods=["GET", "POST"],
                       endpoint="planificador_configuracion")
@require_login
@require_permission(PERM_CONFIG, "ver")
def configuracion():
    u = _current_user()

    if request.method == "POST":
        if not _check_perm(u["rol"], PERM_CONFIG, "crear"):
            abort(403)
        tipo        = request.form.get("tipo", "").strip()
        usuario_id  = request.form.get("usuario_id", "").strip()
        usuario_nombre = request.form.get("usuario_nombre", "").strip()
        rol_config  = request.form.get("rol_config", "").strip()

        tipos_validos = repo.get_tipos_solicitud()
        if not all([tipo, usuario_id, usuario_nombre, rol_config]):
            flash("Todos los campos son requeridos.", "warning")
        elif tipo not in tipos_validos:
            flash("Tipo de solicitud no válido.", "warning")
        elif rol_config not in (ROL_COORDINADOR, ROL_APROBADOR, ROL_MOTORIZADO, ROL_GERENTE_PRESUPUESTO):
            flash("Rol de configuración no válido.", "warning")
        else:
            try:
                repo.insert_config(tipo, int(usuario_id), usuario_nombre, rol_config)
                flash("Configuración guardada.", "success")
            except Exception:
                flash("Error al guardar. Verifique que no exista ya esa combinación.", "danger")

        return redirect(url_for("planificador.planificador_configuracion"))

    config_rows     = repo.get_all_config()
    usuarios        = repo.get_usuarios_for_select()
    tipos_solicitud = repo.get_tipos_solicitud()
    tipo_flags      = repo.get_all_tipo_flags()
    motorizados_tg  = repo.get_motorizados_telegram_status()
    rol_flags       = repo.get_all_rol_flags()

    return render_template(
        "planificador/configuracion.html",
        active_page=ACTIVE_KEY,
        config_rows=config_rows,
        usuarios=usuarios,
        tipos=tipos_solicitud,
        roles_config=[ROL_COORDINADOR, ROL_APROBADOR, ROL_MOTORIZADO, ROL_GERENTE_PRESUPUESTO],
        tipo_flags=tipo_flags,
        motorizados_tg=motorizados_tg,
        rol_flags=rol_flags,
        roles_candidatos_autoaprobar=ROLES_CANDIDATOS_AUTOAPROBAR_VUELO,
    )


@planificador_bp.route("/configuracion/<int:cid>/eliminar", methods=["POST"],
                       endpoint="planificador_config_eliminar")
@require_login
@require_permission(PERM_CONFIG, "eliminar")
def config_eliminar(cid):
    repo.delete_config(cid)
    flash("Configuración eliminada.", "success")
    return redirect(url_for("planificador.planificador_configuracion"))


@planificador_bp.route("/configuracion/telegram-chat-id", methods=["POST"],
                       endpoint="planificador_set_telegram_chat_id")
@require_login
@require_permission(PERM_CONFIG, "editar")
def set_telegram_chat_id():
    """Admin guarda manualmente el telegram_chat_id de un usuario motorizado."""
    usuario_id = request.form.get("usuario_id", "").strip()
    chat_id    = request.form.get("telegram_chat_id", "").strip()
    if not usuario_id:
        flash("Usuario no indicado.", "warning")
        return redirect(url_for("planificador.planificador_configuracion"))
    try:
        repo.update_usuario_telegram_chat_id(int(usuario_id), chat_id or None)
        flash("Chat ID de Telegram actualizado.", "success")
    except Exception as exc:
        flash(f"Error al actualizar: {exc}", "danger")
    return redirect(url_for("planificador.planificador_configuracion"))


@planificador_bp.route("/configuracion/tipo-flags", methods=["POST"],
                       endpoint="planificador_tipo_flags")
@require_login
@require_permission(PERM_CONFIG, "editar")
def tipo_flags_update():
    """Actualiza los flags de configuración por tipo (ej: requiere aprobación gerente)."""
    u = _current_user()
    if not _check_perm(u["rol"], PERM_CONFIG, "editar"):
        abort(403)
    tipos_validos = repo.get_tipos_solicitud()
    for tipo in tipos_validos:
        key = f"req_gerente_{tipo.replace(' ', '_').replace('/', '_')}"
        req_gerente = request.form.get(key) == "1"
        auto_confirmar = request.form.get(f"auto_confirmar_{tipo.replace(' ', '_').replace('/', '_')}") == "1"
        auto_liquidar  = request.form.get(f"auto_liquidar_{tipo.replace(' ', '_').replace('/', '_')}")  == "1"
        repo.set_tipo_flags(tipo, req_gerente, auto_confirmar, auto_liquidar)
    flash("Configuración de tipos actualizada.", "success")
    return redirect(url_for("planificador.planificador_configuracion"))


@planificador_bp.route("/configuracion/rol-flags", methods=["POST"],
                       endpoint="planificador_rol_flags")
@require_login
@require_permission(PERM_CONFIG, "editar")
def rol_flags_update():
    """Actualiza qué roles auto-aprueban el paso de aprobación del jefe
    directo en solicitudes de Vuelo (ej: gerentes que no tienen jefe)."""
    u = _current_user()
    if not _check_perm(u["rol"], PERM_CONFIG, "editar"):
        abort(403)
    for rol in ROLES_CANDIDATOS_AUTOAPROBAR_VUELO:
        key = f"autoaprueba_{rol.replace(' ', '_')}"
        autoaprueba = request.form.get(key) == "1"
        repo.set_rol_flags(rol, autoaprueba)
    flash("Configuración de roles actualizada.", "success")
    return redirect(url_for("planificador.planificador_configuracion"))


# ─────────────────────────────────────────────────────────────
# Reverse geocode (proxy Nominatim – mismo origen, no viola CSP)
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/reverse-geocode", endpoint="planificador_reverse_geocode")
@require_login
def reverse_geocode():
    lat = request.args.get("lat", "").strip()
    lng = request.args.get("lng", "").strip()
    q   = request.args.get("q",   "").strip()
    try:
        if lat and lng:
            url = (f"https://nominatim.openstreetmap.org/reverse"
                   f"?format=json&lat={urllib.parse.quote(lat)}&lon={urllib.parse.quote(lng)}&zoom=18&addressdetails=1")
        elif q:
            url = (f"https://nominatim.openstreetmap.org/search"
                   f"?format=json&q={urllib.parse.quote(q)}&limit=1&addressdetails=1")
        else:
            return jsonify({"error": "Parámetros requeridos: lat+lng o q"}), 400

        req = urllib.request.Request(url, headers={"User-Agent": "SGQ-Quimpac/1.0"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = resp.read().decode("utf-8")
        import json as _json
        parsed = _json.loads(data)
        if isinstance(parsed, list):
            parsed = parsed[0] if parsed else {}
        # Construir dirección limpia desde el objeto address
        addr_obj = parsed.get("address", {})
        parts = []
        road = addr_obj.get("road") or addr_obj.get("pedestrian") or addr_obj.get("street") or ""
        house = addr_obj.get("house_number", "")
        if road:
            parts.append((road + " " + house).strip())
        suburb = addr_obj.get("suburb") or addr_obj.get("neighbourhood") or addr_obj.get("quarter") or ""
        if suburb:
            parts.append(suburb)
        city = addr_obj.get("city") or addr_obj.get("town") or addr_obj.get("municipality") or ""
        if city:
            parts.append(city)
        clean = ", ".join(p for p in parts if p) or parsed.get("display_name", "")
        return jsonify({"address": clean})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ─────────────────────────────────────────────────────────────
# Reporte CSV  (bug fix: iterar .values() no keys del dict)
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/reporte", endpoint="planificador_reporte")
@require_login
@require_permission(PERM_SOLICITUDES, "ver")
def reporte():
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    filters = {
        "estado":      request.args.get("estado", ""),
        "tipo":        request.args.get("tipo",   ""),
        "fecha_desde": request.args.get("fecha_desde", ""),
        "fecha_hasta": request.args.get("fecha_hasta", ""),
    }
    cols, rows = repo.get_solicitudes_para_reporte(filters, u["id"], ctx)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_ALL)
    writer.writerow(cols)
    for row in rows:
        # RowCompat es un dict — iterar .values() para obtener datos, no claves
        writer.writerow([str(v) if v is not None else "" for v in row.values()])

    filename = f"planificador_{date.today()}.csv"
    return Response(
        "﻿" + output.getvalue(),   # BOM UTF-8 para Excel
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ─────────────────────────────────────────────────────────────
# Reporte Excel
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/solicitudes/reporte.xlsx", endpoint="planificador_reporte_excel")
@require_login
@require_permission(PERM_SOLICITUDES, "ver")
def reporte_excel():
    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    filters = {
        "estado":      request.args.get("estado", ""),
        "tipo":        request.args.get("tipo",   ""),
        "fecha_desde": request.args.get("fecha_desde", ""),
        "fecha_hasta": request.args.get("fecha_hasta", ""),
    }
    output = _build_planificador_excel(filters, u["id"], ctx)
    filename = f"planificador_{date.today()}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _build_planificador_excel(filters: dict, usuario_id=None, ctx=None) -> BytesIO:
    """Genera el Excel de solicitudes del planificador con formato."""
    from decimal import Decimal
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl no instalado. Ejecuta: pip install openpyxl")

    cols, rows = repo.get_solicitudes_para_reporte(filters, usuario_id, ctx)

    wb = Workbook()
    ws = wb.active
    ws.title = "Planificador"

    # ── Fila de título ───────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    title_cell = ws["A1"]
    title_cell.value = f"PLANIFICADOR DE SOLICITUDES — SGQ Quimpac · {date.today().strftime('%d/%m/%Y')}"
    title_cell.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", fgColor="1E3A8A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Fila de cabecera ─────────────────────────────────────
    header_fill   = PatternFill("solid", fgColor="DBEAFE")
    header_font   = Font(name="Calibri", bold=True, size=10, color="1E3A8A")
    header_border = Border(
        bottom=Side(style="medium", color="1E3A8A"),
        right=Side(style="thin",   color="CBD5E1"),
    )
    for ci, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=ci, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = header_border
    ws.row_dimensions[2].height = 30

    # ── Colores por estado ────────────────────────────────────
    ESTADO_FILL = {
        "PENDIENTE_COORDINACION":       PatternFill("solid", fgColor="FEF9C3"),
        "PENDIENTE_APROBACION":         PatternFill("solid", fgColor="CFFAFE"),
        "PENDIENTE_APROBACION_GERENTE": PatternFill("solid", fgColor="DBEAFE"),
        "APROBADA":                     PatternFill("solid", fgColor="DCFCE7"),
        "RECHAZADA":                    PatternFill("solid", fgColor="FEE2E2"),
        "COMPLETADA":                   PatternFill("solid", fgColor="F3F4F6"),
    }
    thin_border = Border(
        right=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    # Índice de la columna "Estado" (basado en el nombre de col)
    try:
        estado_col_idx = cols.index("Estado") + 1
    except ValueError:
        estado_col_idx = None

    # ── Datos ────────────────────────────────────────────────
    for ri, row in enumerate(rows, start=3):
        values = list(row.values())
        estado_val = ""
        if estado_col_idx:
            estado_val = str(values[estado_col_idx - 1] or "")
        row_fill = ESTADO_FILL.get(estado_val)

        for ci, v in enumerate(values, start=1):
            if v is None:
                cell_val = ""
            elif isinstance(v, Decimal):
                cell_val = float(v)
            elif isinstance(v, (int, float)):
                cell_val = v
            else:
                cell_val = str(v)
            cell = ws.cell(row=ri, column=ci, value=cell_val)
            cell.font      = Font(name="Calibri", size=9)
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if row_fill:
                cell.fill = row_fill

        # Alternar fila clara si no tiene color de estado
        if not row_fill and ri % 2 == 0:
            alt_fill = PatternFill("solid", fgColor="F8FAFC")
            for ci in range(1, len(values) + 1):
                ws.cell(row=ri, column=ci).fill = alt_fill

        ws.row_dimensions[ri].height = 16

    # ── Anchos de columna ────────────────────────────────────
    col_widths = {
        "N° Solicitud": 12, "Tipo": 22, "Área Solicitante": 22,
        "Descripción": 35, "Lugar / Destino": 28, "Contacto": 18,
        "Prioridad": 11, "Fecha": 12, "Hora Inicio": 11, "Hora Fin": 10,
        "Estado": 22, "Solicitante": 20, "Coordinador": 20,
        "Aprobador": 20, "Obs. Coordinador": 28, "Obs. Aprobador": 28,
        "Ciudad": 14, "Detalle Dirección": 30,
        "Centro de Costo": 22, "Presupuesto Total (Año)": 18,
        "Valor Consumido (Año)": 18, "Gasto Realizado": 16,
        "Fecha Creación": 18, "Última Actualización": 18,
    }
    for ci, col_name in enumerate(cols, start=1):
        w = col_widths.get(col_name, 15)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Freeze panes y auto-filtro ───────────────────────────
    ws.freeze_panes = "A3"
    if rows:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{len(rows) + 2}"

    # ── Pestaña de color ─────────────────────────────────────
    ws.sheet_properties.tabColor = "1E3A8A"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─────────────────────────────────────────────────────────────
# Helper interno de permiso sin decorador
# ─────────────────────────────────────────────────────────────

def _check_perm(rol, opcion, accion):
    from modules.security import has_permission
    return has_permission(rol, opcion, accion)


# ─────────────────────────────────────────────────────────────
# Presupuesto
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/presupuesto", methods=["GET"], endpoint="planificador_presupuesto")
@require_login
@require_permission(PERM_PRESUPUESTO, "ver")
def presupuesto():
    from datetime import date as _date
    anio_actual = _date.today().year
    anio = int(request.args.get("anio", anio_actual))
    empresa_id = request.args.get("empresa_id", type=int)
    tipo_gasto = request.args.get("tipo_gasto", "")

    tipos_gasto = repo.get_tipos_gasto()

    from modules.db import get_db
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, razon_social FROM empresas WHERE activo=1 ORDER BY razon_social")
    empresas = [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]

    # Centros de costo: param_group "Centro de Costo"
    # No se filtra por activo porque la mayoría están marcados como inactivos en param_values
    cur.execute("""
        SELECT pv.id, pv.nombre, COALESCE(pv.valor, '') AS codigo
        FROM param_values pv
        JOIN param_groups pg ON pg.id = pv.group_id
        WHERE pg.nombre = 'Centro de Costo'
        ORDER BY pv.valor, pv.nombre
    """)
    centros_disponibles = [{"id": r[0], "nombre": r[1], "codigo": r[2]}
                           for r in cur.fetchall()]

    def _build_cc_rows(empresa_id, tipo, anio, cur):
        cur.execute("""
            SELECT DISTINCT p.centro_costo_id, pv.nombre AS cc_nombre,
                   COALESCE(pv.valor, '') AS cc_codigo
            FROM planificador_presupuesto p
            JOIN param_values pv ON pv.id = p.centro_costo_id
            WHERE p.empresa_id = ? AND p.anio = ? AND p.tipo_gasto = ?
            ORDER BY cc_codigo, cc_nombre
        """, (empresa_id, anio, tipo))
        centros = [{"id": r[0], "nombre": r[1], "codigo": r[2]} for r in cur.fetchall()]
        rows = []
        for cc in centros:
            cur.execute("""
                SELECT mes, monto_presupuestado, monto_ejecutado
                FROM planificador_presupuesto
                WHERE empresa_id=? AND centro_costo_id=? AND tipo_gasto=? AND anio=?
            """, (empresa_id, cc["id"], tipo, anio))
            by_mes = {r[0]: {"mes": r[0], "monto_presupuestado": float(r[1]),
                              "monto_ejecutado": float(r[2])}
                      for r in cur.fetchall()}
            meses_data = [by_mes.get(m, {"mes": m, "monto_presupuestado": 0.0,
                                          "monto_ejecutado": 0.0})
                          for m in range(1, 13)]
            total_presup = sum(m["monto_presupuestado"] for m in meses_data)
            total_ejec   = sum(m["monto_ejecutado"]     for m in meses_data)
            pct = round(total_ejec / total_presup * 100, 1) if total_presup > 0 else 0
            semaforo = "rojo" if pct >= 100 else ("amarillo" if pct >= 50 else "verde")
            rows.append({
                "cc_id": cc["id"], "cc_nombre": cc["nombre"], "cc_codigo": cc["codigo"],
                "meses": meses_data, "total_presup": total_presup,
                "total_ejec": total_ejec, "pct": pct, "semaforo": semaforo,
            })
        return rows

    # presupuesto_secciones: lista de {tipo_gasto, rows}
    presupuesto_secciones = []
    if empresa_id:
        tipos_a_mostrar = tipos_gasto if not tipo_gasto else [tipo_gasto]
        for t in tipos_a_mostrar:
            rows = _build_cc_rows(empresa_id, t, anio, cur)
            if rows or tipo_gasto:  # si es "Todos" omite secciones vacías
                presupuesto_secciones.append({"tipo": t, "rows": rows})

    # KPIs agregados de todas las secciones visibles (para las tarjetas de resumen)
    todas_las_filas = [row for sec in presupuesto_secciones for row in sec["rows"]]
    kpi_total = sum(r["total_presup"] for r in todas_las_filas)
    kpi_ejec  = sum(r["total_ejec"] for r in todas_las_filas)
    kpi = {
        "total": kpi_total,
        "ejec": kpi_ejec,
        "ejec_pct": round(kpi_ejec / kpi_total * 100, 1) if kpi_total > 0 else 0,
        "disponible": kpi_total - kpi_ejec,
        "alertas": sum(1 for r in todas_las_filas if r["semaforo"] in ("amarillo", "rojo")),
    }

    return render_template(
        "planificador/presupuesto.html",
        anio=anio,
        presupuesto_secciones=presupuesto_secciones,
        kpi=kpi,
        anio_actual=anio_actual,
        empresas=empresas,
        empresa_id=empresa_id,
        tipos_gasto=tipos_gasto,
        tipo_gasto=tipo_gasto,
        centros_disponibles=centros_disponibles,
        meses_nombres=["Ene","Feb","Mar","Abr","May","Jun",
                        "Jul","Ago","Sep","Oct","Nov","Dic"],
    )


@planificador_bp.route("/presupuesto/guardar", methods=["POST"],
                       endpoint="planificador_presupuesto_guardar")
@require_login
@require_permission(PERM_PRESUPUESTO, "editar")
def presupuesto_guardar():
    empresa_id = request.form.get("empresa_id", type=int)
    cc_id = request.form.get("cc_id", type=int)
    tipo_gasto = request.form.get("tipo_gasto", "")
    anio = request.form.get("anio", type=int)

    if not (empresa_id and cc_id and tipo_gasto and anio):
        flash("Datos incompletos.", "danger")
        return redirect(url_for("planificador.planificador_presupuesto"))

    for mes in range(1, 13):
        raw = request.form.get(f"mes_{mes}", "0").replace(",", ".")
        try:
            monto = float(raw)
        except ValueError:
            monto = 0.0
        repo.upsert_presupuesto(empresa_id, cc_id, tipo_gasto, anio, mes, monto)

    flash("Presupuesto guardado correctamente.", "success")
    return redirect(url_for(
        "planificador.planificador_presupuesto",
        anio=anio, empresa_id=empresa_id, tipo_gasto=tipo_gasto,
    ))


# ─────────────────────────────────────────────────────────────
# Indicadores — Voucher
# ─────────────────────────────────────────────────────────────

@planificador_bp.route("/indicadores", methods=["GET"], endpoint="planificador_indicadores")
@require_login
@require_permission(PERM_INDICADORES, "ver")
def indicadores():
    from datetime import date as _date

    u = _current_user()
    ctx = svc.get_user_context(u["id"], u["rol"])
    es_coordinador_voucher = "Voucher" in (ctx.get("tipos_coordinador") or [])
    if not (ctx["es_admin"] or es_coordinador_voucher):
        flash("No tiene permiso para ver los indicadores.", "danger")
        return redirect(url_for("planificador.planificador_solicitudes"))

    hoy = _date.today()
    fecha_desde = request.args.get("fecha_desde") or hoy.replace(day=1).isoformat()
    fecha_hasta = request.args.get("fecha_hasta") or hoy.isoformat()
    area = (request.args.get("area") or "").strip()

    departamentos = repo.get_departamentos()

    kpi          = repo.get_voucher_indicadores_kpi(fecha_desde, fecha_hasta, area)
    por_usuario  = repo.get_voucher_indicadores_por_usuario(fecha_desde, fecha_hasta, area)
    por_depto    = repo.get_voucher_indicadores_por_departamento(fecha_desde, fecha_hasta, area)
    tendencia    = repo.get_voucher_indicadores_tendencia_mensual(fecha_desde, fecha_hasta, area)
    top_rutas    = repo.get_voucher_indicadores_top_rutas(fecha_desde, fecha_hasta, area)

    # Drill-down de "Pend. confirmación": Departamento -> Usuario -> vouchers
    pend_conf_rows = repo.get_voucher_indicadores_pend_confirmacion_detalle(fecha_desde, fecha_hasta, area)
    from collections import OrderedDict as _OrderedDict
    _por_area = _OrderedDict()
    for r in pend_conf_rows:
        area_bucket = _por_area.setdefault(r["area"], {"area": r["area"], "total": 0, "usuarios": _OrderedDict()})
        area_bucket["total"] += 1
        user_bucket = area_bucket["usuarios"].setdefault(
            r["solicitante_nombre"], {"usuario_nombre": r["solicitante_nombre"], "total": 0, "vouchers": []},
        )
        user_bucket["total"] += 1
        user_bucket["vouchers"].append(r)
    pend_conf_detalle = []
    for area_bucket in _por_area.values():
        area_bucket["usuarios"] = list(area_bucket["usuarios"].values())
        pend_conf_detalle.append(area_bucket)

    meses_nombres = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
                      "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    tendencia_labels = [f"{meses_nombres[t['mes'] - 1]} {t['anio']}" for t in tendencia]
    tendencia_vouchers = [t["num_vouchers"] for t in tendencia]
    tendencia_costo    = [round(float(t["costo_total"] or 0), 2) for t in tendencia]

    chart_data = {
        "estado": {
            "labels": ["Confirmados", "No utilizados", "Pend. entrega", "Pend. confirmación"],
            "values": [
                kpi.get("confirmados") or 0,
                kpi.get("no_utilizados") or 0,
                kpi.get("pend_entrega") or 0,
                kpi.get("pend_confirmacion") or 0,
            ],
        },
        "usuarios": {
            "labels": [r["solicitante_nombre"] for r in por_usuario],
            "vouchers": [r["num_vouchers"] for r in por_usuario],
            "costo": [round(float(r["costo_total"] or 0), 2) for r in por_usuario],
        },
        "departamentos": {
            "labels": [r["area"] for r in por_depto],
            "vouchers": [r["num_vouchers"] for r in por_depto],
            "costo": [round(float(r["costo_total"] or 0), 2) for r in por_depto],
        },
        "tendencia": {
            "labels": tendencia_labels,
            "vouchers": tendencia_vouchers,
            "costo": tendencia_costo,
        },
    }

    return render_template(
        "planificador/indicadores.html",
        active_page=ACTIVE_KEY,
        kpi=kpi,
        por_usuario=por_usuario,
        por_depto=por_depto,
        top_rutas=top_rutas,
        pend_conf_detalle=pend_conf_detalle,
        departamentos=departamentos,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        area_sel=area,
        chart_data=chart_data,
    )


# ─────────────────────────────────────────────────────────────
# Registro
# ─────────────────────────────────────────────────────────────

def register_planificador_routes(app):
    app.register_blueprint(planificador_bp)
