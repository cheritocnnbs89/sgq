from __future__ import annotations

import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from flask import (
    Blueprint,
    render_template,
    request,
    session,
    flash,
    url_for,
    redirect,
    abort,
    get_flashed_messages,
    send_file,
)
from modules.security import require_login, require_permission

from . import contratos_repository as repository
from . import contratos_services as services


contratos_bp = Blueprint(
    "contratos",
    __name__,
    url_prefix="/contratos",
    template_folder="templates",
)


def register_contratos_routes(app):
    app.register_blueprint(contratos_bp)


@contratos_bp.after_app_request
def _no_cache_for_contratos(response):
    if request.blueprint == "contratos":
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


@contratos_bp.route("/ingresar", methods=["GET", "POST"])
@require_login
@require_permission("contratos_ingresar", "ver")
def compras_nuevo():
    session["active_page"] = "contratos_ingresar"
    services.ensure_contratos_columns()
    services.ensure_softdelete_columns()
    services.ensure_contrato_archivos_table()

    usuarios, proveedores = services.get_compras_combos()

    if request.method == "POST":
        from modules.security import has_permission

        if not has_permission(session.get("rol"), "contratos_ingresar", "crear"):
            flash("No tiene permiso para crear contratos.", "danger")
            return redirect(url_for("contratos.compras_lista"))

        result = services.create_contrato_from_request()
        if not result["ok"]:
            from modules.security import has_permission as _hp
            can_exportar = _hp(session.get("rol"), "contratos_ingresar", "exportar")
            flash(result["message"], "danger")
            return render_template(
                "compras_ingreso.html",
                mode="create",
                row=result["row_back"],
                usuarios=usuarios,
                proveedores=proveedores,
                archivos=[],
                can_exportar=can_exportar,
                post_url=url_for("contratos.compras_nuevo"),
                back_url=request.form.get("next") or url_for("contratos.compras_lista"),
            )

        flash("Contrato guardado correctamente.", "success")
        return redirect(url_for("contratos.compras_lista"))

    row = services.get_row_create_contrato_default()
    from modules.security import has_permission as _hp
    can_exportar = _hp(session.get("rol"), "contratos_ingresar", "exportar")

    return render_template(
        "compras_ingreso.html",
        mode="create",
        row=row,
        usuarios=usuarios,
        proveedores=proveedores,
        archivos=[],
        can_exportar=can_exportar,
        post_url=url_for("contratos.compras_nuevo"),
        back_url=request.args.get("next") or url_for("contratos.compras_lista"),
    )


@contratos_bp.route("/compras/<int:contrato_id>/editar", methods=["GET", "POST"])
@require_login
@require_permission("contratos_ingresar", "ver")
def compras_editar(contrato_id: int):
    session["active_page"] = "contratos_ingresar"
    services.ensure_contratos_columns()
    services.ensure_softdelete_columns()
    services.ensure_contrato_archivos_table()

    usuarios, proveedores = services.get_compras_combos()
    row_db, row, archivos = services.get_contrato_for_edit(contrato_id)
    if not row_db:
        abort(404)

    if not services.contrato_editable(row_db):
        flash("Este contrato ya fue aprobado y no puede ser editado.", "warning")
        return redirect(url_for("contratos.compras_lista"))

    from modules.security import has_permission as _hp
    can_exportar = _hp(session.get("rol"), "contratos_ingresar", "exportar")

    if request.method == "POST":
        from modules.security import has_permission

        if not has_permission(session.get("rol"), "contratos_ingresar", "editar"):
            flash("No tiene permiso para editar contratos.", "danger")
            return redirect(url_for("contratos.compras_lista"))

        row_estado = repository.fetch_estado_aprobaciones_contrato(contrato_id)
        if row_estado and (
            int(row_estado["aprobado_jefe"] or 0)
            or int(row_estado["aprobado"] or 0)
            or int(row_estado["aprob_gf"] or 0)
        ):
            flash("El contrato fue aprobado recientemente y ya no puede editarse.", "warning")
            return redirect(url_for("contratos.compras_lista"))

        result = services.update_contrato_from_request(contrato_id)
        if not result["ok"]:
            flash(result["message"], "danger")
            return render_template(
                "compras_ingreso.html",
                mode="edit",
                row=result["row_back"],
                usuarios=usuarios,
                proveedores=proveedores,
                archivos=archivos,
                can_exportar=can_exportar,
                post_url=url_for("contratos.compras_editar", contrato_id=contrato_id),
                back_url=request.form.get("next") or url_for("contratos.compras_lista"),
            )

        flash("Contrato actualizado.", "success")
        return redirect(url_for("contratos.compras_lista"))

    return render_template(
        "compras_ingreso.html",
        mode="edit",
        row=row,
        usuarios=usuarios,
        proveedores=proveedores,
        archivos=archivos,
        can_exportar=can_exportar,
        post_url=url_for("contratos.compras_editar", contrato_id=contrato_id),
        back_url=request.args.get("next") or url_for("contratos.compras_lista"),
    )


@contratos_bp.route("/compras/<int:contrato_id>/eliminar", methods=["POST"])
@require_login
@require_permission("contratos_ingresar", "eliminar")
def compras_eliminar(contrato_id: int):
    services.ensure_softdelete_columns()
    row = repository.fetch_contrato_por_id(contrato_id)
    if not row:
        abort(404)

    services.soft_delete_contrato(contrato_id)
    flash("Contrato eliminado.", "success")
    return redirect(url_for("contratos.compras_lista"))

@contratos_bp.route("/compras", methods=["GET"])
@require_login
@require_permission("contratos_ingresar", "ver")
def compras_lista():
    session["active_page"] = "contratos_ingresar"
    services.ensure_contratos_columns()
    services.ensure_softdelete_columns()

    prov = (request.args.get("proveedor") or "").strip()
    pedi = (request.args.get("pedido") or "").strip()
    tipo = ((request.args.get("tipo_pp") or request.args.get("pagare_poliza_filtro") or "").strip().upper())

    rows = repository.list_contratos(proveedor=prov, pedido=pedi, tipo_pp=tipo)
    return render_template("consulta_compras.html", rows=rows)

@contratos_bp.route("/compras/<int:contrato_id>/archivos/fragment", methods=["GET"])
@require_login
@require_permission("contratos_ingresar", "ver")
def compras_archivos_fragment(contrato_id: int):
    services.ensure_contrato_archivos_table()
    archivos = repository.fetch_archivos_contrato(contrato_id)

    from modules.security import has_permission as _hp
    can_exportar = _hp(session.get("rol"), "contratos_ingresar", "exportar")

    return render_template(
        "contrato_archivos_modal_fragment.html",
        archivos=archivos,
        can_exportar=can_exportar,
    )


@contratos_bp.route("/compras/<int:contrato_id>/aprobj/toggle", methods=["POST"], endpoint="toggle_aprobacion_jefe")
@require_login
@require_permission("contratos_ingresar", "aprobar")
def toggle_aprobacion_jefe(contrato_id: int):
    row = repository.fetch_contrato_por_id(contrato_id)
    if not row:
        abort(404)

    services.toggle_jefe_contrato(contrato_id, session.get("usuario_id"))
    flash("Aprobación actualizada.", "success")
    return redirect(request.referrer or url_for("contratos.compras_lista"))


@contratos_bp.route("/contab", methods=["GET"])
@require_login
@require_permission("contratos_garantias", "ver")
def contab_lista():
    session["active_page"] = "contratos_garantias"
    services.ensure_garantias_columns()
    services.ensure_softdelete_columns()

    prov = (request.args.get("proveedor") or "").strip()
    pedi = (request.args.get("pedido") or "").strip()
    estado = (request.args.get("estado") or "").strip()
    requiere = services.get_requiere_renovacion_filter(
        (request.args.get("renovacion") or request.args.get("requiere_renovacion") or "")
    )

    fecha_registro_desde = (request.args.get("fecha_registro_desde") or "").strip()
    fecha_registro_hasta = (request.args.get("fecha_registro_hasta") or "").strip()
    fecha_vencimiento_desde = (request.args.get("fecha_vencimiento_desde") or "").strip()
    fecha_vencimiento_hasta = (request.args.get("fecha_vencimiento_hasta") or "").strip()

    rows = repository.list_garantias(
        proveedor=prov,
        pedido=pedi,
        estado=estado,
        requiere_renovacion=requiere,
        fecha_registro_desde=fecha_registro_desde,
        fecha_registro_hasta=fecha_registro_hasta,
        fecha_vencimiento_desde=fecha_vencimiento_desde,
        fecha_vencimiento_hasta=fecha_vencimiento_hasta,
    )
    return render_template("consulta_contabilidad.html", rows=rows)


@contratos_bp.route("/ver/garantia/<int:garantia_id>/fragment", methods=["GET"])
@require_login
@require_permission("contratos_garantias", "ver")
def ver_garantia_fragment(garantia_id):
    row = repository.fetch_detalle_garantia_fragment(garantia_id)
    if not row:
        return "<div class='alert alert-warning mb-0'>No existe la garantía.</div>", 404

    garantia = dict(row)
    contrato = {
        "pedido": row["pedido"],
        "proveedor": row["proveedor"],
        "objeto": row["objeto"],
        "valor_contrato": row["valor_contrato"],
    }

    return render_template(
        "garantia_detalle_fragment.html",
        garantia=garantia,
        contrato=contrato,
        proveedor=row["proveedor"],
        pedido=row["pedido"],
    )


@contratos_bp.route("/contab/nuevo", methods=["GET", "POST"])
@require_login
@require_permission("contratos_garantias", "crear")
def contab_nuevo():
    session["active_page"] = "contratos_garantias"
    services.ensure_garantias_columns()
    services.ensure_softdelete_columns()

    if request.method == "POST":
        result = services.create_garantia_from_request()
        if not result["ok"]:
            flash(result["message"], "danger")
            contratos = repository.list_contratos_aprobados_para_garantia(limit=300)
            return render_template(
                "contabilidad_ingreso.html",
                mode="create",
                row=result["row_back"],
                contratos=contratos,
                post_url=url_for("contratos.contab_nuevo"),
                back_url=request.form.get("next") or url_for("contratos.contab_lista"),
            )

        flash("Garantía guardada correctamente.", "success")
        return redirect(request.form.get("next") or url_for("contratos.contab_lista"))

    
    contratos = repository.list_contratos_aprobados_para_garantia(limit=300)
    return render_template(
        "contabilidad_ingreso.html",
        mode="create",
        row={},
        contratos=contratos,
        post_url=url_for("contratos.contab_nuevo"),
        back_url=request.args.get("next") or url_for("contratos.contab_lista"),
    )


@contratos_bp.route("/contab/<int:garantia_id>/editar", methods=["GET", "POST"])
@require_login
@require_permission("contratos_garantias", "editar")
def contab_editar(garantia_id: int):
    session["active_page"] = "contratos_garantias"
    services.ensure_garantias_columns()
    services.ensure_softdelete_columns()

    row_db, contratos, archivos = services.get_garantia_for_edit(garantia_id)
    if not row_db:
        abort(404)

    if not services.garantia_editable(row_db):
        flash("Esta garantía ya fue aprobada y no puede ser editada.", "warning")
        return redirect(url_for("contratos.contab_lista"))

    if request.method == "POST":
        rstate = repository.fetch_estado_aprobaciones_garantia(garantia_id)
        if rstate and (
            int(rstate["aprobado_jefe"] or 0)
            or int(rstate["aprobado"] or 0)
            or int(rstate["aprob_gf"] or 0)
        ):
            flash("La garantía fue aprobada recientemente y ya no puede editarse.", "warning")
            return redirect(url_for("contratos.contab_lista"))

        result = services.update_garantia_from_request(garantia_id, row_db["contrato_id"])
        if not result["ok"]:
            flash(result["message"], "danger")
            row_back = dict(row_db)
            row_back.update(result["row_back"])
            return render_template(
                "contabilidad_ingreso.html",
                mode="edit",
                row=row_back,
                contratos=contratos,
                archivos=archivos,
                post_url=url_for("contratos.contab_editar", garantia_id=garantia_id),
                back_url=request.form.get("next") or url_for("contratos.contab_lista"),
            )

        flash("Garantía actualizada.", "success")
        return redirect(url_for("contratos.contab_lista"))

    return render_template(
        "contabilidad_ingreso.html",
        mode="edit",
        row=dict(row_db),
        contratos=contratos,
        archivos=archivos,
        post_url=url_for("contratos.contab_editar", garantia_id=garantia_id),
        back_url=request.args.get("next") or url_for("contratos.contab_lista"),
    )


@contratos_bp.route("/contab/<int:garantia_id>/eliminar", methods=["POST"])
@require_login
@require_permission("contratos_garantias", "eliminar")
def contab_eliminar(garantia_id: int):
    services.ensure_softdelete_columns()
    row = repository.fetch_garantia_activa_por_id(garantia_id)
    if not row:
        abort(404)

    services.soft_delete_garantia(garantia_id)
    flash("Garantía eliminada.", "success")
    return redirect(url_for("contratos.contab_lista"))


@contratos_bp.route("/contab/<int:garantia_id>/aprobj/toggle", methods=["POST"])
@require_login
@require_permission("contratos_garantias", "aprobar")
def toggle_aprobacion_jefe_garantia(garantia_id: int):
    row = repository.fetch_garantia_activa_por_id(garantia_id)
    if not row:
        abort(404)

    services.toggle_jefe_garantia(garantia_id, session.get("usuario_id"))
    flash("Aprobación de Jefe (Garantía) actualizada.", "success")
    return redirect(request.referrer or url_for("contratos.contab_lista"))


@contratos_bp.route("/archivo/<int:archivo_id>/ver", methods=["GET"])
@require_login
@require_permission("contratos_ingresar", "ver")
def ver_archivo_contrato(archivo_id: int):
    services.ensure_contrato_archivos_table()
    row = repository.fetch_archivo_por_id(archivo_id)
    if not row:
        abort(404)

    folder = services.get_contratos_upload_folder()
    file_path = os.path.join(folder, row["filename"])
    if not os.path.exists(file_path):
        abort(404)

    return send_file(
        file_path,
        mimetype="application/pdf",
        as_attachment=False,
        download_name=row["original_name"] or "contrato.pdf",
    )


@contratos_bp.route("/archivo/<int:archivo_id>/descargar", methods=["GET"])
@require_login
@require_permission("contratos_ingresar", "exportar")
def descargar_archivo_contrato(archivo_id: int):
    services.ensure_contrato_archivos_table()
    row = repository.fetch_archivo_por_id(archivo_id)
    if not row:
        abort(404)

    folder = services.get_contratos_upload_folder()
    file_path = os.path.join(folder, row["filename"])
    if not os.path.exists(file_path):
        abort(404)

    return send_file(
        file_path,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=row["original_name"] or "contrato.pdf",
    )


@contratos_bp.route("/aprobacion", methods=["GET"])
@require_login
@require_permission("contratos_aprobaciones", "ver")
def aprobacion():
    from modules.security import has_permission

    session["active_page"] = "contratos_aprobaciones"
    services.ensure_aprob_gf_columns()
    services.ensure_softdelete_columns()

    prov, pedi, tipo, estado, renov = services.get_aprobacion_filters()
    rows = repository.list_aprobacion(
        proveedor=prov,
        pedido=pedi,
        tipo_pp=tipo,
        estado=estado,
        renovacion=renov,
    )
    filtros = {"proveedor": prov, "pedido": pedi, "tipo_pp": (tipo or "")}
    puede_aprobar = has_permission(session.get("rol"), "contratos_aprobaciones", "aprobar")

    return render_template(
        "consulta_aprobacion.html",
        rows=rows,
        filtros=filtros,
        puede_aprobar=puede_aprobar,
    )


@contratos_bp.route("/ver/contrato/<int:contrato_id>/fragment", methods=["GET"])
@require_login
def ver_contrato_fragment(contrato_id: int):
    get_flashed_messages()
    session.pop("_flashes", None)

    from modules.security import has_permission as _hp

    fila = repository.fetch_contrato_por_id(contrato_id)
    if not fila:
        abort(404)

    vista, proveedor, archivos = services.build_contrato_vista(fila)
    can_exportar = _hp(session.get("rol"), "contratos_ingresar", "exportar")

    return render_template(
        "contrato_detalle_fragment.html",
        vista=vista,
        proveedor=proveedor,
        archivos=archivos,
        can_exportar=can_exportar,
    )


@contratos_bp.route("/ver/contrato/<int:contrato_id>/full", methods=["GET"])
@require_login
def ver_contrato_full(contrato_id: int):
    get_flashed_messages()
    session.pop("_flashes", None)

    from modules.security import has_permission as _hp

    fila = repository.fetch_contrato_por_id(contrato_id)
    if not fila:
        abort(404)

    vista, proveedor, archivos = services.build_contrato_vista(fila)
    can_exportar = _hp(session.get("rol"), "contratos_ingresar", "exportar")

    return render_template(
        "contrato_detalle.html",
        vista=vista,
        proveedor=proveedor,
        archivos=archivos,
        can_exportar=can_exportar,
    )


@contratos_bp.route("/ver/contrato/<int:contrato_id>", methods=["GET"])
@require_login
def ver_contrato(contrato_id: int):
    return ver_contrato_fragment(contrato_id)


@contratos_bp.route("/ver/garantia/<int:garantia_id>", methods=["GET"])
@require_login
@require_permission("contratos_aprobaciones", "ver")
def ver_garantia(garantia_id: int):
    services.ensure_aprob_gf_columns()

    get_flashed_messages()
    session.pop("_flashes", None)

    g = repository.fetch_garantia_activa_por_id(garantia_id)
    if not g:
        abort(404)

    vista, pedido, _proveedor = services.build_garantia_vista(g)

    return render_template(
        "garantia_detalle.html",
        vista=vista,
        pedido=pedido,
        hide_flashes=True,
    )


@contratos_bp.route("/aprobacion/toggle_gf", methods=["POST"])
@require_login
@require_permission("contratos_aprobaciones", "aprobar")
def toggle_aprobacion_gf():
    payload = request.get_json(silent=True) or {}
    tipo = (payload.get("tipo") or "").strip().lower()
    rec_id = payload.get("id")
    valor = 1 if payload.get("valor") else 0

    if tipo != "contrato":
        abort(400)

    try:
        contrato_id = int(rec_id)
    except Exception:
        abort(400)

    services.ensure_aprob_gf_columns()
    services.toggle_aprobacion_gf(contrato_id, valor)
    return {"ok": True, "value": valor}


@contratos_bp.route("/aprobacion/contrato/<int:contrato_id>/toggle", methods=["POST"])
@require_login
@require_permission("contratos_aprobaciones", "aprobar")
def toggle_aprobar_contrato(contrato_id: int):
    row = repository.fetch_contrato_por_id(contrato_id)
    if not row:
        abort(404)

    services.toggle_aprobar_contrato(contrato_id, session.get("usuario_id"))
    flash("Contrato actualizado.", "success")
    return redirect(request.referrer or url_for("contratos.aprobacion"))


@contratos_bp.route("/aprobacion/garantia/<int:garantia_id>/toggle", methods=["POST"])
@require_login
@require_permission("contratos_aprobaciones", "aprobar")
def toggle_aprobar_garantia(garantia_id: int):
    row = repository.fetch_garantia_activa_por_id(garantia_id)
    if not row:
        abort(404)

    services.toggle_aprobar_garantia(garantia_id, session.get("usuario_id"))
    flash("Garantía actualizada.", "success")
    return redirect(request.referrer or url_for("contratos.aprobacion"))


@contratos_bp.route("/contab/exportar", methods=["GET"])
@require_login
@require_permission("contratos_garantias", "exportar")
def exportar_garantias_contab():
    prov = (request.args.get("proveedor") or "").strip()
    pedi = (request.args.get("pedido") or "").strip()
    estado = (request.args.get("estado") or "").strip()

    requiere = services.get_requiere_renovacion_filter(
        (request.args.get("renovacion") or request.args.get("requiere_renovacion") or "")
    )

    fecha_registro_desde = (request.args.get("fecha_registro_desde") or "").strip()
    fecha_registro_hasta = (request.args.get("fecha_registro_hasta") or "").strip()
    fecha_vencimiento_desde = (request.args.get("fecha_vencimiento_desde") or "").strip()
    fecha_vencimiento_hasta = (request.args.get("fecha_vencimiento_hasta") or "").strip()

    rows = repository.list_garantias_reporte(
        proveedor=prov,
        pedido=pedi,
        estado=estado,
        requiere_renovacion=requiere,
        fecha_registro_desde=fecha_registro_desde,
        fecha_registro_hasta=fecha_registro_hasta,
        fecha_vencimiento_desde=fecha_vencimiento_desde,
        fecha_vencimiento_hasta=fecha_vencimiento_hasta,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Garantías"

    headers = [
        "ID Garantía",
        "ID Contrato",
        "Pedido",
        "Proveedor",
        "Objeto contrato",
        "Valor contrato",
        "Valor anticipo",
        "Tipo contrato",
        "Fecha suscripción contrato",
        "Fecha terminación contrato",
        "Estado contrato",
        "Aprob. jefe contrato",
        "Aprob. gerencia contrato",
        "Aprob. GF contrato",

        "Tipo garantía",
        "Compañía emisora",
        "Monto póliza",
        "Fecha suscripción garantía",
        "Fecha vencimiento garantía",
        "Fecha vencimiento actual",
        "Vigencia días",
        "Estado garantía",
        "Fecha renovación",
        "Requiere renovación",
        "Estado aprobación garantía",
        "Aprob. jefe garantía",
        "Aprob. gerencia garantía",
        "Aprob. GF garantía",
        "Fecha registro garantía",
        "Fecha actualización garantía",
        "Observaciones garantía",
    ]

    ws.append(headers)

    for row in rows:
        ws.append([
            row["garantia_id"],
            row["contrato_id"],
            row["pedido"],
            row["proveedor"],
            row["objeto"],
            float(row["valor_contrato"] or 0),
            float(row["valor_anticipo"] or 0),
            row["tipo_pp"],
            row["contrato_fecha_suscripcion"],
            row["contrato_fecha_terminacion"],
            row["contrato_status_interno"],
            "Sí" if int(row["contrato_aprobado_jefe"] or 0) else "No",
            "Sí" if int(row["contrato_aprobado"] or 0) else "No",
            "Sí" if int(row["contrato_aprob_gf"] or 0) else "No",

            row["garantia_tipo"],
            row["compania_emisora"],
            float(row["monto_poliza"] or 0),
            row["garantia_fecha_suscripcion"],
            row["garantia_fecha_vencimiento"],
            row["fecha_vencimiento_actual"],
            row["vigencia_dias"],
            row["garantia_estado"],
            row["fecha_renovacion"],
            "Sí" if int(row["requiere_renovacion"] or 0) else "No",
            row["garantia_status_interno"],
            "Sí" if int(row["garantia_aprobado_jefe"] or 0) else "No",
            "Sí" if int(row["garantia_aprobado"] or 0) else "No",
            "Sí" if int(row["garantia_aprob_gf"] or 0) else "No",
            row["garantia_creado_at"],
            row["garantia_actualizado_at"],
            row["garantia_observaciones"],
        ])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 45)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="reporte_garantias_contratos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )