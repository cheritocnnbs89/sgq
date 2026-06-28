from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
import io
import smtplib
import secrets
import logging
from flask import url_for, current_app
from modules.tasks.task_notifications import _send_encuesta_satisfaccion_mail

_log = logging.getLogger(__name__)
import pandas as pd
from flask import session
from openpyxl.styles import Font

from modules.config import ESTADOS
from modules.db import get_db, get_config_value
from modules.tasks.task_notifications import _send_tarea_avance_mail
from modules.tasks.task_repository import (
    repo_obtener_tarea_para_encuesta,
    repo_obtener_resultado_encuesta_email,
    repo_obtener_encuesta_por_tarea,
    repo_insertar_encuesta,
    repo_marcar_encuesta_enviada,
    repo_obtener_encuesta_por_token,
    repo_insertar_respuesta_encuesta,
    repo_finalizar_encuesta,
    repo_listar_encuestas,
    repo_obtener_departamentos_tareas,
    repo_dashboard_tareas,
    repo_eliminar_tarea,
    repo_es_responsable_tarea,
    repo_find_user_id_by_email,
    repo_insertar_tarea,
    repo_insertar_tarea_accion,
    repo_insertar_tarea_responsable_si_no_existe,
    repo_listar_tarea_responsables_map,
    repo_listar_tareas_raw,
    repo_obtener_accion_con_tarea,
    repo_obtener_email_usuario,
    repo_obtener_emails_responsables_tarea,
    repo_obtener_empresas_activas,
    repo_obtener_responsables,
    repo_obtener_solicitantes,
    repo_obtener_tarea_acciones,
    repo_obtener_tarea_detalle,
    repo_obtener_tarea_edicion,
    repo_obtener_tipos_tarea,
    repo_obtener_usuarios_activos,
    repo_actualizar_tarea,
    repo_finalizar_accion,
)
from email.message import EmailMessage
from html import escape
from modules.tasks import task_queries as q
from modules.tasks.task_utils import parse_dt, _extract_email, _norm_email


ESTADOS_NO_EDITABLES = ("Terminado", "Cerrado por sistema")
PREGUNTAS_ENCUESTA = [
    "¿Cómo califica la atención recibida?",
    "¿El tiempo de respuesta fue adecuado?",
    "¿La solución entregada resolvió su requerimiento?",
    "¿El técnico comunicó claramente el avance o solución?",
    "¿Qué tan satisfecho está con la gestión general del ticket?",
]


def svc_crear_y_enviar_encuesta(task_id: int):
    tarea = repo_obtener_tarea_para_encuesta(task_id)
    if not tarea:
        _log.warning("[encuesta] tarea %s no encontrada", task_id)
        return False

    if not tarea.get("solicitante_id"):
        _log.warning("[encuesta] tarea %s sin solicitante_id", task_id)
        return False

    if not tarea.get("solicitante_email"):
        _log.warning(
            "[encuesta] tarea %s: solicitante_id=%s no tiene email registrado en la BD",
            task_id, tarea.get("solicitante_id")
        )
        return False

    existente = repo_obtener_encuesta_por_tarea(task_id)
    conn = get_db()

    if existente:
        estado_enc = (existente.get("estado") or "").strip()

        if estado_enc == "Realizada":
            _log.info("[encuesta] tarea %s ya tiene encuesta respondida (id=%s), no se reenvía", task_id, existente.get("id"))
            return False

        # Verificar si la encuesta es de un cierre anterior (registro obsoleto)
        fecha_cierre = tarea.get("fecha_cierre_real")
        fecha_enc    = existente.get("fecha_creacion") or existente.get("fecha_envio")
        es_obsoleta  = False
        if fecha_cierre and fecha_enc:
            try:
                # Normalizar a datetime para comparar
                if hasattr(fecha_cierre, 'date'):
                    fc = fecha_cierre
                else:
                    fc = datetime.fromisoformat(str(fecha_cierre))
                if hasattr(fecha_enc, 'date'):
                    fe = fecha_enc
                else:
                    fe = datetime.fromisoformat(str(fecha_enc)[:19])
                # Obsoleta si la encuesta fue creada más de 1 hora antes del cierre actual
                if (fc - fe).total_seconds() > 3600:
                    es_obsoleta = True
            except Exception as ex:
                _log.warning("[encuesta] No se pudo comparar fechas: %s", ex)

        if es_obsoleta:
            _log.info("[encuesta] tarea %s: encuesta id=%s es obsoleta (creada %s, tarea cerrada %s) — se reemplaza",
                      task_id, existente.get("id"), fecha_enc, fecha_cierre)
            try:
                cur = conn.cursor()
                cur.execute("DELETE FROM encuestas_satisfaccion WHERE id = ?", (existente.get("id"),))
                conn.commit()
            except Exception as ex:
                _log.warning("[encuesta] No se pudo borrar encuesta obsoleta: %s", ex)
            # Crear nueva encuesta abajo
            token = secrets.token_urlsafe(32)
            try:
                encuesta_id = repo_insertar_encuesta(conn, task_id, tarea.get("solicitante_id"), token)
                conn.commit()
            except Exception as e:
                conn.rollback()
                _log.exception("[encuesta] Error creando nueva encuesta tras borrar obsoleta: %s", e)
                return False
        else:
            # Encuesta pendiente reciente — reenviar el correo con el token existente
            _log.info("[encuesta] tarea %s tiene encuesta pendiente (id=%s), reenviando correo", task_id, existente.get("id"))
            encuesta_id = existente.get("id")
            token = existente.get("token")
            if not token:
                _log.warning("[encuesta] tarea %s: encuesta sin token", task_id)
                return False
    else:
        token = secrets.token_urlsafe(32)
        try:
            encuesta_id = repo_insertar_encuesta(
                conn,
                task_id,
                tarea.get("solicitante_id"),
                token,
            )
            conn.commit()
        except Exception as e:
            conn.rollback()
            _log.exception("[encuesta] Error insertando encuesta para tarea %s: %s", task_id, e)
            return False

    try:
        # url_for requiere contexto de aplicación; fallback a URL manual si falla
        try:
            encuesta_url = url_for("responder_encuesta", token=token, _external=True)
        except RuntimeError:
            base = get_config_value("app_base_url", "").rstrip("/")
            encuesta_url = f"{base}/encuestas/responder/{token}"
            _log.warning("[encuesta] url_for falló, usando base_url=%s", base)

        msg, smtp_host, smtp_port, smtp_user, smtp_pass, use_tls = _send_encuesta_satisfaccion_mail(
            tarea,
            encuesta_url,
        )

        if not msg:
            _log.warning("[encuesta] _send_encuesta_satisfaccion_mail devolvió msg=None — revisar config SMTP (smtp_host, smtp_from)")
            return False

        if not smtp_host:
            _log.warning("[encuesta] smtp_host no configurado")
            return False

        msg["To"] = tarea["solicitante_email"]
        _log.info("[encuesta] Enviando encuesta tarea=%s a=%s via %s:%s tls=%s",
                  task_id, tarea["solicitante_email"], smtp_host, smtp_port, use_tls)

        if use_tls:
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls()
                if smtp_user and smtp_pass:
                    server.login(smtp_user, smtp_pass)
                server.send_message(msg)
        else:
            with smtplib.SMTP_SSL(smtp_host, smtp_port) as server:
                if smtp_user and smtp_pass:
                    server.login(smtp_user, smtp_pass)
                server.send_message(msg)

        conn = get_db()
        repo_marcar_encuesta_enviada(conn, encuesta_id)
        conn.commit()
        _log.info("[encuesta] Encuesta enviada OK tarea=%s encuesta_id=%s", task_id, encuesta_id)
        return True

    except smtplib.SMTPAuthenticationError as e:
        _log.error("[encuesta] Error de autenticación SMTP tarea=%s: %s", task_id, e)
        return False
    except smtplib.SMTPException as e:
        _log.error("[encuesta] Error SMTP tarea=%s: %s", task_id, e)
        return False
    except Exception as e:
        _log.exception("[encuesta] Error inesperado enviando encuesta tarea=%s: %s", task_id, e)
        return False


def _csv_ids_to_set(value):
    if value is None:
        return set()

    result = set()

    for part in str(value).split(","):
        part = part.strip()
        if not part:
            continue

        try:
            result.add(int(part))
        except ValueError:
            continue

    return result


def _safe_float(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    value = str(value).strip()

    if not value:
        return None

    try:
        return float(value)
    except ValueError:
        return None


def _safe_int(value):
    if value is None:
        return None

    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _puede_ver_encuesta(user, row):
    user_id = _safe_int(user.get("id"))
    rol = (user.get("rol") or "").lower().strip()

    if not user_id:
        return False

    if rol == "admin":
        return True

    solicitante_id = _safe_int(row.get("solicitante_id"))
    responsable_id = _safe_int(row.get("responsable_id"))
    responsable_jefe_id = _safe_int(row.get("responsable_jefe_id"))

    responsable_ids = _csv_ids_to_set(row.get("responsable_ids_csv"))
    jefe_ids = _csv_ids_to_set(row.get("jefe_ids_csv"))

    # Solicitante: ve encuestas de tareas que solicitó
    if solicitante_id == user_id:
        return True

    # Técnico: ve encuestas de sus tareas
    if responsable_id == user_id:
        return True

    if user_id in responsable_ids:
        return True

    # Jefe: ve encuestas de su personal
    if rol == "jefe":
        if responsable_jefe_id == user_id:
            return True

        if user_id in jefe_ids:
            return True

    return False


def svc_build_encuestas_context(user, request_args):
    estado = (request_args.get("estado") or "").strip()
    q_text = (request_args.get("q") or "").lower().strip()

    encuestas_raw = repo_listar_encuestas()
    visibles = []

    for row in encuestas_raw:

        r = dict(row)

        if not _puede_ver_encuesta(user, r):
            continue

        if estado and r.get("estado") != estado:
            continue

        texto = " ".join(str(v or "") for v in r.values()).lower()
        if q_text and q_text not in texto:
            continue

        promedio = _safe_float(r.get("promedio"))
        r["promedio_fmt"] = f"{promedio:.1f}" if promedio is not None else "—"

        if r.get("responsables_nombre_csv"):
            r["responsable_nombre"] = r["responsables_nombre_csv"]
        

        tarea_id = _safe_int(r.get("tarea_id"))
        r["tarea_id"] = tarea_id
        r["codigo_tarea"] = f"{tarea_id:07d}" if tarea_id is not None else "—"

        solicitante_id = _safe_int(r.get("solicitante_id"))
        user_id = _safe_int(user.get("id"))

        r["puede_responder"] = (
            user_id is not None
            and solicitante_id == user_id
            and r.get("estado") == "Pendiente"
            and bool(r.get("token"))
        )

        r["puede_crear_encuesta"] = (
            user_id is not None
            and solicitante_id == user_id
            and r.get("estado") == "Pendiente"
            and not r.get("encuesta_id")
        )

        visibles.append(r)

    pendientes = sum(1 for x in visibles if x.get("estado") == "Pendiente")
    realizadas = sum(1 for x in visibles if x.get("estado") == "Realizada")

    return {
        "usuario": user["username"],
        "rol": user["rol"],
        "encuestas": visibles,
        "estado_sel": estado,
        "q": q_text,
        "total": len(visibles),
        "pendientes": pendientes,
        "realizadas": realizadas,
        "active_page": "encuestas",
    }

def svc_build_responder_encuesta_context(token: str):
    encuesta = repo_obtener_encuesta_por_token(token)
    if not encuesta:
        return {"ok": False, "message": "Encuesta no encontrada."}

    return {
        "ok": True,
        "encuesta": encuesta,
        "preguntas": PREGUNTAS_ENCUESTA,
        "active_page": "encuestas_publica",
    }

def _smtp_bool(value):
    return str(value or "").strip().lower() in ("1", "true", "yes", "si", "sí", "on")


def _smtp_config():
    smtp_host = (
        get_config_value("smtp_host", "")
        or get_config_value("SMTP_HOST", "")
        or ""
    ).strip()

    smtp_port_raw = (
        get_config_value("smtp_port", "")
        or get_config_value("SMTP_PORT", "")
        or "587"
    )

    try:
        smtp_port = int(smtp_port_raw)
    except Exception:
        smtp_port = 587

    smtp_user = (
        get_config_value("smtp_user", "")
        or get_config_value("SMTP_USER", "")
        or ""
    ).strip()

    smtp_pass = (
        get_config_value("smtp_pass", "")
        or get_config_value("SMTP_PASS", "")
        or ""
    ).strip()

    smtp_from = (
        get_config_value("smtp_from", "")
        or get_config_value("SMTP_FROM", "")
        or smtp_user
        or "no-reply@sili.local"
    ).strip()

    use_tls = _smtp_bool(
        get_config_value("smtp_use_tls", "")
        or get_config_value("SMTP_USE_TLS", "")
        or "1"
    )

    return smtp_host, smtp_port, smtp_user, smtp_pass, smtp_from, use_tls


def _fmt_fecha_correo(value):
    if not value:
        return "—"

    try:
        return value.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(value)


def _fmt_promedio(value):
    promedio = _safe_float(value)
    return f"{promedio:.1f}" if promedio is not None else "—"


def _color_promedio(promedio):
    value = _safe_float(promedio)

    if value is None:
        return "#16a34a"

    if value >= 4:
        return "#16a34a"

    if value >= 3:
        return "#f59e0b"

    return "#dc2626"


def _calificacion_texto(promedio):
    value = _safe_float(promedio)

    if value is None:
        return "Sin promedio"

    if value >= 4.5:
        return "Excelente"

    if value >= 4:
        return "Muy bueno"

    if value >= 3:
        return "Regular"

    return "Requiere atención"


def _build_resultado_encuesta_email(row):
    promedio = _safe_float(row.get("promedio"))
    promedio_fmt = _fmt_promedio(promedio)
    color = _color_promedio(promedio)
    calificacion = _calificacion_texto(promedio)

    codigo = f"{_safe_int(row.get('tarea_id')) or 0:07d}"
    titulo = escape(str(row.get("titulo") or "—"))
    solicitante = escape(str(row.get("solicitante_nombre") or "—"))
    responsable = escape(str(row.get("responsable_nombre") or "—"))
    comentario = escape(str(row.get("comentario") or "Sin comentario adicional.")).replace("\n", "<br>")
    fecha_respuesta = escape(_fmt_fecha_correo(row.get("fecha_respuesta")))
    fecha_cierre = escape(_fmt_fecha_correo(row.get("fecha_cierre_real")))

    p1 = row.get("p1") or "—"
    p2 = row.get("p2") or "—"
    p3 = row.get("p3") or "—"
    p4 = row.get("p4") or "—"
    p5 = row.get("p5") or "—"

    subject = f"Resultado encuesta de satisfacción — Tarea {codigo}"

    html = f"""
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <title>{escape(subject)}</title>
</head>
<body style="margin:0;padding:0;background:#f3f4f6;font-family:Arial,Helvetica,sans-serif;color:#111827;">
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#f3f4f6;padding:22px 0;">
    <tr>
      <td align="center">
        <table role="presentation" width="680" cellspacing="0" cellpadding="0" style="max-width:680px;width:94%;background:#ffffff;border-radius:10px;overflow:hidden;border:1px solid #e5e7eb;">
          <tr>
            <td style="background:#16a34a;color:#ffffff;padding:18px 22px;">
              <div style="font-size:12px;font-weight:800;letter-spacing:.05em;text-transform:uppercase;">
                Encuesta de satisfacción
              </div>
              <div style="font-size:24px;font-weight:800;margin-top:5px;line-height:1.25;">
                Resultado recibido — Tarea {codigo}
              </div>
              <div style="font-size:13px;margin-top:6px;opacity:.95;">
                El solicitante respondió la encuesta de satisfacción.
              </div>
            </td>
          </tr>

          <tr>
            <td style="padding:20px 22px;">
              <p style="margin:0 0 12px;font-size:14px;line-height:1.55;">
                Hola <strong>{responsable}</strong>,
              </p>

              <p style="margin:0 0 16px;font-size:14px;line-height:1.55;">
                Se registró una respuesta de encuesta para la tarea atendida. A continuación se muestra el resultado:
              </p>

              <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;margin:0 0 16px;">
                <tr>
                  <td style="background:#ecfdf5;border:1px solid #bbf7d0;border-radius:10px;padding:14px;text-align:center;">
                    <div style="font-size:12px;color:#166534;font-weight:800;text-transform:uppercase;letter-spacing:.05em;">
                      Promedio
                    </div>
                    <div style="font-size:38px;font-weight:900;color:{color};line-height:1;margin-top:6px;">
                      {promedio_fmt}
                    </div>
                    <div style="font-size:13px;color:#166534;font-weight:700;margin-top:6px;">
                      {escape(calificacion)}
                    </div>
                  </td>
                </tr>
              </table>

              <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;font-size:13px;">
                <tr>
                  <td style="width:34%;background:#ecfdf5;border-bottom:1px solid #e5e7eb;padding:10px;font-weight:700;color:#064e3b;">Código</td>
                  <td style="border-bottom:1px solid #e5e7eb;padding:10px;">{codigo}</td>
                </tr>
                <tr>
                  <td style="background:#ecfdf5;border-bottom:1px solid #e5e7eb;padding:10px;font-weight:700;color:#064e3b;">Título</td>
                  <td style="border-bottom:1px solid #e5e7eb;padding:10px;">{titulo}</td>
                </tr>
                <tr>
                  <td style="background:#ecfdf5;border-bottom:1px solid #e5e7eb;padding:10px;font-weight:700;color:#064e3b;">Solicitante</td>
                  <td style="border-bottom:1px solid #e5e7eb;padding:10px;">{solicitante}</td>
                </tr>
                <tr>
                  <td style="background:#ecfdf5;border-bottom:1px solid #e5e7eb;padding:10px;font-weight:700;color:#064e3b;">Responsable</td>
                  <td style="border-bottom:1px solid #e5e7eb;padding:10px;">{responsable}</td>
                </tr>
                <tr>
                  <td style="background:#ecfdf5;border-bottom:1px solid #e5e7eb;padding:10px;font-weight:700;color:#064e3b;">Fecha cierre</td>
                  <td style="border-bottom:1px solid #e5e7eb;padding:10px;">{fecha_cierre}</td>
                </tr>
                <tr>
                  <td style="background:#ecfdf5;border-bottom:1px solid #e5e7eb;padding:10px;font-weight:700;color:#064e3b;">Fecha respuesta</td>
                  <td style="border-bottom:1px solid #e5e7eb;padding:10px;">{fecha_respuesta}</td>
                </tr>
                <tr>
                  <td style="background:#ecfdf5;border-bottom:1px solid #e5e7eb;padding:10px;font-weight:700;color:#064e3b;">Puntuaciones</td>
                  <td style="border-bottom:1px solid #e5e7eb;padding:10px;">
                    P1: <strong>{p1}</strong> &nbsp; 
                    P2: <strong>{p2}</strong> &nbsp; 
                    P3: <strong>{p3}</strong> &nbsp; 
                    P4: <strong>{p4}</strong> &nbsp; 
                    P5: <strong>{p5}</strong>
                  </td>
                </tr>
                <tr>
                  <td style="background:#ecfdf5;padding:10px;font-weight:700;color:#064e3b;">Comentario</td>
                  <td style="padding:10px;line-height:1.5;">{comentario}</td>
                </tr>
              </table>

              <div style="margin-top:18px;">
                <a href="{escape(url_for('listar_encuestas', _external=True))}"
                   style="display:inline-block;background:#16a34a;color:#ffffff;text-decoration:none;padding:10px 16px;border-radius:8px;font-size:13px;font-weight:800;">
                  Ver encuestas
                </a>
              </div>

              <p style="margin:18px 0 0;font-size:12px;color:#6b7280;">
                Este correo fue generado automáticamente por el sistema de Sili.
              </p>
            </td>
          </tr>

          <tr>
            <td style="border-top:1px solid #e5e7eb;padding:12px 22px;color:#9ca3af;font-size:11px;">
              Este es un mensaje automático. No responda a este correo.
            </td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>
</html>
"""

    text = f"""
Resultado de encuesta de satisfacción

Tarea: {codigo}
Título: {row.get("titulo") or "—"}
Solicitante: {row.get("solicitante_nombre") or "—"}
Responsable: {row.get("responsable_nombre") or "—"}
Promedio: {promedio_fmt} ({calificacion})
Puntuaciones: P1={p1}, P2={p2}, P3={p3}, P4={p4}, P5={p5}
Comentario: {row.get("comentario") or "Sin comentario adicional."}
Fecha respuesta: {_fmt_fecha_correo(row.get("fecha_respuesta"))}

Este correo fue generado automáticamente por el sistema de Sili.
"""

    return subject, text, html


def svc_enviar_resultado_encuesta(encuesta_id: int):
    row = repo_obtener_resultado_encuesta_email(encuesta_id)

    if not row:
        print("No se encontró resultado de encuesta para enviar correo:", encuesta_id)
        return False

    destinatarios = set()

    responsable_email = (row.get("responsable_email") or "").strip()
    jefe_email = (row.get("jefe_email") or "").strip()

    if responsable_email:
        destinatarios.add(responsable_email)

    # Copia al jefe del técnico, si existe.
    if jefe_email:
        destinatarios.add(jefe_email)

    if not destinatarios:
        print("No hay destinatarios para resultado de encuesta:", encuesta_id)
        return False

    smtp_host, smtp_port, smtp_user, smtp_pass, smtp_from, use_tls = _smtp_config()

    if not smtp_host:
        print("No está configurado smtp_host. No se envía resultado de encuesta.")
        return False

    subject, text, html = _build_resultado_encuesta_email(row)

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = smtp_from
    msg["To"] = ", ".join(sorted(destinatarios))
    msg.set_content(text)
    msg.add_alternative(html, subtype="html")

    try:
        if use_tls:
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls()
                if smtp_user and smtp_pass:
                    server.login(smtp_user, smtp_pass)
                server.send_message(msg)
        else:
            with smtplib.SMTP_SSL(smtp_host, smtp_port) as server:
                if smtp_user and smtp_pass:
                    server.login(smtp_user, smtp_pass)
                server.send_message(msg)

        return True

    except Exception as e:
        print("Error enviando resultado de encuesta:", e)
        return False
def svc_guardar_respuesta_encuesta(token: str, form):
    encuesta = repo_obtener_encuesta_por_token(token)
    if not encuesta:
        return {"ok": False, "message": "Encuesta no encontrada.", "category": "danger"}

    if encuesta["estado"] == "Realizada":
        return {"ok": False, "message": "Esta encuesta ya fue respondida.", "category": "warning"}

    comentario = (form.get("comentario") or "").strip()
    respuestas = []

    for i in range(1, 6):
        raw = (form.get(f"pregunta_{i}") or "").strip()
        try:
            value = int(raw)
        except ValueError:
            return {"ok": False, "message": "Debe responder todas las preguntas.", "category": "warning"}

        if value < 1 or value > 5:
            return {"ok": False, "message": "Las calificaciones deben estar entre 1 y 5.", "category": "warning"}

        respuestas.append((i, value))

    conn = get_db()

    try:
        for pregunta_numero, puntuacion in respuestas:
            repo_insertar_respuesta_encuesta(conn, encuesta["id"], pregunta_numero, puntuacion)

        changed = repo_finalizar_encuesta(conn, encuesta["id"], comentario)
        conn.commit()

        if not changed:
            return {"ok": False, "message": "No se pudo registrar la encuesta.", "category": "warning"}

        # No bloquea el guardado si falla el correo.
        svc_enviar_resultado_encuesta(encuesta["id"])

        return {"ok": True, "message": "Gracias por responder la encuesta.", "category": "success"}

    except Exception as e:
        conn.rollback()
        return {"ok": False, "message": f"Error guardando encuesta: {str(e)}", "category": "danger"}
class TaskServiceError(Exception):
    def __init__(self, message: str, category: str = "danger"):
        super().__init__(message)
        self.message = message
        self.category = category


#def svc_build_dashboard_context(user):
def svc_build_dashboard_context(user, request_args=None):    
    tareas_raw = repo_dashboard_tareas(user)
    request_args = request_args or {}

    fecha_desde_raw = (request_args.get("fecha_desde") or "").strip()
    fecha_hasta_raw = (request_args.get("fecha_hasta") or "").strip()
    depto_sel = (request_args.get("depto") or "").strip()

    fecha_desde = parse_dt(fecha_desde_raw) if fecha_desde_raw else None
    fecha_hasta = parse_dt(fecha_hasta_raw) if fecha_hasta_raw else None

    if fecha_hasta:
        fecha_hasta = fecha_hasta.replace(hour=23, minute=59, second=59)

    tareas_filtradas = []

    for r in tareas_raw:
        t = dict(r)

        if depto_sel and (t.get("departamento") or "") != depto_sel:
            continue

        fecha_base = parse_dt(str(t.get("fecha_cierre_real") or t.get("fecha_creacion") or "").strip())

        if fecha_desde and (not fecha_base or fecha_base < fecha_desde):
            continue

        if fecha_hasta and (not fecha_base or fecha_base > fecha_hasta):
            continue

        tareas_filtradas.append(t)

    tareas_raw = tareas_filtradas

    hoy = date.today()
    week_end = hoy + timedelta(days=(6 - hoy.weekday()))

    conteos = {estado: 0 for estado in ESTADOS}

    overdue_tasks = []
    today_tasks = []
    week_tasks = []
    no_due_tasks = []

    overdue_by_user_count = defaultdict(int)
    overdue_by_depto_count = defaultdict(int)
    compromiso_counts = defaultdict(int)
    horas_cerradas_por_usuario_count = defaultdict(float)
    tareas_cerradas_por_usuario_count = defaultdict(int)
    horas_por_dia_count = defaultdict(float)
    horas_por_depto_count = defaultdict(float)
    tareas_por_depto_mes_count: dict = defaultdict(lambda: defaultdict(int))
    cumplimiento_mensual_count: dict = defaultdict(lambda: {"a_tiempo": 0, "atrasadas": 0})

    for r in tareas_raw:
        t = dict(r)
        estado = t["estado"]

        if estado in conteos:
            conteos[estado] += 1
        else:
            conteos[estado] = conteos.get(estado, 0) + 1

        comp_dt  = parse_dt(t.get("fecha_compromiso"))
        depto_t  = t.get("departamento") or "Sin departamento"

        # Horas = fecha_inicio → fecha_fin (tiempo planificado invertido en la tarea)
        inicio_dt = parse_dt(t.get("fecha_inicio"))
        fin_dt    = parse_dt(t.get("fecha_fin"))

        if estado in ("Terminado", "Cerrado por sistema") and inicio_dt and fin_dt:
            horas = round((fin_dt - inicio_dt).total_seconds() / 3600, 2)
            if horas >= 0:
                owner = t.get("propietario") or "—"
                horas_cerradas_por_usuario_count[owner] += horas
                tareas_cerradas_por_usuario_count[owner] += 1
                horas_por_dia_count[inicio_dt.strftime("%Y-%m-%d")] += horas
                horas_por_depto_count[depto_t] += horas

        # Tareas por departamento por mes (todas las tareas)
        fecha_ref = inicio_dt or parse_dt(str(t.get("fecha_creacion") or "").strip())
        if fecha_ref:
            tareas_por_depto_mes_count[depto_t][fecha_ref.strftime("%Y-%m")] += 1

        # Cumplimiento mensual: cerradas a tiempo vs tardías
        if estado in ("Terminado", "Cerrado por sistema"):
            cierre_dt_c = parse_dt(str(t.get("fecha_cierre_real") or "").strip())
            if cierre_dt_c:
                mes_c = cierre_dt_c.strftime("%Y-%m")
                if comp_dt and cierre_dt_c > comp_dt:
                    cumplimiento_mensual_count[mes_c]["atrasadas"] += 1
                else:
                    cumplimiento_mensual_count[mes_c]["a_tiempo"] += 1

        if comp_dt:
            compromiso_counts[comp_dt.date()] += 1

        t["fecha_compromiso_fmt"] = comp_dt.strftime("%Y-%m-%d %H:%M") if comp_dt else ""
        no_terminada = estado not in ("Terminado", "Cerrado por sistema")

        if comp_dt and no_terminada:
            comp_date = comp_dt.date()
            if comp_date < hoy:
                t["dias_atraso"] = (hoy - comp_date).days
                overdue_tasks.append(t)

                owner = t.get("propietario") or "—"
                depto = t.get("departamento") or "Sin departamento"
                overdue_by_user_count[owner] += 1
                overdue_by_depto_count[depto] += 1
            elif comp_date == hoy:
                t["dias_atraso"] = 0
                today_tasks.append(t)
            elif hoy < comp_date <= week_end:
                t["dias_atraso"] = 0
                week_tasks.append(t)
        elif not comp_dt and no_terminada:
            t["dias_atraso"] = None
            no_due_tasks.append(t)

    overdue_by_user = sorted(
        [{"usuario": u, "total": c} for u, c in overdue_by_user_count.items()],
        key=lambda x: -x["total"],
    )

    overdue_by_depto = sorted(
        [{"departamento": d, "total": c} for d, c in overdue_by_depto_count.items()],
        key=lambda x: -x["total"],
    )

    horas_cerradas_por_usuario = sorted(
        [
            {
                "usuario": u,
                "horas": round(h, 2),
                "tareas": tareas_cerradas_por_usuario_count[u],
                "horas_por_tarea": round(h / tareas_cerradas_por_usuario_count[u], 2)
                    if tareas_cerradas_por_usuario_count[u] else 0,
            }
            for u, h in horas_cerradas_por_usuario_count.items()
        ],
        key=lambda x: -x["horas"],
    )

    dates_sorted = sorted(compromiso_counts.keys())

    # ── Horas por usuario (top 12, para gráfico) ──────────────
    _horas_usr_sorted = sorted(
        horas_cerradas_por_usuario_count.items(), key=lambda x: -x[1]
    )[:12]
    chart_horas_usuario = {
        "labels": [u for u, _ in _horas_usr_sorted],
        "data":   [round(h, 1) for _, h in _horas_usr_sorted],
    }

    # ── Horas por departamento ─────────────────────────────────
    horas_por_depto = sorted(
        [{"departamento": d, "horas": round(h, 2)}
         for d, h in horas_por_depto_count.items()],
        key=lambda x: -x["horas"],
    )
    _DONUT_COLORS = [
        "#3b82f6","#10b981","#f59e0b","#ef4444",
        "#8b5cf6","#ec4899","#06b6d4","#84cc16",
        "#f97316","#64748b","#a21caf","#0ea5e9",
    ]
    chart_horas_depto = {
        "labels": [d["departamento"] for d in horas_por_depto],
        "data":   [d["horas"] for d in horas_por_depto],
        "colors": [_DONUT_COLORS[i % len(_DONUT_COLORS)] for i in range(len(horas_por_depto))],
    }

    # ── Horas por día ──────────────────────────────────────────
    dias_sorted = sorted(horas_por_dia_count.keys())
    chart_horas_dia = {
        "labels": dias_sorted,
        "data":   [round(horas_por_dia_count[d], 1) for d in dias_sorted],
    }

    # ── Tareas por departamento por mes (top 7 deptos) ─────────
    _all_meses = sorted({m for dv in tareas_por_depto_mes_count.values() for m in dv})
    _top_deptos = sorted(
        tareas_por_depto_mes_count,
        key=lambda d: -sum(tareas_por_depto_mes_count[d].values()),
    )[:7]
    _COLORS = ["#3b82f6","#10b981","#f59e0b","#ef4444","#8b5cf6","#ec4899","#06b6d4"]
    chart_depto_mes = {
        "labels": _all_meses,
        "datasets": [
            {
                "label": dep,
                "data":  [tareas_por_depto_mes_count[dep].get(m, 0) for m in _all_meses],
                "color": _COLORS[i % len(_COLORS)],
            }
            for i, dep in enumerate(_top_deptos)
        ],
    }

    # ── Cumplimiento mensual ───────────────────────────────────
    _meses_c = sorted(cumplimiento_mensual_count.keys())
    chart_cumplimiento = {
        "labels":    _meses_c,
        "a_tiempo":  [cumplimiento_mensual_count[m]["a_tiempo"]  for m in _meses_c],
        "atrasadas": [cumplimiento_mensual_count[m]["atrasadas"] for m in _meses_c],
    }

    return {
        "usuario": user["username"],
        "deptos": repo_obtener_departamentos_tareas(),
        "depto_sel": depto_sel,
        "fecha_desde": fecha_desde_raw,
        "fecha_hasta": fecha_hasta_raw,
        "rol": user["rol"],
        "estados": ESTADOS,
        "conteos": conteos,
        "total": len(tareas_raw),
        "overdue_tasks": overdue_tasks,
        "today_tasks": today_tasks,
        "week_tasks": week_tasks,
        "no_due_tasks": no_due_tasks,
        "overdue_by_user": overdue_by_user,
        "overdue_by_depto": overdue_by_depto,
        "horas_cerradas_por_usuario": horas_cerradas_por_usuario,
        "horas_por_depto": horas_por_depto,
        "chart": {
            "status": {
                "labels": list(ESTADOS),
                "data": [conteos.get(e, 0) for e in ESTADOS],
            },
            "overdue_user": {
                "labels": [row["usuario"] for row in overdue_by_user],
                "data": [row["total"] for row in overdue_by_user],
            },
            "overdue_depto": {
                "labels": [row["departamento"] for row in overdue_by_depto],
                "data": [row["total"] for row in overdue_by_depto],
            },
            "timeline": {
                "labels": [d.isoformat() for d in dates_sorted],
                "data": [compromiso_counts[d] for d in dates_sorted],
            },
            "horas_dia":     chart_horas_dia,
            "horas_usuario": chart_horas_usuario,
            "horas_depto":   chart_horas_depto,
            "depto_mes":     chart_depto_mes,
            "cumplimiento":  chart_cumplimiento,
        },
        "active_page": "dashboard",
    }

def svc_build_listar_tareas_context(user, request_args):
    tareas_raw = repo_listar_tareas_raw()
    resp_rows = repo_listar_tarea_responsables_map()

    fecha_desde_raw = (request_args.get("fecha_desde") or "").strip()
    fecha_hasta_raw = (request_args.get("fecha_hasta") or "").strip()

    fecha_desde = parse_dt(fecha_desde_raw) if fecha_desde_raw else None
    fecha_hasta = parse_dt(fecha_hasta_raw) if fecha_hasta_raw else None

    if fecha_hasta:
        fecha_hasta = fecha_hasta.replace(hour=23, minute=59, second=59)

    responsables_map = {}
    for r in resp_rows:
        tid = r["tarea_id"]
        responsables_map.setdefault(tid, []).append(dict(r))

    for t in tareas_raw:
        resp_list = responsables_map.get(t["id"], [])
        ids = set()
        labels = []
        deptos = set()

        for r in resp_list:
            ids.add(r["usuario_id"])

            dept_name = (r["depto_nombre"] or "").strip()
            if dept_name:
                deptos.add(dept_name)

            nombre = (r["nombre_completo"] or "").strip()
            username = (r["username"] or "").strip()

            if nombre:
                labels.append(f"{nombre} ({username})")
            elif username:
                labels.append(username)

        if not ids:
            if t.get("usuario_id"):
                ids.add(t["usuario_id"])

            label_base = (t.get("creador_username") or "").strip()
            nombre_creador = (t.get("creador_nombre") or "").strip()

            if nombre_creador:
                labels.append(f"{nombre_creador} ({label_base})" if label_base else nombre_creador)
            elif label_base:
                labels.append(label_base)

        t["responsable_ids"] = ids
        t["propietario"] = ", ".join(labels) if labels else "Sin asignar"
        t["departamentos_responsables"] = ", ".join(sorted(deptos)) if deptos else "N/A"

    vista = (request_args.get("vista") or "realizar").lower()
    if vista not in ("realizar", "mis"):
        vista = "realizar"

    tareas_filtradas = []

    for t in tareas_raw:
        fecha_base = parse_dt(
            str(
                t.get("fecha_cierre_real")
                or t.get("fecha_fin")
                or t.get("fecha_inicio")
                or t.get("fecha_creacion")
                or ""
            ).strip()
        )

        if fecha_desde and (not fecha_base or fecha_base < fecha_desde):
            continue

        if fecha_hasta and (not fecha_base or fecha_base > fecha_hasta):
            continue

        if user["rol"] in ("admin", "jefe"):
            pasa_rol = True
        else:
            pasa_rol = (user["id"] in t["responsable_ids"]) or (t["creador_id"] == user["id"])

        if not pasa_rol:
            continue

        if vista == "realizar":
            if user["rol"] not in ("admin", "jefe") and user["id"] not in t["responsable_ids"]:
                continue
        else:
            if user["rol"] not in ("admin", "jefe") and t["creador_id"] != user["id"]:
                continue

        tareas_filtradas.append(t)

    try:
        edit_minutes_conf = float(get_config_value("edit_minutes", "5"))
    except Exception:
        edit_minutes_conf = 5.0

    tareas_list = []

    for t in tareas_filtradas:
        editable = False

        if user["rol"] == "admin":
            editable = True
        else:
            if user["id"] in t["responsable_ids"] and t["estado"] not in ESTADOS_NO_EDITABLES:
                editable = True

        t["editable"] = editable

        # Horas de atención = fecha_fin - fecha_inicio
        _fi = parse_dt(str(t.get("fecha_inicio") or "").strip())
        _ff = parse_dt(str(t.get("fecha_fin") or "").strip())
        if _fi and _ff:
            _h = round((_ff - _fi).total_seconds() / 3600, 2)
            t["horas_atencion"] = _h if _h >= 0 else None
        else:
            t["horas_atencion"] = None

        tareas_list.append(t)

    # Verificar si el usuario pertenece a SISTEMAS QP (para mostrar bandeja)
    es_sistemas_qp = False
    try:
        dep_id = user.get("departamento_id")
        if dep_id:
            conn = get_db()
            dep_row = conn.execute(
                "SELECT nombre FROM departamentos WHERE id = ?", (dep_id,)
            ).fetchone()
            if dep_row and dep_row["nombre"].strip().upper() == "SISTEMAS QP":
                es_sistemas_qp = True
    except Exception:
        pass

    # Contar correos pendientes en bandeja (solo si es SISTEMAS QP)
    bandeja_pendientes = 0
    if es_sistemas_qp:
        try:
            conn = get_db()
            row = conn.execute(
                "SELECT COUNT(*) AS c FROM email_tickets_inbox WHERE estado = 'POR_ASIGNAR'"
            ).fetchone()
            bandeja_pendientes = row["c"] if row else 0
        except Exception:
            pass

    return {
        "tareas": tareas_list,
        "usuario": user["username"],
        "rol": user["rol"],
        "user_id": user["id"],
        "active_page": "tareas",
        "edit_minutes": edit_minutes_conf,
        "vista": vista,
        "fecha_desde": fecha_desde_raw,
        "fecha_hasta": fecha_hasta_raw,
        "es_sistemas_qp": es_sistemas_qp,
        "bandeja_pendientes": bandeja_pendientes,
    }




def svc_build_nueva_tarea_context(user, modo: str):
    if modo not in ("para_mi", "asignar"):
        modo = "asignar"

    return {
        "usuario": user["username"],
        "empresas": repo_obtener_empresas_activas(),
        "user_id": user["id"],
        "rol": user["rol"],
        "responsables": repo_obtener_responsables(user, modo),
        "solicitantes": repo_obtener_solicitantes(),
        "tipos_tarea": repo_obtener_tipos_tarea(),
        "modo": modo,
        "modo_asignar": (modo == "asignar"),
        "active_page": "tareas",
    }


def _parse_task_form_dates(fecha_inicio_raw, fecha_comp_raw, fecha_fin_raw):
    fi = fc = ff = None
    try:
        if fecha_inicio_raw:
            fi = datetime.strptime(fecha_inicio_raw, "%Y-%m-%dT%H:%M")
        if fecha_comp_raw:
            fc = datetime.strptime(fecha_comp_raw, "%Y-%m-%dT%H:%M")
        if fecha_fin_raw:
            ff = datetime.strptime(fecha_fin_raw, "%Y-%m-%dT%H:%M")
    except ValueError:
        raise TaskServiceError("Formato de fecha inválido.", "danger")

    if fi and fc and fc < fi:
        raise TaskServiceError(
            "La fecha compromiso no puede ser anterior a la de inicio.",
            "warning",
        )

    return fi, fc, ff


def svc_crear_tarea(user, form):
    modo = (form.get("modo") or "asignar").lower()
    if modo not in ("para_mi", "asignar"):
        modo = "asignar"

    titulo = (form.get("titulo") or "").strip()
    descripcion = (form.get("descripcion") or "").strip()
    tipo_tarea_id = form.get("tipo_tarea_id")
    empresa_id = form.get("empresa_id")

    fi, fc, ff = _parse_task_form_dates(
        (form.get("fecha_inicio") or "").strip(),
        (form.get("fecha_compromiso") or "").strip(),
        (form.get("fecha_fin") or "").strip(),
    )

    now = datetime.now()
    estado = "Por iniciar" if not fi and not fc else ("Atrasada" if fc and fc < now else "En desarrollo")

    fi_str = fi.strftime("%Y-%m-%d %H:%M:%S") if fi else None
    fc_str = fc.strftime("%Y-%m-%d %H:%M:%S") if fc else None
    ff_str = ff.strftime("%Y-%m-%d %H:%M:%S") if ff else None

    solicitante_raw = form.get("solicitante_id")
    solicitante_id = int(solicitante_raw) if solicitante_raw else int(user["id"])

    responsables_raw = form.getlist("responsable_ids")
    if modo == "asignar":
        responsable_ids = [int(r) for r in responsables_raw if str(r).strip()]
        if not responsable_ids:
            responsable_ids = [int(user["id"])]
    else:
        responsable_ids = [int(user["id"])]

    responsable_ids = list(dict.fromkeys(responsable_ids))

    conn = get_db()

    try:
        responsable_principal = responsable_ids[0]
        tarea_id = repo_insertar_tarea(
            conn,
            {
                "titulo": titulo,
                "descripcion": descripcion,
                "estado": estado,
                "fecha_creacion": now.strftime("%Y-%m-%d %H:%M:%S"),
                "fecha_inicio": fi_str,
                "fecha_compromiso": fc_str,
                "fecha_fin": ff_str,
                "usuario_id": responsable_principal,
                "creador_id": user["id"],
                "solicitante_id": solicitante_id,
                "tipo_tarea_id": tipo_tarea_id,
                "empresa_id": empresa_id,
            },
        )

        for rid in responsable_ids:
            repo_insertar_tarea_responsable_si_no_existe(conn, tarea_id, rid)

        conn.commit()
        return {
            "ok": True,
            "message": f"Tarea {tarea_id:07d} creada correctamente.",
            "category": "success",
            "redirect_endpoint": "listar_tareas",
            "redirect_kwargs": {"vista": "realizar"},
        }
    except Exception as e:
        conn.rollback()
        return {
            "ok": False,
            "message": f"Error al guardar la tarea: {str(e)}",
            "category": "danger",
            "redirect_endpoint": "nueva_tarea",
            "redirect_kwargs": {"modo": modo},
        }


def svc_obtener_tarea_para_ver(user, task_id: int):
    tarea = repo_obtener_tarea_detalle(task_id)
    if not tarea:
        return {
            "ok": False,
            "message": "Tarea no encontrada.",
            "category": "warning",
            "redirect_endpoint": "listar_tareas",
            "redirect_kwargs": {},
        }

    if user["rol"] != "admin":
        if tarea["usuario_id"] != user["id"] and tarea["creador_id"] != user["id"]:
            return {
                "ok": False,
                "message": "No tiene permiso para ver esta tarea.",
                "category": "danger",
                "redirect_endpoint": "listar_tareas",
                "redirect_kwargs": {},
            }

    return {
        "ok": True,
        "tarea": tarea,
        "acciones": repo_obtener_tarea_acciones(task_id),
        "puede_anotar": (
            user["rol"] == "admin"
            or tarea["usuario_id"] == user["id"]
            or tarea["creador_id"] == user["id"]
        ),
        "responsables": repo_obtener_usuarios_activos(),
        "usuario": user["username"],
        "rol": user["rol"],
        "active_page": "tareas",
    }


def _enviar_correo_avance_si_aplica(tarea, observacion, detalles, now_str, actor_username, usuario_accion_id):
    try:
        email_asignado_observacion = None
        if usuario_accion_id:
            u_obs = repo_obtener_email_usuario(int(usuario_accion_id))
            if u_obs:
                email_asignado_observacion = u_obs["email"]

        resp_rows = repo_obtener_emails_responsables_tarea(tarea["id"])

        if not resp_rows:
            row_fallback = repo_obtener_email_usuario(tarea["usuario_id"])
            if row_fallback:
                resp_rows = [row_fallback]

        destinatarios_set = {r["email"] for r in resp_rows if r and r["email"]}

        if tarea["solicitante_id"]:
            sol_row = repo_obtener_email_usuario(tarea["solicitante_id"])
            if sol_row and sol_row["email"]:
                destinatarios_set.add(sol_row["email"])

        if email_asignado_observacion:
            destinatarios_set.add(email_asignado_observacion)

        destinatarios = list(destinatarios_set)
        if not destinatarios:
            return

        msg, smtp_host, smtp_port, smtp_user, smtp_pass, use_tls = _send_tarea_avance_mail(
            tarea,
            observacion,
            detalles,
            now_str,
            actor_username=actor_username,
        )

        if not msg or not smtp_host:
            return

        msg["To"] = ", ".join(destinatarios)

        if use_tls:
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls()
                if smtp_user and smtp_pass:
                    server.login(smtp_user, smtp_pass)
                server.send_message(msg)
        else:
            with smtplib.SMTP_SSL(smtp_host, smtp_port) as server:
                if smtp_user and smtp_pass:
                    server.login(smtp_user, smtp_pass)
                server.send_message(msg)
    except Exception as e:
        print("Error enviando correo de avance:", e)

 
def svc_registrar_accion_tarea(user, task_id: int, form):
    tarea = repo_obtener_tarea_detalle(task_id)
    if not tarea:
        return {
            "ok": False,
            "message": "Tarea no encontrada.",
            "category": "warning",
        }

    observacion = (form.get("observacion") or "").strip()
    detalles = (form.get("detalles") or "").strip()
    estado_accion = (form.get("estado_accion") or "").strip() or None

    fecha_fin_tentativa_raw = (form.get("fecha_fin_tentativa") or "").strip()
    if fecha_fin_tentativa_raw:
        try:
            fecha_fin_tentativa_dt = datetime.strptime(fecha_fin_tentativa_raw, "%Y-%m-%dT%H:%M")
            fecha_fin_tentativa = fecha_fin_tentativa_dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            return {
                "ok": False,
                "message": "La fecha fin tentativa tiene un formato inválido.",
                "category": "warning",
            }
    else:
        fecha_fin_tentativa = None

    usuario_accion_raw = (form.get("usuario_accion_id") or "").strip()
    usuario_accion_id = int(usuario_accion_raw) if usuario_accion_raw else None

    if not observacion and not detalles:
        return {
            "ok": False,
            "message": "Escribe al menos una observación o detalle para registrar la acción.",
            "category": "warning",
        }

    now = datetime.now()
    now_str = now.strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()

    try:
        repo_insertar_tarea_accion(
            conn,
            {
                "tarea_id": task_id,
                "usuario_id": user["id"],
                "fecha_accion": now_str,
                "observacion": observacion,
                "detalles": detalles,
                "estado_accion": estado_accion,
                "usuario_asignado_id": usuario_accion_id,
                "fecha_fin_tentativa": fecha_fin_tentativa,
                "fecha_inicio": now_str,
            },
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        return {
            "ok": False,
            "message": f"Error al registrar la acción: {str(e)}",
            "category": "danger",
        }

    _enviar_correo_avance_si_aplica(
        tarea,
        observacion,
        detalles,
        now_str,
        actor_username=user["username"],
        usuario_accion_id=usuario_accion_id,
    )

    return {
        "ok": True,
        "message": "Acción registrada en la tarea.",
        "category": "success",
    }

def svc_reenviar_observacion(accion_id: int, actor_username: str):
    accion = repo_obtener_accion_con_tarea(accion_id)
    if not accion:
        return {
            "ok": False,
            "message": "Acción no encontrada.",
            "category": "danger",
            "redirect_endpoint": None,
            "redirect_kwargs": {},
        }

    try:
        destinatarios_set = set()
        ids_interesados = [
            accion["resp_tarea_id"],
            accion["solicitante_id"],
            accion["creador_id"],
            accion["usuario_asignado_id"],
        ]

        for uid in ids_interesados:
            if uid:
                res = repo_obtener_email_usuario(uid)
                if res and res["email"]:
                    destinatarios_set.add(res["email"])

        if destinatarios_set:
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            msg, smtp_host, smtp_port, smtp_user, smtp_pass, use_tls = _send_tarea_avance_mail(
                accion,
                f"(RECORDATORIO) {accion['observacion']}",
                accion["detalles"] or "",
                now_str,
                actor_username=actor_username,
            )

            if msg and smtp_host:
                del msg["Subject"]
                msg["Subject"] = f"RECORDATORIO: {accion['titulo']}"
                msg["To"] = ", ".join(destinatarios_set)

                if use_tls:
                    with smtplib.SMTP(smtp_host, smtp_port) as server:
                        server.starttls()
                        if smtp_user and smtp_pass:
                            server.login(smtp_user, smtp_pass)
                        server.send_message(msg)
                else:
                    with smtplib.SMTP_SSL(smtp_host, smtp_port) as server:
                        if smtp_user and smtp_pass:
                            server.login(smtp_user, smtp_pass)
                        server.send_message(msg)

                return {
                    "ok": True,
                    "message": f"Recordatorio enviado a: {len(destinatarios_set)} personas.",
                    "category": "success",
                    "redirect_endpoint": "ver_tarea",
                    "redirect_kwargs": {"task_id": accion["tarea_id"]},
                }

        return {
            "ok": False,
            "message": "No hay correos electrónicos válidos para notificar.",
            "category": "warning",
            "redirect_endpoint": "ver_tarea",
            "redirect_kwargs": {"task_id": accion["tarea_id"]},
        }
    except Exception as e:
        print(f"Error reenviando correo: {e}")
        return {
            "ok": False,
            "message": f"Error al enviar: {str(e)}",
            "category": "danger",
            "redirect_endpoint": "ver_tarea",
            "redirect_kwargs": {"task_id": accion["tarea_id"]},
        }


def svc_finalizar_accion(accion_id: int):
    accion = repo_obtener_accion_con_tarea(accion_id)
    if not accion:
        return {
            "ok": False,
            "message": "No se pudo encontrar la acción.",
            "category": "danger",
            "redirect_endpoint": "listar_tareas",
            "redirect_kwargs": {},
        }

    conn = get_db()
    rowcount = repo_finalizar_accion(conn, accion_id)
    conn.commit()

    if rowcount:
        return {
            "ok": True,
            "message": "Estado de la acción actualizado a Finalizado.",
            "category": "success",
            "redirect_endpoint": "ver_tarea",
            "redirect_kwargs": {"task_id": accion["tarea_id"]},
        }

    return {
        "ok": False,
        "message": "No se pudo encontrar la acción.",
        "category": "danger",
        "redirect_endpoint": "listar_tareas",
        "redirect_kwargs": {},
    }


def svc_obtener_tarea_para_editar(user, task_id: int):
    tarea = repo_obtener_tarea_edicion(task_id)
    if not tarea:
        return {
            "ok": False,
            "message": "Tarea no encontrada.",
            "category": "warning",
            "redirect_endpoint": "listar_tareas",
            "redirect_kwargs": {},
        }

    if user["rol"] != "admin" and tarea["estado"] == "Cerrado por sistema":
        return {
            "ok": False,
            "message": "No se puede editar una tarea cerrada por el sistema.",
            "category": "warning",
            "redirect_endpoint": "listar_tareas",
            "redirect_kwargs": {},
        }

    # Tareas originadas desde Bandeja Soporte: solo admin puede editar
    es_de_soporte = bool(tarea.get("inbox_id"))
    if es_de_soporte and user["rol"] != "admin":
        return {
            "ok": False,
            "message": "Esta tarea fue creada automáticamente desde la Bandeja de Soporte y solo puede ser modificada por un administrador.",
            "category": "warning",
            "redirect_endpoint": "listar_tareas",
            "redirect_kwargs": {},
        }

    es_responsable = repo_es_responsable_tarea(task_id, user["id"])
    editable = (user["rol"] == "admin") or (es_responsable and tarea["estado"] not in ESTADOS_NO_EDITABLES)

    if not editable:
        return {
            "ok": False,
            "message": "No tiene permiso para editar esta tarea.",
            "category": "danger",
            "redirect_endpoint": "listar_tareas",
            "redirect_kwargs": {},
        }

    return {
        "ok": True,
        "tarea": tarea,
        "empresas": repo_obtener_empresas_activas(),
        "estados": ESTADOS,
        "usuario": user["username"],
        "rol": user["rol"],
        "is_admin": (user["rol"] == "admin"),
        "es_de_soporte": es_de_soporte,
        "solicitantes": repo_obtener_solicitantes(),
        "tipos_tarea": repo_obtener_tipos_tarea(),
        "active_page": "tareas",
    }


def svc_guardar_edicion_tarea(user, task_id: int, form):
    tarea = repo_obtener_tarea_edicion(task_id)
    if not tarea:
        return {
            "ok": False,
            "message": "Tarea no encontrada.",
            "category": "warning",
            "redirect_endpoint": "listar_tareas",
            "redirect_kwargs": {},
        }

    # Bloqueo tareas de Bandeja Soporte para no-admin
    if bool(tarea.get("inbox_id")) and user["rol"] != "admin":
        return {
            "ok": False,
            "message": "Esta tarea fue creada automáticamente desde la Bandeja de Soporte y solo puede ser modificada por un administrador.",
            "category": "warning",
            "redirect_endpoint": "listar_tareas",
            "redirect_kwargs": {},
        }

    es_responsable = repo_es_responsable_tarea(task_id, user["id"])
    editable = (user["rol"] == "admin") or (es_responsable and tarea["estado"] not in ESTADOS_NO_EDITABLES)
    if not editable:
        return {
            "ok": False,
            "message": "No tiene permiso para editar esta tarea.",
            "category": "danger",
            "redirect_endpoint": "listar_tareas",
            "redirect_kwargs": {},
        }

    ahora = datetime.now()
    titulo = (form.get("titulo") or "").strip()
    descripcion = (form.get("descripcion") or "").strip()
    estado = (form.get("estado", tarea["estado"]) or tarea["estado"]).strip()
    avance = form.get("porcentaje_avance", 0)
    empresa_id = form.get("empresa_id")

    solicitante_raw = (form.get("solicitante_id") or "").strip()
    try:
        solicitante_id = int(solicitante_raw) if solicitante_raw else tarea.get("solicitante_id")
    except ValueError:
        solicitante_id = tarea.get("solicitante_id")

    tipo_tarea_raw = (form.get("tipo_tarea_id") or "").strip()
    try:
        tipo_tarea_id = int(tipo_tarea_raw) if tipo_tarea_raw else tarea.get("tipo_tarea_id")
    except ValueError:
        tipo_tarea_id = tarea.get("tipo_tarea_id")

    fi_raw = (form.get("fecha_inicio") or "").strip()
    fc_raw = (form.get("fecha_compromiso") or "").strip()
    ff_raw = (form.get("fecha_fin") or "").strip()
    fcr_raw = (form.get("fecha_cierre_real") or "").strip()

    try:
        dt_fi = datetime.strptime(fi_raw, "%Y-%m-%dT%H:%M") if fi_raw else None
        dt_fc = datetime.strptime(fc_raw, "%Y-%m-%dT%H:%M") if fc_raw else None
        dt_ff = datetime.strptime(ff_raw, "%Y-%m-%dT%H:%M") if ff_raw else None
        dcr_ff = datetime.strptime(fcr_raw, "%Y-%m-%dT%H:%M") if fcr_raw else None

        fi_str = dt_fi.strftime("%Y-%m-%d %H:%M:%S") if dt_fi else None
        fc_str = dt_fc.strftime("%Y-%m-%d %H:%M:%S") if dt_fc else None
        ff_str = dt_ff.strftime("%Y-%m-%d %H:%M:%S") if dt_ff else None
        _ = dcr_ff.strftime("%Y-%m-%d %H:%M:%S") if dcr_ff else None

        if dt_ff and dt_ff < ahora:
            estado = "Terminado"
    except ValueError:
        return {
            "ok": False,
            "message": "Formato de fecha inválido.",
            "category": "danger",
            "redirect_endpoint": "editar_tarea",
            "redirect_kwargs": {"task_id": task_id},
        }

    fecha_cierre_real_str = tarea.get("fecha_cierre_real")
    if estado == "Terminado":
        if not fecha_cierre_real_str:
            fecha_cierre_real_str = ahora.strftime("%Y-%m-%d %H:%M:%S")
    else:
        fecha_cierre_real_str = None

    conn = get_db()
    estado_anterior = tarea["estado"]
    repo_actualizar_tarea(
        conn,
        task_id,
        {
            "titulo": titulo,
            "descripcion": descripcion,
            "empresa_id": empresa_id,
            "estado": estado,
            "fecha_inicio": fi_str,
            "fecha_compromiso": fc_str,
            "fecha_fin": ff_str,
            "fecha_cierre_real": fecha_cierre_real_str,
            "solicitante_id": solicitante_id,
            "porcentaje_avance": avance,
            "tipo_tarea_id": tipo_tarea_id,
        },
    )
    conn.commit()
    if estado_anterior != "Terminado" and estado == "Terminado":
        try:
            conn.execute(
                "UPDATE email_tickets_inbox SET estado = 'TERMINADA' WHERE tarea_id = ? AND estado = 'ASIGNADA'",
                (task_id,)
            )
            conn.commit()
        except Exception:
            pass
        svc_crear_y_enviar_encuesta(task_id)

    return {
        "ok": True,
        "message": "Tarea actualizada correctamente.",
        "category": "success",
        "redirect_endpoint": "listar_tareas",
        "redirect_kwargs": {},
    }

def svc_eliminar_tarea(user, task_id: int):
    conn = get_db()
    changes = repo_eliminar_tarea(conn, task_id, user)
    conn.commit()

    if changes:
        return {"ok": True, "message": "Tarea eliminada.", "category": "success"}
    return {
        "ok": False,
        "message": "No se pudo eliminar (permiso o inexistente).",
        "category": "warning",
    }

def svc_require_api_key(req):
    expected = (get_config_value("api_inbound_key", "") or "").strip()
    got = (req.headers.get("X-API-Key") or "").strip()
    if not expected:
        return False, "Falta configurar api_inbound_key en tabla configuracion."
    if got != expected:
        return False, "API key inválida."
    return True, ""

def svc_exportar_tareas_excel(request_args):
    texto_q = (request_args.get("q") or "").lower().strip()
    estado_f = (request_args.get("estado") or "").strip()
    prop_f = (
        request_args.get("prop")
        or request_args.get("usuario")
        or request_args.get("creador")
        or ""
    ).lower().strip()
    depto_f = (request_args.get("depto") or "").lower().strip()

    fecha_desde_raw = (request_args.get("fecha_desde") or "").strip()
    fecha_hasta_raw = (request_args.get("fecha_hasta") or "").strip()

    fecha_desde = parse_dt(fecha_desde_raw) if fecha_desde_raw else None
    fecha_hasta = parse_dt(fecha_hasta_raw) if fecha_hasta_raw else None

    if fecha_hasta:
        fecha_hasta = fecha_hasta.replace(hour=23, minute=59, second=59)

    tareas_raw = repo_listar_tareas_raw()
    resp_rows = repo_listar_tarea_responsables_map()

    responsables_map = {}
    for r in resp_rows:
        tid = r["tarea_id"]
        responsables_map.setdefault(tid, []).append(dict(r))

    filas = []

    for t in tareas_raw:
        resp_list = responsables_map.get(t["id"], [])

        responsables = []
        departamentos = set()

        for r in resp_list:
            nombre = (r.get("nombre_completo") or "").strip()
            username = (r.get("username") or "").strip()
            depto = (r.get("depto_nombre") or "").strip()

            if nombre and username:
                responsables.append(f"{nombre} ({username})")
            elif nombre:
                responsables.append(nombre)
            elif username:
                responsables.append(username)

            if depto:
                departamentos.add(depto)

        if not responsables:
            responsables.append(t.get("creador_username") or "")

        responsable_txt = ", ".join(responsables)
        depto_txt = ", ".join(sorted(departamentos)) if departamentos else (t.get("departamento_nombre") or "N/A")

        solicitante_txt = (
            f"{t.get('solicitante_nombre')} ({t.get('solicitante_username')})"
            if t.get("solicitante_nombre")
            else (t.get("solicitante_username") or "")
        )

        creado_por_txt = t.get("creador_nombre") or t.get("creador_username") or ""

        fecha_base = parse_dt(
            str(
                t.get("fecha_cierre_real")
                or t.get("fecha_fin")
                or t.get("fecha_inicio")
                or t.get("fecha_creacion")
                or ""
            ).strip()
        )

        if fecha_desde and (not fecha_base or fecha_base < fecha_desde):
            continue

        if fecha_hasta and (not fecha_base or fecha_base > fecha_hasta):
            continue

        # Horas invertidas = fecha_inicio → fecha_fin (tiempo planificado de la tarea)
        inicio_dt = parse_dt(t.get("fecha_inicio") or t.get("fecha_creacion"))
        fin_dt    = parse_dt(t.get("fecha_fin"))

        horas_soporte = ""
        if inicio_dt and fin_dt:
            horas = round((fin_dt - inicio_dt).total_seconds() / 3600, 2)
            horas_soporte = horas if horas >= 0 else ""

        fila = {
            "Código Tarea": f"{int(t['id']):08d}",
            "Título": t.get("titulo") or "",
            "Descripción Tarea": t.get("descripcion") or "",
            "Estado Global": t.get("estado") or "",
            "Tipo Tarea": t.get("tipo_tarea_nombre") or "",
            "Creado por": creado_por_txt,
            "Solicitante": solicitante_txt,
            "F. Creación": t.get("fecha_creacion") or "",
            "Inicio": t.get("fecha_inicio") or "",
            "Fecha Compromiso": t.get("fecha_compromiso") or "",
            "Fin": t.get("fecha_fin") or "",
            "Fecha Fin Real": t.get("fecha_cierre_real") or "",
            "Horas de Soporte": horas_soporte,
            "Responsable Tarea": responsable_txt,
            "Departamento": depto_txt,
            "Empresa": t.get("empresa_nombre") or "Interno",
            "% Avance": t.get("porcentaje_avance") or 0,
        }

        usuario_texto = " ".join([
            responsable_txt,
            creado_por_txt,
            solicitante_txt,
            t.get("creador_username") or "",
            t.get("creador_nombre") or "",
            t.get("solicitante_username") or "",
            t.get("solicitante_nombre") or "",
        ]).lower()

        row_text = " ".join(str(v).lower() for v in fila.values())

        if texto_q and texto_q not in row_text:
            continue

        if estado_f and fila["Estado Global"] != estado_f:
            continue

        if prop_f and prop_f not in usuario_texto:
            continue

        if depto_f and depto_f not in fila["Departamento"].lower():
            continue

        filas.append(fila)

    df = pd.DataFrame(filas)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Reporte Tareas", startrow=6)

        ws = writer.sheets["Reporte Tareas"]

        label_font = Font(bold=True, size=11)
        value_font = Font(bold=False, size=11)

        ws["A1"] = "Nombre de reporte:"
        ws["B1"] = "Reporte de Tareas Detallado"
        ws["A2"] = "Fecha de generación:"
        ws["B2"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        ws["A3"] = "Usuario de generación:"
        ws["B3"] = session.get("nombre_completo", session.get("username", "Sistema"))

        ws["A1"].font = label_font
        ws["A2"].font = label_font
        ws["A3"].font = label_font
        ws["B1"].font = value_font
        ws["B2"].font = value_font
        ws["B3"].font = value_font

        for column_cells in ws.columns:
            values = [str(cell.value or "") for cell in column_cells]
            length = max(len(v) for v in values)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)

    output.seek(0)
    return output

def svc_api_inbound_email_create_task(payload):
    subject = (payload.get("subject") or "").strip()
    body = (payload.get("body") or payload.get("text") or payload.get("body_text") or "").strip()
    from_email = _extract_email(payload.get("from") or payload.get("from_email"))

    if not subject:
        return {"ok": False, "error": "Falta subject", "status": 400}
    if not from_email:
        return {"ok": False, "error": "Falta from (remitente)", "status": 400}
    if not body:
        body = "(sin cuerpo)"

    default_email = "jchavez@quimpac.com.ec"

    conn = get_db()
    cur = conn.cursor()

    uid_from = repo_find_user_id_by_email(cur, from_email)

    if uid_from:
        responsable_principal = uid_from
        solicitante_id = uid_from
    else:
        uid_def = repo_find_user_id_by_email(cur, _norm_email(default_email))
        if not uid_def:
            return {
                "ok": False,
                "error": f"Remitente {from_email} no existe y el default {default_email} no está en usuarios.",
                "status": 400,
            }

        responsable_principal = uid_def
        solicitante_id = None

    creador_id = uid_from or responsable_principal
    now = datetime.now()

    try:
        tarea_id = repo_insertar_tarea(
            conn,
            {
                "titulo": subject[:250],
                "descripcion": body,
                "estado": "Por iniciar",
                "fecha_creacion": now.strftime("%Y-%m-%d %H:%M:%S"),
                "fecha_inicio": None,
                "fecha_compromiso": None,
                "fecha_fin": None,
                "usuario_id": responsable_principal,
                "creador_id": creador_id,
                "solicitante_id": solicitante_id,
                "tipo_tarea_id": None,
                "empresa_id": None,
            },
        )
        repo_insertar_tarea_responsable_si_no_existe(conn, tarea_id, responsable_principal)
        conn.commit()
    except Exception as e:
        conn.rollback()
        return {"ok": False, "error": str(e), "status": 500}

    return {
        "ok": True,
        "tarea_id": tarea_id,
        "codigo_tarea": f"{tarea_id:07d}",
        "asignado_a_user_id": responsable_principal,
        "from_email": from_email,
        "from_encontrado": bool(uid_from),
        "default_email": default_email,
        "status": 201,
    }



def svc_abrir_o_crear_encuesta_desde_bandeja(user, task_id: int):
    tarea = repo_obtener_tarea_para_encuesta(task_id)

    if not tarea:
        return {
            "ok": False,
            "message": "Tarea no encontrada.",
            "category": "warning",
            "redirect_endpoint": "listar_encuestas",
            "redirect_kwargs": {},
        }

    user_id = _safe_int(user.get("id"))
    solicitante_id = _safe_int(tarea.get("solicitante_id"))

    if user_id != solicitante_id:
        return {
            "ok": False,
            "message": "Solo el solicitante puede responder esta encuesta.",
            "category": "danger",
            "redirect_endpoint": "listar_encuestas",
            "redirect_kwargs": {},
        }

    existente = repo_obtener_encuesta_por_tarea(task_id)

    if existente:
        if existente.get("estado") == "Realizada":
            return {
                "ok": False,
                "message": "Esta encuesta ya fue respondida.",
                "category": "warning",
                "redirect_endpoint": "listar_encuestas",
                "redirect_kwargs": {},
            }

        return {
            "ok": True,
            "message": None,
            "category": None,
            "redirect_endpoint": "responder_encuesta",
            "redirect_kwargs": {"token": existente["token"]},
        }

    if not tarea.get("solicitante_id"):
        return {
            "ok": False,
            "message": "La tarea no tiene solicitante asignado.",
            "category": "warning",
            "redirect_endpoint": "listar_encuestas",
            "redirect_kwargs": {},
        }

    conn = get_db()
    token = secrets.token_urlsafe(32)

    try:
        repo_insertar_encuesta(
            conn,
            task_id,
            tarea.get("solicitante_id"),
            token,
        )
        conn.commit()

        return {
            "ok": True,
            "message": None,
            "category": None,
            "redirect_endpoint": "responder_encuesta",
            "redirect_kwargs": {"token": token},
        }

    except Exception as e:
        conn.rollback()
        return {
            "ok": False,
            "message": f"Error creando encuesta: {str(e)}",
            "category": "danger",
            "redirect_endpoint": "listar_encuestas",
            "redirect_kwargs": {},
        }