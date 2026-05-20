from flask import render_template, request, redirect, url_for, flash, session, send_file
from datetime import datetime, timedelta, date
from io import BytesIO
from collections import defaultdict
import pandas as pd
import io
from datetime import datetime
from flask import send_file, request, session, current_app

from .db import get_db, get_config_value
from .config import ESTADOS
from .security import require_login, require_permission, get_user
from openpyxl import Workbook
from email.message import EmailMessage
import smtplib
from flask import request, jsonify
from datetime import datetime
import re
import pandas as pd
import io
from flask import send_file, request

from .tasks.task_utils import build_user_display_name, parse_dt, _norm_email, _extract_email
from .tasks.task_notifications import _send_tarea_avance_mail, _send_tarea_creada_mail

def register_task_routes(app):

    # --------- Helper: nombre completo de usuario ---------
    def build_user_display_name(user_row) -> str:
        """
        Intenta armar 'Nombre Apellido' usando distintos posibles campos.
        Si no encuentra, usa username.
        """
        d = dict(user_row)
        base = ""

        # 1) campos de nombre completo directos
        for k in ("nombre_completo", "nombrecompleto", "full_name"):
            v = d.get(k)
            if v:
                base = str(v).strip()
                break

        # 2) combinar nombres + apellidos o variantes
        if not base:
            first = None
            last = None
            for k in ("nombres", "nombre", "first_name"):
                if k in d and d[k]:
                    first = str(d[k]).strip()
                    break
            for k in ("apellidos", "apellido", "last_name"):
                if k in d and d[k]:
                    last = str(d[k]).strip()
                    break
            parts = [p for p in (first, last) if p]
            if parts:
                base = " ".join(parts)

        # 3) fallback username
        if not base:
            u = d.get("username")
            if u:
                base = str(u).strip()
            else:
                base = f"ID {d.get('id', '?')}"

        return base




    def _send_tarea_avance_mail(
        tarea_row,
        observacion: str,
        detalles: str,
        fecha_accion_str: str,
        actor_username: str
    ):
        """
        Arma el EmailMessage para notificar un nuevo avance en la tarea.
        Devuelve: (msg, smtp_host, smtp_port, smtp_user, smtp_pass, use_tls)
        Si falta configuración SMTP, msg será None.
        """
        # --- Config SMTP ---
        smtp_host = get_config_value('smtp_host', '')
        smtp_port_raw = get_config_value('smtp_port', '587')
        smtp_user = get_config_value('smtp_user', '')
        smtp_pass = get_config_value('smtp_pass', '')
        smtp_from = get_config_value('smtp_from', smtp_user or '')
        use_tls   = (get_config_value('smtp_tls', '1') == '1')

        try:
            smtp_port = int(smtp_port_raw or 587)
        except Exception:
            smtp_port = 587

        if not (smtp_host and smtp_from):
            # Sin configuración mínima, no armamos nada
            return None, None, None, None, None, None

        # --- Datos de la tarea ---
        tarea = dict(tarea_row)
        codigo_tarea = f"{tarea['id']:07d}"
        titulo = tarea.get('titulo') or 'Sin título'
        estado = tarea.get('estado') or '—'
        fi_str = tarea.get('fecha_inicio') or 'No definida'
        fc_str = tarea.get('fecha_compromiso') or 'No definida'
        ff_str = tarea.get('fecha_fin') or 'No definida'

        obs_txt = observacion or '(sin observación)'
        det_txt = detalles or '(sin detalles)'
        det_html = det_txt.replace("\n", "<br>")

        # --- Asunto ---
        subject = f"[Tareas {codigo_tarea}] Nuevo avance en la tarea: {titulo}"

        # --- Cuerpo texto plano (fallback) ---
        body_text = f"""Se ha registrado un nuevo avance en la tarea {codigo_tarea}.

    Tarea: {titulo}
    Estado actual: {estado}
    Fecha de la acción: {fecha_accion_str}
    Registrado por: {actor_username}

    Observación:
    {obs_txt}

    Detalles:
    {det_txt}

    Este mensaje se generó automáticamente desde el sistema de tareas.
    """

        # --- Cuerpo HTML con iconos ---
        body_html = f"""
    <html>
    <body style="font-family: Arial, sans-serif; font-size:14px; color:#333;">
        <h2 style="margin:0 0 10px; color:#0078d4;">
        🔄 Nuevo avance registrado en la tarea
        </h2>

        <p>Se ha registrado un nuevo avance en la siguiente tarea:</p>

        <div style="border-left:4px solid #0078d4;
                    padding:12px 16px;
                    background:#f5f5f5;
                    margin:16px 0;">
        <p style="margin:0 0 4px;">
            🔢 <strong>Código:</strong> {codigo_tarea}
        </p>
        <p style="margin:0 0 4px;">
            📌 <strong>Título:</strong> {titulo}
        </p>
        <p style="margin:0 0 4px;">
            🎯 <strong>Estado actual:</strong> {estado}
        </p>
        <p style="margin:0 0 4px;">
            👤 <strong>Acción registrada por:</strong> {actor_username}
        </p>
        <p style="margin:0 0 4px;">
            🕒 <strong>Fecha de la acción:</strong> {fecha_accion_str}
        </p>
        <p style="margin:0 0 4px;">
            📅 <strong>Fecha inicio:</strong> {fi_str}
        </p>
        <p style="margin:0 0 4px;">
            ⏱️ <strong>Fecha compromiso:</strong> {fc_str}
        </p>
        <p style="margin:0;">
            ✅ <strong>Fecha fin:</strong> {ff_str}
        </p>
        </div>

        <p style="margin:0 0 4px;">
        📝 <strong>Observación:</strong>
        </p>
        <p style="margin:0 0 12px;">{obs_txt}</p>

        <p style="margin:0 0 4px;">
        📄 <strong>Detalles:</strong>
        </p>
        <p style="margin:0 0 12px;">{det_html}</p>

        <p style="margin-top:16px;">
        👉 Por favor ingresa al sistema de tareas para revisar el detalle completo.
        </p>

        <p style="font-size:12px; color:#777; margin-top:24px;">
        ⚠️ Este mensaje se generó automáticamente, por favor no responder a este correo.
        </p>
    </body>
    </html>
    """

        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = smtp_from
        # el "To" se asigna en ver_tarea con la lista de responsables
        msg.set_content(body_text)
        msg.add_alternative(body_html, subtype="html")

        return msg, smtp_host, smtp_port, smtp_user, smtp_pass, use_tls





    @app.route('/dashboard')
    @require_login
    def dashboard():
        """
        Panel de tareas con foco en vencidas / para hoy / semana, + gráficos.

        - Admin: ve todo.
        - Jefe: ve tareas propias + de su departamento.
        - Usuario: solo sus tareas.
        """
        user = get_user()
        conn = get_db()
        cur = conn.cursor()

        # --- 1. Traer tareas según rol ---
        base_sql = """
            SELECT t.id,
                   t.titulo,
                   t.descripcion,
                   t.estado,
                   t.fecha_creacion,
                   t.fecha_inicio,
                   t.fecha_compromiso,
                   t.fecha_fin,
                   t.usuario_id,
                   u.username AS propietario,
                   d.nombre  AS departamento
            FROM tareas t
            JOIN usuarios u ON t.usuario_id = u.id
            LEFT JOIN departamentos d ON u.departamento_id = d.id
        """
        where = []
        params = []

        if user["rol"] == "admin":
            # sin filtro extra
            pass
        elif user["rol"] == "jefe":
            # propias + de su departamento
            where.append("(u.id = ? OR u.departamento_id = ?)")
            params.extend([user["id"], user.get("departamento_id")])
        else:
            # usuario normal: solo sus tareas
            where.append("u.id = ?")
            params.append(user["id"])

        if where:
            base_sql += " WHERE " + " AND ".join(where)

        base_sql += """
            ORDER BY
                CASE WHEN t.fecha_compromiso IS NULL THEN 1 ELSE 0 END,
                t.fecha_compromiso,
                t.fecha_creacion DESC
        """

        cur.execute(base_sql, params)
        tareas_raw = cur.fetchall()
        conn.close()

        # --- 2. Helpers de fecha ---
        def parse_dt(val):
            if not val:
                return None
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M"):
                try:
                    return datetime.strptime(val, fmt)
                except ValueError:
                    continue
            try:
                return datetime.fromisoformat(val)
            except Exception:
                return None

        hoy = date.today()
        # fin de semana actual (domingo)
        week_end = hoy + timedelta(days=(6 - hoy.weekday()))  # weekday(): 0=lun..6=dom

        # --- 3. Contadores por estado ---
        conteos = {estado: 0 for estado in ESTADOS}

        # Listas detalladas
        overdue_tasks = []      # vencidas
        today_tasks = []        # vencen hoy
        week_tasks = []         # vencen resto de semana
        no_due_tasks = []       # sin fecha compromiso

        # Agregados
        overdue_by_user_count = defaultdict(int)
        overdue_by_depto_count = defaultdict(int)

        # Para gráficos adicionales
        compromiso_counts = defaultdict(int)   # tareas por fecha_compromiso (todas)

        for r in tareas_raw:
            t = dict(r)
            estado = t["estado"]

            # Conteo clásico por estado
            if estado in conteos:
                conteos[estado] += 1
            else:
                conteos[estado] = conteos.get(estado, 0) + 1

            comp_dt = parse_dt(t.get("fecha_compromiso"))
            if comp_dt:
                compromiso_counts[comp_dt.date()] += 1

            t["fecha_compromiso_fmt"] = comp_dt.strftime("%Y-%m-%d %H:%M") if comp_dt else ""
            no_terminada = (estado != "Terminado")

            if comp_dt and no_terminada:
                comp_date = comp_dt.date()
                if comp_date < hoy:
                    # Atrasada
                    t["dias_atraso"] = (hoy - comp_date).days
                    overdue_tasks.append(t)

                    owner = t.get("propietario") or "—"
                    depto = t.get("departamento") or "Sin departamento"
                    overdue_by_user_count[owner] += 1
                    overdue_by_depto_count[depto] += 1

                elif comp_date == hoy:
                    t["dias_atraso"] = 0
                    today_tasks.append(t)

                elif hoy < comp_date <= week_end:
                    t["dias_atraso"] = 0
                    week_tasks.append(t)

            elif not comp_dt and no_terminada:
                # sin fecha compromiso
                t["dias_atraso"] = None
                no_due_tasks.append(t)

        total = len(tareas_raw)

        # Agregados por usuario/departamento
        overdue_by_user = sorted(
            [{"usuario": u, "total": c} for u, c in overdue_by_user_count.items()],
            key=lambda x: -x["total"]
        )
        overdue_by_depto = sorted(
            [{"departamento": d, "total": c} for d, c in overdue_by_depto_count.items()],
            key=lambda x: -x["total"]
        )

        # --- 4. Estructura para gráficos (Chart.js) ---
        # Tareas por estado
        status_labels = list(ESTADOS)
        status_data = [conteos.get(e, 0) for e in ESTADOS]

        # Atrasadas por usuario / depto
        overdue_user_labels = [row["usuario"] for row in overdue_by_user]
        overdue_user_data = [row["total"] for row in overdue_by_user]

        overdue_depto_labels = [row["departamento"] for row in overdue_by_depto]
        overdue_depto_data = [row["total"] for row in overdue_by_depto]

        # Tareas por fecha compromiso (todas)
        dates_sorted = sorted(compromiso_counts.keys())
        timeline_labels = [d.isoformat() for d in dates_sorted]
        timeline_data = [compromiso_counts[d] for d in dates_sorted]

        chart = {
            "status": {
                "labels": status_labels,
                "data": status_data,
            },
            "overdue_user": {
                "labels": overdue_user_labels,
                "data": overdue_user_data,
            },
            "overdue_depto": {
                "labels": overdue_depto_labels,
                "data": overdue_depto_data,
            },
            "timeline": {
                "labels": timeline_labels,
                "data": timeline_data,
            },
        }

        return render_template(
            "dashboard.html",
            usuario=user["username"],
            rol=user["rol"],
            estados=ESTADOS,
            conteos=conteos,
            total=total,
            overdue_tasks=overdue_tasks,
            today_tasks=today_tasks,
            week_tasks=week_tasks,
            no_due_tasks=no_due_tasks,
            overdue_by_user=overdue_by_user,
            overdue_by_depto=overdue_by_depto,
            chart=chart,
            active_page="dashboard",
        )

    @app.route('/tareas')
    @require_login
    @require_permission('tareas', 'ver')
    def listar_tareas():
        user = get_user()
        conn = get_db()
        cur = conn.cursor()

        # 1. Traer tareas base con JOIN para el tipo de tarea y el creador
        cur.execute('''
            SELECT t.id,
                t.titulo,
                t.descripcion,
                t.estado,
                t.fecha_creacion,
                t.fecha_inicio,
                t.fecha_compromiso,
                t.fecha_fin,
                t.fecha_cierre_real,
                t.solicitante_id,                 -- 👈 NUEVO
               COALESCE(dsol.nombre,'') AS departamento_nombre,  

                t.usuario_id,
                t.creador_id,
                t.tipo_tarea_id,
                p.nombre AS tipo_tarea_nombre,
                cu.username        AS creador_username,
                cu.nombre_completo AS creador_nombre,
                t.notificado, porcentaje_avance, 
                e.razon_social AS empresa_nombre  -- Traemos el nombre de la empresa
            FROM tareas t
            LEFT JOIN empresas e ON t.empresa_id = e.id  -- Unión con la tabla empresas
            LEFT JOIN usuarios cu ON t.creador_id = cu.id
            LEFT JOIN param_values p ON t.tipo_tarea_id = p.id
                    
    LEFT JOIN usuarios usol ON usol.id = t.solicitante_id        -- 👈 NUEVO
    LEFT JOIN departamentos dsol ON dsol.id = usol.departamento_id -- 👈 NUEVO

            ORDER BY t.id DESC
        ''')
        tareas_raw = [dict(r) for r in cur.fetchall()]

        # 2. Mapa de responsables por tarea (tabla puente + departamentos)
        cur.execute('''
            SELECT tr.tarea_id,
                ur.id                           AS usuario_id,
                ur.username                     AS username,
                COALESCE(ur.nombre_completo,'') AS nombre_completo,
                COALESCE(d.nombre,'')           AS depto_nombre
            FROM tarea_responsables tr
            JOIN usuarios ur  ON ur.id = tr.usuario_id
            LEFT JOIN departamentos d ON d.id = ur.departamento_id
            ORDER BY tr.tarea_id, ur.username
        ''')
        resp_rows = cur.fetchall()
        
        # Organizar responsables en un diccionario por tarea_id
        responsables_map = {}
        for r in resp_rows:
            tid = r['tarea_id']
            responsables_map.setdefault(tid, []).append(dict(r))

        # 3. Enriquecer cada tarea con labels de responsables y departamentos
        for t in tareas_raw:
            resp_list = responsables_map.get(t['id'], [])
            ids = set()
            labels = []
            deptos = set()

            for r in resp_list:
                ids.add(r['usuario_id'])
                dept_name = (r['depto_nombre'] or '').strip()
                if dept_name:
                    deptos.add(dept_name)
                
                nombre = (r['nombre_completo'] or '').strip()
                if nombre:
                    labels.append(f"{nombre} ({r['username']})")
                else:
                    labels.append(r['username'])

            # Fallback para tareas antiguas sin registros en la tabla puente
            if not ids:
                if t.get('usuario_id'):
                    ids.add(t['usuario_id'])
                # Usamos la info del creador o usuario asignado original
                label_base = (t.get('creador_username') or '').strip()
                nombre_creador = (t.get('creador_nombre') or '').strip()
                if nombre_creador:
                    labels.append(f"{nombre_creador} ({label_base})" if label_base else nombre_creador)
                elif label_base:
                    labels.append(label_base)

            t['responsable_ids'] = ids
            t['propietario'] = ", ".join(labels) if labels else 'Sin asignar'
            t['departamentos_responsables'] = ", ".join(sorted(deptos)) if deptos else 'N/A'


        # 4. Definir la vista y filtrar según el rol y permisos
        vista = (request.args.get('vista') or 'realizar').lower()
        if vista not in ('realizar', 'mis'):
            vista = 'realizar'

        tareas_filtradas = []
        for t in tareas_raw:
            # Lógica de visibilidad
            if user['rol'] in ('admin', 'jefe'):
                pasa_rol = True
            else:
                # Usuario normal: solo lo que creó o donde es responsable
                pasa_rol = (user['id'] in t['responsable_ids']) or (t['creador_id'] == user['id'])

            if not pasa_rol:
                continue

            # Lógica de la pestaña activa (Realizar vs Mis Creadas)
            if vista == 'realizar':
                # Si no es admin/jefe, debe estar en la lista de responsables
                if user['rol'] not in ('admin', 'jefe') and user['id'] not in t['responsable_ids']:
                    continue
            else:  # vista 'mis'
                # Si no es admin/jefe, debe ser el creador
                if user['rol'] not in ('admin', 'jefe') and t['creador_id'] != user['id']:
                    continue

            tareas_filtradas.append(t)

        # 5. Ventana de edición y lista final
        try:
            edit_minutes_conf = float(get_config_value('edit_minutes', '5'))
        except:
            edit_minutes_conf = 5.0

        tareas_list = []
        for t in tareas_filtradas:
            editable = False
            if user['rol'] == 'admin':
                editable = True
            else:
                # Es responsable y la tarea no está finalizada
                if (user['id'] in t['responsable_ids'] and 
                    t['estado'] not in ('Terminado', 'Cerrado por sistema')):
                    editable = True
            
            t['editable'] = editable
            tareas_list.append(t)

        conn.close()

        return render_template(
            'tareas.html',
            tareas=tareas_list,
            usuario=user['username'],
            rol=user['rol'],
            user_id=user['id'],
            active_page='tareas',
            edit_minutes=edit_minutes_conf,
            vista=vista
        )
    
    
    
    
    @app.route('/tareas/nueva', methods=['GET', 'POST'])
    @require_login
    @require_permission('tareas', 'crear')
    def nueva_tarea():
        user = get_user()

        # --- MODO ---
        modo = (request.args.get('modo') or request.form.get('modo') or 'asignar').lower()
        if modo not in ('para_mi', 'asignar'):
            modo = 'asignar'

        # Helper: TODOS los usuarios activos para el campo "Solicitante"
        def obtener_solicitantes():
            conn_i = get_db()
            cur_i = conn_i.cursor()

            cur_i.execute("""
                SELECT id, username, COALESCE(nombre_completo,'') AS nombre_completo
                FROM usuarios WHERE disabled = 0
                ORDER BY nombre_completo, username
            """)
            rows = cur_i.fetchall()
            conn_i.close()
            return rows

        # Helper para obtener posibles responsables
        def obtener_responsables():
            if modo != 'asignar': return []
            conn_i = get_db()
            cur_i = conn_i.cursor()
            
            if user['rol'] == 'admin':
                cur_i.execute("SELECT id, username, COALESCE(nombre_completo,'') AS nombre_completo FROM usuarios WHERE disabled = 0 ORDER BY nombre_completo, username")
            elif user['rol'] == 'jefe':
                cur_i.execute("SELECT id, username, COALESCE(nombre_completo,'') AS nombre_completo FROM usuarios WHERE disabled = 0 AND (id = ? OR jefe_id = ?) ORDER BY nombre_completo, username", (user['id'], user['id']))
            else:
                cur_i.execute("SELECT id, username, COALESCE(nombre_completo,'') AS nombre_completo FROM usuarios WHERE disabled = 0 AND id = ?", (user['id'],))
                
            rows = cur_i.fetchall()
            conn_i.close()
            return rows

        # NUEVO: Helper para obtener Tipos de Tarea (Grupo 11453)
        def obtener_tipos_tarea():
            conn_i = get_db()
            cur_i = conn_i.cursor()
            cur_i.execute("""
                SELECT id, nombre 
                FROM param_values 
                WHERE group_id = 4945 AND activo = 1 
                ORDER BY orden ASC
            """)
            rows = cur_i.fetchall()
            conn_i.close()
            return rows


        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id, razon_social FROM empresas WHERE activo = 1 ORDER BY razon_social ASC")
        empresas = cur.fetchall()
        # ---------------- POST: crear tarea ----------------
        # ---------------- POST: crear tarea ----------------
        if request.method == 'POST':
            # 1. Capturar datos básicos y el nuevo campo empresa_id
            modo_post = (request.form.get('modo') or modo).lower()
            if modo_post in ('para_mi', 'asignar'): 
                modo = modo_post

            titulo = request.form['titulo'].strip()
            descripcion = request.form.get('descripcion', '').strip()
            tipo_tarea_id = request.form.get('tipo_tarea_id')
            empresa_id = request.form.get('empresa_id') # <--- NUEVO CAMPO CAPTURADO

            fecha_inicio_raw = request.form.get('fecha_inicio', '').strip()
            fecha_comp_raw   = request.form.get('fecha_compromiso', '').strip()
            fecha_fin_raw    = request.form.get('fecha_fin', '').strip()
            responsables_raw = request.form.getlist('responsable_ids')

            # 2. Procesamiento de fechas
            fi = fc = ff = None
            try:
                if fecha_inicio_raw: fi = datetime.strptime(fecha_inicio_raw, '%Y-%m-%dT%H:%M')
                if fecha_comp_raw:   fc = datetime.strptime(fecha_comp_raw, '%Y-%m-%dT%H:%M')
                if fecha_fin_raw:    ff = datetime.strptime(fecha_fin_raw, '%Y-%m-%dT%H:%M')
            except ValueError:
                flash('Formato de fecha inválido.', 'danger')
                return redirect(url_for('nueva_tarea', modo=modo))

            if fi and fc and fc < fi:
                flash('La fecha compromiso no puede ser anterior a la de inicio.', 'warning')
                return redirect(url_for('nueva_tarea', modo=modo))

            now = datetime.now()
            # Lógica de estado automático
            estado = 'Por iniciar' if not fi and not fc else ('Atrasada' if fc and fc < now else 'En desarrollo')

            fi_str = fi.strftime('%Y-%m-%d %H:%M:%S') if fi else None
            fc_str = fc.strftime('%Y-%m-%d %H:%M:%S') if fc else None
            ff_str = ff.strftime('%Y-%m-%d %H:%M:%S') if ff else None

            # 3. Identificar Solicitante y Responsables
            solicitante_raw = request.form.get('solicitante_id')
            solicitante_id = int(solicitante_raw) if solicitante_raw else int(user['id'])

            responsable_ids = []
            if modo == 'asignar':
                responsable_ids = [int(r) for r in responsables_raw if r.strip()]
                if not responsable_ids: 
                    responsable_ids = [int(user['id'])]
            else:
                responsable_ids = [int(user['id'])]
            
            responsable_ids = list(dict.fromkeys(responsable_ids)) # Quitar duplicados


            # Consulta para obtener las empresas
   
            try:
                # --- 4) Insertar Tarea con empresa_id ---
                responsable_principal = responsable_ids[0]
                cur.execute('''
                    INSERT INTO tareas (
                        titulo, descripcion, estado, fecha_creacion, fecha_inicio, 
                        fecha_compromiso, fecha_fin, usuario_id, creador_id, 
                        solicitante_id, notificado, tipo_tarea_id, empresa_id
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
                ''', (
                    titulo, descripcion, estado, now.strftime('%Y-%m-%d %H:%M:%S'),
                    fi_str, fc_str, ff_str, responsable_principal, user['id'], 
                    solicitante_id, tipo_tarea_id, empresa_id
                ))

                tarea_id = cur.lastrowid
                codigo_tarea = f"{tarea_id:07d}"

                # --- 5) Tabla Puente Responsables (Multiasignación) ---
                for rid in responsable_ids:
                    cur.execute("""
                        INSERT OR IGNORE INTO tarea_responsables (tarea_id, usuario_id) 
                        VALUES (?, ?)
                    """, (tarea_id, rid))

                # --- 6) Envío de Correo ---
                # Aquí puedes llamar a tu función de envío de correos 
                # similar a la que usamos para las observaciones.
                
                conn.commit()
                flash(f'Tarea {codigo_tarea} creada correctamente.', 'success')
                return redirect(url_for('listar_tareas', vista='realizar'))

            except Exception as e:
                conn.rollback()
                print("Error al crear tarea:", e)
                flash(f"Error al guardar la tarea: {str(e)}", 'danger')
            finally:
                conn.close()
        # ---------------- GET ----------------
        return render_template(
            'nueva_tarea.html',
            usuario=user['username'],
            empresas=empresas,
            user_id=user['id'],
            rol=user['rol'],
            responsables=obtener_responsables(),
            solicitantes=obtener_solicitantes(),
            tipos_tarea=obtener_tipos_tarea(), # 👈 Enviamos los tipos al front
            modo=modo,
            modo_asignar=(modo == 'asignar'),
            active_page='tareas'
        )
    
    @app.route('/tareas/<int:task_id>/ver', methods=['GET', 'POST'])
    @require_login
    @require_permission('tareas', 'ver')
    def ver_tarea(task_id):
        """
        Vista de sólo lectura de la tarea + bitácora de acciones.
        Desde aquí se pueden ir registrando los pasos/avances.
        """
        user = get_user()
        conn = get_db()
        cur = conn.cursor()

        # --- Traer la tarea con responsable y creador ---
        cur.execute("""
            SELECT t.id,
                t.titulo,
                t.descripcion,
                t.estado,
                t.fecha_creacion,
                t.fecha_inicio,
                t.fecha_compromiso,
                t.fecha_fin,
                t.fecha_cierre_real,
                t.usuario_id,
                t.creador_id,
                t.solicitante_id,                         -- 👈 nuevo
                u.username        AS responsable_username,
                u.nombre_completo AS responsable_nombre,
                c.username        AS creador_username,
                c.nombre_completo AS creador_nombre,
                s.username        AS solicitante_username, -- 👈 nuevo
                s.nombre_completo AS solicitante_nombre    -- 👈 nuevo
            FROM tareas t
            JOIN usuarios u ON t.usuario_id = u.id
        LEFT JOIN usuarios c ON t.creador_id = c.id
        LEFT JOIN usuarios s ON t.solicitante_id = s.id     -- 👈 join solicitante
            WHERE t.id = ?
        """, (task_id,))
        tarea = cur.fetchone()


        if not tarea:
            conn.close()
            flash('Tarea no encontrada.', 'warning')
            return redirect(url_for('listar_tareas'))

        # --- Permisos básicos para ver ---
        # Admin siempre puede. 
        # El resto: responsable o creador.
        if user['rol'] != 'admin':
            if tarea['usuario_id'] != user['id'] and tarea['creador_id'] != user['id']:
                conn.close()
                flash('No tiene permiso para ver esta tarea.', 'danger')
                return redirect(url_for('listar_tareas'))

        # --- Registrar una nueva acción (POST) ---
        if request.method == 'POST':
            observacion = (request.form.get('observacion') or '').strip()
            detalles    = (request.form.get('detalles') or '').strip()
            
            # --- NUEVOS CAMPOS CAPTURADOS ---
            estado_accion = request.form.get('estado_accion')
            fecha_fin_tentativa = request.form.get('fecha_fin_tentativa')
            usuario_accion_id = request.form.get('usuario_accion_id') # El del select de responsables

            print('*****************************')

            print(usuario_accion_id)
            # --------------------------------

            if not observacion and not detalles:
                flash('Escribe al menos una observación o detalle para registrar la acción.', 'warning')
                conn.close()
                return redirect(url_for('ver_tarea', task_id=task_id))

            now = datetime.now()
            now_str = now.strftime('%Y-%m-%d %H:%M:%S')

            # 1) Guardar la acción CON LOS NUEVOS CAMPOS
            # En el momento de guardar (POST)
            cur.execute("""
                INSERT INTO tarea_acciones
                (tarea_id, usuario_id, fecha_accion, observacion, detalles, 
                estado_accion, usuario_asignado_id, fecha_fin_tentativa, fecha_inicio)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (task_id, user['id'], now_str, observacion, detalles, 
                estado_accion, usuario_accion_id, fecha_fin_tentativa, now_str)) # Guardamos now_str también en fecha_inicio
            
            conn.commit()








            # 2) Preparar correo a responsables
            try:
                # --- NUEVA LÓGICA: Obtener email del usuario asignado en la OBSERVACIÓN ---
                email_asignado_observacion = None
                if usuario_accion_id:
                    cur.execute("SELECT email FROM usuarios WHERE id = ? AND disabled = 0", (usuario_accion_id,))
                    u_obs = cur.fetchone()
                    if u_obs:
                        email_asignado_observacion = u_obs['email']

                # Obtener todos los responsables de la tarea (tarea_responsables)
                cur.execute("""
                    SELECT u.email,
                           COALESCE(u.nombre_completo, u.username) AS nombre
                      FROM tarea_responsables tr
                      JOIN usuarios u ON u.id = tr.usuario_id
                     WHERE tr.tarea_id = ?
                       AND u.disabled = 0
                """, (task_id,))
                resp_rows = cur.fetchall()

                if not resp_rows:
                    cur.execute("""
                        SELECT email,
                               COALESCE(nombre_completo, username) AS nombre
                          FROM usuarios
                         WHERE id = ?
                           AND disabled = 0
                    """, (tarea['usuario_id'],))
                    row_fallback = cur.fetchone()
                    if row_fallback:
                        resp_rows = [row_fallback]

                # --- Construir lista de destinatarios (Responsables + Solicitante + Asignado en Obs) ---
                destinatarios_set = {
                    r['email'] for r in resp_rows if r['email']
                }

                # 1. Agregar solicitante si existe
                if tarea['solicitante_id']:
                    cur.execute("SELECT email FROM usuarios WHERE id = ? AND disabled = 0", (tarea['solicitante_id'],))
                    sol_row = cur.fetchone()
                    if sol_row and sol_row['email']:
                        destinatarios_set.add(sol_row['email'])

                # 2. Agregar a la persona seleccionada en la observación (Asignado adicional)
                print('correo')
                print(email_asignado_observacion)
                if email_asignado_observacion:
                    destinatarios_set.add(email_asignado_observacion)

                destinatarios = list(destinatarios_set)







 
         
 

                # Si no hay nadie con correo, no enviamos
                if destinatarios:
                    # Preparamos el mensaje base (HTML con iconos)
                    msg, smtp_host, smtp_port, smtp_user, smtp_pass, use_tls = _send_tarea_avance_mail(
                        tarea,
                        observacion,
                        detalles,
                        now_str,
                        actor_username=user['username']
                    )

                    if msg and smtp_host:
                        msg['To'] = ", ".join(destinatarios)

                        if use_tls:
                            with smtplib.SMTP(smtp_host, smtp_port) as server:
                                server.starttls()
                                if smtp_user and smtp_pass:
                                    server.login(smtp_user, smtp_pass)
                                server.send_message(msg)
                        else:
                            with smtplib.SMTP_SSL(smtp_host, smtp_port) as server:
                                if smtp_user and smtp_pass:
                                    server.login(smtp_user, smtp_pass)
                                server.send_message(msg)

            except Exception as e:
                # No romper el flujo si el correo falla
                print("Error enviando correo de avance:", e)

            conn.close()
            flash('Acción registrada en la tarea.', 'success')
            return redirect(url_for('ver_tarea', task_id=task_id))

        # --- Traer historial de acciones ---
        # --- BUSCA ESTA SECCIÓN EN TU PYTHON ---
        cur.execute("""
            SELECT 
                a.id,
                a.tarea_id,
                a.usuario_id,
                a.fecha_accion,
                a.observacion,
                a.detalles,
                a.estado_accion,         -- Ya existe en tu tabla
                a.fecha_inicio,          -- Ya existe en tu tabla
                a.fecha_fin_tentativa,   -- Ya existe en tu tabla
                u.nombre_completo,       -- El que registró la acción
                u.username,
                ua.nombre_completo AS nombre_asignado -- 👈 ESTO ES LO QUE FALTA
            FROM tarea_acciones a
            JOIN usuarios u ON a.usuario_id = u.id
            LEFT JOIN usuarios ua ON a.usuario_asignado_id = ua.id -- 👈 EL JOIN CLAVE
            WHERE a.tarea_id = ?
            ORDER BY a.fecha_accion ASC, a.id ASC
        """, (task_id,))
        acciones = cur.fetchall()
  

        # Quién puede agregar acciones: admin, responsable o creador
        puede_anotar = (
            user['rol'] == 'admin'
            or tarea['usuario_id'] == user['id']
            or tarea['creador_id'] == user['id']
        )
        # --- 2. ¡ESTO ES LO QUE TE FALTA! ---
        # Traer la lista de usuarios para llenar los select de "Solicitante" y "Responsable"
        cur.execute('''
            SELECT id, username, nombre_completo 
            FROM usuarios 
            WHERE disabled = 0 
            ORDER BY nombre_completo ASC, username ASC
        ''')
        # Convertimos a lista de diccionarios para que el HTML lo entienda
        lista_usuarios = [dict(row) for row in cur.fetchall()]

        conn.close()

        return render_template(
            'tarea_detalle.html',
            tarea=tarea,
            acciones=acciones,
            puede_anotar=puede_anotar,
            usuario=user['username'],
            rol=user['rol'],
            active_page='tareas',
            responsables=lista_usuarios  # <--- Esta variable es la que recorre el {% for s in responsables %}
        )

 
    @app.route('/tareas/reenviar-accion/<int:accion_id>', methods=['POST'])
    @require_login
    @require_permission('tareas', 'ver')
    def reenviar_observacion(accion_id):
        import smtplib # Importante asegurarse de que esté importado
        conn = get_db()
        cur = conn.cursor()
        
        # 1. Obtener datos de la acción y la tarea
        cur.execute("""
            SELECT a.*, t.titulo, t.usuario_id as resp_tarea_id, t.solicitante_id, t.creador_id
            FROM tarea_acciones a
            JOIN tareas t ON a.tarea_id = t.id
            WHERE a.id = ?
        """, (accion_id,))
        accion = cur.fetchone()
        
        if not accion:
            conn.close()
            flash("Acción no encontrada.", "danger")
            return redirect(request.referrer)

        try:
            # 2. Recolectar destinatarios
            destinatarios_set = set()
            ids_interesados = [accion['resp_tarea_id'], accion['solicitante_id'], 
                            accion['creador_id'], accion['usuario_asignado_id']]
            
            for uid in ids_interesados:
                if uid:
                    cur.execute("SELECT email FROM usuarios WHERE id = ? AND disabled = 0", (uid,))
                    res = cur.fetchone()
                    if res and res['email']:
                        destinatarios_set.add(res['email'])

            if destinatarios_set:
                now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # 3. Preparar el mensaje usando tu función base
                msg, smtp_host, smtp_port, smtp_user, smtp_pass, use_tls = _send_tarea_avance_mail(
                    accion, 
                    f"(RECORDATORIO) {accion['observacion']}", 
                    accion['detalles'] or "", 
                    now_str, 
                    actor_username=session.get('usuario')
                )
                
                if msg and smtp_host:
                    # --- SOLUCIÓN AL ERROR DE DUPLICADO DE ASUNTO ---
                    del msg['Subject'] 
                    msg['Subject'] = f"RECORDATORIO: {accion['titulo']}" 
                    msg['To'] = ", ".join(destinatarios_set)

                    # --- BLOQUE DE ENVÍO REAL (EL QUE PUSISTE EN TU PREGUNTA) ---
                    if use_tls:
                        with smtplib.SMTP(smtp_host, smtp_port) as server:
                            server.starttls()
                            if smtp_user and smtp_pass:
                                server.login(smtp_user, smtp_pass)
                            server.send_message(msg)
                    else:
                        with smtplib.SMTP_SSL(smtp_host, smtp_port) as server:
                            if smtp_user and smtp_pass:
                                server.login(smtp_user, smtp_pass)
                            server.send_message(msg)
                    
                    flash(f"Recordatorio enviado a: {len(destinatarios_set)} personas.", "success")
            else:
                flash("No hay correos electrónicos válidos para notificar.", "warning")

        except Exception as e:
            print(f"Error reenviando correo: {e}")
            flash(f"Error al enviar: {str(e)}", "danger")
        
        finally:
            conn.close()
        
        return redirect(url_for('ver_tarea', task_id=accion['tarea_id']))
    


    
    @app.route('/tareas/accion/<int:accion_id>/finalizar', methods=['POST'])
    @require_login
    def finalizar_accion(accion_id):
        conn = get_db()
        cur = conn.cursor()
        
        # Buscamos la tarea_id para poder redirigir de vuelta
        cur.execute("SELECT tarea_id FROM tarea_acciones WHERE id = ?", (accion_id,))
        res = cur.fetchone()
        
        if res:
            tarea_id = res['tarea_id']
            # Actualizamos el estado a 'Finalizado'
            cur.execute("""
                UPDATE tarea_acciones 
                SET estado_accion = 'Finalizado' 
                WHERE id = ?
            """, (accion_id,))
            conn.commit()
            flash('Estado de la acción actualizado a Finalizado.', 'success')
            conn.close()
            return redirect(url_for('ver_tarea', task_id=tarea_id))
        
        conn.close()
        flash('No se pudo encontrar la acción.', 'danger')
        return redirect(url_for('listar_tareas'))


    @app.route('/tareas/<int:task_id>/editar', methods=['GET', 'POST'])
    @require_login
    @require_permission('tareas', 'editar')
    def editar_tarea(task_id):
        user = get_user()
        conn = get_db()
        cur = conn.cursor()
    # --- 1. Cargar datos necesarios para el formulario (Siempre al inicio) ---
        cur.execute("SELECT id, razon_social FROM empresas WHERE activo = 1 ORDER BY razon_social ASC")
        empresas = cur.fetchall()

        # 1. Traer la tarea (Incluyendo tipo_tarea_id)
        cur.execute("""
            SELECT id, titulo, descripcion, estado, fecha_inicio, 
                fecha_compromiso, fecha_cierre_real, usuario_id, 
                fecha_fin, creador_id, solicitante_id,
                tipo_tarea_id,porcentaje_avance,EMPRESA_ID
            FROM tareas
            WHERE id=?
        """, (task_id,))
        row = cur.fetchone()

        if not row:
            conn.close()
            flash('Tarea no encontrada.', 'warning')
            return redirect(url_for('listar_tareas'))

        tarea = dict(row)

        # 2. Listas para los combos (Solicitantes y Tipos de Tarea)
        cur.execute("""
            SELECT id, username, COALESCE(nombre_completo,'') AS nombre_completo
            FROM usuarios WHERE disabled = 0
            ORDER BY nombre_completo, username
        """)
        solicitantes = cur.fetchall()

        cur.execute("""
            SELECT id, nombre FROM param_values 
            WHERE group_id = 4945 AND activo = 1 
            ORDER BY orden ASC
        """)
        tipos_tarea = cur.fetchall()

        # Validaciones de permisos
        if user['rol'] != 'admin' and tarea['estado'] == 'Cerrado por sistema':
            conn.close()
            flash('No se puede editar una tarea cerrada por el sistema.', 'warning')
            return redirect(url_for('listar_tareas'))

        cur.execute("SELECT 1 FROM tarea_responsables WHERE tarea_id = ? AND usuario_id = ?", (task_id, user['id']))
        es_responsable = cur.fetchone() is not None

        editable = (user['rol'] == 'admin') or (es_responsable and tarea['estado'] not in ('Terminado', 'Cerrado por sistema'))

        if not editable:
            conn.close()
            flash('No tiene permiso para editar esta tarea.', 'danger')
            return redirect(url_for('listar_tareas'))

        # 3. Procesar el POST
        if request.method == 'POST':
            ahora = datetime.now()
            titulo = request.form['titulo'].strip()
            descripcion = request.form.get('descripcion', '').strip()
            estado = request.form.get('estado', tarea['estado']).strip()
            avance = request.form.get('porcentaje_avance', 0) # Nuevo campo
            # Capturar el nuevo valor
            empresa_id = request.form.get('empresa_id')
            # Inicializar IDs para evitar errores de UndefinedVariable
            solicitante_raw = (request.form.get('solicitante_id') or '').strip()
            try:
                solicitante_id = int(solicitante_raw) if solicitante_raw else tarea.get('solicitante_id')
            except ValueError:
                solicitante_id = tarea.get('solicitante_id')

            tipo_tarea_raw = (request.form.get('tipo_tarea_id') or '').strip()
            try:
                tipo_tarea_id = int(tipo_tarea_raw) if tipo_tarea_raw else tarea.get('tipo_tarea_id')
            except ValueError:
                tipo_tarea_id = tarea.get('tipo_tarea_id')

                        
            tipo_tarea_raw = request.form.get('tipo_tarea_id')
            tipo_tarea_id = int(tipo_tarea_raw) if tipo_tarea_raw else tarea.get('tipo_tarea_id')

            # Fechas
            fi_raw = request.form.get('fecha_inicio', '').strip()
            fc_raw = request.form.get('fecha_compromiso', '').strip()
            ff_raw = request.form.get('fecha_fin', '').strip()

            fcr_raw = request.form.get('fecha_cierre_real', '').strip()

            try:
                dt_fi = datetime.strptime(fi_raw, '%Y-%m-%dT%H:%M') if fi_raw else None
                dt_fc = datetime.strptime(fc_raw, '%Y-%m-%dT%H:%M') if fc_raw else None
                dt_ff = datetime.strptime(ff_raw, '%Y-%m-%dT%H:%M') if ff_raw else None

                dcr_ff = datetime.strptime(fcr_raw, '%Y-%m-%dT%H:%M') if fcr_raw else None

                fi_str = dt_fi.strftime('%Y-%m-%d %H:%M:%S') if dt_fi else None
                fc_str = dt_fc.strftime('%Y-%m-%d %H:%M:%S') if dt_fc else None
                ff_str = dt_ff.strftime('%Y-%m-%d %H:%M:%S') if dt_ff else None
                fcr_str = dcr_ff.strftime('%Y-%m-%d %H:%M:%S') if dcr_ff else None


                # Lógica: Si la fecha fin ya pasó, marcar como Terminado automáticamente
                if dt_ff and dt_ff < ahora:
                    estado = 'Terminado'

            except ValueError:
                flash('Formato de fecha inválido.', 'danger')
                return redirect(url_for('editar_tarea', task_id=task_id))

            # Manejo de la fecha de cierre real
            fecha_cierre_real_str = tarea.get('fecha_cierre_real')
            if estado == 'Terminado':
                if not fecha_cierre_real_str:
                    fecha_cierre_real_str = ahora.strftime('%Y-%m-%d %H:%M:%S')
            else:
                # Si el estado no es terminado, nos aseguramos de que no haya fecha de cierre
                fecha_cierre_real_str = None

            # 4. Actualizar Base de Datos
            cur.execute("""
                UPDATE tareas
                SET titulo = ?, descripcion = ?, empresa_id = ?,estado = ?,
                    fecha_inicio = ?, fecha_compromiso = ?, fecha_fin = ?,
                    fecha_cierre_real = ?, solicitante_id = ?,porcentaje_avance = ?,
                    tipo_tarea_id = ?
                WHERE id = ?
            """, (
                titulo, descripcion,empresa_id, estado, fi_str, fc_str, ff_str,
                fecha_cierre_real_str, solicitante_id, avance,tipo_tarea_id ,task_id
            ))

            conn.commit() # Importante para salvar los cambios
            conn.close()
            flash('Tarea actualizada correctamente.', 'success')
            return redirect(url_for('listar_tareas'))

        # Si es GET
        conn.close()
        return render_template(
            'editar_tarea.html',
            tarea=tarea,
            empresas=empresas,
            estados=ESTADOS,
            usuario=user['username'],
            rol=user['rol'],
            is_admin=(user['rol'] == 'admin'),
            solicitantes=solicitantes,
            tipos_tarea=tipos_tarea,
            active_page='tareas'
        )
  
  
    @app.route('/tareas/<int:task_id>/eliminar', methods=['POST'])
    @require_login
    @require_permission('tareas', 'eliminar')
    def eliminar_tarea(task_id):
        user = get_user()
        conn = get_db()
        cur = conn.cursor()

        # Si no es admin, solo puede eliminar las suyas
        if user['rol'] != 'admin':
            cur.execute("DELETE FROM tareas WHERE id=? AND usuario_id=?", (task_id, user['id']))
        else:
            cur.execute("DELETE FROM tareas WHERE id=?", (task_id,))

        changes = cur.rowcount
        conn.commit()
        conn.close()

        if changes:
            flash('Tarea eliminada.', 'success')
        else:
            flash('No se pudo eliminar (permiso o inexistente).', 'warning')
        return redirect(url_for('listar_tareas'))






  
            


 
    @app.route('/tareas/reporte/excel')
    @require_login
    def exportar_tareas_excel():
        q = request.args.get('q', '').lower()
        estado_f = request.args.get('estado', '')
        prop_f = request.args.get('prop', '').lower()
        depto_f = request.args.get('depto', '').lower()

        conn = get_db()
        
        # Query con los JOINs correctos hacia param_values y departamentos
        query = """
            SELECT 
                t.id AS "Código Tarea",
                t.titulo AS "Título",
                t.descripcion AS "Descripción Tarea",
                t.estado AS "Estado Global",
                pv.nombre AS "Tipo Tarea", -- Aquí obtenemos el nombre del tipo
                u_crea.nombre_completo AS "Creado por",
                t.fecha_creacion AS "F. Creación",
                t.fecha_compromiso AS "F. Compromiso",
                u_res.nombre_completo AS "Responsable Tarea",
                d.nombre AS "Departamento",
                -- Historial de acciones
                a.fecha_accion AS "Fecha Acción",
                u_acc.nombre_completo AS "Acción registrada por",
                a.observacion AS "Resumen Acción",
                a.detalles AS "Detalle Técnico Acción",
                a.estado_accion AS "Estado del Paso",
                u_asig.nombre_completo AS "Asignado en este paso",
                a.fecha_fin_tentativa AS "Fin Tentativo Paso"
            FROM tareas t
            LEFT JOIN param_values pv ON t.tipo_tarea_id = pv.id
            LEFT JOIN usuarios u_crea ON t.creador_id = u_crea.id
            LEFT JOIN usuarios u_res ON t.usuario_id = u_res.id
            LEFT JOIN departamentos d ON u_res.departamento_id = d.id
            LEFT JOIN tarea_acciones a ON t.id = a.tarea_id
            LEFT JOIN usuarios u_acc ON a.usuario_id = u_acc.id
            LEFT JOIN usuarios u_asig ON a.usuario_asignado_id = u_asig.id 
            WHERE 1=1
        """
        
        try:
            df = pd.read_sql_query(query, conn)
        except Exception as e:
            conn.close()
            return f"Error SQL: {str(e)}", 500
        conn.close()

        # --- Filtros de Pandas (Sincronizados con la tabla HTML) ---
        if q:
            df = df[df['Título'].astype(str).str.lower().str.contains(q) | 
                    df['Descripción Tarea'].astype(str).str.lower().str.contains(q)]
        if estado_f:
            df = df[df['Estado Global'] == estado_f]
        if prop_f:
            df = df[df['Responsable Tarea'].astype(str).str.lower().str.contains(prop_f)]
        if depto_f:
            df = df[df['Departamento'].astype(str).str.lower().str.contains(depto_f)]

        # --- Construcción del Excel con Cabecera ---
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Los datos empiezan en la fila 7 para dar espacio a la cabecera extendida
            df.to_excel(writer, index=False, sheet_name='Reporte Tareas', startrow=6)
            
            ws = writer.sheets['Reporte Tareas']
            from openpyxl.styles import Font, Alignment
            
            # Estilos
            label_font = Font(bold=True, size=11)
            value_font = Font(bold=False, size=11)
            title_font = Font(bold=True, size=14)

            # 1. Nombre del reporte (Celda A1: Nombre de reporte, B1: Reporte de Tareas)
            ws['A1'] = "Nombre de reporte:"
            ws['B1'] = "Reporte de Tareas Detallado"
            ws['A1'].font = label_font
            ws['B1'].font = value_font

            # 2. Fecha de generación
            ws['A2'] = "Fecha de generación:"
            ws['B2'] = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            ws['A2'].font = label_font
            
            # 3. Usuario de generación
            ws['A3'] = "Usuario de generación:"
            ws['B3'] = session.get('nombre_completo', session.get('username', 'Sistema'))
            ws['A3'].font = label_font

            # Ajuste automático de columnas
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)

        output.seek(0)
        return send_file(output, 
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, 
                        download_name=f"Reporte_Sili_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")




    def _send_tarea_creada_mail(
        to_email: str,
        solicitante_nombre: str,
        titulo: str,
        descripcion: str,
        fi_str: str,
        fc_str: str,
        ff_str: str,
        responsables_nombres: list[str]
    ):
        """
        Envía un correo al solicitante cuando se crea una nueva tarea.
        Usa la configuración SMTP almacenada en la tabla de configuración (get_config_value).
        """
        try:
            smtp_host = get_config_value('smtp_host', 'localhost')
            smtp_port = int(get_config_value('smtp_port', '587'))
            smtp_user = get_config_value('smtp_user', '')
            smtp_pass = get_config_value('smtp_pass', '')
            smtp_from = get_config_value('smtp_from', smtp_user or 'no-reply@localhost')
            use_tls   = (get_config_value('smtp_tls', '1') == '1')

            if not smtp_host or not smtp_from or not to_email:
                return  # sin config o sin destinatario, no hacemos nada

            resp_txt = ", ".join(responsables_nombres) if responsables_nombres else "Sin responsables asignados"

            def fmt(fecha_str):
                return fecha_str or 'No definida'

            subject = f"Nueva tarea creada: {titulo}"

            body = f"""Hola {solicitante_nombre},

                Se ha creado una nueva tarea en el sistema.

                Título: {titulo}
                Descripción: {descripcion or 'Sin descripción'}

                Responsables: {resp_txt}
                Fecha inicio: {fmt(fi_str)}
                Fecha compromiso: {fmt(fc_str)}
                Fecha fin: {fmt(ff_str)}

                Este es un mensaje automático del sistema de tareas.
                """

            msg = EmailMessage()
            msg['Subject'] = subject
            msg['From'] = smtp_from
            msg['To'] = to_email
            msg.set_content(body)

            # Envío
            if use_tls:
                with smtplib.SMTP(smtp_host, smtp_port) as server:
                    server.starttls()
                    if smtp_user and smtp_pass:
                        server.login(smtp_user, smtp_pass)
                    server.send_message(msg)
            else:
                with smtplib.SMTP_SSL(smtp_host, smtp_port) as server:
                    if smtp_user and smtp_pass:
                        server.login(smtp_user, smtp_pass)
                    server.send_message(msg)

        except Exception as e:
            # Puedes cambiar esto por current_app.logger.exception(...)
            print("Error enviando correo de tarea creada:", e)




    def _norm_email(s: str) -> str:
        return (s or "").strip().lower()

    def _extract_email(value) -> str:
        """
        Acepta:
        - "Nombre <a@b.com>"
        - "a@b.com"
        """
        if not value:
            return ""
        raw = str(value)
        m = re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", raw, flags=re.I)
        return _norm_email(m.group(0)) if m else ""

    def _require_api_key(req) -> tuple[bool, str]:
        expected = (get_config_value("api_inbound_key", "") or "").strip()
        got = (req.headers.get("X-API-Key") or "").strip()
        if not expected:
            return False, "Falta configurar api_inbound_key en tabla configuracion."
        if got != expected:
            return False, "API key inválida."
        return True, ""

    def _find_user_id_by_email(cur, email: str):
        if not email:
            return None
        cur.execute("""
            SELECT id
            FROM usuarios
            WHERE disabled = 0
            AND lower(trim(email)) = ?
            LIMIT 1
        """, (email,))
        row = cur.fetchone()
        return int(row["id"]) if row else None


    @app.route("/api/inbound/email", methods=["POST"])
    def api_inbound_email_create_task():
        # 1) Seguridad (API Key)
        ok, msg = _require_api_key(request)
        if not ok:
            return jsonify({"ok": False, "error": msg}), 401

        # 2) Validar JSON
        if not request.is_json:
            return jsonify({"ok": False, "error": "Content-Type debe ser application/json"}), 415

        payload = request.get_json(silent=True) or {}

        subject = (payload.get("subject") or "").strip()
        body = (payload.get("body") or payload.get("text") or payload.get("body_text") or "").strip()
        from_email = _extract_email(payload.get("from") or payload.get("from_email"))

        if not subject:
            return jsonify({"ok": False, "error": "Falta subject"}), 400
        if not from_email:
            return jsonify({"ok": False, "error": "Falta from (remitente)"}), 400
        if not body:
            body = "(sin cuerpo)"

        # Default responsable
        default_email = "jchavez@quimpac.com.ec"

        conn = get_db()
        cur = conn.cursor()

        # 3) Determinar responsable/solicitante
        uid_from = _find_user_id_by_email(cur, from_email)

        if uid_from:
            responsable_principal = uid_from
            solicitante_id = uid_from
        else:
            uid_def = _find_user_id_by_email(cur, _norm_email(default_email))
            if not uid_def:
                conn.close()
                return jsonify({
                    "ok": False,
                    "error": f"Remitente {from_email} no existe y el default {default_email} no está en usuarios."
                }), 400

            responsable_principal = uid_def
            solicitante_id = None  # no hay usuario válido para el remitente

        creador_id = uid_from or responsable_principal

        now = datetime.now()
        estado = "Por iniciar"

        # 4) Insertar tarea
        cur.execute("""
            INSERT INTO tareas
                (titulo, descripcion, estado, fecha_creacion, fecha_inicio, fecha_compromiso, fecha_fin,
                usuario_id, creador_id, solicitante_id, notificado)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)
        """, (
            subject[:250],
            body,
            estado,
            now.strftime("%Y-%m-%d %H:%M:%S"),
            None, None, None,
            responsable_principal,
            creador_id,
            solicitante_id,
        ))
        tarea_id = cur.lastrowid

        # 5) Tabla puente responsables
        cur.execute("""
            INSERT OR IGNORE INTO tarea_responsables (tarea_id, usuario_id)
            VALUES (?, ?)
        """, (tarea_id, responsable_principal))

        conn.commit()
        conn.close()

        return jsonify({
            "ok": True,
            "tarea_id": tarea_id,
            "codigo_tarea": f"{tarea_id:07d}",
            "asignado_a_user_id": responsable_principal,
            "from_email": from_email,
            "from_encontrado": bool(uid_from),
            "default_email": default_email
        }), 201
